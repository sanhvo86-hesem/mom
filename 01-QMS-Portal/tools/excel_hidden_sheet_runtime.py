#!/usr/bin/env python3
import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def load_json(path: str):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def save_json_stdout(payload):
    sys.stdout.write(json.dumps(payload, ensure_ascii=False))


def ensure_sheet(workbook, name: str):
    if name in workbook.sheetnames:
        sheet = workbook[name]
    else:
        sheet = workbook.create_sheet(title=name)
    sheet.sheet_state = "veryHidden"
    sheet.protection.sheet = True
    return sheet


def set_cell(sheet, cell: str, value):
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    sheet[cell] = "" if value is None else value


def read_cell(sheet, cell: str):
    value = sheet[cell].value
    return "" if value is None else value


def try_parse_json(value):
    if not isinstance(value, str):
        return value
    value = value.strip()
    if not value:
        return ""
    if (value.startswith("{") and value.endswith("}")) or (value.startswith("[") and value.endswith("]")):
        try:
            return json.loads(value)
        except Exception:
            return value
    return value


def init_upload_log(workbook, spec):
    log_name = spec.get("upload_log_sheet", "_QMS_UPLOAD_LOG")
    log_sheet = ensure_sheet(workbook, log_name)
    if log_sheet.max_row == 1 and all(log_sheet.cell(1, idx + 1).value in (None, "") for idx in range(len(spec.get("upload_log_columns", [])))):
      for idx, column in enumerate(spec.get("upload_log_columns", []), start=1):
          log_sheet.cell(1, idx).value = column.get("key", f"col_{idx}")
    return log_sheet


def issue_package(args):
    spec = load_json(args.spec)
    metadata = load_json(args.meta)
    src = Path(args.src)
    dst = Path(args.dst)
    dst.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(src, keep_vba=src.suffix.lower() == ".xlsm")
    verify_sheet = ensure_sheet(workbook, spec.get("sheet_name", "_QMS_VERIFY"))
    for field in spec.get("fields", []):
        key = field.get("key")
        cell = field.get("cell")
        if not key or not cell:
            continue
        set_cell(verify_sheet, cell, metadata.get(key, field.get("value", "")))

    init_upload_log(workbook, spec)
    workbook.save(dst)
    save_json_stdout({"ok": True, "output": str(dst)})


def inspect_package(args):
    spec = load_json(args.spec)
    src = Path(args.src)
    workbook = load_workbook(src, keep_vba=src.suffix.lower() == ".xlsm", data_only=False)
    sheet_name = spec.get("sheet_name", "_QMS_VERIFY")
    if sheet_name not in workbook.sheetnames:
        save_json_stdout({"ok": False, "error": "verify_sheet_missing"})
        return 2

    verify_sheet = workbook[sheet_name]
    payload = {}
    for field in spec.get("fields", []):
        key = field.get("key")
        cell = field.get("cell")
        if not key or not cell:
            continue
        payload[key] = try_parse_json(read_cell(verify_sheet, cell))

    upload_log = []
    log_name = spec.get("upload_log_sheet", "_QMS_UPLOAD_LOG")
    if log_name in workbook.sheetnames:
        log_sheet = workbook[log_name]
        headers = [cell.value for cell in log_sheet[1]]
        for row in log_sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            record = {}
            for idx, value in enumerate(row):
                key = headers[idx] if idx < len(headers) and headers[idx] else f"col_{idx+1}"
                record[str(key)] = try_parse_json(value)
            upload_log.append(record)

    save_json_stdout({"ok": True, "metadata": payload, "upload_log": upload_log})
    return 0


def append_receipt(args):
    spec = load_json(args.spec)
    receipt = load_json(args.receipt)
    src = Path(args.src)
    dst = Path(args.dst)
    dst.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(src, keep_vba=src.suffix.lower() == ".xlsm")
    verify_sheet = ensure_sheet(workbook, spec.get("sheet_name", "_QMS_VERIFY"))
    log_sheet = init_upload_log(workbook, spec)

    fields_by_key = {field.get("key"): field for field in spec.get("fields", [])}
    for key in ("receipt_status", "receipt_version", "latest_stored_filename", "latest_upload_timestamp", "receipt_history_json"):
        field = fields_by_key.get(key)
        if field and field.get("cell"):
            set_cell(verify_sheet, field["cell"], receipt.get(key, ""))

    headers = [cell.value for cell in log_sheet[1]]
    next_row = log_sheet.max_row + 1
    for idx, header in enumerate(headers, start=1):
        set_cell(log_sheet, f"{get_column_letter(idx)}{next_row}", receipt.get(str(header), ""))

    workbook.save(dst)
    save_json_stdout({"ok": True, "output": str(dst)})
    return 0


def get_column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def main():
    parser = argparse.ArgumentParser(description="HESEM Excel hidden sheet runtime helper")
    sub = parser.add_subparsers(dest="command", required=True)

    issue = sub.add_parser("issue")
    issue.add_argument("--src", required=True)
    issue.add_argument("--dst", required=True)
    issue.add_argument("--spec", required=True)
    issue.add_argument("--meta", required=True)

    inspect = sub.add_parser("inspect")
    inspect.add_argument("--src", required=True)
    inspect.add_argument("--spec", required=True)

    append = sub.add_parser("append-receipt")
    append.add_argument("--src", required=True)
    append.add_argument("--dst", required=True)
    append.add_argument("--spec", required=True)
    append.add_argument("--receipt", required=True)

    args = parser.parse_args()
    if args.command == "issue":
        issue_package(args)
        return 0
    if args.command == "inspect":
        return inspect_package(args)
    if args.command == "append-receipt":
        return append_receipt(args)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
