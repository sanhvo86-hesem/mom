#!/usr/bin/env python3
"""
HESEM QMS - Fix Owner & Approved fields in all form headers.
Based on Authority Matrix (ANNEX-QMS-025) and RACI Master Matrix.

Rule: Owner = Department responsible for the process
      Approved = Per authority matrix (all forms approved by GM)
"""
import os
from openpyxl import load_workbook

# ═══════ OWNER/APPROVED MAPPING (per Authority Matrix) ═══════
# Key: form code prefix -> (Owner, Approved)
OWNER_MAP = {
    # FRM-1xx: QMS / Document Control / Governance
    'FRM-101': ('QA / QMS', 'General Manager'),
    'FRM-102': ('QA / QMS', 'General Manager'),
    'FRM-104': ('QA / QMS', 'General Manager'),
    'FRM-105': ('QA / QMS', 'General Manager'),
    'FRM-106': ('QA / QMS', 'General Manager'),
    'FRM-110': ('IT / QMS', 'General Manager'),
    'FRM-111': ('IT / QMS', 'General Manager'),
    'FRM-121': ('QA / QMS', 'General Manager'),
    'FRM-122': ('QA / QMS', 'General Manager'),
    'FRM-123': ('QA / QMS', 'General Manager'),
    'FRM-124': ('QA / QMS', 'General Manager'),
    'FRM-125': ('QA / QMS', 'General Manager'),
    'FRM-131': ('QA / QMS', 'General Manager'),
    'FRM-132': ('Engineering / QA', 'General Manager'),
    'FRM-133': ('Engineering / QA', 'General Manager'),
    'FRM-141': ('IT / QMS', 'General Manager'),
    'FRM-151': ('QA / QMS', 'General Manager'),
    'FRM-161': ('Engineering', 'General Manager'),
    'FRM-162': ('Engineering', 'General Manager'),
    'FRM-163': ('QA / QMS', 'General Manager'),
    'FRM-171': ('QA / QMS', 'General Manager'),
    'FRM-181': ('QA / QMS', 'General Manager'),

    # FRM-2xx: Sales / Order Fulfillment / Customer
    'FRM-201': ('Sales / CS', 'General Manager'),
    'FRM-202': ('Sales / CS', 'General Manager'),
    'FRM-203': ('Planning', 'General Manager'),
    'FRM-204': ('Sales / Planning', 'General Manager'),
    'FRM-205': ('QA / Planning', 'General Manager'),
    'FRM-206': ('Planning / QA', 'General Manager'),
    'FRM-207': ('Production / QA', 'General Manager'),
    'FRM-208': ('Production', 'General Manager'),
    'FRM-209': ('Planning / QA', 'General Manager'),
    'FRM-211': ('QA / CS', 'General Manager'),
    'FRM-212': ('Sales / Engineering', 'General Manager'),
    'FRM-213': ('QA / CS', 'General Manager'),
    'FRM-221': ('SCM / QA', 'General Manager'),

    # FRM-3xx: Engineering
    'FRM-301': ('Engineering', 'General Manager'),
    'FRM-302': ('Engineering', 'General Manager'),
    'FRM-303': ('Engineering / QA', 'General Manager'),
    'FRM-304': ('Engineering / QA', 'General Manager'),
    'FRM-305': ('QA / Engineering', 'General Manager'),
    'FRM-306': ('Engineering', 'General Manager'),
    'FRM-307': ('Engineering', 'General Manager'),
    'FRM-311': ('QA / Engineering', 'General Manager'),

    # FRM-4xx: Supply Chain / Purchasing
    'FRM-401': ('SCM', 'General Manager'),
    'FRM-402': ('SCM / QA', 'General Manager'),
    'FRM-403': ('SCM / QA', 'General Manager'),
    'FRM-404': ('SCM / QA', 'General Manager'),
    'FRM-405': ('SCM / QA', 'General Manager'),
    'FRM-406': ('QA / SCM', 'General Manager'),
    'FRM-408': ('SCM / QA', 'General Manager'),
    'FRM-409': ('QA / SCM', 'General Manager'),
    'FRM-411': ('QA / SCM', 'General Manager'),
    'FRM-413': ('QA / SCM', 'General Manager'),

    # FRM-5xx: Production / Maintenance
    'FRM-501': ('Planning', 'General Manager'),
    'FRM-502': ('Planning', 'General Manager'),
    'FRM-503': ('Planning / Production', 'General Manager'),
    'FRM-504': ('Production', 'General Manager'),
    'FRM-511': ('Production / QA', 'General Manager'),
    'FRM-512': ('Production', 'General Manager'),
    'FRM-513': ('Production', 'General Manager'),
    'FRM-514': ('Production / Engineering', 'General Manager'),
    'FRM-518': ('Engineering / Production', 'General Manager'),
    'FRM-519': ('Production / QA', 'General Manager'),
    'FRM-521': ('Maintenance', 'General Manager'),
    'FRM-522': ('Production / Maintenance', 'General Manager'),
    'FRM-523': ('Production', 'General Manager'),
    'FRM-524': ('Maintenance', 'General Manager'),
    'FRM-525': ('QA / Metrology', 'General Manager'),

    # FRM-6xx: Quality / Metrology / Inspection
    'FRM-601': ('QA', 'General Manager'),
    'FRM-602': ('QA', 'General Manager'),
    'FRM-611': ('QA', 'General Manager'),
    'FRM-612': ('QA', 'General Manager'),
    'FRM-613': ('QA', 'General Manager'),
    'FRM-621': ('QA', 'General Manager'),
    'FRM-631': ('QA', 'General Manager'),
    'FRM-641': ('QA', 'General Manager'),
    'FRM-642': ('QA', 'General Manager'),
    'FRM-643': ('QA', 'General Manager'),
    'FRM-651': ('QA', 'General Manager'),
    'FRM-652': ('QA', 'General Manager'),
    'FRM-653': ('QA', 'General Manager'),
    'FRM-654': ('QA / CS', 'General Manager'),

    # FRM-7xx: Warehouse / Shipping / Cleanroom
    'FRM-701': ('SCM / QA', 'General Manager'),
    'FRM-702': ('SCM / QA', 'General Manager'),
    'FRM-703': ('SCM', 'General Manager'),
    'FRM-704': ('SCM', 'General Manager'),
    'FRM-705': ('SCM', 'General Manager'),
    'FRM-706': ('SCM', 'General Manager'),
    'FRM-707': ('SCM / QA', 'General Manager'),
    'FRM-708': ('Production / QA', 'General Manager'),
    'FRM-709': ('Production / QA', 'General Manager'),
    'FRM-711': ('QA / Production', 'General Manager'),
    'FRM-712': ('QA', 'General Manager'),
    'FRM-713': ('Production / QA', 'General Manager'),
    'FRM-714': ('Production / QA', 'General Manager'),
    'FRM-715': ('Production / QA', 'General Manager'),
    'FRM-721': ('Production / QA', 'General Manager'),

    # FRM-8xx: HR / Training / EHS / Finance
    'FRM-801': ('HR', 'General Manager'),
    'FRM-802': ('HR', 'General Manager'),
    'FRM-803': ('HR / Dept. Manager', 'General Manager'),
    'FRM-804': ('HR / Dept. Manager', 'General Manager'),
    'FRM-805': ('HR', 'General Manager'),
    'FRM-806': ('HR', 'General Manager'),
    'FRM-807': ('HR / Dept. Manager', 'General Manager'),
    'FRM-808': ('HR', 'General Manager'),
    'FRM-809': ('HR / Dept. Manager', 'General Manager'),
    'FRM-811': ('EHS', 'General Manager'),
    'FRM-812': ('EHS', 'General Manager'),
    'FRM-821': ('Finance', 'General Manager'),

    # FRM-9xx: Audit / Management Review
    'FRM-901': ('QMS + Top Management', 'General Manager'),
    'FRM-902': ('QMS + Top Management', 'General Manager'),
    'FRM-911': ('QMS + Top Management', 'General Manager'),
}


def get_form_code(filename):
    """Extract FRM-xxx from filename."""
    parts = filename.split('_')[0]
    return parts


def detect_meta_cols(ws):
    """Find meta value columns by searching for 'Form Code' text."""
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 6:
            cell = ws.cell(row=mg.min_row, column=mg.min_col)
            val = str(cell.value or '').strip()
            if val == 'Form Code':
                # Meta label starts here, value is next merge
                lbl_start = mg.min_col
                lbl_end = mg.max_col
                # Find value merge on same row
                for mg2 in ws.merged_cells.ranges:
                    if mg2.min_row == mg.min_row and mg2.min_col == lbl_end + 1:
                        return lbl_start, lbl_end, mg2.min_col, mg2.max_col
    return None


def find_owner_approved_rows(ws, meta_lbl_start):
    """Find which rows contain Owner and Approved labels."""
    owner_row = approved_row = None
    for r in range(1, 7):
        val = str(ws.cell(row=r, column=meta_lbl_start).value or '').strip()
        if val == 'Owner':
            owner_row = r
        elif val in ('Approved', 'Approved by'):
            approved_row = r
    return owner_row, approved_row


def process_file(filepath):
    fname = os.path.basename(filepath)
    code = get_form_code(fname)

    if code not in OWNER_MAP:
        return fname, "SKIP (no mapping)"

    owner, approved = OWNER_MAP[code]

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, f"ERROR: {e}"

    changes = []
    for sn in wb.sheetnames:
        ws = wb[sn]

        meta = detect_meta_cols(ws)
        if not meta:
            continue

        lbl_s, lbl_e, val_s, val_e = meta
        owner_row, approved_row = find_owner_approved_rows(ws, lbl_s)

        if owner_row:
            current = str(ws.cell(row=owner_row, column=val_s).value or '').strip()
            if current != owner:
                ws.cell(row=owner_row, column=val_s).value = owner
                changes.append(f"[{sn}] Owner: '{current}' -> '{owner}'")

        if approved_row:
            current = str(ws.cell(row=approved_row, column=val_s).value or '').strip()
            if current != approved:
                ws.cell(row=approved_row, column=val_s).value = approved
                changes.append(f"[{sn}] Approved: '{current}' -> '{approved}'")

    if changes:
        try:
            wb.save(filepath)
        except Exception as e:
            wb.close()
            return fname, f"SAVE ERROR: {e}"

    wb.close()
    return fname, changes if changes else "OK (no change needed)"


def main():
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS - Owner & Approved Update")
    print(f"  Source: Authority Matrix (ANNEX-QMS-025)")
    print(f"  Rule: Owner = Process department | Approved = General Manager")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    changed = 0
    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        name, result = process_file(fp)
        if isinstance(result, list) and result:
            changed += 1
            print(f"[{i:3d}] {fn[:50]:50s} UPDATED")
            for r in result:
                print(f"      {r}")
        elif isinstance(result, str) and 'ERROR' in result:
            print(f"[{i:3d}] {fn[:50]:50s} {result}")
        else:
            print(f"[{i:3d}] {fn[:50]:50s} OK")

    print(f"\n{'='*70}")
    print(f"Updated: {changed} files")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
