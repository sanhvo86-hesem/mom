# -*- coding: utf-8 -*-
"""Generate FRM-700, FRM-800, FRM-900 series — 30 forms"""
import sys, os, traceback
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from form_engine import generate_form, _borders_section_row, _borders_table_row
from wave3_log_specs import generate_log_form
from specs_frm700_800_900 import *
import openpyxl

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"
def gop(code):
    for d in os.listdir(BASE):
        t = os.path.join(BASE, d)
        if not os.path.isdir(t): continue
        for f in os.listdir(t):
            if f.startswith(code) and f.endswith('.xlsx'):
                return os.path.join(t, f)
    return None

s = f = 0

def run_checklist(specs, label):
    global s, f
    print(f"\n=== {label}: CHECKLIST/REPORT ===")
    for spec in specs:
        code = spec['code']
        try:
            out = gop(code)
            if not out: print(f"  {code} SKIP"); f += 1; continue
            sha = generate_form(spec, out)
            print(f"  {code:8s} OK  SHA={sha[:12]}"); s += 1
        except Exception as e:
            f += 1; print(f"  {code:8s} FAIL: {e}"); traceback.print_exc()

def run_logs(specs, label):
    global s, f
    print(f"\n=== {label}: LOGS ===")
    for spec in specs:
        code = spec['code']
        try:
            out = gop(code)
            if not out: print(f"  {code} SKIP"); f += 1; continue
            sha = generate_log_form(code, spec['title'], spec['owner'], spec['approved'],
                                    spec['columns'], spec['notice'], spec['data_rows'], out)
            wb = openpyxl.load_workbook(out); ws = wb.active; cols = spec['columns']
            for r in range(7, 15):
                if ws.cell(r, cols[0][0]).value == '#':
                    if 'col_comments' in spec:
                        for idx, cm in spec['col_comments'].items():
                            ws.cell(r, cols[idx][0]).comment = cm
                    break
            wb.save(out)
            print(f"  {code:8s} OK  SHA={sha[:12]} +cmt"); s += 1
        except Exception as e:
            f += 1; print(f"  {code:8s} FAIL: {e}"); traceback.print_exc()

print("=" * 70)
print("FRM-700 + FRM-800 + FRM-900: 30 FORMS")
print("=" * 70)

# FRM-700
run_checklist(FRM_700_CHECKLIST, "FRM-700")
run_logs(FRM_700_LOG, "FRM-700")
run_checklist(FRM_700_LABEL, "FRM-700 LABELS")

# FRM-800
run_checklist(FRM_800_CHECKLIST + FRM_800_WIDE + FRM_800_SIMPLE, "FRM-800")
run_logs(FRM_800_LOG, "FRM-800")
run_checklist(FRM_800_LABEL, "FRM-800 LABELS")

# FRM-900
run_checklist(FRM_900_ALL, "FRM-900")

print(f"\n{'=' * 70}")
print(f"TOTAL: {s} OK, {f} FAIL")
print(f"{'=' * 70}")
