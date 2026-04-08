# -*- coding: utf-8 -*-
import sys, os, traceback
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from form_engine import generate_form
from form_engine import _borders_section_row, _borders_table_row
from wave3_log_specs import generate_log_form
from specs_frm500 import FRM_500_CHECKLIST, FRM_500_LOG
import openpyxl

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"
def get_output_path(code):
    for d in os.listdir(BASE):
        target = os.path.join(BASE, d)
        if not os.path.isdir(target): continue
        for f in os.listdir(target):
            if f.startswith(code) and f.endswith('.xlsx'):
                return os.path.join(target, f)
    return None

print("FRM-500 SERIES: 14 FORMS (excl FRM-511 already done)")
print("=" * 70)
success = failed = 0
for spec in FRM_500_CHECKLIST:
    code = spec['code']
    try:
        out = get_output_path(code)
        if not out: print(f"  {code} SKIP"); failed += 1; continue
        sha = generate_form(spec, out)
        print(f"  {code:8s} OK  SHA={sha[:12]}"); success += 1
    except Exception as e:
        failed += 1; print(f"  {code:8s} FAIL: {e}"); traceback.print_exc()

for spec in FRM_500_LOG:
    code = spec['code']
    try:
        out = get_output_path(code)
        if not out: print(f"  {code} SKIP"); failed += 1; continue
        sha = generate_log_form(code, spec['title'], spec['owner'], spec['approved'],
                                spec['columns'], spec['notice'], spec['data_rows'], out)
        wb = openpyxl.load_workbook(out); ws = wb.active; cols = spec['columns']
        for r in range(7, 15):
            if ws.cell(r, cols[0][0]).value == '#':
                if 'col_comments' in spec:
                    for idx, cm in spec['col_comments'].items():
                        ws.cell(r, cols[idx][0]).comment = cm
                break
        wb.save(out)
        print(f"  {code:8s} OK  SHA={sha[:12]} +cmt"); success += 1
    except Exception as e:
        failed += 1; print(f"  {code:8s} FAIL: {e}"); traceback.print_exc()

print(f"\nFRM-500 COMPLETE: {success} OK, {failed} FAIL")
