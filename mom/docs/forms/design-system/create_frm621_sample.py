# -*- coding: utf-8 -*-
"""
FRM-621 SAMPLE — with cell comments (not bilingual text)
All guidance in Vietnamese as COMMENTS inside cells.
Hover over cell = see instruction.
"""
import sys, os, hashlib
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from form_engine import *
from form_engine import _borders_section_row, _borders_table_row
import openpyxl
from openpyxl.comments import Comment

TMPL = "C:/Users/TEST4/Desktop/frm-000-master-template.xlsx"
nc = 56
fmt = FORMATS['A4L']
wb = openpyxl.load_workbook(TMPL)
src_ws = wb['02-MASTER-A4-LANDSCAPE']
ws = wb.create_sheet('FRM_621', 0)
for i in range(1, nc+1):
    ws.column_dimensions[get_column_letter(i)].width = 2.375

build_header(ws, src_ws, nc, 'AQL INSPECTION RECORD', 'FRM-621', 'QA', 'QA Manager')
embed_logo(ws, fmt['logo_cols'])

zones = get_zones(nc)
CK = zones['check']
PP = zones['pair']

def cmt(text):
    """Create a comment with QMS author"""
    return Comment(text, 'QMS', width=250, height=120)

r = 7
r = spacer(ws, r, nc)

# === S1: REFERENCE ===
r = sect_header(ws, r, nc, '  1.  REFERENCE INFORMATION')
ws.cell(r-1, 1).comment = cmt(
    'Ghi thong tin tham chieu de truy xuat lo hang.\n'
    'Tat ca thong tin lay tu Job Packet.')
ref_r = r
r = data_pair(ws, r, nc, PP, 'Job / WO', 'Part No. / Revision', first=True)
r = data_pair(ws, r, nc, PP, 'Lot Size / Sample Size', 'AQL Level / Insp. Level', last=True)

# Comments on input cells
ws.cell(ref_r, PP[1][0]).comment = cmt('Ghi so Job hoac Work Order tu Job Packet.')
ws.cell(ref_r, PP[3][0]).comment = cmt('Ghi ma chi tiet va phien ban ban ve hien hanh.\nVi du: HES-1234 Rev.C')
ws.cell(ref_r+1, PP[1][0]).comment = cmt(
    'Ghi co lo va co mau theo bang AQL (ISO 2859-1).\n'
    'Vi du: Lot=100, Sample=13 (Level II, AQL 1.0)')
ws.cell(ref_r+1, PP[3][0]).comment = cmt(
    'Ghi muc AQL va muc kiem tra.\n'
    'Vi du: AQL 1.0, Inspection Level II.\n'
    'Tra bang ISO 2859-1 de xac dinh co mau.')

r = spacer(ws, r, nc)

# === S2: INSPECTION RESULTS ===
r = sect_header(ws, r, nc, '  2.  INSPECTION RESULTS')
hdrs = ['#', 'Cat.', 'Characteristic', 'Method / Instrument', 'Result', 'Inspector']
r = col_header(ws, r, nc, CK, hdrs)

# Comments on column header cells
hdr_r = r - 1
ws.cell(hdr_r, CK[1][0]).comment = cmt(
    'Phan loai dac tinh:\n'
    'QUA = Chat luong / Kich thuoc\n'
    'TEC = Ky thuat\n'
    'OPS = Van hanh')
ws.cell(hdr_r, CK[2][0]).comment = cmt(
    'Ghi ro dac tinh kiem tra tu ban ve.\n'
    'Moi dong = 1 dac tinh.\n'
    'Vi du: OD 25.00 +/-0.02, Ra 0.8,\n'
    'GD&T position 0.05')
ws.cell(hdr_r, CK[3][0]).comment = cmt(
    'Ghi phuong phap do va so hieu dung cu.\n'
    'Vi du: Micrometer #M-015\n'
    'Dung cu phai con hieu chuan (kiem tra sticker).')
ws.cell(hdr_r, CK[4][0]).comment = cmt(
    'Chon tu dropdown:\n'
    'PASS = trong dung sai\n'
    'FAIL = ngoai dung sai -> bao QA\n'
    'HOLD = can xem xet them\n'
    'NA = khong ap dung')
ws.cell(hdr_r, CK[5][0]).comment = cmt(
    'Ky tat nguoi kiem tra.\n'
    'Phai duoc dao tao va co nang luc\n'
    '(xem FRM-807 Skills Matrix).')

ck_start = r
items = [
    (1, 'QUA', 'Dimension 1.', 'Measured value vs tolerance.'),
    (2, 'QUA', 'Dimension 2.', 'Measured value vs tolerance.'),
    (3, 'QUA', 'Dimension 3.', 'Measured value vs tolerance.'),
    (4, 'QUA', 'Visual / surface finish.', 'Per drawing requirement.'),
    (5, 'QUA', 'Material / marking verification.', 'Cert match + marking correct.'),
]
for i, (num, cat, item, crit) in enumerate(items):
    check_row(ws, r, nc, CK, num, cat, item, crit, last=False)
    r += 1

# Comments on first data row as example
ws.cell(ck_start, CK[2][0]).comment = cmt(
    'Thay noi dung nay bang kich thuoc thuc te tu ban ve.\n'
    'Vi du: "OD 25.00 +/-0.02"\n'
    'hoac "Chieu dai 50.0 +0/-0.05"')
ws.cell(ck_start, CK[3][0]).comment = cmt(
    'Ghi gia tri do thuc te va ten dung cu.\n'
    'Vi du: "24.99 / Micrometer #M-015"')

# Blank rows with auto-number
blank_count = 10
for j in range(blank_count):
    is_last = (j == blank_count - 1)
    ws.row_dimensions[r].height = 20.0
    _borders_table_row(ws, r, nc, CK, is_last)
    bnum = len(items) + j + 1
    even = (bnum % 2 == 0)
    ifill = fi(P['near']) if even else fi(P['white'])
    zfills = [fi(P['fog']), fi(P['fog']), ifill, fi(P['fog']), fi(P['white']), fi(P['white'])]
    for (s, e), zf in zip(CK, zfills):
        for col in range(s, e+1):
            ws.cell(r, col).fill = zf
    for s, e in CK:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    ws.cell(r, CK[0][0]).value = f'=ROW()-ROW($A${ck_start})+1'
    ws.cell(r, CK[0][0]).font = fo(8, True, 'mid')
    ws.cell(r, CK[0][0]).alignment = AC
    ws.cell(r, CK[4][0]).font = fo(8, False, 'dk')
    ws.cell(r, CK[4][0]).alignment = AC
    ws.cell(r, CK[5][0]).font = fo(8, False, 'dk')
    ws.cell(r, CK[5][0]).alignment = AC
    r += 1
ck_end = r - 1

r = spacer(ws, r, nc)

# === S3: LOT DISPOSITION ===
r = sect_header(ws, r, nc, '  3.  LOT DISPOSITION')
ws.cell(r-1, 1).comment = cmt(
    'Quyet dinh lo hang.\n'
    'Sau khi kiem tra tat ca dac tinh, quyet dinh cho toan lo.')
disp_r = r
r = data_pair(ws, r, nc, PP, 'Accept # / Reject #',
              'Lot Decision: ACCEPT / REJECT / SORT', first=True, last=True)
ws.cell(disp_r, PP[1][0]).comment = cmt(
    'Ghi so luong dat va so luong loi.\n'
    'Vi du: Accept=12, Reject=1')
ws.cell(disp_r, PP[3][0]).comment = cmt(
    'Quyet dinh lo hang:\n'
    'ACCEPT = chap nhan lo, chuyen tiep.\n'
    'REJECT = tu choi lo, lap NCR (FRM-651).\n'
    'SORT = kiem tra 100%% de phan loai.')

r = spacer(ws, r, nc)

# === S4: APPROVAL ===
r = sect_header(ws, r, nc, '  4.  APPROVAL')
r = sig_block(ws, r, nc, ['Inspector', 'QA Manager'], zones['sign2'])
r = spacer(ws, r, nc)

# NOTICE
r = notice_bar(ws, r, nc,
    'NOTICE \u2014 One form per lot inspection. '
    'If REJECT or SORT, raise NCR (FRM-651). '
    'Ref: SOP-605. Retain: 10 years.')

# DV + CF
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

dv = DataValidation(type='list', formula1='=LISTS!$A$2:$A$5', allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
rcl = get_column_letter(CK[4][0])
for rr in range(ck_start, ck_end + 1):
    dv.add(f'{rcl}{rr}')

rng = f'{rcl}{ck_start}:{rcl}{ck_end}'
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"PASS"'],
    fill=fi('C6EFCE'), font=Font(color='006100')))
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"HOLD"'],
    fill=fi('FFEB9C'), font=Font(color='9C6500')))
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"FAIL"'],
    fill=fi('FFC7CE'), font=Font(color='9C0006')))

# Lot Decision DV
dv2 = DataValidation(type='list', formula1='"ACCEPT,REJECT,SORT"', allow_blank=True)
dv2.showErrorMessage = True
ws.add_data_validation(dv2)
dv2.add(f'{get_column_letter(PP[3][0])}{disp_r}')

# Print
ws.page_setup.paperSize = 9
ws.page_setup.orientation = 'landscape'
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# Remove other sheets
for name in list(wb.sheetnames):
    if name not in ['FRM_621', 'LISTS']:
        del wb[name]

# Save
OUT = "C:/Users/TEST4/Desktop/FRM-621_SAMPLE.xlsx"
wb.save(OUT)
sha = hashlib.sha256(open(OUT, 'rb').read()).hexdigest()
print(f"FRM-621 SAMPLE: SHA={sha[:12]}")
print(f"Output: {OUT}")
print()
print("CELL COMMENTS (hover tam giac do o goc cell):")
print("  Section 1 header: giai thich y nghia section")
print("  Job/WO input: 'Ghi so Job hoac Work Order tu Job Packet'")
print("  Part/Rev input: 'Ghi ma chi tiet va phien ban ban ve'")
print("  Lot/Sample input: 'Ghi co lo va co mau theo bang AQL'")
print("  AQL Level input: 'Ghi muc AQL, tra bang ISO 2859-1'")
print("  Col header 'Cat.': giai thich cac loai QUA/TEC/OPS")
print("  Col header 'Characteristic': huong dan ghi dac tinh")
print("  Col header 'Method': huong dan ghi dung cu + ID")
print("  Col header 'Result': giai thich PASS/FAIL/HOLD/NA")
print("  Col header 'Inspector': yeu cau nang luc")
print("  Data row 1 'Characteristic': vi du cu the")
print("  Data row 1 'Method': vi du cu the")
print("  Lot Decision inputs: giai thich ACCEPT/REJECT/SORT")
print("  Section 3 header: giai thich y nghia section")
print("  Blank rows: auto-number formula =ROW()-ROW($A$start)+1")
