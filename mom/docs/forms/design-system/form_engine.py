# -*- coding: utf-8 -*-
"""
HESEM PRECISION FORM ENGINE v2.1
=================================
Generates any QMS form from a spec dict.
Design System v2.1 rules baked in:
  - wrap_text=True, no indent, ALL-cell borders, logo
  - 6-zone checklist (no Ref column)
  - Blue=section outline, Silver=internal
  - Spacer between sections
  - Lean: min ref fields, no unnecessary sections
  - Extra blank rows for data-heavy forms (fit page, no insert needed)
"""
import openpyxl, copy, math, hashlib, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
import openpyxl.worksheet.properties

# =====================================================================
# DESIGN SYSTEM CONSTANTS
# =====================================================================
P = {
    'primary':'1B3A6B','accent':'0078D4','accent2':'005A9E',
    'ice':'EFF6FF','fog':'F1F5F9','white':'FFFFFF',
    'near':'F8FAFC','silver':'E2E8F0',
    'dk':'0F172A','mid':'334155','muted':'64748B','plc':'CBD5E1',
    'tec_bg':'DCFCE7','tec_fg':'14532D','qua_bg':'FEF9C3','qua_fg':'713F12',
    'cap_bg':'FCE7F3','cap_fg':'831843','con_bg':'F3E8FF','con_fg':'4C1D95',
    'com_bg':'DBEAFE','com_fg':'1E3A8A',
}
CAT_COLORS = {
    'COM':(P['com_bg'],P['com_fg']), 'TEC':(P['tec_bg'],P['tec_fg']),
    'QUA':(P['qua_bg'],P['qua_fg']), 'CAP':(P['cap_bg'],P['cap_fg']),
    'CON':(P['con_bg'],P['con_fg']),
    'SET':(P['tec_bg'],P['tec_fg']), 'TOOL':(P['cap_bg'],P['cap_fg']),
    'GAGE':(P['con_bg'],P['con_fg']), 'OPS':(P['com_bg'],P['com_fg']),
    'DIS':(P['cap_bg'],P['cap_fg']), 'DOC':(P['com_bg'],P['com_fg']),
    'SAF':(P['cap_bg'],P['cap_fg']), 'ENV':(P['tec_bg'],P['tec_fg']),
    'HR':(P['com_bg'],P['com_fg']), 'MGT':(P['con_bg'],P['con_fg']),
}

BLUE = Side(style='thin', color=P['accent'])
SILVER = Side(style='thin', color=P['silver'])
NONE_S = Side()
EMU = 9525
LOGO_PATH = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau/00-FORM-DESIGN-SYSTEM/hesem-logo.png"
LOGO_ASPECT = 3.46
TMPL_PATH = "C:/Users/TEST4/Desktop/frm-000-master-template.xlsx"

from openpyxl.comments import Comment as XlComment

def fi(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fo(sz, bold=False, c='dk'): return Font(name='Segoe UI', size=sz, bold=bold, color=P[c])

def cmt(vn_name, *details):
    """Create cell comment: line1=Vietnamese name, line2+=description/examples"""
    text = vn_name
    for d in details:
        text += '\n' + d
    c = XlComment(text, 'QMS')
    c.width = 280
    c.height = max(100, 30 * (1 + len(details)))
    return c
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AT = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Format configs
FORMATS = {
    'A4P': {'ncols': 40, 'logo_cols': 8, 'paper': 9, 'orient': 'portrait'},
    'A4L': {'ncols': 56, 'logo_cols': 11, 'paper': 9, 'orient': 'landscape'},
    'A3P': {'ncols': 56, 'logo_cols': 11, 'paper': 8, 'orient': 'portrait'},
    'A3L': {'ncols': 82, 'logo_cols': 16, 'paper': 8, 'orient': 'landscape'},
}

# Zone definitions per format
def get_zones(nc):
    if nc == 40:
        return {
            'pair': [(1,7),(8,20),(21,27),(28,nc)],
            'check': [(1,2),(3,5),(6,21),(22,31),(32,35),(36,nc)],
            'sign2': [(1,20),(21,nc)],
            'sign3': [(1,13),(14,26),(27,nc)],
        }
    elif nc == 56:
        return {
            'pair': [(1,10),(11,28),(29,38),(39,nc)],
            'check': [(1,3),(4,7),(8,29),(30,43),(44,49),(50,nc)],
            'sign2': [(1,28),(29,nc)],
            'sign3': [(1,18),(19,37),(38,nc)],
        }
    else:  # 82
        return {
            'pair': [(1,14),(15,41),(42,55),(56,nc)],
            'check': [(1,4),(5,10),(11,42),(43,62),(63,70),(71,nc)],
            'sign2': [(1,41),(42,nc)],
            'sign3': [(1,27),(28,54),(55,nc)],
        }

# =====================================================================
# CORE BORDER FUNCTIONS (ALL cells)
# =====================================================================

def _borders_section_row(ws, r, nc, zones, is_first, is_last):
    top = BLUE if is_first else NONE_S
    bot = BLUE if is_last else SILVER
    zone_ends = set(e for s, e in zones[:-1])
    for col in range(1, nc+1):
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == nc else (SILVER if col in zone_ends else NONE_S)
        ws.cell(r, col).border = Border(left=left, right=right, top=top, bottom=bot)

def _borders_table_row(ws, r, nc, zones, is_last):
    bot = BLUE if is_last else SILVER
    zone_ends = set(e for s, e in zones[:-1])
    for col in range(1, nc+1):
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == nc else (SILVER if col in zone_ends else NONE_S)
        ws.cell(r, col).border = Border(left=left, right=right, bottom=bot)

# =====================================================================
# BLOCK BUILDERS
# =====================================================================

def embed_logo(ws, logo_cols):
    zone_w = logo_cols * 19; zone_h = 5 * 20
    lw = math.floor(zone_w * 0.78); lh = math.floor(lw / LOGO_ASPECT)
    if lh > zone_h * 0.78:
        lh = math.floor(zone_h * 0.78); lw = math.floor(lh * LOGO_ASPECT)
    xo = math.floor((zone_w - lw) / 2) * EMU
    yo = math.floor((zone_h - lh) / 2) * EMU
    img = XlImage(LOGO_PATH)
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=0, colOff=xo, row=1, rowOff=yo),
        ext=XDRPositiveSize2D(cx=lw*EMU, cy=lh*EMU))
    ws.add_image(img)

def build_header(ws, src_ws, nc, title, code, owner, approved):
    ws.row_dimensions[1].hidden = True
    for row in range(2, 7):
        ws.row_dimensions[row].height = 15.0
        for col in range(1, nc+1):
            s = src_ws.cell(row, col); d = ws.cell(row, col)
            d.value = s.value; d.font = copy.copy(s.font)
            d.fill = copy.copy(s.fill); d.alignment = copy.copy(s.alignment)
            d.border = copy.copy(s.border)
    for mc in src_ws.merged_cells.ranges:
        ws.merge_cells(str(mc))
    # Find title and meta cells based on format
    if nc == 40:
        tc, mc_start, mv_start = 9, 31, 35
    elif nc == 56:
        tc, mc_start, mv_start = 12, 43, 49
    else:
        tc, mc_start, mv_start = 17, 63, 71
    ws.cell(2, tc).value = title
    ws.cell(2, mv_start).value = code
    ws.cell(3, mv_start).value = 'Rev.01'
    ws.cell(4, mv_start).value = '2026-03-23'
    ws.cell(5, mv_start).value = owner
    ws.cell(6, mv_start).value = approved
    meta_align = Alignment(horizontal='left', vertical='center', indent=0, wrap_text=True)
    for hr in range(2, 7):
        c = ws.cell(hr, mv_start)
        if c.value is not None: c.alignment = meta_align

def spacer(ws, r, nc):
    ws.row_dimensions[r].height = 4.0
    for col in range(1, nc+1):
        ws.cell(r, col).fill = fi(P['white']); ws.cell(r, col).border = Border()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    return r + 1

def sect_header(ws, r, nc, text, sect_cmt=None):
    ws.row_dimensions[r].height = 17.0
    for col in range(1, nc+1):
        ws.cell(r, col).fill = fi(P['ice'])
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == nc else NONE_S
        ws.cell(r, col).border = Border(left=left, right=right, top=BLUE, bottom=BLUE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(r, 1); c.value = text; c.font = fo(9, True, 'primary')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    if sect_cmt: c.comment = sect_cmt
    return r + 1

def data_pair(ws, r, nc, zones, lL, lR, first=False, last=False, cmtL=None, cmtR=None):
    """4-zone data pair. cmtL/cmtR = Comment objects for left/right LABELS."""
    ws.row_dimensions[r].height = 20.0
    _borders_section_row(ws, r, nc, zones, first, last)
    fills = [fi(P['fog']), fi(P['white']), fi(P['fog']), fi(P['white'])]
    for (s, e), zf in zip(zones, fills):
        for col in range(s, e+1): ws.cell(r, col).fill = zf
    for s, e in zones:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    c = ws.cell(r, zones[0][0]); c.value = lL; c.font = fo(8, True, 'mid'); c.alignment = AL
    if cmtL: c.comment = cmtL
    ws.cell(r, zones[1][0]).font = fo(8, False, 'dk'); ws.cell(r, zones[1][0]).alignment = AL
    c = ws.cell(r, zones[2][0]); c.value = lR; c.font = fo(8, True, 'mid'); c.alignment = AL
    if cmtR: c.comment = cmtR
    ws.cell(r, zones[3][0]).font = fo(8, False, 'dk'); ws.cell(r, zones[3][0]).alignment = AL
    return r + 1

def col_header(ws, r, nc, zones, headers):
    ws.row_dimensions[r].height = 17.0
    _borders_section_row(ws, r, nc, zones, True, False)
    for col in range(1, nc+1): ws.cell(r, col).fill = fi(P['accent2'])
    for (s, e), txt in zip(zones, headers):
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
        c = ws.cell(r, s); c.value = txt
        c.font = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF'); c.alignment = AC
    return r + 1

def check_row(ws, r, nc, zones, num, cat, item, criteria, last=False, ck_start=None, item_cmt=None):
    ws.row_dimensions[r].height = 20.0
    _borders_table_row(ws, r, nc, zones, last)
    even = (num % 2 == 0)
    ifill = fi(P['near']) if even else fi(P['white'])
    catc = CAT_COLORS.get(cat)
    zfills = [fi(P['fog']), fi(catc[0]) if catc else fi(P['fog']),
              ifill, fi(P['fog']), fi(P['white']), fi(P['white'])]
    for (s, e), zf in zip(zones, zfills):
        for col in range(s, e+1): ws.cell(r, col).fill = zf
    for s, e in zones:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    # Auto-number formula (not hardcoded int)
    anchor = ck_start if ck_start else r
    ws.cell(r, zones[0][0]).value = f'=ROW()-ROW($A${anchor})+1'
    ws.cell(r, zones[0][0]).font = fo(8, True, 'mid'); ws.cell(r, zones[0][0]).alignment = AC
    c = ws.cell(r, zones[1][0]); c.value = cat; c.alignment = AC
    c.font = Font(name='Segoe UI', size=8, bold=True, color=catc[1]) if catc else fo(8, True, 'mid')
    ic = ws.cell(r, zones[2][0]); ic.value = item; ic.font = fo(8, False, 'dk'); ic.alignment = AL
    if item_cmt: ic.comment = item_cmt
    crit_col = zones[3][0]
    ws.cell(r, crit_col).value = criteria; ws.cell(r, crit_col).font = fo(7, False, 'muted'); ws.cell(r, crit_col).alignment = AL
    ws.cell(r, zones[4][0]).font = fo(8, False, 'dk'); ws.cell(r, zones[4][0]).alignment = AC
    ws.cell(r, zones[5][0]).font = fo(8, False, 'dk'); ws.cell(r, zones[5][0]).alignment = AC
    return r + 1

def blank_check_row(ws, r, nc, zones, num, last=False, ck_start=None):
    """Blank checklist row for user to fill — keeps page height consistent."""
    ws.row_dimensions[r].height = 20.0
    _borders_table_row(ws, r, nc, zones, last)
    even = (num % 2 == 0)
    ifill = fi(P['near']) if even else fi(P['white'])
    zfills = [fi(P['fog']), fi(P['fog']), ifill, fi(P['fog']), fi(P['white']), fi(P['white'])]
    for (s, e), zf in zip(zones, zfills):
        for col in range(s, e+1): ws.cell(r, col).fill = zf
    for s, e in zones:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    # Auto-number formula
    anchor = ck_start if ck_start else r
    ws.cell(r, zones[0][0]).value = f'=ROW()-ROW($A${anchor})+1'
    ws.cell(r, zones[0][0]).font = fo(8, True, 'mid'); ws.cell(r, zones[0][0]).alignment = AC
    ws.cell(r, zones[4][0]).font = fo(8, False, 'dk'); ws.cell(r, zones[4][0]).alignment = AC
    ws.cell(r, zones[5][0]).font = fo(8, False, 'dk'); ws.cell(r, zones[5][0]).alignment = AC
    return r + 1

def textarea(ws, r, nc, label, nrows=3, first=False, last=False):
    ws.row_dimensions[r].height = 17.0
    for col in range(1, nc+1):
        ws.cell(r, col).fill = fi(P['fog'])
        left = BLUE if col == 1 else NONE_S; right = BLUE if col == nc else NONE_S
        top = BLUE if first else NONE_S
        ws.cell(r, col).border = Border(left=left, right=right, top=top, bottom=SILVER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(r, 1); c.value = label; c.font = fo(8, True, 'mid'); c.alignment = AL
    r += 1
    for i in range(nrows):
        is_last = last and (i == nrows - 1)
        ws.row_dimensions[r].height = 20.0
        bot = BLUE if is_last else SILVER
        for col in range(1, nc+1):
            ws.cell(r, col).fill = fi(P['white'])
            left = BLUE if col == 1 else NONE_S; right = BLUE if col == nc else NONE_S
            ws.cell(r, col).border = Border(left=left, right=right, bottom=bot)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
        if i == 0:
            ws.cell(r, 1).font = fo(8, False, 'dk'); ws.cell(r, 1).alignment = AT
        r += 1
    return r

def sig_block(ws, r, nc, labels, sig_zones):
    ws.row_dimensions[r].height = 17.0
    for col in range(1, nc+1): ws.cell(r, col).fill = fi(P['fog'])
    for s, e in sig_zones:
        for col in range(s, e+1):
            ws.cell(r, col).border = Border(
                left=BLUE if col == s else NONE_S, right=BLUE if col == e else NONE_S,
                top=BLUE, bottom=BLUE)
    for (s, e), lab in zip(sig_zones, labels):
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
        ws.cell(r, s).value = lab; ws.cell(r, s).font = fo(8, True, 'mid'); ws.cell(r, s).alignment = AC
    r += 1
    for sr in range(r, r+3): ws.row_dimensions[sr].height = 26.0
    for s, e in sig_zones:
        for sr in range(r, r+3):
            for col in range(s, e+1):
                ws.cell(sr, col).fill = fi(P['white'])
                ws.cell(sr, col).border = Border(
                    left=BLUE if col == s else NONE_S, right=BLUE if col == e else NONE_S,
                    top=BLUE, bottom=BLUE)
        ws.merge_cells(start_row=r, start_column=s, end_row=r+2, end_column=e)
        c = ws.cell(r, s); c.value = 'Name  /  Signature  /  Date'
        c.font = Font(name='Segoe UI', size=8, color=P['plc'])
        c.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
    return r + 3

def notice_bar(ws, r, nc, text):
    ws.row_dimensions[r].height = 24.0
    for col in range(1, nc+1):
        ws.cell(r, col).fill = fi(P['ice'])
        ws.cell(r, col).border = Border(
            left=BLUE if col == 1 else NONE_S, right=BLUE if col == nc else NONE_S,
            top=BLUE, bottom=BLUE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(r, 1); c.value = text
    c.font = Font(name='Segoe UI', size=6, color=P['mid'])
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    return r + 1

# =====================================================================
# MAIN FORM GENERATOR
# =====================================================================

def generate_form(spec, output_path):
    """
    Generate a form from a spec dict.

    spec = {
        'code': 'FRM-511',
        'title': 'SETUP AND FIRST-PIECE RECORD',
        'format': 'A4P',  # A4P/A4L/A3P/A3L
        'owner': 'Production / QA',
        'approved': 'Production Director',
        'ref_fields': [('Job / WO', 'Part No. / Revision'), ('Date', 'Prepared by')],
        'sections': [
            {
                'title': '1.  CHECKLIST SECTION',
                'type': 'checklist',
                'headers': ['#','Cat.','Check Item','Acceptance Criteria','Result','Owner'],
                'items': [(1,'CAT','Item text','Criteria text'), ...],
                'blank_rows': 3,  # extra blank rows for user
            },
            {
                'title': '2.  TEXT SECTION',
                'type': 'textarea',
                'label': 'DESCRIPTION',
                'rows': 4,
            },
            {
                'title': '3.  DISPOSITION',
                'type': 'pairs',
                'fields': [('Disposition','Stakeholders Notified?'), ...],
            },
        ],
        'approval': ['Prepared by', 'Approved by'],  # 2-col or 3-col
        'notice': 'NOTICE text...',
        'dv_result_col': 32,  # column for Result dropdown (or None)
        'cf_result_col': 32,  # column for Result CF (or None)
    }
    """
    fmt = FORMATS[spec['format']]
    nc = fmt['ncols']
    zones = get_zones(nc)

    wb = openpyxl.load_workbook(TMPL_PATH)
    # Select correct master sheet
    src_name = {'A4P':'01-MASTER-A4-PORTRAIT','A4L':'02-MASTER-A4-LANDSCAPE',
                'A3P':'03-MASTER-A3-PORTRAIT','A3L':'04-MASTER-A3-LANDSCAPE'}[spec['format']]
    src_ws = wb[src_name]

    ws = wb.create_sheet(spec['code'].replace('-','_'), 0)
    for i in range(1, nc+1):
        ws.column_dimensions[get_column_letter(i)].width = 2.375

    build_header(ws, src_ws, nc, spec['title'], spec['code'], spec['owner'], spec['approved'])
    embed_logo(ws, fmt['logo_cols'])

    r = 7
    r = spacer(ws, r, nc)

    # Reference section
    ref_sect_title = spec.get('ref_title', '  1.  REFERENCE INFORMATION')
    r = sect_header(ws, r, nc, ref_sect_title)
    if spec.get('ref_sect_cmt'):
        ws.cell(r-1, 1).comment = spec['ref_sect_cmt']
    ref = spec['ref_fields']
    ref_cmts = spec.get('ref_comments', [None] * len(ref))
    for i, (lL, lR) in enumerate(ref):
        rc = ref_cmts[i] if i < len(ref_cmts) else None
        cmtL = rc[0] if rc and len(rc) > 0 else None
        cmtR = rc[1] if rc and len(rc) > 1 else None
        r = data_pair(ws, r, nc, zones['pair'], lL, lR, first=(i==0), last=(i==len(ref)-1), cmtL=cmtL, cmtR=cmtR)
    r = spacer(ws, r, nc)

    # Content sections
    sect_num = 2
    ck_start = ck_end = None

    for section in spec.get('sections', []):
        r = sect_header(ws, r, nc, f'  {sect_num}.  {section["title"]}', sect_cmt=section.get('sect_cmt'))

        if section['type'] == 'checklist':
            hdrs = section.get('headers', ['#','Cat.','Check Item','Acceptance Criteria','Result','Owner'])
            # Allow custom check zones if headers count != standard 6
            if len(hdrs) != len(zones['check']):
                # Auto-generate equal-width zones for custom column count
                n = len(hdrs)
                w = nc // n
                ck_zones = []
                for ci in range(n):
                    s = ci * w + 1
                    e = (ci + 1) * w if ci < n - 1 else nc
                    ck_zones.append((s, e))
            else:
                ck_zones = zones['check']
            r = col_header(ws, r, nc, ck_zones, hdrs)
            hdr_r = r - 1
            # Add comments on column headers if provided
            col_cmts = section.get('col_comments', {})
            for idx, cm in col_cmts.items():
                if idx < len(ck_zones):
                    ws.cell(hdr_r, ck_zones[idx][0]).comment = cm
            ck_start = r
            items = section['items']
            item_cmts = section.get('item_comments', {})
            total_rows = len(items) + section.get('blank_rows', 0)
            for i, (num, cat, item, criteria) in enumerate(items):
                is_last = (i == len(items) - 1) and section.get('blank_rows', 0) == 0
                ic = item_cmts.get(i)
                r = check_row(ws, r, nc, ck_zones, num, cat, item, criteria, last=is_last, ck_start=ck_start, item_cmt=ic)
            for j in range(section.get('blank_rows', 0)):
                bnum = len(items) + j + 1
                is_last = (j == section.get('blank_rows', 0) - 1)
                r = blank_check_row(ws, r, nc, ck_zones, bnum, last=is_last, ck_start=ck_start)
            ck_end = r - 1

        elif section['type'] == 'textarea':
            labels = section.get('labels', [section.get('label', 'DESCRIPTION')])
            for k, lbl in enumerate(labels):
                r = textarea(ws, r, nc, lbl,
                            nrows=section.get('rows', 3),
                            first=(k==0),
                            last=(k==len(labels)-1))

        elif section['type'] == 'pairs':
            fields = section['fields']
            pair_cmts = section.get('pair_comments', [None] * len(fields))
            for i, (lL, lR) in enumerate(fields):
                pc = pair_cmts[i] if i < len(pair_cmts) else None
                cmtL = pc[0] if pc and len(pc) > 0 else None
                cmtR = pc[1] if pc and len(pc) > 1 else None
                r = data_pair(ws, r, nc, zones['pair'], lL, lR, first=(i==0), last=(i==len(fields)-1), cmtL=cmtL, cmtR=cmtR)

        r = spacer(ws, r, nc)
        sect_num += 1

    # Approval
    approval_labels = spec.get('approval', ['Prepared by', 'Approved by'])
    r = sect_header(ws, r, nc, f'  {sect_num}.  APPROVAL')
    sig_type = 'sign3' if len(approval_labels) >= 3 else 'sign2'
    r = sig_block(ws, r, nc, approval_labels, zones[sig_type])
    r = spacer(ws, r, nc)

    # Notice
    r = notice_bar(ws, r, nc, spec.get('notice',
        'NOTICE  \u2014  Enter data only in white input cells. Do not add, delete or resize columns/rows.'))

    # Data Validation + Conditional Formatting
    if ck_start and ck_end and spec.get('dv_result_col'):
        rc = spec['dv_result_col']
        rcl = get_column_letter(rc)
        dv_formula = spec.get('dv_formula', '=LISTS!$A$2:$A$5')
        dv = DataValidation(type="list", formula1=dv_formula, allow_blank=True)
        dv.showErrorMessage = True; ws.add_data_validation(dv)
        for rr in range(ck_start, ck_end+1):
            dv.add(f"{rcl}{rr}")

        rng = f"{rcl}{ck_start}:{rcl}{ck_end}"
        pass_vals = spec.get('cf_pass', ['"PASS"'])
        hold_vals = spec.get('cf_hold', ['"HOLD"'])
        fail_vals = spec.get('cf_fail', ['"FAIL"'])
        for v in pass_vals:
            ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[v],
                fill=fi('C6EFCE'), font=Font(color='006100')))
        for v in hold_vals:
            ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[v],
                fill=fi('FFEB9C'), font=Font(color='9C6500')))
        for v in fail_vals:
            ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[v],
                fill=fi('FFC7CE'), font=Font(color='9C0006')))

    # Print settings
    ws.page_setup.paperSize = fmt['paper']
    ws.page_setup.orientation = fmt['orient']
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0

    # Remove other sheets
    for name in list(wb.sheetnames):
        if name not in [ws.title, 'LISTS']:
            del wb[name]

    # Save
    wb.save(output_path)
    sha = hashlib.sha256(open(output_path, 'rb').read()).hexdigest()
    return sha
