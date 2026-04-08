"""
HESEM QMS Form Upgrade Library v1.0
====================================
Provides reusable functions for upgrading Excel forms while
preserving the master template design system exactly.

Design system rules (immutable):
- Grid: 40 cols A4P (2.33 char-width), 56 cols A4L
- NEVER add/remove/resize columns
- Colors: see COLORS dict below
- Font: Segoe UI throughout
- Row heights: spacer=4pt, section=17pt, colheader=17pt, data=20pt, sig=26pt, notice=24pt
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import hashlib
import json
import shutil
import os

# ============================================================================
# DESIGN SYSTEM CONSTANTS (from frm-000-master-template.xlsx 00-DESIGN-NOTES)
# ============================================================================

COLORS = {
    'PRIMARY':    '1B3A6B',   # Deep steel blue - title text, meta form-code
    'ACCENT':     '0078D4',   # Process blue - ALL borders, subtitle
    'ACCENT2':    '005A9E',   # Deeper blue - column header bg
    'ICE':        'EFF6FF',   # Ice blue - section header bg, notice bar bg
    'FOG':        'F1F5F9',   # Light grey - label cells bg
    'WHITE':      'FFFFFF',   # Pure white - input cells, sig boxes
    'NEAR_WHITE': 'F8FAFC',   # Near-white - even-row stripe
    'SILVER':     'E2E8F0',   # Silver - internal separators
    'DARK_TXT':   '0F172A',   # Near-black - body text
    'LABEL_TXT':  '334155',   # Dark slate - label text
    'SUBTITLE':   '0078D4',   # Process blue - subtitle text
    'MUTED_TXT':  '64748B',   # Muted - company line text
    'PLACEHOLDER':'CBD5E1',   # Light gray - sig placeholder text
    'SEC_HEADER': '1B3A6B',   # Section header text color
    # Category colors (checklist)
    'CAT_COM':    'D8E6FF',   # Commercial - light blue
    'CAT_TEC':    'D7F0E2',   # Technical - light green
    'CAT_QUA':    'FFF3C4',   # Quality - light yellow
    'CAT_CAP':    'F0D7E8',   # Capacity - light pink
    'CAT_CON':    'E0D9FF',   # Condition - light purple
    'CAT_DOC':    'ECECEC',   # Document - light gray
    'CAT_REL':    'D7F0E2',   # Release - light green
    'CAT_CNC':    'D8E6FF',   # CNC-specific - light blue
    # Conditional formatting
    'CF_RED':     'FFCCCC',   # FAIL/CRITICAL bg
    'CF_AMBER':   'FFF3C4',   # HOLD/WARNING bg
    'CF_GREEN':   'D7F0E2',   # PASS bg
    'CF_RED_FONT':'CC0000',   # FAIL text
}

# Borders
THIN_ACCENT = Side(style='thin', color=COLORS['ACCENT'])
THIN_SILVER = Side(style='thin', color=COLORS['SILVER'])
BORDER_OUTLINE = Border(left=THIN_ACCENT, right=THIN_ACCENT, top=THIN_ACCENT, bottom=THIN_ACCENT)
BORDER_INNER = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_SILVER, bottom=THIN_SILVER)
BORDER_FULL = Border(left=THIN_ACCENT, right=THIN_ACCENT, top=THIN_ACCENT, bottom=THIN_ACCENT)
NO_BORDER = Border()

# Fonts
FONT_TITLE      = Font(name='Segoe UI', size=15, bold=True, color=COLORS['PRIMARY'])
FONT_SUBTITLE   = Font(name='Segoe UI', size=8, color=COLORS['SUBTITLE'])
FONT_COMPANY    = Font(name='Segoe UI', size=7, color=COLORS['MUTED_TXT'])
FONT_META_LABEL = Font(name='Segoe UI', size=8, bold=True, color=COLORS['LABEL_TXT'])
FONT_META_VALUE = Font(name='Segoe UI', size=8, color=COLORS['DARK_TXT'])
FONT_META_CODE  = Font(name='Segoe UI', size=8, bold=True, color=COLORS['PRIMARY'])
FONT_SEC_HEADER = Font(name='Segoe UI', size=9, bold=True, color=COLORS['SEC_HEADER'])
FONT_COL_HEADER = Font(name='Segoe UI', size=9, bold=True, color=COLORS['WHITE'])
FONT_LABEL      = Font(name='Segoe UI', size=8, bold=True, color=COLORS['LABEL_TXT'])
FONT_INPUT      = Font(name='Segoe UI', size=8, color=COLORS['DARK_TXT'])
FONT_BODY       = Font(name='Segoe UI', size=8, color=COLORS['DARK_TXT'])
FONT_CAT        = Font(name='Segoe UI', size=8, bold=True, color='475569')
FONT_NOTICE     = Font(name='Segoe UI', size=7, color=COLORS['LABEL_TXT'])
FONT_SIG_PH     = Font(name='Segoe UI', size=8, color=COLORS['PLACEHOLDER'])
FONT_NUM        = Font(name='Segoe UI', size=8, bold=True, color=COLORS['LABEL_TXT'])

# Fills
FILL_ICE        = PatternFill(start_color=COLORS['ICE'], end_color=COLORS['ICE'], fill_type='solid')
FILL_FOG        = PatternFill(start_color=COLORS['FOG'], end_color=COLORS['FOG'], fill_type='solid')
FILL_WHITE      = PatternFill(start_color=COLORS['WHITE'], end_color=COLORS['WHITE'], fill_type='solid')
FILL_NEAR_WHITE = PatternFill(start_color=COLORS['NEAR_WHITE'], end_color=COLORS['NEAR_WHITE'], fill_type='solid')
FILL_COL_HDR    = PatternFill(start_color=COLORS['ACCENT2'], end_color=COLORS['ACCENT2'], fill_type='solid')
FILL_CF_RED     = PatternFill(start_color=COLORS['CF_RED'], end_color=COLORS['CF_RED'], fill_type='solid')
FILL_CF_AMBER   = PatternFill(start_color=COLORS['CF_AMBER'], end_color=COLORS['CF_AMBER'], fill_type='solid')
FILL_CF_GREEN   = PatternFill(start_color=COLORS['CF_GREEN'], end_color=COLORS['CF_GREEN'], fill_type='solid')

# Alignments
ALIGN_CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=False)
ALIGN_LEFT      = Alignment(horizontal='left', vertical='center', wrap_text=False)
ALIGN_LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_SIG       = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
ALIGN_NOTICE    = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)

# Row heights
H_HEADER  = 15.0
H_SPACER  = 4.0
H_SECTION = 17.0
H_COLHDR  = 17.0
H_DATA    = 20.0
H_SIG     = 26.0
H_NOTICE  = 24.0


# ============================================================================
# CELL STYLING FUNCTIONS
# ============================================================================

def style_section_header(ws, row, ncols, text, section_num=None):
    """Apply section header style: ICE bg, 9pt bold PRIMARY text, full-width merge."""
    cell = ws.cell(row=row, column=1)
    prefix = f"  {section_num}.  " if section_num else "  "
    cell.value = f"{prefix}{text}"
    cell.font = FONT_SEC_HEADER
    cell.fill = FILL_ICE
    cell.alignment = ALIGN_LEFT
    cell.border = BORDER_FULL
    ws.row_dimensions[row].height = H_SECTION
    # Merge full width
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def style_col_header(ws, row, col_start, col_end, text):
    """Apply column header style: ACCENT2 bg, 9pt bold WHITE text."""
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = FONT_COL_HEADER
    cell.fill = FILL_COL_HDR
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_FULL
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)


def style_label_cell(ws, row, col_start, col_end, text):
    """Apply label cell style: FOG bg, 8pt bold LABEL_TXT."""
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = FONT_LABEL
    cell.fill = FILL_FOG
    cell.alignment = ALIGN_LEFT
    cell.border = BORDER_FULL
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)


def style_input_cell(ws, row, col_start, col_end, value=None, mandatory=False):
    """Apply input cell style: WHITE bg, 8pt DARK_TXT. If mandatory, use thicker border."""
    cell = ws.cell(row=row, column=col_start)
    if value is not None:
        cell.value = value
    cell.font = FONT_INPUT
    cell.fill = FILL_WHITE
    cell.alignment = ALIGN_LEFT
    if mandatory:
        thick_accent = Side(style='medium', color=COLORS['ACCENT'])
        cell.border = Border(left=thick_accent, right=thick_accent, top=thick_accent, bottom=thick_accent)
    else:
        cell.border = BORDER_FULL
    cell.protection = Protection(locked=False)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)


def style_spacer_row(ws, row, ncols):
    """Apply spacer row: 4pt, white, no borders, full merge."""
    ws.row_dimensions[row].height = H_SPACER
    cell = ws.cell(row=row, column=1)
    cell.fill = FILL_WHITE
    cell.border = NO_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def style_notice_bar(ws, row, ncols, text):
    """Apply notice bar: ICE bg, 7pt text, wrap, indent=2."""
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = FONT_NOTICE
    cell.fill = FILL_ICE
    cell.alignment = ALIGN_NOTICE
    cell.border = BORDER_FULL
    ws.row_dimensions[row].height = H_NOTICE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def clone_checklist_row(ws, source_row, target_row, ncols=40):
    """Clone merge pattern and styling from an existing checklist row to a new row.

    Standard checklist 7-zone pattern (A4P):
    # (A:B) | Cat (C:E) | Item (F:S) | Criteria (T:AE) | Result (AF:AI) | Owner (AJ:AL) | Ref (AM:AN)
    """
    ws.row_dimensions[target_row].height = H_DATA

    # Copy merge pattern from source row
    merges_to_add = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row == source_row and m.max_row == source_row:
            merges_to_add.append((target_row, m.min_col, target_row, m.max_col))

    for sr, sc, er, ec in merges_to_add:
        ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)

    # Copy cell styles from source row
    for col in range(1, ncols + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        tgt.font = copy(src.font)
        tgt.fill = copy(src.fill)
        tgt.alignment = copy(src.alignment)
        tgt.border = copy(src.border)
        tgt.number_format = src.number_format
        tgt.protection = copy(src.protection)


def clone_data_row_pair(ws, source_row, target_row, ncols=40):
    """Clone merge pattern and styling from a data row (2-pair label+input layout)."""
    ws.row_dimensions[target_row].height = H_DATA

    merges_to_add = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row == source_row and m.max_row == source_row:
            merges_to_add.append((target_row, m.min_col, target_row, m.max_col))

    for sr, sc, er, ec in merges_to_add:
        ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)

    for col in range(1, ncols + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        tgt.font = copy(src.font)
        tgt.fill = copy(src.fill)
        tgt.alignment = copy(src.alignment)
        tgt.border = copy(src.border)
        tgt.number_format = src.number_format


# ============================================================================
# DATA VALIDATION FUNCTIONS
# ============================================================================

def add_list_validation(ws, cell_range, formula, allow_blank=True):
    """Add dropdown list validation using named range formula."""
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "Please select from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select from list"
    dv.promptTitle = "Input"
    dv.showInputMessage = True
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def add_date_validation(ws, cell_range, min_date="2024-01-01"):
    """Add date validation with minimum date."""
    dv = DataValidation(type="date", operator="greaterThan", formula1=min_date)
    dv.error = "Please enter a valid date after 2024-01-01."
    dv.errorTitle = "Invalid Date"
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def add_yes_no_validation(ws, cell_range):
    """Add YES/NO dropdown validation."""
    dv = DataValidation(type="list", formula1='"YES,NO,NA"', allow_blank=True)
    dv.error = "Select YES, NO, or NA."
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


# ============================================================================
# CONDITIONAL FORMATTING FUNCTIONS
# ============================================================================

def add_result_conditional_formatting(ws, cell_range):
    """Add conditional formatting for RESULT cells: FAIL→RED, HOLD→AMBER, PASS→GREEN."""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"FAIL"'],
                   fill=FILL_CF_RED, font=Font(color=COLORS['CF_RED_FONT'], bold=True)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"HOLD"'],
                   fill=FILL_CF_AMBER))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"PASS"'],
                   fill=FILL_CF_GREEN))


def add_status_conditional_formatting(ws, cell_range):
    """Add conditional formatting for STATUS cells: OPEN→AMBER, CLOSED→GREEN."""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"OPEN"'],
                   fill=FILL_CF_AMBER))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"CLOSED"'],
                   fill=FILL_CF_GREEN))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"IN PROGRESS"'],
                   fill=PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')))


def add_risk_conditional_formatting(ws, cell_range):
    """Add conditional formatting for RISK cells: CRITICAL→RED, HIGH→AMBER."""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"CRITICAL"'],
                   fill=FILL_CF_RED, font=Font(color=COLORS['CF_RED_FONT'], bold=True)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"HIGH"'],
                   fill=FILL_CF_AMBER))


def add_overdue_conditional_formatting(ws, date_cell_range):
    """Add conditional formatting: overdue dates → RED, within 5 days → AMBER."""
    ws.conditional_formatting.add(date_cell_range,
        FormulaRule(formula=[f'{date_cell_range.split(":")[0]}<TODAY()'],
                    fill=FILL_CF_RED))
    ws.conditional_formatting.add(date_cell_range,
        FormulaRule(formula=[f'AND({date_cell_range.split(":")[0]}>=TODAY(),{date_cell_range.split(":")[0]}<TODAY()+5)'],
                    fill=FILL_CF_AMBER))


# ============================================================================
# WORKBOOK MANAGEMENT FUNCTIONS
# ============================================================================

def add_revision_history_sheet(wb, form_code, current_rev="V0"):
    """Add hidden _REVHIST sheet with revision history table."""
    if '_REVHIST' in wb.sheetnames:
        return wb['_REVHIST']

    ws = wb.create_sheet('_REVHIST')
    ws.sheet_state = 'hidden'

    # Set column widths (simple layout)
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # Headers
    headers = ['Rev', 'Date', 'Description', 'Author', 'Approved By']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_COL_HDR
        cell.alignment = ALIGN_CENTER

    # V0 baseline entry
    ws.cell(row=2, column=1, value=current_rev).font = FONT_BODY
    ws.cell(row=2, column=2, value='2026-03-21').font = FONT_BODY
    ws.cell(row=2, column=3, value='Initial controlled baseline').font = FONT_BODY
    ws.cell(row=2, column=4, value='QMS').font = FONT_BODY
    ws.cell(row=2, column=5, value='General Manager').font = FONT_BODY

    # V1 upgrade entry
    ws.cell(row=3, column=1, value='V1').font = FONT_BODY
    ws.cell(row=3, column=2, value='2026-03-22').font = FONT_BODY
    ws.cell(row=3, column=3, value=f'Upgrade: CNC fields, cross-form linkage, conditional formatting, parent procedure ref').font = FONT_BODY
    ws.cell(row=3, column=4, value='QMS').font = FONT_BODY
    ws.cell(row=3, column=5, value='General Manager').font = FONT_BODY

    return ws


def protect_sheet_formulas(ws, password='hesem2026'):
    """Lock formula cells, unlock input cells, then protect sheet."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.protection = Protection(locked=True)
            elif cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                rgb = cell.fill.fgColor.rgb if cell.fill.fgColor.rgb else ''
                if rgb in ('00FFFFFF', 'FFFFFF', '00F8FAFC', 'F8FAFC'):
                    # White or near-white = input cell → unlock
                    cell.protection = Protection(locked=False)

    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.enable()
    # Allow filtering and selection
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatCells = False


def backup_and_save(wb, filepath):
    """Create backup of original, save modified workbook, return SHA-256."""
    # Backup
    bak_path = filepath + '.bak'
    if os.path.exists(filepath) and not os.path.exists(bak_path):
        shutil.copy2(filepath, bak_path)

    # Save
    wb.save(filepath)

    # Compute SHA-256
    sha = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            sha.update(chunk)
    return sha.hexdigest()


def update_registry(registry_path, form_code, new_sha256, new_rev="V1"):
    """Update form-control-register.json with new checksum and revision."""
    with open(registry_path, 'r', encoding='utf-8') as f:
        registry = json.load(f)

    # Find and update the form entry
    for entry in registry:
        if isinstance(entry, dict) and entry.get('code') == form_code:
            entry['sha256'] = new_sha256
            entry['rev'] = new_rev
            entry['effective_date'] = '2026-03-22'
            if 'notes' not in entry:
                entry['notes'] = []
            entry['notes'].append(f'{new_rev}: CNC upgrade — fields, validation, conditional formatting, cross-form linkage')
            break

    with open(registry_path, 'w', encoding='utf-8') as f:
        json.dump(registry, f, indent=2, ensure_ascii=False)


def bump_revision(ws, new_rev="V1", new_date="2026-03-22"):
    """Update revision and effective date in header meta box.

    Scans first 6 rows for 'Revision' label and updates the value cell
    in the adjacent merged range. Handles merged cells correctly by
    finding the top-left cell of each merge range.
    """
    # Build a map of merged ranges for quick lookup
    merge_map = {}
    for m in ws.merged_cells.ranges:
        if m.min_row <= 6:
            for r in range(m.min_row, m.max_row + 1):
                for c in range(m.min_col, m.max_col + 1):
                    merge_map[(r, c)] = (m.min_row, m.min_col)

    for row in ws.iter_rows(min_row=1, max_row=6, max_col=ws.max_column):
        for cell in row:
            if cell.value and str(cell.value).strip() == 'Revision':
                # Find the value cell - scan right for the next merge range start
                for offset in range(1, 10):
                    val_col = cell.column + offset
                    coord = (cell.row, val_col)
                    if coord in merge_map:
                        top_r, top_c = merge_map[coord]
                        if top_c != cell.column:  # Different merge group
                            ws.cell(row=top_r, column=top_c).value = new_rev
                            break
                    elif val_col <= ws.max_column:
                        try:
                            ws.cell(row=cell.row, column=val_col).value = new_rev
                            break
                        except AttributeError:
                            continue
            elif cell.value and str(cell.value).strip() == 'Eff. Date':
                for offset in range(1, 10):
                    val_col = cell.column + offset
                    coord = (cell.row, val_col)
                    if coord in merge_map:
                        top_r, top_c = merge_map[coord]
                        if top_c != cell.column:
                            ws.cell(row=top_r, column=top_c).value = new_date
                            break
                    elif val_col <= ws.max_column:
                        try:
                            ws.cell(row=cell.row, column=val_col).value = new_date
                            break
                        except AttributeError:
                            continue


# ============================================================================
# LISTS SHEET HELPERS
# ============================================================================

def ensure_lists_values(ws_lists, col_header, values, start_col=None):
    """Ensure a column in _LISTS has the specified header and values.
    If column exists, update values. If not, add new column."""
    # Find existing column
    for col in range(1, ws_lists.max_column + 2):
        cell = ws_lists.cell(row=1, column=col)
        if cell.value == col_header:
            # Update values
            for i, v in enumerate(values, 2):
                ws_lists.cell(row=i, column=col, value=v)
            return col
        if cell.value is None:
            # Use this empty column
            ws_lists.cell(row=1, column=col, value=col_header)
            for i, v in enumerate(values, 2):
                ws_lists.cell(row=i, column=col, value=v)
            return col
    return None


def add_named_range(wb, name, sheet_name, min_col, min_row, max_col, max_row):
    """Add a workbook-level named range."""
    from openpyxl.workbook.defined_name import DefinedName
    col_letter_min = get_column_letter(min_col)
    col_letter_max = get_column_letter(max_col)
    ref = f"'{sheet_name}'!${col_letter_min}${min_row}:${col_letter_max}${max_row}"
    # Remove existing if present
    if name in wb.defined_names:
        del wb.defined_names[name]
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)
