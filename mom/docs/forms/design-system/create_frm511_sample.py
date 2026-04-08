# -*- coding: utf-8 -*-
"""
CREATE FRM-511 Setup and First-Piece Record
Using master template v2.0 from Desktop as base.
"""
import openpyxl, copy, hashlib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import openpyxl.worksheet.properties

P = {
    'primary':'1B3A6B','accent':'0078D4','accent2':'005A9E',
    'ice':'EFF6FF','fog':'F1F5F9','white':'FFFFFF',
    'near':'F8FAFC','silver':'E2E8F0',
    'dk':'0F172A','mid':'334155','muted':'64748B','plc':'CBD5E1',
    'tec_bg':'DCFCE7','tec_fg':'14532D',
    'qua_bg':'FEF9C3','qua_fg':'713F12',
    'cap_bg':'FCE7F3','cap_fg':'831843',
    'con_bg':'F3E8FF','con_fg':'4C1D95',
    'com_bg':'DBEAFE','com_fg':'1E3A8A',
}
NC = 40
TA = Side(style='thin', color=P['accent'])
TS = Side(style='thin', color=P['silver'])

def fi(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fo(sz, bold=False, c='dk'): return Font(name='Segoe UI', size=sz, bold=bold, color=P[c])

AL = Alignment(horizontal='left', vertical='center', indent=1)
AC = Alignment(horizontal='center', vertical='center')

# Load master
TMPL = "C:/Users/TEST4/Desktop/frm-000-master-template.xlsx"
wb = openpyxl.load_workbook(TMPL)
src = wb['01-MASTER-A4-PORTRAIT']
ws = wb.create_sheet('FRM-511', 0)

# Col widths
for i in range(1, NC+1):
    ws.column_dimensions[get_column_letter(i)].width = 2.375

# Row 1 hidden
ws.row_dimensions[1].hidden = True

# Header rows 2-6: copy from master
for row in range(2, 7):
    ws.row_dimensions[row].height = 15.0
    for col in range(1, NC+1):
        s = src.cell(row, col)
        d = ws.cell(row, col)
        d.value = s.value
        d.font = copy.copy(s.font)
        d.fill = copy.copy(s.fill)
        d.alignment = copy.copy(s.alignment)
        d.border = copy.copy(s.border)

for mc in src.merged_cells.ranges:
    ws.merge_cells(str(mc))

ws['I2'].value = 'SETUP AND FIRST-PIECE RECORD'
ws['AI2'].value = 'FRM-511'
ws['AI3'].value = 'Rev.01'
ws['AI4'].value = '2026-03-23'
ws['AI5'].value = 'Production / QA'
ws['AI6'].value = 'Production Director'

# --- HELPERS ---
def spacer(r):
    ws.row_dimensions[r].height = 4.0
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    ws.cell(r,1).fill = fi(P['white'])
    return r+1

def sect(r, text):
    ws.row_dimensions[r].height = 17.0
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r,1)
    c.value = text
    c.font = fo(9, True, 'primary')
    c.fill = fi(P['ice'])
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = Border(left=TA, right=TA, top=TA, bottom=TA)
    return r+1

def dpair(r, lL, lR, first=False, last=False):
    ws.row_dimensions[r].height = 20.0
    for s,e in [(1,7),(8,20),(21,27),(28,NC)]:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    t = TA if first else TS
    b = TA if last else TS
    c=ws.cell(r,1); c.value=lL; c.font=fo(8,True,'mid'); c.fill=fi(P['fog']); c.alignment=AL
    c.border=Border(left=TA, right=TS, top=t, bottom=b)
    c=ws.cell(r,8); c.font=fo(8,False,'dk'); c.fill=fi(P['white']); c.alignment=AL
    c.border=Border(right=TS, top=t, bottom=b)
    c=ws.cell(r,21); c.value=lR; c.font=fo(8,True,'mid'); c.fill=fi(P['fog']); c.alignment=AL
    c.border=Border(right=TS, top=t, bottom=b)
    c=ws.cell(r,28); c.font=fo(8,False,'dk'); c.fill=fi(P['white']); c.alignment=AL
    c.border=Border(right=TA, top=t, bottom=b)
    return r+1

def colhdr(r, hdrs):
    ws.row_dimensions[r].height = 17.0
    for s,e,txt in hdrs:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
        c=ws.cell(r,s); c.value=txt
        c.font=Font(name='Segoe UI', size=9, bold=True, color='FFFFFF')
        c.fill=fi(P['accent2']); c.alignment=AC
        lb = TA if s==1 else TS
        rb = TA if e==NC else TS
        c.border=Border(left=lb, right=rb, top=TA, bottom=TS)
    return r+1

def ckrow(r, num, cat, item, crit, catc=None, last=False):
    ws.row_dimensions[r].height = 20.0
    zn = [(1,2),(3,5),(6,19),(20,31),(32,35),(36,38),(39,40)]
    for s,e in zn:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    b = TA if last else TS
    even = (num%2==0)
    ifill = fi(P['near']) if even else fi(P['white'])

    c=ws.cell(r,1); c.value=num; c.font=fo(8,True,'mid'); c.fill=fi(P['fog']); c.alignment=AC
    c.border=Border(left=TA, right=TS, bottom=b)

    c=ws.cell(r,3); c.value=cat; c.alignment=AC
    if catc:
        c.font=Font(name='Segoe UI',size=8,bold=True,color=catc[1])
        c.fill=PatternFill(start_color=catc[0],end_color=catc[0],fill_type='solid')
    else:
        c.font=fo(8,True,'mid'); c.fill=fi(P['fog'])
    c.border=Border(right=TS, bottom=b)

    c=ws.cell(r,6); c.value=item; c.font=fo(8,False,'dk'); c.fill=ifill; c.alignment=AL
    c.border=Border(right=TS, bottom=b)

    c=ws.cell(r,20); c.value=crit; c.font=fo(7,False,'muted'); c.fill=fi(P['fog']); c.alignment=AL
    c.border=Border(right=TS, bottom=b)

    for s in [32,36,39]:
        c=ws.cell(r,s); c.font=fo(8,False,'dk'); c.fill=fi(P['white']); c.alignment=AC
        rb = TA if s==39 else TS
        c.border=Border(right=rb, bottom=b)
    return r+1

# === BUILD ===
r = 7
r = spacer(r)  # r=8

# SECTION 1
r = sect(r, '  1.  SETUP / FIRST-PIECE HEADER')  # r=9
r = dpair(r, 'Job / WO', 'Part No. / Revision', first=True)
r = dpair(r, 'Machine / Cell', 'Operation / Route Step')
r = dpair(r, 'Setup Technician', 'QA Reviewer')
r = dpair(r, 'CNC Program / Rev', 'Fixture ID')
r = dpair(r, 'Tool List Ref / Count', 'Setup Time (min)')
r = dpair(r, 'Coolant Type / Conc.', 'Offset Baseline Verified?')
r = dpair(r, 'Drawing Rev = Program Rev?', 'Evidence Folder')
r = dpair(r, 'Ref: SOP-504 / WI-517', 'Linked FRM-302 / FRM-519', last=True)
# r should be ~18

r = spacer(r)

# SECTION 2
r = sect(r, '  2.  SETUP & FIRST-PIECE VERIFICATION')
CKH = [(1,2,'#'),(3,5,'Cat.'),(6,19,'Setup / FP Check Item'),
       (20,31,'Acceptance / Control Rule'),(32,35,'Result'),
       (36,38,'Owner'),(39,40,'Ref.')]
r = colhdr(r, CKH)
ck_start = r

CT = {'SET':(P['tec_bg'],P['tec_fg']),'TOOL':(P['cap_bg'],P['cap_fg']),
      'GAGE':(P['con_bg'],P['con_fg']),'QUA':(P['qua_bg'],P['qua_fg']),
      'OPS':(P['com_bg'],P['com_fg'])}

items = [
    (1,'SET','Setup baseline matches released package (FRM-302).','FRM-302 and released packet align.'),
    (2,'SET','Machine / fixture / datums are correct.','Actual setup matches baseline condition.'),
    (3,'TOOL','Tool list, offsets and wear comp. are loaded.','Tool data matches FRM-302 tool list.'),
    (4,'TOOL','Tool condition is acceptable (no chip/wear).','Visual + touch-off check before run.'),
    (5,'GAGE','Required gages / programs available and in cal.','Cal sticker current; CMM program rev matches.'),
    (6,'QUA','First-piece dimensions are within tolerance.','All CTQ dims measured and recorded.'),
    (7,'QUA','Surface finish / burr / edge break acceptable.','Ra check per drawing requirement.'),
    (8,'QUA','GD&T / profile / position verified (if applicable).','CMM report attached or referenced.'),
    (9,'OPS','Abnormal findings are documented.','Any discrepancy has hold or action logic.'),
    (10,'OPS','Run condition after first-piece is explicit.','Pass-to-run or hold decision is visible.'),
]
for i,(num,cat,item,crit) in enumerate(items):
    r = ckrow(r, num, cat, item, crit, CT.get(cat), last=(i==len(items)-1))
ck_end = r - 1

r = spacer(r)

# SECTION 3: GATE
r = sect(r, '  3.  RUN DECISION / RELEASE GATE')

# Gate header bar
ws.row_dimensions[r].height = 17.0
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
c=ws.cell(r,1); c.value='GATE  \u2014  After setup and first-piece verification'
c.font=Font(name='Segoe UI',size=9,bold=True,color='FFFFFF')
c.fill=fi(P['accent2']); c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
c.border=Border(left=TA, right=TA, top=TA, bottom=TS)
r += 1

r = dpair(r, 'Run Decision', 'Summary / Notes', first=True)
r = dpair(r, 'Open Conditions before Run', 'Escalation if Not Closed')
r = dpair(r, 'Action Owner', 'Completion Deadline', last=True)

r = spacer(r)

# SECTION 4: APPROVAL
r = sect(r, '  4.  APPROVAL')

# 3-col sig
ws.row_dimensions[r].height = 17.0
sigs = [(1,13,'Setup Technician'),(14,26,'QA Reviewer'),(27,40,'Production Supervisor')]
for s,e,lab in sigs:
    ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    c=ws.cell(r,s); c.value=lab; c.font=fo(8,True,'mid'); c.fill=fi(P['fog']); c.alignment=AC
    c.border=Border(left=TA, right=TA, top=TA, bottom=TA)
r += 1
sig_start = r
for sr in range(r, r+3):
    ws.row_dimensions[sr].height = 26.0
for s,e,_ in sigs:
    ws.merge_cells(start_row=r, start_column=s, end_row=r+2, end_column=e)
    c=ws.cell(r,s)
    c.value='Name  /  Signature  /  Date'
    c.font=Font(name='Segoe UI',size=8,color=P['plc'])
    c.fill=fi(P['white'])
    c.alignment=Alignment(horizontal='center',vertical='bottom')
    c.border=Border(left=TA,right=TA,top=TA,bottom=TA)
r += 3

r = spacer(r)

# NOTICE
ws.row_dimensions[r].height = 24.0
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
c=ws.cell(r,1)
c.value=('NOTICE  \u2014  Enter data only in white input cells.  Do not add, delete or resize columns/rows.  '
         'One form per setup event.  Ref: SOP-504 / WI-517.  Retain: 10 years per QMS schedule.')
c.font=Font(name='Segoe UI',size=6,color=P['mid'])
c.fill=fi(P['ice'])
c.alignment=Alignment(horizontal='left',vertical='center',indent=2,wrap_text=True)
c.border=Border(left=TA,right=TA,top=TA,bottom=TA)

# DATA VALIDATIONS
dv = DataValidation(type="list", formula1="=LISTS!$A$2:$A$5", allow_blank=True)
dv.error = "Select: PASS / HOLD / FAIL / NA"
dv.showErrorMessage = True
ws.add_data_validation(dv)
for rr in range(ck_start, ck_end+1):
    dv.add(f"AF{rr}")

# CONDITIONAL FORMATTING
rng = f"AF{ck_start}:AF{ck_end}"
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"PASS"'],
    fill=PatternFill(start_color='C6EFCE',end_color='C6EFCE',fill_type='solid'),
    font=Font(color='006100')))
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"HOLD"'],
    fill=PatternFill(start_color='FFEB9C',end_color='FFEB9C',fill_type='solid'),
    font=Font(color='9C6500')))
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"FAIL"'],
    fill=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid'),
    font=Font(color='9C0006')))

# PRINT
ws.page_setup.paperSize = 9
ws.page_setup.orientation = 'portrait'
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# Remove other sheets
for name in list(wb.sheetnames):
    if name not in ['FRM-511', 'LISTS']:
        del wb[name]

# SAVE
OUT = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau/05-FRM-500/FRM-511_Setup_and_First_Piece_Record.xlsx"
wb.save(OUT)
sha = hashlib.sha256(open(OUT,'rb').read()).hexdigest()

wb2 = openpyxl.load_workbook(OUT)
ws2 = wb2['FRM-511']
print(f"FRM-511 CREATED:")
print(f"  Rows: {ws2.max_row}")
print(f"  Merged: {len(ws2.merged_cells.ranges)}")
print(f"  DVs: {len(ws2.data_validations.dataValidation)}")
print(f"  SHA: {sha[:16]}...")
print(f"  Output: {OUT}")
