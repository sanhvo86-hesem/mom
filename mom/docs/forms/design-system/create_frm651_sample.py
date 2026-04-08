# -*- coding: utf-8 -*-
"""
FRM-651 NCR Report — Sample using Design System v2.0
Different from FRM-511: Report form with free-text blocks, NOT checklist-heavy.
Demonstrates: text area sections, multi-section layout, 2 visible tabs.
ALL RULES: wrap_text=True, no indent, borders on ALL 40 cells, logo.
"""
import openpyxl, copy, hashlib, math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
import openpyxl.worksheet.properties

P = {
    'primary':'1B3A6B','accent':'0078D4','accent2':'005A9E',
    'ice':'EFF6FF','fog':'F1F5F9','white':'FFFFFF',
    'near':'F8FAFC','silver':'E2E8F0',
    'dk':'0F172A','mid':'334155','muted':'64748B','plc':'CBD5E1',
    'tec_bg':'DCFCE7','tec_fg':'14532D','qua_bg':'FEF9C3','qua_fg':'713F12',
    'cap_bg':'FCE7F3','cap_fg':'831843','con_bg':'F3E8FF','con_fg':'4C1D95',
    'com_bg':'DBEAFE','com_fg':'1E3A8A',
}
NC = 40
BLUE = Side(style='thin', color=P['accent'])
SILVER = Side(style='thin', color=P['silver'])
NONE_S = Side()
EMU = 9525
LOGO_PATH = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau/00-FORM-DESIGN-SYSTEM/hesem-logo.png"
LOGO_ASPECT = 3.46

def fi(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fo(sz, bold=False, c='dk'): return Font(name='Segoe UI', size=sz, bold=bold, color=P[c])
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AT = Alignment(horizontal='left', vertical='top', wrap_text=True)  # for text areas

def embed_logo(ws, logo_cols=8):
    zone_w = logo_cols * 19
    zone_h = 5 * 20
    lw = math.floor(zone_w * 0.78)
    lh = math.floor(lw / LOGO_ASPECT)
    if lh > zone_h * 0.78:
        lh = math.floor(zone_h * 0.78); lw = math.floor(lh * LOGO_ASPECT)
    xo = math.floor((zone_w - lw) / 2) * EMU
    yo = math.floor((zone_h - lh) / 2) * EMU
    img = XlImage(LOGO_PATH)
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=0, colOff=xo, row=1, rowOff=yo),
        ext=XDRPositiveSize2D(cx=lw*EMU, cy=lh*EMU))
    ws.add_image(img)

# =====================================================================
# ROW BORDER HELPERS — borders on ALL 40 cells
# =====================================================================
def _set_row_borders(ws, r, zones, is_first, is_last):
    top = BLUE if is_first else NONE_S
    bot = BLUE if is_last else SILVER
    zone_ends = set(e for s, e in zones[:-1])
    for col in range(1, NC+1):
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == NC else (SILVER if col in zone_ends else NONE_S)
        ws.cell(r, col).border = Border(left=left, right=right, top=top, bottom=bot)

def _set_row_borders_table(ws, r, zones, is_last):
    bot = BLUE if is_last else SILVER
    zone_ends = set(e for s, e in zones[:-1])
    for col in range(1, NC+1):
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == NC else (SILVER if col in zone_ends else NONE_S)
        ws.cell(r, col).border = Border(left=left, right=right, bottom=bot)

# =====================================================================
# BLOCK BUILDERS
# =====================================================================
PAIR_ZONES = [(1,7),(8,20),(21,27),(28,NC)]
CK_ZONES = [(1,2),(3,5),(6,21),(22,31),(32,35),(36,NC)]

def spacer(ws, r):
    ws.row_dimensions[r].height = 4.0
    for col in range(1, NC+1):
        ws.cell(r, col).fill = fi(P['white']); ws.cell(r, col).border = Border()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    return r+1

def sect(ws, r, text):
    ws.row_dimensions[r].height = 17.0
    for col in range(1, NC+1):
        ws.cell(r, col).fill = fi(P['ice'])
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == NC else NONE_S
        ws.cell(r, col).border = Border(left=left, right=right, top=BLUE, bottom=BLUE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1); c.value = text; c.font = fo(9, True, 'primary')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    return r+1

def dpair(ws, r, lL, lR, first=False, last=False):
    ws.row_dimensions[r].height = 20.0
    _set_row_borders(ws, r, PAIR_ZONES, first, last)
    for (s,e), zf in zip(PAIR_ZONES, [fi(P['fog']),fi(P['white']),fi(P['fog']),fi(P['white'])]):
        for col in range(s, e+1): ws.cell(r, col).fill = zf
    for s, e in PAIR_ZONES:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    ws.cell(r,1).value=lL; ws.cell(r,1).font=fo(8,True,'mid'); ws.cell(r,1).alignment=AL
    ws.cell(r,8).font=fo(8,False,'dk'); ws.cell(r,8).alignment=AL
    ws.cell(r,21).value=lR; ws.cell(r,21).font=fo(8,True,'mid'); ws.cell(r,21).alignment=AL
    ws.cell(r,28).font=fo(8,False,'dk'); ws.cell(r,28).alignment=AL
    return r+1

def textarea(ws, r, label, nrows=3, first=False, last=False):
    """Full-width label row + multi-row white text area."""
    # Label sub-header
    ws.row_dimensions[r].height = 17.0
    for col in range(1, NC+1):
        ws.cell(r, col).fill = fi(P['fog'])
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == NC else NONE_S
        top = BLUE if first else NONE_S
        ws.cell(r, col).border = Border(left=left, right=right, top=top, bottom=SILVER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1); c.value = label; c.font = fo(8, True, 'mid'); c.alignment = AL
    r += 1
    # Text area rows
    for i in range(nrows):
        is_last_row = last and (i == nrows - 1)
        ws.row_dimensions[r].height = 20.0
        bot = BLUE if is_last_row else SILVER
        for col in range(1, NC+1):
            ws.cell(r, col).fill = fi(P['white'])
            left = BLUE if col == 1 else NONE_S
            right = BLUE if col == NC else NONE_S
            ws.cell(r, col).border = Border(left=left, right=right, bottom=bot)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        if i == 0:
            ws.cell(r, 1).font = fo(8, False, 'dk'); ws.cell(r, 1).alignment = AT
        r += 1
    return r

def colhdr(ws, r, hdrs):
    ws.row_dimensions[r].height = 17.0
    zones = [(s,e) for s,e,_ in hdrs]
    _set_row_borders(ws, r, zones, True, False)
    for col in range(1, NC+1): ws.cell(r, col).fill = fi(P['accent2'])
    for s, e, txt in hdrs:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
        c = ws.cell(r, s); c.value = txt
        c.font = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF'); c.alignment = AC
    return r+1

def ckrow(ws, r, num, cat, item, crit, catc=None, last=False):
    ws.row_dimensions[r].height = 20.0
    _set_row_borders_table(ws, r, CK_ZONES, last)
    even = (num % 2 == 0)
    ifill = fi(P['near']) if even else fi(P['white'])
    zone_fills = [fi(P['fog']), fi(catc[0]) if catc else fi(P['fog']),
                  ifill, fi(P['fog']), fi(P['white']), fi(P['white']), fi(P['white'])]
    for (s,e), zf in zip(CK_ZONES, zone_fills):
        for col in range(s, e+1): ws.cell(r, col).fill = zf
    for s, e in CK_ZONES:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    ws.cell(r,1).value=num; ws.cell(r,1).font=fo(8,True,'mid'); ws.cell(r,1).alignment=AC
    c=ws.cell(r,3); c.value=cat; c.alignment=AC
    c.font=Font(name='Segoe UI',size=8,bold=True,color=catc[1]) if catc else fo(8,True,'mid')
    ws.cell(r,6).value=item; ws.cell(r,6).font=fo(8,False,'dk'); ws.cell(r,6).alignment=AL
    ws.cell(r,22).value=crit; ws.cell(r,22).font=fo(7,False,'muted'); ws.cell(r,22).alignment=AL
    for s in [32,36,39]:
        ws.cell(r,s).font=fo(8,False,'dk'); ws.cell(r,s).alignment=AC
    return r+1

def sig3(ws, r, labels):
    SZ = [(1,13),(14,26),(27,NC)]
    ws.row_dimensions[r].height = 17.0
    for col in range(1, NC+1): ws.cell(r, col).fill = fi(P['fog'])
    for s, e in SZ:
        for col in range(s, e+1):
            ws.cell(r,col).border = Border(
                left=BLUE if col==s else NONE_S, right=BLUE if col==e else NONE_S,
                top=BLUE, bottom=BLUE)
    for (s,e),lab in zip(SZ, labels):
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
        ws.cell(r,s).value=lab; ws.cell(r,s).font=fo(8,True,'mid'); ws.cell(r,s).alignment=AC
    r += 1
    for sr in range(r, r+3): ws.row_dimensions[sr].height = 26.0
    for s, e in SZ:
        for sr in range(r, r+3):
            for col in range(s, e+1):
                ws.cell(sr,col).fill = fi(P['white'])
                ws.cell(sr,col).border = Border(
                    left=BLUE if col==s else NONE_S, right=BLUE if col==e else NONE_S,
                    top=BLUE, bottom=BLUE)
        ws.merge_cells(start_row=r, start_column=s, end_row=r+2, end_column=e)
        c=ws.cell(r,s); c.value='Name  /  Signature  /  Date'
        c.font=Font(name='Segoe UI',size=8,color=P['plc'])
        c.alignment=Alignment(horizontal='center',vertical='bottom',wrap_text=True)
    return r+3

def notice_bar(ws, r, text):
    ws.row_dimensions[r].height = 24.0
    for col in range(1, NC+1):
        ws.cell(r,col).fill = fi(P['ice'])
        ws.cell(r,col).border = Border(
            left=BLUE if col==1 else NONE_S, right=BLUE if col==NC else NONE_S,
            top=BLUE, bottom=BLUE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c=ws.cell(r,1); c.value=text
    c.font=Font(name='Segoe UI',size=6,color=P['mid'])
    c.alignment=Alignment(horizontal='left',vertical='center',indent=0,wrap_text=True)
    return r+1

def build_header(ws, src, title, code, owner, approved):
    """Copy header from master src, update content. Apply to any visible tab."""
    ws.row_dimensions[1].hidden = True
    for row in range(2, 7):
        ws.row_dimensions[row].height = 15.0
        for col in range(1, NC+1):
            s = src.cell(row, col); d = ws.cell(row, col)
            d.value = s.value; d.font = copy.copy(s.font)
            d.fill = copy.copy(s.fill); d.alignment = copy.copy(s.alignment)
            d.border = copy.copy(s.border)
    for mc in src.merged_cells.ranges:
        ws.merge_cells(str(mc))
    ws['I2'].value = title; ws['AI2'].value = code; ws['AI3'].value = 'Rev.01'
    ws['AI4'].value = '2026-03-23'; ws['AI5'].value = owner; ws['AI6'].value = approved
    # Fix meta values: no indent, wrap
    meta_align = Alignment(horizontal='left', vertical='center', indent=0, wrap_text=True)
    for hr in range(2, 7):
        c = ws.cell(hr, 35)
        if c.value is not None: c.alignment = meta_align

# =====================================================================
# LOAD TEMPLATE + BUILD
# =====================================================================
TMPL = "C:/Users/TEST4/Desktop/frm-000-master-template.xlsx"
wb = openpyxl.load_workbook(TMPL)
src = wb['01-MASTER-A4-PORTRAIT']

# =====================================================================
# TAB 1: NCR_FORM (main form face)
# =====================================================================
ws1 = wb.create_sheet('NCR_FORM', 0)
for i in range(1, NC+1): ws1.column_dimensions[get_column_letter(i)].width = 2.375
build_header(ws1, src, 'NCR REPORT', 'FRM-651', 'QA', 'QA Manager')
embed_logo(ws1, 8)

r = 7
r = spacer(ws1, r)

# S1: REFERENCE
r = sect(ws1, r, '  1.  REFERENCE INFORMATION')
r = dpair(ws1, r, 'NCR ID', 'Source', first=True)
r = dpair(ws1, r, 'Severity', 'Containment Status')
r = dpair(ws1, r, 'Job / Lot / Part / Revision', 'Suspect Qty')
r = dpair(ws1, r, 'Prepared by / Date-Time', 'Record Status')
r = dpair(ws1, r, 'Inspector / Quality Owner', 'Discovery Point')
r = dpair(ws1, r, 'Quantity Affected', 'Quantity Contained / Sorted')
r = dpair(ws1, r, 'Cost of NC (scrap+rework+delay)', 'Customer Notification Required?')
r = dpair(ws1, r, 'Linked FRM-641 / FRM-406 / FRM-213', 'Ref: SOP-605 / Retain: 10 yrs', last=True)
r = spacer(ws1, r)

# S2: EVENT DETAIL (text areas)
r = sect(ws1, r, '  2.  EVENT DETAIL / CONTROL CONTEXT')
r = textarea(ws1, r, 'NONCONFORMITY DESCRIPTION / REQUIREMENT REFERENCE', nrows=3, first=True)
r = textarea(ws1, r, 'OBJECTIVE EVIDENCE / RISK TO DOWNSTREAM OR SHIPPED PRODUCT', nrows=3, last=True)
r = spacer(ws1, r)

# S3: CONTAINMENT TABLE
r = sect(ws1, r, '  3.  CONTAINMENT / DISPOSITION SUMMARY')
CKH = [(1,2,'#'),(3,5,'Cat.'),(6,21,'Containment / Disposition Item'),
       (22,31,'Requirement / Evidence'),(32,35,'Status'),(36,NC,'Owner')]
r = colhdr(ws1, r, CKH)
ck_start = r
CT = {'CON':(P['con_bg'],P['con_fg']),'DIS':(P['cap_bg'],P['cap_fg']),
      'QUA':(P['qua_bg'],P['qua_fg'])}
items = [
    (1,'CON','Suspect material/parts segregated and tagged HOLD.','Physical isolation confirmed.'),
    (2,'CON','Downstream / shipped product assessed.','Affected lots identified and traced.'),
    (3,'CON','Work-in-progress stopped or rerouted.','Production notified; WIP tagged.'),
    (4,'DIS','Disposition decision recorded.','USE-AS-IS / REWORK / SCRAP / RTV.'),
    (5,'DIS','Rework instruction issued (if applicable).','Method, acceptance, re-inspect defined.'),
]
for i,(num,cat,item,crit) in enumerate(items):
    r = ckrow(ws1, r, num, cat, item, crit, CT.get(cat), last=(i==4))
ck_end = r - 1
r = spacer(ws1, r)

# S4: FINAL DECISION
r = sect(ws1, r, '  4.  FINAL DECISION / CLOSURE')
r = dpair(ws1, r, 'Disposition', 'Affected Stakeholders Notified?', first=True)
r = dpair(ws1, r, 'Linked SCAR / CAPA Ref', 'Closeout Status', last=True)
r = spacer(ws1, r)

# S5: APPROVAL
r = sect(ws1, r, '  5.  APPROVAL / SIGN-OFF')
r = sig3(ws1, r, ['Prepared by', 'Reviewed by (QA)', 'Approved by'])
r = spacer(ws1, r)
r = notice_bar(ws1, r,
    'NOTICE  \u2014  Complete in English. No back-dating, history deletion, proxy-signing. '
    'One form per NCR event. Ref: SOP-605 / WI-605. Retain: 10 years per QMS schedule.')

# =====================================================================
# TAB 2: ACTIONS (visible action tracker tab)
# =====================================================================
ws2 = wb.create_sheet('ACTIONS', 1)
for i in range(1, NC+1): ws2.column_dimensions[get_column_letter(i)].width = 2.375
build_header(ws2, src, 'NCR REPORT  \u2014  ACTION TRACKER', 'FRM-651', 'QA', 'QA Manager')
embed_logo(ws2, 8)

r = 7
r = spacer(ws2, r)

r = sect(ws2, r, '  1.  CORRECTIVE ACTIONS ARISING FROM NCR')
ACT_ZONES = [(1,2),(3,12),(13,24),(25,29),(30,33),(34,37),(38,NC)]
ACT_HDR = [(1,2,'#'),(3,12,'Issue / Root Cause'),(13,24,'Required Action'),
           (25,29,'Owner'),(30,33,'Due Date'),(34,37,'Status'),(38,NC,'Verified')]
r = colhdr(ws2, r, ACT_HDR)
act_start = r
for i in range(1, 6):
    ws2.row_dimensions[r].height = 20.0
    is_last = (i == 5)
    _set_row_borders_table(ws2, r, ACT_ZONES, is_last)
    ifill = fi(P['near']) if i % 2 == 0 else fi(P['white'])
    for (s,e) in ACT_ZONES:
        for col in range(s, e+1):
            ws2.cell(r, col).fill = fi(P['fog']) if s == 1 else ifill
    for s, e in ACT_ZONES:
        ws2.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    ws2.cell(r,1).value = i; ws2.cell(r,1).font = fo(8,True,'mid'); ws2.cell(r,1).alignment = AC
    for s in [3,13,25,30,34,38]:
        ws2.cell(r,s).font = fo(8,False,'dk'); ws2.cell(r,s).alignment = AL if s < 25 else AC
    r += 1
act_end = r - 1

r = spacer(ws2, r)
r = notice_bar(ws2, r,
    'NOTICE  \u2014  Track all corrective actions arising from this NCR. '
    'Link to FRM-652 CAPA if systemic root cause identified.')

# =====================================================================
# DV + CF
# =====================================================================
ws_lists = wb['LISTS']
# Result/Status DV on NCR_FORM containment table
dv1 = DataValidation(type="list", formula1="=LISTS!$C$2:$C$5", allow_blank=True)
dv1.showErrorMessage = True; ws1.add_data_validation(dv1)
for rr in range(ck_start, ck_end+1): dv1.add(f"AF{rr}")

# Status DV on ACTIONS tab
dv2 = DataValidation(type="list", formula1="=LISTS!$C$2:$C$5", allow_blank=True)
dv2.showErrorMessage = True; ws2.add_data_validation(dv2)
for rr in range(act_start, act_end+1): dv2.add(f"AH{rr}")

# CF on both tabs
for ws_t, rng in [(ws1, f"AF{ck_start}:AF{ck_end}"), (ws2, f"AH{act_start}:AH{act_end}")]:
    ws_t.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"CLOSED"'],
        fill=fi('C6EFCE'), font=Font(color='006100')))
    ws_t.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"IN PROGRESS"'],
        fill=fi('FFEB9C'), font=Font(color='9C6500')))
    ws_t.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"OPEN"'],
        fill=fi('FFC7CE'), font=Font(color='9C0006')))

# Print settings both tabs
for ws_t in [ws1, ws2]:
    ws_t.page_setup.paperSize = 9; ws_t.page_setup.orientation = 'portrait'
    ws_t.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws_t.page_setup.fitToWidth = 1; ws_t.page_setup.fitToHeight = 0

# Remove non-form sheets
for name in list(wb.sheetnames):
    if name not in ['NCR_FORM', 'ACTIONS', 'LISTS']:
        del wb[name]

# SAVE
OUT = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau/06-FRM-600/FRM-651_NCR_Report.xlsx"
wb.save(OUT)
sha = hashlib.sha256(open(OUT,'rb').read()).hexdigest()
print(f"FRM-651 CREATED: SHA={sha[:16]}...")
print(f"  Tab 1: NCR_FORM (visible)")
print(f"  Tab 2: ACTIONS (visible)")
print(f"  Tab 3: LISTS (hidden)")
print(f"  Output: {OUT}")
