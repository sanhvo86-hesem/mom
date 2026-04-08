# -*- coding: utf-8 -*-
"""
UPDATE MASTER TEMPLATE v2.0 -> v2.1
Consolidate all approved design rules into Design Notes + Example.
"""
import openpyxl, sys
from openpyxl.styles import Alignment
sys.stdout.reconfigure(encoding='utf-8')

TMPL = "C:/Users/TEST4/Desktop/frm-000-master-template.xlsx"
wb = openpyxl.load_workbook(TMPL)
ws = wb['00-DESIGN-NOTES']

# === UPDATE EXISTING ROWS ===

ws.cell(36, 2).value = (
    "4 merged zones per row: left-label(1-7) | left-input(8-20) | right-label(21-27) | right-input(28-40)\n"
    "Label: fill=#F1F5F9 (fog), 8pt bold #334155, left, NO indent, wrap_text=True.\n"
    "Input: fill=#FFFFFF (white), 8pt regular #0F172A, left, NO indent, wrap_text=True.\n"
    "ALL 40 cells in row must have borders set (openpyxl merged-cell requirement)."
)

ws.cell(37, 2).value = (
    "7-zone row. # cell: fog bg, bold, auto-number formula.\n"
    "Cat cell: category colour bg, bold text.\n"
    "Item: white(odd)/#F8FAFC(even), 8pt, left, NO indent, wrap_text=True.\n"
    "Criteria: fog bg, 7pt muted, left, NO indent, wrap_text=True.\n"
    "Result/Owner/Ref: white, 8pt, centre, wrap_text=True.\n"
    "ALL 40 cells in row must have borders set."
)

ws.cell(38, 2).value = (
    "h=4pt, fill=#FFFFFF, NO borders on ALL 40 cells.\n"
    "One spacer between every section block.\n"
    "Must clear borders on every cell (not just merge-start)."
)

ws.cell(39, 2).value = (
    "Label row (fog, 8pt bold, centre, wrap_text=True) + 3 sig rows x 26pt.\n"
    "Box: white, 'Name / Signature / Date' in CBD5E1, centre/bottom, wrap.\n"
    "THIN accent border around each individual box.\n"
    "ALL cells in each box must have borders set."
)

ws.cell(40, 2).value = (
    "h=24pt, fill=#EFF6FF (ice), 6-7pt #334155, NO indent, wrap_text=True.\n"
    "THIN accent border. No col-1 accent bar."
)

ws.cell(59, 2).value = "Segoe UI 8pt BOLD #334155 (mid-txt). Left, NO indent, wrap_text=True."
ws.cell(60, 2).value = "Segoe UI 8pt regular #0F172A (dark-txt). Left, NO indent, wrap_text=True."
ws.cell(63, 2).value = "Segoe UI 7pt regular #64748B (muted). Left, NO indent, wrap_text=True."
ws.cell(64, 2).value = "Segoe UI 6-7pt regular #334155 (mid-txt). Left, NO indent, wrap_text=True."

ws.cell(67, 2).value = (
    "ALL borders use #0078D4 (accent) and #E2E8F0 (silver) only.\n"
    "CRITICAL: openpyxl requires borders on ALL individual cells in a row,\n"
    "including cells inside merged ranges. Set borders BEFORE merging.\n"
    "Section outline = BLUE on all 4 edges. Internal dividers = SILVER.\n"
    "Spacer rows = NO borders on any cell."
)

ws.cell(68, 2).value = (
    "Section header: all 4 sides BLUE.\n"
    "First data row: top=BLUE. Last data row: bottom=BLUE.\n"
    "Left edge (col 1) of ALL data rows: left=BLUE.\n"
    "Right edge (col NC) of ALL data rows: right=BLUE.\n"
    "Each signature box: all 4 sides BLUE. Notice bar: all 4 sides BLUE."
)

ws.cell(69, 2).value = (
    "Bottom of middle data rows (not first, not last): bottom=SILVER.\n"
    "Right of zone-end columns WITHIN row (internal separators): right=SILVER.\n"
    "Meta row separators in header. Hairline under title."
)

print("Updated rows: 36-40, 59-60, 63-64, 67-69")

# === ADD NEW SECTION 16: MULTI-TAB + GLOBAL RULES ===

# R106 is merged A106:B106 — unmerge it first, then shift version history
ws.unmerge_cells('A106:B106')

# Shift version history R106-108 -> R113-115
for r in [108, 107, 106]:
    for c in [1, 2]:
        src_cell = ws.cell(r, c)
        if not isinstance(src_cell, openpyxl.cell.cell.MergedCell):
            ws.cell(r+7, c).value = src_cell.value

# Merge new section header
ws.merge_cells('A106:B106')
ws.cell(106, 1).value = "16. GLOBAL FORM RULES (v2.1)"

ws.cell(107, 1).value = "wrap_text global rule"
ws.cell(107, 2).value = (
    "ALL content cells must have wrap_text=True.\n"
    "If content exceeds cell width, it wraps within the cell.\n"
    "Prevents overflow and ensures print fidelity.\n"
    "Only exception: hidden data sheets (LISTS, REF_IMPORT etc.)."
)

ws.cell(108, 1).value = "No-indent global rule"
ws.cell(108, 2).value = (
    "ALL body cells (labels, inputs, criteria, meta values) use indent=0.\n"
    "No left padding saves space in narrow columns.\n"
    "Exception: section headers keep indent=1 for visual hierarchy."
)

ws.cell(109, 1).value = "Multi-tab header rule"
ws.cell(109, 2).value = (
    "EVERY visible sheet must have the full header block (rows 1-6 + row 7 spacer).\n"
    "Same form code, revision, owner, approved on all tabs.\n"
    "Title may add tab suffix: e.g. 'NCR REPORT - ACTION TRACKER'.\n"
    "Logo must be embedded on EVERY visible tab with same sizing rules."
)

ws.cell(110, 1).value = "Border ALL-cells rule"
ws.cell(110, 2).value = (
    "openpyxl: borders must be set on ALL individual cells in a row,\n"
    "including cells inside merged ranges. Set borders BEFORE merging.\n"
    "This is required for Excel to render borders correctly on merged cells."
)

ws.cell(111, 1).value = "Logo sizing formula (all formats)"
ws.cell(111, 2).value = (
    "zone_w = logo_cols x 19px. zone_h = 5 x 20px = 100px.\n"
    "logo_w = floor(zone_w x 0.78). logo_h = floor(logo_w / 3.46).\n"
    "Guard: if logo_h > zone_h x 0.78 then scale from height.\n"
    "x_off = floor((zone_w - logo_w) / 2) x 9525 EMU.\n"
    "y_off = floor((zone_h - logo_h) / 2) x 9525 EMU.\n"
    "A4P:8cols->118x34. A4L/A3P:11cols->163x47. A3L:16cols->237x68."
)

ws.cell(112, 1).value = "Section border anatomy"
ws.cell(112, 2).value = (
    "A section = header row + data rows, enclosed in BLUE border box.\n"
    "Header: BLUE all 4 sides. First data row: top=BLUE.\n"
    "Middle rows: top=NONE, bottom=SILVER, left=BLUE, right=BLUE.\n"
    "Last row: bottom=BLUE. Internal col separators: SILVER.\n"
    "Spacer (4pt white, no border) sits BETWEEN sections."
)

# Version history
ws.merge_cells('A113:B113')
ws.cell(113, 1).value = "17. VERSION HISTORY"

ws.cell(114, 1).value = "v2.1 (current)"
ws.cell(114, 2).value = (
    "wrap_text=True on ALL content cells. indent=0 on all body cells.\n"
    "Border ALL-cells rule for openpyxl merged-cell rendering.\n"
    "Multi-tab rules: header + logo on every visible sheet.\n"
    "Logo sizing formula for all 4 paper formats.\n"
    "Section border anatomy documented. Notice bar: 6pt, no indent."
)

ws.cell(115, 1).value = "v2.0"
ws.cell(115, 2).value = (
    "Multi-format: A4P / A4L / A3P / A3L masters.\n"
    "Design library: 13 groups A-M. Design notes: full spec."
)

# Title version bump
ws.cell(1, 1).value = "HESEM PRECISION DESIGN SYSTEM  -  MASTER SPECIFICATION  v2.1"
print("Added section 16 + version bump v2.0 -> v2.1")

# === UPDATE EXAMPLE FRM-202 ===
ws_ex = wb['05-EXAMPLE-FRM-202']
WL = Alignment(horizontal='left', vertical='center', indent=0, wrap_text=True)
WC = Alignment(horizontal='center', vertical='center', indent=0, wrap_text=True)

for r in list(range(9, 15)) + list(range(18, 30)) + list(range(32, 35)):
    for col in range(1, 41):
        c = ws_ex.cell(r, col)
        if c.value is not None:
            old_h = c.alignment.horizontal if c.alignment else 'left'
            c.alignment = WC if old_h == 'center' else WL

for col in range(1, 41):
    c = ws_ex.cell(37, col)
    if c.value: c.alignment = WC

c = ws_ex.cell(42, 1)
if c.value:
    c.alignment = Alignment(horizontal='left', vertical='center', indent=0, wrap_text=True)

print("Example FRM-202 updated: wrap_text + no-indent")

# === SAVE ===
wb.save(TMPL)
print(f"\nSaved: {TMPL}")
