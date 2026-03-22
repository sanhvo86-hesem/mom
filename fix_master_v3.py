#!/usr/bin/env python3
"""
HESEM QMS - Master Template Compliance Fix v3
═══════════════════════════════════════════════
Source of truth: frm-000-master-template.xlsx (00-FORM-DESIGN-SYSTEM)

CORRECT PALETTE (verified from actual file):
  #1B3A6B  PRIMARY     title text, meta form-code bold
  #0078D4  ACCENT      ALL borders, subtitle text
  #005A9E  ACCENT2     column header bg
  #EFF6FF  ICE         section header bg, notice bg
  #F1F5F9  FOG         label cells, # cells, meta labels
  #FFFFFF  WHITE       input cells, logo zone, title zone
  #F8FAFC  NEAR-WHITE  even-row stripe
  #E2E8F0  SILVER      internal separators
  #0F172A  DARK-TXT    body/input text
  #334155  MID-TXT     label text, meta label text, notice text
  #64748B  MUTED-TXT   tagline, criteria

HEADER: Row 1 HIDDEN, Rows 2-6 header, Row 7 = 4pt white spacer NO borders
BORDERS: Only THIN, two colors: #0078D4 (accent) and #E2E8F0 (silver)
"""
import os, sys, copy, zipfile, shutil, tempfile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

# ═══════════════════════════════════════════════════════════════════
# MASTER TEMPLATE CONSTANTS (from frm-000-master-template.xlsx)
# ═══════════════════════════════════════════════════════════════════
FONT = "Segoe UI"

# Colors
PRIMARY    = "1B3A6B"
ACCENT     = "0078D4"
ACCENT2    = "005A9E"
ICE        = "EFF6FF"
FOG        = "F1F5F9"
WHITE      = "FFFFFF"
NEAR_WHITE = "F8FAFC"
SILVER     = "E2E8F0"
DARK_TXT   = "0F172A"
MID_TXT    = "334155"
MUTED_TXT  = "64748B"

# Fills
F_WHITE     = PatternFill('solid', fgColor=WHITE)
F_ICE       = PatternFill('solid', fgColor=ICE)
F_FOG       = PatternFill('solid', fgColor=FOG)
F_ACCENT2   = PatternFill('solid', fgColor=ACCENT2)
F_NEAR_WHT  = PatternFill('solid', fgColor=NEAR_WHITE)
F_NONE      = PatternFill(fill_type=None)

# Borders (ONLY THIN, two colors)
B_ACCENT = Side(style='thin', color=ACCENT)
B_SILVER = Side(style='thin', color=SILVER)
B_NONE   = Side(style=None)

# Fonts
FN_TITLE    = Font(name=FONT, size=15, bold=True, color=PRIMARY)
FN_SUBTITLE = Font(name=FONT, size=8, bold=False, color=ACCENT)
FN_TAGLINE  = Font(name=FONT, size=7, bold=False, color=MUTED_TXT)
FN_META_LBL = Font(name=FONT, size=8, bold=True, color=MID_TXT)
FN_META_VAL = Font(name=FONT, size=8, bold=False, color=DARK_TXT)
FN_META_FC  = Font(name=FONT, size=8, bold=True, color=PRIMARY)  # Form Code value
FN_SEC_HDR  = Font(name=FONT, size=9, bold=True, color=PRIMARY)
FN_COL_HDR  = Font(name=FONT, size=9, bold=True, color=WHITE)
FN_LABEL    = Font(name=FONT, size=8, bold=True, color=MID_TXT)
FN_BODY     = Font(name=FONT, size=8, bold=False, color=DARK_TXT)
FN_CRITERIA = Font(name=FONT, size=7, bold=False, color=MUTED_TXT)
FN_NOTICE   = Font(name=FONT, size=7, bold=False, color=MID_TXT)
FN_AUTONUM  = Font(name=FONT, size=8, bold=True, color=PRIMARY)

# Alignments
A_CENTER = Alignment(horizontal='center', vertical='center', wrapText=True)
A_LEFT   = Alignment(horizontal='left', vertical='center', wrapText=True, indent=1)
A_RIGHT  = Alignment(horizontal='right', vertical='center', wrapText=True)

# Row heights
H_HEADER   = 15.0   # Rows 2-6
H_SPACER   = 4.0    # Row 7
H_SECTION  = 17.0   # Section headers
H_COL_HDR  = 17.0   # Column headers
H_DATA     = 20.0   # Data rows
H_SPACER_S = 4.0    # Inter-section spacers
H_SIG      = 26.0   # Signature rows
H_NOTICE   = 26.0   # Notice bars

# Column width
COL_WIDTH = 2.33

# NS for XML
NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'

# Logo file path (in design system folder)
LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
    '04-Bieu-Mau', '00-FORM-DESIGN-SYSTEM', 'hesem-logo.png')


# ═══════════════════════════════════════════════════════════════════
# HEADER ZONE MAPPING
# ═══════════════════════════════════════════════════════════════════
def get_zones(ncols):
    """Return header zone column boundaries (1-indexed)."""
    if ncols <= 40:  # A4P
        return {
            'logo': (1, 8), 'title': (9, 30),
            'meta_lbl': (31, 34), 'meta_val': (35, 40),
            'ncols': 40
        }
    elif ncols <= 56:  # A4L / A3P
        return {
            'logo': (1, 11), 'title': (12, 42),
            'meta_lbl': (43, 48), 'meta_val': (49, 56),
            'ncols': 56
        }
    else:  # A3L
        return {
            'logo': (1, 16), 'title': (17, 62),
            'meta_lbl': (63, 70), 'meta_val': (71, 82),
            'ncols': 82
        }


def detect_ncols(ws):
    """Detect grid width from merges or max_column."""
    max_col = 0
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 7:
            max_col = max(max_col, mg.max_col)
    if max_col == 0:
        max_col = ws.max_column or 40
    if max_col <= 44:
        return 40
    elif max_col <= 60:
        return 56
    else:
        return 82


def fill_hex(cell):
    """Get fill color hex from cell."""
    if cell.fill and cell.fill.fill_type == 'solid' and cell.fill.fgColor:
        rgb = str(cell.fill.fgColor.rgb or '')
        if len(rgb) == 8:
            return rgb[2:].upper()
        if len(rgb) == 6:
            return rgb.upper()
    return None


# ═══════════════════════════════════════════════════════════════════
# COLOR TRANSFORM MAP (old wrong colors → correct master template)
# ═══════════════════════════════════════════════════════════════════
FILL_MAP = {
    # Old section header fills → ICE
    '1565C0': ICE, '2C9CD7': ICE,
    # Old column header fills → ACCENT2
    '0C2D48': ACCENT2,
    # Old label fills → FOG
    'F0F7FC': FOG, 'E4F0F9': FOG,
    # Old stripe → NEAR_WHITE
    'F7FBFF': NEAR_WHITE,
    # Old notice → ICE
    'FFF8E1': ICE,
}

FONT_COLOR_MAP = {
    # Old body text → DARK_TXT
    '1A2733': DARK_TXT,
    # Old dark navy → PRIMARY
    '0C2D48': PRIMARY,
    # Old subtitle → ACCENT
    '1565C0': ACCENT,
    # Old muted → MUTED_TXT
    '7A8FA6': MUTED_TXT,
    # Old notice text → MID_TXT
    '7B4F00': MID_TXT,
    # Old sig text → MUTED_TXT
    'BBCCCC': MUTED_TXT,
}


def fix_cell_colors(cell):
    """Transform old colors to master template colors."""
    changed = False

    # Fix fill
    fh = fill_hex(cell)
    if fh and fh in FILL_MAP:
        new_color = FILL_MAP[fh]
        cell.fill = PatternFill('solid', fgColor=new_color)
        changed = True

        # If fill changed from dark to light, fix font color too
        # (white text on dark bg → dark text on light bg)
        if fh in ('1565C0', '2C9CD7') and cell.font:
            fc = None
            if cell.font.color and cell.font.color.rgb:
                fc = str(cell.font.color.rgb)
                if len(fc) == 8:
                    fc = fc[2:]
            if fc and fc.upper() in ('FFFFFF', 'FFFFFFFF'):
                # Section header: white → PRIMARY
                cell.font = Font(
                    name=FONT, size=cell.font.size or 9,
                    bold=True, color=PRIMARY
                )
            elif fc and fc.upper() in ('000000', 'FF000000', '00000000'):
                cell.font = Font(
                    name=FONT, size=cell.font.size or 9,
                    bold=cell.font.bold, color=PRIMARY
                )

    # Fix font color
    if cell.font and cell.font.color and cell.font.color.rgb:
        fc = str(cell.font.color.rgb)
        if len(fc) == 8:
            fc = fc[2:]
        fc = fc.upper()
        if fc in FONT_COLOR_MAP:
            new_fc = FONT_COLOR_MAP[fc]
            cell.font = Font(
                name=FONT,
                size=cell.font.size,
                bold=cell.font.bold,
                italic=cell.font.italic,
                underline=cell.font.underline,
                strike=cell.font.strikethrough,
                color=new_fc
            )
            changed = True

    return changed


def fix_cell_borders(cell):
    """Transform old border colors to master template."""
    if not cell.border:
        return False

    old_to_new = {
        '2C9CD7': ACCENT, '0078D4': ACCENT,
        'C5D9E8': SILVER, 'E2E8F0': SILVER,
        'F9A825': ACCENT,  # old notice border → accent
        '0C2D48': ACCENT,
        '1565C0': ACCENT,
    }

    changed = False
    sides = {}
    for sn in ['top', 'bottom', 'left', 'right']:
        side = getattr(cell.border, sn, None)
        if side and side.style:
            # Convert all medium → thin
            new_style = 'thin' if side.style in ('medium', 'thick') else side.style
            if new_style != side.style:
                changed = True

            # Fix color
            if side.color and side.color.rgb:
                sc = str(side.color.rgb)
                if len(sc) == 8:
                    sc = sc[2:]
                sc = sc.upper()
                if sc in old_to_new:
                    sides[sn] = Side(style=new_style, color=old_to_new[sc])
                    changed = True
                else:
                    sides[sn] = Side(style=new_style, color=side.color)
            else:
                sides[sn] = Side(style=new_style, color=ACCENT)
                changed = True
        else:
            sides[sn] = side or B_NONE

    if changed:
        cell.border = Border(
            top=sides.get('top', B_NONE),
            bottom=sides.get('bottom', B_NONE),
            left=sides.get('left', B_NONE),
            right=sides.get('right', B_NONE)
        )
    return changed


def fix_cell_font(cell):
    """Fix font family to Segoe UI."""
    if cell.font and cell.font.name and cell.font.name != FONT:
        cell.font = Font(
            name=FONT,
            size=cell.font.size,
            bold=cell.font.bold,
            italic=cell.font.italic,
            underline=cell.font.underline,
            strike=cell.font.strikethrough,
            color=cell.font.color
        )
        return True
    return False


def fix_header_structure(ws, ncols):
    """Ensure rows 1-7 follow master template structure."""
    z = get_zones(ncols)
    changed = False

    # Row 1: HIDDEN
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[1].height = None

    # Rows 2-6: 15pt
    for r in range(2, 7):
        if ws.row_dimensions[r].height != H_HEADER:
            ws.row_dimensions[r].height = H_HEADER
            changed = True

    # Row 7: 4pt white spacer, NO borders
    ws.row_dimensions[7].height = H_SPACER
    for c in range(1, z['ncols'] + 1):
        cell = ws.cell(row=7, column=c)
        try:
            cell.fill = F_WHITE
            cell.border = Border()
            cell.value = None
        except AttributeError:
            pass  # MergedCell - skip

    return changed


def fix_header_cells(ws, ncols):
    """Fix header cell formatting (zones 1-3)."""
    z = get_zones(ncols)
    n = z['ncols']

    def safe_set(cell, **kwargs):
        """Safely set cell properties, skip MergedCell."""
        try:
            for k, v in kwargs.items():
                setattr(cell, k, v)
        except AttributeError:
            pass

    # Logo zone (fill white)
    for r in range(2, 7):
        for c in range(z['logo'][0], z['logo'][1] + 1):
            safe_set(ws.cell(row=r, column=c), fill=F_WHITE)

    # Title zone formatting
    for r in range(2, 5):
        for c in range(z['title'][0], z['title'][1] + 1):
            cell = ws.cell(row=r, column=c)
            try:
                if cell.value:
                    cell.font = FN_TITLE
                    cell.alignment = A_CENTER
                cell.fill = F_WHITE
            except AttributeError:
                pass

    # Subtitle (row 5 in title zone)
    for c in range(z['title'][0], z['title'][1] + 1):
        cell = ws.cell(row=5, column=c)
        try:
            if cell.value:
                cell.font = FN_SUBTITLE
                cell.alignment = A_CENTER
            cell.fill = F_WHITE
        except AttributeError:
            pass

    # Tagline (row 6 in title zone)
    for c in range(z['title'][0], z['title'][1] + 1):
        cell = ws.cell(row=6, column=c)
        try:
            if cell.value:
                cell.font = FN_TAGLINE
                cell.alignment = A_CENTER
            cell.fill = F_WHITE
        except AttributeError:
            pass

    # Meta labels
    for r in range(2, 7):
        for c in range(z['meta_lbl'][0], z['meta_lbl'][1] + 1):
            cell = ws.cell(row=r, column=c)
            try:
                cell.fill = F_FOG
                cell.alignment = A_RIGHT
                if cell.value:
                    cell.font = FN_META_LBL
            except AttributeError:
                pass

    # Meta values
    for r in range(2, 7):
        for c in range(z['meta_val'][0], z['meta_val'][1] + 1):
            cell = ws.cell(row=r, column=c)
            try:
                cell.fill = F_WHITE
                cell.alignment = A_LEFT
                if cell.value:
                    cell.font = FN_META_FC if r == 2 else FN_META_VAL
            except AttributeError:
                pass

    # Header borders
    for r in range(2, 7):
        for c in range(1, n + 1):
            cell = ws.cell(row=r, column=c)
            top = B_ACCENT if r == 2 else B_NONE
            bottom = B_ACCENT if r == 6 else B_SILVER if c >= z['meta_lbl'][0] else B_NONE
            left = B_ACCENT if c == 1 else B_NONE
            right = B_ACCENT if c == n else B_NONE
            if c == z['logo'][1]:
                right = B_ACCENT
            if c == z['title'][1]:
                right = B_ACCENT
            if c == z['meta_lbl'][1]:
                right = B_SILVER
            safe_set(cell, border=Border(top=top, bottom=bottom, left=left, right=right))


def fix_logo_images(ws, ncols):
    """Remove all images and add single centered logo."""
    # Remove all existing images
    ws._images = []

    # Add logo if file exists
    logo_path = LOGO_PATH
    if not os.path.exists(logo_path):
        # Try PNG version
        for ext in ['.png', '.svg', '.jpg', '.jpeg']:
            alt = logo_path.rsplit('.', 1)[0] + ext
            if os.path.exists(alt):
                logo_path = alt
                break

    if not os.path.exists(logo_path):
        return False

    try:
        img = XlImage(logo_path)

        # Set size based on paper format
        if ncols <= 40:  # A4P
            img.width = 118
            img.height = 34
        elif ncols <= 56:  # A4L / A3P
            img.width = 163
            img.height = 46
        else:  # A3L
            img.width = 237
            img.height = 68

        # Anchor at A2 (row index 1 in 0-based)
        img.anchor = 'A2'
        ws.add_image(img)
        return True
    except Exception as e:
        print(f"    Logo error: {e}")
        return False


def fix_sheet(ws, ncols, is_main=True):
    """Fix a single sheet's formatting."""
    z = get_zones(ncols)
    n = z['ncols']
    issues = []

    # 1. Column widths
    for c in range(1, n + 1):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTH

    # 2. Header structure
    fix_header_structure(ws, ncols)
    fix_header_cells(ws, ncols)
    issues.append("header")

    # 3. Logo
    if fix_logo_images(ws, ncols):
        issues.append("logo")

    # 4. Gridlines
    try:
        ws.sheet_view.showGridLines = False
    except:
        pass

    # 5. Page setup
    try:
        pm = ws.page_margins
        pm.left = pm.right = pm.top = pm.bottom = 0.197
        pm.header = pm.footer = 0.1
        ps = ws.page_setup
        ps.paperSize = 9 if ncols <= 56 else 8
        ps.fitToWidth = 1
        ps.fitToHeight = 0
    except:
        pass

    # 6. Remove freeze panes
    try:
        ws.freeze_panes = None
    except:
        pass

    # 7. Fix body area (rows 8+)
    max_row = ws.max_row or 50
    color_count = 0
    border_count = 0
    font_count = 0

    # Build merged secondary set
    merged_secondary = set()
    for mg in ws.merged_cells.ranges:
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r != mg.min_row or c != mg.min_col:
                    merged_secondary.add((r, c))

    for row_cells in ws.iter_rows(min_row=8, max_row=max_row):
        for cell in row_cells:
            if (cell.row, cell.column) in merged_secondary:
                continue

            try:
                # Fix colors
                if fix_cell_colors(cell):
                    color_count += 1
                # Fix borders
                if fix_cell_borders(cell):
                    border_count += 1
                # Fix font
                if fix_cell_font(cell):
                    font_count += 1
            except AttributeError:
                pass  # MergedCell

    if color_count:
        issues.append(f"{color_count} colors")
    if border_count:
        issues.append(f"{border_count} borders")
    if font_count:
        issues.append(f"{font_count} fonts")

    # 8. Fix notice bars
    notice_count = 0
    for row in range(8, max_row + 1):
        cell = ws.cell(row=row, column=1)
        val = str(cell.value or '').upper()
        if val.startswith('NOTICE') or val.startswith('NOTE:'):
            ws.row_dimensions[row].height = H_NOTICE
            for c in range(1, n + 1):
                nc = ws.cell(row=row, column=c)
                try:
                    nc.fill = F_ICE
                    nc.font = FN_NOTICE
                    nc.alignment = Alignment(horizontal='left', vertical='center',
                                             wrapText=True, indent=2)
                    nc.border = Border(
                        top=B_ACCENT, bottom=B_ACCENT,
                        left=B_ACCENT if c == 1 else B_NONE,
                        right=B_ACCENT if c == n else B_NONE
                    )
                except AttributeError:
                    pass
            notice_count += 1
    if notice_count:
        issues.append(f"{notice_count} notices")

    # 9. Fix named styles
    try:
        for nst in ws.parent._named_styles:
            if nst.font and nst.font.name != FONT:
                nst.font = Font(name=FONT, size=nst.font.size,
                               bold=nst.font.bold, color=nst.font.color)
    except:
        pass

    return issues


def fix_sheetviews_xml(filepath):
    """Fix sheetViews XML corruption."""
    fixed = False
    tmp = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            zf.extractall(tmp)

        ws_dir = os.path.join(tmp, 'xl', 'worksheets')
        if os.path.isdir(ws_dir):
            for sf in os.listdir(ws_dir):
                if not sf.endswith('.xml'):
                    continue
                sp = os.path.join(ws_dir, sf)
                ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
                ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
                tree = ET.parse(sp)
                root = tree.getroot()
                svs = root.find(f'{NS}sheetViews')
                if svs is None:
                    continue
                for sv in svs.findall(f'{NS}sheetView'):
                    pane = sv.find(f'{NS}pane')
                    if pane is not None:
                        sv.remove(pane)
                        fixed = True
                    for sel in sv.findall(f'{NS}selection'):
                        if sel.get('pane') and sel.get('pane') != 'topLeft':
                            sv.remove(sel)
                            fixed = True
                    if not sv.findall(f'{NS}selection'):
                        se = ET.SubElement(sv, f'{NS}selection')
                        se.set('activeCell', 'A1')
                        se.set('sqref', 'A1')
                tree.write(sp, xml_declaration=True, encoding='UTF-8')

        if fixed:
            with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                for rd, ds, fs in os.walk(tmp):
                    for ff in fs:
                        fp = os.path.join(rd, ff)
                        zf.write(fp, os.path.relpath(fp, tmp))
    except Exception as e:
        print(f"    XML error: {e}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    return fixed


def process_workbook(filepath):
    """Process one workbook."""
    fname = os.path.basename(filepath)
    all_issues = []

    # Phase 1: XML fix
    if fix_sheetviews_xml(filepath):
        all_issues.append("  XML: pane fix")

    # Phase 2: openpyxl fixes
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, [f"  ERROR: {e}"]

    for idx, sn in enumerate(wb.sheetnames):
        ws = wb[sn]
        ncols = detect_ncols(ws)

        sheet_issues = fix_sheet(ws, ncols, is_main=(idx == 0))

        if sheet_issues:
            all_issues.append(f"  [{sn}] {', '.join(sheet_issues)}")

    try:
        wb.save(filepath)
        all_issues.append("  >> SAVED")
    except Exception as e:
        all_issues.append(f"  >> SAVE ERROR: {e}")

    wb.close()

    # Phase 3: Post-save XML fix (openpyxl may reintroduce panes)
    fix_sheetviews_xml(filepath)

    return fname, all_issues


def main():
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS - MASTER TEMPLATE COMPLIANCE v3")
    print(f"  Source: frm-000-master-template.xlsx")
    print(f"  Palette: #1B3A6B/#0078D4/#005A9E/#EFF6FF/#F1F5F9")
    print(f"  Borders: THIN only, accent #0078D4 + silver #E2E8F0")
    print(f"  Row 1: HIDDEN | Row 7: 4pt white spacer")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    v = '-v' in sys.argv or '--verbose' in sys.argv
    fixed = errors = 0

    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        print(f"[{i:3d}/{len(files)}] {fn[:55]:55s}", end=" ", flush=True)

        name, issues = process_workbook(fp)

        has_err = any('ERROR' in str(x) for x in issues)
        if has_err:
            print("ERROR")
            errors += 1
        else:
            print("OK")
            fixed += 1

        if v:
            for iss in issues:
                print(iss)

    print(f"\n{'='*70}")
    print(f"DONE: {fixed} fixed, {errors} errors out of {len(files)}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
