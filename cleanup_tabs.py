#!/usr/bin/env python3
"""
HESEM QMS - Delete non-operational tabs.
Keep ONLY tabs needed for:
  1. Shop floor operation (filling in data)
  2. ISO/AS9100 compliance (mandatory records)
  3. Dropdown validation (LISTS - hidden)

DELETE everything else: Evidence, Import, Reference, Print, Review summaries etc.
"""
import os, sys
from openpyxl import load_workbook

# Tabs to ALWAYS DELETE (case-insensitive partial match)
DELETE_PATTERNS = [
    'evidence',       # Evidence ref tabs - admin metadata
    'ref_import',     # Import guidance - not operational
    'import_ref',     # Same
    '_import',        # Import data
    'ref_links',      # Reference links - not operational
    'source_ref',     # Source reference - not operational
    'print_control',  # Print system
    'print_face',     # Print system
    'print_log',      # Print system
    'printctl',       # Print control
    'generator',      # Generator system
    'calc_summary',   # Calculation backend
    'result_detail',  # Result detail backend
    'raw_data',       # Raw data backend
    'chart_data',     # Chart data backend
    'review_summary', # Periodic review - admin
    'filter_view',    # System view
    'annex_map',      # System mapping
    'rankref',        # Reference lookup
    'reactionref',    # Reference lookup
    'roleref',        # Reference lookup
    'certref',        # Reference lookup
    'sessionref',     # Reference lookup
    'taskref',        # Reference lookup
    'criteria',       # Criteria reference
    'expiryq',        # Expiry query
    'revhist',        # Revision history (kept in doc control)
    'rev_history',    # Same
    'signoffs',       # Redundant with main form sign-off
    'inputrefs',      # Input references
    'codes',          # Code reference
]

# Tabs to ALWAYS KEEP (exact or partial match)
KEEP_PATTERNS = [
    'lists', 'zlist', '_list',  # Dropdowns - essential
    'gate', 'main', 'form', 'log', 'register', 'track',  # Main forms
    'action', 'finding', 'capa', '8d', 'ncr',  # CAPA/Actions
    'audit', 'inspect', 'setup', 'dispatch', 'ship',  # Operations
    'cost', 'score', 'scar', 'survey',  # Functional forms
    'plan', 'matrix', 'attend', 'ojt', 'assess',  # HR forms
    'incident', 'event', 'crash', 'hold',  # Event forms
    'data_log', 'verify', 'clean', 'leak', 'entry',  # Operational
    'tool_list', 'sketch', 'step', 'comparison',  # Support
    'containment', 'shortage', 'mismatch', 'blocker',  # Exception
    'risk', 'impact', 'legend', 'gap',  # Analytical
    'aging', 'bucket', 'summary', 'trend',  # Reporting
    'milestone', 'kpi', 'decision', 'recovery',  # Management
    'baseline', 'sample', 'function_test',  # Engineering
    'obligation', 'control_overlay', 'char_result',  # Quality
    'matl_proc', 'revalidation', 'fp_result', 'adjust',  # Production
    'changeover', 'smed', 'pm', 'safety',  # Maintenance
    'capability', 'spc', 'grr', 'bls', 'attr_msa',  # MSA/SPC
    'aql', 'spec_char', 'certlog', 'lighting',  # Specific
    'dev_plan', 'invoice', 'shipref', 'target',  # Misc operational
    'open_item', 'closure', 'return', 'audience',  # Misc
    'pilot', 'followup', 'impl', 'comm',  # Process
    'swot', 'csr', 'climate', 'interested',  # Planning
    'lesson', 'config', 'disruption', 'access',  # System
    'po_data', 'po_exception', 'supplier',  # Purchasing
    'flowdown', 'incoming', 'outsource',  # External
    'period_review', 'score_input', 'scorecard',  # Scoring
]


def should_delete(tab_name):
    """Determine if a tab should be deleted."""
    lower = tab_name.lower()

    # Never delete LISTS tabs
    if 'list' in lower:
        return False

    # Check delete patterns
    for pattern in DELETE_PATTERNS:
        if pattern in lower:
            return True

    return False


def process_workbook(filepath):
    """Process a single workbook - delete non-operational tabs."""
    fname = os.path.basename(filepath)
    code = fname.split('_')[0]

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, 0, 0, f"LOAD ERROR: {e}"

    original_count = len(wb.sheetnames)
    deleted = []

    # Never delete the first sheet
    sheets_to_check = wb.sheetnames[1:]  # Skip first

    for sn in sheets_to_check:
        if should_delete(sn):
            # Safety: don't delete if it's the only visible sheet
            visible = [s for s in wb.sheetnames if wb[s].sheet_state == 'visible']
            if len(visible) <= 1:
                continue
            try:
                del wb[sn]
                deleted.append(sn)
            except Exception as e:
                pass

    if deleted:
        try:
            wb.save(filepath)
        except Exception as e:
            wb.close()
            return fname, original_count, 0, f"SAVE ERROR: {e}"

    wb.close()
    return fname, original_count, len(deleted), deleted


def main():
    base = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print("HESEM QMS - TAB CLEANUP (Delete non-operational tabs)")
    print(f"Files: {len(files)}")
    print(f"{'='*70}")

    total_deleted = 0
    total_original = 0

    for i, fp in enumerate(files, 1):
        fname, orig, ndel, info = process_workbook(fp)
        total_original += orig
        total_deleted += ndel

        if ndel > 0:
            deleted_names = ', '.join(info) if isinstance(info, list) else info
            print(f"[{i:3d}] {fname[:45]:45s} {orig}>{orig-ndel} tabs  DEL: {deleted_names}")

    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"  Original total tabs: {total_original}")
    print(f"  Deleted tabs:        {total_deleted}")
    print(f"  Remaining tabs:      {total_original - total_deleted}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
