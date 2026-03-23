"""
FRM-202 Contract Review Checklist - Full Upgrade V0 -> V1
CNC Job-Order Precision Machining / Semiconductor Equipment Parts
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

from upgrade_lib import (
    COLORS, FONT_LABEL, FONT_INPUT, FONT_BODY, FONT_CAT, FONT_NUM,
    FILL_ICE, FILL_FOG, FILL_WHITE, FILL_NEAR_WHITE, FILL_COL_HDR,
    FILL_CF_RED, FILL_CF_AMBER, FILL_CF_GREEN,
    THIN_ACCENT, THIN_SILVER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_LEFT_WRAP,
    H_DATA,
    add_result_conditional_formatting, add_status_conditional_formatting,
    add_revision_history_sheet, backup_and_save, bump_revision,
    ensure_lists_values,
)

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"
FORM_PATH = f"{BASE}/02-FRM-200/FRM-202_Contract_Review_Checklist.xlsx"

print(f"Loading {FORM_PATH}...")
wb = openpyxl.load_workbook(FORM_PATH)
ws = wb['FRM-202_GATE']
ws_lists = wb['_LISTS']
ws_actions = wb['ACTIONS']
NCOLS = 40

# ============================================================================
# 1. INSERT 7 ROWS after R31 (last checklist item), before R32 (spacer)
# ============================================================================
print("Step 1: Inserting 7 new rows after R31...")
INSERT_AT = 32
TOTAL_INSERT = 7
ws.insert_rows(INSERT_AT, amount=TOTAL_INSERT)

# After insertion: R32-R38 = new, R39 = old spacer, R40+ shifted

# ============================================================================
# 2. CNC-SPECIFIC CHECKLIST ITEMS (R32-R36)
# ============================================================================
print("Step 2: Adding 5 CNC-specific checklist items...")

CK_ZONES = [(1,2), (3,5), (6,19), (20,31), (32,35), (36,38), (39,40)]

def make_ck_row(row, num, cat, cat_color, item, criteria, even=False):
    for s, e in CK_ZONES:
        if e > s:
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    ws.row_dimensions[row].height = H_DATA

    c = ws.cell(row=row, column=1)
    c.value = num; c.font = FONT_NUM; c.fill = FILL_FOG
    c.alignment = ALIGN_CENTER
    c.border = Border(left=THIN_ACCENT, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=3)
    c.value = cat; c.font = FONT_CAT
    c.fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type='solid')
    c.alignment = ALIGN_CENTER
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=6)
    c.value = item; c.font = FONT_BODY
    c.fill = FILL_NEAR_WHITE if even else FILL_WHITE
    c.alignment = ALIGN_LEFT_WRAP
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=20)
    c.value = criteria; c.font = FONT_BODY; c.fill = FILL_FOG
    c.alignment = ALIGN_LEFT_WRAP
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    for col in [32, 36, 39]:
        c = ws.cell(row=row, column=col)
        c.font = FONT_INPUT; c.fill = FILL_WHITE; c.alignment = ALIGN_CENTER
        c.border = Border(left=THIN_SILVER, right=THIN_SILVER if col != 39 else THIN_ACCENT,
                          top=THIN_ACCENT, bottom=THIN_ACCENT)
        c.protection = Protection(locked=False)

cnc_items = [
    (11, 'TEC', COLORS['CAT_TEC'],
     'CNC program revision matches released drawing revision.',
     'Program rev, drawing rev, setup sheet rev all align.', False),
    (12, 'TEC', COLORS['CAT_TEC'],
     'Tight-tolerance features identified (< 0.025 mm).',
     'Feature count flagged; measurement method confirmed.', True),
    (13, 'QUA', COLORS['CAT_QUA'],
     'Semiconductor cleanliness class confirmed.',
     'Clean-room / clean-pack / vacuum-bag level specified or NA.', False),
    (14, 'QUA', COLORS['CAT_QUA'],
     'Material cert / CoC / test-report requirement specified.',
     'Cert type, mill cert, RoHS, REACH, or conflict minerals.', True),
    (15, 'CAP', COLORS['CAT_CAP'],
     'Outsource processes (HT/plate/anod) confirmed with lead time.',
     'Supplier approved, lead time fits delivery promise.', False),
]

for i, (num, cat, cc, item, crit, even) in enumerate(cnc_items):
    make_ck_row(32 + i, num, cat, cc, item, crit, even)
    print(f"  + Item #{num}: {item[:50]}...")

# ============================================================================
# 3. CROSS-FORM LINKAGE + PARENT PROCEDURE REF (R37-R38)
# ============================================================================
print("Step 3: Adding cross-form linkage and parent procedure ref...")

REF_ZONES = [(1,7), (8,20), (21,27), (28,40)]

def make_ref_row(row, l_label, r_label, l_value=None, r_value=None, locked=False):
    for s, e in REF_ZONES:
        if e > s:
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    ws.row_dimensions[row].height = H_DATA

    c = ws.cell(row=row, column=1)
    c.value = l_label; c.font = FONT_LABEL; c.fill = FILL_FOG
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_ACCENT, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=8)
    c.value = l_value; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)

    c = ws.cell(row=row, column=21)
    c.value = r_label; c.font = FONT_LABEL; c.fill = FILL_FOG
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=28)
    c.value = r_value; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_ACCENT, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)


make_ref_row(37, 'Linked FRM-203 Job #', 'Linked FRM-204 Kickoff #')
make_ref_row(38, 'Ref: SOP-201 / WI-201~203', 'Retain: 10 years',
             l_value='SOP-201 (RFQ, Contract Review, Order Kickoff)',
             r_value='Per QMS retention schedule', locked=True)

# ============================================================================
# 4. UPDATE GATE FORMULAS (now AF22:AF36 instead of AF22:AF31)
# ============================================================================
print("Step 4: Updating gate formulas for 15 items...")

ws.cell(row=15, column=28).value = (
    '=IF(COUNTA(AF22:AF36)=0,"",IF(COUNTIF(AF22:AF36,"FAIL")'
    '+COUNTIF(AF22:AF36,"HOLD")>0,"HOLD",'
    'IF(COUNTIF(AF22:AF36,"PASS")=COUNTA(AF22:AF36),"APPROVED","REVIEW")))'
)
ws.cell(row=18, column=28).value = (
    '=IF(COUNTA(AF22:AF36)=0,"",IF(COUNTIF(AF22:AF36,"FAIL")'
    '+COUNTIF(AF22:AF36,"HOLD")>0,"OPEN","CLOSED"))'
)

# ============================================================================
# 5. DATA VALIDATIONS for new items
# ============================================================================
print("Step 5: Adding data validations...")

for r in range(32, 37):
    dv = DataValidation(type="list", formula1="=VAL_RESULT_CORE", allow_blank=True)
    dv.error = "Select PASS, HOLD, FAIL, or NA."
    dv.errorTitle = "Invalid"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"AF{r}")

# ============================================================================
# 6. CONDITIONAL FORMATTING
# ============================================================================
print("Step 6: Adding conditional formatting...")

add_result_conditional_formatting(ws, "AF22:AI36")
add_result_conditional_formatting(ws, "AB15:AN15")
add_status_conditional_formatting(ws, "AB18:AN18")
add_status_conditional_formatting(ws_actions, "G17:G26")

# ============================================================================
# 7. ENRICH _LISTS
# ============================================================================
print("Step 7: Enriching _LISTS...")

ensure_lists_values(ws_lists, 'VAL_YES_NO', ['YES', 'NO', 'NA'])
ensure_lists_values(ws_lists, 'VAL_DISCOVERY_STAGE',
    ['INCOMING', 'SETUP', 'FIRST PIECE', 'IN-PROCESS',
     'FINAL INSPECTION', 'PACKAGING', 'SHIPPING', 'CUSTOMER'])
ensure_lists_values(ws_lists, 'VAL_CK_CATEGORY',
    ['COM', 'TEC', 'QUA', 'CAP', 'CON', 'DOC', 'REL', 'CNC'])

# ============================================================================
# 8. REVISION HISTORY + BUMP REV
# ============================================================================
print("Step 8: Adding revision history and bumping V0 -> V1...")

add_revision_history_sheet(wb, 'FRM-202')
bump_revision(ws, 'V1', '2026-03-22')
bump_revision(ws_actions, 'V1', '2026-03-22')

# ============================================================================
# 9. PROTECT FORMULAS
# ============================================================================
print("Step 9: Protecting formula cells...")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            cell.protection = Protection(locked=True)

# ============================================================================
# 10. UPDATE NOTICE BAR
# ============================================================================
print("Step 10: Updating notice bar...")

for r in range(ws.max_row, max(ws.max_row - 10, 1), -1):
    c = ws.cell(row=r, column=1)
    if c.value and 'NOTICE' in str(c.value):
        c.value = (
            "NOTICE  -  Enter data only in designated white input cells. "
            "Do not add, delete or resize columns/rows. "
            "File naming: FRM-202_{Customer}_{PO}.xlsx. "
            "Ref: SOP-201 / WI-201, WI-202, WI-203. "
            "Retain 10 years per QMS retention schedule."
        )
        print(f"  Updated notice at R{r}")
        break

# ============================================================================
# 11. SAVE
# ============================================================================
print("Step 11: Saving...")
sha = backup_and_save(wb, FORM_PATH)

print(f"""
{'='*70}
FRM-202 CONTRACT REVIEW CHECKLIST - UPGRADE COMPLETE
{'='*70}
Version:    V0 -> V1
Date:       2026-03-22

Changes:
  + 5 CNC-specific checklist items (#11-#15)
  + Cross-form linkage: FRM-203, FRM-204
  + Parent procedure ref: SOP-201 / WI-201~203
  + Records retention: 10 years
  + Conditional formatting: FAIL/RED, HOLD/AMBER, PASS/GREEN
  + Gate formulas updated for 15 items
  + Data validations on all new items
  + _LISTS enriched with YES_NO, DISCOVERY_STAGE, CK_CATEGORY
  + Revision history sheet
  + Formula cells protected

SHA-256: {sha}
{'='*70}
""")
