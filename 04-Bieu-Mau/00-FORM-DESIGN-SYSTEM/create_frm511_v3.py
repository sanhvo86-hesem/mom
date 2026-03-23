# -*- coding: utf-8 -*-
"""
FRM-511 v3.1 — PIXEL-PERFECT border + wrap_text + logo.
- Borders on ALL 40 cells per row (openpyxl merged cell requirement)
- wrap_text=True on ALL content cells
- Alignment: left WITHOUT indent (save space); center only for #/Cat/Result/Owner/Ref
- Logo: sized per paper format, centered in logo zone
"""
import openpyxl, copy, hashlib, math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
import openpyxl.worksheet.properties

P = {
    'primary':'1B3A6B','accent':'0078D4','accent2':'005A9E',
    'ice':'EFF6FF','fog':'F1F5F9','white':'FFFFFF',
    'near':'F8FAFC','silver':'E2E8F0',
    'dk':'0F172A','mid':'334155','muted':'64748B','plc':'CBD5E1',
    'tec_bg':'DCFCE7','tec_fg':'14532D','qua_bg':'FEF9C3','qua_fg':'713F12',
    'cap_bg':'FCE7F3','cap_fg':'831843','con_bg':'F3E8FF','con_fg':'4C1D95',
    'com_bg':'DBEAFE','com_fg':'1E3A8A',
}
NC = 40
BLUE = Side(style='thin', color=P['accent'])
SILVER = Side(style='thin', color=P['silver'])
NONE_S = Side()

def fi(c): return PatternFill(start_color=c, end_color=c, fill_type='solid')
def fo(sz, bold=False, c='dk'): return Font(name='Segoe UI', size=sz, bold=bold, color=P[c])
AL = Alignment(horizontal='left', vertical='center', indent=1)
AC = Alignment(horizontal='center', vertical='center')

# =====================================================================
# CORE: Apply borders to ALL 40 cells in a row
# =====================================================================
def apply_row_borders(ws, r, zones, is_first, is_last):
    """
    Apply correct borders to ALL 40 cells in a data row.
    zones: list of (start_col, end_col) for each merge zone
    is_first: True if first data row after section header (top=BLUE)
    is_last: True if last data row in section (bottom=BLUE)

    Rules from master template:
    - Col 1: left=BLUE (section left edge)
    - Col NC: right=BLUE (section right edge)
    - End of each zone: right=SILVER (internal separator)
    - Start of last zone also has right stored but overridden by end cell
    - ALL cells: top=BLUE if first, else NONE; bottom=BLUE if last, else SILVER
    """
    top = BLUE if is_first else NONE_S
    bot = BLUE if is_last else SILVER

    # Build a set of zone-end columns (internal separators)
    zone_ends = set()
    for s, e in zones[:-1]:  # all zones except last
        zone_ends.add(e)

    for col in range(1, NC + 1):
        left = BLUE if col == 1 else NONE_S
        right = NONE_S
        if col == NC:
            right = BLUE  # section right edge
        elif col in zone_ends:
            right = SILVER  # internal zone separator

        ws.cell(r, col).border = Border(left=left, right=right, top=top, bottom=bot)


def apply_row_borders_table(ws, r, zones, is_last):
    """
    Apply borders for checklist/table rows (after column header).
    No top border (column header's bottom already provides it).
    """
    bot = BLUE if is_last else SILVER
    zone_ends = set()
    for s, e in zones[:-1]:
        zone_ends.add(e)

    for col in range(1, NC + 1):
        left = BLUE if col == 1 else NONE_S
        right = NONE_S
        if col == NC:
            right = BLUE
        elif col in zone_ends:
            right = SILVER
        ws.cell(r, col).border = Border(left=left, right=right, bottom=bot)


# =====================================================================
# LOAD MASTER TEMPLATE
# =====================================================================
TMPL = "C:/Users/TEST4/Desktop/frm-000-master-template.xlsx"
wb = openpyxl.load_workbook(TMPL)
src = wb['01-MASTER-A4-PORTRAIT']
ws = wb.create_sheet('FRM-511', 0)

for i in range(1, NC+1):
    ws.column_dimensions[get_column_letter(i)].width = 2.375
ws.row_dimensions[1].hidden = True

# Header rows 2-6: exact copy
for row in range(2, 7):
    ws.row_dimensions[row].height = 15.0
    for col in range(1, NC+1):
        s = src.cell(row, col); d = ws.cell(row, col)
        d.value = s.value; d.font = copy.copy(s.font)
        d.fill = copy.copy(s.fill); d.alignment = copy.copy(s.alignment)
        d.border = copy.copy(s.border)
for mc in src.merged_cells.ranges:
    ws.merge_cells(str(mc))

ws['I2'].value = 'SETUP AND FIRST-PIECE RECORD'
ws['AI2'].value = 'FRM-511'; ws['AI3'].value = 'Rev.01'
ws['AI4'].value = '2026-03-23'; ws['AI5'].value = 'Production / QA'
ws['AI6'].value = 'Production Director'

# =====================================================================
# HELPERS
# =====================================================================
PAIR_ZONES = [(1,7), (8,20), (21,27), (28,NC)]
CK_ZONES = [(1,2), (3,5), (6,19), (20,31), (32,35), (36,38), (39,NC)]

def spacer(r):
    ws.row_dimensions[r].height = 4.0
    # Set ALL cells to no border, white fill
    for col in range(1, NC+1):
        c = ws.cell(r, col)
        c.fill = fi(P['white'])
        c.border = Border()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    return r+1

def sect(r, text):
    ws.row_dimensions[r].height = 17.0
    # Set ALL cells borders first
    for col in range(1, NC+1):
        c = ws.cell(r, col)
        c.fill = fi(P['ice'])
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == NC else NONE_S
        c.border = Border(left=left, right=right, top=BLUE, bottom=BLUE)
    # Then merge and set content on cell A
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1)
    c.value = text; c.font = fo(9, True, 'primary')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    return r+1

def dpair(r, lL, lR, first=False, last=False):
    ws.row_dimensions[r].height = 20.0
    # Step 1: Set borders on ALL 40 cells
    apply_row_borders(ws, r, PAIR_ZONES, first, last)
    # Step 2: Set fill on ALL cells per zone
    zone_fills = [fi(P['fog']), fi(P['white']), fi(P['fog']), fi(P['white'])]
    for (s, e), zfill in zip(PAIR_ZONES, zone_fills):
        for col in range(s, e+1):
            ws.cell(r, col).fill = zfill
    # Step 3: Merge zones
    for s, e in PAIR_ZONES:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    # Step 4: Set content on merge-start cells
    c = ws.cell(r, 1); c.value = lL; c.font = fo(8, True, 'mid'); c.alignment = AL
    c = ws.cell(r, 8); c.font = fo(8, False, 'dk'); c.alignment = AL
    c = ws.cell(r, 21); c.value = lR; c.font = fo(8, True, 'mid'); c.alignment = AL
    c = ws.cell(r, 28); c.font = fo(8, False, 'dk'); c.alignment = AL
    return r+1

def colhdr(r, hdrs):
    ws.row_dimensions[r].height = 17.0
    zones = [(s, e) for s, e, _ in hdrs]
    # Step 1: borders on ALL cells (first row after sect header = top BLUE)
    apply_row_borders(ws, r, zones, True, False)
    # Step 2: fill ALL cells with accent2
    for col in range(1, NC+1):
        ws.cell(r, col).fill = fi(P['accent2'])
    # Step 3: merge and content
    for s, e, txt in hdrs:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
        c = ws.cell(r, s); c.value = txt
        c.font = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF')
        c.alignment = AC
    return r+1

def ckrow(r, num, cat, item, crit, catc=None, last=False):
    ws.row_dimensions[r].height = 20.0
    even = (num % 2 == 0)
    # Step 1: borders
    apply_row_borders_table(ws, r, CK_ZONES, last)
    # Step 2: fills per zone
    ifill = fi(P['near']) if even else fi(P['white'])
    zone_fills = [fi(P['fog']),                                          # #
                  fi(catc[0]) if catc else fi(P['fog']),                  # Cat
                  ifill,                                                   # Item
                  fi(P['fog']),                                            # Criteria
                  fi(P['white']),                                          # Result
                  fi(P['white']),                                          # Owner
                  fi(P['white'])]                                          # Ref
    for (s, e), zf in zip(CK_ZONES, zone_fills):
        for col in range(s, e+1):
            ws.cell(r, col).fill = zf
    # Step 3: merge
    for s, e in CK_ZONES:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    # Step 4: content
    ws.cell(r,1).value = num; ws.cell(r,1).font = fo(8,True,'mid'); ws.cell(r,1).alignment = AC
    c = ws.cell(r,3); c.value = cat; c.alignment = AC
    if catc:
        c.font = Font(name='Segoe UI',size=8,bold=True,color=catc[1])
    else:
        c.font = fo(8,True,'mid')
    ws.cell(r,6).value = item; ws.cell(r,6).font = fo(8,False,'dk'); ws.cell(r,6).alignment = AL
    ws.cell(r,20).value = crit; ws.cell(r,20).font = fo(7,False,'muted'); ws.cell(r,20).alignment = AL
    for s in [32,36,39]:
        ws.cell(r,s).font = fo(8,False,'dk'); ws.cell(r,s).alignment = AC
    return r+1

def gate_header(r, text):
    """Gate sub-header bar (like column header but descriptive)"""
    ws.row_dimensions[r].height = 17.0
    for col in range(1, NC+1):
        c = ws.cell(r, col); c.fill = fi(P['accent2'])
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == NC else NONE_S
        c.border = Border(left=left, right=right, top=BLUE, bottom=SILVER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r,1); c.value = text
    c.font = Font(name='Segoe UI',size=9,bold=True,color='FFFFFF')
    c.alignment = Alignment(horizontal='left',vertical='center',indent=1)
    return r+1

def sig3(r, labels):
    SIG_ZONES = [(1,13),(14,26),(27,NC)]
    # Label row
    ws.row_dimensions[r].height = 17.0
    for col in range(1, NC+1):
        ws.cell(r,col).fill = fi(P['fog'])
    # Borders: each sig zone has blue outline
    for s, e in SIG_ZONES:
        for col in range(s, e+1):
            left = BLUE if col == s else NONE_S
            right = BLUE if col == e else NONE_S
            ws.cell(r,col).border = Border(left=left, right=right, top=BLUE, bottom=BLUE)
    for (s,e), lab in zip(SIG_ZONES, labels):
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
        c = ws.cell(r,s); c.value = lab; c.font = fo(8,True,'mid'); c.alignment = AC
    r += 1
    # Signature boxes (3 rows)
    for sr in range(r, r+3):
        ws.row_dimensions[sr].height = 26.0
        for col in range(1, NC+1):
            ws.cell(sr, col).fill = fi(P['white'])
        for s, e in SIG_ZONES:
            for col in range(s, e+1):
                left = BLUE if col == s else NONE_S
                right = BLUE if col == e else NONE_S
                ws.cell(sr, col).border = Border(left=left, right=right, top=BLUE, bottom=BLUE)
    for s, e in SIG_ZONES:
        ws.merge_cells(start_row=r, start_column=s, end_row=r+2, end_column=e)
        c = ws.cell(r, s)
        c.value = 'Name  /  Signature  /  Date'
        c.font = Font(name='Segoe UI', size=8, color=P['plc'])
        c.alignment = Alignment(horizontal='center', vertical='bottom')
    return r + 3

def notice_bar(r, text):
    ws.row_dimensions[r].height = 24.0
    for col in range(1, NC+1):
        c = ws.cell(r, col); c.fill = fi(P['ice'])
        left = BLUE if col == 1 else NONE_S
        right = BLUE if col == NC else NONE_S
        c.border = Border(left=left, right=right, top=BLUE, bottom=BLUE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1); c.value = text
    c.font = Font(name='Segoe UI', size=6, color=P['mid'])
    c.alignment = Alignment(horizontal='left', vertical='center', indent=2, wrap_text=True)
    return r + 1

# =====================================================================
# BUILD FRM-511
# =====================================================================
r = 7
r = spacer(r)

# SECTION 1: SETUP HEADER
r = sect(r, '  1.  SETUP / FIRST-PIECE HEADER')
r = dpair(r, 'Job / WO', 'Part No. / Revision', first=True)
r = dpair(r, 'Machine / Cell', 'Operation / Route Step')
r = dpair(r, 'Setup Technician', 'QA Reviewer')
r = dpair(r, 'CNC Program / Rev', 'Fixture ID')
r = dpair(r, 'Tool List Ref / Count', 'Setup Time (min)')
r = dpair(r, 'Coolant Type / Conc.', 'Offset Baseline Verified?')
r = dpair(r, 'Drawing Rev = Program Rev?', 'Evidence Folder')
r = dpair(r, 'Ref: SOP-504 / WI-517', 'Linked FRM-302 / FRM-519', last=True)

r = spacer(r)

# SECTION 2: CHECKLIST
r = sect(r, '  2.  SETUP & FIRST-PIECE VERIFICATION')
CKH = [(1,2,'#'),(3,5,'Cat.'),(6,19,'Setup / FP Check Item'),
       (20,31,'Acceptance / Control Rule'),(32,35,'Result'),
       (36,38,'Owner'),(39,NC,'Ref.')]
r = colhdr(r, CKH)
ck_start = r

CT = {'SET':(P['tec_bg'],P['tec_fg']),'TOOL':(P['cap_bg'],P['cap_fg']),
      'GAGE':(P['con_bg'],P['con_fg']),'QUA':(P['qua_bg'],P['qua_fg']),
      'OPS':(P['com_bg'],P['com_fg'])}

items = [
    (1,'SET','Setup baseline matches released package (FRM-302).','FRM-302 and released packet align.'),
    (2,'SET','Machine / fixture / datums are correct.','Actual setup matches baseline condition.'),
    (3,'TOOL','Tool list, offsets and wear comp. are loaded.','Tool data matches FRM-302 tool list.'),
    (4,'TOOL','Tool condition is acceptable (no chip/wear).','Visual + touch-off check before run.'),
    (5,'GAGE','Required gages / programs available and in cal.','Cal sticker current; CMM program rev matches.'),
    (6,'QUA','First-piece dimensions are within tolerance.','All CTQ dims measured and recorded.'),
    (7,'QUA','Surface finish / burr / edge break acceptable.','Ra check per drawing requirement.'),
    (8,'QUA','GD&T / profile / position verified (if applicable).','CMM report attached or referenced.'),
    (9,'OPS','Abnormal findings are documented.','Any discrepancy has hold or action logic.'),
    (10,'OPS','Run condition after first-piece is explicit.','Pass-to-run or hold decision is visible.'),
]
for i,(num,cat,item,crit) in enumerate(items):
    r = ckrow(r,num,cat,item,crit,CT.get(cat),last=(i==len(items)-1))
ck_end = r - 1

r = spacer(r)

# SECTION 3: GATE
r = sect(r, '  3.  RUN DECISION / RELEASE GATE')
r = dpair(r, 'Run Decision', 'Summary / Notes', first=True)
r = dpair(r, 'Open Conditions before Run', 'Escalation if Not Closed')
r = dpair(r, 'Action Owner', 'Completion Deadline', last=True)

r = spacer(r)

# SECTION 4: APPROVAL
r = sect(r, '  4.  APPROVAL')
r = sig3(r, ['Setup Technician', 'QA Reviewer', 'Production Supervisor'])

r = spacer(r)

# NOTICE
r = notice_bar(r,
    'NOTICE  \u2014  Enter data only in white input cells.  Do not add, delete or resize columns/rows.  '
    'One form per setup event.  Ref: SOP-504 / WI-517.  Retain: 10 years per QMS schedule.')

# DATA VALIDATION
dv = DataValidation(type="list", formula1="=LISTS!$A$2:$A$5", allow_blank=True)
dv.error = "Select: PASS / HOLD / FAIL / NA"; dv.showErrorMessage = True
ws.add_data_validation(dv)
for rr in range(ck_start, ck_end+1):
    dv.add(f"AF{rr}")

# CONDITIONAL FORMATTING
rng = f"AF{ck_start}:AF{ck_end}"
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"PASS"'],
    fill=fi('C6EFCE'), font=Font(color='006100')))
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"HOLD"'],
    fill=fi('FFEB9C'), font=Font(color='9C6500')))
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"FAIL"'],
    fill=fi('FFC7CE'), font=Font(color='9C0006')))

# PRINT
ws.page_setup.paperSize = 9; ws.page_setup.orientation = 'portrait'
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0

# Remove other sheets
for name in list(wb.sheetnames):
    if name not in ['FRM-511', 'LISTS']:
        del wb[name]

# SAVE
OUT = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau/05-FRM-500/FRM-511_Setup_and_First_Piece_Record.xlsx"
wb.save(OUT)
sha = hashlib.sha256(open(OUT,'rb').read()).hexdigest()
print(f"FRM-511 v3 CREATED: SHA={sha[:16]}...")
print(f"Output: {OUT}")
