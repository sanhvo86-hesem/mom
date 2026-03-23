# -*- coding: utf-8 -*-
"""Generate FRM-100 series with expert-level content + VN comments"""
import sys, os, traceback
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from form_engine import generate_form
from form_engine import _borders_section_row, _borders_table_row
from wave3_log_specs import generate_log_form
from specs_frm100 import FRM_100_CHECKLIST, FRM_100_REPORT, FRM_100_LOG

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"

def get_output_path(code):
    for d in os.listdir(BASE):
        target = os.path.join(BASE, d)
        if not os.path.isdir(target): continue
        for f in os.listdir(target):
            if f.startswith(code) and f.endswith('.xlsx'):
                return os.path.join(target, f)
    return None

print("=" * 70)
print("FRM-100 SERIES: 18 FORMS (expert rebuild with VN comments)")
print("=" * 70)
success = failed = 0

# Checklist + Report forms (use generate_form)
for spec in FRM_100_CHECKLIST + FRM_100_REPORT:
    code = spec['code']
    try:
        out = get_output_path(code)
        if not out:
            print(f"  {code} SKIP"); failed += 1; continue
        sha = generate_form(spec, out)
        print(f"  {code:8s} OK  SHA={sha[:12]}")
        success += 1
    except Exception as e:
        failed += 1
        print(f"  {code:8s} FAIL: {e}")
        traceback.print_exc()

# Log forms (use generate_log_form with comments)
for spec in FRM_100_LOG:
    code = spec['code']
    try:
        out = get_output_path(code)
        if not out:
            print(f"  {code} SKIP"); failed += 1; continue
        cols = spec['columns']
        sha = generate_log_form(code, spec['title'], spec['owner'], spec['approved'],
                                cols, spec['notice'], spec['data_rows'], out)
        # Add comments to column headers
        import openpyxl
        from form_engine import cmt as cmt_fn
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Find header row (row with # in first zone)
        hdr_row = None
        for r in range(7, 15):
            if ws.cell(r, cols[0][0]).value == '#':
                hdr_row = r
                break
        if hdr_row and 'col_comments' in spec:
            for idx, cm in spec['col_comments'].items():
                ws.cell(hdr_row, cols[idx][0]).comment = cm
        wb.save(out)
        print(f"  {code:8s} OK  SHA={sha[:12]} +comments")
        success += 1
    except Exception as e:
        failed += 1
        print(f"  {code:8s} FAIL: {e}")
        traceback.print_exc()

print(f"\nFRM-100 COMPLETE: {success} OK, {failed} FAIL")
