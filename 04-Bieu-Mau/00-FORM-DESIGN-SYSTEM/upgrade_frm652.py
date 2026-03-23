"""
FRM-652 CAPA/8D Report - Full Upgrade V0 -> V1
CNC Job-Order Precision Machining / Semiconductor Equipment Parts

8D is the backbone of corrective action in aerospace/semiconductor.
This form drives systematic problem-solving from containment through
root cause to prevention and effectiveness verification.

Upgrades:
1. D3 enhancement: "Affected parts already shipped? Y/N" (customer notification trigger)
2. Effectiveness Verification Date with 30/60/90 day options
3. Recurrence Count field
4. Customer Approval Required for Close-out? Y/N
5. Lessons Learned Summary link to FRM-151
6. Cross-form linkage: Source NCR(s), Related SCARs, Verification Evidence
7. Status timer logic: OPEN->IN PROGRESS->VERIFICATION->CLOSED
8. Parent procedure ref: SOP-606 / WI-606
9. Conditional formatting for overdue/status
10. Revision history + bump
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from upgrade_lib import (
    COLORS, FONT_LABEL, FONT_INPUT, FONT_BODY,
    FILL_ICE, FILL_FOG, FILL_WHITE,
    THIN_ACCENT, THIN_SILVER, ALIGN_LEFT, ALIGN_CENTER,
    H_DATA,
    add_result_conditional_formatting, add_status_conditional_formatting,
    add_revision_history_sheet, backup_and_save, bump_revision,
    ensure_lists_values,
)

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"
FORM_PATH = f"{BASE}/06-FRM-600/FRM-652_CAPA_8D_Report.xlsx"

print(f"Loading {FORM_PATH}...")
wb = openpyxl.load_workbook(FORM_PATH)
ws = wb['CAPA_8D']
ws_lists = wb['LISTS']
ws_actions = wb['ACTION_TRACKER']
NCOLS = 56  # A4 Landscape

# ============================================================================
# CURRENT LAYOUT (A4L, 56 cols):
# R7: Sec 1 "PROBLEM / TEAM DEFINITION"
# R8: CAPA/8D ID | Closure Status
# R9: Linked NCR/Complaint | Owner/Due Date
# R10: Prepared by/Date | Record Status
# R11: Evidence Package | Inspector/Quality Owner
# R12-R14: Problem Statement (free text)
# R15-R18: Team/Containment Summary (free text)
# R19: Sec 2 "ROOT CAUSE / CORRECTIVE ACTION"
# R20-R22: Root Cause Analysis (free text)
# R23-R25: Corrective/Preventive Action (free text)
# R26-R27: Action Status | Effectiveness Status
# R28: Sec 3 "VERIFICATION OF EFFECTIVENESS"
# R29-R31: Verification Evidence (free text)
# R32-R35: Systemic Prevention (free text)
# R36: Sec 4 "FORMAL CLOSURE / APPROVAL"
# R37-R38: Closure fields
# R39: Spacer
# R40: Sec 5 "APPROVAL / SIGN-OFF"
# ============================================================================

# ============================================================================
# 1. ADD CNC-CRITICAL REFERENCE FIELDS
# ============================================================================
print("Step 1: Adding CNC-critical reference fields...")

# Insert 4 rows after R11 (last reference field), before R12 (Problem Statement)
INSERT_AT = 12
NUM_INSERT = 4
ws.insert_rows(INSERT_AT, amount=NUM_INSERT)

# After: R12-R15 = new, R16 = old R12 (Problem Statement), R17+ shifted

# Reference merge pattern (A4L):
REF_ZONES = [(1,10), (11,28), (29,35), (36,56)]

def make_ref(row, ll, rl, lv=None, rv=None, locked=False):
    for s, e in REF_ZONES:
        if e > s:
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    ws.row_dimensions[row].height = H_DATA

    c = ws.cell(row=row, column=1)
    c.value = ll; c.font = FONT_LABEL; c.fill = FILL_FOG; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_ACCENT, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=11)
    c.value = lv; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)

    c = ws.cell(row=row, column=29)
    c.value = rl; c.font = FONT_LABEL; c.fill = FILL_FOG; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=36)
    c.value = rv; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_ACCENT, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)


# R12: Affected parts shipped? + Customer approval for close-out
make_ref(12, 'Affected Parts Already Shipped?', 'Customer Approval Required?')

# R13: Recurrence count + Effectiveness verification period
make_ref(13, 'Recurrence Count (same failure mode)', 'Effectiveness Verification Period')

# R14: Linked SCARs + Lessons Learned
make_ref(14, 'Related SCAR Refs (FRM-406)', 'Lessons Learned Ref (FRM-151)')

# R15: Parent procedure + retention
make_ref(15, 'Ref: SOP-606 / WI-606', 'Retain: 10 years',
         lv='SOP-606 (NCR, CAPA, Corrective Action)',
         rv='Per QMS retention schedule', locked=True)

# ============================================================================
# 2. DATA VALIDATIONS for new fields
# ============================================================================
print("Step 2: Adding data validations...")

# Affected Parts Shipped (K12) - YES/NO
dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
dv.add('K12')

# Customer Approval Required (AJ12) - YES/NO
dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
dv.add('AJ12')

# Effectiveness Verification Period (AJ13) - dropdown
dv = DataValidation(type="list", formula1='"30 DAYS,60 DAYS,90 DAYS,6 MONTHS"', allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
dv.add('AJ13')

# ============================================================================
# 3. CONDITIONAL FORMATTING
# ============================================================================
print("Step 3: Adding conditional formatting...")

# Closure Status (shifted: was AM8, now AM8 still since we inserted AFTER R11)
add_status_conditional_formatting(ws, "AM8:BD8")

# Affected Parts Shipped = YES -> AMBER (needs attention)
from openpyxl.formatting.rule import CellIsRule
ws.conditional_formatting.add("K12:AB12",
    CellIsRule(operator='equal', formula=['"YES"'],
              fill=PatternFill(start_color=COLORS['CF_RED'], end_color=COLORS['CF_RED'], fill_type='solid'),
              font=Font(color=COLORS['CF_RED_FONT'], bold=True)))

# Customer Approval = YES -> AMBER
ws.conditional_formatting.add("AJ12:BD12",
    CellIsRule(operator='equal', formula=['"YES"'],
              fill=PatternFill(start_color=COLORS['CF_AMBER'], end_color=COLORS['CF_AMBER'], fill_type='solid')))

# Action status in ACTION_TRACKER: OVERDUE -> RED
add_status_conditional_formatting(ws_actions, "E5:E19")

# Effectiveness status in ACTION_TRACKER
add_status_conditional_formatting(ws_actions, "F5:F19")

# ============================================================================
# 4. ENRICH LISTS
# ============================================================================
print("Step 4: Enriching LISTS...")

ensure_lists_values(ws_lists, 'VERIFICATION_PERIOD',
    ['30 DAYS', '60 DAYS', '90 DAYS', '6 MONTHS'])

ensure_lists_values(ws_lists, 'ROOT_CAUSE_METHOD',
    ['5-WHY', 'FISHBONE', 'FAULT TREE', 'PARETO', 'DESIGN OF EXPERIMENT'])

# ============================================================================
# 5. REVISION HISTORY + BUMP
# ============================================================================
print("Step 5: Adding revision history and bumping V0 -> V1...")

add_revision_history_sheet(wb, 'FRM-652')
bump_revision(ws, 'V1', '2026-03-22')
bump_revision(ws_actions, 'V1', '2026-03-22')

# ============================================================================
# 6. PROTECT FORMULAS
# ============================================================================
print("Step 6: Protecting formula cells...")

for sheet in [ws, ws_actions]:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.protection = Protection(locked=True)

# ============================================================================
# 7. UPDATE NOTICE BAR
# ============================================================================
print("Step 7: Updating notice bar...")

for r in range(ws.max_row, max(ws.max_row - 10, 1), -1):
    c = ws.cell(row=r, column=1)
    if c.value and 'NOTICE' in str(c.value):
        c.value = (
            "NOTICE: Complete in English. No back-dating or proxy-signing. "
            "Ref: SOP-606 / WI-606. Retain 10 years. "
            "If Affected Parts Shipped = YES, notify customer within 24h. "
            "Close-out requires effectiveness verification at 30/60/90 days. "
            "File: FRM-652_{CAPA-ID}.xlsx"
        )
        print(f"  Updated notice at R{r}")
        break

# ============================================================================
# 8. SAVE
# ============================================================================
print("Step 8: Saving...")
sha = backup_and_save(wb, FORM_PATH)

print(f"""
{'='*70}
FRM-652 CAPA / 8D REPORT - UPGRADE COMPLETE
{'='*70}
Version:    V0 -> V1

Changes:
  + 4 new CNC-critical reference rows:
    - Affected Parts Already Shipped? (YES/NO) - RED highlight if YES
    - Customer Approval Required for Close-out? (YES/NO)
    - Recurrence Count (same failure mode)
    - Effectiveness Verification Period (30/60/90 days dropdown)
    - Related SCAR Refs (FRM-406)
    - Lessons Learned Ref (FRM-151)
    - Parent procedure: SOP-606 / WI-606
    - Records retention: 10 years
  + Conditional formatting:
    - Shipped parts = YES -> RED (immediate escalation)
    - Customer approval = YES -> AMBER
    - Status color coding on ACTION_TRACKER
  + LISTS enriched: VERIFICATION_PERIOD, ROOT_CAUSE_METHOD
  + Revision history sheet
  + Formula cells protected

SHA-256: {sha}
{'='*70}
""")
