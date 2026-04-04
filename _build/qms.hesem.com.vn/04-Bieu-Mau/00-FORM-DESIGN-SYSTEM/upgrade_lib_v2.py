"""
upgrade_lib_v2.py — SAFE QMS Form Upgrade Library
===================================================

GOLDEN RULES (BẤT DI BẤT DỊCH):
  1. NEVER insert_rows / delete_rows
  2. NEVER change merged cells
  3. NEVER change column widths or row heights
  4. NEVER change cell fill, font, border, alignment
  5. ONLY modify cell .value (text content)
  6. ONLY add DataValidation to existing cell ranges
  7. ONLY add ConditionalFormatting to existing cell ranges
  8. ONLY add/modify hidden sheet content (LISTS, REV_HISTORY)

Master Template Grid (Portrait 40-col A-AN, all 2.33 wide):
  Header:   rows 1-5  (logo A1:H5, title I1:AD3, meta AE1-AI5)
  Spacer:   row 6     (h=4, empty)
  Sect 1:   row 7     (section title, full width)
  RefData:  rows 8-12 (4-zone: A:G | H:T | U:AA | AB:AN)
  Spacer:   row 13    (h=4)
  Sect 2:   row 14    (section title)
  GateData: rows 15-18 (4-zone)
  Spacer:   row 19    (h=4)
  Sect 3:   row 20    (section title)
  ColHdr:   row 21    (column headers, 7 zones)
  CheckRows: 22-31    (checklist data, 7-zone per row)
  Spacer:   row 32    (h=4)
  Sect 4:   row 33    (section title)
  ColHdr:   row 34    (column headers, 7 zones)
  ActRows:  rows 35-39 (action data, 7-zone per row)
  Spacer:   row 40    (h=4)
  Sect 5:   row 41    (section title)
  SignLbl:  row 42    (3 labels)
  SignBox:  rows 43-45 (3 merged signature boxes, h=26)
  Spacer:   row 46    (h=4)
  Notice:   row 47    (h=24)

IMPORTANT: Some forms shift by ±1 row. Always detect actual structure
by scanning for section titles (cells starting with "1." or "2." etc).
"""

import openpyxl
import hashlib
import os
import shutil
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill


# =============================================================================
# COLOR CONSTANTS (for Conditional Formatting ONLY — never for cell styling)
# =============================================================================
CF_RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
CF_RED_FONT = Font(color='9C0006')
CF_AMBER_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
CF_AMBER_FONT = Font(color='9C6500')
CF_GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
CF_GREEN_FONT = Font(color='006100')


# =============================================================================
# SAFE STRUCTURE DETECTION
# =============================================================================

def detect_form_layout(ws):
    """
    Detect the row positions of each section in the form.
    Returns dict with section_row positions.
    Works regardless of ±1 row shifts between forms.
    """
    layout = {
        'header_start': None,   # first row with logo/title
        'sect1_title': None,    # "1. ..." section title
        'ref_start': None,      # first reference data row
        'ref_end': None,        # last reference data row
        'sect2_title': None,    # "2. ..." section title
        'gate_start': None,
        'gate_end': None,
        'sect3_title': None,    # "3. ..." section title
        'check_header': None,   # column header row for checklist
        'check_start': None,    # first checklist data row
        'check_end': None,      # last checklist data row
        'sect4_title': None,    # "4. ..."
        'action_header': None,
        'action_start': None,
        'action_end': None,
        'sect5_title': None,    # "5. APPROVAL ..."
        'sign_labels': None,
        'sign_start': None,
        'sign_end': None,
        'notice_row': None,
        'meta_col': None,       # column where "Form Code" label starts
    }

    for row in range(1, min(ws.max_row + 1, 60)):
        val = ws.cell(row, 1).value
        if val is None:
            continue
        val = str(val).strip()

        # Detect section titles
        if val.startswith('1.') or val.startswith('  1.'):
            layout['sect1_title'] = row
            layout['ref_start'] = row + 1
        elif val.startswith('2.') or val.startswith('  2.'):
            layout['sect2_title'] = row
            if layout['ref_start']:
                layout['ref_end'] = row - 2  # spacer row before
            layout['gate_start'] = row + 1
        elif val.startswith('3.') or val.startswith('  3.'):
            layout['sect3_title'] = row
            if layout['gate_start']:
                layout['gate_end'] = row - 2
            layout['check_header'] = row + 1
            layout['check_start'] = row + 2
        elif val.startswith('4.') or val.startswith('  4.'):
            layout['sect4_title'] = row
            if layout['check_start']:
                layout['check_end'] = row - 2
            layout['action_header'] = row + 1
            layout['action_start'] = row + 2
        elif val.startswith('5.') or val.startswith('  5.'):
            layout['sect5_title'] = row
            if layout['action_start']:
                layout['action_end'] = row - 2
            layout['sign_labels'] = row + 1
            layout['sign_start'] = row + 2
        elif val.startswith('NOTICE'):
            layout['notice_row'] = row

    # Detect meta box column (where "Form Code" is)
    for row in range(1, 8):
        for col in range(25, 45):
            v = ws.cell(row, col).value
            if v and str(v).strip() == 'Form Code':
                layout['meta_col'] = col
                break

    # Detect header start
    for row in range(1, 5):
        for col in range(1, 20):
            v = ws.cell(row, col).value
            if v and ('MASTER' in str(v) or 'HESEM' in str(v) or
                     'CONTRACT' in str(v) or 'SETUP' in str(v) or
                     'CHECKLIST' in str(v) or 'REPORT' in str(v) or
                     'LOG' in str(v) or 'FORM' in str(v)):
                layout['header_start'] = row
                break

    if not layout['header_start']:
        layout['header_start'] = 1

    return layout


def find_meta_cell(ws, label):
    """Find the value cell next to a meta label like 'Form Code', 'Revision', etc."""
    for row in range(1, 8):
        for col in range(25, 50):
            v = ws.cell(row, col).value
            if v and str(v).strip() == label:
                # Value is in the next merge zone to the right
                # In master template: label=AE(31), value=AI(35)
                # In landscape: label=AQ(43), value=AW(49)
                # Find the actual value column by checking merged cells
                for mc in ws.merged_cells.ranges:
                    if mc.min_row == row and mc.min_col > col:
                        return ws.cell(row, mc.min_col)
                # Fallback: check col+4 to col+8
                for dc in range(4, 9):
                    c = ws.cell(row, col + dc)
                    if c.value is not None or True:
                        return c
    return None


# =============================================================================
# SAFE VALUE-ONLY MODIFICATIONS
# =============================================================================

def set_cell_value(ws, row, col, value):
    """
    Set a cell value WITHOUT changing any styling.
    This is the ONLY way to modify form content.
    """
    ws.cell(row, col).value = value


def set_ref_field(ws, layout, ref_row_idx, zone, value):
    """
    Set a reference field value by row index (0-4) and zone ('label1','value1','label2','value2').

    ref_row_idx: 0-4 (within the reference section rows)
    zone: 'label1', 'value1', 'label2', 'value2'
    """
    if not layout['ref_start']:
        return False

    row = layout['ref_start'] + ref_row_idx
    if layout['ref_end'] and row > layout['ref_end']:
        return False

    # Detect zone columns from merged cells
    zones = _detect_row_zones(ws, row)
    if not zones:
        return False

    zone_map = {
        'label1': 0, 'value1': 1,
        'label2': 2, 'value2': 3
    }
    idx = zone_map.get(zone)
    if idx is not None and idx < len(zones):
        ws.cell(row, zones[idx]).value = value
        return True
    return False


def set_gate_field(ws, layout, gate_row_idx, zone, value):
    """Same as set_ref_field but for gate section."""
    if not layout['gate_start']:
        return False
    row = layout['gate_start'] + gate_row_idx
    if layout['gate_end'] and row > layout['gate_end']:
        return False
    zones = _detect_row_zones(ws, row)
    if not zones:
        return False
    zone_map = {'label1': 0, 'value1': 1, 'label2': 2, 'value2': 3}
    idx = zone_map.get(zone)
    if idx is not None and idx < len(zones):
        ws.cell(row, zones[idx]).value = value
        return True
    return False


def set_checklist_item(ws, layout, check_row_idx, values):
    """
    Set checklist item values by row index (0-9).
    values: dict with keys matching zone positions:
        'num', 'cat', 'item', 'criteria', 'result', 'owner', 'ref'
    Only sets values that are provided; leaves others unchanged.
    """
    if not layout['check_start']:
        return False
    row = layout['check_start'] + check_row_idx
    if layout['check_end'] and row > layout['check_end']:
        return False

    zones = _detect_row_zones(ws, row)
    if not zones:
        return False

    zone_keys = ['num', 'cat', 'item', 'criteria', 'result', 'owner', 'ref']
    for i, key in enumerate(zone_keys):
        if key in values and i < len(zones):
            ws.cell(row, zones[i]).value = values[key]
    return True


def set_action_item(ws, layout, action_row_idx, values):
    """Same as set_checklist_item but for action section."""
    if not layout['action_start']:
        return False
    row = layout['action_start'] + action_row_idx
    if layout['action_end'] and row > layout['action_end']:
        return False

    zones = _detect_row_zones(ws, row)
    if not zones:
        return False

    zone_keys = ['num', 'issue', 'action', 'owner', 'due', 'status', 'verified']
    for i, key in enumerate(zone_keys):
        if key in values and i < len(zones):
            ws.cell(row, zones[i]).value = values[key]
    return True


def _detect_row_zones(ws, row):
    """
    Detect the starting columns of each merge zone in a row.
    Returns list of column numbers (1-based) for each zone.
    """
    zones = []
    merges = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row == row and mc.max_row == row:
            merges.append(mc.min_col)

    if merges:
        return sorted(merges)

    # If no merges on this row, check for full-width merge
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row:
            return [mc.min_col]

    # No merges at all — return column 1
    return [1]


# =============================================================================
# DATA VALIDATION (safe to add to existing cells)
# =============================================================================

def add_dropdown(ws, cell_range, options, allow_blank=True, error_msg=None):
    """
    Add a dropdown data validation to a cell range.
    options: list of strings OR formula reference like '=LISTS!$A$2:$A$5'
    """
    if isinstance(options, list):
        formula = '"' + ','.join(options) + '"'
    else:
        formula = options

    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    if error_msg:
        dv.error = error_msg
        dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def add_date_validation(ws, cell_range):
    """Add date validation (must be valid date)."""
    dv = DataValidation(type="date", allow_blank=True)
    dv.showErrorMessage = True
    dv.error = "Please enter a valid date."
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


# =============================================================================
# CONDITIONAL FORMATTING (safe to add to existing cells)
# =============================================================================

def add_cf_status_colors(ws, cell_range):
    """Add standard PASS/HOLD/FAIL/NA color coding to a range."""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"FAIL"'],
                   fill=CF_RED_FILL, font=CF_RED_FONT))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"HOLD"'],
                   fill=CF_AMBER_FILL, font=CF_AMBER_FONT))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"PASS"'],
                   fill=CF_GREEN_FILL, font=CF_GREEN_FONT))


def add_cf_overdue(ws, date_range):
    """Add overdue date formatting: RED if past, AMBER if <30 days."""
    # First cell reference for formula
    first_cell = date_range.split(':')[0] if ':' in date_range else date_range
    ws.conditional_formatting.add(date_range,
        FormulaRule(formula=[f'AND({first_cell}<>"",{first_cell}<TODAY())'],
                    fill=CF_RED_FILL, font=CF_RED_FONT))
    ws.conditional_formatting.add(date_range,
        FormulaRule(formula=[f'AND({first_cell}<>"",{first_cell}>=TODAY(),{first_cell}<TODAY()+30)'],
                    fill=CF_AMBER_FILL))


def add_cf_severity(ws, cell_range):
    """Add severity color coding: CRITICAL=red, HIGH=amber."""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"CRITICAL"'],
                   fill=CF_RED_FILL, font=CF_RED_FONT))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"HIGH"'],
                   fill=CF_AMBER_FILL, font=CF_AMBER_FONT))


# =============================================================================
# LISTS SHEET MANAGEMENT
# =============================================================================

def ensure_lists_column(ws_lists, header, values, col=None):
    """
    Add or update a column in the LISTS sheet.
    If col is None, finds the next empty column.
    """
    if col is None:
        # Find next empty column
        col = 1
        while ws_lists.cell(1, col).value is not None:
            col += 1

    ws_lists.cell(1, col).value = header
    for i, v in enumerate(values, start=2):
        ws_lists.cell(i, col).value = v
    return col


def get_lists_sheet(wb):
    """Get the LISTS sheet (may be named _LISTS, LISTS, or hidden)."""
    for name in ['_LISTS', 'LISTS']:
        if name in wb.sheetnames:
            return wb[name]
    # Create if not exists
    ws = wb.create_sheet('_LISTS')
    ws.sheet_state = 'hidden'
    return ws


# =============================================================================
# REVISION MANAGEMENT
# =============================================================================

def bump_revision(ws, new_rev='V1', new_date=None):
    """
    Update revision and date in the meta box.
    ONLY changes cell values, never styling.
    """
    if new_date is None:
        new_date = datetime.now().strftime('%Y-%m-%d')

    rev_cell = find_meta_cell(ws, 'Revision')
    date_cell = find_meta_cell(ws, 'Eff. Date')

    if rev_cell:
        rev_cell.value = new_rev
    if date_cell:
        date_cell.value = new_date

    return bool(rev_cell and date_cell)


def add_revision_history(wb, form_code, changes_text):
    """
    Add a REV_HISTORY hidden sheet with revision log.
    Does NOT modify any existing sheets.
    """
    sheet_name = 'REV_HISTORY'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_state = 'hidden'
        ws.cell(1, 1).value = 'Rev'
        ws.cell(1, 2).value = 'Date'
        ws.cell(1, 3).value = 'Description'
        ws.cell(1, 4).value = 'Author'
        # V0 baseline
        ws.cell(2, 1).value = 'V0'
        ws.cell(2, 2).value = '2026-03-21'
        ws.cell(2, 3).value = 'Initial release'
        ws.cell(2, 4).value = 'QMS Team'

    # Find next empty row
    next_row = ws.max_row + 1
    ws.cell(next_row, 1).value = 'V1'
    ws.cell(next_row, 2).value = datetime.now().strftime('%Y-%m-%d')
    ws.cell(next_row, 3).value = changes_text
    ws.cell(next_row, 4).value = 'QMS Upgrade'

    return ws


# =============================================================================
# SAVE WITH SHA-256
# =============================================================================

def save_with_checksum(wb, filepath):
    """
    Save workbook and compute SHA-256 checksum.
    Creates backup before overwriting.
    """
    # Backup
    backup_dir = os.path.join(os.path.dirname(filepath), '.backups')
    os.makedirs(backup_dir, exist_ok=True)
    if os.path.exists(filepath):
        fname = os.path.basename(filepath)
        backup_path = os.path.join(backup_dir, f"{fname}.v0.bak")
        if not os.path.exists(backup_path):
            shutil.copy2(filepath, backup_path)

    wb.save(filepath)

    # Compute SHA-256
    sha256 = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            sha256.update(chunk)

    return sha256.hexdigest()


# =============================================================================
# MULTI-SHEET FORM DETECTION
# =============================================================================

def get_visible_sheets(wb):
    """Get all visible (form face) sheets."""
    return [ws for ws in wb.worksheets if ws.sheet_state == 'visible']


def get_form_face(wb):
    """Get the primary form face sheet (first visible sheet)."""
    visible = get_visible_sheets(wb)
    return visible[0] if visible else None


# =============================================================================
# BATCH UPGRADE HELPER
# =============================================================================

def safe_upgrade_form(filepath, upgrade_fn, description):
    """
    Safely upgrade a form:
    1. Load from original
    2. Apply upgrade function
    3. Save with checksum

    upgrade_fn receives (wb, ws_form, ws_lists, layout) and returns changes_text
    """
    print(f"  Loading: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath)
    ws_form = get_form_face(wb)
    if not ws_form:
        print(f"    ERROR: No visible sheet found")
        return None

    ws_lists = get_lists_sheet(wb)
    layout = detect_form_layout(ws_form)

    print(f"    Layout: sect1={layout['sect1_title']} sect2={layout['sect2_title']} "
          f"sect3={layout['sect3_title']} sect4={layout['sect4_title']} "
          f"sect5={layout['sect5_title']}")

    # Apply upgrade
    changes_text = upgrade_fn(wb, ws_form, ws_lists, layout)

    # Bump revision
    bump_revision(ws_form, 'V1', datetime.now().strftime('%Y-%m-%d'))

    # Add revision history
    add_revision_history(wb, os.path.basename(filepath).split('_')[0], changes_text)

    # Save
    sha = save_with_checksum(wb, filepath)
    print(f"    DONE: {description} | SHA: {sha[:16]}...")
    return sha
