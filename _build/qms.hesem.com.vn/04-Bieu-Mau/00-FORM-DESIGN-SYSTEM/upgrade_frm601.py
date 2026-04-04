"""
FRM-601 Calibration Log - Full Upgrade V0 -> V1

HIGH SEVERITY: ISO 7.1.5.2 requires evaluating validity of previous
measurement results when a gage is found out of tolerance. Missing this
is a potential MAJOR finding in AS9100D audits.

Upgrades:
1. Impact Assessment column (MANDATORY per ISO 7.1.5.2)
2. As-Found / As-Left result distinction
3. Environmental conditions (temp/humidity) field
4. Measurement Uncertainty of calibration
5. Overdue conditional formatting on Next Due dates
6. Cross-form linkage to FRM-525 Gage Register
7. Parent procedure ref: SOP-601
8. Revision history + bump
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

from upgrade_lib import (
    COLORS, FONT_LABEL, FONT_INPUT, FONT_BODY, FONT_COL_HEADER, FONT_NUM,
    FILL_ICE, FILL_FOG, FILL_WHITE, FILL_COL_HDR,
    FILL_CF_RED, FILL_CF_AMBER, FILL_CF_GREEN,
    THIN_ACCENT, THIN_SILVER, ALIGN_LEFT, ALIGN_CENTER,
    H_DATA,
    add_status_conditional_formatting,
    add_revision_history_sheet, backup_and_save, bump_revision,
    ensure_lists_values,
)

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"
FORM_PATH = f"{BASE}/06-FRM-600/FRM-601_Calibration_Log.xlsx"

print(f"Loading {FORM_PATH}...")
wb = openpyxl.load_workbook(FORM_PATH)
ws_data = wb['CAL_DATA_LOG']
ws_form = wb['CAL_LOG']
ws_lists = wb['LISTS']

# ============================================================================
# 1. ENHANCE CAL_DATA_LOG with critical columns
#    Current: A=#, B=Event Date, C=Equipment ID, D=Cal Source, E=Cert/Result,
#             F=Next Due, G=Status, H=Use Stop, I=Owner Dept, J=Evidence Ref
#    Add: K=As-Found, L=As-Left, M=Impact Assessment, N=Env Conditions, O=Uncertainty
# ============================================================================
print("Step 1: Adding critical columns to CAL_DATA_LOG...")

new_headers = {
    11: ('As-Found Result', 14),      # K
    12: ('As-Left Result', 14),       # L
    13: ('Impact Assessment', 22),     # M - widest, needs room for text
    14: ('Env. (Temp/Humid)', 16),    # N
    15: ('U (k=2)', 10),              # O - measurement uncertainty
}

for col, (header, width) in new_headers.items():
    c = ws_data.cell(row=6, column=col)
    c.value = header
    c.font = Font(name='Segoe UI', size=8, bold=True, color=COLORS['WHITE'])
    c.fill = PatternFill(start_color='0F5FA6', end_color='0F5FA6', fill_type='solid')
    c.alignment = ALIGN_CENTER
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    ws_data.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Add data validation for Impact Assessment
dv_impact = DataValidation(
    type="list",
    formula1='"NO IMPACT - within tolerance,REVIEW REQUIRED - marginal,RECALL REQUIRED - out of tolerance,NA - initial cal"',
    allow_blank=True
)
dv_impact.error = "Select impact assessment level."
dv_impact.showErrorMessage = True
ws_data.add_data_validation(dv_impact)
dv_impact.add('M7:M36')

# Add data validation for As-Found/As-Left
for col_letter in ['K', 'L']:
    dv = DataValidation(type="list", formula1='"PASS,FAIL,ADJUSTED,NA"', allow_blank=True)
    dv.showErrorMessage = True
    ws_data.add_data_validation(dv)
    dv.add(f'{col_letter}7:{col_letter}36')

# ============================================================================
# 2. CONDITIONAL FORMATTING on CAL_DATA_LOG
# ============================================================================
print("Step 2: Adding conditional formatting...")

# As-Found = FAIL -> RED (out of tolerance, needs impact assessment)
ws_data.conditional_formatting.add("K7:K36",
    CellIsRule(operator='equal', formula=['"FAIL"'],
              fill=FILL_CF_RED, font=Font(color=COLORS['CF_RED_FONT'], bold=True)))

# Impact Assessment = RECALL REQUIRED -> RED
ws_data.conditional_formatting.add("M7:M36",
    CellIsRule(operator='equal', formula=['"RECALL REQUIRED - out of tolerance"'],
              fill=FILL_CF_RED, font=Font(color=COLORS['CF_RED_FONT'], bold=True)))

ws_data.conditional_formatting.add("M7:M36",
    CellIsRule(operator='equal', formula=['"REVIEW REQUIRED - marginal"'],
              fill=FILL_CF_AMBER))

# Next Due date overdue -> RED, within 30 days -> AMBER
ws_data.conditional_formatting.add("F7:F36",
    FormulaRule(formula=['AND(F7<>"",F7<TODAY())'],
               fill=FILL_CF_RED, font=Font(color=COLORS['CF_RED_FONT'])))

ws_data.conditional_formatting.add("F7:F36",
    FormulaRule(formula=['AND(F7<>"",F7>=TODAY(),F7<TODAY()+30)'],
               fill=FILL_CF_AMBER))

# Status column color coding
ws_data.conditional_formatting.add("G7:G36",
    CellIsRule(operator='equal', formula=['"OVERDUE"'],
              fill=FILL_CF_RED, font=Font(color=COLORS['CF_RED_FONT'], bold=True)))
ws_data.conditional_formatting.add("G7:G36",
    CellIsRule(operator='equal', formula=['"OUT OF CAL"'],
              fill=FILL_CF_RED))
ws_data.conditional_formatting.add("G7:G36",
    CellIsRule(operator='equal', formula=['"VERIFIED"'],
              fill=FILL_CF_GREEN))

# Use Stop = YES -> RED
ws_data.conditional_formatting.add("H7:H36",
    CellIsRule(operator='equal', formula=['"Y"'],
              fill=FILL_CF_RED, font=Font(color=COLORS['CF_RED_FONT'], bold=True)))

# ============================================================================
# 3. ENHANCE CAL_LOG form face (56-col grid)
# ============================================================================
print("Step 3: Enhancing CAL_LOG form face...")

# Insert 2 rows after R12 for cross-form linkage and parent procedure
ws_form.insert_rows(13, amount=2)

REF_ZONES = [(1,10), (11,28), (29,35), (36,56)]

def make_ref(ws, row, ll, rl, lv=None, rv=None, locked=False):
    for s, e in REF_ZONES:
        if e > s:
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    ws.row_dimensions[row].height = H_DATA
    c = ws.cell(row=row, column=1)
    c.value = ll; c.font = FONT_LABEL; c.fill = FILL_FOG; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_ACCENT, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c = ws.cell(row=row, column=11)
    c.value = lv; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)
    c = ws.cell(row=row, column=29)
    c.value = rl; c.font = FONT_LABEL; c.fill = FILL_FOG; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c = ws.cell(row=row, column=36)
    c.value = rv; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_ACCENT, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)

make_ref(ws_form, 13, 'Linked FRM-525 Gage Reg.', 'Environmental Conditions')
make_ref(ws_form, 14, 'Ref: SOP-601 / Cal Procedure', 'Retain: 10 years',
         lv='SOP-601 (Calibration and Measurement Traceability)',
         rv='Per QMS retention schedule', locked=True)

# ============================================================================
# 4. ENRICH LISTS
# ============================================================================
print("Step 4: Enriching LISTS...")

ensure_lists_values(ws_lists, 'IMPACT_ASSESSMENT',
    ['NO IMPACT', 'REVIEW REQUIRED', 'RECALL REQUIRED', 'NA'])

ensure_lists_values(ws_lists, 'AS_FOUND_LEFT',
    ['PASS', 'FAIL', 'ADJUSTED', 'NA'])

# ============================================================================
# 5. UPDATE DATA LOG header/guidance text
# ============================================================================
print("Step 5: Updating guidance text...")

ws_data.cell(row=3, column=1).value = (
    'Controlled row-level entry. Use approved dropdowns only. Do not delete history. '
    'CRITICAL: When As-Found = FAIL, complete Impact Assessment column (ISO 7.1.5.2 mandatory). '
    'Record environmental conditions for precision calibrations.'
)

# ============================================================================
# 6. REVISION HISTORY + BUMP
# ============================================================================
print("Step 6: Adding revision history and bumping V0 -> V1...")

add_revision_history_sheet(wb, 'FRM-601')
bump_revision(ws_form, 'V1', '2026-03-23')

# ============================================================================
# 7. SAVE
# ============================================================================
print("Step 7: Saving...")
sha = backup_and_save(wb, FORM_PATH)

print(f"""
{'='*70}
FRM-601 CALIBRATION LOG - UPGRADE COMPLETE
{'='*70}
Version:    V0 -> V1

HIGH SEVERITY GAPS FIXED:
  + Impact Assessment column (ISO 7.1.5.2 MANDATORY)
    - NO IMPACT / REVIEW REQUIRED / RECALL REQUIRED / NA
    - RED highlight when RECALL REQUIRED
  + As-Found / As-Left result distinction
    - PASS/FAIL/ADJUSTED/NA for each
    - As-Found = FAIL triggers RED highlight

Additional upgrades:
  + Environmental Conditions field (temp/humidity)
  + Measurement Uncertainty U(k=2) column
  + Next Due date overdue formatting (RED if past, AMBER if < 30 days)
  + Use Stop = YES -> RED highlight
  + Status color coding (OVERDUE/OUT OF CAL -> RED, VERIFIED -> GREEN)
  + Cross-form linkage: FRM-525 Gage Register
  + Parent procedure: SOP-601
  + Records retention: 10 years
  + Revision history sheet

SHA-256: {sha}
{'='*70}
""")
