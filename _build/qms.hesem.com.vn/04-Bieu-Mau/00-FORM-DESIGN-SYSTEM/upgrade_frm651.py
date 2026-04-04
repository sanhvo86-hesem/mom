"""
FRM-651 NCR Report - Full Upgrade V0 -> V1
CNC Job-Order Precision Machining / Semiconductor Equipment Parts

NCR is used daily in any CNC job shop. This form must capture:
- What went wrong (nonconformity description)
- How bad (severity, quantity, cost)
- What was done immediately (containment)
- What happens next (disposition, CAPA linkage)
- Who else needs to know (customer notification)

Upgrades:
1. Cost of Nonconformity (scrap cost, rework hours, delivery delay)
2. Quantity Affected / Quantity Contained
3. Customer Notification Required? Y/N
4. Recurrence Check (has this defect occurred before?)
5. Risk Level dropdown (LOW/MEDIUM/HIGH/CRITICAL)
6. Cross-form linkage: FRM-652 CAPA, FRM-641 Final Inspection, FRM-406 SCAR, FRM-213 RMA
7. Escape Point Analysis dropdown
8. Parent procedure ref: SOP-606 / WI-606
9. Conditional formatting
10. Revision history + bump
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from upgrade_lib import (
    COLORS, FONT_LABEL, FONT_INPUT, FONT_BODY,
    FILL_ICE, FILL_FOG, FILL_WHITE,
    THIN_ACCENT, THIN_SILVER, ALIGN_LEFT, ALIGN_CENTER,
    H_DATA,
    add_result_conditional_formatting, add_status_conditional_formatting,
    add_risk_conditional_formatting,
    add_revision_history_sheet, backup_and_save, bump_revision,
    ensure_lists_values,
)

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"
FORM_PATH = f"{BASE}/06-FRM-600/FRM-651_NCR_Report.xlsx"

print(f"Loading {FORM_PATH}...")
wb = openpyxl.load_workbook(FORM_PATH)
ws = wb['NCR_FORM']
ws_lists = wb['LISTS']
ws_actions = wb['ACTIONS']
NCOLS = 40  # A4 Portrait

# ============================================================================
# CURRENT LAYOUT (A4P, 40 cols):
# R7: Sec 1 "REFERENCE INFORMATION"
# R8: NCR ID | Source
# R9: Severity | Containment Status
# R10: Job/Lot/Part/Rev | Suspect Qty
# R11: Prepared by/Date | Record Status
# R12: Evidence Package Ref | Linked CAPA/Complaint
# R13: Inspector/Quality Owner | Discovery Point
# R14: Spacer
# R15: Sec 2 "EVENT DETAIL / CONTROL CONTEXT"
# R16-R19: NONCONFORMITY DESCRIPTION (free text area)
# R20-R23: OBJECTIVE EVIDENCE / RISK (free text area)
# R23: Spacer
# R24: Sec 3 "CONTAINMENT / DISPOSITION SUMMARY"
# R25: Column headers
# R26-R30: 5 disposition items
# R31: Spacer
# R32: Sec 4 "FINAL DECISION / CLOSURE"
# R33-R34: Disposition fields
# R35: Spacer
# R36: Sec 5 "APPROVAL / SIGN-OFF"
# ============================================================================

# ============================================================================
# 1. UPGRADE REFERENCE FIELDS
#    Repurpose and add CNC-critical fields
# ============================================================================
print("Step 1: Upgrading reference fields...")

# R9: Keep Severity but repurpose right side
# Currently: Severity | Containment Status -> keep as is, good fields

# R12: Currently Evidence Package Ref | Linked CAPA/Complaint
# Already has CAPA linkage, good

# R13: Currently Inspector/Quality Owner | Discovery Point
# Discovery Point is good but needs dropdown - already has validation

# Insert 4 new reference rows after R13 for CNC-critical fields
INSERT_REF = 14  # Before current spacer R14
NUM_INSERT = 4
ws.insert_rows(INSERT_REF, amount=NUM_INSERT)

# After: R14-R17 = new, R18 = old spacer, R19+ shifted

# Reference merge pattern (A4P, 40 cols):
# A:G(1-7)=left-label | H:T(8-20)=left-input | U:AA(21-27)=right-label | AB:AN(28-40)=right-input
REF_ZONES = [(1,7), (8,20), (21,27), (28,40)]

def make_ref(row, ll, rl, lv=None, rv=None, locked=False):
    for s, e in REF_ZONES:
        if e > s:
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    ws.row_dimensions[row].height = H_DATA

    c = ws.cell(row=row, column=1)
    c.value = ll; c.font = FONT_LABEL; c.fill = FILL_FOG; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_ACCENT, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=8)
    c.value = lv; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)

    c = ws.cell(row=row, column=21)
    c.value = rl; c.font = FONT_LABEL; c.fill = FILL_FOG; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=28)
    c.value = rv; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_ACCENT, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)


# R14: Quantity Affected + Quantity Contained
make_ref(14, 'Quantity Affected', 'Quantity Contained / Sorted')

# R15: Cost of NC + Risk Level
make_ref(15, 'Cost of NC (scrap+rework+delay)', 'Risk Level')

# R16: Customer Notification + Recurrence Check
make_ref(16, 'Customer Notification Required?', 'Recurrence? Previous NCR #')

# R17: Cross-form linkage + Parent procedure
make_ref(17, 'Linked FRM-641 / FRM-406 / FRM-213', 'Ref: SOP-606 / WI-606',
         rv='SOP-606 (NCR, CAPA, Corrective Action)', locked=True)

# ============================================================================
# 2. ADD DATA VALIDATIONS for new fields
# ============================================================================
print("Step 2: Adding data validations...")

# Risk Level dropdown (AB15 = col 28, row 15)
dv = DataValidation(type="list", formula1="'LISTS'!$C$2:$C$5", allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
dv.add('AB15')

# Customer Notification Required (H16)
dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
dv.add('H16')

# Recurrence (AB16) - YES/NO
dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
dv.add('AB16')

# ============================================================================
# 3. CONDITIONAL FORMATTING
# ============================================================================
print("Step 3: Adding conditional formatting...")

# Containment status items (shifted +4): now AF30:AF34
add_status_conditional_formatting(ws, "AF30:AI34")

# Risk Level (AB15)
add_risk_conditional_formatting(ws, "AB15:AN15")

# Severity (H9) - using risk colors
add_risk_conditional_formatting(ws, "H9:T9")

# Customer Notification - RED if YES (needs attention)
from openpyxl.formatting.rule import CellIsRule
ws.conditional_formatting.add("H16:T16",
    CellIsRule(operator='equal', formula=['"YES"'],
              fill=PatternFill(start_color=COLORS['CF_AMBER'], end_color=COLORS['CF_AMBER'], fill_type='solid')))

# ============================================================================
# 4. ENRICH LISTS
# ============================================================================
print("Step 4: Enriching LISTS...")

# Add ESCAPE_POINT column
ensure_lists_values(ws_lists, 'ESCAPE_POINT',
    ['INCOMING', 'SETUP', 'FIRST PIECE', 'IN-PROCESS',
     'FINAL INSPECTION', 'PACKAGING', 'SHIPPING', 'CUSTOMER SITE'])

# Add COST_CATEGORY
ensure_lists_values(ws_lists, 'COST_CATEGORY',
    ['SCRAP', 'REWORK', 'SORT', 'REINSPECT', 'DELIVERY DELAY',
     'CUSTOMER RETURN', 'WARRANTY', 'OTHER'])

# ============================================================================
# 5. REVISION HISTORY + BUMP
# ============================================================================
print("Step 5: Adding revision history and bumping V0 -> V1...")

add_revision_history_sheet(wb, 'FRM-651')
bump_revision(ws, 'V1', '2026-03-22')
bump_revision(ws_actions, 'V1', '2026-03-22')

# ============================================================================
# 6. PROTECT FORMULAS
# ============================================================================
print("Step 6: Protecting formula cells...")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            cell.protection = Protection(locked=True)

# ============================================================================
# 7. UPDATE NOTICE BAR
# ============================================================================
print("Step 7: Updating notice bar...")

for r in range(ws.max_row, max(ws.max_row - 10, 1), -1):
    c = ws.cell(row=r, column=1)
    if c.value and 'NOTICE' in str(c.value):
        c.value = (
            "NOTICE: Complete in English. No back-dating, history deletion, "
            "proxy-signing or vague comments. "
            "Ref: SOP-606 / WI-606. Retain 10 years. "
            "If Customer Notification Required = YES, escalate to Sales/CS within 24h. "
            "File: FRM-651_{NCR-ID}.xlsx"
        )
        print(f"  Updated notice at R{r}")
        break

# ============================================================================
# 8. SAVE
# ============================================================================
print("Step 8: Saving...")
sha = backup_and_save(wb, FORM_PATH)

print(f"""
{'='*70}
FRM-651 NCR REPORT - UPGRADE COMPLETE
{'='*70}
Version:    V0 -> V1

Changes:
  + 4 new CNC-critical reference rows:
    - Quantity Affected / Quantity Contained
    - Cost of NC (scrap + rework + delay)
    - Risk Level (LOW/MEDIUM/HIGH/CRITICAL dropdown)
    - Customer Notification Required? (YES/NO)
    - Recurrence Check / Previous NCR #
    - Cross-form linkage: FRM-641, FRM-406, FRM-213
    - Parent procedure: SOP-606 / WI-606
  + Conditional formatting: Risk/Severity color coding, Customer Notification highlight
  + LISTS enriched: ESCAPE_POINT, COST_CATEGORY
  + Revision history sheet
  + Formula cells protected

SHA-256: {sha}
{'='*70}
""")
