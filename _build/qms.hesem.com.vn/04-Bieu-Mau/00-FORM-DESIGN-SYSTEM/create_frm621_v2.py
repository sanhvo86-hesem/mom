# -*- coding: utf-8 -*-
"""
FRM-621 SAMPLE v2 — Comments on LABELS (not input cells)
Comment format:
  Line 1: Vietnamese translation of label
  Line 2+: Description + examples
"""
import sys, os, hashlib
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from form_engine import *
from form_engine import _borders_section_row, _borders_table_row
import openpyxl
from openpyxl.comments import Comment

TMPL = "C:/Users/TEST4/Desktop/frm-000-master-template.xlsx"
nc = 56
fmt = FORMATS['A4L']
wb = openpyxl.load_workbook(TMPL)
src_ws = wb['02-MASTER-A4-LANDSCAPE']
ws = wb.create_sheet('FRM_621', 0)
for i in range(1, nc+1):
    ws.column_dimensions[get_column_letter(i)].width = 2.375

build_header(ws, src_ws, nc, 'AQL INSPECTION RECORD', 'FRM-621', 'QA', 'QA Manager')
embed_logo(ws, fmt['logo_cols'])

zones = get_zones(nc)
CK = zones['check']
PP = zones['pair']

def cmt(line1_vn, *details):
    """Comment: line 1 = Vietnamese meaning, line 2+ = description/examples"""
    text = line1_vn
    for d in details:
        text += '\n' + d
    c = Comment(text, 'QMS')
    c.width = 280
    c.height = 150
    return c

r = 7
r = spacer(ws, r, nc)

# === S1: REFERENCE ===
r = sect_header(ws, r, nc, '  1.  REFERENCE INFORMATION')
# Comment on section header label
ws.cell(r-1, 1).comment = cmt(
    'Thông tin tham chiếu',
    'Ghi thông tin cơ bản để truy xuất lô hàng.',
    'Tất cả thông tin lấy từ Job Packet / Lệnh sản xuất.')

ref_r = r
r = data_pair(ws, r, nc, PP, 'Job / WO', 'Part No. / Revision', first=True)
r = data_pair(ws, r, nc, PP, 'Lot Size / Sample Size', 'AQL Level / Insp. Level', last=True)

# Comments on LABEL cells (column 1 = left label, column 21/29 = right label)
ws.cell(ref_r, PP[0][0]).comment = cmt(
    'Số công việc / Lệnh sản xuất',
    'Ghi số Job hoặc Work Order từ Job Packet.',
    'Ví dụ: JOB-2026-0145 hoặc WO-00382')
ws.cell(ref_r, PP[2][0]).comment = cmt(
    'Mã chi tiết / Phiên bản bản vẽ',
    'Ghi mã part và revision hiện hành.',
    'Ví dụ: HES-1234 Rev.C')
ws.cell(ref_r+1, PP[0][0]).comment = cmt(
    'Cỡ lô / Cỡ mẫu',
    'Ghi tổng số lượng lô hàng và số mẫu kiểm tra.',
    'Tra bảng ISO 2859-1 theo AQL Level để xác định cỡ mẫu.',
    'Ví dụ: Lot Size = 100, Sample = 13 (Level II, AQL 1.0)')
ws.cell(ref_r+1, PP[2][0]).comment = cmt(
    'Mức AQL / Mức kiểm tra',
    'AQL = Acceptable Quality Level (mức chất lượng chấp nhận).',
    'Inspection Level thường dùng Level II (tiêu chuẩn).',
    'Ví dụ: AQL 1.0, Inspection Level II')

r = spacer(ws, r, nc)

# === S2: INSPECTION RESULTS ===
r = sect_header(ws, r, nc, '  2.  INSPECTION RESULTS')
ws.cell(r-1, 1).comment = cmt(
    'Kết quả kiểm tra',
    'Ghi từng đặc tính cần kiểm tra từ bản vẽ.',
    'Mỗi dòng = 1 đặc tính. Điền đầy đủ Result cho mỗi dòng.')

hdrs = ['#', 'Cat.', 'Characteristic', 'Method / Instrument', 'Result', 'Inspector']
r = col_header(ws, r, nc, CK, hdrs)

# Comments on column header LABELS
hdr_r = r - 1
ws.cell(hdr_r, CK[0][0]).comment = cmt(
    'Số thứ tự',
    'Tự động tăng dần. Không cần điền.')
ws.cell(hdr_r, CK[1][0]).comment = cmt(
    'Phân loại đặc tính',
    'QUA = Chất lượng / Kích thước',
    'TEC = Kỹ thuật / Vật liệu',
    'OPS = Vận hành / Ngoại quan')
ws.cell(hdr_r, CK[2][0]).comment = cmt(
    'Đặc tính kiểm tra',
    'Ghi rõ kích thước hoặc yêu cầu từ bản vẽ.',
    'Bao gồm giá trị danh nghĩa và dung sai.',
    'Ví dụ: OD 25.00 ±0.02',
    'Ví dụ: Ra 0.8 max',
    'Ví dụ: Position ⌀0.05 @ MMC')
ws.cell(hdr_r, CK[3][0]).comment = cmt(
    'Phương pháp đo / Dụng cụ',
    'Ghi phương pháp đo và số hiệu dụng cụ.',
    'Dụng cụ phải còn hiệu chuẩn (kiểm tra sticker).',
    'Ví dụ: Micrometer #M-015',
    'Ví dụ: CMM Program REV02',
    'Ví dụ: Profilometer #R-003')
ws.cell(hdr_r, CK[4][0]).comment = cmt(
    'Kết quả',
    'Chọn từ dropdown:',
    'PASS = Trong dung sai, đạt yêu cầu',
    'FAIL = Ngoài dung sai → báo QA lập NCR',
    'HOLD = Cần xem xét thêm, chờ quyết định',
    'NA = Không áp dụng cho đặc tính này')
ws.cell(hdr_r, CK[5][0]).comment = cmt(
    'Người kiểm tra',
    'Ký tắt hoặc ghi tên người thực hiện kiểm tra.',
    'Người kiểm tra phải được đào tạo và có năng lực.',
    '(Xem FRM-807 Skills Matrix)')

ck_start = r
items = [
    (1, 'QUA', 'Dimension 1.', 'Measured value vs tolerance.'),
    (2, 'QUA', 'Dimension 2.', 'Measured value vs tolerance.'),
    (3, 'QUA', 'Dimension 3.', 'Measured value vs tolerance.'),
    (4, 'QUA', 'Visual / surface finish.', 'Per drawing requirement.'),
    (5, 'QUA', 'Material / marking verification.', 'Cert match + marking correct.'),
]
for i, (num, cat, item, crit) in enumerate(items):
    check_row(ws, r, nc, CK, num, cat, item, crit, last=False, ck_start=ck_start)
    r += 1

# Blank rows with auto-number
blank_count = 10
for j in range(blank_count):
    is_last = (j == blank_count - 1)
    ws.row_dimensions[r].height = 20.0
    _borders_table_row(ws, r, nc, CK, is_last)
    bnum = len(items) + j + 1
    even = (bnum % 2 == 0)
    ifill = fi(P['near']) if even else fi(P['white'])
    zfills = [fi(P['fog']), fi(P['fog']), ifill, fi(P['fog']), fi(P['white']), fi(P['white'])]
    for (s, e), zf in zip(CK, zfills):
        for col in range(s, e+1):
            ws.cell(r, col).fill = zf
    for s, e in CK:
        ws.merge_cells(start_row=r, start_column=s, end_row=r, end_column=e)
    ws.cell(r, CK[0][0]).value = f'=ROW()-ROW($A${ck_start})+1'
    ws.cell(r, CK[0][0]).font = fo(8, True, 'mid')
    ws.cell(r, CK[0][0]).alignment = AC
    ws.cell(r, CK[4][0]).font = fo(8, False, 'dk')
    ws.cell(r, CK[4][0]).alignment = AC
    ws.cell(r, CK[5][0]).font = fo(8, False, 'dk')
    ws.cell(r, CK[5][0]).alignment = AC
    r += 1
ck_end = r - 1

r = spacer(ws, r, nc)

# === S3: LOT DISPOSITION ===
r = sect_header(ws, r, nc, '  3.  LOT DISPOSITION')
ws.cell(r-1, 1).comment = cmt(
    'Quyết định lô hàng',
    'Sau khi kiểm tra tất cả đặc tính, quyết định cho toàn lô.',
    'Nếu REJECT hoặc SORT → lập NCR (FRM-651).')

disp_r = r
r = data_pair(ws, r, nc, PP, 'Accept # / Reject #',
              'Lot Decision: ACCEPT / REJECT / SORT', first=True, last=True)
# Comments on LABELS
ws.cell(disp_r, PP[0][0]).comment = cmt(
    'Số chấp nhận / Số từ chối',
    'Ghi số lượng chi tiết đạt và số lượng chi tiết lỗi.',
    'Ví dụ: Accept = 12, Reject = 1')
ws.cell(disp_r, PP[2][0]).comment = cmt(
    'Quyết định lô hàng',
    'ACCEPT = Chấp nhận lô, chuyển tiếp sản xuất / giao hàng.',
    'REJECT = Từ chối toàn bộ lô → lập NCR (FRM-651).',
    'SORT = Kiểm tra 100% toàn bộ lô để phân loại đạt/lỗi.')

r = spacer(ws, r, nc)

# === S4: APPROVAL ===
r = sect_header(ws, r, nc, '  4.  APPROVAL')
ws.cell(r-1, 1).comment = cmt(
    'Phê duyệt',
    'Người kiểm tra ký xác nhận kết quả.',
    'QA Manager ký phê duyệt quyết định lô hàng.')
r = sig_block(ws, r, nc, ['Inspector', 'QA Manager'], zones['sign2'])
r = spacer(ws, r, nc)

# NOTICE
r = notice_bar(ws, r, nc,
    'NOTICE \u2014 One form per lot inspection. '
    'If REJECT or SORT, raise NCR (FRM-651). '
    'Ref: SOP-605. Retain: 10 years.')

# DV + CF
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

dv = DataValidation(type='list', formula1='=LISTS!$A$2:$A$5', allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
rcl = get_column_letter(CK[4][0])
for rr in range(ck_start, ck_end + 1):
    dv.add(f'{rcl}{rr}')

rng = f'{rcl}{ck_start}:{rcl}{ck_end}'
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"PASS"'],
    fill=fi('C6EFCE'), font=Font(color='006100')))
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"HOLD"'],
    fill=fi('FFEB9C'), font=Font(color='9C6500')))
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"FAIL"'],
    fill=fi('FFC7CE'), font=Font(color='9C0006')))

dv2 = DataValidation(type='list', formula1='"ACCEPT,REJECT,SORT"', allow_blank=True)
dv2.showErrorMessage = True
ws.add_data_validation(dv2)
dv2.add(f'{get_column_letter(PP[3][0])}{disp_r}')

# Print
ws.page_setup.paperSize = 9
ws.page_setup.orientation = 'landscape'
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

for name in list(wb.sheetnames):
    if name not in ['FRM_621', 'LISTS']:
        del wb[name]

OUT = "C:/Users/TEST4/Desktop/FRM-621_v3.xlsx"
wb.save(OUT)
sha = hashlib.sha256(open(OUT, 'rb').read()).hexdigest()
print(f"FRM-621 v2: SHA={sha[:12]}")
print(f"Output: {OUT}")
