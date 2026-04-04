"""
FRM-511 Setup and First-Piece Record - Full Upgrade V0 -> V1
CNC Job-Order Precision Machining / Semiconductor Equipment Parts

This is THE most critical shop-floor form in a CNC job shop.
Every setup error costs time, material, and potentially a scrapped part.

Upgrades:
1. CNC-critical reference fields: Program Name/Rev, Fixture ID, Coolant,
   Offset Values, Drawing-Program Rev cross-check
2. 6 new CNC-specific checklist items
3. Parent procedure ref: SOP-504 / WI-517, WI-519
4. Cross-form linkage: FRM-302 Setup Sheet, FRM-519 Pre-Run Check
5. Setup Time field for SMED analysis
6. FP_RESULTS enhancements: CMM Program/Rev, Measurement Uncertainty
7. Conditional formatting
8. Revision history + bump V0->V1
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

from upgrade_lib import (
    COLORS, FONT_LABEL, FONT_INPUT, FONT_BODY, FONT_CAT, FONT_NUM,
    FILL_ICE, FILL_FOG, FILL_WHITE, FILL_NEAR_WHITE, FILL_COL_HDR,
    THIN_ACCENT, THIN_SILVER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_LEFT_WRAP,
    H_DATA,
    add_result_conditional_formatting, add_status_conditional_formatting,
    add_revision_history_sheet, backup_and_save, bump_revision,
    ensure_lists_values,
)

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"
FORM_PATH = f"{BASE}/05-FRM-500/FRM-511_Setup_and_First_Piece_Record.xlsx"

print(f"Loading {FORM_PATH}...")
wb = openpyxl.load_workbook(FORM_PATH)
ws = wb['FRM-511_SETUP']
ws_lists = wb['_LISTS']
ws_fp = wb['FP_RESULTS']
NCOLS = 56  # A4 Landscape

# ============================================================================
# CURRENT LAYOUT (A4L, 56 cols):
# R1-R5: Header (15pt)
# R6: Spacer (4pt)
# R7: Sec 1 "SETUP / FIRST-PIECE HEADER" (17pt)
# R8-R12: Reference fields (5 rows, 20pt each)
#   R8: Job/WO | Part No./Revision
#   R9: Machine/Cell | Operation/Route Step
#   R10: Setup Technician | QA Reviewer
#   R11: Tooling/Fixture Ref | First-Piece ID
#   R12: Evidence Folder | Record Status
# R13: Spacer
# R14: Sec 2 "RUN DECISION / RELEASE BLOCK"
# R15-R18: Gate fields
# R19: Spacer
# R20: Sec 3 "SETUP AND FIRST-PIECE CHECK"
# R21: Column headers
# R22-R31: 10 checklist items
# R32: Spacer
# R33: Sec 4 "ADJUSTMENTS / OPEN ACTIONS"
# ...
# ============================================================================

# ============================================================================
# 1. REPURPOSE EXISTING REFERENCE FIELDS for CNC-critical data
#    R11 currently: "Tooling / Fixture Ref." | "First-Piece ID"
#    R12 currently: "Evidence Folder" | "Record Status"
#    These are generic. We'll make them CNC-specific and add new rows.
# ============================================================================

print("Step 1: Upgrading reference fields for CNC context...")

# R11: Change labels to CNC-specific
ws.cell(row=11, column=1).value = 'CNC Program / Rev'
ws.cell(row=11, column=29).value = 'Fixture ID'
# Note: col 29 = AC in A4L (label starts at AC for right-side fields)

# R12: Keep Evidence Folder but change right side
ws.cell(row=12, column=29).value = 'Setup Time (min)'

# Insert 3 new reference rows after R12 for additional CNC fields
INSERT_REF = 13  # Insert before current R13 (spacer)
NUM_REF_INSERT = 3
ws.insert_rows(INSERT_REF, amount=NUM_REF_INSERT)

# After insertion: R13-R15 = new, R16 = old spacer, R17+ shifted
# New reference field rows use same merge pattern as R8-R12

# Reference field merge pattern (A4L, 56 cols):
# A:J(1-10)=left-label | K:AB(11-28)=left-input | AC:AI(29-35)=right-label | AJ:BD(36-56)=right-input
REF_ZONES_L = [(1,10), (11,28), (29,35), (36,56)]

def make_ref_row_l(row, l_label, r_label, l_value=None, r_value=None, locked=False):
    """Make a reference field row in A4L (56-col) layout."""
    for s, e in REF_ZONES_L:
        if e > s:
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    ws.row_dimensions[row].height = H_DATA

    c = ws.cell(row=row, column=1)
    c.value = l_label; c.font = FONT_LABEL; c.fill = FILL_FOG
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_ACCENT, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=11)
    c.value = l_value; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)

    c = ws.cell(row=row, column=29)
    c.value = r_label; c.font = FONT_LABEL; c.fill = FILL_FOG
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=36)
    c.value = r_value; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_ACCENT, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)


# R13: Coolant + Offset values
make_ref_row_l(13, 'Coolant Type / Concentration', 'Offset Values Recorded?')

# R14: Drawing-Program Rev cross-check + linked forms
make_ref_row_l(14, 'Drawing Rev = Program Rev?', 'Linked FRM-302 Setup Sheet #')

# R15: Parent procedure ref + linked pre-run check
make_ref_row_l(15, 'Ref: SOP-504 / WI-517, WI-519', 'Linked FRM-519 Pre-Run #',
               l_value='SOP-504 (Setup, First-Piece, Run Release)',
               r_value=None, locked=False)
# Lock the left label+value for parent procedure
ws.cell(row=15, column=11).font = Font(name='Segoe UI', size=8, color=COLORS['SUBTITLE'])
ws.cell(row=15, column=11).fill = FILL_FOG
ws.cell(row=15, column=11).protection = Protection(locked=True)

# Add YES/NO validation for "Drawing Rev = Program Rev?" and "Offset Values Recorded?"
for cell_ref in ['AJ14', 'AJ13']:
    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_ref)

# ============================================================================
# 2. INSERT CNC-SPECIFIC CHECKLIST ITEMS
#    After insertion, checklist is at R25-R34 (was R22-R31, shifted by 3)
#    Insert 6 new items after R34, before spacer R35
# ============================================================================

print("Step 2: Adding 6 CNC-specific checklist items...")

# Current checklist rows after ref insertion: R25-R34 (shifted +3)
CK_LAST = 34  # Last original checklist item (was R31, now R34)
INSERT_CK = CK_LAST + 1  # Insert at R35
NUM_CK_INSERT = 6
ws.insert_rows(INSERT_CK, amount=NUM_CK_INSERT)

# New items at R35-R40

# Checklist merge zones (A4L, 56 cols):
# A:C(1-3)=# | D:G(4-7)=Cat | H:Z(8-26)=Item | AA:AQ(27-43)=Criteria | AR:AW(44-49)=Result | AX:AZ(50-52)=Owner | BA:BD(53-56)=Ref
CK_ZONES_L = [(1,3), (4,7), (8,26), (27,43), (44,49), (50,52), (53,56)]

def make_ck_row_l(row, num, cat, cat_color, item, criteria, even=False):
    for s, e in CK_ZONES_L:
        if e > s:
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    ws.row_dimensions[row].height = H_DATA

    c = ws.cell(row=row, column=1)
    c.value = num; c.font = FONT_NUM; c.fill = FILL_FOG
    c.alignment = ALIGN_CENTER
    c.border = Border(left=THIN_ACCENT, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=4)
    c.value = cat; c.font = FONT_CAT
    c.fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type='solid')
    c.alignment = ALIGN_CENTER
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=8)
    c.value = item; c.font = FONT_BODY
    c.fill = FILL_NEAR_WHITE if even else FILL_WHITE
    c.alignment = ALIGN_LEFT_WRAP
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=27)
    c.value = criteria; c.font = FONT_BODY; c.fill = FILL_FOG
    c.alignment = ALIGN_LEFT_WRAP
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    for col in [44, 50, 53]:
        c = ws.cell(row=row, column=col)
        c.font = FONT_INPUT; c.fill = FILL_WHITE; c.alignment = ALIGN_CENTER
        c.border = Border(left=THIN_SILVER, right=THIN_SILVER if col != 53 else THIN_ACCENT,
                          top=THIN_ACCENT, bottom=THIN_ACCENT)
        c.protection = Protection(locked=False)


cnc_setup_items = [
    (11, 'CNC', COLORS['CAT_CNC'],
     'CNC program name and revision match released drawing.',
     'Program file, drawing rev, and setup sheet rev are identical.', False),
    (12, 'CNC', COLORS['CAT_CNC'],
     'Tool offsets verified against setup sheet values.',
     'Length and diameter offsets match FRM-302. No manual adjustment without log.', True),
    (13, 'CNC', COLORS['CAT_CNC'],
     'Coolant type, concentration, and flow are correct.',
     'Coolant matches material spec. Flow rate adequate for operation.', False),
    (14, 'CNC', COLORS['CAT_CNC'],
     'Workholding / clamping force verified for part geometry.',
     'No part shift risk. Jaw marks / clamping deformation acceptable.', True),
    (15, 'GAGE', COLORS['CAT_TEC'],
     'CMM program revision matches current drawing revision.',
     'CMM program ID, revision, and datum strategy are documented.', False),
    (16, 'QUA', COLORS['CAT_QUA'],
     'First-piece measurement uncertainty is acceptable for tolerance.',
     'Gage R&R / MSA adequate. U(k=2) < 25% of tolerance band.', True),
]

for i, (num, cat, cc, item, crit, even) in enumerate(cnc_setup_items):
    make_ck_row_l(35 + i, num, cat, cc, item, crit, even)
    print(f"  + Item #{num}: {item[:50]}...")

# ============================================================================
# 3. ADD VALIDATIONS for new checklist items
# ============================================================================
print("Step 3: Adding data validations...")

for r in range(35, 41):
    dv = DataValidation(type="list", formula1="=VAL_RESULT_CORE", allow_blank=True)
    dv.error = "Select PASS, HOLD, FAIL, or NA."
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"AR{r}")

# ============================================================================
# 4. UPDATE GATE FORMULAS (now AR25:AR40 instead of AR22:AR31)
# ============================================================================
print("Step 4: Updating gate formulas for 16 items...")

# Gate cells shifted: R15→R18, R18→R21 (shifted by 3 from ref insert)
# Find gate formula cells
for r in range(14, 25):
    c = ws.cell(row=r, column=39)  # AM column
    if c.value and isinstance(c.value, str) and 'COUNTA' in c.value:
        # Update range
        old_val = c.value
        new_val = old_val.replace('AR22:AR31', 'AR25:AR40')
        c.value = new_val
        print(f"  Updated gate formula at R{r}: AR22:AR31 -> AR25:AR40")

# ============================================================================
# 5. CONDITIONAL FORMATTING
# ============================================================================
print("Step 5: Adding conditional formatting...")

add_result_conditional_formatting(ws, "AR25:AW40")
add_status_conditional_formatting(ws, "AM21:BD21")

# ============================================================================
# 6. ENHANCE FP_RESULTS sheet
# ============================================================================
print("Step 6: Enhancing FP_RESULTS sheet...")

# Add columns for CMM Program/Rev and Measurement Uncertainty
# Current columns in FP_RESULTS (row 16 headers):
# A=#, B=Characteristic, C=Requirement, D=Actual Result, E=Result, F=Reviewer, G=Evidence, H=Note

# We need to check if there's room to add columns or use existing H (Note)
# Since we can't change grid, we'll use the Note column and add info to the header guidance
ws_fp.cell(row=15, column=1).value = (
    'One line per first-piece result. Include CMM program rev in Evidence Ref. '
    'Note measurement uncertainty U(k=2) in Note column for tight-tolerance features.'
)

# Update column H header to be more specific
ws_fp.cell(row=16, column=8).value = 'U(k=2) / Note'

# ============================================================================
# 7. ENRICH _LISTS
# ============================================================================
print("Step 7: Enriching _LISTS...")

ensure_lists_values(ws_lists, 'VAL_YES_NO', ['YES', 'NO', 'NA'])
ensure_lists_values(ws_lists, 'VAL_RISK_LEVEL_CORE', ['LOW', 'MEDIUM', 'HIGH', 'CRITICAL'])
ensure_lists_values(ws_lists, 'VAL_CK_CATEGORY',
    ['SET', 'TOOL', 'GAGE', 'QUA', 'OPS', 'CON', 'REL', 'CNC'])

# ============================================================================
# 8. REVISION HISTORY + BUMP
# ============================================================================
print("Step 8: Adding revision history and bumping V0 -> V1...")

add_revision_history_sheet(wb, 'FRM-511')
bump_revision(ws, 'V1', '2026-03-22')
bump_revision(ws_fp, 'V1', '2026-03-22')

# ============================================================================
# 9. PROTECT FORMULAS
# ============================================================================
print("Step 9: Protecting formula cells...")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            cell.protection = Protection(locked=True)

# ============================================================================
# 10. UPDATE NOTICE BAR
# ============================================================================
print("Step 10: Updating notice bar...")

for r in range(ws.max_row, max(ws.max_row - 10, 1), -1):
    c = ws.cell(row=r, column=1)
    if c.value and 'NOTICE' in str(c.value):
        c.value = (
            "NOTICE  -  Enter data only in designated white input cells. "
            "Do not add, delete or resize columns/rows. "
            "File naming: FRM-511_{JobNo}_{PartNo}.xlsx. "
            "Ref: SOP-504 / WI-517, WI-519. "
            "Retain 10 years per QMS retention schedule. "
            "Verify program rev = drawing rev before first cut."
        )
        print(f"  Updated notice at R{r}")
        break

# ============================================================================
# 11. SAVE
# ============================================================================
print("Step 11: Saving...")
sha = backup_and_save(wb, FORM_PATH)

print(f"""
{'='*70}
FRM-511 SETUP AND FIRST-PIECE RECORD - UPGRADE COMPLETE
{'='*70}
Version:    V0 -> V1
Date:       2026-03-22

Changes:
  + Reference fields upgraded for CNC:
    - CNC Program / Rev (was Tooling/Fixture generic)
    - Fixture ID
    - Setup Time (min) for SMED tracking
    - Coolant Type / Concentration
    - Offset Values Recorded? (YES/NO dropdown)
    - Drawing Rev = Program Rev? (YES/NO cross-check gate)
    - Linked FRM-302 Setup Sheet #
    - Linked FRM-519 Pre-Run Check #
    - Parent procedure ref: SOP-504 / WI-517, WI-519
  + 6 CNC-specific checklist items (#11-#16):
    #11 CNC program matches drawing revision
    #12 Tool offsets verified against setup sheet
    #13 Coolant type and concentration correct
    #14 Workholding / clamping force verified
    #15 CMM program revision matches drawing
    #16 First-piece measurement uncertainty acceptable
  + FP_RESULTS: U(k=2) column guidance added
  + Conditional formatting: FAIL/RED, HOLD/AMBER, PASS/GREEN
  + Gate formulas updated for 16 items
  + _LISTS enriched with YES_NO, RISK_LEVEL, CK_CATEGORY
  + Revision history sheet
  + Formula cells protected

SHA-256: {sha}
{'='*70}
""")
