"""
FRM-641 Final Inspection Report - Full Upgrade V0 -> V1
CNC Job-Order Precision Machining / Semiconductor Equipment Parts

This is THE quality gate before shipping. Every part leaving the facility
must pass through this form. For semiconductor equipment parts, this form
is the last line of defense against nonconforming product reaching the customer.

Upgrades:
1. CNC-critical fields: CMM Report Attached, Measurement Uncertainty notation,
   Concession/Deviation Ref, Customer Source Inspection, Packaging Instruction Ref
2. Cross-form linkage: FRM-651 NCR, FRM-642 CoC, FRM-302 Setup Sheet
3. Safety/Special Characteristics flag guidance
4. Parent procedure ref: SOP-605 / WI-605
5. Records retention notation: 10 years
6. Conditional formatting on Result cells
7. LISTS standardization -> _LISTS with named ranges
8. RESULT_DETAIL enhancements: U(k=2) column
9. Revision history + bump V0->V1
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

from upgrade_lib import (
    COLORS, FONT_LABEL, FONT_INPUT, FONT_BODY, FONT_CAT, FONT_NUM,
    FILL_ICE, FILL_FOG, FILL_WHITE, FILL_NEAR_WHITE,
    THIN_ACCENT, THIN_SILVER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_LEFT_WRAP,
    H_DATA,
    add_result_conditional_formatting, add_status_conditional_formatting,
    add_revision_history_sheet, backup_and_save, bump_revision,
    ensure_lists_values,
)

BASE = "C:/Users/TEST4/qms.hesem.com.vn/04-Bieu-Mau"
FORM_PATH = f"{BASE}/06-FRM-600/FRM-641_Final_Inspection_Report.xlsx"

print(f"Loading {FORM_PATH}...")
wb = openpyxl.load_workbook(FORM_PATH)
ws = wb['FINAL_INSP']
ws_lists = wb['LISTS']
ws_detail = wb['RESULT_DETAIL']
NCOLS = 56  # A4 Landscape

# ============================================================================
# CURRENT LAYOUT (A4L, 56 cols):
# R1-R5: Header
# R6: Spacer
# R7: Sec 1 "LOT / PLAN / REFERENCE INFORMATION"
# R8-R12: Reference fields
# R13: Spacer
# R14: Sec 2 "RESULT / SAMPLE EXECUTION SUMMARY"
# R15: Col headers
# R16-R22: 7 inspection result rows
# R23: Spacer
# R24: Sec 4 "DISPOSITION / RELEASE"
# R25-R27: Disposition fields
# R28: Spacer
# R29: Sec 5 "APPROVAL / SIGN-OFF"
# R30-R33: Approval block
# R34: Spacer
# R35: Notice bar
# ============================================================================

# ============================================================================
# 1. UPGRADE REFERENCE FIELDS for CNC context
# ============================================================================
print("Step 1: Upgrading reference fields...")

# R11: Currently "Inspection Basis | Evidence Package Ref"
# Change to CNC-specific
ws.cell(row=11, column=1).value = 'Inspection Basis / CMM Program'
ws.cell(row=11, column=29).value = 'CoC Number / FRM-642 Ref'

# Insert 3 new reference rows after R12
INSERT_REF = 13
ws.insert_rows(INSERT_REF, amount=3)

# After: R13-R15 = new, R16 = old spacer, R17+ shifted

# Reference merge pattern (A4L):
# A:J(1-10)=left-label | K:AB(11-28)=left-input | AC:AI(29-35)=right-label | AJ:BD(36-56)=right-input
REF_ZONES = [(1,10), (11,28), (29,35), (36,56)]

def make_ref(row, ll, rl, lv=None, rv=None, locked=False):
    for s, e in REF_ZONES:
        if e > s:
            ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    ws.row_dimensions[row].height = H_DATA

    c = ws.cell(row=row, column=1)
    c.value = ll; c.font = FONT_LABEL; c.fill = FILL_FOG; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_ACCENT, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=11)
    c.value = lv; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)

    c = ws.cell(row=row, column=29)
    c.value = rl; c.font = FONT_LABEL; c.fill = FILL_FOG; c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_SILVER, top=THIN_ACCENT, bottom=THIN_ACCENT)

    c = ws.cell(row=row, column=36)
    c.value = rv; c.font = FONT_INPUT
    c.fill = FILL_FOG if locked else FILL_WHITE
    c.alignment = ALIGN_LEFT
    c.border = Border(left=THIN_SILVER, right=THIN_ACCENT, top=THIN_ACCENT, bottom=THIN_ACCENT)
    c.protection = Protection(locked=locked)


# R13: CMM Report + Customer Source Inspection
make_ref(13, 'CMM Report Attached?', 'Customer Source Insp. Required?')

# R14: Cross-form linkage
make_ref(14, 'Linked NCR Refs (FRM-651)', 'Linked Setup Sheet (FRM-302)')

# R15: Parent procedure + retention
make_ref(15, 'Ref: SOP-605 / WI-605', 'Retain: 10 years',
         lv='SOP-605 (Final Inspection and Release)',
         rv='Per QMS retention schedule', locked=True)

# YES/NO validations for new fields
for cell_ref in ['AJ13']:
    dv = DataValidation(type="list", formula1='"YES,NO,NA"', allow_blank=True)
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_ref)

# CMM Report: YES/NO
dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
dv.showErrorMessage = True
ws.add_data_validation(dv)
dv.add('K13')

# ============================================================================
# 2. UPGRADE DISPOSITION BLOCK
#    After ref insertion (+3): R24->R27, R25->R28, etc.
#    Disposition section now at R27
# ============================================================================
print("Step 2: Upgrading disposition block...")

# Insert 2 rows in disposition section for new fields
# Find disposition section (shifted +3)
# Old R25 (Disposition) -> now R28
# Old R26 (Cleanliness) -> now R29
# Old R27 (Quality Authority) -> now R30
# Insert after R30 for concession/deviation ref and packaging ref

DISP_INSERT = 31  # After the Quality Authority row
ws.insert_rows(DISP_INSERT, amount=2)

# R31: Concession/Deviation + Packaging Ref
make_ref(31, 'Concession / Deviation Ref', 'Packaging Instruction Ref (FRM-707/709)')

# R32: Safety characteristics flag
make_ref(32, 'Safety/Special Char. Verified?', 'Measurement Uncertainty Noted?')

# YES/NO for these
for cell_ref in ['K31', 'K32', 'AJ32']:
    dv = DataValidation(type="list", formula1='"YES,NO,NA"', allow_blank=True)
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_ref)

# ============================================================================
# 3. CONDITIONAL FORMATTING
# ============================================================================
print("Step 3: Adding conditional formatting...")

# Result column for inspection items (shifted +3): AR19:AR25
add_result_conditional_formatting(ws, "AR19:AW25")

# Disposition cell (shifted)
add_result_conditional_formatting(ws, "K28:AB28")

# ============================================================================
# 4. ENHANCE RESULT_DETAIL sheet
# ============================================================================
print("Step 4: Enhancing RESULT_DETAIL sheet...")

ws_detail.cell(row=2, column=1).value = (
    'Final inspection event record. Consumes outputs from FRM-621/631/643/712/707/709. '
    'For tight-tolerance features, note U(k=2) in Notes column. '
    'Flag Safety/Special Characteristics with SC prefix in Characteristic column.'
)

# Add column header for measurement uncertainty guidance
if ws_detail.cell(row=4, column=6).value == 'Notes':
    ws_detail.cell(row=4, column=6).value = 'Notes / U(k=2)'

# ============================================================================
# 5. STANDARDIZE LISTS sheet
#    Currently uses 'LISTS' with direct cell references like 'LISTS'!$A$2:$A$5
#    Rename to _LISTS would break existing validations, so we keep 'LISTS'
#    but add new columns
# ============================================================================
print("Step 5: Enriching LISTS sheet...")

ensure_lists_values(ws_lists, 'RISK_LEVEL', ['LOW', 'MEDIUM', 'HIGH', 'CRITICAL'])

# Check for DISCOVERY_STAGE
ensure_lists_values(ws_lists, 'DISCOVERY_STAGE',
    ['INCOMING', 'SETUP', 'FIRST PIECE', 'IN-PROCESS',
     'FINAL INSPECTION', 'PACKAGING', 'SHIPPING', 'CUSTOMER'])

# ============================================================================
# 6. REVISION HISTORY + BUMP
# ============================================================================
print("Step 6: Adding revision history and bumping V0 -> V1...")

add_revision_history_sheet(wb, 'FRM-641')
bump_revision(ws, 'V1', '2026-03-22')

# ============================================================================
# 7. PROTECT FORMULAS
# ============================================================================
print("Step 7: Protecting formula cells...")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            cell.protection = Protection(locked=True)

# ============================================================================
# 8. UPDATE NOTICE BAR
# ============================================================================
print("Step 8: Updating notice bar...")

for r in range(ws.max_row, max(ws.max_row - 10, 1), -1):
    c = ws.cell(row=r, column=1)
    if c.value and 'NOTICE' in str(c.value):
        c.value = (
            "NOTICE: Complete in English. No back-dating, history deletion, "
            "proxy-signing or vague comments. "
            "Ref: SOP-605 / WI-605. Retain 10 years. "
            "Flag SC dimensions. Note U(k=2) for tight tolerances. "
            "File: FRM-641_{JobNo}_{PartNo}.xlsx"
        )
        print(f"  Updated notice at R{r}")
        break

# ============================================================================
# 9. SAVE
# ============================================================================
print("Step 9: Saving...")
sha = backup_and_save(wb, FORM_PATH)

print(f"""
{'='*70}
FRM-641 FINAL INSPECTION REPORT - UPGRADE COMPLETE
{'='*70}
Version:    V0 -> V1
Date:       2026-03-22

Changes:
  + Reference fields upgraded:
    - CMM Report Attached? (YES/NO)
    - Customer Source Inspection Required? (YES/NO)
    - CoC Number / FRM-642 Ref
    - Inspection Basis / CMM Program
    - Linked NCR Refs (FRM-651)
    - Linked Setup Sheet (FRM-302)
    - Parent procedure: SOP-605 / WI-605
    - Records retention: 10 years
  + Disposition block enhanced:
    - Concession / Deviation Ref
    - Packaging Instruction Ref (FRM-707/709)
    - Safety/Special Char. Verified? (YES/NO)
    - Measurement Uncertainty Noted? (YES/NO)
  + RESULT_DETAIL: U(k=2) guidance added
  + Conditional formatting: FAIL/RED, HOLD/AMBER, PASS/GREEN
  + LISTS enriched: RISK_LEVEL, DISCOVERY_STAGE
  + Revision history sheet
  + Formula cells protected

SHA-256: {sha}
{'='*70}
""")
