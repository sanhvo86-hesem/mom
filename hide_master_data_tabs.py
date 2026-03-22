#!/usr/bin/env python3
"""
HESEM QMS - Hide Master Data Tabs
Sets non-operational tabs to 'veryHidden' (only accessible via VBA/developer).
Keeps operational tabs visible for shop floor use.
"""
import os, sys, glob

from openpyxl import load_workbook

# ═══════ MASTER DATA TAB PATTERNS (to hide) ═══════
# These are system/reference tabs, not used during daily operations
HIDE_PATTERNS = [
    # Dropdown / validation data
    '_lists', 'zlists', 'lists',
    # Import / reference data
    '_import', 'ref_import', 'import_ref', 'ref_links',
    # Evidence / audit trail (admin only)
    'evidence_ref', 'evidence',
    # Print system
    'print_control', 'generator', 'print_face',
    # Calculation / summary backends
    'calc_summary', 'source_ref', 'result_detail', 'raw_data',
    'chart_data',
    # Review summaries (admin periodic review)
    'review_summary',
    # Reference tabs in FRM-1xx series
    'rankref', 'reactionref', 'revhist', 'roleref',
    'certref', 'sessionref', 'taskref', 'criteria',
    'expiryq', 'printctl',
    # Import/filter views
    'filter_view', 'annex_map',
]

# ═══════ OPERATIONAL TAB PATTERNS (keep visible) ═══════
# These are used by operators during daily work
KEEP_VISIBLE_PATTERNS = [
    # Main form (always first tab)
    'gate', 'log', 'track', 'main', 'register', 'checklist',
    'dispatch', 'cost', 'setup', 'audit', 'assess',
    'form', 'record', 'report', 'index', 'plan',
    'invoice', 'survey', 'certificate', 'matrix',
    # Action/CAPA tabs
    'action', 'capa', 'finding', 'risk',
    # Operational support (used during work)
    'milestone', 'blocker', 'open_item', 'review',
    'cost_line', 'tool_list', 'sketch', 'step',
    'shortage', 'hold', 'sample', 'comparison',
    'containment', 'function_test', 'mismatch',
    'trend', 'gap', 'legend', 'score',
    'overlay', 'baseline', 'signoff',
    'decision', 'input', 'kpi',
    # Data entry tabs (operators write to these)
    'data_log', 'batch', 'verify', 'test',
]


def should_hide(sheet_name, sheet_index, total_sheets):
    """Determine if a tab should be hidden."""
    lower = sheet_name.lower()

    # RULE 1: First tab is ALWAYS visible (main form)
    if sheet_index == 0:
        return False

    # RULE 2: Check explicit hide patterns
    for pat in HIDE_PATTERNS:
        if pat in lower or lower == pat:
            return True

    # RULE 3: Check explicit keep-visible patterns
    for pat in KEEP_VISIBLE_PATTERNS:
        if pat in lower:
            return False

    # RULE 4: Tabs starting with _ are system tabs
    if lower.startswith('_'):
        return True

    # RULE 5: Default - keep visible (safer to show than hide)
    return False


def process(filepath):
    fname = os.path.basename(filepath)
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, f"ERROR: {e}", 0, 0

    hidden_count = 0
    visible_count = 0
    changes = []

    for idx, sn in enumerate(wb.sheetnames):
        ws = wb[sn]
        hide = should_hide(sn, idx, len(wb.sheetnames))

        if hide:
            if ws.sheet_state != 'veryHidden':
                ws.sheet_state = 'veryHidden'
                changes.append(f"  HIDE: {sn}")
            hidden_count += 1
        else:
            if ws.sheet_state != 'visible':
                ws.sheet_state = 'visible'
                changes.append(f"  SHOW: {sn}")
            visible_count += 1

    try:
        wb.save(filepath)
    except Exception as e:
        wb.close()
        return fname, f"SAVE ERROR: {e}", visible_count, hidden_count

    wb.close()
    return fname, changes, visible_count, hidden_count


def main():
    base = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS - HIDE MASTER DATA TABS")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    v = '--verbose' in sys.argv
    total_vis = total_hid = total_changed = 0
    errors = 0

    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        print(f"[{i:3d}/{len(files)}] {fn}...", end=" ", flush=True)
        name, changes, vis, hid = process(fp)

        if isinstance(changes, str):
            print(f"FAIL: {changes}")
            errors += 1
        else:
            n_changes = len(changes)
            print(f"OK ({vis}V/{hid}H" + (f", {n_changes} changed)" if n_changes else ")"))
            total_vis += vis
            total_hid += hid
            if n_changes:
                total_changed += 1
            if v and changes:
                for c in changes:
                    print(c)

    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"  Files processed: {len(files)}")
    print(f"  Files changed:   {total_changed}")
    print(f"  Total visible:   {total_vis} tabs")
    print(f"  Total hidden:    {total_hid} tabs")
    print(f"  Errors:          {errors}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
