#!/usr/bin/env python3
"""
HESEM QMS - Tab Consolidation (World-Class Standard)

Based on case studies: Toyota A3, Bosch 8D, AS9102, TSMC.

3 Categories:
  A: Checklist/Gate/Event → 1 visible tab (merge all support into main)
  B: Event + CAPA/Actions → 2 visible tabs (main + actions)
  C: Rolling Log → 1-2 visible tabs (log + optional summary)

All other tabs → veryHidden
"""
import os, sys, glob
from openpyxl import load_workbook

# ═══════ CATEGORY B FORMS ═══════
# These KEEP a separate Actions/CAPA tab (different lifecycle)
# Pattern: event investigation + separate action tracking
CATEGORY_B_FORMS = {
    # NCR / CAPA / Problem solving (Bosch 8D model)
    'FRM-651', 'FRM-652', 'FRM-653',
    # SCAR (supplier corrective action)
    'FRM-406',
    # Crash/Incident reports (event + follow-up)
    'FRM-522', 'FRM-811',
    # Audit checklists (checklist + findings)
    'FRM-901', 'FRM-902',
    'FRM-163', 'FRM-111', 'FRM-409',
}

# ═══════ CATEGORY C FORMS ═══════
# Rolling logs with many data rows - keep data tab + optional summary
CATEGORY_C_FORMS = {
    'FRM-101', 'FRM-203', 'FRM-208',
    'FRM-401', 'FRM-503', 'FRM-512', 'FRM-513',
    'FRM-523', 'FRM-524', 'FRM-525',
    'FRM-601', 'FRM-602', 'FRM-631',
    'FRM-701', 'FRM-708',
    'FRM-806', 'FRM-807', 'FRM-809',
}

# ═══════ TABS THAT ARE ALWAYS OPERATIONAL (visible) ═══════
OPERATIONAL_TAB_KEYWORDS = [
    'action', 'finding', 'capa', '8d',     # CAPA/Actions (Cat B only)
    'risk', 'blocker', 'open_item',         # Risk tracking
    'milestone', 'cost_line', 'tool_list',  # Operational support
    'sketch', 'step', 'containment',        # Work instructions
    'shortage', 'hold', 'mismatch',         # Exception handling
    'kpi', 'score', 'trend',                # KPI/metrics
    'decision', 'recovery',                 # Decision tracking
    'comparison', 'sample', 'function_test',# Testing
    'obligation', 'impact',                 # Review
    'rev_history',                          # Document control
]

# ═══════ TABS THAT ARE ALWAYS MASTER DATA (hidden) ═══════
MASTER_DATA_TAB_KEYWORDS = [
    '_lists', 'zlists', 'lists',            # Dropdowns
    '_import', 'ref_import', 'import_ref',  # Import data
    'ref_links',                            # Reference links
    'evidence_ref', 'evidence',             # Evidence (admin)
    'print_control', 'generator', 'print_face', 'print_log',  # Print system
    'calc_summary', 'source_ref',           # Calculation backends
    'result_detail', 'raw_data', 'chart_data',  # Data backends
    'review_summary',                       # Periodic review (admin)
    'filter_view', 'annex_map',             # System views
    'rankref', 'reactionref', 'roleref',    # Reference lookups
    'certref', 'sessionref', 'taskref',     # Reference lookups
    'criteria', 'expiryq', 'printctl',      # System
    'revhist',                              # Rev history (hidden in FRM-1xx)
]


def get_form_code(filename):
    """Extract FRM-xxx from filename."""
    parts = filename.split('_')[0]
    return parts


def classify_form(filename):
    """Classify form into Category A, B, or C."""
    code = get_form_code(filename)
    if code in CATEGORY_B_FORMS:
        return 'B'
    if code in CATEGORY_C_FORMS:
        return 'C'
    return 'A'


def is_master_data(sheet_name):
    """Check if tab is master data (should be hidden)."""
    lower = sheet_name.lower()
    for kw in MASTER_DATA_TAB_KEYWORDS:
        if kw in lower or lower == kw:
            return True
    if lower.startswith('_'):
        return True
    return False


def is_operational(sheet_name):
    """Check if tab is operational (should be visible)."""
    lower = sheet_name.lower()
    for kw in OPERATIONAL_TAB_KEYWORDS:
        if kw in lower:
            return True
    return False


def process(filepath):
    fname = os.path.basename(filepath)
    category = classify_form(fname)

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, category, f"ERROR: {e}", 0, 0

    changes = []
    vis_count = 0
    hid_count = 0

    for idx, sn in enumerate(wb.sheetnames):
        ws = wb[sn]
        lower = sn.lower()

        # RULE 1: First tab always visible (main form)
        if idx == 0:
            if ws.sheet_state != 'visible':
                ws.sheet_state = 'visible'
                changes.append(f"  SHOW: {sn} (main form)")
            vis_count += 1
            continue

        # RULE 2: Master data tabs → veryHidden
        if is_master_data(sn):
            if ws.sheet_state != 'veryHidden':
                ws.sheet_state = 'veryHidden'
                changes.append(f"  HIDE: {sn} (master data)")
            hid_count += 1
            continue

        # RULE 3: Category-specific logic
        if category == 'A':
            # Category A: ONLY main tab visible, everything else hidden
            # Exception: if it's clearly an operational support tab
            if is_operational(sn):
                # Even in Cat A, some operational tabs stay visible
                # But gate/checklist forms shouldn't have many
                if ws.sheet_state != 'visible':
                    ws.sheet_state = 'visible'
                    changes.append(f"  SHOW: {sn} (operational)")
                vis_count += 1
            else:
                if ws.sheet_state != 'veryHidden':
                    ws.sheet_state = 'veryHidden'
                    changes.append(f"  HIDE: {sn} (Cat A: single-tab)")
                hid_count += 1

        elif category == 'B':
            # Category B: Main + Actions/Findings visible
            if is_operational(sn):
                if ws.sheet_state != 'visible':
                    ws.sheet_state = 'visible'
                    changes.append(f"  SHOW: {sn} (Cat B: actions/findings)")
                vis_count += 1
            else:
                if ws.sheet_state != 'veryHidden':
                    ws.sheet_state = 'veryHidden'
                    changes.append(f"  HIDE: {sn} (Cat B: support)")
                hid_count += 1

        elif category == 'C':
            # Category C: Main log + operational support visible
            if is_operational(sn):
                if ws.sheet_state != 'visible':
                    ws.sheet_state = 'visible'
                    changes.append(f"  SHOW: {sn} (Cat C: support)")
                vis_count += 1
            else:
                if ws.sheet_state != 'veryHidden':
                    ws.sheet_state = 'veryHidden'
                    changes.append(f"  HIDE: {sn} (Cat C: non-operational)")
                hid_count += 1

    try:
        wb.save(filepath)
    except Exception as e:
        wb.close()
        return fname, category, f"SAVE ERROR: {e}", vis_count, hid_count

    wb.close()
    return fname, category, changes, vis_count, hid_count


def main():
    base = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS - TAB CONSOLIDATION (World-Class Standard)")
    print(f"  Cat A: 1 visible tab (checklist/gate/event)")
    print(f"  Cat B: 2 visible tabs (event + CAPA/actions)")
    print(f"  Cat C: 1-2 visible tabs (rolling log + summary)")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    v = '--verbose' in sys.argv
    stats = {'A': 0, 'B': 0, 'C': 0}
    total_vis = total_hid = total_changed = errors = 0

    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        name, cat, changes, vis, hid = process(fp)
        stats[cat] += 1

        status = "OK"
        if isinstance(changes, str):
            status = changes
            errors += 1
        else:
            n = len(changes)
            if n: total_changed += 1

        print(f"[{i:3d}/{len(files)}] [{cat}] {fn[:45]:45s} {vis}V/{hid}H {status}")
        total_vis += vis
        total_hid += hid

        if v and isinstance(changes, list) and changes:
            for c in changes:
                print(c)

    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"  Category A (1 tab):  {stats['A']} forms")
    print(f"  Category B (2 tabs): {stats['B']} forms")
    print(f"  Category C (log):    {stats['C']} forms")
    print(f"  Total visible tabs:  {total_vis}")
    print(f"  Total hidden tabs:   {total_hid}")
    print(f"  Files changed:       {total_changed}")
    print(f"  Errors:              {errors}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
