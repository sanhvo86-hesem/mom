#!/usr/bin/env python3
"""Batch 2: Auto-fill more Vietnamese translations for yellow-marked words."""
import openpyxl
from openpyxl.styles import PatternFill

BATCH2 = {
    # ── Vietnamese words (not English) ──
    'chai': 'chai',
    'khoang': 'khoang',
    'Trao': 'Trao',
    'bulong': 'bu-lông',
    'cali': 'cali (đo)',

    # ── Keep English (codes/abbreviations) ──
    'upsert': 'cập nhật/thêm mới (upsert)',
    'thru': 'xuyên qua (thru)',
    'APAR': 'APAR',
    'GUID': 'GUID',
    'CSAT': 'CSAT',
    'MTBF': 'MTBF',
    'PESTLE': 'PESTLE',
    'YYYYMM': 'YYYYMM',
    'NCREL': 'NCREL',
    'INDUSTRIAL': 'CÔNG NGHIỆP',
    'STOREKEEPER': 'THỦ KHO',
    'CLERK': 'NHÂN VIÊN VĂN PHÒNG',
    'ADMINISTRATOR': 'QUẢN TRỊ VIÊN',
    'KPIs': 'KPI',
    'CAPASource': 'CAPASource',
    'Partho': 'Partho',

    # ── Manufacturing / CNC ──
    'case': 'trường hợp / vỏ',
    'mating': 'lắp ghép',
    'humidity': 'độ ẩm',
    'screw': 'vít',
    'grip': 'kẹp / nắm',
    'parting': 'cắt đứt (parting)',
    'chips': 'phoi',
    'interference': 'lắp chặt (interference fit)',
    'retract': 'rút về',
    'locator': 'chốt định vị',
    'seat': 'bề mặt tựa / ổ',
    'etch': 'khắc axit (etch)',
    'spacer': 'miếng đệm',
    'flange': 'mặt bích (flange)',
    'endmill': 'dao phay ngón (endmill)',
    'gauges': 'dưỡng đo',
    'surfaces': 'bề mặt',
    'body': 'thân',
    'bare': 'trần / không phủ',
    'splash': 'bắn / văng',
    'tuning': 'tinh chỉnh',
    'milling': 'phay',
    'sharp': 'sắc / bén',
    'dimensional': 'kích thước',
    'roughness': 'độ nhám',
    'cushioning': 'đệm / chống sốc',
    'chiller': 'máy làm lạnh',
    'gloves': 'găng tay',
    'extraction': 'hút / chiết xuất',
    'loose': 'lỏng',
    'ergonomic': 'công thái học',
    'door': 'cửa',
    'wall': 'tường',
    'camera': 'camera',
    'wire': 'dây',
    'electrical': 'điện',
    'prototype': 'mẫu thử (prototype)',
    'proto': 'mẫu thử',
    'tacky': 'dính',
    'backout': 'hoàn tác / rút lại',

    # ── Quality / Process ──
    'metrological': 'đo lường',
    'prior': 'trước / ưu tiên',
    'forced': 'bắt buộc',
    'subfolder': 'thư mục con',
    'rest': 'còn lại / nghỉ',
    'levels': 'cấp độ',
    'versioned': 'có phiên bản',
    'equivalence': 'tương đương',
    'beyond': 'vượt quá',
    'topology': 'cấu trúc liên kết',
    'addendum': 'phụ lục bổ sung',
    'rebaseline': 'thiết lập lại mốc chuẩn',
    'custom': 'tùy chỉnh',
    'kanban': 'kanban (bảng quản lý)',
    'database': 'cơ sở dữ liệu',
    'practical': 'thực tế',
    'trends': 'xu hướng',
    'provided': 'được cung cấp',
    'many': 'nhiều',
    'plus': 'cộng thêm',
    'membership': 'tư cách thành viên',
    'obligations': 'nghĩa vụ',
    'attempt': 'nỗ lực / thử',
    'obsoleting': 'đưa vào lỗi thời',
    'that': 'rằng / đó',
    'attributable': 'có thể quy cho',
    'realism': 'thực tế',
    'subset': 'tập con',

    # ── Lean ──
    'Overproduction': 'Sản xuất thừa',
    'Defects': 'Khuyết tật',
    'defects': 'khuyết tật',
    'Motion': 'Thao tác thừa',
    'Milieu': 'Môi trường',
    'Facilitate': 'Hỗ trợ / Tạo điều kiện',
    'Facilitation': 'Hỗ trợ / Tạo điều kiện',
    'Enduring': 'Bền vững',
    'Proactive': 'Chủ động',
    'play': 'đóng vai / thực hành',

    # ── Roles / Organization ──
    'Tech': 'Kỹ thuật',
    'Pick': 'Chọn / Lấy hàng',
    'pick': 'chọn / lấy hàng',
    'Guides': 'Hướng dẫn',
    'Other': 'Khác',
    'Printed': 'Đã in',
    'Throttling': 'Điều tiết',
    'Unmanaged': 'Không được quản lý',
    'unmanaged': 'không được quản lý',
    'Ballooned': 'Đã đánh số balloon',
    'ballooned': 'đã đánh số balloon',
    'Scoreboard': 'Bảng điểm',
    'cheatsheet': 'bảng tóm tắt nhanh',
    'Editorial': 'Biên tập',
    'Consistency': 'Tính nhất quán',
    'consistency': 'tính nhất quán',
    'Flows': 'Luồng',
    'attainment': 'đạt được',
    'signage': 'biển báo',
    'escort': 'hộ tống / đưa đi',
    'surveillance': 'giám sát',
    'affecting': 'ảnh hưởng',
    'observations': 'quan sát',
    'Facilities': 'Cơ sở vật chất',
    'recurring': 'định kỳ / lặp lại',
    'standardize': 'chuẩn hóa',
    'questions': 'câu hỏi',
    'apps': 'ứng dụng',
    'shortcut': 'phím tắt / lối tắt',
    'analytics': 'phân tích',
    'physical': 'vật lý',
    'verbal': 'bằng lời',
    'received': 'đã nhận',
    'corrected': 'đã sửa',
    'trackers': 'công cụ theo dõi',
    'bridge': 'cầu nối',
    'Workforce': 'Lực lượng lao động',
    'vacancy': 'vị trí tuyển dụng',
    'interview': 'phỏng vấn',
    'partner': 'đối tác',
    'periodic': 'định kỳ',
    'compromise': 'thỏa hiệp',
    'remake': 'làm lại',
    'Incoterm': 'Incoterm',
    'owned': 'sở hữu',
    'unclear': 'không rõ ràng',
    'exclusion': 'loại trừ',
    'courier': 'chuyển phát',
    'ethics': 'đạo đức',
    'expedites': 'đẩy nhanh',
    'technician': 'kỹ thuật viên',
    'suspicion': 'nghi ngờ',
    'damaged': 'bị hư hỏng',
    'productivity': 'năng suất',
    'Replica': 'Bản sao',
    'guidance': 'hướng dẫn',
    'checkpoints': 'điểm kiểm tra',
    'workload': 'khối lượng công việc',
    'legacy': 'kế thừa / cũ',
    'concerns': 'lo ngại',
    'segregated': 'đã cách ly',
    'burden': 'gánh nặng / chi phí',
    'advance': 'tạm ứng / nâng cao',
    'calculation': 'tính toán',
    'queries': 'truy vấn',
    'hypercare': 'hỗ trợ tăng cường',
    'tickets': 'phiếu yêu cầu',
    'minute': 'phút / biên bản',
    'justification': 'lý giải / biện minh',
    'golden': 'vàng / chuẩn',
    'continue': 'tiếp tục',
    'clarify': 'làm rõ',
    'opened': 'đã mở',
    'revised': 'đã sửa đổi',
    'detailed': 'chi tiết',
    'clarity': 'sự rõ ràng',
    'buffer': 'bộ đệm / vùng đệm',
    'preconditions': 'điều kiện tiên quyết',
    'measured': 'đã đo',
    'doubt': 'nghi ngờ',
    'formalize': 'chính thức hóa',
    'Supersedure': 'Thay thế',
    'supersedure': 'thay thế',
    'fixed': 'cố định / đã sửa',
    'house': 'nhà / nội bộ',
    'migration': 'di chuyển / chuyển đổi',
    'continuation': 'tiếp tục',
    'recount': 'đếm lại',
    'resources': 'nguồn lực',
    'artefacts': 'sản phẩm / tạo tác',
    'fitness': 'sự phù hợp',
    'Borrowed': 'Mượn',
    'routine': 'thường xuyên / quy trình',
    'long': 'dài',
    'unless': 'trừ khi',
    'executive': 'điều hành',
    'attachments': 'tài liệu đính kèm',
    'wrap': 'bọc / quấn',
    'language': 'ngôn ngữ',
    'cartons': 'thùng carton',
    'protocol': 'quy trình / giao thức',
    'personhip': 'tư cách cá nhân',

    # ── More common words ──
    'across': 'xuyên suốt',
    'along': 'dọc theo',
    'among': 'trong số',
    'another': 'khác',
    'around': 'xung quanh',
    'because': 'bởi vì',
    'being': 'đang',
    'both': 'cả hai',
    'during': 'trong suốt',
    'either': 'hoặc',
    'enough': 'đủ',
    'every': 'mỗi',
    'fully': 'đầy đủ',
    'further': 'thêm / xa hơn',
    'given': 'cho trước / được cung cấp',
    'however': 'tuy nhiên',
    'instead': 'thay vì',
    'itself': 'chính nó',
    'likely': 'có khả năng',
    'must': 'phải',
    'once': 'một khi',
    'other': 'khác',
    'rather': 'đúng hơn',
    'shall': 'phải',
    'should': 'nên',
    'since': 'kể từ',
    'still': 'vẫn',
    'such': 'như vậy',
    'their': 'của họ',
    'them': 'họ',
    'these': 'những',
    'those': 'những',
    'upon': 'dựa trên',
    'whether': 'liệu',
    'while': 'trong khi',
    'whose': 'của ai',
    'would': 'sẽ',
}

def main():
    wb = openpyxl.load_workbook('tools/remaining-english-words.xlsx')
    ws = wb.active
    yes_fill = PatternFill('solid', fgColor='C6EFCE')
    yellow_fill = PatternFill('solid', fgColor='FFFF00')

    filled = 0
    still_yellow = 0

    for row_idx in range(2, ws.max_row + 1):
        word = ws.cell(row=row_idx, column=2).value
        vi = ws.cell(row=row_idx, column=4).value
        note = ws.cell(row=row_idx, column=6).value

        if not word:
            continue

        word_str = str(word).strip()

        # Only process yellow-marked rows
        if note and 'Cần anh' in str(note):
            translation = BATCH2.get(word_str, BATCH2.get(word_str.lower(), ''))
            if translation:
                ws.cell(row=row_idx, column=4).value = translation
                ws.cell(row=row_idx, column=4).fill = yes_fill
                ws.cell(row=row_idx, column=5).fill = yes_fill
                ws.cell(row=row_idx, column=6).value = ''
                filled += 1
            else:
                still_yellow += 1

    wb.save('tools/remaining-english-words.xlsx')
    print(f'Batch 2 done!')
    print(f'  Filled: {filled}')
    print(f'  Still yellow: {still_yellow}')

if __name__ == '__main__':
    main()
