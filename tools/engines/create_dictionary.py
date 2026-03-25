import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=11)
cell_align = Alignment(vertical='center', wrap_text=True)
yes_fill = PatternFill('solid', fgColor='C6EFCE')
no_fill = PatternFill('solid', fgColor='FFC7CE')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ['STT', 'Thuật ngữ tiếng Anh', 'Bản dịch tiếng Việt', 'Dịch (Yes/No)', 'Ghi chú']

def setup_sheet(ws, data):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for i, (en, vi, translate, note) in enumerate(data, 2):
        ws.cell(row=i, column=1, value=i-1).font = cell_font
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=i, column=1).border = thin_border
        ws.cell(row=i, column=2, value=en).font = cell_font
        ws.cell(row=i, column=2).alignment = cell_align
        ws.cell(row=i, column=2).border = thin_border
        ws.cell(row=i, column=3, value=vi).font = cell_font
        ws.cell(row=i, column=3).alignment = cell_align
        ws.cell(row=i, column=3).border = thin_border
        c = ws.cell(row=i, column=4, value=translate)
        c.font = cell_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = yes_fill if translate == 'Yes' else no_fill
        c.border = thin_border
        ws.cell(row=i, column=5, value=note).font = cell_font
        ws.cell(row=i, column=5).alignment = cell_align
        ws.cell(row=i, column=5).border = thin_border
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 48
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:E{ws.max_row}'

# ===== Sheet 1: Vai trò & Chức danh =====
ws1 = wb.active
ws1.title = 'Vai tro & Chuc danh'
roles = [
    ('General Director', 'Tổng Giám đốc', 'Yes', 'Chức danh cao nhất'),
    ('Deputy General Director', 'Phó Tổng Giám đốc', 'Yes', ''),
    ('Quality Manager', 'Trưởng phòng Chất lượng', 'Yes', 'Quản lý hệ thống QMS'),
    ('Production Manager', 'Trưởng phòng Sản xuất', 'Yes', ''),
    ('Engineering Manager', 'Trưởng phòng Kỹ thuật', 'Yes', ''),
    ('Purchasing Manager', 'Trưởng phòng Mua hàng', 'Yes', ''),
    ('Warehouse Manager', 'Trưởng phòng Kho', 'Yes', ''),
    ('HR Manager', 'Trưởng phòng Nhân sự', 'Yes', ''),
    ('Sales Manager', 'Trưởng phòng Kinh doanh', 'Yes', ''),
    ('Finance Manager', 'Trưởng phòng Tài chính', 'Yes', ''),
    ('Maintenance Manager', 'Trưởng phòng Bảo trì', 'Yes', ''),
    ('EHS Manager', 'Trưởng phòng An toàn - Môi trường', 'Yes', 'EHS giữ nguyên viết tắt'),
    ('IT Administrator', 'Quản trị viên CNTT', 'Yes', 'IT giữ nguyên viết tắt'),
    ('QMS Engineer', 'Kỹ sư QMS', 'Yes', 'QMS giữ nguyên viết tắt'),
    ('QA Engineer', 'Kỹ sư QA', 'Yes', 'QA giữ nguyên viết tắt'),
    ('QC Engineer', 'Kỹ sư QC', 'Yes', 'QC giữ nguyên viết tắt'),
    ('Process Engineer', 'Kỹ sư quy trình', 'Yes', ''),
    ('Document Responsible Person', 'Người phụ trách tài liệu', 'Yes', 'Vai trò trong hệ thống kiểm soát tài liệu'),
    ('Lead Department', 'Bộ phận chủ trì', 'Yes', 'Phòng ban chịu trách nhiệm chính'),
    ('Responsible Person', 'Người phụ trách', 'Yes', 'Đầu mối phụ trách công việc'),
    ('Foreman', 'Quản đốc', 'Yes', 'Quản lý phân xưởng sản xuất'),
    ('Team Leader', 'Tổ trưởng', 'Yes', 'Trưởng tổ/nhóm sản xuất'),
    ('Shift Leader', 'Trưởng ca', 'Yes', 'Phụ trách ca sản xuất'),
    ('Cell Leader', 'Tổ trưởng chuyền', 'Yes', 'Phụ trách 1 chuyền/cell sản xuất'),
    ('Worker', 'Công nhân', 'Yes', ''),
    ('Operator', 'Người vận hành', 'Yes', 'Vận hành máy CNC, thiết bị'),
    ('Specialist', 'Chuyên viên', 'Yes', ''),
    ('Inspector', 'Người kiểm tra', 'Yes', 'Kiểm tra viên chất lượng'),
    ('Performer', 'Người thực hiện', 'Yes', 'Vai trò thực thi công việc'),
    ('Author', 'Người soạn', 'Yes', 'Người soạn thảo tài liệu'),
    ('Reviewer', 'Người rà soát', 'Yes', 'Vai trò xem xét/rà soát'),
    ('Approver', 'Người phê duyệt', 'Yes', 'Vai trò phê duyệt tài liệu/quyết định'),
    ('Cross-reviewer', 'Người rà soát chéo', 'Yes', 'Rà soát bởi phòng ban khác'),
    ('End user', 'Người dùng cuối', 'Yes', 'Người sử dụng tài liệu/hệ thống cuối cùng'),
    ('Warehouse Team Leader', 'Tổ trưởng Kho', 'Yes', ''),
    ('QC Team Leader', 'Tổ trưởng QC', 'Yes', 'QC giữ nguyên viết tắt'),
]
setup_sheet(ws1, roles)

# ===== Sheet 2: Thuật ngữ QMS =====
ws2 = wb.create_sheet('Thuat ngu QMS')
qms_terms = [
    ('document', 'tài liệu', 'Yes', 'Tài liệu quy trình, hướng dẫn'),
    ('record', 'hồ sơ', 'Yes', 'Bằng chứng hoạt động đã thực hiện'),
    ('requirement', 'yêu cầu', 'Yes', ''),
    ('approval', 'phê duyệt', 'Yes', ''),
    ('release', 'phát hành', 'Yes', 'Ban hành tài liệu'),
    ('review', 'rà soát', 'Yes', 'Xem xét, đánh giá'),
    ('inspection', 'kiểm tra', 'Yes', ''),
    ('evidence', 'bằng chứng', 'Yes', ''),
    ('compliance', 'tuân thủ', 'Yes', 'Sự phù hợp với yêu cầu'),
    ('deviation', 'sai lệch', 'Yes', 'Độ lệch so với tiêu chuẩn'),
    ('traceability', 'truy xuất nguồn gốc', 'Yes', ''),
    ('calibration', 'hiệu chuẩn', 'Yes', 'Hiệu chuẩn thiết bị đo'),
    ('competence', 'năng lực', 'Yes', ''),
    ('training', 'đào tạo', 'Yes', ''),
    ('process', 'quy trình', 'Yes', ''),
    ('revision', 'phiên bản', 'Yes', 'Lần sửa đổi tài liệu'),
    ('input', 'đầu vào', 'Yes', ''),
    ('output', 'đầu ra', 'Yes', ''),
    ('operation', 'vận hành', 'Yes', ''),
    ('control gate', 'cổng kiểm soát', 'Yes', 'Điểm kiểm soát trong quy trình'),
    ('mandatory hold point', 'điểm dừng bắt buộc', 'Yes', 'Phải dừng chờ phê duyệt'),
    ('hold point', 'điểm chặn', 'Yes', 'Điểm kiểm soát chờ xác nhận'),
    ('point-of-use', 'điểm sử dụng', 'Yes', 'Nơi sử dụng tài liệu'),
    ('cross-review', 'rà soát chéo', 'Yes', 'Rà soát bởi phòng ban khác'),
    ('master copy', 'bản gốc', 'Yes', 'Bản tài liệu gốc duy nhất'),
    ('release copy', 'bản phát hành', 'Yes', 'Bản được phân phối sử dụng'),
    ('controlled copy', 'bản kiểm soát', 'Yes', 'Bản được kiểm soát phân phối'),
    ('internal audit', 'đánh giá nội bộ', 'Yes', ''),
    ('management review', 'xem xét của lãnh đạo', 'Yes', ''),
    ('change control', 'kiểm soát thay đổi', 'Yes', ''),
    ('continual improvement', 'cải tiến liên tục', 'Yes', ''),
    ('corrective action', 'hành động khắc phục', 'Yes', ''),
    ('preventive action', 'hành động phòng ngừa', 'Yes', ''),
    ('root cause', 'nguyên nhân gốc', 'Yes', 'Nguyên nhân cốt lõi'),
    ('nonconformity', 'sự không phù hợp', 'Yes', ''),
    ('risk assessment', 'đánh giá rủi ro', 'Yes', ''),
    ('effectiveness', 'hiệu lực', 'Yes', 'Mức độ đạt kết quả mong muốn'),
    ('verification', 'xác minh', 'Yes', 'Xác nhận đáp ứng yêu cầu quy định'),
    ('validation', 'xác nhận giá trị sử dụng', 'Yes', 'Xác nhận đáp ứng yêu cầu sử dụng thực tế'),
    ('obsolete', 'lỗi thời', 'Yes', 'Tài liệu không còn hiệu lực'),
    ('superseded', 'thay thế', 'Yes', 'Đã bị phiên bản mới thay thế'),
    ('retention', 'lưu giữ', 'Yes', 'Thời hạn lưu giữ hồ sơ'),
    ('disposition', 'xử lý', 'Yes', 'Quyết định xử lý (phế phẩm, làm lại...)'),
    ('containment', 'ngăn chặn', 'Yes', 'Hành động ngăn chặn tạm thời'),
    ('escalation', 'báo cáo vượt cấp', 'Yes', 'Chuyển lên cấp cao hơn giải quyết'),
    ('authority', 'thẩm quyền', 'Yes', 'Quyền hạn quyết định'),
    ('accountability', 'trách nhiệm giải trình', 'Yes', ''),
    ('responsibility', 'trách nhiệm', 'Yes', ''),
    ('Per issuance decision', 'Theo quyết định ban hành', 'Yes', 'Cụm từ metadata'),
]
setup_sheet(ws2, qms_terms)

# ===== Sheet 3: Đối tượng & Vận hành =====
ws3 = wb.create_sheet('Doi tuong & Van hanh')
objects = [
    ('form', 'biểu mẫu', 'Yes', 'Mẫu ghi chép, biểu mẫu'),
    ('checklist', 'bảng kiểm', 'Yes', 'Danh sách kiểm tra'),
    ('log', 'nhật ký', 'Yes', 'Sổ ghi chép'),
    ('register', 'sổ theo dõi', 'Yes', 'Sổ đăng ký/theo dõi'),
    ('tracking register', 'bảng theo dõi', 'Yes', ''),
    ('job dossier', 'hồ sơ công việc', 'Yes', 'Bộ hồ sơ sản xuất theo đơn hàng'),
    ('production', 'sản xuất', 'Yes', ''),
    ('quality', 'chất lượng', 'Yes', ''),
    ('customer', 'khách hàng', 'Yes', ''),
    ('supplier', 'nhà cung cấp', 'Yes', ''),
    ('complaint', 'khiếu nại', 'Yes', 'Phản hồi/khiếu nại từ khách hàng'),
    ('incident', 'sự cố', 'Yes', ''),
    ('equipment', 'thiết bị', 'Yes', ''),
    ('maintenance', 'bảo trì', 'Yes', ''),
    ('tooling', 'dao cụ', 'Yes', 'Dụng cụ cắt gọt CNC'),
    ('material', 'nguyên vật liệu', 'Yes', ''),
    ('scrap', 'phế phẩm', 'Yes', 'Sản phẩm loại bỏ'),
    ('rework', 'làm lại', 'Yes', 'Gia công lại sản phẩm lỗi'),
    ('hold', 'tạm giữ', 'Yes', 'Trạng thái tạm giữ hàng chờ xử lý'),
    ('lot', 'lô hàng', 'Yes', ''),
    ('batch', 'lô hàng', 'Yes', ''),
    ('warehouse', 'kho', 'Yes', ''),
    ('delivery', 'giao hàng', 'Yes', ''),
    ('packaging', 'đóng gói', 'Yes', ''),
    ('measurement', 'đo lường', 'Yes', ''),
    ('workshop', 'xưởng', 'Yes', 'Phân xưởng sản xuất'),
    ('production line', 'dây chuyền sản xuất', 'Yes', ''),
    ('readiness level', 'mức sẵn sàng', 'Yes', 'Mức độ sẵn sàng vận hành'),
    ('recall', 'thu hồi', 'Yes', 'Thu hồi tài liệu/sản phẩm'),
    ('scope', 'phạm vi', 'Yes', ''),
    ('purpose', 'mục đích', 'Yes', ''),
    ('procedure', 'thủ tục', 'Yes', 'Quy trình thực hiện'),
    ('standard', 'tiêu chuẩn', 'Yes', ''),
    ('specification', 'quy cách kỹ thuật', 'Yes', 'Thông số kỹ thuật sản phẩm'),
    ('drawing', 'bản vẽ', 'Yes', ''),
    ('work instruction', 'hướng dẫn công việc', 'Yes', 'WI giữ nguyên viết tắt'),
    ('acceptance', 'nghiệm thu', 'Yes', 'Chấp nhận sản phẩm/kết quả'),
    ('rejection', 'từ chối', 'Yes', 'Loại bỏ sản phẩm không đạt'),
    ('concession', 'nhượng bộ', 'Yes', 'Chấp nhận sản phẩm lệch tiêu chuẩn'),
    ('legal hold', 'giữ pháp lý', 'Yes', 'Giữ tài liệu theo yêu cầu pháp lý'),
    ('emergency release', 'phát hành khẩn cấp', 'Yes', ''),
]
setup_sheet(ws3, objects)

# ===== Sheet 4: Giữ nguyên tiếng Anh =====
ws4 = wb.create_sheet('Giu nguyen tieng Anh')
keep_english = [
    ('QMS', 'QMS', 'No', 'Viết tắt quốc tế - Quality Management System'),
    ('QA', 'QA', 'No', 'Viết tắt - Quality Assurance'),
    ('QC', 'QC', 'No', 'Viết tắt - Quality Control'),
    ('NCR', 'NCR', 'No', 'Viết tắt - Non-Conformance Report'),
    ('CAPA', 'CAPA', 'No', 'Viết tắt - Corrective & Preventive Action'),
    ('DCR', 'DCR', 'No', 'Viết tắt - Document Change Request'),
    ('SOP', 'SOP', 'No', 'Viết tắt - Standard Operating Procedure'),
    ('WI', 'WI', 'No', 'Viết tắt - Work Instruction'),
    ('FRM', 'FRM', 'No', 'Viết tắt - Form'),
    ('ANNEX', 'ANNEX', 'No', 'Mã tài liệu phụ lục'),
    ('ISO', 'ISO', 'No', 'Tổ chức tiêu chuẩn quốc tế'),
    ('AS9100D', 'AS9100D', 'No', 'Tiêu chuẩn hàng không vũ trụ'),
    ('KPI', 'KPI', 'No', 'Viết tắt - Key Performance Indicator'),
    ('RACI', 'RACI', 'No', 'Ma trận phân công trách nhiệm'),
    ('SSOT', 'SSOT', 'No', 'Viết tắt - Single Source of Truth'),
    ('SoR', 'SoR', 'No', 'Viết tắt - System of Record'),
    ('FAI', 'FAI', 'No', 'Viết tắt - First Article Inspection'),
    ('FMEA', 'FMEA', 'No', 'Viết tắt - Failure Mode & Effects Analysis'),
    ('PFMEA', 'PFMEA', 'No', 'Viết tắt - Process FMEA'),
    ('SPC', 'SPC', 'No', 'Viết tắt - Statistical Process Control'),
    ('MSA', 'MSA', 'No', 'Viết tắt - Measurement System Analysis'),
    ('IPQC', 'IPQC', 'No', 'Viết tắt - In-Process Quality Control'),
    ('OTD', 'OTD', 'No', 'Viết tắt - On-Time Delivery'),
    ('FPY', 'FPY', 'No', 'Viết tắt - First Pass Yield'),
    ('COPQ', 'COPQ', 'No', 'Viết tắt - Cost of Poor Quality'),
    ('CTQ', 'CTQ', 'No', 'Viết tắt - Critical to Quality'),
    ('RFQ', 'RFQ', 'No', 'Viết tắt - Request for Quotation'),
    ('PO', 'PO', 'No', 'Viết tắt - Purchase Order'),
    ('BOM', 'BOM', 'No', 'Viết tắt - Bill of Materials'),
    ('CoC', 'CoC', 'No', 'Viết tắt - Certificate of Conformance'),
    ('CoA', 'CoA', 'No', 'Viết tắt - Certificate of Analysis'),
    ('POD', 'POD', 'No', 'Viết tắt - Proof of Delivery'),
    ('CNC', 'CNC', 'No', 'Viết tắt - Computer Numerical Control'),
    ('ERP', 'ERP', 'No', 'Viết tắt - Enterprise Resource Planning'),
    ('IT', 'IT', 'No', 'Viết tắt - Information Technology'),
    ('HR', 'HR', 'No', 'Viết tắt - Human Resources'),
    ('EHS', 'EHS', 'No', 'Viết tắt - Environment, Health & Safety'),
    ('Setup', 'Setup', 'No', 'Thuật ngữ ngành - cài đặt/chuẩn bị máy'),
    ('Traveler', 'Traveler', 'No', 'Thuật ngữ ngành - phiếu theo dõi công đoạn'),
    ('Balloon', 'Balloon', 'No', 'Thuật ngữ ngành - đánh số trên bản vẽ kỹ thuật'),
    ('Epicor', 'Epicor', 'No', 'Tên phần mềm ERP'),
    ('SharePoint', 'SharePoint', 'No', 'Tên nền tảng Microsoft'),
    ('M365', 'M365', 'No', 'Viết tắt - Microsoft 365'),
    ('Power Automate', 'Power Automate', 'No', 'Tên sản phẩm Microsoft'),
]
setup_sheet(ws4, keep_english)

# ===== Sheet 5: Phòng ban =====
ws5 = wb.create_sheet('Phong ban')
departments = [
    ('Quality Department', 'Phòng Chất lượng', 'Yes', ''),
    ('Production Department', 'Phòng Sản xuất', 'Yes', ''),
    ('Engineering Department', 'Phòng Kỹ thuật', 'Yes', ''),
    ('Purchasing Department', 'Phòng Mua hàng', 'Yes', ''),
    ('Warehouse Department', 'Phòng Kho', 'Yes', ''),
    ('Sales Department', 'Phòng Kinh doanh', 'Yes', ''),
    ('Finance Department', 'Phòng Tài chính', 'Yes', ''),
    ('HR Department', 'Phòng Nhân sự', 'Yes', 'HR giữ nguyên viết tắt'),
    ('Maintenance Department', 'Phòng Bảo trì', 'Yes', ''),
    ('EHS Department', 'Phòng An toàn - Môi trường', 'Yes', 'EHS giữ nguyên viết tắt'),
    ('IT Department', 'Phòng CNTT', 'Yes', 'IT giữ nguyên viết tắt'),
    ('Planning Department', 'Phòng Kế hoạch', 'Yes', ''),
    ('All departments', 'Tất cả các phòng ban', 'Yes', ''),
]
setup_sheet(ws5, departments)

# ===== Sheet 6: Metadata labels =====
ws6 = wb.create_sheet('Metadata labels')
metadata = [
    ('Code:', 'Mã:', 'Yes', 'Label metadata tài liệu'),
    ('Version:', 'Phiên bản:', 'Yes', ''),
    ('Effective Date:', 'Ngày hiệu lực:', 'Yes', ''),
    ('Owner:', 'Chủ sở hữu:', 'Yes', 'Phòng ban sở hữu tài liệu'),
    ('Approved by:', 'Phê duyệt bởi:', 'Yes', ''),
    ('Classification:', 'Phân loại:', 'Yes', ''),
    ('Retention:', 'Lưu giữ:', 'Yes', ''),
    ('Scope:', 'Phạm vi:', 'Yes', ''),
    ('Purpose:', 'Mục đích:', 'Yes', ''),
    ('References:', 'Tài liệu tham chiếu:', 'Yes', ''),
    ('Definitions:', 'Định nghĩa:', 'Yes', ''),
    ('Responsibilities:', 'Trách nhiệm:', 'Yes', ''),
    ('Procedure:', 'Quy trình:', 'Yes', ''),
    ('Related documents:', 'Tài liệu liên quan:', 'Yes', ''),
    ('Revision history:', 'Lịch sử sửa đổi:', 'Yes', ''),
]
setup_sheet(ws6, metadata)

out = 'tools/qms-terminology-dictionary.xlsx'
wb.save(out)
total = len(roles) + len(qms_terms) + len(objects) + len(keep_english) + len(departments) + len(metadata)
print(f'Created: {out}')
print(f'Sheet 1 - Vai tro & Chuc danh: {len(roles)} entries')
print(f'Sheet 2 - Thuat ngu QMS: {len(qms_terms)} entries')
print(f'Sheet 3 - Doi tuong & Van hanh: {len(objects)} entries')
print(f'Sheet 4 - Giu nguyen tieng Anh: {len(keep_english)} entries')
print(f'Sheet 5 - Phong ban: {len(departments)} entries')
print(f'Sheet 6 - Metadata labels: {len(metadata)} entries')
print(f'TOTAL: {total} entries')
