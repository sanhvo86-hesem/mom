from __future__ import annotations

import html
import json
import os
import re
from pathlib import Path

import openpyxl

from department_handbook_data import DEPARTMENT_TERMS, HANDBOOKS


ROOT = Path(__file__).resolve().parents[3]
HANDBOOK_DIR = ROOT / "02-Tai-Lieu-He-Thong/03-Organization/02-Department-Handbooks"
WORKBOOK_PATH = ROOT / "tools/data/qms-terminology-dictionary.xlsx"
REGISTRY_PATH = ROOT / "tools/data/role-registry-job-order-cnc.json"


def load_registry() -> dict:
    return json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))


def rel_path(source_dir: Path, target: str) -> str:
    return os.path.relpath(ROOT / target, source_dir).replace("\\", "/")


def escape(text: str) -> str:
    return html.escape(text, quote=True)


def normalize_tokens(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return list(value)
    return [value]


def code_chip(source_dir: Path, registry: dict, code: str) -> str:
    roles = registry["roles"]
    departments = registry["departments"]
    if code in roles:
        role = roles[code]
        title = f"{role['title_en']} ({role['title_vi']})"
        href = rel_path(source_dir, role["jd_path"])
        return (
            f'<a class="entity-link role-link" href="{escape(href)}" title="{escape(title)}">'
            f'<span class="entity-code role-code">{escape(code)}</span></a>'
        )
    if code in departments:
        dept = departments[code]
        title = f"{dept['title_en']} ({dept['title_vi']})"
        href = rel_path(source_dir, dept["handbook_path"])
        return (
            f'<a class="entity-link dept-link" href="{escape(href)}" title="{escape(title)}">'
            f'<span class="entity-code dept-code">{escape(code)}</span></a>'
        )
    return f'<span class="entity-code">{escape(code)}</span>'


def chips(source_dir: Path, registry: dict, values, separator: str = " / ") -> str:
    items = normalize_tokens(values)
    rendered = [code_chip(source_dir, registry, item) for item in items]
    return f'<span class="entity-cluster role-cluster">{separator.join(rendered)}</span>'


def linked_tag(source_dir: Path, doc: dict) -> str:
    href = rel_path(source_dir, doc["path"])
    return f'<span class="inline-tag"><a href="{escape(href)}">{escape(doc["label"])}</a></span>'


def render_table(headers: list[str], rows: list[list[str]]) -> str:
    head_html = "".join(f"<th>{escape(header)}</th>" for header in headers)
    body_rows = []
    for row in rows:
        body_rows.append("<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>")
    return (
        '<div class="table-card"><table class="table">'
        f"<thead><tr>{head_html}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody></table></div>"
    )


def render_list(items: list[str]) -> str:
    return '<ul class="tight">' + "".join(f"<li>{item}</li>" for item in items) + "</ul>"


def render_metric_grid(cards: list[dict]) -> str:
    blocks = []
    for card in cards:
        blocks.append(
            '<div class="metric-card">'
            f'<div class="value">{escape(card["value"])}</div>'
            f'<div class="label">{escape(card["label"])}</div>'
            "</div>"
        )
    return f'<div class="metric-grid">{"".join(blocks)}</div>'


def render_auth_grid(cards: list[dict]) -> str:
    blocks = []
    for card in cards:
        blocks.append(
            '<div class="auth-item">'
            f'<b>{escape(card["title"])}</b>'
            f'<p>{escape(card["body"])}</p>'
            "</div>"
        )
    return f'<div class="auth-grid">{"".join(blocks)}</div>'


def render_process_strip(items: list[str]) -> str:
    return '<div class="process-strip">' + "".join(
        f'<span class="pill">{escape(item)}</span>' for item in items
    ) + "</div>"


def render_backup_cards(cards: list[dict]) -> str:
    blocks = []
    for card in cards:
        blocks.append(
            '<div class="backup-card">'
            f'<b>{escape(card["title"])}</b>'
            f'<p>{escape(card["body"])}</p>'
            "</div>"
        )
    return "".join(blocks)


def render_actor_cell(source_dir: Path, registry: dict, value) -> str:
    items = normalize_tokens(value)
    if items and all(
        isinstance(item, str) and (item in registry["roles"] or item in registry["departments"])
        for item in items
    ):
        return chips(source_dir, registry, items)
    return escape(str(value))


def render_docs_group(source_dir: Path, groups: list[dict]) -> str:
    rows = []
    for group in groups:
        docs_html = "<br>".join(
            f'<a href="{escape(rel_path(source_dir, doc["path"]))}">{escape(doc["label"])}</a>'
            for doc in group["docs"]
        )
        rows.append([escape(group["group"]), docs_html])
    return render_table(["Nhóm", "Tài liệu"], rows)


def render_output_rows(source_dir: Path, registry: dict, rows: list[dict]) -> str:
    table_rows = []
    for row in rows:
        table_rows.append(
            [
                escape(row["name"]),
                escape(row["description"]),
                chips(source_dir, registry, row["owner"]),
                chips(source_dir, registry, row["decision"]),
                escape(row["system"]),
            ]
        )
    return render_table(
        ["Đầu ra / hồ sơ", "Ý nghĩa vận hành", "Owner cấp chức năng", "Vai trò quyết định", "Hệ thống / nơi lưu"],
        table_rows,
    )


def render_kpis(source_dir: Path, registry: dict, rows: list[dict]) -> str:
    table_rows = []
    for row in rows:
        table_rows.append(
            [
                escape(row["name"]),
                chips(source_dir, registry, row["owner"]),
                escape(row["target"]),
                escape(row["source"]),
                escape(row["reaction"]),
            ]
        )
    return render_table(
        ["KPI", "Owner", "Ngưỡng / mục tiêu", "Nguồn dữ liệu", "Khi lệch ngưỡng phải làm gì"],
        table_rows,
    )


def render_interfaces(source_dir: Path, registry: dict, rows: list[dict]) -> str:
    table_rows = []
    for row in rows:
        table_rows.append(
            [
                chips(source_dir, registry, row["with"]),
                escape(row["receive"]),
                escape(row["handoff"]),
                chips(source_dir, registry, row["func_owner"]),
                chips(source_dir, registry, row["decision"]),
            ]
        )
    return render_table(
        ["Interface với", "Nhận từ họ", "Bàn giao cho họ", "Owner cấp chức năng", "Vai trò quyết định / escalations"],
        table_rows,
    )


def render_data_table(rows: list[dict]) -> str:
    table_rows = []
    for row in rows:
        table_rows.append(
            [
                escape(row["data"]),
                escape(row["source"]),
                escape(row["frequency"]),
                escape(row["decision"]),
            ]
        )
    return render_table(
        ["Dữ liệu tối thiểu", "Nguồn chuẩn", "Chu kỳ cập nhật", "Dùng để ra quyết định gì"],
        table_rows,
    )


def render_competence_rows(source_dir: Path, registry: dict, rows: list[dict]) -> str:
    table_rows = []
    for row in rows:
        table_rows.append(
            [
                chips(source_dir, registry, row["role"]),
                escape(row["skill"]),
                escape(row["evidence"]),
                escape(row["requalify"]),
            ]
        )
    return render_table(
        ["Nhóm vai trò", "Năng lực bắt buộc", "Cách chứng minh", "Khi phải tái xác nhận"],
        table_rows,
    )


def render_risks(source_dir: Path, registry: dict, rows: list[dict]) -> str:
    table_rows = []
    for row in rows:
        table_rows.append(
            [
                escape(row["risk"]),
                escape(row["signal"]),
                escape(row["first_hour"]),
                chips(source_dir, registry, row["escalation"]),
            ]
        )
    return render_table(
        ["Rủi ro / kích hoạt", "Dấu hiệu nhận biết", "Hành động trong 60 phút đầu", "Vai trò leo thang"],
        table_rows,
    )


def render_iso_items(items: list[dict]) -> str:
    blocks = []
    for item in items:
        blocks.append(
            '<div class="req">'
            '<span class="req-tag shall">PHẢI</span>'
            f'<div>{item["text"]} <span class="iso-clause">§{escape(item["clause"])}</span></div>'
            "</div>"
        )
    return "".join(blocks)


def render_preface(source_dir: Path, registry: dict, data: dict) -> str:
    role_tags = "".join(
        f'<span class="inline-tag">{chips(source_dir, registry, role)}</span>'
        for role in data["roles"]
    )
    docs = "".join(linked_tag(source_dir, doc) for doc in data["primary_docs"])
    subfunctions = "".join(
        f'<span class="inline-tag">{chips(source_dir, registry, dept)}</span>'
        for dept in data.get("subfunctions", [])
    )
    scope_line = (
        f"<p><b>Phân hệ thuộc phạm vi</b></p><div>{subfunctions}</div>"
        if subfunctions
        else "<p><b>Phân hệ thuộc phạm vi</b></p><div><span class=\"mini-note\">Không tách phân hệ riêng trong phiên bản hiện tại.</span></div>"
    )
    return (
        '<div class="grid-2">'
        '<div>'
        '<p><b>Vai trò thuộc phạm vi</b></p>'
        f"<div>{role_tags}</div>"
        f"{scope_line}"
        "</div>"
        '<div>'
        '<p><b>Tài liệu điều hành chính</b></p>'
        f"<div>{docs}</div>"
        "</div>"
        "</div>"
    )


def page_wrapper(source_dir: Path, registry: dict, data: dict) -> str:
    owner_html = chips(source_dir, registry, data["code"])
    approver_html = chips(source_dir, registry, data["approver"])
    preface_html = render_preface(source_dir, registry, data)
    related_docs_html = render_docs_group(source_dir, data["related_docs"])
    coverage_gap_html = ""
    if data.get("coverage_gap"):
        coverage_gap_html = (
            '<div class="note-soft"><b>Coverage gap đang được khóa tạm</b><br>'
            + "<br>".join(escape(item) for item in data["coverage_gap"])
            + "</div>"
        )
    return f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{escape(data["code"])} — {escape(data["title"])} | HESEM QMS</title>
<link rel="stylesheet" href="{escape(rel_path(source_dir, 'assets/style.css'))}">
<style>
.toc{{border:1px solid var(--ln);border-radius:var(--r);padding:16px;background:var(--bg2);margin:18px 0 24px;}}
.toc-title{{font-size:12px;font-weight:700;color:var(--navy);text-transform:uppercase;letter-spacing:.4px;margin-bottom:10px;}}
.toc-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:8px;}}
.toc-grid a{{display:block;padding:8px 10px;border:1px solid var(--ln);border-radius:6px;background:var(--bg);font-size:12px;color:var(--ink);}}
.metric-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:12px;}}
.metric-card{{border:1px solid var(--ln);border-radius:var(--r);padding:14px;background:var(--bg);}}
.metric-card .value{{font-size:18px;font-weight:700;color:var(--navy);}}
.metric-card .label{{font-size:11px;color:var(--ink3);margin-top:4px;text-transform:uppercase;letter-spacing:.3px;}}
.auth-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px;margin:10px 0 16px;}}
.auth-item{{padding:12px 14px;border:1px solid var(--ln);border-radius:8px;background:var(--bg2);}}
.process-strip{{display:flex;flex-wrap:wrap;gap:8px;margin:12px 0 18px;}}
.pill{{display:inline-block;padding:4px 8px;border-radius:999px;background:var(--blue-l);color:var(--navy);font-size:11px;font-weight:700;margin:0 6px 6px 0;}}
.tight{{margin:0;padding-left:18px;}}
.tight li{{margin:4px 0;}}
.mini-note{{font-size:12px;color:var(--ink2);}}
@media(max-width:960px){{.auth-grid{{grid-template-columns:1fr;}}}}
</style>
</head>
<body>
<div class="container"><div class="page"><div class="page-body">
<div class="form-header">
<div class="fh-left">
<a class="brand-logo" href="{escape(rel_path(source_dir, '01-QMS-Portal/portal.html'))}"><img alt="HESEM Logo" src="{escape(rel_path(source_dir, 'assets/hesem-logo.svg'))}"></a>
<div class="fh-company">
<a href="{escape(rel_path(source_dir, '01-QMS-Portal/portal.html'))}">HESEM ENGINEERING</a>
<span>Tài liệu hệ thống</span>
</div>
</div>
<div class="title">
<strong>{escape(data["title"])}</strong>
<span class="sub-vn">{escape(data["subtitle"])}</span>
</div>
<div class="meta">
<div class="row"><span><b>Code:</b></span><span>{escape(data["code"])}</span></div>
<div class="row"><span><b>Version:</b></span><span>V0</span></div>
<div class="row"><span><b>Effective Date:</b></span><span>Theo quyết định ban hành</span></div>
<div class="row"><span><b>Owner:</b></span><span>{owner_html}</span></div>
<div class="row"><span><b>Approved by:</b></span><span>{approver_html}</span></div>
</div>
</div>
<div class="doc-content" id="docContent"><div class="form-sheet">
<div class="iso-map">
<div class="iso-title">Operating intent / nguyên tắc điều hành</div>
{render_iso_items(data["iso_map"])}
</div>
<div class="preface-block">{preface_html}</div>
<div class="toc">
<div class="toc-title">Mục lục</div>
<div class="toc-grid">
<a href="#d1">1. Mục tiêu phòng ban</a>
<a href="#d2">2. Phạm vi</a>
<a href="#d3">3. Trách nhiệm bắt buộc</a>
<a href="#d4">4. Quyền hạn</a>
<a href="#d5">5. Đầu ra / hồ sơ</a>
<a href="#d6">6. KPI</a>
<a href="#d7">7. Interface liên phòng ban</a>
<a href="#d8">8. Tài liệu liên quan</a>
<a href="#d9">9. Mô hình vận hành &amp; ranh giới vai trò</a>
<a href="#d10">10. Nhịp điều hành, dữ liệu &amp; bằng chứng</a>
<a href="#d11">11. Năng lực &amp; deputy</a>
<a href="#d12">12. Rủi ro &amp; escalations</a>
</div>
</div>
<div class="note-blue"><b>Hệ thống 8 cổng (G0→G7):</b> Sổ tay này mô tả ranh giới chức năng của {escape(data["short_vi"])} trong mô hình job-order CNC. Khi tài liệu khác cần chỉ trách nhiệm cấp phòng ban hoặc phân hệ ổn định, phải dùng mã {escape(data["code"])} hoặc mã phân hệ liên quan; chỉ dùng role code khi tài liệu đi tới phê duyệt, giữ quyền HOLD / RELEASE hoặc trách nhiệm cá nhân.</div>
<h2 class="h2" id="d1">1. Mục tiêu phòng ban</h2>
<p>{escape(data["purpose"])}</p>
{render_metric_grid(data["metric_cards"])}
<h2 class="h2" id="d2">2. Phạm vi</h2>
<p>{escape(data["scope"])}</p>
{render_table(["Nhóm hoạt động", "Bao gồm", "Không thay thế / không vượt quyền"], [[escape(row["group"]), escape(row["include"]), escape(row["exclude"])] for row in data["scope_rows"]])}
<h2 class="h2" id="d3">3. Trách nhiệm bắt buộc</h2>
{render_list(data["responsibilities"])}
<h2 class="h2" id="d4">4. Quyền hạn</h2>
{render_auth_grid(data["authorities"])}
<h2 class="h2" id="d5">5. Đầu ra / hồ sơ</h2>
{render_output_rows(source_dir, registry, data["outputs"])}
<h2 class="h2" id="d6">6. KPI</h2>
{render_kpis(source_dir, registry, data["kpis"])}
<h2 class="h2" id="d7">7. Interface liên phòng ban</h2>
{render_interfaces(source_dir, registry, data["interfaces"])}
<h2 class="h2" id="d8">8. Tài liệu liên quan</h2>
{related_docs_html}
<h2 class="h2" id="d9">9. Mô hình vận hành &amp; ranh giới vai trò</h2>
{render_process_strip(data["operating_model"])}
<p>{escape(data["boundary_intro"])}</p>
{render_table(["Giao điểm khó", "Chủ thể chính", "Ranh giới bắt buộc"], [[escape(row["point"]), render_actor_cell(source_dir, registry, row["owner"]), escape(row["boundary"])] for row in data["boundaries"]])}
{coverage_gap_html}
<h2 class="h2" id="d10">10. Nhịp điều hành, dữ liệu &amp; bằng chứng</h2>
{render_list(data["rhythm_notes"])}
{render_data_table(data["data_table"])}
<h2 class="h2" id="d11">11. Năng lực &amp; deputy</h2>
<p>{escape(data["competence_intro"])}</p>
{render_competence_rows(source_dir, registry, data["competence_rows"])}
{render_backup_cards(data["deputies"])}
<h2 class="h2" id="d12">12. Rủi ro &amp; escalations</h2>
{render_risks(source_dir, registry, data["risks"])}
<div class="footer-note">Sổ tay này được xây để dùng như tài liệu ranh giới chức năng, không phải bản tóm tắt SOP. Khi thay đổi mô hình vận hành, điểm bàn giao liên phòng ban, vai trò quyết định hoặc coverage gap, phải rà lại handbook, từ điển department code và các tài liệu tổ chức liên đới trong cùng đợt.</div>
</div></div></div></div></div>
</body>
</html>
"""


def render_index(registry: dict) -> str:
    cards = []
    source_dir = HANDBOOK_DIR
    for data in HANDBOOKS:
        href = Path(data["path"]).name
        tags = "".join(f'<span class="tag">{escape(tag)}</span>' for tag in data["index_tags"])
        next_docs = ", ".join(
            f'<a href="{escape(rel_path(source_dir, doc["path"]))}">{escape(doc["label"])}</a>'
            for doc in data["index_next_docs"]
        )
        cards.append(
            '<div class="dept-card">'
            f'<h3><a href="{escape(href)}">{escape(data["title"])}</a></h3>'
            f'<div class="mini-note"><b>{escape(data["code"])}</b> — {escape(data["subtitle"])}</div>'
            f"<div style=\"margin:8px 0 0\">{tags}</div>"
            '<ul class="tight">'
            f'<li>{escape(data["index_intro"])}</li>'
            f"<li>Điểm vào tiếp theo: {next_docs}.</li>"
            "</ul></div>"
        )
    return f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Department Handbooks Index | HESEM QMS</title>
<link rel="stylesheet" href="../../../assets/style.css">
<style>
.page-body h1{{font-size:24px;line-height:1.35;margin:0 0 10px;}}
.card-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:14px;margin:14px 0;}}
.dept-card{{border:1px solid var(--ln);border-radius:var(--r);padding:14px;background:#fff;}}
.dept-card h3{{margin:0 0 6px;font-size:16px;color:var(--ink);}}
.tag{{display:inline-block;padding:2px 8px;border-radius:999px;border:1px solid var(--ln);background:#eef7ff;color:var(--navy);font-size:11px;font-weight:700;margin:0 6px 6px 0;}}
ul.tight{{margin:8px 0;padding-left:18px;}}
ul.tight li{{margin:4px 0;}}
@media (max-width:960px){{.card-grid{{grid-template-columns:1fr;}}.page-body{{padding:24px 18px 36px;}}}}
</style>
</head>
<body>
<div class="container"><div class="page"><div class="page-body">
<div class="form-header">
<div class="fh-left">
<a class="brand-logo" href="../../../01-QMS-Portal/portal.html"><img alt="HESEM Logo" src="../../../assets/hesem-logo.svg"></a>
<div class="fh-company"><a href="../../../01-QMS-Portal/portal.html">HESEM ENGINEERING</a><span>Tài liệu hệ thống • Department Handbooks</span></div>
</div>
<div class="title"><strong>Department Handbooks Index</strong><span class="sub-vn">Bản đồ sổ tay phòng ban và phân hệ ổn định cho mô hình job-order CNC</span></div>
</div>
<div class="note-blue"><b>Dùng trang này khi</b><br>Cần xác định tài liệu gốc của một phòng ban hoặc phân hệ ổn định trước khi đọc SOP/WI/FRM. Thứ tự mặc định là <b>Department Handbook → SOP → WI → FRM → ANNEX</b>. Handbook trả lời câu hỏi “đây là mandate của phòng ban nào”, còn JD trả lời câu hỏi “ai là người giữ quyết định cá nhân”.</div>
<h1>Department Handbooks Index</h1>
<div class="keyline"><strong>Quy tắc đọc:</strong> chọn đúng handbook theo chức năng hoặc phân hệ ổn định trước. Nếu nội dung đang nói về giao điểm giữa các phòng ban, đọc handbook của bên tạo đầu ra trước rồi mở tiếp handbook của bên nhận bàn giao để khóa ranh giới.</div>
<div class="card-grid">{"".join(cards)}</div>
<div class="table-card"><table class="table"><thead><tr><th>Điểm vào chung</th><th>Link</th></tr></thead><tbody>
<tr><td>Thư viện JD</td><td><a href="../03-Job-Descriptions/index.html">Job Descriptions</a></td></tr>
<tr><td>Ma trận authority</td><td><a href="../04-RACI-Authority/authority-matrix.html">Authority Matrix</a></td></tr>
<tr><td>Ma trận RACI master</td><td><a href="../04-RACI-Authority/raci-master-matrix.html">RACI Master Matrix</a></td></tr>
<tr><td>Role &amp; gate boundary</td><td><a href="../../../03-Tai-Lieu-Van-Hanh/03-Reference/05-ANNEX-500/annex-503-cnc-operating-model-and-role-boundary.html">ANNEX-503</a></td></tr>
<tr><td>Department boundary standard</td><td><a href="../../../core-standards/20-department-boundary-handbook-codes.md">Core Standard 20</a></td></tr>
</tbody></table></div>
</div></div></div>
</body>
</html>
"""


def write_handbooks(registry: dict) -> None:
    for data in HANDBOOKS:
        path = ROOT / data["path"]
        html_text = page_wrapper(path.parent, registry, data)
        path.write_text(html_text, encoding="utf-8")
    (HANDBOOK_DIR / "index.html").write_text(render_index(registry), encoding="utf-8")


def update_department_dictionary(workbook_path: Path, registry: dict) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    handbook_by_code = {item["code"]: item for item in HANDBOOKS}
    note_by_title = {item["en"]: item["note"] for item in DEPARTMENT_TERMS}

    ws_terms = wb["Phong ban"]
    ws_terms.delete_rows(1, ws_terms.max_row)
    ws_terms.append(["STT", "Thuật ngữ tiếng Anh", "Bản dịch tiếng Việt", "Dịch (Yes/No)", "Ghi chú"])
    for index, row in enumerate(DEPARTMENT_TERMS, start=1):
        ws_terms.append([index, row["en"], row["vi"], "Yes", row["note"]])

    ws_codes = wb["Department code & handbook link"]
    ws_codes.delete_rows(1, ws_codes.max_row)
    ws_codes.append(
        [
            "STT",
            "Department code",
            "Department title English",
            "Tên tiếng Việt chuẩn",
            "Loại",
            "Parent department",
            "Handbook path",
            "Lead roles",
            "Mandate summary",
            "Boundary / use note",
            "Coverage gap / expansion trigger",
        ]
    )
    for index, code in enumerate(sorted(registry["departments"].keys()), start=1):
        dept = registry["departments"][code]
        handbook = handbook_by_code.get(code)
        mandate = handbook["purpose"] if handbook else note_by_title.get(dept["title_en"], "")
        boundary_note = note_by_title.get(dept["title_en"], "")
        gap_note = " | ".join(handbook.get("coverage_gap", [])) if handbook else ""
        ws_codes.append(
            [
                index,
                code,
                dept["title_en"],
                dept["title_vi"],
                dept["type"],
                dept.get("parent_department"),
                dept["handbook_path"],
                ", ".join(dept["lead_roles"]),
                mandate,
                boundary_note,
                gap_note,
            ]
        )

    wb.save(workbook_path)


def audit_handbook_data(registry: dict) -> list[str]:
    issues = []
    department_codes = set(registry["departments"].keys())
    role_codes = set(registry["roles"].keys())
    for data in HANDBOOKS:
        if data["code"] not in department_codes:
            issues.append(f"{data['code']}: handbook code missing in registry")
        for index, row in enumerate(data.get("interfaces", []), start=1):
            for field in ("with", "func_owner"):
                values = normalize_tokens(row.get(field))
                invalid = [value for value in values if value not in department_codes]
                if invalid:
                    issues.append(
                        f"{data['code']}: interfaces[{index}].{field} must use department codes -> {', '.join(invalid)}"
                    )
            decision_values = normalize_tokens(row.get("decision"))
            invalid_decision = [value for value in decision_values if value in department_codes]
            if invalid_decision:
                issues.append(
                    f"{data['code']}: interfaces[{index}].decision must use role codes -> {', '.join(invalid_decision)}"
                )
        for index, row in enumerate(data.get("outputs", []), start=1):
            decision_values = normalize_tokens(row.get("decision"))
            invalid_decision = [value for value in decision_values if value in department_codes]
            if invalid_decision:
                issues.append(
                    f"{data['code']}: outputs[{index}].decision must use role codes -> {', '.join(invalid_decision)}"
                )
            owner_values = normalize_tokens(row.get("owner"))
            unknown_owner = [value for value in owner_values if value not in department_codes and value not in role_codes]
            if unknown_owner:
                issues.append(
                    f"{data['code']}: outputs[{index}].owner has unknown codes -> {', '.join(unknown_owner)}"
                )
        for index, row in enumerate(data.get("risks", []), start=1):
            escalation_values = normalize_tokens(row.get("escalation"))
            invalid_escalation = [value for value in escalation_values if value in department_codes]
            if invalid_escalation:
                issues.append(
                    f"{data['code']}: risks[{index}].escalation must use role codes -> {', '.join(invalid_escalation)}"
                )
    return issues


def audit_handbooks() -> list[str]:
    issues = []
    forbidden = [
        "Customer Dịch vụ",
        "Department Head",
        "Quy trình Owner",
        "Process Owner",
        "all functions",
        "Team Leader / Supervisor",
        "QA/QMS",
        "Kinh doanh and",
        "các trưởng bộ phận",
        "các phòng chức năng",
        "các phòng liên quan",
        "bộ phận chuyên môn",
        "line manager",
        "line owner",
        "D-QMS",
    ]
    for path in HANDBOOK_DIR.glob("*.html"):
        text = path.read_text(encoding="utf-8")
        for phrase in forbidden:
            if phrase in text:
                issues.append(f"{path.name}: forbidden phrase -> {phrase}")
        for href in re.findall(r'href="([^"]+)"', text):
            if href.startswith("http") or href.startswith("#"):
                continue
            target = (path.parent / href).resolve()
            if not target.exists():
                issues.append(f"{path.name}: broken link -> {href}")
    return issues


def main() -> None:
    registry = load_registry()
    data_issues = audit_handbook_data(registry)
    if data_issues:
        raise SystemExit("\n".join(data_issues))
    write_handbooks(registry)
    update_department_dictionary(WORKBOOK_PATH, registry)
    issues = audit_handbooks()
    if issues:
        raise SystemExit("\n".join(issues))
    print(f"Generated {len(HANDBOOKS)} department handbooks, index and workbook updates.")


if __name__ == "__main__":
    main()
