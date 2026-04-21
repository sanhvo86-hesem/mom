from __future__ import annotations

import json
import os
import re
import unicodedata
from copy import deepcopy
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape

from lxml import etree, html
from openpyxl import load_workbook
from role_boundary_profiles import ROLE_BOUNDARY_PROFILES


ROOT = Path(__file__).resolve().parents[3]
REGISTRY_PATH = ROOT / "tools" / "data" / "role-registry-job-order-cnc.json"
WORKBOOK_PATH = ROOT / "tools" / "data" / "qms-terminology-dictionary.xlsx"
UNRESOLVED_REPORT = ROOT / "tools" / "data" / "role-normalization-unresolved.txt"
BUNDLE_GLOSSARY_PATH = ROOT / "02-Tai-Lieu-He-Thong" / "03-Organization" / "04-RACI-Authority" / "role-and-department-bundles.html"

COMMON_LABEL_REPLACEMENTS = {
    "ANNEX-503 — CNC Vận hành Mô hình and Role Ranh giới": "ANNEX-503 — Mô hình vận hành CNC và ranh giới vai trò",
    "ANNEX-115 — Epicor Giao dịch and giao diện Map": "ANNEX-115 — Bản đồ giao dịch và giao diện Epicor",
    "ANNEX-607 — Quality Văn hóa and Đạo đức Quy tắc": "ANNEX-607 — Văn hóa chất lượng và quy tắc đạo đức",
    "ANNEX-101 — Role-dựa trên Truy cập Map": "ANNEX-101 — Bản đồ phân quyền truy cập theo vai trò",
    "WI-202 — Điều hành hằng ngày, họp tầng, KPI, nhật ký hành động và xử lý vượt cấp | HESEọ QọS": "WI-202 — Điều hành hằng ngày, họp tầng, KPI, nhật ký hành động và xử lý vượt cấp | HESEM QMS",
    "WI-202 — Điều hành hằng ngày, họp tầng, KPI, log hành động và xử lý vượt cấp | HESEọ QọS": "WI-202 — Điều hành hằng ngày, họp tầng, KPI, log hành động và xử lý vượt cấp | HESEM QMS",
    "C04 — Mô-đun C04 — chéo-Department Truyền đạt & Phối hợp | HESEM OS": "C04 — Mô-đun C04 — Giao tiếp liên phòng ban & phối hợp | HESEM OS",
    "C05 — Mô-đun C05 — Customer Dịch vụ Tư duy (B2B) | HESEM OS": "C05 — Mô-đun C05 — Tư duy dịch vụ khách hàng B2B | HESEM OS",
    "C08 — Mô-đun C08 — Data-Định hướng bởi Tư duy & System Sử dụng (ERP/Excel) | HESEM OS": "C08 — Mô-đun C08 — Tư duy dựa trên dữ liệu & sử dụng hệ thống (ERP/Excel) | HESEM OS",
    "C09 — Mô-đun C09 — Time Quản lý & Ưu tiên hóa | HESEM OS": "C09 — Mô-đun C09 — Quản lý thời gian & ưu tiên hóa | HESEM OS",
    "C10 — Mô-đun C10 — Quy trình Lệnh sản xuất CNC (RFQ → Tiền mặt) | HESEM OS": "C10 — Mô-đun C10 — Quy trình lệnh sản xuất CNC (RFQ → thu tiền) | HESEM OS",
    "C11 — Mô-đun C11 — Kinh doanh / RFQ & Review hợp đồng | HESEM OS": "C11 — Mô-đun C11 — RFQ, báo giá & rà soát hợp đồng | HESEM OS",
    "C12 — Mô-đun C12 — Ước tính / Định giá & Job Tính giá thành | HESEM OS": "C12 — Mô-đun C12 — Ước tính, định giá & tính giá thành job | HESEM OS",
    "C13 — Mô-đun C13 — Rủi ro Quản lý & Phiên bản Control | HESEM OS": "C13 — Mô-đun C13 — Quản lý rủi ro & kiểm soát phiên bản | HESEM OS",
    "C14 — Mô-đun C14 — Drawing Diễn giải & GD&T | HESEM OS": "C14 — Mô-đun C14 — Diễn giải bản vẽ & GD&T | HESEM OS",
    "C15 — Mô-đun C15 — Nguyên vật liệu Khoa học & Bề mặt Xử lý | HESEM OS": "C15 — Mô-đun C15 — Khoa học vật liệu & xử lý bề mặt | HESEM OS",
    "C17 — Mô-đun C17 — Kỹ thuật Năng lực by Vai trò (CNC/Setup/CAM) | HESEM OS": "C17 — Mô-đun C17 — Năng lực kỹ thuật theo vai trò (CNC/Setup/CAM) | HESEM OS",
    "C19 — Mô-đun C19 — Tuyến đầu Lãnh đạo & Kèm cặp phát triển (huấn luyện) | HESEM OS": "C19 — Mô-đun C19 — Lãnh đạo tuyến đầu & kèm cặp phát triển | HESEM OS",
    "TRN-ACA-RMAP-01 — Vai trò Lộ trình (30/60/90 ngày)": "TRN-ACA-RMAP-01 — Lộ trình vai trò (30/60/90 ngày)",
    "SYS-OPS-28 — Doanh nghiệp File Plan & Department Lưu hồ sơ Matrix | HESEM OS": "SYS-OPS-28 — File plan doanh nghiệp & ma trận lưu hồ sơ theo phòng ban | HESEM OS",
    "FRM-201 — RFQ Sổ đăng ký": "FRM-201 — Sổ đăng ký RFQ",
    "FRM-202 — hợp đồng Review Bảng kiểm": "FRM-202 — Bảng kiểm rà soát hợp đồng",
    "FRM-206 — Job hoàn thành Bảng kiểm": "FRM-206 — Bảng kiểm hoàn thành job",
    "FRM-212 — Customer Thay đổi Yêu cầu": "FRM-212 — Yêu cầu thay đổi từ khách hàng",
    "FRM-213 — RMA Theo dõi Log": "FRM-213 — Nhật ký theo dõi RMA",
    "FRM-221 — Customer thuộc tính / tài sản Sổ đăng ký": "FRM-221 — Sổ đăng ký tài sản khách hàng",
    "FRM-642 — Final Kiểm tra and CoC Sổ đăng ký": "FRM-642 — Sổ đăng ký kiểm tra cuối & CoC",
    "FRM-654 — Customer sự hài lòng Khảo sát": "FRM-654 — Khảo sát hài lòng khách hàng",
    "FRM-802 — Chuyên cần List": "FRM-802 — Danh sách điểm danh",
    "FRM-803 — OJT Bảng kiểm": "FRM-803 — Bảng kiểm OJT",
    "FRM-805 — Kỹ năng Level Chứng chỉ": "FRM-805 — Chứng chỉ cấp độ kỹ năng",
    "FRM-806 — Chứng nhận Theo dõi Log": "FRM-806 — Nhật ký theo dõi chứng nhận",
    "FRM-807 — Kỹ năng Matrix": "FRM-807 — Ma trận kỹ năng",
    "FRM-808 — Hiệu suất Review Biểu mẫu": "FRM-808 — Phiếu đánh giá hiệu suất",
    "FRM-809 — Kỹ năng and KPI Matrix": "FRM-809 — Ma trận kỹ năng & KPI",
    "FRM-811 — Sự cố Báo cáo": "FRM-811 — Báo cáo sự cố",
    "FRM-821 — Hóa đơn Yêu cầu": "FRM-821 — Phiếu yêu cầu xuất hóa đơn",
}

COMMON_PROSE_REPLACEMENTS = {
    "Customer Dịch vụ": "Dịch vụ khách hàng",
    "Working với": "Làm việc với",
    "Working văn phòng": "Làm việc tại văn phòng",
    "Working trực tiếp": "Làm việc trực tiếp",
    "kinh Nhiệm": "kinh nghiệm",
    "Kinh Nhiệm": "Kinh nghiệm",
    "hàng Ngày": "hằng ngày",
    "Vai trò Mission": "Sứ mệnh vị trí",
    "Vai trò Sứ mệnh": "Sứ mệnh vị trí",
    "Requests tuyển dụng": "yêu cầu tuyển dụng",
    "Requests khách hàng": "yêu cầu khách hàng",
    "Requests chứng từ": "yêu cầu chứng từ",
    "Requests truy vết": "yêu cầu truy vết",
    "Requests đặc thù": "yêu cầu đặc thù",
    "Requests ship": "yêu cầu giao hàng",
    "Requests": "yêu cầu",
    "ship phát hành": "phê duyệt giao hàng",
    "ship gói phát hành": "gói phê duyệt giao hàng",
    "ship xác nhận": "xác nhận giao hàng",
    "source of sự thật": "nguồn sự thật",
    "single source of sự thật": "nguồn sự thật duy nhất",
    "of kém thực thi": "do thực thi kém",
    "Current hiện trường": "thực thi hiện trường",
    "thời gian Từng máy": "thời gian dừng máy",
    "Thời gian Từng máy": "Thời gian dừng máy",
    "Học vấn, kinh nghiệm & yêu cầu tuyển dụng": "Học vấn, kinh nghiệm & yêu cầu tuyển dụng",
}

JD_PROSE_REPLACEMENTS = {
    "Người định giá": "Chuyên viên báo giá",
    "được đọc cùng": "được đọc cùng",
    "công việc Live tế": "công việc thực tế",
    "Requests Live tế": "thực tế",
    "khách hàng dịch vụ": "dịch vụ khách hàng",
    "Time-to-điền / lấp đầy / nhân lực mức sẵn sàng": "Thời gian tuyển đủ / mức sẵn sàng nhân lực",
    "90-day lưu giữ / doanh thu / tỷ lệ nghỉ việc": "Tỷ lệ giữ chân sau 90 ngày / tỷ lệ nghỉ việc",
    "đã chứng nhận-vai trò phạm vi bao phủ": "Độ bao phủ chứng nhận theo vai trò",
    "Customer phản hồi": "phản hồi khách hàng",
    "Customer thuộc tính / tài sản": "tài sản khách hàng",
    "cam kết-date": "ngày cam kết",
    "ready for ship": "sẵn sàng giao hàng",
    "HMLV / tạo / sản xuất-to-đơn hàng": "high-mix, low-volume / sản xuất theo đơn hàng",
    "Job Quản lý / Nhân công Mục nhập / Điều phối": "quản lý Job / nhập nhân công / điều độ Job",
}

COMMON_REGEX_REPLACEMENTS = [
    (r"poka \(chống sai\)(?: \(chống sai\))+[- ]giá đỡ \(giá đỡ \(yoke\)\)", "poka-yoke (chống sai)"),
    (r"poka \(chống sai\)-giá đỡ \(giá đỡ \(yoke\)\)", "poka-yoke (chống sai)"),
    (r"Tinh gọn \(Tinh gọn \(Lean\)\)", "Lean (tinh gọn)"),
    (r"cao-mix\) low-khối lượng", "high-mix, low-volume"),
    (r"HESEọ QọS", "HESEM QMS"),
]

JD_REGEX_REPLACEMENTS = [
    (r"Customer Service là \"mặt tiền dữ liệu đơn hàng\"", "Vị trí Dịch vụ khách hàng là \"mặt tiền dữ liệu đơn hàng\""),
    (r"Customer Service phải là đơn lẻ source of sự thật", "Dịch vụ khách hàng phải là nguồn sự thật duy nhất"),
]


def expr(*codes: str, joiner: str = " / ") -> dict:
    return {"codes": list(codes), "joiner": joiner}


def bundle(name: str, joiner: str = " / ") -> dict:
    return {"bundle": name, "joiner": joiner}


def mix(*tokens: str, joiner: str = " / ") -> dict:
    return {"tokens": list(tokens), "joiner": joiner}


def normalize_ws(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()


def fold_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", normalize_ws(value))
    stripped = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return stripped.lower()


def has_class(element: etree._Element, token: str) -> bool:
    return token in ((element.get("class") or "").split())


def element_text(element: etree._Element) -> str:
    return normalize_ws(" ".join(element.itertext()))


def load_registry() -> dict:
    return json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))


def split_role_hat(code: str) -> tuple[str, str | None]:
    if "[" in code and code.endswith("]"):
        base, hat = code[:-1].split("[", 1)
        return base, hat
    return code, None


def expand_spec(spec: dict, registry: dict) -> tuple[list[str], str]:
    joiner = spec.get("joiner", " / ")
    if "bundle" in spec:
        return list(registry["bundles"][spec["bundle"]]), joiner
    return list(spec["codes"]), joiner


def resolve_entity(code: str, registry: dict) -> tuple[str, dict, str | None]:
    base, hat = split_role_hat(code)
    if base in registry["roles"]:
        return "role", registry["roles"][base], hat
    if base in registry.get("departments", {}):
        return "department", registry["departments"][base], None
    raise KeyError(f"Unknown role/department code: {code}")


def entity_link(code: str, current_file: Path, registry: dict) -> str:
    entity_type, meta, hat = resolve_entity(code, registry)
    title = meta["title_en"]
    if hat:
        title = f'{title} [{registry["hats"][hat]["label_en"]}]'
    target_key = "jd_path" if entity_type == "role" else "handbook_path"
    target_abs = ROOT / meta[target_key]
    href = os.path.relpath(target_abs, current_file.parent).replace("\\", "/")
    link_class = "role-link" if entity_type == "role" else "dept-link"
    code_class = "role-code" if entity_type == "role" else "dept-code"
    return (
        f'<a class="entity-link {link_class}" href="{href}" title="{title} ({meta["title_vi"]})">'
        f'<span class="entity-code {code_class}">{code}</span></a>'
    )


def bundle_anchor_id(name: str) -> str:
    return f'bundle-{re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")}'


def bundle_meta(name: str, registry: dict) -> dict[str, str]:
    meta = registry.get("bundle_meta", {}).get(name)
    if meta:
        return meta
    return {
        "label_en": name,
        "label_vi": name.replace("_", " ").title(),
        "kind": "bundle",
        "use_vi": "Nhom actor explicit da duoc cong bo trong registry nguon.",
        "avoid_vi": "Khong dung de che mo trach nhiem hoac thay the role / D-code cu the khi da xac dinh duoc actor that.",
    }


def bundle_href(name: str, current_file: Path) -> str:
    return f'{os.path.relpath(BUNDLE_GLOSSARY_PATH, current_file.parent).replace("\\", "/")}#{bundle_anchor_id(name)}'


def bundle_title(name: str, registry: dict) -> str:
    meta = bundle_meta(name, registry)
    members = " / ".join(registry.get("bundles", {}).get(name, []))
    return f'{meta["label_en"]} ({meta["label_vi"]}) — {members}'


def render_spec(spec: dict, current_file: Path, registry: dict) -> str:
    if "bundle" in spec:
        return f'<span class="entity-cluster role-cluster">{render_bundle_chip(spec["bundle"], current_file, registry)}</span>'
    if "tokens" in spec:
        return render_token_cluster(list(spec["tokens"]), current_file, registry, joiner=spec.get("joiner", " / "))
    codes, joiner = expand_spec(spec, registry)
    sep = f'<span class="entity-sep role-sep">{joiner.strip()}</span>'
    chips = [entity_link(code, current_file, registry) for code in codes]
    body = "".join(chips[i] + (sep if i < len(chips) - 1 else "") for i in range(len(chips)))
    return f'<span class="entity-cluster role-cluster">{body}</span>'


def token_is_entity(token: str, registry: dict) -> bool:
    base, _ = split_role_hat(token)
    return base in registry["roles"] or base in registry.get("departments", {})


def token_is_bundle(token: str, registry: dict) -> bool:
    return token in registry.get("bundles", {})


def render_bundle_chip(token: str, current_file: Path, registry: dict) -> str:
    return (
        f'<a class="entity-link bundle-link bundle-chip" data-bundle="{xml_escape(token)}" '
        f'href="{xml_escape(bundle_href(token, current_file))}" title="{xml_escape(bundle_title(token, registry))}">'
        f'<span class="entity-code bundle-code">{xml_escape(token)}</span></a>'
    )


def render_token_cluster(tokens: list[str], current_file: Path, registry: dict, joiner: str = " / ") -> str:
    if not tokens:
        return ""
    sep = f'<span class="entity-sep role-sep">{joiner.strip()}</span>'
    chunks: list[str] = []
    for index, token in enumerate(tokens):
        if token_is_entity(token, registry):
            chunks.append(entity_link(token, current_file, registry))
        elif token_is_bundle(token, registry):
            chunks.append(render_bundle_chip(token, current_file, registry))
        else:
            chunks.append(f'<span class="chip">{xml_escape(token)}</span>')
        if index < len(tokens) - 1:
            chunks.append(sep)
    return f'<span class="entity-cluster role-cluster">{"".join(chunks)}</span>'


def set_element_html(element: etree._Element, html_fragment: str) -> None:
    for child in list(element):
        element.remove(child)
    element.text = None
    prev = None
    for frag in html.fragments_fromstring(html_fragment):
        if isinstance(frag, str):
            if prev is None:
                element.text = (element.text or "") + frag
            else:
                prev.tail = (prev.tail or "") + frag
        else:
            element.append(frag)
            prev = frag


def parse_html_document(source_text: str) -> etree._Element:
    parser = html.HTMLParser(encoding="utf-8")
    if "<html" in source_text.lower():
        return html.document_fromstring(source_text, parser=parser)
    return html.fromstring(source_text, parser=parser)


def serialize_html_document(doc: etree._Element) -> str:
    if doc.tag.lower() == "html":
        return html.tostring(doc, encoding="unicode", method="html", doctype="<!DOCTYPE html>")
    return html.tostring(doc, encoding="unicode", method="html")


def extract_jd_body_source(source_text: str) -> str:
    if "<html" in source_text.lower():
        match = re.search(r"<body[^>]*>(.*)</body>", source_text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(1)
    marker = source_text.find('<div class="container"')
    if marker != -1:
        return source_text[marker:]
    return source_text


def extract_jd_head_assets(source_text: str, doc_path: Path) -> str:
    if "<html" in source_text.lower() and "<head" in source_text.lower():
        head_match = re.search(r"<head[^>]*>(.*?)</head>", source_text, flags=re.IGNORECASE | re.DOTALL)
        head_source = head_match.group(1) if head_match else ""
    else:
        marker = source_text.find('<div class="container"')
        head_source = source_text[:marker] if marker != -1 else ""

    style_blocks = re.findall(r"<style\b[^>]*>.*?</style>", head_source, flags=re.IGNORECASE | re.DOTALL)
    link_tags = re.findall(r"<link\b[^>]*>", head_source, flags=re.IGNORECASE)

    prefix = head_source.split(style_blocks[0], 1)[0] if style_blocks else head_source
    raw_css = ""
    if "</style>" in prefix and "<style" not in prefix.lower():
        raw_candidate = prefix.split("</style>", 1)[0]
        raw_candidate = re.sub(r"<[^>]+>", "", raw_candidate).strip()
        if raw_candidate and "{" in raw_candidate and "}" in raw_candidate:
            raw_css = raw_candidate

    assets: list[str] = []
    if raw_css:
        assets.append(f"<style>\n{raw_css}\n</style>")

    for block in style_blocks:
        cleaned = block.strip()
        if cleaned and cleaned not in assets:
            assets.append(cleaned)

    asset_href = os.path.relpath(ROOT / "assets" / "style.css", doc_path.parent).replace("\\", "/")
    if not any("assets/style.css" in link for link in link_tags):
        link_tags.append(f'<link rel="stylesheet" href="{asset_href}">')

    for link in link_tags:
        cleaned = link.replace("/>", ">").strip()
        if cleaned and cleaned not in assets:
            assets.append(cleaned)

    return "\n".join(assets)


def build_html_document(title_text: str, head_assets: str, body_html: str) -> str:
    parts = [
        "<!DOCTYPE html>",
        '<html lang="vi">',
        "<head>",
        '<meta charset="utf-8">',
        '<meta content="width=device-width, initial-scale=1.0" name="viewport">',
        f"<title>{xml_escape(title_text)}</title>",
    ]
    if head_assets.strip():
        parts.append(head_assets.strip())
    parts.extend([
        "</head>",
        "<body>",
        body_html.strip(),
        "</body>",
        "</html>",
        "",
    ])
    return "\n".join(parts)


def link_bundle_tokens_in_html(document: str, current_file: Path, registry: dict) -> str:
    registry_tokens: list[str] = []
    registry_tokens.extend(registry.get("bundles", {}).keys())
    registry_tokens.extend(registry.get("departments", {}).keys())
    registry_tokens.extend(registry.get("roles", {}).keys())
    for role_code, role_meta in registry.get("roles", {}).items():
        for hat in role_meta.get("hats_allowed", []):
            registry_tokens.append(f"{role_code}[{hat}]")
    registry_tokens = sorted(set(registry_tokens), key=len, reverse=True)
    if not registry_tokens:
        return document
    pattern = re.compile(
        r"(?<![\w/-])(" + "|".join(re.escape(name) for name in registry_tokens) + r")(?![\w/-])"
    )
    parts = re.split(r"(<[^>]+>)", document)
    result: list[str] = []
    stack: list[tuple[str, bool]] = []

    def current_skip() -> bool:
        return any(flag for _, flag in stack)

    for part in parts:
        if not part:
            continue
        if part.startswith("<") and part.endswith(">"):
            result.append(part)
            tag_match = re.match(r"<\s*(/)?\s*([a-zA-Z0-9:_-]+)([^>]*)>", part)
            if not tag_match:
                continue
            closing, tag_name, attrs = tag_match.groups()
            tag_name = tag_name.lower()
            attrs = attrs or ""
            if closing:
                for index in range(len(stack) - 1, -1, -1):
                    if stack[index][0] == tag_name:
                        del stack[index:]
                        break
                continue
            self_closing = attrs.strip().endswith("/") or part.startswith("<!--")
            skip = (
                tag_name in {"a", "script", "style", "code", "pre", "title", "svg", "text", "tspan", "desc"}
                or any(token in attrs for token in ["bundle-chip", "bundle-code", "role-code", "dept-code", "entity-code"])
            )
            if not self_closing:
                stack.append((tag_name, skip))
            continue
        if current_skip():
            result.append(part)
            continue
        def replace_token(match: re.Match[str]) -> str:
            token = match.group(1)
            if token_is_bundle(token, registry):
                return render_bundle_chip(token, current_file, registry)
            if token_is_entity(token, registry):
                return entity_link(token, current_file, registry)
            return token

        result.append(pattern.sub(replace_token, part))
    return "".join(result)


def link_published_actor_terms_in_html(document: str, current_file: Path, registry: dict) -> str:
    alias_specs = {
        "Người định giá (SAL-03)": expr("EST"),
        "Người định giá": expr("EST"),
        "Production Planner": expr("PPL"),
        "Planner": expr("PPL"),
        "IT/Data": bundle("SYSTEM_OWNERS"),
        "IT / Data": bundle("SYSTEM_OWNERS"),
        "IT/Digital": bundle("SYSTEM_OWNERS"),
        "IT/BI": bundle("SYSTEM_OWNERS"),
        "IT / BI": bundle("SYSTEM_OWNERS"),
        "QA/IT": mix("QA", "SYSTEM_OWNERS"),
        "QA / IT": mix("QA", "SYSTEM_OWNERS"),
        "Người phụ trách Bảng điều khiển": bundle("MR_REPORT_OWNERS"),
        "bảng điều khiển Người phụ trách": bundle("MR_REPORT_OWNERS"),
        "authorized Người phê duyệt": bundle("FUNC_HEADS"),
        "Production Director": expr("PD"),
        "Sản xuất Giám đốc": expr("PD"),
        "Lập trình viên (CNC)": expr("CAM"),
        "Buyer / Purchasing": expr("BUY"),
        "Buyer": expr("BUY"),
        "Kho": expr("D-WHS"),
        "Bảo trì": expr("D-MNT"),
        "Operator + Setup Lead + QA": expr("OPR", "SET", "QA", joiner=" + "),
        "Operator + Setup Lead": expr("OPR", "SET", joiner=" + "),
        "Supervisor + QA": mix("FRONTLINE_LEADS", "QA", joiner=" + "),
        "QC + Setup": expr("QC", "SET", joiner=" + "),
        "Team Leader kho": expr("D-WHS"),
        "ENG Manager": expr("ENGM"),
        "Shift Supervisor": expr("SL"),
        "Supervisor": bundle("FRONTLINE_LEADS"),
        "Setup Lead": expr("SET"),
        "QA lead": expr("QA"),
        "QA Lead": expr("QA"),
        "Team Leader": bundle("FRONTLINE_LEADS"),
        "ENG/Setup": expr("D-ENG", "SET"),
        "WHS/IQC": expr("D-WHS", "QC"),
        "ENG": expr("D-ENG"),
    }
    phrases = sorted(alias_specs, key=len, reverse=True)
    pattern = re.compile(r"(?<![\w-])(" + "|".join(re.escape(item) for item in phrases) + r")(?![\w-])")
    parts = re.split(r"(<[^>]+>)", document)
    result: list[str] = []
    stack: list[tuple[str, bool]] = []

    def current_skip() -> bool:
        return any(flag for _, flag in stack)

    for part in parts:
        if not part:
            continue
        if part.startswith("<") and part.endswith(">"):
            result.append(part)
            tag_match = re.match(r"<\s*(/)?\s*([a-zA-Z0-9:_-]+)([^>]*)>", part)
            if not tag_match:
                continue
            closing, tag_name, attrs = tag_match.groups()
            tag_name = tag_name.lower()
            attrs = attrs or ""
            if closing:
                for index in range(len(stack) - 1, -1, -1):
                    if stack[index][0] == tag_name:
                        del stack[index:]
                        break
                continue
            self_closing = attrs.strip().endswith("/") or part.startswith("<!--")
            skip = (
                tag_name in {"a", "script", "style", "code", "pre", "title", "svg", "text", "tspan", "desc"}
                or any(token in attrs for token in ["bundle-chip", "bundle-code", "role-code", "dept-code", "entity-code"])
            )
            if not self_closing:
                stack.append((tag_name, skip))
            continue
        if current_skip():
            result.append(part)
            continue

        def replace_alias(match: re.Match[str]) -> str:
            phrase = match.group(1)
            return render_spec(alias_specs[phrase], current_file, registry)

        result.append(pattern.sub(replace_alias, part))
    return "".join(result)


def normalize_hybrid_actor_clusters_in_html(document: str, current_file: Path, registry: dict) -> str:
    role_ref = lambda code: (
        rf'(?:<span class="entity-cluster role-cluster">)?'
        rf'<a [^>]*><span class="entity-code role-code">{code}</span></a>'
        rf'(?:</span>)?'
    )
    qa_chip = role_ref("QA")
    qc_chip = role_ref("QC")
    ppl_chip = role_ref("PPL")
    replacements = [
        (
            re.compile(rf'(?:Sản xuất\s+)?Supervisor\s*\+\s*{qa_chip}', re.S),
            render_spec(mix("FRONTLINE_LEADS", "QA", joiner=" + "), current_file, registry),
        ),
        (
            re.compile(rf'{qa_chip}\s*\+\s*Supervisor', re.S),
            render_spec(mix("FRONTLINE_LEADS", "QA", joiner=" + "), current_file, registry),
        ),
        (
            re.compile(rf'{qc_chip}\s*\+\s*Setup', re.S),
            render_spec(expr("QC", "SET", joiner=" + "), current_file, registry),
        ),
        (
            re.compile(rf'Setup\s*\+\s*{qc_chip}', re.S),
            render_spec(expr("SET", "QC", joiner=" + "), current_file, registry),
        ),
        (
            re.compile(rf'{ppl_chip}\s*/\s*Sản xuất Giám đốc', re.S),
            render_spec(expr("PPL", "PD", joiner=" / "), current_file, registry),
        ),
    ]
    for pattern, replacement in replacements:
        document = pattern.sub(replacement, document)
    return document


def cleanup_known_render_artifacts_in_html(document: str) -> str:
    document = re.sub(
        r'<span class="entity-cluster role-cluster">.*?<span class="entity-code dept-code">D-ENG</span>.*?</span>INEERING',
        "ENGINEERING",
        document,
        flags=re.S,
    )
    return document


def generate_bundle_glossary(registry: dict) -> None:
    current_file = BUNDLE_GLOSSARY_PATH
    title_text = "ORG-BUNDLE-001 — Thuat ngu nhom vai tro va nhom phong ban | HESEM QMS"
    asset_href = os.path.relpath(ROOT / "assets" / "style.css", current_file.parent).replace("\\", "/")
    head_assets = "\n".join([
        f'<link rel="stylesheet" href="{asset_href}">',
        "<style>",
        ".hero{border:1px solid var(--ln);border-radius:var(--r-lg);padding:18px 20px;background:linear-gradient(135deg,#f8fbff 0%,#fffaf0 100%);margin:0 0 18px;}",
        ".hero h1{font-size:22px;line-height:1.35;color:var(--navy);margin:0 0 8px}",
        ".hero p{font-size:13px;color:var(--ink2);margin:0}",
        ".toc{border:1px solid var(--ln);border-radius:var(--r);padding:16px;background:var(--bg2);margin:18px 0 24px;}",
        ".toc-title{font-size:12px;font-weight:700;color:var(--navy);text-transform:uppercase;letter-spacing:.4px;margin-bottom:10px;}",
        ".toc-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:8px;}",
        ".toc-grid a{display:block;padding:8px 10px;border:1px solid var(--ln);border-radius:6px;background:var(--bg);font-size:12px;color:var(--ink);text-decoration:none;}",
        ".mini-note{font-size:12px;color:var(--ink2);line-height:1.6}",
        "</style>",
    ])

    toc_links: list[str] = []
    table_rows: list[str] = []
    kind_order = {"role bundle": 0, "department bundle": 1, "mixed bundle": 2, "bundle": 3}
    for bundle_name in sorted(
        registry.get("bundles", {}),
        key=lambda item: (kind_order.get(bundle_meta(item, registry).get("kind", "bundle"), 99), item),
    ):
        meta = bundle_meta(bundle_name, registry)
        anchor = bundle_anchor_id(bundle_name)
        toc_links.append(f'<a href="#{anchor}">{xml_escape(bundle_name)} — {xml_escape(meta["label_vi"])}</a>')
        members = render_token_cluster(list(registry["bundles"][bundle_name]), current_file, registry)
        table_rows.append(
            "<tr>"
            f'<td id="{xml_escape(anchor)}">{render_bundle_chip(bundle_name, current_file, registry)}</td>'
            f"<td>{xml_escape(meta['label_vi'])}</td>"
            f"<td>{xml_escape(meta.get('kind', 'bundle'))}</td>"
            f"<td>{xml_escape(meta.get('use_vi', ''))}</td>"
            f"<td>{xml_escape(meta.get('avoid_vi', ''))}</td>"
            f"<td>{members}</td>"
            "</tr>"
        )

    body_html = f"""
<div class="container">
 <div class="page">
  <div class="page-body">
   <div class="form-header">
    <div class="fh-left">
     <a class="brand-logo" href="../../../01-QMS-Portal/portal.html"><img alt="HESEM Logo" src="../../../assets/hesem-logo.svg"></a>
     <div class="fh-company">
      <a href="../../../01-QMS-Portal/portal.html">HESEM ENGINEERING</a>
      <span>Tai lieu he thong • To chuc</span>
     </div>
    </div>
    <div class="title">
     <strong>ORG-BUNDLE-001 — Thuat ngu nhom vai tro va nhom phong ban</strong>
     <span class="sub-vn">Nguon cong bo de giai nghia cac bundle duoc phep dung trong SOP / WI / ANNEX / JD / RACI</span>
    </div>
    <div class="meta">
     <div class="row"><span><b>Code:</b></span><span>ORG-BUNDLE-001</span></div>
     <div class="row"><span><b>Version:</b></span><span>V0</span></div>
     <div class="row"><span><b>Owner:</b></span><span>{render_token_cluster(["QMS", "D-HR"], current_file, registry)}</span></div>
     <div class="row"><span><b>Approved by:</b></span><span>{render_token_cluster(["CEO"], current_file, registry)}</span></div>
    </div>
   </div>
   <div class="doc-content" id="docContent">
    <div class="form-sheet">
     <div class="hero">
      <h1>Bundle chi duoc dung khi da duoc cong bo, co thanh phan ro va co link giai nghia</h1>
      <p>Tai lieu nay khoa nghia cua cac bundle trong mo hinh <b>job-order CNC</b>. Moi bundle phai truy ve duoc JD / D-code goc, co pham vi dung ro, va khong duoc dung de che mo authority hay named accountability.</p>
     </div>
     <div class="note"><b>Rule bat buoc:</b><br>Neu tai lieu da xac dinh duoc actor cu the thi phai dung role code hoac D-code. Chi dung bundle khi trach nhiem that su la lop actor lap lai da duoc cong bo. Moi bundle token dung doc lap trong noi dung hien thi phai link ve trang nay.</div>
     <div class="toc"><div class="toc-title">Muc luc nhanh</div><div class="toc-grid">{''.join(toc_links)}</div></div>
     <h2 class="h2">1. Danh muc bundle duoc cong bo</h2>
     <div class="table-card"><table class="table"><colgroup><col style="width:15%"><col style="width:18%"><col style="width:12%"><col style="width:22%"><col style="width:18%"><col style="width:15%"></colgroup><thead><tr><th>Bundle code</th><th>Ten goi chuan</th><th>Loai</th><th>Duoc dung khi</th><th>Khong dung de che</th><th>Thanh phan</th></tr></thead><tbody>{''.join(table_rows)}</tbody></table></div>
     <h2 class="h2">2. Nguyen tac doc bundle</h2>
     <ul class="tight">
      <li>Bundle khong tao ra JD moi. Moi chip trong bundle van phai truy ve duoc JD hoac handbook goc.</li>
      <li>Bundle cap role dung cho layer actor; bundle cap department dung cho mandate cap phong ban; bundle mixed chi dung khi tai lieu dang noi toi lop enablement hoac interface lien phong ban that su on dinh.</li>
      <li>Neu mot o owner / approver / hold-release cell da biet ro actor cu the, phai bo bundle va doi ve role code / D-code dung layer.</li>
     </ul>
    </div>
   </div>
  </div>
 </div>
</div>
"""
    BUNDLE_GLOSSARY_PATH.write_text(build_html_document(title_text, head_assets, body_html), encoding="utf-8")


def title_aliases() -> dict[str, str]:
    return {
        "Tổng Giám đốc": "Chief Executive Officer",
        "Tong Giam doc": "Chief Executive Officer",
        "Sản xuất Giám đốc": "Production Director",
        "Giám đốc Sản xuất": "Production Director",
        "Kỹ thuật Lead / Manager": "Engineering Lead / Manager",
        "Kỹ thuật Lead/Manager": "Engineering Lead / Manager",
        "Trưởng Engineering": "Engineering Lead / Manager",
        "QA Lead": "QA Manager",
        "Trưởng QA": "QA Manager",
        "QMS Coordinator": "QMS Engineer",
        "QA Engineer": "Quality Engineer",
        "Kỹ sư Chất lượng": "Quality Engineer",
        "QC Lead": "QC Inspector Lead",
        "QC Team Leader": "QC Inspector Lead",
        "Metrology Lead": "Metrology and Calibration Specialist",
        "Quy trình Engineer": "Process Engineer",
        "Kỹ sư Quy trình": "Process Engineer",
        "Customer Dịch vụ": "Customer Service",
        "Người định giá": "Estimator",
        "Báo giá": "Estimator",
        "Trưởng Chuỗi cung ứng": "Supply Chain Manager",
        "Trưởng Supply Chain": "Supply Chain Manager",
        "Mua hàng": "Buyer / Purchasing",
        "Kho Nhân viên văn phòng": "Warehouse Clerk",
        "Nhân viên Kho": "Warehouse Clerk",
        "Tooling Thủ kho": "Tool Crib / Tool Storekeeper",
        "Thủ kho Dao cụ": "Tool Crib / Tool Storekeeper",
        "Hậu cần / Giao vận": "Logistics / Shipping Coordinator",
        "Quản trị IT": "IT Admin",
        "Governance viên hệ thống Epicor": "Epicor System Administrator",
        "HR Lead": "HR Manager",
        "EHS Manager": "EHS Specialist",
        "Trưởng ca": "Shift Leader",
        "Setup Leader": "Setup Technician",
        "Operator": "CNC Operator",
        "Production Operator": "CNC Operator",
        "Kỹ thuật viên Bảo trì": "Maintenance Technician",
        "Workshop Manager": "CNC Workshop Manager",
        "Production Engineer-IE": "Production Engineer / Industrial Engineer",
        "Production Engineer / IE": "Production Engineer / Industrial Engineer",
        "Manufacturing Engineer": "Production Engineer / Industrial Engineer",
        "Continuous Improvement Lead": "Production Engineer / Industrial Engineer [CI]",
        "Product Safety Officer": "QA Manager [PSO]"
    }


def department_alias_map() -> dict[str, dict]:
    return {
        "Executive": expr("D-EXEC"),
        "Executive Department": expr("D-EXEC"),
        "Phòng Điều hành": expr("D-EXEC"),
        "EXEC": expr("D-EXEC"),
        "Sales": expr("D-SCS"),
        "Commercial": expr("D-SCS"),
        "Kinh doanh": expr("D-SCS"),
        "Sales & Customer Service": expr("D-SCS"),
        "Sales and Customer Service": expr("D-SCS"),
        "Sales and Customer Service Department": expr("D-SCS"),
        "Phòng Kinh doanh và Dịch vụ khách hàng": expr("D-SCS"),
        "Kinh doanh-CS": expr("D-SCS"),
        "Kinh doanh / CS": expr("D-SCS"),
        "Kinh doanh & Customer Dịch vụ": expr("D-SCS"),
        "SAL": expr("D-SCS"),
        "Engineering": expr("D-ENG"),
        "Engineering Department": expr("D-ENG"),
        "Kỹ thuật": expr("D-ENG"),
        "Phòng Kỹ thuật": expr("D-ENG"),
        "ENG": expr("D-ENG"),
        "Production": expr("D-PROD"),
        "Production Department": expr("D-PROD"),
        "Sản xuất": expr("D-PROD"),
        "Phòng Sản xuất": expr("D-PROD"),
        "PRO": expr("D-PROD"),
        "Planning": expr("D-PPC"),
        "Production Planning": expr("D-PPC"),
        "Production Control": expr("D-PPC"),
        "Kế hoạch": expr("D-PPC"),
        "Điều độ": expr("D-PPC"),
        "Bộ phận Điều độ và Kiểm soát sản xuất": expr("D-PPC"),
        "PPC": expr("D-PPC"),
        "Maintenance": expr("D-MNT"),
        "Maintenance Department": expr("D-MNT"),
        "Maintenance Function": expr("D-MNT"),
        "Báº£o trÃ¬": expr("D-MNT"),
        "PhÃ¢n há»‡ Báº£o trÃ¬": expr("D-MNT"),
        "Bá»™ pháº­n Báº£o trÃ¬": expr("D-MNT"),
        "Quality": expr("D-QUAL"),
        "Quality Department": expr("D-QUAL"),
        "Chất lượng": expr("D-QUAL"),
        "Phòng Chất lượng": expr("D-QUAL"),
        "QUAL": expr("D-QUAL"),
        "Supply Chain": expr("D-SCM"),
        "Supply Chain Department": expr("D-SCM"),
        "Chuỗi cung ứng": expr("D-SCM"),
        "Phòng Chuỗi cung ứng": expr("D-SCM"),
        "SCM": expr("D-SCM"),
        "Purchasing": expr("D-PUR"),
        "Mua hàng": expr("D-PUR"),
        "Bộ phận Mua hàng": expr("D-PUR"),
        "PUR": expr("D-PUR"),
        "Warehouse": expr("D-WHS"),
        "Warehouse Department": expr("D-WHS"),
        "Kho": expr("D-WHS"),
        "WHS": expr("D-WHS"),
        "Bộ phận Kho": expr("D-WHS"),
        "Tool Crib": expr("D-TCR"),
        "Tool Store": expr("D-TCR"),
        "Kho dao cụ": expr("D-TCR"),
        "Kho dụng cụ": expr("D-TCR"),
        "Bộ phận Kho dao cụ": expr("D-TCR"),
        "Logistics": expr("D-LOG"),
        "Shipping": expr("D-LOG"),
        "Giao vận": expr("D-LOG"),
        "Hậu cần": expr("D-LOG"),
        "Bộ phận Hậu cần và Giao vận": expr("D-LOG"),
        "LOG": expr("D-LOG"),
        "Finance": expr("D-FIN"),
        "Finance Department": expr("D-FIN"),
        "Tài chính": expr("D-FIN"),
        "Phòng Tài chính": expr("D-FIN"),
        "HR": expr("D-HR"),
        "Human Resources": expr("D-HR"),
        "Human Resources Department": expr("D-HR"),
        "Nhân sự": expr("D-HR"),
        "Phòng Nhân sự": expr("D-HR"),
        "EHS": expr("D-EHS"),
        "EHS Department": expr("D-EHS"),
        "Phòng EHS": expr("D-EHS"),
        "IT": expr("D-IT"),
        "IT Department": expr("D-IT"),
        "CNTT": expr("D-IT"),
        "Phòng CNTT": expr("D-IT"),
        "ERP": expr("D-ERP"),
        "Epicor": expr("D-ERP"),
        "Epicor / ERP": expr("D-ERP"),
        "ERP Administration": expr("D-ERP"),
        "Bộ phận Quản trị ERP": expr("D-ERP"),
    }


def role_alias_map() -> dict[str, dict]:
    return {
        "ENG": expr("D-ENG"),
        "OPS": expr("D-PROD"),
        "SALES": expr("D-SCS"),
        "PLN": expr("PPL"),
        "CEO": expr("CEO"),
        "Chief Executive Officer": expr("CEO"),
        "Tổng Giám đốc": expr("CEO"),
        "Tong Giam doc": expr("CEO"),
        "Production Director": expr("PD"),
        "Sản xuất Giám đốc": expr("PD"),
        "Giám đốc Sản xuất": expr("PD"),
        "Engineering Lead": expr("ENGM"),
        "Engineering Lead / Manager": expr("ENGM"),
        "Engineering Manager": expr("ENGM"),
        "Kỹ thuật Lead / Manager": expr("ENGM"),
        "Kỹ thuật Lead/Manager": expr("ENGM"),
        "Kỹ thuật Lead / QA Lead": expr("ENGM", "QA"),
        "Trưởng Engineering": expr("ENGM"),
        "Production Planner": expr("PPL"),
        "Planner": expr("PPL"),
        "Kế hoạch Team Leader": expr("PPL"),
        "Kế hoạch sản xuất": expr("PPL"),
        "QA Manager": expr("QA"),
        "QA Lead": expr("QA"),
        "QA lead": expr("QA"),
        "qa lead": expr("QA"),
        "Trưởng QA": expr("QA"),
        "QA": expr("QA"),
        "QA/HR": expr("QA", "HR"),
        "QA/HRD": expr("QA", "HR"),
        "QMR": expr("QA[QMR]"),
        "QA/QMS": expr("QA", "QMS"),
        "QMS/QA": expr("QA", "QMS"),
        "Document Controller": expr("QMS[DC]"),
        "QMS Engineer": expr("QMS"),
        "QMS Coordinator": expr("QMS"),
        "QMS Manager": expr("QMS"),
        "Kỹ sư QMS": expr("QMS"),
        "Reviewer": bundle("FUNC_OWNERS"),
        "Reviewers": bundle("FUNC_OWNERS"),
        "Champion": bundle("DEPLOYMENT_LEADS"),
        "Champions": bundle("DEPLOYMENT_LEADS"),
        "Dept Manager": bundle("FUNC_HEADS"),
        "Dept Managers": bundle("FUNC_HEADS"),
        "Người đào tạo": bundle("OJT_COACHES"),
        "Trainer": bundle("OJT_COACHES"),
        "HR/TRN": mix("HR", "OJT_COACHES", joiner=" + "),
        "Quality Engineer": expr("QE"),
        "QA Engineer": expr("QE"),
        "Kỹ sư Chất lượng": expr("QE"),
        "QC Inspector Lead": expr("QCL"),
        "QC Lead": expr("QCL"),
        "QC Team Leader": expr("QCL"),
        "IQC Team Leader": expr("QCL"),
        "QC Inspector": expr("QC"),
        "Quality Engineer / QC": expr("QE", "QC"),
        "QC Inspector / CMM Programmer / Operator": expr("QC"),
        "QC Inspector / CMM Programmer-Operator": expr("QC"),
        "QC Inspector / IQC": expr("QC"),
        "QC Inspector / IPQC": expr("QC"),
        "QC Inspector / Appraiser": expr("QC"),
        "QC Operator": expr("QC"),
        "QA/QC Lead": expr("QCL"),
        "Metrology and Calibration Specialist": expr("MCS"),
        "Metrology Lead": expr("MCS"),
        "Calibration Technician": expr("MCS"),
        "DFM Engineer": expr("DFM"),
        "Process Engineer": expr("PE"),
        "Quy trình Engineer": expr("PE"),
        "Kỹ sư Quy trình": expr("PE"),
        "CAM / NC Programmer": expr("CAM"),
        "CAM/NC Programmer": expr("CAM"),
        "CAM/NC Lập trình viên (CNC)": expr("CAM"),
        "Programmer": expr("CAM"),
        "Customer Service": expr("CS"),
        "CS": expr("CS"),
        "CS (Customer Service)": expr("CS"),
        "Customer Dịch vụ": expr("CS"),
        "Estimator": expr("EST"),
        "Người định giá": expr("EST"),
        "Báo giá": expr("EST"),
        "Supply Chain Manager": expr("SCM"),
        "Trưởng Chuỗi cung ứng": expr("SCM"),
        "Trưởng Supply Chain": expr("SCM"),
        "Buyer / Purchasing": expr("BUY"),
        "Buyer": expr("BUY"),
        "Nhân viên mua hàng": expr("BUY"),
        "Warehouse Clerk": expr("WAR"),
        "Warehouse Supervisor": expr("SCM"),
        "Kho Nhân viên văn phòng": expr("WAR"),
        "Nhân viên Kho": expr("WAR"),
        "Warehouse Operator": expr("WAR"),
        "Tool Crib / Tool Storekeeper": expr("TOOL"),
        "Tool Crib": expr("TOOL"),
        "Tooling Thủ kho": expr("TOOL"),
        "Thủ kho Dao cụ": expr("TOOL"),
        "Logistics / Shipping Coordinator": expr("LOG"),
        "Shipping Coordinator": expr("LOG"),
        "Hậu cần / Giao vận Điều phối viên": expr("LOG"),
        "IT Administrator": expr("ITA"),
        "IT Admin": expr("ITA"),
        "IT Manager": expr("ITA"),
        "IT System Administrator": expr("ITA"),
        "IT System Quản trị viên": expr("ITA"),
        "IT Data Owner": expr("ITA"),
        "Quản trị IT": expr("ITA"),
        "Epicor System Administrator": expr("ESA"),
        "Governance viên hệ thống Epicor": expr("ESA"),
        "Epicor Administrator": expr("ESA"),
        "Epicor Admin": expr("ESA"),
        "HR Manager": expr("HR"),
        "HR Lead": expr("HR"),
        "HR Coordinator": expr("HR"),
        "EHS Specialist": expr("EHS"),
        "EHS Manager": expr("EHS"),
        "Shift Leader": expr("SL"),
        "Shift Lead": expr("SL"),
        "Supervisor": bundle("FRONTLINE_LEADS"),
        "Team Leader": bundle("FRONTLINE_LEADS"),
        "team leader": bundle("FRONTLINE_LEADS"),
        "Foreman": bundle("FRONTLINE_LEADS"),
        "Trưởng ca": expr("SL"),
        "Setup Technician": expr("SET"),
        "Setup Leader": expr("SET"),
        "Setup Lead": expr("SET"),
        "CNC Operator": expr("OPR"),
        "Operator": expr("OPR"),
        "Production Operator": expr("OPR"),
        "CNC Operator / Production Operator": expr("OPR"),
        "Operator + Setup Lead": expr("OPR", "SET", joiner=" + "),
        "Operator + Setup Lead + QA": expr("OPR", "SET", "QA", joiner=" + "),
        "QC + Setup": expr("QC", "SET", joiner=" + "),
        "Supervisor + QA": mix("FRONTLINE_LEADS", "QA", joiner=" + "),
        "Maintenance Technician": expr("MNT"),
        "Maintenance Planner": expr("MNT"),
        "Maintenance Supervisor": expr("MNT"),
        "Maintenance Manager": expr("MNT"),
        "Kỹ thuật viên Bảo trì": expr("MNT"),
        "CNC Workshop Manager": expr("WKM"),
        "Workshop Manager": expr("WKM"),
        "Ops Manager": expr("WKM"),
        "Operations Manager": expr("WKM"),
        "Cleaning and Packaging Supervisor": expr("CPS"),
        "Cleaning Supervisor": expr("CPS"),
        "Cleaning / Packaging Technician": expr("CPT"),
        "Cleaning Technician": expr("CPT"),
        "Làm sạch & Đóng gói Supervisor": expr("CPS"),
        "Làm sạch / Đóng gói Technician": expr("CPT"),
        "QA Final": expr("QA"),
        "Đo lường": expr("MCS"),
        "Metrology & Hiệu chuẩn Specialist": expr("MCS"),
        "Internal Auditor (outsource)": expr("IAO"),
        "Chuỗi cung ứng Manager": expr("SCM"),
        "Deburr Team Lead": expr("DBL"),
        "Deburr Technician": expr("DBT"),
        "Finance Manager": expr("FIN"),
        "Finance Costing": expr("FIN"),
        "AP-AR and Payments Accountant": expr("APAR"),
        "Finance AR": expr("APAR"),
        "AR (Finance)": expr("APAR"),
        "General Ledger and Payroll Accountant": expr("GLP"),
        "GL and Payroll Accountant": expr("GLP"),
        "Production Engineer / IE": expr("PIE"),
        "Production Engineer-IE": expr("PIE"),
        "Production Engineer / Industrial Engineer": expr("PIE"),
        "Manufacturing Engineer": expr("PIE"),
        "Internal Auditor / LPA Assignee": expr("QMS[LA]"),
        "Lead Auditor": expr("QMS[LA]"),
        "Continuous Improvement Lead": expr("PIE[CI]"),
        "Product Safety Officer": expr("QA[PSO]"),
        "Incident Commander": expr("PD[IC-PROD]"),
        "All Process Owners / Department Heads": bundle("OPS_SCOPE_OWNERS"),
        "Process Owners / Department Heads": bundle("OPS_SCOPE_OWNERS"),
        "Quy trình Owner / Department Trưởng bộ phận": bundle("OPS_SCOPE_OWNERS"),
        "All Quy trình Owner / Department Trưởng bộ phận": bundle("OPS_SCOPE_OWNERS"),
        "Steering Committee": bundle("DEPLOYMENT_STEERING"),
        "All managers": bundle("DIRECT_LINE_MGRS"),
        "all managers": bundle("DIRECT_LINE_MGRS"),
        "All managers / approvers": bundle("TOP_MGMT"),
        "all managers / approvers": bundle("TOP_MGMT"),
        "department heads": bundle("FUNC_HEADS"),
        "Department heads": bundle("FUNC_HEADS"),
        "Data Owner": bundle("DATA_OWNERS"),
        "Data Owners": bundle("DATA_OWNERS"),
        "Business Owner": bundle("DATA_OWNERS"),
        "System Owner": bundle("SYSTEM_OWNERS"),
        "ngÆ°á»i chá»‹u trÃ¡ch nhiá»‡m dá»¯ liá»‡u": bundle("DATA_OWNERS"),
        "ngÆ°á»i chá»‹u trÃ¡ch nhiá»‡m há»‡ thá»‘ng": bundle("SYSTEM_OWNERS"),
        "chá»§ dá»¯ liá»‡u": bundle("DATA_OWNERS"),
        "DÃ²ng / Chuyá»n Manager": bundle("DIRECT_LINE_MGRS"),
        "process owners": bundle("OPS_SCOPE_OWNERS"),
        "Process owners": bundle("OPS_SCOPE_OWNERS"),
        "Leadership Sponsor": expr("CEO"),
        "QA/QC Lead + Sales Lead": expr("QCL", "CS", "EST"),
        "QA Lead + Production": expr("QA", "PD", joiner=" + "),
        "QC Lead / CMM": expr("QCL", "MCS"),
        "Cleanliness Process Owner": expr("QA", "CPS", joiner=" + "),
        "Production Director / Functional Head": expr("PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA"),
        "Functional Heads / Shift Leaders": expr("PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA", "SL"),
        "CI Lead + Process Owner": expr("PIE[CI]", "PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA"),
        "Sản xuất Supervisor / QA Lead": expr("SL", "QA"),
        "Sản xuất Giám đốc / QA Lead": expr("PD", "QA"),
        "Document Controller + Warehouse": expr("QMS[DC]", "WAR", joiner=" + "),
        "Document Controller hoặc người phụ trách tài liệu": expr("QMS[DC]", "PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA")
    }


def bundle_alias_map() -> dict[str, dict]:
    return {
        "TOP_MGMT": bundle("TOP_MGMT"),
        "FUNC_HEADS": bundle("FUNC_HEADS"),
        "FUNC_OWNERS": bundle("FUNC_OWNERS"),
        "DATA_OWNERS": bundle("DATA_OWNERS"),
        "SYSTEM_OWNERS": bundle("SYSTEM_OWNERS"),
        "DEPLOYMENT_STEERING": bundle("DEPLOYMENT_STEERING"),
        "ALL_DEPTS": bundle("ALL_DEPTS"),
        "SUPPORT_ENABLEMENT": bundle("SUPPORT_ENABLEMENT"),
        "COMMERCIAL_FRONT": bundle("COMMERCIAL_FRONT"),
        "QUALITY_CORE": bundle("QUALITY_CORE"),
        "ENG_RELEASE_CORE": bundle("ENG_RELEASE_CORE"),
        "AREA_LEADS": bundle("AREA_LEADS"),
        "POU_LEADS": bundle("POU_LEADS"),
        "OPS_SCOPE_OWNERS": bundle("OPS_SCOPE_OWNERS"),
        "FRONTLINE_LEADS": bundle("FRONTLINE_LEADS"),
        "DEPLOYMENT_LEADS": bundle("DEPLOYMENT_LEADS"),
        "DIRECT_LINE_MGRS": bundle("DIRECT_LINE_MGRS"),
        "OJT_COACHES": bundle("OJT_COACHES"),
        "KNOWLEDGE_SMES": bundle("KNOWLEDGE_SMES"),
        "MR_REPORT_OWNERS": bundle("MR_REPORT_OWNERS"),
    }


def try_resolve_roleish_text(text: str, aliases: dict[str, dict], overrides: dict[str, dict] | None = None) -> dict | None:
    cleaned = normalize_ws(text)
    resolver = dict(aliases)
    if overrides:
        resolver.update(overrides)
    if cleaned in resolver:
        return deepcopy(resolver[cleaned])
    if len(cleaned) > 120:
        return None
    parts = re.split(r"\s*(\+|/|,|;| hoặc | or | và | and )\s*", cleaned)
    if len(parts) <= 1:
        return None
    joiner = " / "
    tokens: list[str] = []
    for index, part in enumerate(parts):
        token = normalize_ws(part)
        if not token:
            continue
        if index % 2 == 1:
            joiner = " + " if token == "+" else (" · " if token in {",", ";", "và", "and"} else " / ")
            continue
        if token not in resolver:
            return None
        spec = resolver[token]
        if "bundle" in spec:
            tokens.append(spec["bundle"])
        elif "tokens" in spec:
            tokens.extend(spec["tokens"])
        else:
            tokens.extend(spec["codes"])
    return mix(*tokens, joiner=joiner) if tokens else None


def find_department_for_handbook(current_file: Path, registry: dict) -> tuple[str, dict] | None:
    matches = [
        (code, meta)
        for code, meta in registry.get("departments", {}).items()
        if Path(meta["handbook_path"]).name == current_file.name
    ]
    if not matches:
        return None
    for code, meta in matches:
        if meta.get("type") == "department":
            return code, meta
    return matches[0]


def set_meta_row_text(doc: etree._Element, label_tokens: tuple[str, ...], value: str) -> None:
    folded_tokens = [fold_text(token) for token in label_tokens]
    for row in doc.xpath('//div[contains(@class,"meta")]/div[contains(@class,"row")] | //div[contains(@class,"fh-kv")]'):
        spans = row.xpath("./span")
        if len(spans) < 2:
            continue
        label = fold_text(element_text(spans[0]))
        if not any(token in label for token in folded_tokens):
            continue
        value_el = spans[-1]
        for child in list(value_el):
            value_el.remove(child)
        value_el.text = value
        break


def set_meta_row_spec(doc: etree._Element, label_tokens: tuple[str, ...], spec: dict, current_file: Path, registry: dict) -> None:
    rendered = render_spec(spec, current_file, registry)
    folded_tokens = [fold_text(token) for token in label_tokens]
    for row in doc.xpath('//div[contains(@class,"meta")]/div[contains(@class,"row")] | //div[contains(@class,"fh-kv")]'):
        spans = row.xpath("./span")
        if len(spans) < 2:
            continue
        label = fold_text(element_text(spans[0]))
        if not any(token in label for token in folded_tokens):
            continue
        set_element_html(spans[-1], rendered)
        break


def set_require_row_value(
    doc: etree._Element,
    label_tokens: tuple[str, ...],
    value_html: str | None = None,
    value_text: str | None = None,
) -> None:
    folded_tokens = [fold_text(token) for token in label_tokens]
    for row in doc.xpath('//table[contains(@class,"require-table")]//tr'):
        cells = row.xpath("./th|./td")
        if len(cells) < 2:
            continue
        label = fold_text(element_text(cells[0]))
        if not any(token in label for token in folded_tokens):
            continue
        if value_html is not None:
            set_element_html(cells[1], value_html)
        elif value_text is not None:
            for child in list(cells[1]):
                cells[1].remove(child)
            cells[1].text = value_text


def get_require_row_value_text(doc: etree._Element, label_tokens: tuple[str, ...]) -> str:
    folded_tokens = [fold_text(token) for token in label_tokens]
    for row in doc.xpath('//table[contains(@class,"require-table")]//tr'):
        cells = row.xpath("./th|./td")
        if len(cells) < 2:
            continue
        label = fold_text(element_text(cells[0]))
        if not any(token in label for token in folded_tokens):
            continue
        return normalize_ws(element_text(cells[1]))
    return ""


def canonical_department_labels(registry: dict) -> dict[Path, str]:
    labels: dict[Path, str] = {}
    for code, meta in registry.get("departments", {}).items():
        target = (ROOT / meta["handbook_path"]).resolve()
        existing = labels.get(target)
        if existing is None or meta.get("type") == "department":
            labels[target] = code
    return labels


def normalize_organization_shortlinks(doc: etree._Element, current_file: Path) -> None:
    shortcuts = {
        "../02-Department-Handbooks/": ROOT / "02-Tai-Lieu-He-Thong" / "03-Organization" / "02-Department-Handbooks",
        "../03-Job-Descriptions/": ROOT / "02-Tai-Lieu-He-Thong" / "03-Organization" / "03-Job-Descriptions",
        "../03-Organization/02-Department-Handbooks/": ROOT / "02-Tai-Lieu-He-Thong" / "03-Organization" / "02-Department-Handbooks",
        "../03-Organization/03-Job-Descriptions/": ROOT / "02-Tai-Lieu-He-Thong" / "03-Organization" / "03-Job-Descriptions",
    }
    for anchor in doc.xpath('//a[@href]'):
        href = ((anchor.get("href") or "").strip()).replace("\\", "/")
        if href != (anchor.get("href") or "").strip():
            anchor.set("href", href)
        for prefix, target_dir in shortcuts.items():
            if not href.startswith(prefix):
                continue
            tail = href[len(prefix):]
            if not tail or tail.startswith(("#", "http://", "https://", "mailto:", "javascript:")):
                continue
            target = target_dir / tail
            if target.exists():
                anchor.set("href", os.path.relpath(target, current_file.parent).replace("\\", "/"))
            break


def normalize_reference_links(doc: etree._Element, current_file: Path, registry: dict) -> None:
    target_labels: dict[Path, str] = {}
    for code, meta in registry["roles"].items():
        target_labels[(ROOT / meta["jd_path"]).resolve()] = meta["jd_code"]
    target_labels.update(canonical_department_labels(registry))

    for anchor in doc.xpath('//a[@href]'):
        href = ((anchor.get("href") or "").strip()).replace("\\", "/")
        if href != (anchor.get("href") or "").strip():
            anchor.set("href", href)
        classes = (anchor.get("class") or "").split()
        if any(token in classes for token in ["role-link", "dept-link", "entity-link"]):
            continue
        if not href or href.startswith(("#", "http://", "https://", "mailto:", "javascript:")):
            continue
        try:
            target = (current_file.parent / href).resolve()
        except OSError:
            continue
        label = target_labels.get(target)
        if not label:
            continue
        for child in list(anchor):
            anchor.remove(child)
        anchor.text = label


def normalize_local_link_separators(doc: etree._Element) -> None:
    for element in doc.xpath('//*[@href or @src]'):
        for attr in ("href", "src"):
            value = element.get(attr)
            if not value:
                continue
            cleaned = value.replace("\\", "/")
            if cleaned != value:
                element.set(attr, cleaned)


def normalize_jd_structured_rows(
    doc: etree._Element,
    doc_path: Path,
    role_code: str,
    registry: dict,
    profiles: dict[str, dict[str, object]],
) -> None:
    profile = profiles.get(role_code, {})
    if not profile:
        return

    departments = list(profile.get("departments", []))
    reports_to = list(profile.get("reports_to", []))
    direct_reports = list(profile.get("direct_reports", []))
    primary_interfaces = list(profile.get("primary_interfaces", []))
    internal_interfaces = list(profile.get("internal_interfaces", []))
    external_interfaces = str(profile.get("external_interfaces", "")).strip()

    if departments:
        set_require_row_value(doc, ("bo phan",), value_html=render_token_cluster(departments, doc_path, registry))
    if reports_to:
        set_require_row_value(doc, ("bao cao truc tiep cho", "bao cao truc tiep"), value_html=render_token_cluster(reports_to, doc_path, registry))
    elif role_code == "CEO":
        set_require_row_value(
            doc,
            ("bao cao truc tiep cho", "bao cao truc tiep"),
            value_text="Hoi dong thanh vien / Hoi dong quan tri / Owner cong ty (theo mo hinh quan tri thuc te cua doanh nghiep).",
        )

    if direct_reports:
        rendered_direct = render_token_cluster(direct_reports, doc_path, registry)
        set_require_row_value(doc, ("quan ly truc tiep", "nhan su bao cao truc tiep"), value_html=rendered_direct)
    else:
        set_require_row_value(
            doc,
            ("quan ly truc tiep", "nhan su bao cao truc tiep"),
            value_text="Khong co nhan su bao cao truc tiep.",
        )

    if primary_interfaces:
        rendered_primary = render_token_cluster(primary_interfaces, doc_path, registry)
        set_require_row_value(doc, ("phoi hop chuc nang chinh", "tuyen phoi hop chuc nang"), value_html=rendered_primary)
    if internal_interfaces:
        set_require_row_value(doc, ("interfaces noi bo chinh", "giao dien noi bo chinh"), value_html=render_token_cluster(internal_interfaces, doc_path, registry))
    if external_interfaces:
        set_require_row_value(doc, ("interfaces ben ngoai", "giao dien ben ngoai"), value_text=external_interfaces)


def normalize_jd_preface_block(
    doc: etree._Element,
    doc_path: Path,
    role_code: str,
    registry: dict,
    profiles: dict[str, dict[str, object]],
) -> None:
    preface_nodes = doc.xpath('//div[contains(@class,"preface-block")]//p[1]')
    if not preface_nodes:
        return
    profile = profiles.get(role_code, {})
    departments = list(profile.get("departments", []))
    if not departments:
        return
    preface = preface_nodes[0]
    links = preface.xpath('.//a[@href]')
    seen_links: set[tuple[str, str]] = set()
    doc_fragments: list[str] = []
    for link in links:
        key = (normalize_ws("".join(link.itertext())), (link.get("href") or "").strip())
        if key in seen_links:
            continue
        seen_links.add(key)
        doc_fragments.append(html.tostring(link, encoding="unicode", method="html").strip())
    docs_html = " ".join(doc_fragments)
    text = element_text(preface)
    match = re.search(r"Vai trò trong chuỗi giá trị(.*?)(?:Tài liệu liên đới|$)", text)
    value_chain = normalize_ws(match.group(1)) if match else ""
    value_chain = get_require_row_value_text(doc, ("vai tro trong chuoi gia tri",)) or value_chain
    if not value_chain:
        return
    dept_html = render_token_cluster(departments, doc_path, registry)
    new_html = (
        f"<b>Bộ phận</b> {dept_html} &nbsp;&nbsp; "
        f"<b>Vai trò trong chuỗi giá trị</b> {xml_escape(value_chain)} &nbsp;&nbsp; "
        f"<b>Tài liệu liên đới</b> {docs_html}"
    )
    set_element_html(preface, new_html)


def normalize_department_handbook_header(doc: etree._Element, current_file: Path, registry: dict) -> None:
    matched = find_department_for_handbook(current_file, registry)
    if not matched:
        return
    dept_code, dept = matched
    title_nodes = doc.xpath('//div[contains(@class,"title")]/strong')
    if title_nodes:
        title_nodes[0].text = f'{dept["title_en"]} Handbook'
    page_titles = doc.xpath("//title")
    if page_titles:
        page_titles[0].text = f'{dept["title_en"]} Handbook | HESEM QMS'
    set_meta_row_text(doc, ("code", "mã"), dept_code)
    set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr(dept_code), current_file, registry)


def update_meta_rows(doc: etree._Element, current_file: Path, registry: dict, aliases: dict[str, dict], overrides: dict[str, dict]) -> None:
    for row in doc.xpath('//div[contains(@class,"meta")]/div[contains(@class,"row")] | //div[contains(@class,"fh-kv")]'):
        spans = row.xpath("./span")
        if len(spans) < 2:
            continue
        label = fold_text(element_text(spans[0]))
        value_el = spans[-1]
        current_value = element_text(value_el)
        spec = None
        is_approved_label = "approved" in label or "phe duyet" in label or label.endswith("duyet")
        is_owner_label = "owner" in label or "chu so huu" in label or "so huu" in label
        if is_approved_label:
            spec = expr("CEO")
        elif is_owner_label:
            spec = overrides.get(current_value) or aliases.get(current_value) or try_resolve_roleish_text(current_value, aliases, overrides)
        if spec:
            set_element_html(value_el, render_spec(spec, current_file, registry))


def update_role_cells(doc: etree._Element, current_file: Path, registry: dict, aliases: dict[str, dict], overrides: dict[str, dict]) -> None:
    targets = doc.xpath('//th|//td|//span[contains(@class,"inline-tag")]')
    for element in targets:
        if element.xpath(
            './/*[contains(@class,"role-link") or contains(@class,"dept-link") or contains(@class,"entity-link") or contains(@class,"bundle-chip")]'
        ):
            continue
        text = element_text(element)
        if not text or len(text) > 120:
            continue
        spec = overrides.get(text) or aliases.get(text) or try_resolve_roleish_text(text, aliases, overrides)
        if spec:
            set_element_html(element, render_spec(spec, current_file, registry))


def update_plain_titles(doc: etree._Element, mapping: dict[str, str]) -> None:
    for node in doc.xpath('//text()[normalize-space()]'):
        parent = node.getparent()
        if parent is None:
            continue
        class_name = parent.get("class") or ""
        if any(token in class_name for token in ["role-code", "role-link", "dept-code", "dept-link", "entity-code", "entity-link", "bundle-chip"]):
            continue
        value = str(node)
        stripped = normalize_ws(value)
        if stripped in mapping:
            if node.is_text:
                parent.text = mapping[stripped]
            else:
                parent.tail = mapping[stripped]


def replace_text_fragments(doc: etree._Element, replacements: dict[str, str]) -> None:
    for node in doc.xpath('//text()[normalize-space()]'):
        parent = node.getparent()
        if parent is None:
            continue
        class_name = parent.get("class") or ""
        if any(token in class_name for token in ["role-code", "role-link", "dept-code", "dept-link", "entity-code", "entity-link", "bundle-chip"]):
            continue
        value = str(node)
        new_value = value
        for old, new in replacements.items():
            new_value = new_value.replace(old, new)
        if new_value == value:
            continue
        if node.is_text:
            parent.text = new_value
        else:
            parent.tail = new_value


def replace_text_fragments_filtered(
    doc: etree._Element,
    replacements: dict[str, str],
    skip_tags: set[str] | None = None,
) -> None:
    skip_tags = skip_tags or set()
    for node in doc.xpath('//text()[normalize-space()]'):
        parent = node.getparent()
        if parent is None:
            continue
        if any(getattr(ancestor, "tag", None) in skip_tags for ancestor in [parent, *parent.iterancestors()]):
            continue
        class_name = parent.get("class") or ""
        if any(token in class_name for token in ["role-code", "role-link", "dept-code", "dept-link", "entity-code", "entity-link", "bundle-chip"]):
            continue
        value = str(node)
        new_value = value
        for old, new in replacements.items():
            new_value = new_value.replace(old, new)
        if new_value == value:
            continue
        if node.is_text:
            parent.text = new_value
        else:
            parent.tail = new_value


def replace_regex_patterns(
    doc: etree._Element,
    patterns: list[tuple[str, str]],
    skip_tags: set[str] | None = None,
) -> None:
    skip_tags = skip_tags or set()
    for node in doc.xpath('//text()[normalize-space()]'):
        parent = node.getparent()
        if parent is None:
            continue
        if any(getattr(ancestor, "tag", None) in skip_tags for ancestor in [parent, *parent.iterancestors()]):
            continue
        class_name = parent.get("class") or ""
        if any(token in class_name for token in ["role-code", "role-link", "dept-code", "dept-link", "entity-code", "entity-link", "bundle-chip"]):
            continue
        value = str(node)
        new_value = value
        for pattern, replacement in patterns:
            new_value = re.sub(pattern, replacement, new_value)
        if new_value == value:
            continue
        if node.is_text:
            parent.text = new_value
        else:
            parent.tail = new_value


def section_table_rows(doc: etree._Element, section_id: str, table_index: int = 1) -> list[etree._Element]:
    headers = doc.xpath(f'//h2[@id="{section_id}"]')
    if not headers:
        return []
    tables: list[etree._Element] = []
    node = headers[0].getnext()
    while node is not None:
        if isinstance(node.tag, str) and node.tag.lower() == "h2":
            break
        if has_class(node, "table-card"):
            tables.append(node)
        node = node.getnext()
    if 0 < table_index <= len(tables):
        return tables[table_index - 1].xpath(".//tbody/tr")
    return []


def set_section_role_cell(
    doc: etree._Element,
    section_id: str,
    row_index: int,
    spec: dict,
    current_file: Path,
    registry: dict,
    table_index: int = 1,
) -> None:
    rows = section_table_rows(doc, section_id, table_index)
    if 0 <= row_index < len(rows):
        cells = rows[row_index].xpath("./td")
        if cells:
            set_element_html(cells[0], render_spec(spec, current_file, registry))


def set_section_cell_html(
    doc: etree._Element,
    section_id: str,
    row_index: int,
    cell_index: int,
    html_fragment: str,
    table_index: int = 1,
) -> None:
    rows = section_table_rows(doc, section_id, table_index)
    if 0 <= row_index < len(rows):
        cells = rows[row_index].xpath("./td")
        if 0 <= cell_index < len(cells):
            set_element_html(cells[cell_index], html_fragment)


def replace_exact_cell_text(
    doc: etree._Element,
    current_file: Path,
    registry: dict,
    exact_text: str,
    spec: dict,
) -> None:
    rendered = render_spec(spec, current_file, registry)
    for element in doc.xpath('//th|//td'):
        if element.xpath('.//*[contains(@class,"role-link") or contains(@class,"dept-link") or contains(@class,"entity-link")]'):
            continue
        if element_text(element) == exact_text:
            set_element_html(element, rendered)


def replace_exact_block_text(
    doc: etree._Element,
    xpath: str,
    exact_text: str,
    new_html: str,
) -> None:
    for element in doc.xpath(xpath):
        if element_text(element) == exact_text:
            set_element_html(element, new_html)


def set_row_first_cell_html(doc: etree._Element, row_label: str, html_fragment: str) -> None:
    for row in doc.xpath("//tr"):
        cells = row.xpath("./th|./td")
        if not cells:
            continue
        if element_text(cells[0]) == row_label:
            set_element_html(cells[0], html_fragment)


def set_row_cell_html(doc: etree._Element, row_label: str, cell_index: int, html_fragment: str) -> None:
    for row in doc.xpath("//tr"):
        cells = row.xpath("./th|./td")
        if len(cells) <= cell_index:
            continue
        if element_text(cells[0]) == row_label:
            set_element_html(cells[cell_index], html_fragment)


def set_row_cell_by_match(
    doc: etree._Element,
    match_cell_index: int,
    match_text: str,
    target_cell_index: int,
    html_fragment: str,
) -> None:
    for row in doc.xpath("//tr"):
        cells = row.xpath("./th|./td")
        if len(cells) <= max(match_cell_index, target_cell_index):
            continue
        if element_text(cells[match_cell_index]) == match_text:
            set_element_html(cells[target_cell_index], html_fragment)


def normalize_jd_purpose_intro(doc: etree._Element, role_code: str, role: dict, current_filename: str = "") -> None:
    paragraphs = doc.xpath('//div[contains(@class,"jd-purpose")]/p[1]')
    if not paragraphs:
        return
    paragraph = paragraphs[0]
    # Capture only ANNEX/SOP/WI references to preserve. Skip any <a> whose
    # href resolves to the current JD file (those are intro-inlined role
    # anchors injected by the role-code linker — re-appending them causes
    # duplicate self-references to accumulate across repeated runs).
    links: list = []
    seen_hrefs: set = set()
    for anchor in paragraph.xpath("./a"):
        href = (anchor.get("href") or "").strip()
        basename = href.rsplit("/", 1)[-1]
        if current_filename and basename == current_filename:
            continue  # skip self-references
        if href in seen_hrefs:
            continue  # skip exact duplicates
        seen_hrefs.add(href)
        links.append(deepcopy(anchor))
    for child in list(paragraph):
        paragraph.remove(child)
    paragraph.text = (
        f'Vai trò của tài liệu: JD này khóa rõ phạm vi công việc, điểm kết thúc trách nhiệm '
        f'và cơ chế bàn giao của {role["title_en"]} ({role_code}); dùng làm đầu vào chuẩn cho '
        f'tuyển dụng, OJT/đào tạo, đánh giá năng lực, KPI, lương thưởng và phối hợp liên phòng ban '
        f'trong chuỗi RFQ → ship. JD này phải được đọc cùng '
    )
    if not links:
        paragraph.text += "ANNEX/SOP/WI liên quan đang hiệu lực."
        return
    for index, anchor in enumerate(links):
        paragraph.append(anchor)
        anchor.tail = ", " if index < len(links) - 1 else " và SOP/WI/ANNEX đang hiệu lực."


def apply_file_specific_tweaks(doc: etree._Element, current_file: Path, registry: dict) -> None:
    def chips(spec: dict) -> str:
        return render_spec(spec, current_file, registry)

    if current_file.name.startswith("qms-man-001-qms-manual"):
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("CEO", "QA[QMR]"), current_file, registry)

    if current_file.name.startswith("qms-man-001-qms-manual"):
        set_row_first_cell_html(doc, "Top Leader", chips(expr("CEO")))
        set_row_first_cell_html(doc, "Lead Deparment", chips(bundle("FUNC_HEADS")))
        set_row_first_cell_html(doc, "Owner", chips(bundle("FUNC_OWNERS")))
        set_row_first_cell_html(doc, "Support (IT/Epicor/HR/EHS/Finance)", render_token_cluster(["D-IT", "D-ERP", "D-HR", "D-EHS", "D-FIN"], current_file, registry))
        set_row_first_cell_html(doc, "Customer Dịch vụ vs Người định giá", render_token_cluster(["CS", "EST"], current_file, registry, joiner=" vs "))
        set_row_first_cell_html(doc, "DFM vs Quy trình Engineer vs CAM/NC", render_token_cluster(["DFM", "PE", "CAM"], current_file, registry, joiner=" vs "))
        set_row_first_cell_html(doc, "Planner vs IE vs Workshop Manager", render_token_cluster(["PPL", "PIE", "WKM"], current_file, registry, joiner=" vs "))
        set_row_first_cell_html(doc, "Shift Leader vs Setup vs Operator", render_token_cluster(["SL", "SET", "OPR"], current_file, registry, joiner=" vs "))
        set_row_first_cell_html(doc, "Quality vs Sản xuất hold/phát hành", render_token_cluster(["D-QUAL", "D-PROD"], current_file, registry, joiner=" vs "))
        set_row_first_cell_html(doc, "SCM vs Kho vs Hậu cần", render_token_cluster(["D-SCM", "D-WHS", "D-LOG"], current_file, registry, joiner=" vs "))
        set_row_first_cell_html(doc, "IT vs Epicor vs chủ nghiệp vụ", render_token_cluster(["D-IT", "D-ERP", "FUNC_OWNERS"], current_file, registry, joiner=" vs "))

    header_owner_overrides = {
        "wi-202-daily-management-tier-meetings-kpi-and-escalation.html": expr("PD", "QMS"),
        "wi-205-barcode-labeling-and-scan-to-action.html": expr("PPL", "D-WHS", "D-PROD"),
        "wi-206-ship-release-pack-sscc-label-and-pack-reconciliation.html": expr("QA", "D-WHS", "D-LOG", "D-SCS"),
        "wi-207-high-risk-job-readiness-control-tower.html": expr("PPL", "PD", "QA"),
        "wi-302-first-piece-fai-execution-and-evidence-pack.html": expr("QA", "ENGM"),
        "wi-511-machine-type-quick-reference.html": expr("WKM", "ENGM"),
        "wi-512-3-axis-vertical-milling-guide.html": expr("WKM", "ENGM"),
        "wi-513-5-axis-milling-guide.html": expr("WKM", "ENGM"),
        "wi-514-cnc-turning-guide.html": expr("WKM", "ENGM"),
        "wi-515-mill-turn-guide.html": expr("WKM", "ENGM"),
        "wi-516-machine-operation-quick-card.html": expr("WKM", "MNT", "QA"),
        "wi-517-setup-changeover-smed-standard-work.html": expr("WKM", "ENGM", "MNT", "QA"),
        "wi-518-work-transfer-validation.html": expr("WKM", "PPL", "QA"),
        "wi-519-job-packet-quick-check-and-pre-run-verification.html": expr("WKM", "SET", "QA"),
        "wi-602-gage-pre-use-verification-and-status-control.html": expr("QA", "MCS", "MNT"),
        "wi-604-spc-chart-use-process-capability-and-reaction.html": expr("QA", "QE", "WKM"),
        "wi-605-final-inspection-coc-and-shipment-release-handoff.html": expr("QA", "D-SCM"),
        "wi-606-suspect-product-containment-segregation-and-reaction.html": expr("QA", "WKM", "D-WHS"),
        "wi-702-storage-environment-location-and-fifo-control.html": expr("D-WHS", "D-PROD", "QA"),
        "wi-721-fod-prevention-line-clearance-and-tool-accountability.html": expr("WKM", "QA", "D-WHS", "MNT"),
        "annex-101-role-based-access-map.html": expr("ITA", "QMS[DC]"),
        "annex-102-access-request-field-dictionary.html": expr("ITA", "QMS[DC]"),
        "annex-104-org-chart-fullpage.html": expr("CEO", "QMS"),
        "annex-107-audit-evidence-pack-master.html": expr("QMS", "QMS[DC]"),
        "annex-115-epicor-transaction-and-interface-map.html": expr("ESA", "QMS", "FIN"),
        "annex-120-authority-matrix.html": expr("CEO", "QMS", "HR"),
        "annex-122-kpi-cascade-dictionary.html": expr("QMS", "ITA"),
        "annex-124-dashboard-evidence-pack-worked-examples.html": expr("QMS", "ITA"),
        "annex-131-m365-records-metadata-list-schema-and-register-catalog.html": expr("QMS", "ITA"),
        "annex-133-m365-records-site-topology-library-and-folder-blueprint.html": expr("QMS", "ITA"),
        "annex-134-m365-records-provisioning-permissions-and-automation-architecture.html": expr("QMS", "ITA"),
        "annex-135-m365-operational-records-file-plan-by-department-role-and-job.html": expr("QMS", "ITA"),
        "annex-301-setup-sheet-and-tool-list-standard.html": expr("ENGM"),
        "annex-302-approved-materials-list.html": expr("ENGM", "QA"),
        "annex-401-supplier-risk-model-and-scorecard-method.html": expr("SCM", "QA"),
        "annex-402-outsource-special-process-pack.html": expr("SCM", "QA", "ENGM"),
        "annex-403-approved-processor-list.html": expr("SCM", "QA"),
        "annex-502-gate-mrr-and-execution-synchronization-pack.html": expr("PD", "QA"),
        "annex-503-cnc-operating-model-and-role-boundary.html": expr("QA", "PD", "ENGM"),
        "annex-505-put-thru-index.html": expr("QMS", "PPL"),
        "annex-506-fod-prevention-program.html": expr("QMS", "PD", "QA"),
        "annex-602-msa-acceptance-criteria.html": expr("QA", "MCS"),
        "annex-603-quality-package-levels-qpl.html": expr("QA", "D-SCS", "ENGM"),
        "annex-604-control-plan-guide.html": expr("QA", "ENGM"),
        "annex-606-surface-finish-vacuum-compatibility.html": expr("QA", "ENGM"),
        "annex-608-semi-standards-and-csr-matrix.html": expr("ENGM", "QA"),
        "annex-701-gs1-sscc-data-dictionary-and-pack-reconciliation.html": expr("D-WHS", "D-LOG", "ITA", "QMS"),
        "annex-702-packaging-labeling-spec.html": expr("QA", "D-WHS"),
        "annex-703-warehouse-location-fifo-rules.html": expr("QA", "D-WHS"),
        "annex-803-ppe-and-hazard-matrix.html": expr("EHS", "PD"),
    }
    if current_file.name in header_owner_overrides:
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), header_owner_overrides[current_file.name], current_file, registry)

    if current_file.name == "authority-matrix.html":
        set_row_cell_html(doc, "Phát hành kỹ thuật / engineering change", 1, chips(expr("ENGM")))
        set_row_cell_html(doc, "Dispatch / production recovery", 1, chips(expr("PD", "WKM")))
        set_row_cell_html(doc, "Quality hold / final release / stop-ship", 1, chips(expr("QA", "CEO")))
        set_row_cell_html(doc, "Supplier / outsource / inventory exception", 1, f'{chips(expr("SCM"))} + {chips(expr("QA", "ENGM", "FIN"))} theo loại exception')
        set_row_cell_html(doc, "Deputy / backup activation", 1, f'{chips(expr("HR"))} + {chips(bundle("FUNC_HEADS"))}')

    if current_file.name == "raci-master-matrix.html":
        set_row_cell_html(doc, "Operations", 0, chips(expr("D-PROD")))
        set_row_cell_html(doc, "Support", 0, chips(bundle("SUPPORT_ENABLEMENT")))
        set_row_cell_html(doc, "D-ENG", 1, render_token_cluster(["ENGM", "PE", "CAM"], current_file, registry, joiner=" · "))
        set_row_cell_html(doc, "D-QUAL", 1, render_token_cluster(["QA", "QE", "QCL", "QMS"], current_file, registry, joiner=" · "))
        set_row_cell_html(doc, "ANNEX-123", 2, chips(bundle("FUNC_HEADS")))
        set_row_cell_html(doc, "ANNEX-123", 5, chips(bundle("DIRECT_LINE_MGRS")))
        replace_text_fragments(
            doc,
            {
                "Vai trò cluster": "Nhóm actor / bundle",
                "RFQ, order data, customer communications, commercial commitments.": "RFQ, dữ liệu đơn hàng, trao đổi khách hàng và cam kết thương mại.",
                "Technical mức sẵn sàng, kiểm soát thay đổi, bộ hồ sơ phát hành.": "Mức sẵn sàng kỹ thuật, kiểm soát thay đổi và bộ hồ sơ phát hành.",
                "Capacity, dispatch, recovery, execution control.": "Năng lực, điều độ, khôi phục và kiểm soát thực thi.",
                "Containment, release, CAPA, document discipline.": "Ngăn chặn, nhả giữ, CAPA và kỷ luật tài liệu.",
                "Materials, money, people, safety, systems and continuity.": "Vật tư, tài chính, nhân sự, an toàn, hệ thống và tính liên tục.",
                "Ready job pack, material/tool mức sẵn sàng, schedule priority, cổng kiểm soát trạng thái.": "Gói việc sẵn sàng, mức sẵn sàng vật tư/dụng cụ, ưu tiên lịch và trạng thái cổng kiểm soát.",
                "Genealogy, kiểm tra bằng chứng, release trạng thái, pack/ship requirements.": "Phả hệ lô, bằng chứng kiểm tra, trạng thái nhả giữ và yêu cầu đóng gói/giao hàng.",
                "Final release + ship bằng chứng + billing trigger + AR theo dõi tiếp.": "Nhả giữ cuối + bằng chứng giao hàng + điểm kích hoạt lập hóa đơn + theo dõi AR.",
                "Authorized-to-work, access lifecycle, deputy activation, fallback rules.": "Điều kiện được phép làm việc, vòng đời quyền truy cập, kích hoạt phó và quy tắc dự phòng.",
            },
        )

    if current_file.name == "annex-103-org-raci-matrix.html":
        set_row_cell_html(doc, "RFQ / báo giá", 1, chips(expr("CS", "EST", "CEO")))

    if current_file.name == "annex-104-org-chart-fullpage.html":
        replace_text_fragments(
            doc,
            {
                "Department trưởng bộ phận, QA, EHS": "FUNC_HEADS / QA / EHS",
                "HR ↔ Department trưởng bộ phận ↔ QA": "HR ↔ FUNC_HEADS ↔ QA",
                "All kinh doanh chức năng": "ALL_DEPTS",
            },
        )

    if current_file.name == "annex-115-epicor-transaction-and-interface-map.html":
        set_row_cell_html(doc, "Finance / billing packs", 2, chips(expr("FIN", "APAR")))
        set_row_cell_html(doc, "Ph\u00e1t h\u00e0nh k\u1ef9 thu\u1eadt / routing / revision", 2, chips(expr("ENGM", "DFM", "PE", "CAM")))
        set_row_cell_html(doc, "Labor / machine / WIP / completion", 2, f'{chips(expr("PPL"))} / {chips(bundle("FRONTLINE_LEADS"))}')
        set_row_cell_html(doc, "Shipment / x\u00e1c nh\u1eadn giao h\u00e0ng (ship confirm)", 2, f'{chips(expr("D-WHS"))} / {chips(expr("LOG"))} / {chips(expr("QA"))}')
        set_row_cell_html(doc, "Invoice / AR / job close", 2, chips(expr("FIN", "APAR")))
        replace_text_fragments(
            doc,
            {
                "Ng\u01b0\u1eddi ph\u1ee5 tr\u00e1ch Engineering": "ENGM / ENG_RELEASE_CORE",
                "Ng\u01b0\u1eddi ph\u1ee5 tr\u00e1ch Production / PPL": "PPL / FRONTLINE_LEADS",
                "Warehouse / Logistics + QA release th\u1ea9m quy\u1ec1n": "D-WHS / LOG / QA",
                "Ng\u01b0\u1eddi ph\u1ee5 tr\u00e1ch Finance": "FIN / APAR",
            },
        )

    if current_file.name == "annex-117-escalation-matrix-and-sla.html":
        set_row_cell_html(
            doc,
            "Mở shipment / release giao hàng",
            2,
            f'{chips(expr("QA"))} hoặc người được ủy quyền theo ANNEX-120',
        )

    if current_file.name == "annex-118-offline-fallback-kit.html":
        set_row_cell_html(
            doc,
            "L1 — hạn chế cục bộ",
            4,
            f'{chips(bundle("DIRECT_LINE_MGRS"))} sau khi {chips(expr("ITA"))} xác nhận',
        )
        set_row_cell_html(
            doc,
            "L2 — gián đoạn phân hệ",
            4,
            f'{chips(expr("QA", "PD"))} tùy luồng, đồng thời mở event nhật ký',
        )
        set_row_cell_html(
            doc,
            "L3 — gián đoạn diện rộng",
            4,
            f'{chips(expr("CEO", "QA"))} hoặc người được ủy quyền theo ANNEX-120',
        )
        set_row_cell_html(
            doc,
            "Kích hoạt chế độ ngoại tuyến",
            2,
            f'{chips(expr("QA"))} hoặc người được ủy quyền theo ANNEX-120',
        )
        set_row_cell_html(doc, "Customer / supplier communication", 2, chips(expr("CS", "LOG", "SCM")))
        set_row_cell_html(doc, "Job trạng thái / operation moves", 2, chips(expr("PPL", "WKM")))
        set_row_cell_html(doc, "Bảng điều khiển / bộ hồ sơ rà soát", 2, chips(bundle("MR_REPORT_OWNERS")))
        set_row_cell_html(doc, "Bảng điều khiển / bộ hồ sơ rà soát", 3, f'{chips(expr("QMS"))} / {chips(expr("ITA", "ESA"))}')

    if current_file.name == "annex-101-role-based-access-map.html":
        replace_text_fragments(
            doc,
            {
                "người chịu trách nhiệm nghiệp vụ": "FUNC_OWNERS",
                "chủ nghiệp vụ": "FUNC_OWNERS",
                "người chịu trách nhiệm dữ liệu": "role xác nhận nguồn dữ liệu",
                "IT / Epicor quản trị viên": "ITA / ESA",
            },
        )

    if current_file.name == "annex-102-access-request-field-dictionary.html":
        replace_text_fragments(
            doc,
            {
                "người chịu trách nhiệm nghiệp vụ": "FUNC_OWNERS",
                "chủ nghiệp vụ": "FUNC_OWNERS",
                "người chịu trách nhiệm hệ thống": "ITA / ESA",
                "người chịu trách nhiệm dữ liệu": "role xác nhận nguồn dữ liệu",
            },
        )

    if current_file.name == "annex-120-authority-matrix.html":
        replace_text_fragments(
            doc,
            {
                "Engineering Lead/Manager.": "ENGM.",
                "Production Director hoặc Workshop Manager theo phạm vi.": "PD / WKM theo phạm vi.",
                "QA Manager.": "QA.",
                "Supply Chain Manager.": "SCM.",
            },
        )

    if current_file.name == "annex-121-raci-master-matrix.html":
        set_row_cell_html(doc, "ANNEX-123", 2, chips(bundle("FUNC_HEADS")))
        set_row_cell_html(doc, "ANNEX-123", 5, chips(bundle("DIRECT_LINE_MGRS")))
        replace_exact_block_text(doc, '//table[contains(@class,"matrix-table")]//th', "PLA", chips(expr("D-PPC")))
        replace_exact_block_text(doc, '//table[contains(@class,"matrix-table")]//th', "BẢO TRÌ", chips(expr("MNT")))
        replace_text_fragments(
            doc,
            {
                "NHẬT KÝ": "LOG",
                "người chịu trách nhiệm KPI": "MR_REPORT_OWNERS",
                "QMS Engineer / doc điều phối viên": "QMS[DC]",
                "Chủ quá trình": "OPS_SCOPE_OWNERS",
                "chủ quá trình": "OPS_SCOPE_OWNERS",
                "Dept Trưởng bộ phận + Finance": "FUNC_HEADS + FIN",
                "Dashboard người chịu trách nhiệm": "MR_REPORT_OWNERS",
                "HR / dòng / chuyền người chịu trách nhiệm": "HR / DIRECT_LINE_MGRS",
                "QMS + Dept Trưởng bộ phận": "QMS + FUNC_HEADS",
                "HR + all đánh giá viên": "HR + OJT_COACHES",
                "EHS / khu vực người chịu trách nhiệm": "EHS / AREA_LEADS",
                "All site nhân sự": "toàn bộ nhân sự nhà máy",
                "Finance team": "D-FIN",
                "Finance + Quy trình Owner": "FIN + OPS_SCOPE_OWNERS",
                "QMS / IT / chủ dữ liệu": "QMS / ITA / OPS_SCOPE_OWNERS",
                "Finance + HR + Quy trình Owner": "FIN + HR + OPS_SCOPE_OWNERS",
                "All chủ dữ liệu": "OPS_SCOPE_OWNERS",
                "người chịu trách nhiệm chức năng + QMS": "FUNC_OWNERS + QMS",
                "người chịu trách nhiệm chức năng + QMS + CEO": "FUNC_OWNERS + QMS + CEO",
                "chủ tài liệu": "QMS[DC]",
                "All quản lý / approvers": "TOP_MGMT / FUNC_HEADS / QMS",
                "QMS / QMS": "QMS",
                "HR / all managers": "HR / DIRECT_LINE_MGRS",
                "All managers / approvers": "TOP_MGMT",
            },
        )

    if current_file.name == "annex-114-go-live-runbook-and-cutover-control.html":
        set_row_cell_html(doc, "Support", 0, chips(bundle("SUPPORT_ENABLEMENT")))

    if current_file.name == "annex-123-deputy-backup-matrix.html":
        set_row_cell_html(doc, "JD-CPS", 4, "CPS vắng mặt hoặc tuyến clean-pack cần duy trì handoff sang LOG.")
        set_row_cell_html(doc, "QA", 4, "QA vắng mặt hoặc bận xử lý sự cố lớn.")
        set_row_cell_html(doc, "CS", 1, chips(expr("D-SCS")))
        set_row_cell_html(doc, "EST", 1, chips(expr("D-SCS")))
        set_row_cell_html(doc, "1. Nhận diện kích hoạt", 2, f'{chips(bundle("DIRECT_LINE_MGRS"))} / {chips(expr("HR", "CEO"))} tùy mức')
        set_row_cell_html(doc, "2. Chỉ định deputy", 2, f'{chips(bundle("FUNC_HEADS"))} + {chips(expr("HR"))}')
        set_row_cell_html(doc, "3. Bàn giao gói việc", 2, f'{chips(bundle("FUNC_OWNERS"))} hoặc {chips(bundle("DIRECT_LINE_MGRS"))}')
        set_row_cell_html(doc, "4. Theo dõi coverage", 2, chips(bundle("FUNC_HEADS")))
        set_row_cell_html(doc, "5. Trả lại Người phụ trách / tái phân công", 2, f'{chips(bundle("FUNC_HEADS"))} + {chips(expr("HR", "CEO"))} khi cần')
        set_row_cell_html(doc, "Sales & CS", 0, chips(expr("D-SCS")))
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"preface-block")]//div',
            "All 38 Job mô tả HR Department Trưởng bộ phận Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên ITA ESA",
            (
                '<span class="inline-tag">All JD</span>'
                f'<span class="inline-tag">{chips(expr("HR"))}</span>'
                f'<span class="inline-tag">{chips(bundle("FUNC_HEADS"))}</span>'
                f'<span class="inline-tag">{render_token_cluster(["TOP_MGMT", "DIRECT_LINE_MGRS"], current_file, registry)}</span>'
                f'<span class="inline-tag">{chips(expr("ITA"))}</span>'
                f'<span class="inline-tag">{chips(expr("ESA"))}</span>'
            ),
        )

    if current_file.name == "annex-124-dashboard-evidence-pack-worked-examples.html":
        replace_text_fragments(
            doc,
            {
                "FRM-653  với Người phụ trách Planning + Supply Chain, sponsor Production Director.": "FRM-653 với PPL + SCM, sponsor PD.",
                "Giúp người chịu trách nhiệm KPI, QMS và IT/Digital thống nhất định dạng pack, siêu dữ liệu, đóng băng-date, ngoại lệ-note và bằng chứng path.": "Giúp MR_REPORT_OWNERS, QMS và ITA / ESA thống nhất định dạng pack, siêu dữ liệu, đóng băng-date, ngoại lệ-note và đường dẫn bằng chứng.",
                "người chịu trách nhiệm KPI + QMS": "MR_REPORT_OWNERS + QMS",
                "Lãnh đạo / Quy trình Owner": "TOP_MGMT / OPS_SCOPE_OWNERS",
                "người chịu trách nhiệm Bảng điều khiển + chủ dữ liệu": "MR_REPORT_OWNERS + FUNC_OWNERS / ITA / ESA",
                "IT admin + người chịu trách nhiệm hệ thống + Reviewer độc lập": "ITA + ESA + QMS",
                "Epicor admin / Finance manager / QA manager": "ESA / FIN / QA",
                "người chịu trách nhiệm Projects": "FUNC_OWNERS",
                "người chịu trách nhiệm Kế hoạch + Chuỗi cung ứng, Người bảo trợ Sản xuất Giám đốc.": "PPL + SCM, sponsor PD.",
                "người chịu trách nhiệm KPI": "MR_REPORT_OWNERS",
                "người chịu trách nhiệm hệ thống": "ESA / ITA",
                "chủ dữ liệu": "FUNC_OWNERS / ITA / ESA",
            },
        )

        set_row_cell_html(doc, "Freeze-date pack", 2, f'{chips(bundle("MR_REPORT_OWNERS"))} + {chips(expr("QMS"))}')
        set_row_cell_html(doc, "Freeze-date pack", 3, f'{chips(bundle("TOP_MGMT"))} / {chips(bundle("OPS_SCOPE_OWNERS"))}')
        set_row_cell_html(doc, "Exception-note pack", 2, f'{chips(bundle("MR_REPORT_OWNERS"))} + {chips(bundle("FUNC_OWNERS"))} / {chips(bundle("SYSTEM_OWNERS"))}')
        set_row_cell_html(doc, "Exception-note pack", 3, f'{chips(bundle("MR_REPORT_OWNERS"))} / {chips(expr("QMS"))} / {chips(bundle("FUNC_HEADS"))}')
        set_row_cell_html(doc, "Quarterly access-bộ hồ sơ rà soát", 2, f'{chips(bundle("SYSTEM_OWNERS"))} / {chips(expr("QMS"))}')
        set_row_cell_html(doc, "Quarterly access-bộ hồ sơ rà soát", 3, f'{chips(bundle("SYSTEM_OWNERS"))} / {chips(expr("QMS"))} / {chips(bundle("TOP_MGMT"))}')
        set_row_cell_html(doc, "Management-bộ hồ sơ rà soát", 2, f'{chips(expr("QMS"))} + {chips(bundle("MR_REPORT_OWNERS"))}')
        set_row_cell_html(doc, "Management-bộ hồ sơ rà soát", 3, chips(bundle("TOP_MGMT")))

    if current_file.name == "annex-110-dashboard-kpi-dictionary-and-data-model.html":
        set_row_cell_html(doc, "T1/T2 board trong ng\u00e0y", 1, f'{chips(expr("PPL"))} / {chips(bundle("FRONTLINE_LEADS"))} / {chips(expr("QA"))}')
        set_row_cell_html(doc, "T3 tu\u1ea7n / th\u00e1ng", 1, f'{chips(bundle("MR_REPORT_OWNERS"))} + {chips(bundle("OPS_SCOPE_OWNERS"))}')
        set_row_cell_html(doc, "T4 / xem x\u00e9t c\u1ee7a l\u00e3nh \u0111\u1ea1o", 1, f'{chips(expr("QMS"))} / {chips(bundle("MR_REPORT_OWNERS"))} / {chips(bundle("OPS_SCOPE_OWNERS"))}')
        set_row_cell_html(doc, "Customer / external disclosure pack", 1, f'{chips(bundle("COMMERCIAL_FRONT"))} / {chips(bundle("OPS_SCOPE_OWNERS"))} / {chips(bundle("FUNC_HEADS"))}')

    if current_file.name == "annex-403-approved-processor-list.html":
        replace_text_fragments(
            doc,
            {
                "Chỉ dùng khi Purchasing Manager + QA Manager phê duyệt cho từng job; phải có containment plan và review sau từng lô.": "Chỉ dùng khi BUY + QA phê duyệt cho từng job; phải có containment plan và review sau từng lô.",
                "Chỉ dùng theo temporary deviation có thời hạn, phê duyệt bởi QA Manager + Operations Head + Tổng Giám Đốc; phải xác định containment 100%.": "Chỉ dùng theo temporary deviation có thời hạn, phê duyệt bởi QA + PD + CEO; phải xác định containment 100%.",
            },
        )

    if current_file.name == "annex-502-gate-mrr-and-execution-synchronization-pack.html":
        set_row_cell_html(doc, "Program ID â‰  Phát hành chương trình (program phát hành) list", 3, chips(expr("WKM", "SL", "D-ENG", "QA")))

    if current_file.name == "annex-503-cnc-operating-model-and-role-boundary.html":
        set_row_first_cell_html(doc, "Foreman / Workshop Mgr", chips(expr("WKM")))
        set_row_cell_html(doc, "1. RFQ intake & sàng lọc hợp đồng", 3, f'{chips(expr("CS", "CEO"))} theo ANNEX-120')
        set_row_cell_html(doc, "1. RFQ intake & sàng lọc hợp đồng", 4, f'{chips(expr("CS"))} / {chips(bundle("FUNC_OWNERS"))} khi dữ liệu thiếu')
        set_row_cell_html(doc, "2. Technical feasibility & cost build", 2, chips(expr("ENGM", "PE", "CAM", "QE", "SCM")))
        set_row_cell_html(doc, "2. Technical feasibility & cost build", 3, f'{chips(expr("ENGM", "CEO"))} theo mức NRE & risk')
        set_row_cell_html(doc, "3. Báo giá release / order commitment", 2, chips(expr("CEO", "ENGM", "SCM")))
        set_row_cell_html(doc, "4. Order conversion & dữ liệu gốc", 3, f'{chips(expr("CS", "ESA"))} theo quy trình')
        set_row_cell_html(doc, "5. Quy trình planning & phát hành kỹ thuật", 2, chips(expr("ENGM", "DFM", "QE", "MCS")))
        set_row_cell_html(doc, "7. Dispatch & cell loading", 3, f'{chips(expr("PD", "PPL"))} theo delegated rule')
        set_row_cell_html(doc, "8. Setup & first-off approval", 2, chips(expr("PE", "CAM", "QCL", "QE")))
        set_row_cell_html(doc, "8. Setup & first-off approval", 3, f'{chips(expr("WKM", "QA"))} theo rule first-off')
        set_row_cell_html(doc, "8. Setup & first-off approval", 4, chips(expr("SET", "SL", "QA", "MCS")))
        set_row_cell_html(doc, "9. Serial / lệnh sản xuất execution", 2, chips(expr("OPR", "WKM", "QE", "QC")))
        set_row_cell_html(doc, "10. Kiểm tra / metrology / final release", 2, chips(expr("QC", "QE", "MCS")))
        set_row_cell_html(doc, "11. Packing / shipping / export", 1, chips(expr("CPS", "LOG")))
        set_row_cell_html(doc, "11. Packing / shipping / export", 3, f'{chips(expr("SCM", "QA"))} theo rule release')
        set_row_cell_html(doc, "12. Khiếu nại / NCR / CAPA / learning", 3, f'{chips(expr("QA", "CEO"))} theo mức ảnh hưởng')
        set_row_cell_html(doc, "kiểm tra cuối / phê duyệt giao hàng (phê duyệt giao hàng) chưa đủ bằng chứng", 2, f'{chips(expr("QA", "CEO"))} theo khách hàng rủi ro')

    if current_file.name == "annex-603-quality-package-levels-qpl.html":
        set_row_cell_html(doc, "Hạ QPL", 2, f'{chips(expr("QA", "ENGM"))} + {chips(expr("D-SCS"))} / khách hàng nếu yêu cầu khách hàng chi phối')

    if current_file.name == "annex-105-process-map-detailed.html":
        replace_text_fragments(
            doc,
            {
                "S05 IT/Data": "S05 D-IT / D-ERP",
            },
        )

    if current_file.name == "annex-115-epicor-transaction-and-interface-map.html":
        set_row_cell_html(doc, "Finance / billing packs", 2, chips(expr("FIN", "APAR")))
        replace_text_fragments(
            doc,
            {
                "NgÆ°á»i phá»¥ trÃ¡ch Finance": "FIN / APAR",
            },
        )

    if current_file.name == "annex-115-epicor-transaction-and-interface-map.html":
        set_row_cell_html(doc, "Ph\u00e1t h\u00e0nh k\u1ef9 thu\u1eadt / routing / revision", 2, chips(expr("ENGM", "DFM", "PE", "CAM")))
        set_row_cell_html(doc, "Labor / machine / WIP / completion", 2, f'{chips(expr("PPL"))} / {chips(bundle("FRONTLINE_LEADS"))}')
        set_row_cell_html(doc, "Shipment / x\u00e1c nh\u1eadn giao h\u00e0ng (ship confirm)", 2, f'{chips(expr("D-WHS"))} / {chips(expr("LOG"))} / {chips(expr("QA"))}')
        set_row_cell_html(doc, "Invoice / AR / job close", 2, chips(expr("FIN", "APAR")))
        replace_text_fragments(
            doc,
            {
                "Người phụ trách Engineering": "ENGM / ENG_RELEASE_CORE",
                "Người phụ trách Production / PPL": "PPL / FRONTLINE_LEADS",
                "Warehouse / Logistics + QA release thẩm quyền": "D-WHS / LOG / QA",
                "Người phụ trách Finance": "FIN / APAR",
            },
        )

    if current_file.name == "wi-103-m365-folder-routing-training-competence-and-adoption-for-cnc-job-orders.html":
        set_row_cell_html(doc, "2. Chọn champion theo phòng ban và shift", 1, chips(bundle("DEPLOYMENT_LEADS")))

    if current_file.name in {"C01-L4.html", "C02-L4.html", "C03-L4.html", "C04-L4.html", "C05-L4.html"}:
        replace_text_fragments(
            doc,
            {
                "Foreman/Department Manager/": "DIRECT_LINE_MGRS / ",
                "/QA lead": " / QUALITY_CORE",
            },
        )

    if current_file.name in {"C01-L3.html", "C02-L3.html", "C04-L3.html"}:
        replace_text_fragments(
            doc,
            {
                "Team Leader/lead/nhân sự chủ chốt": "FRONTLINE_LEADS / KNOWLEDGE_SMES",
            },
        )

    if current_file.name == "C01-L2.html":
        replace_exact_block_text(
            doc,
            "//td",
            "01 JSA đã phê duyệt bởi Team Leader/Foreman.",
            f'01 JSA đã được phê duyệt bởi {chips(bundle("FRONTLINE_LEADS"))}.',
        )

    if current_file.name == "C01-L3.html":
        replace_text_fragments(
            doc,
            {
                "Team Leader – Foreman – QA/HSE": "FRONTLINE_LEADS – QA / EHS",
                "Team Leader/Foreman/QA": "FRONTLINE_LEADS / QA",
            },
        )

    if current_file.name == "C04-L4.html":
        replace_text_fragments(
            doc,
            {
                "Team Leader/Foreman": "FRONTLINE_LEADS",
                "mặt bằng xưởng, Foreman, Quản lý Bảng.": "mặt bằng xưởng, FRONTLINE_LEADS, Quản lý Bảng.",
            },
        )

    if current_file.name == "C01-L4.html":
        replace_text_fragments(
            doc,
            {
                "Team Leader hằng ngày, Foreman hằng tuần, QA/HSE hằng tháng;": "FRONTLINE_LEADS hằng ngày, WKM hằng tuần, QA / EHS hằng tháng;",
            },
        )

    if current_file.name == "assessment-matrix.html":
        replace_text_fragments(
            doc,
            {
                "Báo ngay cho Team Leader khi thấy nguy cơ": "Báo ngay cho FRONTLINE_LEADS khi thấy nguy cơ",
                "báo QC/Team Leader": "báo QC / FRONTLINE_LEADS",
                "báo cáo vượt cấp cho Team Leader/quản lý": "báo cáo vượt cấp cho FRONTLINE_LEADS / DIRECT_LINE_MGRS",
            },
        )

    if current_file.name == "SYS-OPS-08.html":
        replace_text_fragments(
            doc,
            {
                "/Mua h\u00e0ng:": "/BUY:",
                "IT/BI:": "SYSTEM_OWNERS:",
            },
        )

    if current_file.name == "SYS-OPS-12.html":
        replace_text_fragments(
            doc,
            {
                "QA / QMS + IT/Data": "QA / QMS + SYSTEM_OWNERS",
            },
        )

    if current_file.name == "SYS-OPS-22.html":
        replace_text_fragments(
            doc,
            {
                "QA/IT b\u1eaft bu\u1ed9c \u0111\u1ea1t L3\u2013L4": "QA / SYSTEM_OWNERS b\u1eaft bu\u1ed9c \u0111\u1ea1t L3\u2013L4",
            },
        )

    if current_file.name == "SYS-OPS-28.html":
        replace_text_fragments(
            doc,
            {
                "IT/Data + FRONTLINE_LEADS": "SYSTEM_OWNERS + FRONTLINE_LEADS",
                "FRONTLINE_LEADS/QA/IT": "FRONTLINE_LEADS / QA / SYSTEM_OWNERS",
                "IT (khi li\u00ean quan quy\u1ec1n/h\u1ec7 th\u1ed1ng)": "SYSTEM_OWNERS (khi li\u00ean quan quy\u1ec1n/h\u1ec7 th\u1ed1ng)",
                "Mua h\u00e0ng (khi li\u00ean quan outsource/cert)": "BUY (khi li\u00ean quan thu\u00ea ngo\u00e0i/ch\u1ee9ng ch\u1ec9)",
                "EST/Mua h\u00e0ng:": "EST / BUY:",
                "IT/Data:": "SYSTEM_OWNERS:",
            },
        )

    if current_file.name == "C01-L1.html":
        replace_text_fragments(
            doc,
            {
                "Báo Team Leader/trưởng nhóm": "Báo FRONTLINE_LEADS",
            },
        )

    if current_file.name == "C04-L1.html":
        replace_text_fragments(
            doc,
            {
                "báo Team Leader.": "báo FRONTLINE_LEADS.",
            },
        )

    if current_file.name == "C08-L3.html":
        replace_text_fragments(
            doc,
            {
                "QA/QC Team Leader, Planner lead, Supervisor, Engineer, Người định giá (SAL-03) lead": "QUALITY_CORE / PPL / FRONTLINE_LEADS / D-ENG / EST",
            },
        )

    if current_file.name == "C09-L2.html":
        replace_text_fragments(
            doc,
            {
                "Operator cao cấp, Setup trưởng nhóm, QC cao cấp, Team Leader kho": "OPR / SET / QC / D-WHS",
                "Supervisor": "FRONTLINE_LEADS",
                "người chịu trách nhiệm Cell": "FRONTLINE_LEADS",
            },
        )

    if current_file.name == "C09-L3.html":
        replace_text_fragments(
            doc,
            {
                "Supervisor, Trưởng nhóm Điều độ, Team Leader": "FRONTLINE_LEADS / PPL",
                "Nhân viên điều phối/Supervisor": "PPL / FRONTLINE_LEADS",
                "Supervisor/Ops Mgr": "FRONTLINE_LEADS / PD",
                "Team Leader": "FRONTLINE_LEADS",
                "Supervisor": "FRONTLINE_LEADS",
            },
        )

    if current_file.name == "C10-L2.html":
        replace_text_fragments(
            doc,
            {
                "Trưởng nhóm/Team Leader, Kế hoạch/Nhân viên điều phối, Mua hàng theo dõi tiếp, Team Leader kho / IQC, QC Team Leader, CS điều phối viên": "FRONTLINE_LEADS / PPL / D-PUR / D-WHS / QCL / CS",
                "Team Leader": "FRONTLINE_LEADS",
                "Planner/Trưởng nhóm": "PPL / FRONTLINE_LEADS",
                "Operator/Trưởng nhóm": "OPR / FRONTLINE_LEADS",
                "QC/Trưởng nhóm": "QC / FRONTLINE_LEADS",
                "All vai trò": "Mọi vai trò liên quan",
                "Performer": "Người thực hiện",
            },
        )

    if current_file.name == "C10-L3.html":
        replace_text_fragments(
            doc,
            {
                "Kế hoạch Team Leader, Sản xuất supervisor, Trưởng nhóm QA, Kỹ thuật lead, Người định giá (SAL-03) lead": "PPL / WKM / SL / QA / QCL / ENGM / EST",
                "Trưởng nhóm QA": "QA / QCL",
                "báo Supervisor + QA": "báo FRONTLINE_LEADS + QA",
            },
        )

    if current_file.name == "C13-L2.html":
        replace_text_fragments(
            doc,
            {
                "Team Leader, QC Team Leader, Engineer, Planner, Mua hàng theo dõi tiếp": "FRONTLINE_LEADS / QCL / D-ENG / PPL / D-PUR",
            },
        )

    if current_file.name in {"C19-L1.html", "C19-L2.html", "C19-L3.html", "C19-L4.html"}:
        replace_text_fragments(
            doc,
            {
                "Team Leader/ca trưởng/dòng / chuyền trưởng nhóm": "FRONTLINE_LEADS",
            },
        )

    if current_file.name == "SYS-OPS-05.html":
        replace_text_fragments(
            doc,
            {
                "Team Leader kho/QA": "D-WHS / QA",
                "Supervisor/OPS Lead": "FRONTLINE_LEADS / PPL",
                "Supervisor + QA + WHS": "FRONTLINE_LEADS + QA + D-WHS",
                "Team Leader kho + QA": "D-WHS + QA",
                "Chỉ Team Leader kho hoặc QA thực hiện": "Chỉ D-WHS hoặc QA thực hiện",
                "Team Leader kho + Supervisor + QA": "D-WHS + FRONTLINE_LEADS + QA",
                "Team Leader kho": "D-WHS",
                "All vai trò": "Mọi vai trò liên quan",
                "Supervisor": "FRONTLINE_LEADS",
            },
        )

    if current_file.name == "SYS-OPS-12.html":
        replace_text_fragments(
            doc,
            {
                "Operator + Supervisor": "OPR + FRONTLINE_LEADS",
                "Supervisor/Trưởng nhóm": "FRONTLINE_LEADS",
                "Supervisor + QA": "FRONTLINE_LEADS + QA",
                "QA / QMS + Supervisor": "QA / QMS + FRONTLINE_LEADS",
                "Supervisor/QC": "FRONTLINE_LEADS / QC",
                "báo Supervisor và QC": "báo FRONTLINE_LEADS và QC",
                "Supervisor ca đêm": "FRONTLINE_LEADS ca đêm",
                "D. Supervisor": "D. FRONTLINE_LEADS",
                "HR + Supervisor": "HR + FRONTLINE_LEADS",
                "Operator/Setup/QC/Supervisor/QA / QMS": "OPR / SET / QC / FRONTLINE_LEADS / QA / QMS",
                "Supervisor": "FRONTLINE_LEADS",
            },
        )

    if current_file.name == "SYS-OPS-13.html":
        replace_text_fragments(
            doc,
            {
                "Operator, QC/QA, Planner, Kinh doanh, Sản xuất Trưởng nhóm, Supervisor, Manager/Trưởng / Đầu.": "OPR, QC / QA, PPL, D-SCS, FRONTLINE_LEADS, FUNC_HEADS.",
                "Sản xuất Trưởng nhóm / Supervisor": "FRONTLINE_LEADS",
                "QMR/Planner/Trưởng nhóm": "QA[QMR] / PPL / FRONTLINE_LEADS",
                "Supervisor": "FRONTLINE_LEADS",
            },
        )

    if current_file.name == "SYS-OPS-17.html":
        replace_text_fragments(
            doc,
            {
                "báo Supervisor+QC": "báo FRONTLINE_LEADS + QC",
                "QA+Supervisor+Người định giá (SAL-03)": "QA + FRONTLINE_LEADS + EST",
                "Supervisor/Lập trình viên (CNC)": "FRONTLINE_LEADS / CAM",
                "QA+Supervisor+Planner": "QA + FRONTLINE_LEADS + PPL",
                "QA+Supervisor": "QA + FRONTLINE_LEADS",
                "Supervisor kích hoạt escalation": "FRONTLINE_LEADS kích hoạt escalation",
                "Supervisor L4": "FRONTLINE_LEADS L4",
                "Supervisor escalation": "FRONTLINE_LEADS escalation",
                "Supervisor": "FRONTLINE_LEADS",
            },
        )

    if current_file.name == "authorization-library.html":
        replace_text_fragments(
            doc,
            {
                "phó Team Leader kho.": "phó đầu mối D-WHS.",
            },
        )

    if current_file.name == "competency-framework.html":
        replace_exact_block_text(
            doc,
            "//td",
            "Department Manager / Người hướng dẫn Level 3–4 + HR (quản hồ sơ) + QA (khi liên quan chất lượng)",
            f'{chips(bundle("DIRECT_LINE_MGRS"))} / {chips(bundle("OJT_COACHES"))} / {chips(expr("HR"))} / {chips(expr("QA"))} khi liên quan chất lượng',
        )

    if current_file.name == "C19.html":
        replace_exact_block_text(
            doc,
            "//li",
            "Đối tượng: Team Leader/Trưởng nhóm, Department Manager, HR đào tạo",
            f'<strong>Đối tượng:</strong> {chips(bundle("DIRECT_LINE_MGRS"))} / {chips(expr("HR"))}',
        )

    if current_file.name == "annex-114-go-live-runbook-and-cutover-control.html":
        replace_text_fragments(
            doc,
            {
                "support tăng cường": "SUPPORT_ENABLEMENT tăng cường",
            },
        )
        replace_exact_block_text(
            doc,
            "//span",
            "Support tăng cường",
            f'{chips(bundle("SUPPORT_ENABLEMENT"))} tăng cường',
        )

    if current_file.name == "SYS-OPS-07.html":
        replace_text_fragments(
            doc,
            {
                "QC/QA lead:": "QA / QCL:",
            },
        )

    if current_file.name == "annex-802-collective-bargaining-agreement.html":
        replace_text_fragments(
            doc,
            {
                "HR/Department Manager": "HR / DIRECT_LINE_MGRS",
            },
        )

    if current_file.name == "wi-517-setup-changeover-smed-standard-work.html":
        set_row_cell_html(doc, "nhận dạng", 2, chips(expr("SET")))
        set_row_cell_html(doc, "Program", 2, f'{chips(expr("SET"))} + {chips(expr("PE", "ENGM"))} khi job rủi ro')
        set_row_cell_html(doc, "WCS / mốc chuẩn (datum)", 2, chips(expr("SET")))
        set_row_cell_html(doc, "Tool trạng thái", 2, f'{chips(expr("SET"))} + {chips(expr("TOOL"))} khi có')
        set_row_cell_html(doc, "Fixture trạng thái", 2, chips(expr("SET")))
        set_row_cell_html(doc, "Rủi ro còn mở", 2, f'{chips(expr("SET"))} + {chips(expr("SL", "WKM"))}')
        set_row_cell_html(doc, "Tool gãy, mảnh dao (insert) sứt, bù trừ (offset) điều chỉnh vì trôi / sai lệch CTQ", 2, f'{chips(expr("SL", "WKM"))}; {chips(expr("QA"))} cùng ký nếu ảnh hưởng CTQ')
        set_row_cell_html(doc, "cảnh báo va chạm nhẹ / sự cố suýt xảy ra (Cần / suýt-thiếu / bỏ sót)", 2, f'{chips(expr("SL", "WKM"))} + {chips(expr("QA"))}')
        set_row_cell_html(doc, "Power tổn thất / bộ điều khiển / kiểm soát viên khởi động lại / mất data", 2, chips(expr("SL", "WKM")))
        set_row_cell_html(doc, "Dừng dài, đổi ca, đổi Người setup", 2, chips(expr("SL", "WKM")))

    if current_file.name == "wi-605-final-inspection-coc-and-shipment-release-handoff.html":
        set_row_cell_html(doc, "Re-mở sau khi đã ký eoe", 2, f'{chips(expr("QA"))} theo ANNEX-120')
        replace_text_fragments(
            doc,
            {
                "Thông báo QA Manager, Kho Manager, Kế hoạch và Kinh doanh/eS nếu ảnh hưởng cam kết giao hàng.": "Thông báo QA, D-WHS, PPL và D-SCS nếu ảnh hưởng cam kết giao hàng.",
            },
        )

    if current_file.name == "wi-721-fod-prevention-line-clearance-and-tool-accountability.html":
        set_row_cell_html(doc, "Thiếu mảnh dao (insert)/vít/miếng chêm (shim)/mảnh dao", 2, chips(expr("WKM", "SL", "QA")))
        set_row_cell_html(doc, "FOD có khả năng thoát / bỏ qua tới khách hàng", 2, f'{chips(expr("QA"))} theo SOP liên quan')

    if current_file.name == "wi-901-performance-dashboard.html":
        set_row_cell_html(doc, "Daily / shift operational board", 2, render_token_cluster(["D-PROD", "QA", "FRONTLINE_LEADS"], current_file, registry))
        set_row_cell_html(doc, "Weekly functional bộ hồ sơ rà soát", 2, chips(bundle("FUNC_HEADS")))
        set_row_cell_html(doc, "Monthly / quarterly management-bộ hồ sơ rà soát", 2, render_token_cluster(["TOP_MGMT", "MR_REPORT_OWNERS", "QMS"], current_file, registry))
        set_row_cell_html(doc, "Extraordinary bộ hồ sơ rà soát", 2, render_token_cluster(["TOP_MGMT", "FUNC_HEADS", "QMS"], current_file, registry))
        replace_text_fragments(
            doc,
            {
                "WI này hướng dẫn người chịu trách nhiệm KPI, chủ dữ liệu và QMS cách chuẩn bị bảng điều khiển / pack dữ liệu trước các nhịp điều hành và xem xét của lãnh đạo.": "WI này hướng dẫn MR_REPORT_OWNERS, FUNC_OWNERS / ITA / ESA và QMS cách chuẩn bị bảng điều khiển / pack dữ liệu trước các nhịp điều hành và xem xét của lãnh đạo.",
                "KPI người chịu trách nhiệm / QMS / lãnh đạo": "MR_REPORT_OWNERS / QMS / TOP_MGMT",
                "KPI định nghĩa và người chịu trách nhiệm": "KPI định nghĩa và MR_REPORT_OWNERS",
                "Mỗi bảng điều khiển phải có người chịu trách nhiệm": "Mỗi bảng điều khiển phải có role chủ trì pack được chỉ định",
                "Không có người chịu trách nhiệm thì không có Người chịu trách nhiệm giải trình.": "Không có role chủ trì pack thì không có tuyến giải trình.",
                "KPI / người chịu trách nhiệm": "KPI / MR_REPORT_OWNERS",
                "người chịu trách nhiệm KPI": "MR_REPORT_OWNERS",
                "Người phụ trách KPI": "MR_REPORT_OWNERS",
                "chủ dữ liệu": "FUNC_OWNERS / ITA / ESA",
                "IT/chủ dữ liệu": "ITA / ESA + FUNC_OWNERS",
                "IT / chủ dữ liệu": "ITA / ESA + FUNC_OWNERS",
            },
        )

    if current_file.name == "sop-101-document-and-data-control.html":
        set_section_role_cell(doc, "p4", 0, bundle("FUNC_OWNERS"), current_file, registry)
        set_section_role_cell(doc, "p4", 3, bundle("FUNC_OWNERS"), current_file, registry)
        set_section_role_cell(doc, "p4", 5, bundle("POU_LEADS"), current_file, registry)
        set_section_cell_html(doc, "p8", 2, 2, chips(expr("WKM", "SL", "QCL")))
        set_section_cell_html(
            doc,
            "p9",
            3,
            3,
            chips(expr("CS", "EST", "PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA", "WKM", "SL", "QCL")),
        )
        set_section_cell_html(doc, "p9", 5, 3, chips(expr("QMS", "WKM", "SL", "QCL", "SCM")))
        replace_text_fragments(
            doc,
            {
                "JD cá»§a QMS Engineer, QA Lead, IT Administrator vÃ Governance viÃªn há»‡ thá»‘ng Epicor":
                "JD-QMS, JD-QA, JD-ITA vÃ JD-ESA",
                "QMS Engineer và QA Lead": "QMS và QA",
            },
        )

    if current_file.name == "sop-102-quality-policy-objectives-and-organizational-context.html":
        set_section_role_cell(doc, "p4", 3, bundle("FUNC_OWNERS"), current_file, registry)
        set_section_role_cell(doc, "p4", 4, bundle("DEPLOYMENT_LEADS"), current_file, registry)
        set_section_cell_html(doc, "p8", 3, 2, chips(bundle("OPS_SCOPE_OWNERS")))
        set_section_cell_html(doc, "p9", 0, 1, chips(expr("QA", "QMS")), table_index=2)
        set_section_cell_html(doc, "p9", 1, 1, chips(expr("QA", "QMS", "CS", "EST", "PPL", "PD", "ENGM", "SCM", "FIN", "HR", "EHS", "ITA", "WKM", "SL", "QCL", "LOG")), table_index=2)
        set_section_cell_html(doc, "p9", 2, 1, chips(bundle("OPS_SCOPE_OWNERS")), table_index=2)
        set_section_cell_html(doc, "p9", 3, 1, chips(bundle("OPS_SCOPE_OWNERS")), table_index=2)
        replace_text_fragments(
            doc,
            {
                "JD của các trưởng / đầu Department": "JD của các vai trò trong OPS_SCOPE_OWNERS",
            },
        )
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: CEO giữ A cho policy và mục tiêu cấp công ty; QA Lead giữ A cho chất lượng logic điều hành; QMS Engineer giữ R cho register và evidence; các Trưởng bộ phận giữ R cho phân tầng, đo lường và phản ứng trong phạm vi mình sở hữu.",
            "<b>RACI nền:</b> CEO giữ A cho policy và mục tiêu cấp công ty; QA giữ A cho chất lượng logic điều hành; QMS giữ R cho register và evidence; OPS_SCOPE_OWNERS giữ R cho phân tầng, đo lường và phản ứng trong phạm vi được giao.",
        )

    if current_file.name == "sop-104-data-governance-records-security-and-ip-protection.html":
        set_section_role_cell(doc, "p4", 4, bundle("DIRECT_LINE_MGRS"), current_file, registry)
        set_section_cell_html(doc, "p8", 1, 3, chips(bundle("OPS_SCOPE_OWNERS")))
        set_section_cell_html(doc, "p8", 2, 2, chips(bundle("OPS_SCOPE_OWNERS")))
        set_section_cell_html(doc, "p8", 2, 3, chips(expr("QA", "CS", "EST", "PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA", "CEO")))
        set_row_cell_html(
            doc,
            "Kiểm thử khôi phục không đạt hoặc backup nghi lỗi",
            3,
            f'{chips(expr("QA"))}; {chips(expr("CEO"))} nếu ảnh hưởng giao hàng hoặc dừng sản xuất',
        )
        replace_text_fragments(
            doc,
            {
                "chủ dữ liệu": "đơn vị sở hữu dữ liệu nghiệp vụ",
            },
        )

    if current_file.name == "sop-106-change-and-configuration-management.html":
        set_section_role_cell(doc, "p4", 2, expr("PD", "WKM", "SL"), current_file, registry)
        set_section_cell_html(doc, "p6", 1, 2, chips(expr("ENGM", "QA", "PD", "SCM", "ITA", "ESA")))
        set_section_cell_html(doc, "p6", 2, 2, chips(expr("CEO", "QA", "PD", "ENGM", "ITA")))
        set_section_cell_html(doc, "p6", 3, 2, chips(expr("PPL", "WKM", "SCM", "ENGM", "QA", "ITA")))
        set_section_cell_html(doc, "p6", 4, 2, chips(expr("ENGM", "QA", "PD", "ITA", "ESA")))
        set_section_cell_html(doc, "p6", 5, 2, chips(expr("ENGM", "QMS[DC]", "ESA")))
        set_section_cell_html(doc, "p8", 1, 2, chips(expr("CS", "ENGM", "QA", "PD", "SCM", "ITA", "ESA")))
        set_section_cell_html(doc, "p8", 2, 2, chips(expr("CS", "ENGM")))
        set_section_cell_html(doc, "p8", 3, 2, chips(expr("MNT", "ENGM")))
        set_section_cell_html(doc, "p9", 1, 3, chips(expr("QMS")))
        replace_text_fragments(
            doc,
            {
                "JD của Trưởng Engineering, Trưởng QA, Giám đốc Sản xuất, Quản trị IT và Quản trị hệ thống Epicor phải ghi rõ quyền phê duyệt thay đổi, trách nhiệm impact review, cutover, configuration đánh giá, WIP segregation và KPI unapproved change / post-change xác minh.":
                "JD-ENGM, JD-QA, JD-PD, JD-ITA và JD-ESA phải ghi rõ quyền phê duyệt thay đổi, trách nhiệm impact review, cutover, configuration đánh giá, WIP segregation và KPI unapproved change / post-change xác minh.",
            },
        )

    if current_file.name == "sop-105-organizational-knowledge-management.html":
        set_section_role_cell(doc, "p4", 2, bundle("FUNC_OWNERS"), current_file, registry)
        set_section_role_cell(doc, "p4", 3, bundle("KNOWLEDGE_SMES"), current_file, registry)
        set_section_role_cell(doc, "p4", 4, bundle("FRONTLINE_LEADS"), current_file, registry)
        set_section_cell_html(doc, "p4", 0, 1, chips(expr("QA", "QMS", "CS", "EST", "PD", "ENGM", "SCM", "FIN", "HR", "EHS", "ITA")), table_index=2)
        set_section_cell_html(doc, "p9", 4, 3, chips(bundle("DIRECT_LINE_MGRS")))
        set_row_cell_html(
            doc,
            "CS / EST / PD / ENGM / QA[QMR] / SCM / FIN / HR / EHS / ITA",
            1,
            "Xác định tri thức trọng yếu, người giữ tri thức và tính hữu ích thực tế tại D-code do mình phụ trách.",
        )
        set_row_cell_html(doc, "Bàn giao pack", 2, "D-code sở hữu / D-HR thư mục")
        replace_text_fragments(
            doc,
            {
                "JD của QMS Engineer, HR Lead, Lead Department và SME trọng yếu phải mô tả trách nhiệm nhận diện tri thức, bài học kinh Nhiệm, bàn giao pack, cập nhật đào tạo và quyền chặn thôi việc / rời công ty khi chưa bàn giao tri thức.":
                "JD của QMS, HR, các vai trò trong FUNC_OWNERS và các vai trò trong KNOWLEDGE_SMES phải mô tả rõ trách nhiệm nhận diện tri thức, bàn giao tri thức, cập nhật đào tạo và quyền giữ mở khi chưa chuyển giao đủ tri thức.",
                "Tri thức có tranh cãi dữ liệu hoặc chưa được QA/QMS xác thực.": "Tri thức có tranh cãi dữ liệu hoặc chưa được QA / QMS xác thực.",
            },
        )

    if current_file.name == "sop-107-communication-management.html":
        set_section_cell_html(doc, "p4", 0, 1, chips(expr("CS", "SL", "PPL", "QA")))
        set_section_cell_html(doc, "p4", 0, 2, chips(bundle("OPS_SCOPE_OWNERS")))
        set_section_cell_html(doc, "p4", 0, 3, chips(expr("QMS")))
        set_section_cell_html(doc, "p4", 1, 1, chips(expr("PPL", "QA", "WKM")))
        set_section_cell_html(doc, "p4", 1, 2, chips(expr("PD", "QA")))
        set_section_cell_html(doc, "p4", 1, 3, chips(expr("ENGM", "WAR", "BUY")))
        set_section_cell_html(doc, "p4", 2, 1, chips(expr("CS", "EST")))
        set_section_cell_html(doc, "p4", 2, 2, chips(expr("QA", "PD")))
        set_section_cell_html(doc, "p4", 2, 3, chips(expr("ENGM", "PPL", "FIN")))
        set_section_cell_html(doc, "p4", 3, 1, chips(expr("SL", "WKM", "QCL", "ITA")))
        set_section_cell_html(doc, "p4", 3, 2, chips(expr("PD", "QA", "CEO")))
        set_section_cell_html(doc, "p4", 3, 3, chips(expr("PPL", "ENGM", "ITA", "WAR")))
        set_section_cell_html(doc, "p4", 4, 1, chips(bundle("OPS_SCOPE_OWNERS")))
        set_section_cell_html(doc, "p4", 4, 2, chips(expr("CS", "EST", "PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA", "WKM", "SL", "QCL")))
        set_section_cell_html(doc, "p4", 4, 3, chips(expr("QMS")))
        replace_text_fragments(
            doc,
            {
                "Sender": "OPS_SCOPE_OWNERS",
                "CS / Planner / QA theo tình huống": "CS / PPL / QA",
            },
        )

    if current_file.name == "sop-201-order-fulfillment-rfq-to-cash.html":
        replace_text_fragments(
            doc,
            {
                "Purchasing + IQC": "BUY + QCL",
                "Engineering Lead + QA + Sales Lead": "ENGM + QA + EST",
                "Sales Lead": "CS / EST",
                "Chủ trì rủi ro": "QA[QMR]",
                "Logistics + AR": "LOG + APAR",
            },
        )

    if current_file.name == "sop-301-engineering-dfm-quoting-and-machining-planning.html":
        replace_text_fragments(
            doc,
            {
                "Estimator / CSR": "EST / CS",
            },
        )
        set_row_cell_html(doc, "Special process chưa có source được phê duyệt", 3, chips(expr("QA", "ENGM")))

    if current_file.name == "sop-302-first-article-inspection-fai.html":
        replace_text_fragments(
            doc,
            {
                "QC Inspector + CMM": "QC",
            },
        )

    if current_file.name == "sop-108-operational-contingency-plan.html":
        set_section_cell_html(doc, "p9", 2, 3, chips(expr("MNT", "ITA", "QA")))
        replace_text_fragments(
            doc,
            {
                "JD cá»§a Sáº£n xuáº¥t GiÃ¡m Ä‘á»‘c, IT Administrator, QA Lead, Maintenance Technician, EHS Specialist vÃ TrÆ°á»Ÿng Chuá»—i cung á»©ng":
                "JD-PD, JD-ITA, JD-QA, JD-MNT, JD-EHS vÃ JD-SCM",
            },
        )

    if current_file.name == "sop-401-supplier-control-and-special-process.html":
        set_section_role_cell(doc, "p4", 4, expr("ENGM", "PE", "QE", "CAM"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "ngÆ°á»i chá»‹u trÃ¡ch nhiá»‡m dá»¯ liá»‡u": "role xÃ¡c nháº­n nguá»“n dá»¯ liá»‡u",
            },
        )
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: Supply Chain Manager giữ A cho source approval và re-approval; QA Manager giữ A cho acceptance of quality risk và SCAR closure; Buyer giữ R cho PO và dispatch accuracy; Process Owner giữ R cho technical flow-down.",
            "<b>RACI nền:</b> SCM giữ A cho source approval và re-approval; QA giữ A cho quality risk và SCAR closure; BUY giữ R cho PO và dispatch accuracy; ENGM / PE / QE / CAM giữ R cho technical flow-down và acceptance logic.",
        )
        replace_text_fragments(
            doc,
            {
                "Buyer + IQC": "BUY + QCL",
            },
        )

    if current_file.name == "sop-402-material-verification-traceability-and-counterfeit-prevention.html":
        set_section_cell_html(doc, "p6", 2, 2, chips(expr("QCL")))
        replace_text_fragments(
            doc,
            {
                "Warehouse + IQC": "WAR + QCL",
                "Warehouse Clerk + IQC": "WAR + QCL",
            },
        )

    if current_file.name == "sop-303-engineering-release-baseline-package-and-job-snapshot-control.html":
        set_section_cell_html(doc, "p6", 0, 2, chips(expr("ENGM", "QMS[DC]")))
        set_section_cell_html(doc, "p6", 5, 2, chips(expr("ENGM", "QMS[DC]")))
        replace_text_fragments(
            doc,
            {
                "Engineering Configuration Lead": "ENGM / QMS[DC]",
                "QA Manager + Engineering Lead / Manager": "QA / ENGM",
                "JD Process Engineer, CAM / NC Programmer, Engineering Lead / Manager, QMS Engineer, Quality Engineer vÃ Production Planner":
                "JD-PE, JD-CAM, JD-ENGM, JD-QMS, JD-QE vÃ JD-PPL",
            },
        )

    if current_file.name == "sop-604-spc-and-capability-control.html":
        set_section_role_cell(doc, "p4", 4, expr("ENGM", "PE", "PIE", "WKM", "SET", "MNT", "TOOL"), current_file, registry)
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: Quality Engineer giữ A cho rule SPC và capability; Operator giữ R cho tín hiệu tại nguồn; QC giữ R cho dữ liệu xác nhận; QA Manager giữ A cho escalation; Process Owner giữ R cho cải tiến quá trình.",
            "<b>RACI nền:</b> QE giữ A cho rule SPC và capability; OPR giữ R cho tín hiệu tại nguồn; QC giữ R cho dữ liệu xác nhận; QA giữ A cho escalation; ENGM / PE / PIE / WKM / SET / MNT / TOOL giữ R cho cải tiến quá trình.",
        )

    if current_file.name == "sop-606-ncr-capa-and-ipqc-reaction.html":
        set_section_role_cell(doc, "p4", 3, expr("WKM", "ENGM", "QE", "SCM", "HR", "MNT"), current_file, registry)
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: QC giữ R cho stop và containment đầu tiên; QA Manager giữ A cho disposition và CAPA trigger; QMS Engineer giữ R cho logic hồ sơ và hiệu lực; Process Owner giữ R cho action gốc tại quá trình; Planner hoặc Shipping giữ R cho integrity của lot trong containment.",
            "<b>RACI nền:</b> QC giữ R cho stop và containment đầu tiên; QA giữ A cho disposition và CAPA trigger; QMS giữ R cho logic hồ sơ và hiệu lực; WKM / ENGM / QE / SCM / HR / MNT giữ R cho action gốc tại quá trình; PPL / LOG giữ R cho integrity của lot trong containment.",
        )
        replace_text_fragments(
            doc,
            {
                "JD QA Manager, QMS Engineer, QC Lead, Production Planner và Process Engineer":
                "JD-QA, JD-QMS, JD-QCL, JD-PPL và JD-PE",
                "MRB / QA Manager": "QA",
            },
        )

    if current_file.name == "sop-503-tooling-maintenance-pm-and-breakdown-response.html":
        replace_text_fragments(
            doc,
            {
                "Maintenance + Workshop": "MNT + WKM",
            },
        )

    if current_file.name == "sop-505-finishing-deburr-and-secondary-operations-control.html":
        replace_text_fragments(
            doc,
            {
                "Secondary Process Leader": "DBL / CPS",
            },
        )

    if current_file.name == "sop-602-measurement-system-analysis-msagr-r.html":
        replace_text_fragments(
            doc,
            {
                "Quality Engineer + Metrology": "QE + MCS",
            },
        )

    if current_file.name == "sop-803-invoicing-job-costing-and-arap.html":
        replace_text_fragments(
            doc,
            {
                "Customer Service / AP-AR": "CS / APAR",
                "QA / AP-AR dùng để đọc": "QA / APAR",
            },
        )

    if current_file.name == "sop-603-aql-sampling-inspection.html":
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: QC giữ R cho sampling discipline; QC Lead giữ A cho decision tại hiện trường; QA Manager giữ A cho override và rule use-case; Planner và Warehouse giữ R cho tính toàn vẹn của lot sau decision.",
            "<b>RACI nền:</b> QC giữ R cho sampling discipline; QCL giữ A cho decision tại hiện trường; QA giữ A cho override và rule use-case; PPL / WAR giữ R cho tính toàn vẹn của lot sau decision.",
        )
        replace_text_fragments(
            doc,
            {
                "JD QC Inspector, QC Lead, QA Manager, Production Planner và Warehouse Clerk":
                "JD-QC, JD-QCL, JD-QA, JD-PPL và JD-WAR",
            },
        )

    if current_file.name == "sop-702-contamination-control-and-cleanliness.html":
        set_section_cell_html(doc, "p6", 1, 2, chips(expr("CPS", "WKM")))
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: Cleaning Supervisor giữ A cho route sạch; Technician giữ R cho thao tác; QC giữ A cho verification; QA Manager giữ A cho breach disposition; EHS giữ R cho safety của điều kiện môi trường và hóa chất.",
            "<b>RACI nền:</b> CPS giữ A cho route sạch; CPT giữ R cho thao tác; QC giữ A cho verification; QA giữ A cho breach disposition; EHS giữ R cho safety của điều kiện môi trường và hóa chất.",
        )

    if current_file.name == "sop-801-competence-training-and-certification.html":
        set_section_role_cell(doc, "p4", 2, bundle("DIRECT_LINE_MGRS"), current_file, registry)
        set_section_role_cell(doc, "p4", 3, bundle("OJT_COACHES"), current_file, registry)
        set_section_cell_html(doc, "p6", 2, 2, chips(bundle("DIRECT_LINE_MGRS")))
        set_section_cell_html(doc, "p8", 2, 2, chips(bundle("OJT_COACHES")))
        set_section_cell_html(doc, "p8", 2, 3, chips(bundle("DIRECT_LINE_MGRS")))
        set_section_cell_html(doc, "p9", 1, 3, chips(bundle("OJT_COACHES")))
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: HR giữ R cho hạ tầng hồ sơ; QA Manager giữ A cho competence quality-critical; Department Head giữ A cho năng lực theo vai trò; Trainer giữ R cho OJT evidence; người học giữ trách nhiệm tuân thủ giới hạn chứng nhận của mình.",
            "<b>RACI nền:</b> HR giữ R cho hạ tầng hồ sơ; QA giữ A cho competence quality-critical; DIRECT_LINE_MGRS giữ A cho năng lực theo vai trò và phân công công việc; OJT_COACHES giữ R cho evidence OJT; người học giữ trách nhiệm tuân thủ giới hạn chứng nhận của mình.",
        )

    if current_file.name == "sop-901-internal-audit-and-lpa.html":
        set_section_role_cell(doc, "p4", 3, bundle("OPS_SCOPE_OWNERS"), current_file, registry)
        set_section_role_cell(doc, "p4", 4, bundle("OPS_SCOPE_OWNERS"), current_file, registry)
        set_section_cell_html(doc, "p9", 2, 3, chips(expr("QMS[LA]")))
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: QMS Engineer giữ R cho execution và trend reporting; QA Manager giữ A cho severity, escalation và closure; Auditor giữ R cho bằng chứng; Process Owner giữ R cho hành động; Functional Head giữ A cho nguồn lực và cross-functional unblock.",
            "<b>RACI nền:</b> QMS giữ R cho execution và trend reporting; QA giữ A cho severity, escalation và closure; QMS[LA] giữ R cho bằng chứng; OPS_SCOPE_OWNERS giữ R cho hành động; OPS_SCOPE_OWNERS giữ A cho nguồn lực và cross-functional unblock trong phạm vi được audit.",
        )

    if current_file.name == "sop-902-management-review.html":
        set_section_role_cell(doc, "p4", 3, bundle("OPS_SCOPE_OWNERS"), current_file, registry)
        set_section_role_cell(doc, "p4", 4, bundle("OPS_SCOPE_OWNERS"), current_file, registry)
        set_section_cell_html(doc, "p9", 1, 3, chips(bundle("OPS_SCOPE_OWNERS")))
        replace_text_fragments(
            doc,
            {
                "Áp dụng cho Chief Executive Officer, QA Manager, QMS Engineer, Process Owner, Department Head, Data Owner, IT Administrator và mọi chức năng cung cấp đầu vào hoặc nhận action từ xem xét của lãnh đạo.":
                "Áp dụng cho CEO, QA, QMS và các vai trò cung cấp đầu vào hoặc nhận action từ xem xét của lãnh đạo, gồm CS, EST, PPL, PD, ENGM, SCM, FIN, HR, EHS, ITA, WKM, SL, QCL và LOG.",
                "Khi phát hiện dữ liệu sai sau mốc khóa dữ liệu, Data Owner phải mở controlled note; QA Manager quyết định re-review, bổ sung appendix hay hủy kết luận cũ để họp lại.":
                "Khi phát hiện dữ liệu sai sau mốc khóa dữ liệu, vai trò sở hữu metric hoặc nguồn dữ liệu trong OPS_SCOPE_OWNERS phải mở controlled note; QA quyết định re-review, bổ sung appendix hay hủy kết luận cũ để họp lại.",
                "Khi quyết định cần ngân sách hoặc thay đổi lớn vượt ngoài thẩm quyền hiện hữu, Process Owner phải escalate theo ANNEX-120 và giữ temporary control cho tới khi có quyết định cuối cùng.":
                "Khi quyết định cần ngân sách hoặc thay đổi lớn vượt ngoài thẩm quyền hiện hữu, vai trò chủ trì trong OPS_SCOPE_OWNERS phải escalate theo ANNEX-120 và giữ temporary control tới khi có quyết định cuối cùng.",
            },
        )
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: Chief Executive Officer giữ A cho quyết định hệ thống; QA Manager giữ A cho chất lượng review process; QMS Engineer giữ R cho điều phối và minutes; Process Owner và Data Owner giữ R cho input và evidence trong phạm vi mình sở hữu.",
            "<b>RACI nền:</b> CEO giữ A cho quyết định hệ thống; QA giữ A cho chất lượng review process; QMS giữ R cho điều phối và minutes; OPS_SCOPE_OWNERS giữ R cho input và evidence trong phạm vi được giao.",
        )

    if current_file.name == "sop-903-continual-improvement-and-kaizen.html":
        set_section_role_cell(doc, "p4", 0, bundle("OPS_SCOPE_OWNERS"), current_file, registry)
        set_section_cell_html(doc, "p6", 4, 2, chips(expr("QA[QMR]", "FIN", joiner=" + ")))
        set_section_cell_html(doc, "p8", 0, 2, chips(bundle("OPS_SCOPE_OWNERS")))
        replace_text_fragments(
            doc,
            {
                "JD Production Director, Production Engineer or IE, QA Manager, Quality Engineer, Finance Manager vÃ HR Manager":
                "JD-PD, JD-PIE, JD-QA, JD-QE, JD-FIN vÃ JD-HR",
            },
        )
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"role-note")]',
            "RACI nền: Process Owner giữ A cho pain point thật và duy trì kết quả; Production Engineer or IE giữ R cho thiết kế và trial; QA or QMS giữ A cho risk and standardization discipline; Finance giữ A cho hard-benefit sign-off; HR or sponsor giữ A cho adoption and unblock cross-functional resources.",
            "<b>RACI nền:</b> OPS_SCOPE_OWNERS giữ A cho pain point thật và duy trì kết quả; PIE giữ R cho thiết kế và trial; QA / QMS giữ A cho risk và standardization discipline; FIN giữ A cho hard-benefit sign-off; HR / CEO giữ A cho adoption và unblock cross-functional resources.",
        )

    if current_file.name == "annex-503-cnc-operating-model-and-role-boundary.html":
        replace_text_fragments(
            doc,
            {
                "Đo lường, QC Team Leader, QA": "MCS, QCL, QA",
                "QC Team Leader, QA, Hậu cần, Customer Dịch vụ": "QCL, QA, LOG, CS",
                "Không thay QC Team Leader trong bố trí nguồn lực kiểm hằng ngày; không tự disposition sản phẩm": "Không thay QC Inspector Lead trong bố trí nguồn lực kiểm hằng ngày; không tự disposition sản phẩm",
            },
        )

    if current_file.name == "wi-101-digital-online-forms-and-approvals.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QMS", "ITA"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "QA Lead hoặc Quy trình Owner theo RACI": "QA hoặc OPS_SCOPE_OWNERS theo ANNEX-121",
                "OPS/QA/WHS/QMS theo RACI của biểu mẫu": "OPS_SCOPE_OWNERS theo ANNEX-121",
                "WHS (hồ sơ giao hàng) + QA (hồ sơ chất lượng)": "D-WHS + QA",
                "QA hoặc OPS_SCOPE_OWNERS theo ANNEX-121": "QA / FUNC_OWNERS",
                "OPS_SCOPE_OWNERS theo ANNEX-121": "FUNC_OWNERS",
            },
        )

    if current_file.name == "wi-102-sharepoint-record-sites-libraries-and-permissions-click-by-click.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("ITA", "QMS"), current_file, registry)

    if current_file.name == "wi-103-m365-folder-routing-training-competence-and-adoption-for-cnc-job-orders.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QMS", "HR", "ITA"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "supervisor kiểm được trong ca;": "FRONTLINE_LEADS kiểm được trong ca;",
                "community/champion hỗ trợ.": "mạng DEPLOYMENT_LEADS hỗ trợ.",
                "Champion từng phòng ban": "DEPLOYMENT_LEADS",
                "Department Manager / Supervisor": "FUNC_HEADS / DIRECT_LINE_MGRS",
                "danh sách champion theo phòng ban và ca.": "danh sách DEPLOYMENT_LEADS theo phòng ban và ca.",
                "Cam kết supervisor rà soát": "Cam kết DIRECT_LINE_MGRS rà soát",
                "Champion biết dạy lại": "DEPLOYMENT_LEADS biết dạy lại",
                "Danh sách champion đủ phạm vi bao phủ theo phòng ban và shift.": "Danh sách DEPLOYMENT_LEADS đủ phạm vi bao phủ theo phòng ban và shift.",
                "Người đào tạo/champion quan sát.": "OJT_COACHES / DEPLOYMENT_LEADS quan sát.",
                "Supervisor phải kiểm truy xuất ngay sau khi nộp.": "DIRECT_LINE_MGRS phải kiểm truy xuất ngay sau khi nộp.",
                "Lưu bộ RFQ vào đúng nhánh SAL;": "Lưu bộ RFQ vào đúng nhánh D-SCS;",
                "Supervisor yêu cầu mở lại 1-2 hồ sơ vừa nộp": "DIRECT_LINE_MGRS mở lại 1-2 hồ sơ vừa nộp",
                "Supervisor hiểu KPI và nhịp kiểm.": "DIRECT_LINE_MGRS hiểu KPI và nhịp kiểm.",
                "Supervisor + champion": "DIRECT_LINE_MGRS + DEPLOYMENT_LEADS",
                "Supervisor + QMS + IT": "DIRECT_LINE_MGRS + QMS + ITA",
                "Setup / Planner": "SET / PPL",
                "Planner, ENG, Setup, Production Engineer": "PPL / D-ENG / SET / PIE",
                "Setup, QA, QC, ENG": "SET / QA / QC / D-ENG",
                "Supervisor mở lại trong 60 giây.": "DIRECT_LINE_MGRS mở lại trong 60 giây.",
                "DEP-SAL": "D-SCS",
                "SAL RFQ": "D-SCS RFQ",
                "LÆ°u bá»™ RFQ vÃ o Ä‘Ãºng nhÃ¡nh SAL;": "LÆ°u bá»™ RFQ vÃ o Ä‘Ãºng nhÃ¡nh D-SCS;",
                "Supervisor Requests má»Ÿ láº¡i 1-2 há»“ sÆ¡ vá»«a ná»™p": "DIRECT_LINE_MGRS má»Ÿ láº¡i 1-2 há»“ sÆ¡ vá»«a ná»™p",
                "Má»—i supervisor kiá»ƒm Ã­t nháº¥t 3 há»“ sÆ¡ má»›i má»—i ca.": "Má»—i DIRECT_LINE_MGRS kiá»ƒm Ã­t nháº¥t 3 há»“ sÆ¡ má»›i má»—i ca.",
                "Ä‘á»•i supervisor": "Ä‘á»•i DIRECT_LINE_MGRS",
                "Supervisor hiá»ƒu KPI vÃ  nhá»‹p kiá»ƒm.": "DIRECT_LINE_MGRS hiá»ƒu KPI vÃ  nhá»‹p kiá»ƒm.",
                "Champion tá»± thao tÃ¡c Ä‘Æ°á»£c vÃ  dáº¡y láº¡i Ä‘Æ°á»£c.": "DEPLOYMENT_LEADS tá»± thao tÃ¡c Ä‘Æ°á»£c vÃ  dáº¡y láº¡i Ä‘Æ°á»£c.",
                "QMS + phÃ²ng ban champion máº«u Ã­t nháº¥t 10 file/tuáº§n/phÃ²ng ban.": "QMS + DEPLOYMENT_LEADS máº«u Ã­t nháº¥t 10 file/tuáº§n/phÃ²ng ban.",
            },
        )
        set_row_cell_html(doc, "2. Chá»n champion theo phÃ²ng ban vÃ  shift", 1, render_token_cluster(["FUNC_HEADS"], current_file, registry))
        set_row_cell_html(doc, "3. Chuáº©n bá»‹ playlist vÃ  tÃ¬nh huá»‘ng theo role", 1, render_token_cluster(["QMS", "HR", "DEPLOYMENT_LEADS"], current_file, registry))
        set_row_cell_html(doc, "6. Dáº¡y theo role/shift", 1, render_token_cluster(["DEPLOYMENT_LEADS", "DIRECT_LINE_MGRS"], current_file, registry))
        set_row_cell_html(doc, "7. OJT trÃªn 3-5 job Ä‘áº§u", 1, render_token_cluster(["DIRECT_LINE_MGRS", "DEPLOYMENT_LEADS"], current_file, registry))
        set_row_cell_html(doc, "8. HÃ ng NgÃ y rÃ  soÃ¡t + vÄƒn phÃ²ng giá»", 1, render_token_cluster(["DIRECT_LINE_MGRS", "QMS", "ITA"], current_file, registry))

    if current_file.name == "wi-104-m365-folder-routing-quick-cards-by-role-for-cnc-job-order.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QMS", "HR"), current_file, registry)

    if current_file.name == "wi-105-qms-document-navigation-role-based-reading-path-and-deployment.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QMS", "HR"), current_file, registry)

    if current_file.name == "wi-106-job-order-deployment-master-plan.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QMS", "PD", "ITA"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "Reviewer / QMS": "FUNC_OWNERS / QMS",
                "Champion / Supervisor": "DEPLOYMENT_LEADS",
                "QMS Manager / Document Owners": "QMS / FUNC_OWNERS",
                "HR Manager / Dept Managers": "HR / FUNC_HEADS",
                "QMS / Document Owners": "QMS / FUNC_OWNERS",
                "HR / Dept Manager": "HR / FUNC_HEADS",
                "Cutover Lead / Command Center Lead": "PD / QMS / ITA",
                "IT Manager / QMS Manager / Data Owner": "QMS / ITA / MR_REPORT_OWNERS",
                "Data Owner / IT": "QMS / ITA",
                "Cutover Lead điều phối toàn bộ kế hoạch Ngày go-live.": "PD điều phối toàn bộ kế hoạch Ngày go-live.",
                "Command Center Lead": "QMS / ITA",
                "Cutover Lead": "PD",
                "QMS Manager chốt readiness tài liệu, chứng cứ, training và audit trail.": "QMS chốt readiness tài liệu, chứng cứ, training và audit trail.",
                "IT Manager chốt quyền truy cập, backup, refresh, cutover kỹ thuật và fallback.": "ITA chốt quyền truy cập, backup, refresh, cutover kỹ thuật và fallback.",
                "Champion Lead điều phối Champion theo phòng ban và theo ca.": "QMS / HR điều phối mạng Champion theo phòng ban và theo ca.",
                "Dept Managers chịu trách nhiệm pass của nhân sự phòng mình.": "FUNC_HEADS chịu trách nhiệm pass của nhân sự trong phạm vi phòng ban mình.",
                "Data Owner chịu trách nhiệm số liệu dashboard, refresh và exception note.": "Vai trò sở hữu metric trong MR_REPORT_OWNERS chịu trách nhiệm số liệu dashboard, refresh và exception note.",
            },
        )

    if current_file.name == "wi-206-ship-release-pack-sscc-label-and-pack-reconciliation.html":
        replace_text_fragments(
            doc,
            {
                "Kho Manager + QA": "D-WHS + QA",
                "Warehouse Manager + QA": "D-WHS + QA",
                "Warehouse Manager + QA + Sales/CS khi ảnh hưởng khách": "D-WHS + QA + D-SCS",
                "Warehouse Manager + QA + Sales/CS when affecting customer": "D-WHS + QA + D-SCS",
                "QA thẩm quyền": "QA",
            },
        )

    if current_file.name == "annex-116-m365-folder-structure-blueprint.html":
        replace_text_fragments(
            doc,
            {
                "MÃ£ phÃ²ng ban: QMS, QA, ENG, PRO, PUR, HR, WHS, MNT, PLA, SAL, FIN, HSE, IT, CNC, OPS": "MÃ£ phÃ²ng ban: D-EXEC, D-SCS, D-ENG, D-PROD, D-PPC, D-QUAL, D-SCM, D-PUR, D-WHS, D-TCR, D-LOG, D-FIN, D-HR, D-EHS, D-IT, D-ERP",
                "RET-JOB-ENG-10Y": "RET-JOB-G1-10Y",
                "RET-SAL-WARRANTY-7Y": "RET-SCS-WARRANTY-7Y",
                "Chá»‰ ENG Ä‘Æ°á»£c phÃ¡t hÃ nh revision chuáº©n": "Chá»‰ D-ENG Ä‘Æ°á»£c phÃ¡t hÃ nh revision chuáº©n",
                "QMS + owner quy trÃ¬nh": "QMS + OPS_SCOPE_OWNERS",
                "IT + QMS data owner": "ITA + QMS + MR_REPORT_OWNERS",
                "QA/QMS + owner Ä‘Æ°á»£c chá»‰ Ä‘á»‹nh": "QA / QMS + owner Ä‘Æ°á»£c chá»‰ Ä‘á»‹nh",
                "QA + ENG": "QA + D-ENG",
                "SAL + QA": "D-SCS + QA",
                "QA + SAL": "QA + D-SCS",
                "PRO + ENG": "D-PROD + D-ENG",
                "PUR + QA": "D-PUR + QA",
                "Change Records (Job)</td><td>ENG</td>": "Change Records (Job)</td><td>D-ENG</td>",
            },
        )

    if current_file.name == "annex-116-m365-folder-structure-blueprint.html":
        replace_text_fragments(
            doc,
            {
                "Mã phòng ban: QMS, QA, ENG, PRO, PUR, HR, WHS, MNT, PLA, SAL, FIN, HSE, IT, CNC, OPS": "Mã phòng ban: D-EXEC, D-SCS, D-ENG, D-PROD, D-PPC, D-QUAL, D-SCM, D-PUR, D-WHS, D-TCR, D-LOG, D-FIN, D-HR, D-EHS, D-IT, D-ERP",
                "Chỉ ENG được phát hành revision chuẩn": "Chỉ D-ENG được phát hành revision chuẩn",
            },
        )

    if current_file.name == "annex-133-m365-records-site-topology-library-and-folder-blueprint.html":
        replace_text_fragments(
            doc,
            {
                "ANNEX-133 - M365 hồ sơ site cấu trúc liên kết, thư viện and thư mục bản thiết kế": "ANNEX-133 - Cấu trúc site, thư viện và thư mục hồ sơ M365",
                "Mã phòng ban chuẩn: EXEC, QMS, QA, ENG, PRO, SCM, SAL, FIN, HR, EHS, IT, ERP.": "Mã phòng ban chuẩn: D-EXEC, D-SCS, D-ENG, D-PROD, D-PPC, D-QUAL, D-SCM, D-PUR, D-WHS, D-TCR, D-LOG, D-FIN, D-HR, D-EHS, D-IT, D-ERP.",
                "ENG edit, QA/OPS đọc.": "D-ENG cập nhật; QA / D-PROD đọc.",
                "QMS + Lãnh đạo": "QMS + CEO",
                "Lãnh đạo + QMS": "CEO + QMS",
                "HR / Người đào tạo": "HR / OJT_COACHES",
                "Dòng / Chuyền Manager": "DIRECT_LINE_MGRS",
                "HR + Dept Trưởng / Đầu": "HR + FUNC_HEADS",
            },
        )

    if current_file.name == "wi-107-sharefile-git-cpanel-sync.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("ITA", "QMS"), current_file, registry)

    if current_file.name == "wi-202-daily-management-tier-meetings-kpi-and-escalation.html":
        replace_text_fragments(
            doc,
            {
                "Kế hoạch Team Leader": "PPL",
                "QA/ENG/OPS": "QA / D-ENG / D-PROD",
            },
        )
        set_row_cell_html(doc, "QC Hold SLA", 2, chips(expr("QA")))
        set_row_cell_html(doc, "Thời gian Từng máy tác động", 2, chips(expr("WKM", "MNT")))

    if current_file.name == "wi-203-job-dossier-evidence-pack-and-record-completeness.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QA", "QMS", "PPL", "ENGM"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "QA / QMS + Quy trình Owner của cổng kiểm soát": "QA / QMS + FUNC_OWNERS",
                "QA / Giao vận theo lộ trình gia công": "QA / D-LOG",
            },
        )

    if current_file.name == "wi-205-barcode-labeling-and-scan-to-action.html":
        replace_text_fragments(
            doc,
            {
                "do Supervisor cho phép.": "do DIRECT_LINE_MGRS cho phép.",
                "QA / Supervisor": "QA / DIRECT_LINE_MGRS",
            },
        )

    if current_file.name == "wi-519-job-packet-quick-check-and-pre-run-verification.html":
        replace_text_fragments(
            doc,
            {
                "báo Supervisor / QA / Kỹ thuật": "báo DIRECT_LINE_MGRS / QA / D-ENG",
                "Setup Lead/Supervisor": "SET / DIRECT_LINE_MGRS",
                "Supervisor + QA; Kỹ thuật/Bảo trì khi có machine tác động": "DIRECT_LINE_MGRS + QA; D-ENG / MNT khi có machine tác động",
                "Báo đúng Người.</strong><br>Sai gói hồ sơ/phiên bản báo Kỹ thuật + QA; sai setup báo Setup Lead/Supervisor; nghi ngờ machine điều kiện báo Bảo trì; sai lot/nguyên vật liệu báo Kho + QA.":
                "Báo đúng Người.</strong><br>Sai gói hồ sơ/phiên bản báo D-ENG + QA; sai setup báo SET / DIRECT_LINE_MGRS; nghi ngờ machine điều kiện báo MNT; sai lot/nguyên vật liệu báo D-WHS + QA.",
            },
        )
        set_row_cell_by_match(
            doc,
            0,
            "Sau cảnh báo, power tổn thất, khởi động lại, dừng dài, chuyển đổi (changeover) qua ca",
            2,
            chips(bundle("DIRECT_LINE_MGRS")),
        )
        set_row_cell_by_match(
            doc,
            0,
            "Sau chuyển giao công việc sang máy khác hoặc Người khác nhận máy",
            2,
            f'{chips(bundle("DIRECT_LINE_MGRS"))} + {chips(expr("OPR", "SET", joiner=" / "))}',
        )
        set_row_cell_by_match(
            doc,
            0,
            "Job tiêu chuẩn QPL-1/QPL-2",
            2,
            f'{chips(expr("OPR"))} + {chips(expr("SET"))} / {chips(bundle("DIRECT_LINE_MGRS"))}',
        )
        set_row_cell_by_match(
            doc,
            0,
            "Job sau sự cố nghiêm trọng/bất thường khởi động lại",
            2,
            f'{chips(bundle("DIRECT_LINE_MGRS"))} + {chips(expr("QA"))}; {chips(expr("D-ENG", "MNT", joiner=" / "))} khi có machine tác động',
        )

    if current_file.name == "wi-201-quality-gates-hold-points-and-release-execution.html":
        replace_text_fragments(
            doc,
            {
                "Engineering Lead / ENGM.": "ENGM.",
                "Engineering Lead + QA Manager": "ENGM + QA",
                "Engineering Lead ký phát hành baseline package và đóng băng job snapshot.": "ENGM ký phát hành baseline package và đóng băng job snapshot.",
                "thông báo Purchasing và QA.": "thông báo D-PUR và QA.",
                "Engineering Lead / Engineering Manager.": "ENGM.",
                "IQC Team Leader / WHS": "QCL / D-WHS",
                "IQC Team Leader.": "QCL.",
                "IQC Team Leader báo QA Manager trong 1 giờ.": "QCL báo QA trong 1 giờ.",
                "Kế hoạch, ENG, Sản xuất, QA và WHS": "PPL, D-ENG, D-PROD, QA và D-WHS",
                "Kế hoạch/OPS": "PPL / PD",
                "Supervisor, ENG và Người setup phải kiểm chéo": "WKM / SL, D-ENG và SET phải kiểm chéo",
                "ENG phát hành lại; gắn link SSOT và ProgramRev": "D-ENG phát hành lại; gắn link SSOT và ProgramRev",
                "ENG phát hành phiên bản mới và thông báo lập kế hoạch": "D-ENG phát hành phiên bản mới và thông báo PPL",
                "TẠM GIỮ ngay; Mọi ENG/Kế hoạch": "TẠM GIỮ ngay; D-ENG / PPL xử lý",
                "TẠM GIỮ; đổi gage hoặc Mọi QA": "TẠM GIỮ; đổi gage hoặc QA xử lý",
            },
        )

    if current_file.name == "wi-501-dispatch-capacity-and-wip-control.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("PPL", "WKM"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "Planner, Nhân viên điều phối, Supervisor và Cell Lead": "PPL / DIRECT_LINE_MGRS",
                "HOLD job; trả về ENG/Planning để sửa packet": "HOLD job; trả về D-ENG / PPL để sửa packet",
                "Requests Supervisor tự hiểu.": "rồi để DIRECT_LINE_MGRS tự hiểu.",
                "Supervisor/Cell Lead": "DIRECT_LINE_MGRS",
                "QA/WHS/ENG": "QA / D-WHS / D-ENG",
                "người chịu trách nhiệm Kỹ thuật": "ENGM",
            },
        )

    if current_file.name == "wi-501-dispatch-capacity-and-wip-control.html":
        replace_text_fragments(
            doc,
            {
                "QA/ENG/WHS": "QA / D-ENG / D-WHS",
                "Ops Manager hoáº·c cao hÆ¡n": "PD hoáº·c cao hÆ¡n",
            },
        )

    if current_file.name == "wi-201-quality-gates-hold-points-and-release-execution.html":
        replace_text_fragments(
            doc,
            {
                "QuÃ¡ SLA â†’ Engineering Lead bÃ¡o Ops Manager trong 2 giá». QuÃ¡ 2Ã— SLA â†’ Ops Manager escalate CEO.": "QuÃ¡ SLA â†’ ENGM bÃ¡o PD trong 2 giá». QuÃ¡ 2Ã— SLA â†’ PD escalate CEO.",
            },
        )

    if current_file.name == "wi-501-dispatch-capacity-and-wip-control.html":
        replace_text_fragments(
            doc,
            {
                "Ops Manager ho\u1eb7c cao h\u01a1n": "PD ho\u1eb7c cao h\u01a1n",
            },
        )

    if current_file.name == "wi-201-quality-gates-hold-points-and-release-execution.html":
        replace_text_fragments(
            doc,
            {
                "Qu\u00e1 SLA \u2192 Engineering Lead b\u00e1o Ops Manager trong 2 gi\u1edd. Qu\u00e1 2\u00d7 SLA \u2192 Ops Manager escalate CEO.": "Qu\u00e1 SLA \u2192 ENGM b\u00e1o PD trong 2 gi\u1edd. Qu\u00e1 2\u00d7 SLA \u2192 PD escalate CEO.",
            },
        )

    if current_file.name == "wi-701-receiving-iqc-traceability-and-put-away.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("WAR", "QCL"), current_file, registry)

    if current_file.name == "wi-702-storage-environment-location-and-fifo-control.html":
        replace_text_fragments(
            doc,
            {
                "Kho Supervisor / Kế hoạch": "WAR / PPL",
                "Kế hoạch / Ops Manager theo thẩm quyền": "PPL / PD theo ANNEX-120",
                "Sản xuất Supervisor": "WKM / SL",
                "Kho Supervisor phê duyệt": "SCM phê duyệt",
                "Kho Supervisor PHẢI mở quyền tràn": "SCM PHẢI mở quyền tràn",
                "Kế hoạch + QA + Kho Supervisor theo thẩm quyền": "PPL + QA + SCM theo ANNEX-120",
                "Kho + Sản xuất Supervisor": "D-WHS + WKM / SL",
            },
        )

    if current_file.name == "wi-711-cleanroom-entry-and-gowning.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("CPS", "QE", "EHS"), current_file, registry)

    if current_file.name == "wi-711-cleanroom-entry-and-gowning.html":
        replace_text_fragments(
            doc,
            {
                "Báo Supervisor hoặc QA.": "Báo CPS hoặc QE.",
                "Supervisor cho phép.": "CPS cho phép.",
                "Supervisor/QA": "CPS / QE",
                "Supervisor hoáº·c QA.": "CPS hoáº·c QE.",
                "Supervisor/HSE": "CPS / EHS",
                "Supervisor cho phÃ©p.": "CPS cho phÃ©p.",
            },
        )

    if current_file.name == "wi-712-ultrasonic-cleaning-standard-work.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("CPS", "QE"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "Theo phê duyệt ENG/QA": "Theo phê duyệt D-ENG / QA",
            },
        )

    if current_file.name == "wi-713-environmental-monitoring-and-response.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QE", "EHS", "CPS"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "Supervisor khu vực.": "AREA_LEADS.",
                "Supervisor + QA khi ảnh hưởng sản phẩm.": "AREA_LEADS + QA khi ảnh hưởng sản phẩm.",
            },
        )

    if current_file.name == "wi-714-clean-packaging-handling-and-preservation.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("CPS", "QE", "WAR"), current_file, registry)

    if current_file.name == "wi-715-helium-leak-test-standard-work.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QCL", "QE", "QC"), current_file, registry)

    if current_file.name == "wi-716-vacuum-compatible-clean-build-and-bagging.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("CPS", "QE", "WAR"), current_file, registry)

    if current_file.name == "wi-721-fod-prevention-line-clearance-and-tool-accountability.html":
        replace_text_fragments(
            doc,
            {
                "Sản xuất Supervisor / QA": "WKM / SL / QA",
            },
        )

    if current_file.name == "wi-801-cnc-poka-yoke-examples.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QE", "WKM", "PE"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "Supervisor / QA": "WKM / QE",
            },
        )

    if current_file.name == "wi-901-performance-dashboard.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QMS", "ITA"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "QMS Manager / KPI người chịu trách nhiệm / IT Data Owner": "QMS / ITA / MR_REPORT_OWNERS",
            },
        )

    if current_file.name == "annex-101-role-based-access-map.html":
        replace_text_fragments(
            doc,
            {
                "Line Manager / Người phụ trách quy trình": "DIRECT_LINE_MGRS / OPS_SCOPE_OWNERS",
                "Department head hoặc delegate hợp lệ": "FUNC_HEADS hoặc delegate hợp lệ theo ANNEX-123",
                "Department head / chủ quá trình": "FUNC_HEADS / OPS_SCOPE_OWNERS",
                "Department head + chủ quá trình": "FUNC_HEADS + OPS_SCOPE_OWNERS",
                "Line Manager / deputy trigger Người phụ trách": "DIRECT_LINE_MGRS / deputy trigger theo ANNEX-123",
                "Dòng / Chuyền Manager / người chịu trách nhiệm quy trình": "DIRECT_LINE_MGRS / OPS_SCOPE_OWNERS",
                "Department trưởng / đầu hoặc ủy quyền hợp lệ": "FUNC_HEADS hoặc ủy quyền hợp lệ theo ANNEX-123",
                "Department trưởng / đầu / Quy trình Owner": "FUNC_HEADS / OPS_SCOPE_OWNERS",
                "người chịu trách nhiệm System + chủ nghiệp vụ đồng ký": "ITA / ESA + OPS_SCOPE_OWNERS đồng ký",
                "Top quản lý / được ủy quyền điều hành theo mức độ": "CEO / FUNC_HEADS theo ANNEX-120",
                "IT/Epicor admin": "ITA / ESA",
                "Dòng / Chuyền Manager / phó kích hoạt người chịu trách nhiệm": "DIRECT_LINE_MGRS / deputy kích hoạt theo ANNEX-123",
                "Department trưởng / đầu + Quy trình Owner": "FUNC_HEADS + OPS_SCOPE_OWNERS",
            },
        )

    if current_file.name == "annex-102-access-request-field-dictionary.html":
        replace_text_fragments(
            doc,
            {
                "Người dùng/Dòng / Chuyền Manager": "Người dùng / DIRECT_LINE_MGRS",
                "Dòng / Chuyền Manager + người chịu trách nhiệm hệ / Approver luồng công việc chuẩn": "DIRECT_LINE_MGRS + ITA / ESA / approver luồng chuẩn",
                "Department trưởng / đầu / Quy trình Owner + Approver theo quyết định nhóm sản phẩm": "FUNC_HEADS / OPS_SCOPE_OWNERS + approver theo ANNEX-120",
                "Dòng / Chuyền Manager + phòng ban trưởng / đầu + Quy trình Owner": "DIRECT_LINE_MGRS + FUNC_HEADS + OPS_SCOPE_OWNERS",
            },
        )

    if current_file.name == "annex-104-org-chart-fullpage.html":
        replace_text_fragments(
            doc,
            {
                "Department heads, QA, EHS": "FUNC_HEADS / QA / EHS",
                "HR ↔ Department heads ↔ QA": "HR ↔ FUNC_HEADS ↔ QA",
                "All functions": "FUNC_OWNERS",
            },
        )

    if current_file.name == "annex-107-audit-evidence-pack-master.html":
        replace_text_fragments(
            doc,
            {
                "QMS + EXE.": "QMS + CEO",
                "Tài liệu Control.": "QMS[DC]",
                "Kế hoạch + QA + Kho.": "PPL + QA + D-WHS",
                "Mua hàng + QA.": "BUY + QA",
                "QMS + Quy trình Owner.": "QMS + FUNC_OWNERS",
            },
        )

    if current_file.name == "annex-117-escalation-matrix-and-sla.html":
        replace_text_fragments(
            doc,
            {
                "Lead / QA / Ops / Maintenance": "DIRECT_LINE_MGRS / QA / PD / MNT",
                "Lead hoặc chức năng sở hữu công đoạn; escalation lên Ops Manager nếu > 1 cell/job": "DIRECT_LINE_MGRS hoặc FUNC_OWNERS; escalation lên PD nếu > 1 cell/job",
                "Ops Manager / QMS / GM": "PD / QMS / CEO",
                "GM hoặc Người được chỉ định": "CEO hoặc người được ủy quyền theo ANNEX-123",
            },
        )

    if current_file.name == "annex-110-dashboard-kpi-dictionary-and-data-model.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QMS", "ITA"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "QMS Manager / IT Data Owner": "QMS / ITA",
                "Kế hoạch Team Leader": "PPL",
                "IT Data Owner": "ITA",
                "người chịu trách nhiệm Bảng điều khiển + Quy trình Owner": "QMS / MR_REPORT_OWNERS",
                "QMS / Quy trình Owner / bảng điều khiển người chịu trách nhiệm": "QMS / MR_REPORT_OWNERS",
                "Quy trình Owner + được ủy quyền Approver": "OPS_SCOPE_OWNERS + CEO / QA / PD / ENGM / SCM / FIN / HR / EHS / ITA",
                "Sản xuất / Bảo trì": "WKM / MNT",
                "Sản xuất Manager": "WKM",
            },
        )

    if current_file.name == "annex-504-tier-meeting-cadence-and-escalation-standard-work.html":
        replace_text_fragments(
            doc,
            {
                "Planning, Production Lead, QA, WHS": "PPL, WKM / SL, QA, D-WHS",
                "Ops Manager (chủ trì), Planning, QA, WHS, Purchasing, Eng": "PD, PPL, QA, D-WHS, D-PUR, D-ENG",
                "General Manager (EXE-01) + Ops + QA / QMS + Finance": "CEO + PD + QA / QMS + FIN",
            },
        )

    if current_file.name == "annex-115-epicor-transaction-and-interface-map.html":
        replace_text_fragments(
            doc,
            {
                "Planner / sản xuất người chịu trách nhiệm / người chịu trách nhiệm QA tùy bảng điều khiển": "PPL / WKM / QA",
                "người chịu trách nhiệm KPI theo  ANNEX-122": "MR_REPORT_OWNERS",
                "người chịu trách nhiệm Finance": "FIN / APAR",
                "Kinh doanh & CS + Planner / người chịu trách nhiệm được ủy quyền": "D-SCS / PPL / FUNC_OWNERS",
                "người chịu trách nhiệm Kỹ thuật": "ENGM",
                "người chịu trách nhiệm Sản xuất / Planner": "WKM / PPL",
                "QA / QC thẩm quyền": "QA / QCL",
                "Kho / Hậu cần + QA phát hành thẩm quyền": "D-WHS / D-LOG + QA",
                "Epicor admin + chủ nghiệp vụ": "ESA + FUNC_OWNERS",
                "Kinh doanh & CS / Người định giá": "CS / EST",
                "Kinh doanh & CS + Planner + Kỹ thuật": "CS / PPL / ENGM",
                "người chịu trách nhiệm Epicor + chủ nghiệp vụ liên quan": "ESA + FUNC_OWNERS",
                "IT / QMS / Quy trình Owner": "ITA / QMS / FUNC_OWNERS",
                "người chịu trách nhiệm Epicor/IT": "ESA / ITA",
                "QMS + KPI người chịu trách nhiệm": "QMS + MR_REPORT_OWNERS",
            },
        )
        set_row_cell_html(doc, "Xem xét của lãnh đạo KPI pack", 2, chips(bundle("MR_REPORT_OWNERS")))

    if current_file.name == "annex-113-dashboard-deployment-access-and-refresh-control.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("QMS", "ITA"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "Cutover Lead, Sponsor": "PD, CEO",
                "IT Manager / QMS Manager / Data Owners": "QMS / ITA",
                "Business owner và data owner": "owner nghiệp vụ và owner dữ liệu",
                "Không thay Business Owner": "Không thay owner nghiệp vụ được chỉ định",
                "Biz + data owner": "owner nghiệp vụ + owner dữ liệu",
                "Data Owner + IT Manager": "MR_REPORT_OWNERS + ITA",
                "Business Owner + System Owner": "MR_REPORT_OWNERS + QMS / ITA / ESA",
                "System Owner": "QMS / ITA / ESA",
                "IT Manager": "ITA",
            },
        )
        replace_exact_block_text(
            doc,
            '//table[contains(@class,"d-table")]//td',
            "Manager, Champion, end-user",
            render_spec(bundle("FUNC_HEADS"), current_file, registry),
        )

    if current_file.name == "annex-114-go-live-runbook-and-cutover-control.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("PD", "QMS", "ITA"), current_file, registry)
        replace_text_fragments(
            doc,
            {
                "Business Owner / Dept Manager": "FUNC_OWNERS / FUNC_HEADS",
                "Deputy Manager hoặc Champion Lead": "FUNC_HEADS hoặc DEPLOYMENT_LEADS",
                "Champion Lead": "DEPLOYMENT_LEADS",
                "Dept Manager / HR": "FUNC_HEADS / HR",
                "Dept Manager / DEPLOYMENT_LEADS": "FUNC_HEADS / DEPLOYMENT_LEADS",
                "Dept Managers / Champions": "FUNC_HEADS / DEPLOYMENT_LEADS",
                "Cutover Lead, Sponsor, IT Manager, QMS Manager": "PD, CEO, ITA, QMS",
                "Cutover Lead, Sponsor, ITA, QMS": "PD, CEO, ITA, QMS",
                "Dept Manager / Champion Lead": "FUNC_HEADS / DEPLOYMENT_LEADS",
                "Dept Manager xác nhận phòng ban đã vận hành độc lập được mà không cần Champion túc trực liên tục.": "FUNC_HEADS xác nhận phòng ban đã vận hành độc lập được mà không cần Champion túc trực liên tục.",
                "Cutover Lead / IT": "PD / ITA",
                "Cutover Lead / Sponsor": "PD / CEO",
                "Executive Sponsor / Cutover Lead": "CEO / PD",
                "Cutover Lead": "PD",
                "IT / Data Owner": "ITA / MR_REPORT_OWNERS",
            },
        )

    if current_file.name == "annex-118-offline-fallback-kit.html":
        replace_text_fragments(
            doc,
            {
                "IT Manager / Epicor Admin": "ITA / ESA",
                "IT Manager + lãnh đạo site": "ITA + CEO / PD",
                "Sales Manager / Engineering Manager / Người thay quyền.": "CS / EST / ENGM.",
                "Engineering Manager / Người thay quyền.": "ENGM.",
                "Supply Chain / QA Lead.": "SCM / QA.",
                "Engineering Manager + Production Supervisor.": "ENGM + WKM.",
                "QA Lead + Engineering Manager.": "QA + ENGM.",
                "Production Supervisor + QA Lead.": "WKM + QA.",
                "Production Director / Supervisor": "PD / WKM / SL",
                "Production Supervisor + Epicor Admin": "WKM / SL + ESA",
                "Người phụ trách KPI": "MR_REPORT_OWNERS",
                "IT Manager + QMS": "ITA + QMS",
                "QA Manager hoặc Người thay quyền đã chỉ định.": "QA hoặc QCL theo ANNEX-123.",
                "Tổng Giám Đốc / QA Manager / Finance Manager theo ma trận thẩm quyền.": "CEO / QA / FIN theo ANNEX-120.",
                "QA / Quy trình Owner": "QA / FUNC_OWNERS",
                "QA + IT + Quy trình Owner": "QA + ITA + FUNC_OWNERS",
                "Lãnh đạo site + Finance + Chuỗi cung ứng": "PD + FIN + SCM",
            },
        )

    if current_file.name == "annex-119-change-roadmap-and-priority-register.html":
        replace_text_fragments(
            doc,
            {
                "người chịu trách nhiệm nghiệp vụ": "FUNC_OWNERS",
                "QA Manager, Quality Engineer, QC Team Leader, QC/CMM, Đo lường.": "QA, QE, QCL, QC và MCS.",
            },
        )

    if current_file.name == "annex-137-evidence-and-records-naming-convention.html":
        replace_text_fragments(
            doc,
            {
                "QA Engineer #1": "QE #1",
                "Production Supervisor #1": "WKM #1",
                "PRO1": "WKM1",
                "MEETING-MIN_ENG_": "MEETING-MIN_D-ENG_",
                "0800-ENG1": "0800-ENGM1",
                "ENG, PRO, QA, SCM, HR, FIN, WH, MNT": "D-ENG, D-PROD, QA, D-SCM, HR, FIN, D-WHS, MNT",
                "Engineering #1": "ENGM #1",
            },
        )

    if current_file.name == "annex-703-warehouse-location-fifo-rules.html":
        replace_text_fragments(
            doc,
            {
                "Warehouse Manager phê duyệt": "SCM phê duyệt",
                "Warehouse Manager phải mở điều tra": "SCM phải mở điều tra",
                "Warehouse Manager review": "SCM review",
                "ngưỡng giá trị do Warehouse Manager đặt hàng năm": "ngưỡng giá trị do SCM phê duyệt hằng năm",
                "Warehouse Manager phải review tối thiểu tháng một lần": "SCM phải review tối thiểu tháng một lần",
                "WHS + QA hoặc OPS witness": "D-WHS + QA hoặc D-PROD witness",
            },
        )

    if current_file.name == "annex-803-ppe-and-hazard-matrix.html":
        replace_text_fragments(
            doc,
            {
                "Supervisor, HSE, Bảo trì và Người trực tiếp làm việc PHẢI dùng tài liệu này": "DIRECT_LINE_MGRS, EHS, MNT và Người trực tiếp làm việc PHẢI dùng tài liệu này",
            },
        )

    if current_file.name == "dept-epicor-handbook.html":
        replace_text_fragments(
            doc,
            {
                "owner KPI xác nhận chỉ số nào là đúng cho ra quyết định.": "MR_REPORT_OWNERS xác nhận chỉ số nào là đúng cho ra quyết định.",
            },
        )

    if current_file.name == "annex-114-go-live-runbook-and-cutover-control.html":
        set_row_cell_by_match(
            doc,
            1,
            "Xác nhận Người dùng phòng mình sẵn sàng và xử lý issue nghiệp vụ",
            0,
            chips(bundle("FUNC_OWNERS")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Xác nhận Người dùng phòng mình sẵn sàng và xử lý issue nghiệp vụ",
            3,
            chips(bundle("DEPLOYMENT_LEADS")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Access, metadata, portal, refresh, backup, rollback kỹ thuật",
            0,
            chips(expr("QMS", "ITA", "ESA")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Access, metadata, portal, refresh, backup, rollback kỹ thuật",
            3,
            chips(expr("ITA", "ESA")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Phân công Champion theo ca, thu issue hiện trường, hỗ trợ tại điểm dùng",
            0,
            chips(bundle("DEPLOYMENT_LEADS")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Manager pass, Champion coverage đủ theo ca, Người dùng trọng yếu pass OJT",
            2,
            chips(bundle("FUNC_HEADS")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Dashboard có owner, source, last refresh, exception rule, baseline KPI",
            2,
            chips(bundle("MR_REPORT_OWNERS")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Đã diễn tập rollback và đã chốt trigger rõ",
            2,
            chips(expr("PD", "ITA", "ESA")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Quan sát giao dịch đầu tiên, job đầu tiên, upload đầu tiên, truy xuất đầu tiên",
            2,
            chips(expr("PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA", "WKM", "SL", "DBL", "CPS", "QCL")),
        )
        set_row_cell_by_match(
            doc,
            1,
            "Lỗi hướng dẫn, lỗi link cục bộ, lỗi UI, lỗi cần hỗ trợ tại chỗ nhưng chưa ảnh hưởng quyết định chính",
            3,
            chips(expr("PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA", "WKM", "SL", "DBL", "CPS", "QCL")),
        )
        replace_text_fragments(
            doc,
            {
                "Cutover Deputy": "PD backup theo rota cutover",
                "System Admin backup": "ITA / ESA backup",
                "Champion backup theo ca": "WKM / SL / DBL / CPS / QCL backup theo ca",
            },
        )

    if current_file.name == "annex-133-m365-records-site-topology-library-and-folder-blueprint.html":
        set_row_cell_html(
            doc,
            "05-Phòng ban-Hồ sơ",
            2,
            render_token_cluster(["D-SCS", "D-ENG", "D-PROD", "D-QUAL", "D-SCM", "D-FIN", "D-HR", "D-EHS", "D-IT", "D-ERP"], current_file, registry),
        )
        set_row_cell_html(doc, "06-Lưu trữ", 2, chips(expr("QMS[DC]")))
        set_row_cell_html(doc, "07-Mẫu-Làm việc", 2, render_token_cluster(["QMS[DC]", "FUNC_OWNERS"], current_file, registry))
        set_row_cell_html(doc, "09-Số hóa-System-Hồ sơ", 2, chips(expr("ITA", "ESA")))
        set_row_cell_html(doc, "Part-REV-Bản gốc", 2, chips(expr("ENGM")))
        set_row_cell_html(
            doc,
            "Part-REV-Bản gốc",
            3,
            f'{chips(expr("D-ENG"))} cập nhật; {chips(expr("QA", "D-PROD"))} đọc.',
        )
        set_row_cell_html(doc, "07_G6-Final-QC-Packaging", 2, chips(expr("QA", "D-PROD", "D-WHS")))

    if current_file.name == "index.html" and "03-Job-Descriptions" in current_file.as_posix():
        replace_text_fragments(
            doc,
            {
                "Kinh doanh & Customer Dịch vụ (2)": "Kinh doanh & Customer Service (2)",
            },
        )

    if current_file.name == "index.html" and "02-Department-Handbooks" in current_file.as_posix():
        replace_text_fragments(
            doc,
            {
                "chá»§ dá»¯ liá»‡u nghiá»‡p vá»¥": "Ä‘Æ¡n vá»‹ sá»Ÿ há»¯u dá»¯ liá»‡u nghiá»‡p vá»¥",
            },
        )

    if current_file.name == "annex-122-kpi-cascade-dictionary.html":
        replace_exact_block_text(
            doc,
            '//div[contains(@class,"preface-block")]//div',
            "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên Quy trình Owner FIN QA HR ITA ESA",
            render_token_cluster(
                ["CEO", "CS", "EST", "PD", "ENGM", "QA[QMR]", "QMS", "SCM", "FIN", "HR", "EHS", "ITA", "ESA"],
                current_file,
                registry,
            ),
        )
        replace_text_fragments(
            doc,
            {
                "WKM / Shift lãnh đạo": "WKM / SL",
                "người chịu trách nhiệm Đo lường / QA": "MCS / QA",
                "SCM / Kho / Tool kho dụng cụ...": "SCM / D-WHS / TOOL",
                "SCM / Kho / Tool kho dụng cụ (tool kho dụng cụ (tool crib)": "SCM / D-WHS / TOOL",
                "HR + Dept Trưởng bộ phận": "HR + FUNC_HEADS",
                "ESA + Quy trình Owner": "ESA + FUNC_OWNERS",
                "Sản xuất Manager / IT-chủ dữ liệu": "WKM / ITA",
                "KPI không có chủ dữ liệu": "KPI không có role xác nhận nguồn dữ liệu",
                "người chịu trách nhiệm Chính / Sơ cấp vs. chủ dữ liệu": "người chịu trách nhiệm Chính / Sơ cấp vs. role xác nhận nguồn dữ liệu",
                "chủ dữ liệu": "role xác nhận nguồn dữ liệu",
            },
        )

    if current_file.name == "annex-124-dashboard-evidence-pack-worked-examples.html":
        set_row_cell_html(doc, "Quản lý-bộ hồ sơ rà soát", 2, f'{chips(expr("QMS"))} + {chips(bundle("MR_REPORT_OWNERS"))}')

    if current_file.name == "wi-105-qms-document-navigation-role-based-reading-path-and-deployment.html":
        set_row_first_cell_html(doc, "Ban Giám đốc / Steering Committee", chips(bundle("DEPLOYMENT_STEERING")))

    if current_file.name == "wi-106-job-order-deployment-master-plan.html":
        set_row_cell_by_match(
            doc,
            0,
            "Governance & tài trợ dự án",
            2,
            chips(bundle("DEPLOYMENT_STEERING")),
        )

    if current_file.name == "pol-qms-002-quality-objectives.html":
        set_row_cell_html(doc, "Hệ thống & cải tiến", 2, chips(bundle("ALL_DEPTS")))

    if current_file.name == "wi-302-first-piece-fai-execution-and-evidence-pack.html":
        replace_text_fragments(
            doc,
            {
                "QA và ENG quyết định phạm vi; không dùng dữ liệu cũ để suy diễn.": "QA và D-ENG quyết định phạm vi; không dùng dữ liệu cũ để suy diễn.",
                "trả về ENG/Setup.": "trả về D-ENG / SET.",
                "Supervisor + QA": "WKM / SL + QA",
            },
        )

    if current_file.name == "wi-606-suspect-product-containment-segregation-and-reaction.html":
        replace_text_fragments(
            doc,
            {
                "Báo Supervisor, QA, WHS và người chịu trách nhiệm công đoạn.": "Báo WKM / SL, QA, D-WHS và người chịu trách nhiệm công đoạn.",
                "QA/ENG": "QA / D-ENG",
            },
        )

    if current_file.name == "annex-106-iso9001-matrix-full.html":
        replace_text_fragments(
            doc,
            {
                "QMS + Ban lãnh đạo": "QMS + CEO",
                "Kinh doanh/Program + QMS": "D-SCS + QMS",
                "QMS + Chủ quá trình": "QMS + OPS_SCOPE_OWNERS",
                "Tổng / Chung Manager": "CEO",
                "Tổng / Chung Manager + QMS": "CEO + QMS",
                "Tổng / Chung Manager + QMS + HR": "CEO + QMS + HR",
                "Chủ quá trình + QMS": "OPS_SCOPE_OWNERS + QMS",
                "Top quản lý + Chủ quá trình": "CEO + OPS_SCOPE_OWNERS",
                "Quy trình Owner + QA/ENG": "OPS_SCOPE_OWNERS + QA + D-ENG",
                "HR / Sản xuất / QA / HSE": "HR / D-PROD / QA / EHS",
                "QMS + HR + Chủ quá trình": "QMS + HR + OPS_SCOPE_OWNERS",
                "HR + Dòng / Chuyền quản lý": "HR + DIRECT_LINE_MGRS",
                "Chủ quá trình": "OPS_SCOPE_OWNERS",
                "QMS + chủ dữ liệu": "QMS + MR_REPORT_OWNERS",
                "Kinh doanh/Program": "D-SCS",
                "QA + Quy trình Owner": "QA + OPS_SCOPE_OWNERS",
                "Line managers": "DIRECT_LINE_MGRS",
                "QMS Engineer / doc điều phối viên": "QMS[DC]",
                "Dept Trưởng bộ phận + Finance": "FUNC_HEADS + FIN",
                "Dept Trưởng bộ phận": "FUNC_HEADS",
                "HR / dòng / chuyền người chịu trách nhiệm": "HR / DIRECT_LINE_MGRS",
                "EHS / khu vực người chịu trách nhiệm": "EHS / AREA_LEADS",
                "Finance team": "D-FIN",
                "QMS / IT / chủ dữ liệu": "QMS / ITA / MR_REPORT_OWNERS",
                "người chịu trách nhiệm chức năng + QMS": "FUNC_OWNERS + QMS",
                "người chịu trách nhiệm chức năng + QMS + CEO": "FUNC_OWNERS + QMS + CEO",
                "chủ tài liệu": "QMS[DC]",
            },
        )

    if current_file.name == "annex-106-iso9001-matrix-full.html":
        replace_text_fragments(
            doc,
            {
                "QA/ENG": "QA / D-ENG",
            },
        )

    if current_file.name == "annex-105-process-map-detailed.html":
        replace_text_fragments(
            doc,
            {
                "Sales → ENG": "D-SCS → D-ENG",
                "ENG → Planning": "D-ENG → D-PPC",
                "Sales → ENG → Planning; Customer phản hồi → QA / QMS.": "D-SCS → D-ENG → D-PPC; phản hồi khách hàng → QA / QMS.",
                "Kinh doanh → ENG": "D-SCS → D-ENG",
                "ENG → Kế hoạch": "D-ENG → D-PPC",
                "Kinh doanh → ENG → Kế hoạch; Customer phản hồi → QA / QMS.": "D-SCS → D-ENG → D-PPC; phản hồi khách hàng → QA / QMS.",
                "ENG → Planning / Production / QA.": "D-ENG → D-PPC / D-PROD / QA.",
                "ENG → Kế hoạch / Sản xuất / QA.": "D-ENG → D-PPC / D-PROD / QA.",
                "Kinh doanh â†’ ENG": "D-SCS â†’ D-ENG",
                "Sales â†’ ENG": "D-SCS â†’ D-ENG",
                "ENG â†’ Káº¿ hoáº¡ch": "D-ENG â†’ D-PPC",
                "Kinh doanh â†’ ENG â†’ Káº¿ hoáº¡ch; Customer pháº£n há»“i â†’ QA / QMS.": "D-SCS â†’ D-ENG â†’ D-PPC; Customer pháº£n há»“i â†’ QA / QMS.",
                "Sales â†’ ENG â†’ Planning; Customer pháº£n há»“i â†’ QA / QMS.": "D-SCS â†’ D-ENG â†’ D-PPC; Customer pháº£n há»“i â†’ QA / QMS.",
                "ENG â†’ Káº¿ hoáº¡ch / Sáº£n xuáº¥t / QA.": "D-ENG â†’ D-PPC / D-PROD / QA.",
                "PUR / WHS / QA / ENG.": "D-PUR / D-WHS / QA / D-ENG.",
            },
        )

    if current_file.name == "annex-116-m365-folder-structure-blueprint.html":
        replace_text_fragments(
            doc,
            {
                "{Dept} = QMS, QA, ENG, PRO, PUR, HR, WHS, MNT, PLA, SAL, FIN, HSE, IT, CNC, OPS": "{Dept} = D-SCS, D-ENG, D-PROD, D-PPC, D-QUAL, D-SCM, D-PUR, D-WHS, D-TCR, D-LOG, D-FIN, D-HR, D-EHS, D-IT, D-ERP",
                "Planning, QA, ENG, Production, WHS theo nhu cầu": "D-PPC / QA / D-ENG / D-PROD / D-WHS theo nhu cầu",
                "ENG cập nhật, QA/OPS đọc": "D-ENG cập nhật, QA / D-PROD đọc",
                "QMS + Người phụ trách quy trình": "QMS + OPS_SCOPE_OWNERS",
                "IT + QMS chủ dữ liệu": "ITA + QMS + MR_REPORT_OWNERS",
                "Dept Head": "FUNC_HEADS",
            },
        )

    if current_file.name == "annex-133-m365-records-site-topology-library-and-folder-blueprint.html":
        replace_text_fragments(
            doc,
            {
                "DEP-SAL": "D-SCS",
                "SAL RFQ": "D-SCS RFQ",
                "DEP-ENG": "D-ENG",
                "ENG team": "D-ENG",
            },
        )

    if current_file.name == "annex-133-m365-records-site-topology-library-and-folder-blueprint.html":
        replace_text_fragments(
            doc,
            {
                "MÃ£ phÃ²ng ban chuáº©n: EXEC, QMS, QA, ENG, PRO, SCM, SAL, FIN, HR, EHS, IT, ERP.": "MÃ£ phÃ²ng ban chuáº©n: D-EXEC, D-SCS, D-ENG, D-PROD, D-PPC, D-QUAL, D-SCM, D-PUR, D-WHS, D-TCR, D-LOG, D-FIN, D-HR, D-EHS, D-IT, D-ERP.",
                "ENG edit, QA/OPS Ä‘á»c.": "D-ENG edit, QA / D-PROD Ä‘á»c.",
            },
        )

    if current_file.name == "annex-134-m365-records-provisioning-permissions-and-automation-architecture.html":
        replace_text_fragments(
            doc,
            {
                "SAL khu vực": "D-SCS khu vực",
                "QMS-SAL-Biên tập viên": "QMS-D-SCS-Biên tập viên",
                "05-SAL": "05-D-SCS",
            },
        )

    if current_file.name == "annex-135-m365-operational-records-file-plan-by-department-role-and-job.html":
        replace_text_fragments(
            doc,
            {
                "ANNEX-135 - M365 operational records file plan by department, role and job": "ANNEX-135 - File plan hồ sơ vận hành M365 theo phòng ban, vai trò và công việc",
                "Engineering Lead, DFM Engineer, Estimator interface": "ENGM / DFM / EST interface",
                "6.4 ENG": "6.4 D-ENG",
                "ENG team": "D-ENG",
                "6.7 SAL": "6.7 D-SCS",
                "SAL RFQ interfaces": "D-SCS RFQ interfaces",
                "DEP-SAL": "D-SCS",
                "DEP-ENG": "D-ENG",
            },
        )

    if current_file.name == "annex-135-m365-operational-records-file-plan-by-department-role-and-job.html":
        replace_text_fragments(
            doc,
            {
                "Giữ fixture validation notes ngoài ENG zone": "Giữ fixture validation notes ngoài vùng D-ENG",
                "Kế hoạch, QA, ENG, PRO, SCM theo gate; external share chỉ qua approved package": "D-PPC, QA, D-ENG, D-PROD, D-SCM theo gate; external share chỉ qua approved package",
                "Giá»¯ fixture validation notes ngoÃ i ENG zone": "Giá»¯ fixture validation notes ngoÃ i vÃ¹ng D-ENG",
                "Káº¿ hoáº¡ch, QA, ENG, PRO, SCM theo gate; external share chá»‰ qua approved package": "D-PPC, QA, D-ENG, D-PROD, D-SCM theo gate; external share chá»‰ qua approved package",
            },
        )

    if current_file.name == "annex-302-approved-materials-list.html":
        replace_text_fragments(
            doc,
            {
                "Engineering + QA + khi áp dụng: Customer": "D-ENG + QA + Customer khi áp dụng",
                "Customer + Engineering + QA": "Customer + D-ENG + QA",
                "ENG Manager + QA Manager phê duyệt §9 quy trình.": "ENGM + QA phê duyệt §9 quy trình.",
                "Engineering phải": "D-ENG phải",
                "Engineering mở review": "D-ENG mở review",
                "Engineering, DFM, báo giá và hoạch định gia công.": "D-ENG, DFM, EST và hoạch định gia công.",
                "ENG Manager": "ENGM",
                "Purchasing Manager": "SCM",
                "General Manager (EXE-01)": "CEO",
            },
        )

    if current_file.name == "annex-302-approved-materials-list.html":
        replace_text_fragments(
            doc,
            {
                "confirm với Purchasing. ENG approval": "confirm với D-PUR. D-ENG approval",
                "với ENG approval riêng.": "với D-ENG approval riêng.",
                "ENG review; thường OK": "D-ENG review; thường OK",
                "Downgrade — ENG+QA+Customer": "Downgrade — D-ENG+QA+Customer",
                "D-D-ENG + QA required": "D-ENG + QA required",
                "Auto upgrade / ENG+QA / ENG+QA+Customer / Block": "Auto upgrade / D-ENG+QA / D-ENG+QA+Customer / Block",
                "confirm vá»›i Purchasing. ENG approval": "confirm vá»›i D-PUR. D-ENG approval",
                "CL-B structural non-wetted: ENG approval.": "CL-B structural non-wetted: D-ENG approval.",
                "ENG approval required.": "D-ENG approval required.",
                "Engineering PHáº¢I cáº­p nháº­t": "D-ENG PHáº¢I cáº­p nháº­t",
                "Chá»‰ CL-C non-wetted vá»›i ENG approval riÃªng.": "Chá»‰ CL-C non-wetted vá»›i D-ENG approval riÃªng.",
                "ENG + QA required": "D-ENG + QA required",
                "ENG review; thÆ°á»ng OK": "D-ENG review; thÆ°á»ng OK",
                "ENG + Customer approval": "D-ENG + Customer approval",
                "Downgrade â€” ENG+QA+Customer": "Downgrade â€” D-ENG+QA+Customer",
            },
        )

    if current_file.name == "annex-501-dispatch-capacity-wip-rules.html":
        replace_text_fragments(
            doc,
            {
                "Planner, Nhân viên điều phối, Supervisor và Ops Manager": "PPL / DIRECT_LINE_MGRS / PD",
                "Ops Manager phải ra quyết định bằng văn bản": "PD phải ra quyết định bằng văn bản",
                "Ops Manager hoặc cao hơn theo ma trận thẩm quyền": "PD hoặc cao hơn theo ANNEX-120",
                "Planner Lead / Ops Manager": "PPL / PD",
                "Planner / Nhân viên điều phối": "PPL",
                "chờ ENG": "chờ D-ENG",
                "Requests Supervisor tự hiểu": "để DIRECT_LINE_MGRS tự hiểu",
                "Planner + Supervisor": "PPL + DIRECT_LINE_MGRS",
                "ENG phát hành controlled copy": "D-ENG phát hành controlled copy",
                "ENG release set mới": "D-ENG release set mới",
                "QA/WHS/ENG": "QA / D-WHS / D-ENG",
                "được ENG/QA chấp thuận": "được D-ENG / QA chấp thuận",
            },
        )

    if current_file.name == "annex-502-gate-mrr-and-execution-synchronization-pack.html":
        replace_text_fragments(
            doc,
            {
                "Sản xuất Supervisor / QA": "WKM / SL / QA",
            },
        )

    if current_file.name == "annex-801-competency-levels-and-certification-rules.html":
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr("HR", "QA", "PD"), current_file, registry)
        set_row_cell_html(doc, "C04", 3, render_token_cluster(["FRONTLINE_LEADS", "PPL", "QA", "D-WHS", "D-ENG"], current_file, registry))
        set_row_cell_html(doc, "C06", 3, render_token_cluster(["FRONTLINE_LEADS", "QA", "D-ENG", "FUNC_HEADS"], current_file, registry))
        set_row_cell_html(doc, "C10", 3, render_token_cluster(["D-PROD", "QA", "D-PPC", "D-WHS", "D-ENG"], current_file, registry))
        set_row_cell_html(doc, "C13", 3, render_token_cluster(["D-ENG", "QA", "FRONTLINE_LEADS", "D-PUR"], current_file, registry))
        set_row_cell_html(doc, "C14", 3, render_token_cluster(["OPR", "SET", "QA", "D-ENG", "CAM"], current_file, registry))
        set_row_cell_html(doc, "C15", 3, render_token_cluster(["D-ENG", "QA", "SET", "D-PUR"], current_file, registry))
        set_row_cell_html(doc, "C16", 3, render_token_cluster(["QA", "QC", "SET", "QCL"], current_file, registry))
        set_row_cell_html(doc, "C17", 3, render_token_cluster(["OPR", "SET", "CAM", "MNT"], current_file, registry))
        set_row_cell_html(doc, "C19", 3, render_token_cluster(["DIRECT_LINE_MGRS", "OJT_COACHES", "FUNC_HEADS"], current_file, registry))
        replace_text_fragments(
            doc,
            {
                "supervisor xác nhận.": "DIRECT_LINE_MGRS xác nhận.",
                "Trưởng bộ phận và QA/HR": "FUNC_HEADS và QA / HR",
                "QA/Trưởng bộ phận": "QA / FUNC_HEADS",
                "trưởng bộ phận + QA/HR": "FUNC_HEADS + QA / HR",
            },
        )

    if current_file.name == "sop-802-incident-near-miss-and-ehs.html":
        role_rows = doc.xpath('//h2[@id="p4"]/following-sibling::div[contains(@class,"table-card")][1]//table/tbody/tr')
        if len(role_rows) >= 2:
            cells = role_rows[1].xpath("./td")
            if cells:
                set_element_html(cells[0], render_spec(bundle("AREA_LEADS"), current_file, registry))
        replace_text_fragments(
            doc,
            {
                "Employee / Witness": "Người chứng kiến / người phát hiện",
                "Supervisor": "cấp quản lý hiện trường",
                "scene control": "kiểm soát hiện trường",
                "return-to-use": "mở lại sử dụng",
                "follow-up": "theo dõi sau sự cố",
                "product-impact": "ảnh hưởng tới sản phẩm",
                "quality flow": "luồng chất lượng",
                "unsafe condition": "điều kiện không an toàn",
            },
        )

    if current_file.name == "annex-110-dashboard-kpi-dictionary-and-data-model.html":
        replace_text_fragments(
            doc,
            {
                "Mỗi KPI có 01 người chịu trách nhiệm nghiệp vụ chịu trách nhiệm hành động và 01 người chịu trách nhiệm dữ liệu chịu trách nhiệm quy trình / luồng công việc/làm mới.": "Mỗi KPI phải chỉ rõ 01 role trong MR_REPORT_OWNERS chịu trách nhiệm hành động và 01 role hoặc chức năng xác nhận nguồn dữ liệu (ITA / ESA hoặc FUNC_OWNERS tùy source).",
                "người chịu trách nhiệm nghiệp vụ và người chịu trách nhiệm dữ liệu ký chấp thuận": "MR_REPORT_OWNERS và role xác nhận nguồn dữ liệu ký chấp thuận",
                "người chịu trách nhiệm nghiệp vụ": "role chủ trì KPI",
                "người chịu trách nhiệm dữ liệu": "role xác nhận nguồn dữ liệu",
                "IT chủ dữ liệu": "ITA / ESA",
            },
        )

    if current_file.name == "annex-503-cnc-operating-model-and-role-boundary.html":
        set_row_first_cell_html(doc, "Đo lường & Hiệu chuẩn Specialist", chips(expr("MCS")))
        set_row_first_cell_html(doc, "Nội bộ Đánh giá viên (Outsource)", chips(expr("IAO")))
        replace_text_fragments(
            doc,
            {
                "người chịu trách nhiệm nghiệp vụ": "FUNC_OWNERS",
                "Quy trình Owner": "OPS_SCOPE_OWNERS",
            },
        )

    if current_file.name == "annex-504-tier-meeting-cadence-and-escalation-standard-work.html":
        replace_text_fragments(
            doc,
            {
                "Ban điều hành nhà máy + chủ quy trình": "TOP_MGMT + OPS_SCOPE_OWNERS",
                "Trưởng nhóm Kế hoạch": "PPL",
                "chủ quá trình liên quan": "OPS_SCOPE_OWNERS",
            },
        )

    if current_file.name == "wi-101-digital-online-forms-and-approvals.html":
        replace_text_fragments(
            doc,
            {
                "Quy trình Owner": "OPS_SCOPE_OWNERS",
            },
        )

    if current_file.name == "TRN-OPS-03.html":
        set_meta_row_spec(doc, ("owner", "chá»§ sá»Ÿ há»¯u"), expr("QA", "QMS"), current_file, registry)
        set_meta_row_spec(doc, ("approved", "phÃª duyá»‡t"), expr("CEO"), current_file, registry)
        for row in doc.xpath('//div[contains(@class,"fh-kv")]'):
            spans = row.xpath("./span")
            if len(spans) < 2:
                continue
            label = fold_text(element_text(spans[0]))
            if "owner" in label or "chu so huu" in label:
                set_element_html(spans[-1], render_token_cluster(["HR", "FUNC_HEADS", "QA", "QMS"], current_file, registry, joiner=" + "))
            elif "approved" in label or "phe duyet" in label:
                set_element_html(spans[-1], render_spec(expr("CEO"), current_file, registry))

    if current_file.name == "annex-607-quality-culture-and-ethics-rules.html":
        replace_text_fragments(
            doc,
            {
                "Kênh 1: Shift Leader / Trưởng bộ phận trực tiếp.": "Kênh 1: DIRECT_LINE_MGRS.",
            },
        )


    if current_file.name == "ojt-tracker.html":
        replace_text_fragments(
            doc,
            {
                " / WHS / PUR / ": " / D-WHS / D-PUR / ",
                "? Production Planner": "? PPL",
                "? Kho / Nh?n h?ng": "? WAR / D-WHS",
                "? Mua h?ng / Gia c?ng ngo?i": "? BUY / D-PUR",
            },
        )

    if current_file.name == "role-roadmaps.html":
        replace_text_fragments(
            doc,
            {
                "Production Engineer / IE": "PIE",
                "Chu?i cung ?ng Manager": "SCM",
                "OJT Planner": "OJT PPL",
                "OJT Mua h?ng": "OJT BUY / D-PUR",
                "OJT Kho": "OJT WAR / D-WHS",
                "D-WHS / Nh?n h?ng; H?u c?n": "D-WHS / Nh?n h?ng; D-LOG",
            },
        )

    if current_file.name == "index.html" and "10-Training-Academy\01-Competency-System" in str(current_file):
        replace_text_fragments(
            doc,
            {
                "S?n xu?t matrix": "D-PROD matrix",
                "Quality matrix": "D-QUAL matrix",
                "Kho matrix": "D-WHS matrix",
            },
        )

    if current_file.name == "training-matrix.html":
        replace_text_fragments(
            doc,
            {
                "Kho / Giao v?n": "D-WHS / D-LOG",
                "TRN-MTX-07 ? Training Matrix ? B?o tr? | HESEM OS": "TRN-MTX-07 ? Training Matrix ? D-MNT | HESEM OS",
            },
        )

    if current_file.name == "skill-matrix-bonus.html":
        replace_text_fragments(
            doc,
            {
                "D-WHS & Giao v?n": "D-WHS & D-LOG",
                "Kho/H?u c?n": "D-WHS / D-LOG",
            },
        )

    if current_file.name == "C15-L2.html":
        replace_text_fragments(
            doc,
            {
                "Mua h?ng ? ": "D-PUR ? ",
                " ? K? ho?ch": " ? PPL",
            },
        )

    if current_file.name == "sop-201-order-fulfillment-rfq-to-cash.html":
        replace_text_fragments(
            doc,
            {
                "B?o gi?, K? thu?t, ho?ch ??nh, Mua h?ng, s?n xu?t, QA/QC, ": "EST, D-ENG, PPL, D-PUR, D-PROD, QA / QC, ",
                " v? t?i ch?nh ??u c?": " v? D-FIN ??u c?",
            },
        )


def normalize_jd_file(
    doc_path: Path,
    registry: dict,
    aliases: dict[str, dict],
    titles: dict[str, str],
    phrase_replacements: dict[str, str],
    profiles: dict[str, dict[str, object]],
) -> None:
    source_text = doc_path.read_text(encoding="utf-8")
    body_source = extract_jd_body_source(source_text)
    doc = html.fromstring(body_source, parser=html.HTMLParser(encoding="utf-8"))

    role_code = None
    role = None
    for code, meta in registry["roles"].items():
        if Path(meta["jd_path"]).name == doc_path.name:
            role_code = code
            role = meta
            break
    if role_code is None or role is None:
        return
    profile = profiles.get(role_code, {})
    jd_departments = list(profile.get("departments", []))

    title_text = f'{role["jd_code"]} — {role["title_en"]} | HESEM QMS'
    strong_text = f'{role["jd_code"]} — {role["title_en"]}'

    title_el = doc.find(".//title")
    if title_el is None:
        head = doc.find(".//head")
        if head is not None:
            title_el = etree.SubElement(head, "title")
    if title_el is not None:
        title_el.text = title_text

    strong = doc.xpath('//div[contains(@class,"title")]/strong')
    if strong:
        strong[0].text = strong_text
    subtitle = doc.xpath('//div[contains(@class,"title")]/span[contains(@class,"sub-vn")]')
    if subtitle:
        subtitle[0].text = role["title_vi"]

    update_meta_rows(doc, doc_path, registry, aliases, {"HR Manager": expr("HR"), "Chief Executive Officer": expr("CEO")})
    if jd_departments:
        set_meta_row_spec(doc, ("owner", "chủ sở hữu"), expr(*jd_departments), doc_path, registry)
    set_meta_row_spec(doc, ("approved", "phe duyet", "phê duyệt"), expr("CEO"), doc_path, registry)

    for row in doc.xpath('//div[contains(@class,"meta")]/div[contains(@class,"row")]'):
        spans = row.xpath("./span")
        if len(spans) < 2:
            continue
        label = fold_text(element_text(spans[0]))
        if "code" in label or label.startswith("ma"):
            value_el = spans[-1]
            for child in list(value_el):
                value_el.remove(child)
            value_el.text = role["jd_code"]

    position_row = None
    stale_rows: list[etree._Element] = []
    for row in doc.xpath('//table[contains(@class,"require-table")]//tr'):
        cells = row.xpath("./th|./td")
        if len(cells) < 2:
            continue
        label = fold_text(element_text(cells[0]))
        value_el = cells[1]
        if label.startswith("ma vi tri"):
            for child in list(value_el):
                value_el.remove(child)
            value_el.text = role["jd_code"]
            position_row = row
        elif label.startswith("ma vai tro dung trong sop/raci") or label.startswith("mu quan tri co the gan"):
            stale_rows.append(row)
        elif label.startswith("chuc danh theo tai lieu"):
            for child in list(value_el):
                value_el.remove(child)
            value_el.text = role["title_en"]

    for stale_row in stale_rows:
        parent = stale_row.getparent()
        if parent is not None:
            parent.remove(stale_row)

    tail_row = position_row
    if tail_row is not None:
        role_row = html.fragment_fromstring("<tr><th>Mã vai trò dùng trong SOP/RACI</th><td></td></tr>")
        tail_row.addnext(role_row)
        set_element_html(role_row.xpath("./td")[0], render_spec(expr(role_code), doc_path, registry))
        tail_row = role_row

    hats = role.get("hats_allowed", [])
    if hats and tail_row is not None:
        hat_row = html.fragment_fromstring("<tr><th>Mũ quản trị có thể gắn</th><td></td></tr>")
        tail_row.addnext(hat_row)
        set_element_html(hat_row.xpath("./td")[0], render_spec(expr(*[f"{role_code}[{hat}]" for hat in hats]), doc_path, registry))

    update_role_cells(doc, doc_path, registry, aliases, {})
    update_plain_titles(doc, titles)
    replace_text_fragments(doc, phrase_replacements)
    normalize_jd_structured_rows(doc, doc_path, role_code, registry, profiles)
    normalize_jd_preface_block(doc, doc_path, role_code, registry, profiles)
    replace_text_fragments(doc, jd_role_replacements(role_code, role))
    apply_jd_specific_tweaks(doc, role_code, doc_path, registry)
    replace_text_fragments(doc, COMMON_LABEL_REPLACEMENTS)
    replace_text_fragments_filtered(doc, COMMON_PROSE_REPLACEMENTS, {"a"})
    replace_text_fragments_filtered(doc, JD_PROSE_REPLACEMENTS, {"a"})
    replace_regex_patterns(doc, COMMON_REGEX_REPLACEMENTS, {"a"})
    replace_regex_patterns(doc, JD_REGEX_REPLACEMENTS, {"a"})
    update_role_cells(doc, doc_path, registry, aliases, {})
    normalize_organization_shortlinks(doc, doc_path)
    normalize_reference_links(doc, doc_path, registry)
    normalize_local_link_separators(doc)
    normalize_jd_purpose_intro(doc, role_code, role, doc_path.name)
    replace_text_fragments_filtered(doc, COMMON_PROSE_REPLACEMENTS, {"a"})
    replace_text_fragments_filtered(doc, JD_PROSE_REPLACEMENTS, {"a"})
    replace_regex_patterns(doc, COMMON_REGEX_REPLACEMENTS, {"a"})
    replace_regex_patterns(doc, JD_REGEX_REPLACEMENTS, {"a"})
    set_require_row_value(doc, ("chuc danh theo tai lieu",), value_text=role["title_en"])

    container = doc.xpath('//div[contains(@class,"container")]')
    body_html = html.tostring(container[0] if container else doc, encoding="unicode", method="html")
    head_assets = extract_jd_head_assets(source_text, doc_path)
    document = build_html_document(title_text, head_assets, body_html)
    document = link_bundle_tokens_in_html(document, doc_path, registry)
    document = link_published_actor_terms_in_html(document, doc_path, registry)
    document = normalize_hybrid_actor_clusters_in_html(document, doc_path, registry)
    document = cleanup_known_render_artifacts_in_html(document)
    doc_path.write_text(document, encoding="utf-8")


def normalize_controlled_file(
    doc_path: Path,
    registry: dict,
    aliases: dict[str, dict],
    titles: dict[str, str],
    overrides: dict[str, dict],
    file_overrides: dict[str, dict[str, dict]],
    phrase_replacements: dict[str, str],
) -> None:
    doc = parse_html_document(doc_path.read_text(encoding="utf-8"))
    local_overrides = dict(overrides)
    local_overrides.update(file_overrides.get(doc_path.name, {}))
    update_meta_rows(doc, doc_path, registry, aliases, local_overrides)
    update_role_cells(doc, doc_path, registry, aliases, local_overrides)
    update_plain_titles(doc, titles)
    normalize_department_handbook_header(doc, doc_path, registry)
    replace_text_fragments(doc, phrase_replacements)
    apply_file_specific_tweaks(doc, doc_path, registry)
    replace_text_fragments(doc, COMMON_LABEL_REPLACEMENTS)
    replace_text_fragments_filtered(doc, COMMON_PROSE_REPLACEMENTS, {"a"})
    replace_regex_patterns(doc, COMMON_REGEX_REPLACEMENTS, {"a"})
    update_meta_rows(doc, doc_path, registry, aliases, local_overrides)
    update_role_cells(doc, doc_path, registry, aliases, local_overrides)
    normalize_organization_shortlinks(doc, doc_path)
    normalize_reference_links(doc, doc_path, registry)
    normalize_local_link_separators(doc)
    document = serialize_html_document(doc)
    document = link_bundle_tokens_in_html(document, doc_path, registry)
    document = link_published_actor_terms_in_html(document, doc_path, registry)
    document = normalize_hybrid_actor_clusters_in_html(document, doc_path, registry)
    document = cleanup_known_render_artifacts_in_html(document)
    doc_path.write_text(document, encoding="utf-8")


def jd_role_replacements(role_code: str, role_meta: dict) -> dict[str, str]:
    replacements: dict[str, str] = {}
    hats = role_meta.get("hats_allowed", [])
    if "DC" in hats:
        replacements["Document Controller"] = f"{role_code}[DC]"
    if "LA" in hats:
        replacements["Lead Auditor"] = f"{role_code}[LA]"
    if "CI" in hats:
        replacements["Continuous Improvement Lead"] = f"{role_code}[CI]"
    ic_hats = [hat for hat in hats if hat.startswith("IC-")]
    if ic_hats:
        replacements["Incident Commander"] = f"{role_code}[{ic_hats[0]}]"
    if role_code == "QA":
        replacements["QA Engineer"] = "QE"
    return replacements


def apply_jd_specific_tweaks(doc: etree._Element, role_code: str, current_file: Path, registry: dict) -> None:
    def chips(spec: dict) -> str:
        return render_spec(spec, current_file, registry)

    profile = ROLE_BOUNDARY_PROFILES.get(role_code, {})
    reports_to = list(profile.get("reports_to", []))
    departments = list(profile.get("departments", []))
    if reports_to and departments:
        manager_html = render_token_cluster(reports_to, current_file, registry)
        dept_html = render_token_cluster(departments, current_file, registry)
        for para in doc.xpath('//div[contains(@class,"backup-card")]/p'):
            text = normalize_ws(element_text(para))
            if text.startswith("Chính / Sơ cấp Backup:"):
                set_element_html(
                    para,
                    f"<b>Chính / Sơ cấp Backup:</b> {manager_html} chỉ định backup theo ANNEX-123; người thay thế phải được OJT, có bàn giao và được phép nhận việc trong phạm vi được giao.",
                )
            elif text.startswith("Thứ cấp / Phụ Backup:"):
                set_element_html(
                    para,
                    f"<b>Thứ cấp / Phụ Backup:</b> {manager_html} phối hợp {dept_html} bố trí nguồn lực dự phòng theo ca/kế hoạch để duy trì công việc và truy vết hồ sơ.",
                )

    if role_code == "DBT":
        replace_text_fragments(
            doc,
            {
                "Team Leader/Quality": "DBL / QA",
            },
        )

    if role_code == "CPT":
        set_row_first_cell_html(doc, "Làm sạch & Đóng gói Supervisor / ca sau", f'{chips(expr("CPS"))} / ca sau')
        replace_text_fragments(
            doc,
            {
                "Supervisor/QA/Hậu cần": "CPS / QA / LOG",
                "Supervisor xác nhận": "CPS xác nhận",
                "Supervisor hoặc Hậu cần": "CPS / LOG",
            },
        )

    if role_code == "CPS":
        set_require_row_value(doc, ("cap vai tro",), value_text="Supervisor")

    if role_code == "DBL":
        set_require_row_value(doc, ("cap vai tro",), value_text="Lead")

    if role_code == "SL":
        set_require_row_value(doc, ("cap vai tro",), value_text="Lead")
        replace_text_fragments(
            doc,
            {
                "Team Leader Ca – Ca 1 / Ca 2": "Shift Leader - Ca 1 / Ca 2",
                "Team Leader / Shift Leader / trưởng nhóm hiện trường": "SL / DBL / CPS / QCL hoặc trưởng nhóm hiện trường tương đương",
            },
        )

    if role_code == "QC":
        replace_text_fragments(
            doc,
            {
                "QC Team Leader/QE": "QCL / QE",
                "QC Team Leader/QA": "QCL / QA",
                "QC Team Leader / QA": "QCL / QA",
                "Sản xuất / QA / QC Team Leader": "Sản xuất / QA / QCL",
            },
        )

    if role_code == "QMS":
        replace_text_fragments(
            doc,
            {
                "QA Engineer / QMS": "QE / QMS",
            },
        )

    if role_code == "QE":
        replace_text_fragments(
            doc,
            {
                "QC Team Leader / QC team": "QCL / đội QC",
                "QA Engineer / QMS": "QE / QMS",
            },
        )

    if role_code == "PPL":
        set_row_first_cell_html(doc, "Chuỗi cung ứng / Nhân viên mua hàng / Kho / Tool kho dụng cụ (tool kho dụng cụ (tool kho dụng cụ (tool crib))", render_token_cluster(["D-SCM", "BUY", "D-WHS", "TOOL"], current_file, registry))
        set_row_first_cell_html(doc, "CNC Workshop Manager / Shift lãnh đạo", render_token_cluster(["WKM", "SL"], current_file, registry))

    if role_code == "WAR":
        set_row_first_cell_html(doc, "Sản xuất / Planner / Tool kho dụng cụ (tool kho dụng cụ (tool kho dụng cụ (tool crib))", render_token_cluster(["D-PROD", "PPL", "TOOL"], current_file, registry))

    if role_code == "HR":
        set_row_first_cell_html(doc, "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên", chips(bundle("DEPLOYMENT_STEERING")))
        set_row_first_cell_html(doc, "Department Trưởng bộ phận", chips(bundle("FUNC_HEADS")))
        replace_text_fragments(
            doc,
            {
                "lên Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên.": "lên CEO.",
            },
        )

    if role_code == "GLP":
        set_row_first_cell_html(doc, "Sản xuất Giám đốc / phòng ban trưởng bộ phận", chips(bundle("FUNC_HEADS")))

    if role_code == "PD":
        set_row_first_cell_html(
            doc,
            "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên",
            render_token_cluster(["CEO", "ENGM", "PPL", "WKM", "CPS", "MNT", "PIE"], current_file, registry),
        )

    if role_code == "QA":
        set_row_first_cell_html(doc, "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên", chips(expr("CEO")))
        set_row_first_cell_html(doc, "Sản xuất Giám đốc / Kỹ thuật Lead", render_token_cluster(["PD", "ENGM"], current_file, registry))
        set_row_first_cell_html(doc, "Customer Dịch vụ / khách hàng", render_token_cluster(["CS", "EST", "Khách hàng"], current_file, registry))

    if role_code == "QE":
        set_row_first_cell_html(doc, "Customer Dịch vụ / khách hàng", render_token_cluster(["CS", "EST", "Khách hàng"], current_file, registry))

    if role_code == "SCM":
        set_row_first_cell_html(doc, "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên", chips(expr("CEO")))
        set_row_first_cell_html(
            doc,
            "Nhân viên mua hàng / Kho / Tool kho dụng cụ (tool kho dụng cụ (tool kho dụng cụ (tool crib)) / Hậu cần",
            render_token_cluster(["BUY", "WAR", "TOOL", "LOG"], current_file, registry),
        )

    if role_code == "CS":
        set_row_first_cell_html(
            doc,
            "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên / key tài khoản",
            render_token_cluster(["CEO", "ENGM", "PPL", "QA", "SCM", "FIN", "WKM", "LOG"], current_file, registry),
        )
        replace_text_fragments(
            doc,
            {
                "lên Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên và các người chịu trách nhiệm liên quan": "lên CEO và các role chịu trách nhiệm liên quan",
            },
        )

    if role_code == "FIN":
        set_row_first_cell_html(doc, "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên", chips(expr("CEO")))
        replace_text_fragments(
            doc,
            {
                "lên Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên.": "lên CEO.",
            },
        )

    if role_code == "EHS":
        set_row_first_cell_html(doc, "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên", render_token_cluster(["CEO", "PD", "WKM", "HR", "QA"], current_file, registry))
        replace_text_fragments(
            doc,
            {
                "lên Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên": "lên CEO / PD",
            },
        )

    if role_code == "ITA":
        replace_text_fragments(
            doc,
            {
                "không là chủ dữ liệu của các giao dịch kinh doanh/sản xuất.": "không giữ quyền sở hữu nội dung dữ liệu nghiệp vụ của các D-code chức năng trong OPS_SCOPE_OWNERS.",
                "Phối hợp Epicor Admin/QMS/người chịu trách nhiệm nghiệp vụ": "Phối hợp ESA / QMS / FUNC_OWNERS",
                "chủ / gốc-data kinh doanh": "nội dung dữ liệu nghiệp vụ",
                "người chịu trách nhiệm nghiệp vụ": "FUNC_OWNERS",
                "người chịu trách nhiệm xác nhận tiếp nhận": "role xác nhận tiếp nhận",
            },
        )
        set_row_first_cell_html(doc, "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên", render_token_cluster(["CEO", "ESA"], current_file, registry))
        set_row_first_cell_html(
            doc,
            "All phòng ban / Người dùng",
            render_token_cluster(["D-SCS", "D-ENG", "D-PROD", "D-QUAL", "D-SCM", "D-FIN", "D-HR", "D-EHS", "D-IT", "Người dùng"], current_file, registry),
        )
        set_row_first_cell_html(doc, "QMS / Finance / HR khi liên quan", render_token_cluster(["QMS", "D-FIN", "D-HR"], current_file, registry))
        replace_text_fragments(
            doc,
            {
                "lên Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên và Governance viên hệ thống Epicor.": "lên CEO / ESA.",
            },
        )

    if role_code == "IAO":
        set_row_first_cell_html(
            doc,
            "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên / Xem xét của lãnh đạo",
            render_token_cluster(["QA", "QMS", "CEO"], current_file, registry),
        )
        set_row_first_cell_html(doc, "HR / Training / QMS", render_token_cluster(["HR", "OJT_COACHES", "QMS"], current_file, registry))
        set_row_first_cell_html(doc, "đã đánh giá Department", chips(bundle("ALL_DEPTS")))

    if role_code == "ESA":
        replace_text_fragments(
            doc,
            {
                "chủ / gốc-data rule": "quy tắc dữ liệu nguồn nghiệp vụ",
                "Các bộ phận vẫn là chủ dữ liệu; Epicor Admin bảo vệ đúng logic hệ thống, quyền và truy vết.": "các D-code chức năng và role trong OPS_SCOPE_OWNERS vẫn sở hữu nội dung dữ liệu nghiệp vụ; ESA bảo vệ logic hệ thống, quyền và truy vết.",
                "System admin không thay chủ nghiệp vụ; vai trò là chuyển yêu cầu nghiệp vụ thành rào chắn / quy định bảo vệ, cấu hình và dữ liệu vận hành tin cậy.": "ESA không thay FUNC_OWNERS; vai trò là chuyển yêu cầu nghiệp vụ thành rào chắn / quy định bảo vệ, cấu hình và dữ liệu vận hành tin cậy.",
                "người chịu trách nhiệm nghiệp vụ": "FUNC_OWNERS",
                "chủ dữ liệu": "FUNC_OWNERS",
                "chủ / gốc-data quản trị": "quản trị dữ liệu nguồn nghiệp vụ",
            },
        )
        set_row_first_cell_html(
            doc,
            "Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên / key Quy trình Owner",
            render_token_cluster(
                ["CEO", "PD", "ENGM", "QA[QMR]", "SCM", "FIN", "HR", "EHS", "ITA", "QMS"],
                current_file,
                registry,
            ),
        )
        replace_text_fragments(
            doc,
            {
                "lên Trưởng / Giám đốc Điều hành Cán bộ / Nhân viên theo mức độ.": "lên ITA / CEO theo mức độ.",
            },
        )

    if role_code == "ESA":
        replace_text_fragments(
            doc,
            {
                "ngÆ°á»i chá»‹u trÃ¡ch nhiá»‡m dá»¯ liá»‡u": "role xÃ¡c nháº­n nguá»“n dá»¯ liá»‡u",
                "khÃ³a ngÆ°á»i chá»‹u trÃ¡ch nhiá»‡m dá»¯ liá»‡u": "khÃ³a role xÃ¡c nháº­n nguá»“n dá»¯ liá»‡u",
            },
        )

    if role_code == "GLP":
        replace_text_fragments(
            doc,
            {
                "ngÆ°á»i chá»‹u trÃ¡ch nhiá»‡m dá»¯ liá»‡u nguá»“n": "role xÃ¡c nháº­n nguá»“n dá»¯ liá»‡u",
            },
        )

    if role_code == "QCL":
        set_require_row_value(doc, ("cap vai tro",), value_text="Lead")


def refresh_workbook(registry: dict, profiles: dict[str, dict[str, object]]) -> None:
    wb = load_workbook(WORKBOOK_PATH)

    ws = wb["Vai tro & Chuc danh"]
    ws.delete_rows(2, ws.max_row)
    headers = ["STT", "Thuật ngữ tiếng Anh", "Bản dịch tiếng Việt", "Dịch (Yes/No)", "Ghi chú"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    for index, code in enumerate(sorted(registry["roles"]), start=1):
        role = registry["roles"][code]
        profile = profiles.get(code, {})
        department_codes = ", ".join(profile.get("departments", []))
        ws.cell(row=index + 1, column=1).value = index
        ws.cell(row=index + 1, column=2).value = role["title_en"]
        ws.cell(row=index + 1, column=3).value = role["title_vi"]
        ws.cell(row=index + 1, column=4).value = "Yes"
        ws.cell(row=index + 1, column=5).value = f'{code} | {role["jd_code"]} | {department_codes or role["department"]}'

    detail_name = "Role code & JD link"
    if detail_name in wb.sheetnames:
        del wb[detail_name]
    detail = wb.create_sheet(detail_name)
    detail_headers = [
        "STT",
        "Role code",
        "Job title English",
        "Chuc danh tieng Viet chuan",
        "Legacy department label",
        "Department codes",
        "Reports to role code",
        "Direct reports role codes",
        "JD code",
        "JD path",
        "Governance hats",
    ]
    for col, header in enumerate(detail_headers, 1):
        detail.cell(row=1, column=col).value = header
    for index, code in enumerate(sorted(registry["roles"]), start=1):
        role = registry["roles"][code]
        profile = profiles.get(code, {})
        detail.cell(row=index + 1, column=1).value = index
        detail.cell(row=index + 1, column=2).value = code
        detail.cell(row=index + 1, column=3).value = role["title_en"]
        detail.cell(row=index + 1, column=4).value = role["title_vi"]
        detail.cell(row=index + 1, column=5).value = role["department"]
        detail.cell(row=index + 1, column=6).value = ", ".join(profile.get("departments", []))
        detail.cell(row=index + 1, column=7).value = ", ".join(profile.get("reports_to", []))
        detail.cell(row=index + 1, column=8).value = ", ".join(profile.get("direct_reports", []))
        detail.cell(row=index + 1, column=9).value = role["jd_code"]
        detail.cell(row=index + 1, column=10).value = role["jd_path"]
        detail.cell(row=index + 1, column=11).value = ", ".join(role.get("hats_allowed", []))

    hats_name = "Governance hats"
    if hats_name in wb.sheetnames:
        del wb[hats_name]
    hats = wb.create_sheet(hats_name)
    for col, header in enumerate(["STT", "Hat code", "Label English", "Nhan de tieng Viet", "Host roles"], 1):
        hats.cell(row=1, column=col).value = header
    for index, hat_code in enumerate(sorted(registry["hats"]), start=1):
        hat = registry["hats"][hat_code]
        hats.cell(row=index + 1, column=1).value = index
        hats.cell(row=index + 1, column=2).value = hat_code
        hats.cell(row=index + 1, column=3).value = hat["label_en"]
        hats.cell(row=index + 1, column=4).value = hat["label_vi"]
        hats.cell(row=index + 1, column=5).value = ", ".join(hat["host_roles"])

    dept_name = "Department code & handbook link"
    if dept_name in wb.sheetnames:
        del wb[dept_name]
    depts = wb.create_sheet(dept_name)
    dept_headers = [
        "STT",
        "Department code",
        "Department title English",
        "Ten tieng Viet chuan",
        "Loai",
        "Parent department",
        "Handbook path",
        "Lead roles",
    ]
    for col, header in enumerate(dept_headers, 1):
        depts.cell(row=1, column=col).value = header
    for index, code in enumerate(sorted(registry.get("departments", {})), start=1):
        dept = registry["departments"][code]
        depts.cell(row=index + 1, column=1).value = index
        depts.cell(row=index + 1, column=2).value = code
        depts.cell(row=index + 1, column=3).value = dept["title_en"]
        depts.cell(row=index + 1, column=4).value = dept["title_vi"]
        depts.cell(row=index + 1, column=5).value = dept.get("type", "department")
        depts.cell(row=index + 1, column=6).value = dept.get("parent_department")
        depts.cell(row=index + 1, column=7).value = dept["handbook_path"]
        depts.cell(row=index + 1, column=8).value = ", ".join(dept.get("lead_roles", []))

    bundle_name = "Bundle glossary"
    if bundle_name in wb.sheetnames:
        del wb[bundle_name]
    bundles_ws = wb.create_sheet(bundle_name)
    bundle_headers = [
        "STT",
        "Bundle code",
        "Label English",
        "Ten / dien giai tieng Viet",
        "Loai",
        "Thanh phan",
        "Duoc dung khi",
        "Khong dung de che",
        "Glossary path",
    ]
    for col, header in enumerate(bundle_headers, 1):
        bundles_ws.cell(row=1, column=col).value = header
    for index, code in enumerate(sorted(registry.get("bundles", {})), start=1):
        meta = bundle_meta(code, registry)
        bundles_ws.cell(row=index + 1, column=1).value = index
        bundles_ws.cell(row=index + 1, column=2).value = code
        bundles_ws.cell(row=index + 1, column=3).value = meta.get("label_en")
        bundles_ws.cell(row=index + 1, column=4).value = meta.get("label_vi")
        bundles_ws.cell(row=index + 1, column=5).value = meta.get("kind")
        bundles_ws.cell(row=index + 1, column=6).value = ", ".join(registry["bundles"][code])
        bundles_ws.cell(row=index + 1, column=7).value = meta.get("use_vi")
        bundles_ws.cell(row=index + 1, column=8).value = meta.get("avoid_vi")
        bundles_ws.cell(row=index + 1, column=9).value = str(BUNDLE_GLOSSARY_PATH.relative_to(ROOT)).replace("\\", "/")

    wb.save(WORKBOOK_PATH)


def scan_unresolved(paths: list[Path]) -> str:
    hints = [
        "All Process Owners / Department Heads",
        "Process Owners / Department Heads",
        "All Quy tr??nh Owner / Department Tr?????ng b??? ph???n",
        "Quy tr??nh Owner / Department Tr?????ng b??? ph???n",
        "Process Owner",
        "Department Head",
        "Department head",
        "Functional Head",
        "Responsible Person",
        "Document Controller",
        "Data Owner",
        "Data Owners",
        "IT Data Owner",
        "KPI owner",
        "Top Management",
        "QA/QMS",
        "QMS/QA",
        "QA Lead",
        "QC Lead",
        "QA Engineer",
        "QMS Manager",
        "IT Manager",
        "Sales Manager",
        "Engineering Manager",
        "Production Supervisor",
        "Lead Auditor",
        "Warehouse Supervisor",
        "Metrology Lead",
        "Incident Commander",
        "Continuous Improvement Lead",
        "Commercial Responsible Person",
        "Team Leader",
        "Area Supervisor",
        "Team Leader / Supervisor",
        "Trainer / Mentor",
        "Champion Lead",
        "Dept Managers",
        "Program Lead",
        "Cutover Lead",
        "Command Center Lead",
        "Executive Sponsor",
        "Publisher / IT",
        "QA Final",
        "QMS-Owners / QMS-IT-Administrators",
        "QMS + L??nh ?????o",
        "L??nh ?????o + QMS",
        "D??ng / Chuy???n Manager",
        "HR / Ng?????i ????o t???o",
        "HR + Dept Tr?????ng / ?????u",
        "SAL",
        "ENG",
        "PRO/QA",
        "WHS/SAL",
        "Business Owner",
        "System Owner",
        "Change Owner",
        "Approval Board",
        "Engineering Configuration Lead",
        "Maintenance / Engineering",
        "ng?????i ch???u tr??ch nhi???m KPI",
        "ng?????i ch???u tr??ch nhi???m ch???c n??ng",
        "ng?????i ch???u tr??ch nhi???m d??? li???u",
        "ng?????i ch???u tr??ch nhi???m nghi???p v???",
        "ng?????i ch???u tr??ch nhi???m h??? th???ng",
        "ch??? d??? li???u",
        "Finance team",
        "All site nh??n s???",
        "All ch??? d??? li???u",
        "QMS Engineer / doc ??i???u ph???i vi??n",
        "QA/HR",
        "QA/HRD",
        "QA/Tr?????ng b??? ph???n",
        "Tr?????ng b??? ph???n v?? QA/HR",
        "Shift Leader / Tr?????ng b??? ph???n tr???c ti???p",
        "L??nh ?????o / Quy tr??nh Owner",
        "Dept Tr?????ng b??? ph???n",
        "Reviewer ?????c l???p",
        "Supervisor / Lead",
        "Request ?????u m???i ph??? tr??ch",
        "Foreman",
        "Department Manager",
        "QA lead",
        "QA Lead",
        "Team Leader/Foreman",
        "Department Manager / Foreman",
        "HR/TRN",
        "SALES/PLN",
        "OPS/QA",
        "Setup Lead",
        "Operator + Setup Lead",
        "Operator + Setup Lead + QA",
        "Supervisor + QA",
        "QC + Setup",
    ]
    lines: list[str] = []
    parser = html.HTMLParser(encoding="utf-8")
    for path in sorted(paths):
        doc = html.fromstring(path.read_text(encoding="utf-8"), parser=parser)
        focus = doc.xpath(
            '//div[contains(@class,"meta")]'
            ' | //div[contains(@class,"preface-block")]'
            ' | //div[contains(@class,"role-note")]'
            ' | //div[contains(@class,"note-blue")]'
            ' | //p'
            ' | //li'
            ' | //th'
            ' | //td'
            ' | //span[contains(@class,"inline-tag")]'
        )
        text = "\n".join(element_text(node) for node in focus if 0 < len(element_text(node)) <= 220)
        hits = []
        for hint in hints:
            if hint in {"ENG", "SAL"}:
                pattern = rf"(?<![\w-]){re.escape(hint)}(?![\w-])"
            else:
                pattern = rf"(?<!\w){re.escape(hint)}(?!\w)"
            if re.search(pattern, text):
                hits.append(hint)
        if hits:
            lines.append(f"{path.relative_to(ROOT)} :: {', '.join(hits)}")
    report = "\n".join(lines)
    UNRESOLVED_REPORT.write_text(report, encoding="utf-8")
    return report


def main() -> None:
    registry = load_registry()
    profiles = ROLE_BOUNDARY_PROFILES
    aliases = department_alias_map()
    aliases.update(role_alias_map())
    aliases.update(bundle_alias_map())
    for code in registry["roles"]:
        aliases.setdefault(code, expr(code))
    for code in registry.get("departments", {}):
        aliases.setdefault(code, expr(code))
    titles = title_aliases()
    phrase_replacements = {
        "Responsible Person": "ng?????i ch???u tr??ch nhi???m",
        "Top Management": "Ban l??nh ?????o",
        "Top Qu???n l??": "Ban l??nh ?????o",
        "QA/QMS": "QA / QMS",
        "QMS/QA": "QA / QMS",
        "QMS Manager": "QMS",
        "IT Manager": "ITA",
        "IT System Administrator": "ITA",
        "IT System Governance vi??n": "ITA",
        "Engineering Manager": "ENGM",
        "QA Lead": "QA",
        "QC Lead": "QCL",
        "QA Engineer": "QE",
        "Department Tr?????ng b??? ph???n": "FUNC_HEADS",
        "Dept Tr?????ng b??? ph???n": "FUNC_HEADS",
        "Quy tr??nh Owner": "OPS_SCOPE_OWNERS",
        "ch??? qu?? tr??nh": "OPS_SCOPE_OWNERS",
        "ng?????i ch???u tr??ch nhi???m KPI": "MR_REPORT_OWNERS",
        "ng?????i ch???u tr??ch nhi???m ch???c n??ng": "FUNC_OWNERS",
        "Data Owner": "DATA_OWNERS",
        "Data Owners": "DATA_OWNERS",
        "Business Owner": "DATA_OWNERS",
        "System Owner": "SYSTEM_OWNERS",
        "Department Qu??????n l????": "FUNC_HEADS",
        "ng??????????i ch???????u tr????ch nhi???????m nghi???????p v??????": "FUNC_OWNERS",
        "ng??????????i ch???????u tr????ch nhi???????m h??????? th???????ng": "SYSTEM_OWNERS",
        "ng??????????i ch???????u tr????ch nhi???????m d?????? li???????u": "DATA_OWNERS",
        "HR / Department Qu???n l?? / QA": "HR + FUNC_HEADS + QA",
        "QA / QMS + Ops xu???t s???c": "QA + QMS + PIE[CI]",
        "OPS + WHS + QA / QMS + IT": "D-PROD + D-WHS + QA + QMS + D-IT",
        "Finance team": "D-FIN",
        "Tr?????ng nh??m K??? ho???ch": "PPL",
        "QMS Engineer / doc ??i???u ph???i vi??n": "QMS[DC]",
        "All site nh??n s???": "to??n b??? nh??n s??? nh?? m??y",
        "All ch??? d??? li???u": "DATA_OWNERS",
        "QA/HR": "QA + HR",
        "QA/HRD": "QA + HR",
        "HR/TRN": "HR + OJT_COACHES",
        "SALES/PLN": "D-SCS / PPL",
        "OPS/QA": "D-PROD / QA",
        "Epicor Kinetic (Epicor) (Epicor) (Epicor)": "Epicor Kinetic / Epicor ERP",
    }
    overrides = {
        "QMS Engineer / QA Lead": expr("QMS[DC]", "QA[QMR]", joiner=" + "),
        "QA Lead / Quality Engineer": expr("QA", "QE", joiner=" + "),
        "IT Administrator / Governance vi??n h??? th???ng Epicor": expr("ITA", "ESA", joiner=" + "),
        "QMS Engineer / HR Lead": expr("QMS[DC]", "HR", joiner=" + "),
        "K??? thu???t Lead / QA Lead": expr("ENGM", "QA", joiner=" + "),
        "CS / QA Lead": expr("CS", "QA", joiner=" + "),
        "Engineering Lead / Estimator": expr("ENGM", "EST", joiner=" + "),
        "Quality Engineer / QA Manager": expr("QE", "QA", joiner=" + "),
        "Engineering Lead / Process Engineer": expr("ENGM", "PE", joiner=" + "),
        "Supply Chain Manager / QA Manager": expr("SCM", "QA", joiner=" + "),
        "QA Manager / Supply Chain Manager": expr("QA", "SCM", joiner=" + "),
        "Production Planner / Production Director": expr("PPL", "PD", joiner=" + "),
        "CNC Workshop Manager / Shift Leader": expr("WKM", "SL", joiner=" + "),
        "CNC Workshop Manager / Maintenance Technician": expr("WKM", "MNT", joiner=" + "),
        "Engineering Lead / CNC Workshop Manager / QA Manager": expr("ENGM", "WKM", "QA", joiner=" / "),
        "Deburr Team Lead / QA Manager": expr("DBL", "QA", joiner=" + "),
        "Metrology and Calibration Specialist / QA Manager": expr("MCS", "QA", joiner=" + "),
        "Quality Engineer / Metrology and Calibration Specialist": expr("QE", "MCS", joiner=" + "),
        "QA Manager / QC Inspector Lead": expr("QA", "QCL", joiner=" + "),
        "QA Manager / QMS Engineer": expr("QA", "QMS", joiner=" + "),
        "Supply Chain Manager / Warehouse Clerk": expr("SCM", "WAR", joiner=" + "),
        "QA Manager / Cleaning and Packaging Supervisor": expr("QA", "CPS", joiner=" + "),
        "QA Manager / CNC Workshop Manager": expr("QA[PSO]", "WKM", joiner=" + "),
        "HR Manager / QA Manager": expr("HR", "QA", joiner=" + "),
        "EHS Specialist / HR Manager": expr("EHS", "HR", joiner=" + "),
        "Finance Manager / AP-AR and Payments Accountant": expr("FIN", "APAR", joiner=" + "),
        "QA Manager / Production Engineer-IE": expr("QA", "PIE[CI]", joiner=" + "),
        "QMS Engineer / QA Manager": expr("QMS[LA]", "QA[QMR]", joiner=" + "),
        "Production Engineer-IE / QA Manager": expr("PIE[CI]", "QA[QMR]", joiner=" + "),
        "QA + QMS + PIE[CI]": expr("QA", "QMS", "PIE[CI]", joiner=" + "),
        "Document Controller": expr("QMS[DC]"),
        "Lead Auditor": expr("QMS[LA]"),
        "Continuous Improvement Lead": expr("PIE[CI]"),
        "Incident Commander": expr("PD[IC-PROD]"),
        "Engineering Lead/Manager": expr("ENGM"),
        "Maintenance / Engineering": expr("D-MNT", "D-ENG"),
        "Engineering / Maintenance": expr("D-ENG", "D-MNT"),
        "SALES/PLN": expr("D-SCS", "PPL"),
        "PRO/QA": expr("D-PROD", "QA"),
        "OPS/QA": expr("D-PROD", "QA"),
        "WHS/SAL": expr("D-WHS", "D-SCS"),
        "QA/HR": expr("QA", "HR"),
        "QA/HRD": expr("QA", "HR"),
        "HR/TRN": mix("HR", "OJT_COACHES", joiner=" + "),
        "WHS Supervisor": expr("D-WHS"),
        "Warehouse Manager": expr("D-WHS"),
        "Kho Manager": expr("D-WHS"),
        "Giao v???n Lead": expr("D-LOG"),
        "Shipping Lead": expr("D-LOG"),
        "Mua h??ng Lead": expr("D-PUR"),
        "Kinh doanh Manager": expr("D-SCS"),
        "Kinh doanh-Customer D???ch v??? Manager": expr("D-SCS"),
        "Program / K??? ho???ch": expr("PD", "PPL"),
        "QMS + L??nh ?????o": expr("QMS", "CEO", joiner=" + "),
        "L??nh ?????o + QMS": expr("CEO", "QMS", joiner=" + "),
        "QMS-Owners / QMS-IT-Administrators": expr("QMS", "ITA"),
        "Publisher / IT": expr("QMS", "ITA"),
        "Executive Sponsor": expr("CEO"),
        "Cutover Lead": expr("PD"),
        "Program Lead": expr("PD"),
        "Command Center Lead": expr("QMS", "ITA"),
        "Champion / Supervisor": bundle("DEPLOYMENT_LEADS"),
        "Champions / Supervisors": bundle("DEPLOYMENT_LEADS"),
        "Executive Sponsor / Cutover Lead": expr("CEO", "PD"),
        "Cutover Lead / Sponsor": expr("PD", "CEO"),
        "Command Center Lead / QMS": expr("QMS", "ITA"),
        "Command Center Lead / QMS / IT": expr("QMS", "ITA"),
    }
    file_overrides = {
        "sop-802-incident-near-miss-and-ehs.html": {
            "Top Management + EHS": expr("PD", "EHS", joiner=" + "),
        },
        "sop-106-change-and-configuration-management.html": {
            "Change Owner": expr("ENGM", "QA", "PD", "SCM", "ITA", "ESA"),
            "Approval Board": expr("CEO", "QA", "PD", "ENGM", "ITA"),
            "Change Owner + QA Manager": expr("ENGM", "QA", "PD", "ITA", "ESA"),
            "Engineering Configuration Lead": expr("ENGM", "QMS[DC]", "ESA"),
        },
        "competency-assessment-guide.html": {
            "Department Manager / Foreman": bundle("DIRECT_LINE_MGRS"),
            "Ng?????i h?????ng d???n L3???L4": bundle("OJT_COACHES"),
            "HR/Training": expr("HR"),
        },
        "training-matrix-production.html": {
            "Foreman": expr("WKM"),
            "S???n xu???t Tr?????ng nh??m": expr("SL"),
            "M??i bavia/Ho??n thi???n": expr("DBL", "DBT", "CPS", "CPT"),
        },
        "training-matrix-planning-purchasing.html": {
            "Department Manager Chu???i cung ???ng": expr("SCM"),
            "Planner/b??? l???p l???ch": expr("PPL"),
        },
        "training-matrix-warehouse.html": {
            "Team Leader kho": expr("D-WHS"),
            "H???u c???n / Giao v???n Nh??n vi??n": expr("LOG"),
        },
        "drill-setup-firstarticle-ir.html": {
            "Setup/Operator": expr("SET", "OPR", joiner=" + "),
            "Setup": expr("SET"),
            "QC + Setup": expr("QC", "SET", joiner=" + "),
            "Supervisor + QA": mix("FRONTLINE_LEADS", "QA", joiner=" + "),
            "Setup/QC": expr("SET", "QC", joiner=" + "),
        },
        "drill-safety-5s-hazards.html": {
            "Supervisor": bundle("FRONTLINE_LEADS"),
            "Team": bundle("FRONTLINE_LEADS"),
        },
        "drill-ncr-capa-response.html": {
            "Supervisor + QC": mix("FRONTLINE_LEADS", "QC", joiner=" + "),
            "QA/ENG/OPS": expr("QA", "D-ENG", "D-PROD", joiner=" / "),
        },
        "wi-519-job-packet-quick-check-and-pre-run-verification.html": {
            "Operator + Setup Lead": expr("OPR", "SET", joiner=" + "),
            "Setup Lead": expr("SET"),
            "Operator + Setup Lead + QA": expr("OPR", "SET", "QA", joiner=" + "),
        },
        "annex-135-m365-operational-records-file-plan-by-department-role-and-job.html": {
            "Setup Technician, CNC Workshop Manager, Production Engineer/IE": expr("SET", "WKM", "PIE[CI]"),
            "Shift Leader, CNC Operator, Deburr/Clean-Pack supervisors": expr("SL", "OPR", "DBL", "CPS"),
            "Deburr Team Lead, Cleaning and Packaging Supervisor, technicians": expr("DBL", "DBT", "CPS", "CPT"),
        },
    }

    jd_root = ROOT / "02-Tai-Lieu-He-Thong" / "03-Organization" / "03-Job-Descriptions"
    sop_root = ROOT / "03-Tai-Lieu-Van-Hanh" / "01-SOPs"
    annex_targets = [
        ROOT / "03-Tai-Lieu-Van-Hanh" / "03-Reference" / "01-ANNEX-100" / "12-ANNEX-120-Authority-KPI-and-Deputy-Control" / "annex-120-authority-matrix.html",
        ROOT / "03-Tai-Lieu-Van-Hanh" / "03-Reference" / "05-ANNEX-500" / "annex-503-cnc-operating-model-and-role-boundary.html",
    ]

    jd_files = list(jd_root.rglob("jd-*.html"))
    sop_files = list(sop_root.rglob("sop-*.html"))
    controlled_files = sop_files + annex_targets

    for path in jd_files:
        normalize_jd_file(path, registry, aliases, titles, phrase_replacements, profiles)
    for path in controlled_files:
        normalize_controlled_file(path, registry, aliases, titles, overrides, file_overrides, phrase_replacements)

    all_html_files: list[Path] = []
    for root in [ROOT / "03-Tai-Lieu-Van-Hanh", ROOT / "02-Tai-Lieu-He-Thong", ROOT / "10-Training-Academy"]:
        for path in root.rglob("*.html"):
            all_html_files.append(path)
            if path in jd_files or path in controlled_files:
                continue
            normalize_controlled_file(path, registry, aliases, titles, overrides, file_overrides, phrase_replacements)

    generate_bundle_glossary(registry)
    refresh_workbook(registry, profiles)
    report = scan_unresolved(sorted(set(all_html_files + jd_files + controlled_files)))
    print("UPDATED ROLE SYSTEM")
    print("UNRESOLVED: 0" if not report else f"UNRESOLVED: see {UNRESOLVED_REPORT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
