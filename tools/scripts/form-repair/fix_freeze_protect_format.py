#!/usr/bin/env python3
"""
HESEM QMS - Remove freeze panes, sheet protection, and audit format compliance.
Does NOT touch colors, borders, fonts, or logo - only removes freeze/protect
and reports format deviations.
"""
import os, sys, zipfile, shutil, tempfile
from xml.etree import ElementTree as ET
from openpyxl import load_workbook

NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
FONT = "Segoe UI"
COL_WIDTH = 2.33


def fix_xml_panes(filepath):
    """Remove frozen panes at XML level to prevent recovery errors."""
    tmp = tempfile.mkdtemp()
    fixed = False
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            zf.extractall(tmp)
        ws_dir = os.path.join(tmp, 'xl', 'worksheets')
        if os.path.isdir(ws_dir):
            for sf in os.listdir(ws_dir):
                if not sf.endswith('.xml'):
                    continue
                sp = os.path.join(ws_dir, sf)
                ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
                ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
                tree = ET.parse(sp)
                root = tree.getroot()
                svs = root.find(f'{NS}sheetViews')
                if svs is None:
                    continue
                for sv in svs.findall(f'{NS}sheetView'):
                    pane = sv.find(f'{NS}pane')
                    if pane is not None:
                        sv.remove(pane)
                        fixed = True
                    for sel in sv.findall(f'{NS}selection'):
                        if sel.get('pane') and sel.get('pane') != 'topLeft':
                            sv.remove(sel)
                            fixed = True
                    if not sv.findall(f'{NS}selection'):
                        se = ET.SubElement(sv, f'{NS}selection')
                        se.set('activeCell', 'A1')
                        se.set('sqref', 'A1')
                tree.write(sp, xml_declaration=True, encoding='UTF-8')
        if fixed:
            with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                for rd, ds, fs in os.walk(tmp):
                    for ff in fs:
                        fp2 = os.path.join(rd, ff)
                        zf.write(fp2, os.path.relpath(fp2, tmp))
    except Exception as e:
        print(f"    XML error: {e}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    return fixed


def process_workbook(filepath):
    fname = os.path.basename(filepath)
    issues = []

    # Phase 1: XML pane fix
    if fix_xml_panes(filepath):
        issues.append("xml_pane_fix")

    # Phase 2: openpyxl fixes
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, [f"ERROR: {e}"], False

    changed = False

    for sn in wb.sheetnames:
        ws = wb[sn]
        sheet_fixes = []

        # 1. Remove freeze panes
        try:
            if ws.freeze_panes:
                ws.freeze_panes = None
                sheet_fixes.append("unfreeze")
                changed = True
        except:
            pass

        # 2. Remove sheet protection
        if ws.protection.sheet:
            ws.protection = ws.protection.__class__()
            sheet_fixes.append("unprotect")
            changed = True

        # 3. Hide gridlines
        try:
            if hasattr(ws, 'sheet_view') and ws.sheet_view.showGridLines != False:
                ws.sheet_view.showGridLines = False
                sheet_fixes.append("gridlines_hidden")
                changed = True
        except:
            pass

        # 4. Audit format compliance (report only, don't fix)
        # Check fonts
        wrong_fonts = set()
        for row_cells in ws.iter_rows(max_row=min(ws.max_row or 50, 100)):
            for cell in row_cells:
                try:
                    if cell.font and cell.font.name and cell.font.name != FONT:
                        if cell.value is not None:
                            wrong_fonts.add(cell.font.name)
                except:
                    pass

        if wrong_fonts:
            sheet_fixes.append(f"fonts:{','.join(wrong_fonts)}")

        if sheet_fixes:
            issues.append(f"  [{sn}] {', '.join(sheet_fixes)}")

    if changed:
        try:
            wb.save(filepath)
            issues.append("  >> SAVED")
        except Exception as e:
            issues.append(f"  >> SAVE ERROR: {e}")

    wb.close()

    # Phase 3: post-save XML fix
    if changed:
        fix_xml_panes(filepath)

    return fname, issues, changed


def main():
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS - Unfreeze + Unprotect + Format Audit")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    v = '-v' in sys.argv
    total_freeze = total_protect = total_font = 0
    fixed_files = 0

    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        name, issues, changed = process_workbook(fp)

        # Count
        for iss in issues:
            if 'unfreeze' in str(iss):
                total_freeze += 1
            if 'unprotect' in str(iss):
                total_protect += 1
            if 'fonts:' in str(iss):
                total_font += 1

        if changed:
            fixed_files += 1

        status = "FIXED" if changed else "OK"
        print(f"[{i:3d}/{len(files)}] {fn[:55]:55s} {status}")

        if v:
            for iss in issues:
                print(iss)

    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"  Files processed:    {len(files)}")
    print(f"  Files changed:      {fixed_files}")
    print(f"  Freeze removed:     {total_freeze} sheets")
    print(f"  Protection removed: {total_protect} sheets")
    print(f"  Wrong fonts found:  {total_font} sheets")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
