#!/usr/bin/env python3
"""
HESEM QMS - Comprehensive Form Fix v2
Fixes ALL known issues:
  1. sheetViews corruption (recovery errors)
  2. Broken/missing headers on support tabs
  3. Wrong fonts (must be Segoe UI)
  4. Frozen panes removal
  5. Notice bar formatting
"""
import os, sys, copy, zipfile, shutil, tempfile, re
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════ MASTER TEMPLATE CONSTANTS ═══════
FONT_FAMILY = "Segoe UI"

# Colors
HESEM_BLUE  = "2C9CD7"
DARK_NAVY   = "0C2D48"
MED_BLUE    = "1565C0"
WHITE       = "FFFFFF"
LABEL_BG    = "F0F7FC"
META_LABEL  = "E4F0F9"
THIN_BORDER_COLOR = "C5D9E8"
BODY_TEXT   = "1A2733"
SUBTITLE_CLR= "1565C0"
LGRAY       = "7A8FA6"
NOTICE_BG   = "FFF8E1"
GOLD_TEXT   = "7B4F00"
NOTICE_BORDER_CLR = "F9A825"
STRIPE_BG   = "F7FBFF"

# Fills
FILL_HESEM_BLUE = PatternFill('solid', fgColor=HESEM_BLUE)
FILL_DARK_NAVY  = PatternFill('solid', fgColor=DARK_NAVY)
FILL_MED_BLUE   = PatternFill('solid', fgColor=MED_BLUE)
FILL_WHITE      = PatternFill('solid', fgColor=WHITE)
FILL_LABEL      = PatternFill('solid', fgColor=LABEL_BG)
FILL_META_LABEL = PatternFill('solid', fgColor=META_LABEL)
FILL_NOTICE     = PatternFill('solid', fgColor=NOTICE_BG)
FILL_STRIPE     = PatternFill('solid', fgColor=STRIPE_BG)

# Fonts
FONT_TITLE     = Font(name=FONT_FAMILY, size=14, bold=True, color=DARK_NAVY)
FONT_SUBTITLE  = Font(name=FONT_FAMILY, size=8, bold=False, color=SUBTITLE_CLR)
FONT_TAGLINE   = Font(name=FONT_FAMILY, size=7, bold=False, color=LGRAY)
FONT_META_LBL  = Font(name=FONT_FAMILY, size=8, bold=True, color=DARK_NAVY)
FONT_META_VAL  = Font(name=FONT_FAMILY, size=8, bold=False, color=BODY_TEXT)
FONT_SECTION   = Font(name=FONT_FAMILY, size=9, bold=True, color=WHITE)
FONT_COL_HDR   = Font(name=FONT_FAMILY, size=9, bold=True, color=WHITE)
FONT_LABEL     = Font(name=FONT_FAMILY, size=8, bold=True, color=DARK_NAVY)
FONT_BODY      = Font(name=FONT_FAMILY, size=8, bold=False, color=BODY_TEXT)
FONT_NOTICE    = Font(name=FONT_FAMILY, size=7, bold=False, color=GOLD_TEXT)
FONT_AUTONUM   = Font(name=FONT_FAMILY, size=8, bold=True, color=DARK_NAVY)

# Borders
MEDIUM_SIDE = Side(style='medium', color=HESEM_BLUE)
THIN_SIDE   = Side(style='thin', color=THIN_BORDER_COLOR)
NOTICE_SIDE = Side(style='thin', color=NOTICE_BORDER_CLR)
NO_SIDE     = Side(style=None)

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrapText=True)
ALIGN_LEFT   = Alignment(horizontal='left', vertical='center', wrapText=True)
ALIGN_RIGHT  = Alignment(horizontal='right', vertical='center', wrapText=True)

# ═══════ XMLNS for XML editing ═══════
NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'


def detect_ncols(ws):
    """Detect grid width from merged cells in header area."""
    max_col = 0
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 6:
            max_col = max(max_col, mg.max_col)
    if max_col == 0:
        max_col = ws.max_column or 40
    if max_col <= 44:
        return 40
    elif max_col <= 60:
        return 56
    else:
        return 82


def get_header_zones(ncols):
    """Return column boundaries for header zones."""
    if ncols <= 40:
        return {'logo_end': 8, 'title_start': 9, 'title_end': 30,
                'meta_lbl_start': 31, 'meta_lbl_end': 34,
                'meta_val_start': 35, 'meta_val_end': 40}
    elif ncols <= 56:
        return {'logo_end': 11, 'title_start': 12, 'title_end': 42,
                'meta_lbl_start': 43, 'meta_lbl_end': 48,
                'meta_val_start': 49, 'meta_val_end': 56}
    else:
        return {'logo_end': 16, 'title_start': 17, 'title_end': 62,
                'meta_lbl_start': 63, 'meta_lbl_end': 70,
                'meta_val_start': 71, 'meta_val_end': 82}


def has_valid_header(ws, ncols):
    """Check if sheet has valid header structure (rows 1-6)."""
    zones = get_header_zones(ncols)

    # Check if header merges exist
    header_merges = [mg for mg in ws.merged_cells.ranges if mg.min_row <= 5]
    if len(header_merges) < 3:
        return False

    # Check rule stripe (row 6)
    rd6 = ws.row_dimensions.get(6)
    if rd6 and rd6.height and rd6.height <= 3:
        return True

    # Check for title text in header area
    for mg in header_merges:
        cell = ws.cell(row=mg.min_row, column=mg.min_col)
        if cell.value and isinstance(cell.value, str) and len(cell.value) > 5:
            return True

    return False


def read_main_tab_header(ws, ncols):
    """Read header info from main tab to replicate on support tabs."""
    header_info = {
        'title': '', 'subtitle': '', 'tagline': '',
        'meta': {}  # {row: (label, value)}
    }
    zones = get_header_zones(ncols)

    # Read title
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 3 and mg.min_col >= zones['title_start'] and mg.max_col <= zones['title_end']:
            val = ws.cell(row=mg.min_row, column=mg.min_col).value
            if val:
                header_info['title'] = str(val)
                break

    # Read subtitle (row 4)
    for mg in ws.merged_cells.ranges:
        if mg.min_row == 4 and mg.min_col >= zones['title_start']:
            val = ws.cell(row=4, column=mg.min_col).value
            if val:
                header_info['subtitle'] = str(val)
                break

    # Read tagline (row 5)
    for mg in ws.merged_cells.ranges:
        if mg.min_row == 5 and mg.min_col >= zones['title_start']:
            val = ws.cell(row=5, column=mg.min_col).value
            if val:
                header_info['tagline'] = str(val)
                break

    # Read meta (rows 1-5, meta zone)
    for row in range(1, 6):
        lbl = ws.cell(row=row, column=zones['meta_lbl_start']).value
        val = ws.cell(row=row, column=zones['meta_val_start']).value
        if lbl:
            header_info['meta'][row] = (str(lbl), str(val or ''))

    return header_info


def build_header_on_sheet(ws, ncols, header_info):
    """Build full header structure on a sheet."""
    zones = get_header_zones(ncols)

    # Set column widths
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 2.33

    # Set row heights for header
    for r in range(1, 6):
        ws.row_dimensions[r].height = 15.0
    ws.row_dimensions[6].height = 1.5  # Rule stripe

    # Logo zone: merge A1:logo_end,5
    logo_end_letter = get_column_letter(zones['logo_end'])
    ws.merge_cells(f'A1:{logo_end_letter}5')
    ws['A1'].fill = FILL_WHITE

    # Title zone: merge title_start,1:title_end,3
    ts = get_column_letter(zones['title_start'])
    te = get_column_letter(zones['title_end'])
    ws.merge_cells(f'{ts}1:{te}3')
    title_cell = ws.cell(row=1, column=zones['title_start'])
    title_cell.value = header_info['title']
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_CENTER

    # Subtitle (row 4)
    ws.merge_cells(f'{ts}4:{te}4')
    sub_cell = ws.cell(row=4, column=zones['title_start'])
    sub_cell.value = header_info.get('subtitle', 'Quality Management System · Controlled Document')
    sub_cell.font = FONT_SUBTITLE
    sub_cell.alignment = ALIGN_CENTER

    # Tagline (row 5)
    ws.merge_cells(f'{ts}5:{te}5')
    tag_cell = ws.cell(row=5, column=zones['title_start'])
    tag_cell.value = header_info.get('tagline', 'HESEM Engineering · ISO 9001 / AS9100 · Semiconductor Equipment Parts')
    tag_cell.font = FONT_TAGLINE
    tag_cell.alignment = ALIGN_CENTER

    # Meta zone
    mls = get_column_letter(zones['meta_lbl_start'])
    mle = get_column_letter(zones['meta_lbl_end'])
    mvs = get_column_letter(zones['meta_val_start'])
    mve = get_column_letter(zones['meta_val_end'])

    meta_labels = {
        1: 'Form Code', 2: 'Revision', 3: 'Eff. Date',
        4: 'Owner', 5: 'Approved'
    }

    for row in range(1, 6):
        # Label
        ws.merge_cells(f'{mls}{row}:{mle}{row}')
        lbl_cell = ws.cell(row=row, column=zones['meta_lbl_start'])
        meta = header_info.get('meta', {}).get(row)
        if meta:
            lbl_cell.value = meta[0]
        else:
            lbl_cell.value = meta_labels.get(row, '')
        lbl_cell.font = FONT_META_LBL
        lbl_cell.fill = FILL_META_LABEL
        lbl_cell.alignment = ALIGN_RIGHT

        # Value
        ws.merge_cells(f'{mvs}{row}:{mve}{row}')
        val_cell = ws.cell(row=row, column=zones['meta_val_start'])
        if meta:
            val_cell.value = meta[1]
        val_cell.font = FONT_META_VAL
        val_cell.fill = FILL_WHITE
        val_cell.alignment = ALIGN_LEFT

    # Rule stripe (row 6)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=6, column=c)
        cell.fill = FILL_HESEM_BLUE
        cell.border = Border()

    # Header borders (medium outline)
    for r in range(1, 6):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            top = MEDIUM_SIDE if r == 1 else NO_SIDE
            bottom = MEDIUM_SIDE if r == 5 else NO_SIDE
            left = MEDIUM_SIDE if c == 1 else NO_SIDE
            right = MEDIUM_SIDE if c == ncols else NO_SIDE

            # Zone dividers
            if c == zones['logo_end']:
                right = MEDIUM_SIDE
            if c == zones['title_end']:
                right = MEDIUM_SIDE
            if c == zones['meta_lbl_end']:
                right = THIN_SIDE

            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def fix_sheetviews_xml(filepath):
    """Fix sheetViews XML corruption that causes recovery errors."""
    fixed = False
    tmp_dir = tempfile.mkdtemp()

    try:
        # Extract xlsx (it's a zip)
        with zipfile.ZipFile(filepath, 'r') as zf:
            zf.extractall(tmp_dir)

        # Fix each sheet XML
        xl_ws_dir = os.path.join(tmp_dir, 'xl', 'worksheets')
        if os.path.isdir(xl_ws_dir):
            for sheet_file in os.listdir(xl_ws_dir):
                if not sheet_file.endswith('.xml'):
                    continue
                sheet_path = os.path.join(xl_ws_dir, sheet_file)

                # Parse XML
                ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
                ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
                tree = ET.parse(sheet_path)
                root = tree.getroot()

                # Find sheetViews
                sheet_views = root.find(f'{NS}sheetViews')
                if sheet_views is None:
                    continue

                for sv in sheet_views.findall(f'{NS}sheetView'):
                    # Remove pane elements (freeze pane source of corruption)
                    pane = sv.find(f'{NS}pane')
                    if pane is not None:
                        sv.remove(pane)
                        fixed = True

                    # Remove selection elements that reference panes
                    for sel in sv.findall(f'{NS}selection'):
                        pane_attr = sel.get('pane')
                        if pane_attr and pane_attr != 'topLeft':
                            sv.remove(sel)
                            fixed = True

                    # Ensure at least one selection exists
                    selections = sv.findall(f'{NS}selection')
                    if not selections:
                        sel_elem = ET.SubElement(sv, f'{NS}selection')
                        sel_elem.set('activeCell', 'A1')
                        sel_elem.set('sqref', 'A1')
                        fixed = True

                tree.write(sheet_path, xml_declaration=True, encoding='UTF-8')

        if fixed:
            # Repack into xlsx
            with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                for root_dir, dirs, files in os.walk(tmp_dir):
                    for f in files:
                        full_path = os.path.join(root_dir, f)
                        arc_name = os.path.relpath(full_path, tmp_dir)
                        zf.write(full_path, arc_name)

    except Exception as e:
        print(f"    XML fix error: {e}")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    return fixed


def fix_all_fonts(ws):
    """Fix all fonts to Segoe UI, skip non-primary merged cells."""
    merged_secondary = set()
    for mg in ws.merged_cells.ranges:
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r != mg.min_row or c != mg.min_col:
                    merged_secondary.add((r, c))

    count = 0
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if (cell.row, cell.column) in merged_secondary:
                continue
            if cell.font and cell.font.name and cell.font.name != FONT_FAMILY:
                has_content = cell.value is not None
                has_fill = cell.fill and cell.fill.fill_type is not None
                has_border = cell.border and any(
                    getattr(cell.border, s) and getattr(cell.border, s).style
                    for s in ['top', 'bottom', 'left', 'right']
                )
                if has_content or has_fill or has_border:
                    cell.font = Font(
                        name=FONT_FAMILY,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strikethrough,
                        color=cell.font.color
                    )
                    count += 1
    return count


def fix_named_styles(wb):
    """Fix Normal style font to Segoe UI."""
    try:
        for ns in wb._named_styles:
            if ns.font and ns.font.name and ns.font.name != FONT_FAMILY:
                ns.font = Font(
                    name=FONT_FAMILY,
                    size=ns.font.size,
                    bold=ns.font.bold,
                    italic=ns.font.italic,
                    color=ns.font.color
                )
                return True
    except:
        pass
    return False


def fix_notice_formatting(ws, ncols):
    """Fix notice bar formatting (amber background, small gold text)."""
    count = 0
    for row in range(1, (ws.max_row or 50) + 1):
        cell = ws.cell(row=row, column=1)
        val = str(cell.value or '').upper()

        # Detect notice rows
        if val.startswith('NOTICE') or val.startswith('NOTE:') or val.startswith('⚠'):
            ws.row_dimensions[row].height = 26.0
            for c in range(1, ncols + 1):
                nc = ws.cell(row=row, column=c)
                nc.fill = FILL_NOTICE
                nc.font = FONT_NOTICE
                nc.alignment = ALIGN_LEFT
                nc.border = Border(
                    top=NOTICE_SIDE, bottom=NOTICE_SIDE,
                    left=NOTICE_SIDE if c == 1 else NO_SIDE,
                    right=NOTICE_SIDE if c == ncols else NO_SIDE
                )
            count += 1
    return count


def remove_freeze_panes(ws):
    """Remove freeze panes from worksheet."""
    try:
        if ws.freeze_panes:
            ws.freeze_panes = None
            return True
    except:
        pass
    return False


def fix_gridlines(ws):
    """Hide gridlines."""
    try:
        if hasattr(ws, 'sheet_view') and ws.sheet_view.showGridLines != False:
            ws.sheet_view.showGridLines = False
            return True
    except:
        pass
    return False


def fix_page_setup(ws, ncols):
    """Fix page margins and setup."""
    try:
        pm = ws.page_margins
        pm.left = pm.right = 0.197
        pm.top = pm.bottom = 0.197
        pm.header = pm.footer = 0.1

        ps = ws.page_setup
        ps.paperSize = 9 if ncols <= 56 else 8
        ps.fitToWidth = 1
        ps.fitToHeight = 0
    except:
        pass


def process_workbook(filepath):
    """Process a single workbook - fix all issues."""
    fname = os.path.basename(filepath)
    issues = []

    # STEP 1: Fix sheetViews XML (prevents recovery errors)
    xml_fixed = fix_sheetviews_xml(filepath)
    if xml_fixed:
        issues.append("  XML: sheetViews pane corruption fixed")

    # STEP 2: Open and fix with openpyxl
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, [f"  ERROR: Cannot open: {e}"]

    # Fix named styles
    if fix_named_styles(wb):
        issues.append("  Normal style -> Segoe UI")

    # Get main tab info
    main_ws = wb[wb.sheetnames[0]]
    main_ncols = detect_ncols(main_ws)
    main_header = read_main_tab_header(main_ws, main_ncols)

    for sn in wb.sheetnames:
        ws = wb[sn]
        ncols = detect_ncols(ws)
        ws_issues = []

        # Fix freeze panes
        if remove_freeze_panes(ws):
            ws_issues.append("freeze removed")

        # Fix gridlines
        if fix_gridlines(ws):
            ws_issues.append("gridlines hidden")

        # Fix page setup
        fix_page_setup(ws, ncols)

        # Check if header is valid
        if not has_valid_header(ws, ncols) and ws.max_row and ws.max_row > 1:
            # Header is broken - check if this is a support tab
            if sn != wb.sheetnames[0]:
                # Need to rebuild header
                # First, check if there's content below that we need to preserve
                # Get all existing data and merges
                existing_data = {}
                existing_merges = list(ws.merged_cells.ranges)

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            existing_data[(cell.row, cell.column)] = {
                                'value': cell.value,
                                'font': copy.copy(cell.font),
                                'fill': copy.copy(cell.fill),
                                'alignment': copy.copy(cell.alignment),
                                'border': copy.copy(cell.border),
                            }

                # Determine if content starts at row 1 (no header) or row 7+ (has header)
                first_content_row = 1
                for r in range(1, 10):
                    cell = ws.cell(row=r, column=1)
                    if cell.value:
                        first_content_row = r
                        break

                # If content starts at row 1, we need to shift everything down by 6
                if first_content_row <= 6:
                    shift = 7 - first_content_row

                    # Unmerge all
                    for mg in existing_merges:
                        try:
                            ws.unmerge_cells(str(mg))
                        except:
                            pass

                    # Shift data down
                    shifted_data = {}
                    for (r, c), d in existing_data.items():
                        new_r = r + shift
                        shifted_data[(new_r, c)] = d

                    # Clear old positions
                    for r in range(1, (ws.max_row or 50) + 10):
                        for c in range(1, ncols + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.value = None
                            cell.fill = PatternFill(fill_type=None)
                            cell.font = Font(name=FONT_FAMILY)
                            cell.border = Border()
                            cell.alignment = Alignment()

                    # Write shifted data
                    for (r, c), d in shifted_data.items():
                        cell = ws.cell(row=r, column=c)
                        cell.value = d['value']
                        cell.font = d['font']
                        cell.fill = d['fill']
                        cell.alignment = d['alignment']
                        cell.border = d['border']

                    # Re-merge shifted ranges
                    for mg in existing_merges:
                        try:
                            new_mg = f"{get_column_letter(mg.min_col)}{mg.min_row + shift}:{get_column_letter(mg.max_col)}{mg.max_row + shift}"
                            ws.merge_cells(new_mg)
                        except:
                            pass

                    ws_issues.append(f"content shifted down {shift} rows")

                # Build header
                build_header_on_sheet(ws, ncols, main_header)
                ws_issues.append("header rebuilt from main tab")

        # Fix all fonts
        font_count = fix_all_fonts(ws)
        if font_count > 0:
            ws_issues.append(f"{font_count} fonts->Segoe UI")

        # Fix notice formatting
        notice_count = fix_notice_formatting(ws, ncols)
        if notice_count > 0:
            ws_issues.append(f"{notice_count} notice bars fixed")

        if ws_issues:
            issues.append(f"  [{sn}] {', '.join(ws_issues)}")

    # Save
    try:
        wb.save(filepath)
        issues.append("  >> SAVED")
    except Exception as e:
        issues.append(f"  >> SAVE ERROR: {e}")

    wb.close()
    return fname, issues


def main():
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS - COMPREHENSIVE FORM FIX v2")
    print(f"  1. Fix sheetViews XML (recovery errors)")
    print(f"  2. Rebuild broken headers")
    print(f"  3. Fix fonts -> Segoe UI")
    print(f"  4. Remove freeze panes")
    print(f"  5. Fix notice bars")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    fixed_count = 0
    error_count = 0

    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        print(f"[{i:3d}/{len(files)}] {fn[:55]:55s}", end=" ", flush=True)

        name, issues = process_workbook(fp)

        has_error = any('ERROR' in str(x) for x in issues)
        has_fix = any(x for x in issues if 'SAVED' not in x and 'ERROR' not in x)

        if has_error:
            print("ERROR")
            error_count += 1
        elif has_fix:
            print("FIXED")
            fixed_count += 1
        else:
            print("OK")

        # Print issues in verbose mode
        if '--verbose' in sys.argv or '-v' in sys.argv:
            for iss in issues:
                print(iss)

    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"  Total: {len(files)}")
    print(f"  Fixed: {fixed_count}")
    print(f"  Errors: {error_count}")
    print(f"  OK: {len(files) - fixed_count - error_count}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
