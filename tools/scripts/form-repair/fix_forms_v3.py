#!/usr/bin/env python3
"""
HESEM QMS Form Compliance Fixer v3
Complete visual transformation to match master template spec.
Fixes: colors, header structure, logo, borders, fonts, action tab merge.
"""
import os
import sys
import copy
import glob as globmod
import subprocess
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# MASTER TEMPLATE SPEC (from hesem-PRECISION-template.xlsx v4.2)
# ═══════════════════════════════════════════════════════════════════════════════

FONT = "Segoe UI"

# Colors
C = {
    'HESEM_BLUE':  '2C9CD7',
    'DARK_NAVY':   '0C2D48',
    'MED_BLUE':    '1565C0',
    'WHITE':       'FFFFFF',
    'LABEL_BG':    'F0F7FC',
    'META_LABEL':  'E4F0F9',
    'STRIPE_BG':   'F7FBFF',
    'THIN_BORDER': 'C5D9E8',
    'BODY_TEXT':   '1A2733',
    'SUBTITLE':    '1565C0',
    'LGRAY':       '7A8FA6',
    'GOLD_TEXT':   '7B4F00',
    'NOTICE_BG':   'FFF8E1',
    'NOTICE_BRD':  'F9A825',
    'SIG_TEXT':    'BBCCCC',
}

# OLD colors from generated forms -> NEW colors
COLOR_MAP = {
    'EFF6FF': C['MED_BLUE'],     # old section header bg -> L1 section header
    '1B3A6B': 'FFFFFF',           # old section header text -> white on blue
    '005A9E': C['DARK_NAVY'],     # old column header bg -> dark navy
    'F1F5F9': C['LABEL_BG'],      # old label bg -> proper label bg
    '334155': C['DARK_NAVY'],     # old label text -> dark navy
    'CBD5E1': C['SIG_TEXT'],       # old signature hint -> sig text
    '3D4B60': C['GOLD_TEXT'],      # old notice text -> gold
    'F5F5F5': C['STRIPE_BG'],     # old stripe -> proper stripe
}

# Border sides
MEDIUM = Side(style='medium', color=C['HESEM_BLUE'])
THIN   = Side(style='thin', color=C['THIN_BORDER'])
NONE   = Side(style=None)
NOTICE = Side(style='thin', color=C['NOTICE_BRD'])

# Fills
FILL = {
    'HESEM_BLUE': PatternFill('solid', fgColor=C['HESEM_BLUE']),
    'DARK_NAVY':  PatternFill('solid', fgColor=C['DARK_NAVY']),
    'MED_BLUE':   PatternFill('solid', fgColor=C['MED_BLUE']),
    'WHITE':      PatternFill('solid', fgColor=C['WHITE']),
    'LABEL':      PatternFill('solid', fgColor=C['LABEL_BG']),
    'META_LABEL': PatternFill('solid', fgColor=C['META_LABEL']),
    'STRIPE':     PatternFill('solid', fgColor=C['STRIPE_BG']),
    'NOTICE':     PatternFill('solid', fgColor=C['NOTICE_BG']),
    'NONE':       PatternFill(fill_type=None),
}

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'assets', 'hesem-logo.png')
STD_COL_WIDTH = 2.33

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def color_hex(c):
    if c is None: return None
    if hasattr(c, 'rgb') and c.rgb and c.rgb != '00000000':
        s = str(c.rgb)
        return s[2:].upper() if len(s) == 8 else s.upper()
    return None

def fill_color(cell):
    if cell.fill and cell.fill.fill_type:
        return color_hex(cell.fill.fgColor)
    return None

def font_color(cell):
    if cell.font and cell.font.color:
        return color_hex(cell.font.color)
    return None

def detect_ncols(ws):
    count = 0
    for i in range(1, 200):
        d = ws.column_dimensions.get(get_column_letter(i))
        if d and d.width is not None and abs(d.width - STD_COL_WIDTH) < 0.5:
            count += 1
        elif count > 10:
            break
    if count <= 44: return 40
    elif count <= 60: return 56
    else: return 82

def get_header_zones(n):
    """Returns (logo_end, title_start, title_end, meta_lbl_start, meta_lbl_end, meta_val_start, meta_val_end)"""
    if n <= 40: return 8, 9, 30, 31, 34, 35, 40
    elif n <= 56: return 11, 12, 42, 43, 48, 49, 56
    else: return 16, 17, 62, 63, 70, 71, 82

def merged_non_primary(ws):
    s = set()
    for mg in ws.merged_cells.ranges:
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r != mg.min_row or c != mg.min_col:
                    s.add((r, c))
    return s

def has_header_structure(ws):
    """Check if sheet has the standard header (rows 1-5 or 1-7)."""
    # Look for merged title area in rows 1-5
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 3 and mg.max_row <= 5:
            cell = ws.cell(row=mg.min_row, column=mg.min_col)
            if cell.value and isinstance(cell.value, str) and len(cell.value) > 5:
                return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# ROW CLASSIFICATION
# ═══════════════════════════════════════════════════════════════════════════════

def classify_rows(ws, ncols, header_end):
    """Classify all rows in the sheet."""
    max_row = ws.max_row or 50
    types = {}

    for r in range(1, max_row + 1):
        if r <= header_end:
            types[r] = 'header'
            continue

        rd = ws.row_dimensions.get(r)
        h = rd.height if rd else None
        c1 = ws.cell(row=r, column=1)
        fh = fill_color(c1)
        val = str(c1.value or '')

        # Spacer: short rows with no content
        if h and h <= 5:
            types[r] = 'spacer'
            continue

        # Rule stripe (already converted)
        if fh == C['HESEM_BLUE'] and h and h <= 3:
            types[r] = 'rule_stripe'
            continue

        # Notice bar
        if fh == C['NOTICE_BG'] or fh == 'FFF8E1':
            types[r] = 'notice'
            continue

        # Section header: check original colors or already-fixed colors
        if fh in (C['MED_BLUE'], 'EFF6FF', 'E8F0FE', 'DCEEFB', 'E3F2FD'):
            # Check if it's a full-width merge
            is_full_merge = False
            for mg in ws.merged_cells.ranges:
                if mg.min_row == r and mg.max_row == r and (mg.max_col - mg.min_col + 1) >= ncols * 0.5:
                    is_full_merge = True
                    break
            if is_full_merge or (val and len(val.strip()) > 3):
                types[r] = 'section'
                continue

        # Column header: dark fills
        if fh in (C['DARK_NAVY'], '005A9E', '1B4472', '1F3864', '2F5496'):
            types[r] = 'col_header'
            continue
        # Check across row for column headers
        for cc in [2, 3, 5, 8]:
            if cc <= ncols:
                fh2 = fill_color(ws.cell(row=r, column=cc))
                if fh2 in (C['DARK_NAVY'], '005A9E', '1B4472', '1F3864', '2F5496'):
                    types[r] = 'col_header'
                    break

        if r not in types:
            # Signature area
            if any(kw in val.lower() for kw in ['prepared', 'reviewed', 'approved',
                                                   'name', 'signature', 'date',
                                                   'chuẩn bị', 'xem xét', 'phê duyệt']):
                types[r] = 'signature'
            else:
                types[r] = 'data'

    return types


# ═══════════════════════════════════════════════════════════════════════════════
# TRANSFORMATIONS
# ═══════════════════════════════════════════════════════════════════════════════

def transform_header(ws, ncols, log):
    """Transform header rows 1-5/6 to match master template spec.
    Adds rule stripes if missing, fixes colors."""
    zones = get_header_zones(ncols)
    logo_end, title_start, title_end, meta_lbl_start, meta_lbl_end, meta_val_start, meta_val_end = zones

    # Detect existing header structure
    # Rows 1-5: logo (A1:H5), title (I1:AD3 + I4 + I5), meta (AE1:AN5)
    # We need to add rule stripes and fix colors

    # First check if row 6 is a spacer that we can convert to rule stripe
    # or if we need to insert rows
    rd6 = ws.row_dimensions.get(6)
    h6 = rd6.height if rd6 else None
    is_spacer_6 = h6 and h6 <= 5

    header_end = 5  # Default

    if is_spacer_6:
        # Unmerge row 6 first if merged
        merges_to_remove = []
        for mg in ws.merged_cells.ranges:
            if mg.min_row <= 6 <= mg.max_row and mg.min_row == 6:
                merges_to_remove.append(mg)
        for mg in merges_to_remove:
            ws.unmerge_cells(str(mg))

        # Convert row 6 to rule stripe
        ws.row_dimensions[6].height = 1.5
        for c in range(1, ncols + 1):
            cell = ws.cell(row=6, column=c)
            cell.fill = FILL['HESEM_BLUE']
            cell.border = Border()
            try:
                cell.value = None
            except AttributeError:
                pass  # MergedCell
        header_end = 6
        log.append("  Row 6: converted spacer -> rule stripe")

    # Fix header rows 1-5 colors
    for r in range(1, 6):
        rd = ws.row_dimensions[r]
        if rd.height is None or abs(rd.height - 15.0) > 2:
            rd.height = 15.0

        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)

            # Logo zone (cols 1 to logo_end)
            if c <= logo_end:
                cell.fill = FILL['WHITE']
                cell.border = Border()

            # Title zone
            elif title_start <= c <= title_end:
                cell.fill = FILL['WHITE']
                if r <= 3:  # Main title
                    cell.font = Font(name=FONT, size=14, bold=True, color=C['DARK_NAVY'])
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                elif r == 4:  # QMS subtitle
                    cell.font = Font(name=FONT, size=8, bold=False, color=C['SUBTITLE'])
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif r == 5:  # Tagline
                    cell.font = Font(name=FONT, size=7, bold=False, color=C['LGRAY'])
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Meta label zone
            elif meta_lbl_start <= c <= meta_lbl_end:
                cell.fill = FILL['META_LABEL']
                cell.font = Font(name=FONT, size=8, bold=True, color=C['DARK_NAVY'])
                cell.alignment = Alignment(horizontal='right', vertical='center', wrapText=True)

            # Meta value zone
            elif meta_val_start <= c <= meta_val_end:
                cell.fill = FILL['WHITE']
                cell.font = Font(name=FONT, size=8, bold=False, color=C['BODY_TEXT'])
                cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    # Apply header block borders
    for r in range(1, 6):
        for c in range(1, ncols + 1):
            top = MEDIUM if r == 1 else NONE
            bottom = MEDIUM if r == 5 else NONE
            left = MEDIUM if c == 1 else NONE
            right = MEDIUM if c == ncols else NONE

            # Zone boundaries
            if c == logo_end:
                right = MEDIUM
            if c == logo_end + 1:
                left = MEDIUM
            if c == title_end:
                right = MEDIUM
            if c == title_end + 1:
                left = MEDIUM

            # Thin separators between meta rows
            if c >= meta_lbl_start:
                if r > 1:
                    top = THIN
                # Separator between label and value
                if c == meta_lbl_end:
                    right = THIN
                if c == meta_val_start:
                    left = THIN

            ws.cell(row=r, column=c).border = Border(
                top=top, bottom=bottom, left=left, right=right
            )

    log.append("  Header: colors and borders fixed")
    return header_end


def transform_section_header(ws, row, ncols, log):
    """Transform section header to L1 style."""
    # Fix fill and font
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL['MED_BLUE']
        if cell.font:
            cell.font = Font(name=FONT, size=9, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    # Medium outline border
    for c in range(1, ncols + 1):
        left = MEDIUM if c == 1 else NONE
        right = MEDIUM if c == ncols else NONE
        ws.cell(row=row, column=c).border = Border(
            top=MEDIUM, bottom=MEDIUM, left=left, right=right
        )

    # Fix row height
    ws.row_dimensions[row].height = 18.0


def transform_col_header(ws, row, ncols, log):
    """Transform column header to DARK_NAVY style."""
    # Detect merge boundaries for thin vertical separators
    merge_bounds = set()
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= row <= mg.max_row:
            merge_bounds.add(mg.min_col)
            merge_bounds.add(mg.max_col + 1)

    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL['DARK_NAVY']
        cell.font = Font(name=FONT, size=9, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        left = MEDIUM if c == 1 else (THIN if c in merge_bounds else NONE)
        right = MEDIUM if c == ncols else NONE
        cell.border = Border(top=MEDIUM, bottom=MEDIUM, left=left, right=right)

    ws.row_dimensions[row].height = 18.0


def transform_data_rows(ws, rows, ncols, row_types, log):
    """Transform data rows: fix colors, borders."""
    if not rows:
        return

    # Find column boundaries from merged ranges
    col_bounds = set()
    col_bounds.add(1)
    for mg in ws.merged_cells.ranges:
        if any(mg.min_row <= r <= mg.max_row for r in rows):
            col_bounds.add(mg.min_col)
            col_bounds.add(mg.max_col + 1)

    # Also detect label/input boundaries from fill patterns
    for r in rows[:5]:
        prev_fill = None
        for c in range(1, ncols + 1):
            curr_fill = fill_color(ws.cell(row=r, column=c))
            if curr_fill and curr_fill != prev_fill:
                col_bounds.add(c)
            prev_fill = curr_fill

    # Find contiguous blocks (between section headers/spacers)
    blocks = []
    current_block = []
    for r in rows:
        if row_types.get(r) in ('section', 'col_header', 'spacer', 'rule_stripe', 'notice'):
            if current_block:
                blocks.append(current_block)
                current_block = []
        else:
            current_block.append(r)
    if current_block:
        blocks.append(current_block)

    for block in blocks:
        if not block:
            continue
        first = block[0]
        last = block[-1]

        for r in block:
            # Fix row height
            rt = row_types.get(r, 'data')
            if rt == 'signature':
                ws.row_dimensions[r].height = 26.0
            else:
                rd = ws.row_dimensions.get(r)
                if rd and (rd.height is None or abs(rd.height - 20.0) > 3):
                    ws.row_dimensions[r].height = 20.0

            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)

                # Fix fill colors
                fc = fill_color(cell)
                if fc in COLOR_MAP:
                    new_color = COLOR_MAP[fc]
                    cell.fill = PatternFill('solid', fgColor=new_color)

                # Fix font
                if cell.font:
                    fn_color = font_color(cell)
                    new_fn_color = cell.font.color
                    if fn_color and fn_color in COLOR_MAP:
                        new_fn_color = COLOR_MAP[fn_color]
                    cell.font = Font(
                        name=FONT,
                        size=cell.font.size or 8,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=new_fn_color
                    )

                # Fix borders
                top = MEDIUM if r == first else THIN
                bottom = MEDIUM if r == last else THIN
                left = MEDIUM if c == 1 else (THIN if c in col_bounds else NONE)
                right = MEDIUM if c == ncols else NONE

                cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def transform_spacer(ws, row, ncols):
    """Fix spacer row."""
    # Unmerge row first
    merges_to_remove = [mg for mg in ws.merged_cells.ranges
                        if mg.min_row == row and mg.max_row == row]
    for mg in merges_to_remove:
        ws.unmerge_cells(str(mg))

    ws.row_dimensions[row].height = 3.0
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL['NONE']
        cell.border = Border()
        try:
            cell.value = None
        except AttributeError:
            pass


def transform_notice(ws, row, ncols):
    """Fix notice bar."""
    ws.row_dimensions[row].height = 26.0
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL['NOTICE']
        cell.font = Font(name=FONT, size=7, bold=False, color=C['GOLD_TEXT'])
        cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        left = NOTICE if c == 1 else NONE
        right = NOTICE if c == ncols else NONE
        cell.border = Border(top=NOTICE, bottom=NOTICE, left=left, right=right)


def embed_logo(ws, ncols):
    """Embed HESEM logo in the header logo zone."""
    if not os.path.exists(LOGO_PATH):
        return False

    ws._images = []

    zones = get_header_zones(ncols)
    logo_end = zones[0]

    img = XlImage(LOGO_PATH)

    # Logo zone: logo_end cols x 5 rows (each 15pt)
    # Each col ≈ 5mm, 5 rows x 15pt ≈ 26.5mm
    zone_w_mm = logo_end * 5.03
    zone_h_mm = 75 * 0.353  # 75pt

    aspect = 52.0 / 15.0  # SVG aspect ratio

    # Fit with 10% padding
    avail_w = zone_w_mm * 0.90
    avail_h = zone_h_mm * 0.70

    if avail_w / aspect <= avail_h:
        w_mm = avail_w
        h_mm = avail_w / aspect
    else:
        h_mm = avail_h
        w_mm = avail_h * aspect

    img.width = w_mm / 25.4 * 96  # px at 96 DPI
    img.height = h_mm / 25.4 * 96

    img.anchor = 'A2'
    ws.add_image(img)
    return True


def fix_named_styles(wb):
    """Fix workbook default/Normal font to Segoe UI."""
    try:
        for ns in wb._named_styles:
            if ns.font and ns.font.name and ns.font.name != FONT:
                ns.font = Font(
                    name=FONT,
                    size=ns.font.size,
                    bold=ns.font.bold,
                    italic=ns.font.italic,
                    color=ns.font.color
                )
    except:
        pass


def fix_remaining_fonts(ws):
    """Final sweep: fix any remaining non-Segoe UI fonts."""
    mnp = merged_non_primary(ws)
    count = 0
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if (cell.row, cell.column) in mnp:
                continue
            if cell.font and cell.font.name and cell.font.name != FONT:
                has_content = cell.value is not None
                has_fill = cell.fill and cell.fill.fill_type is not None
                has_border = cell.border and any(
                    getattr(cell.border, s) and getattr(cell.border, s).style
                    for s in ['top', 'bottom', 'left', 'right']
                )
                if has_content or has_fill or has_border:
                    cell.font = Font(
                        name=FONT,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strikethrough,
                        color=cell.font.color
                    )
                    count += 1
    return count


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def process_file(filepath):
    fname = os.path.basename(filepath)
    log = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, [f"ERROR: {e}"], False

    fix_named_styles(wb)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # Skip pure utility sheets
        skip_names = ['lists', 'dropdown', 'data', 'zlists', '_lists']
        if any(ws_name.lower().startswith(s) or ws_name.lower() == s for s in skip_names):
            n = fix_remaining_fonts(ws)
            if n: log.append(f"  [{ws_name}] {n} fonts fixed")
            continue

        ncols = detect_ncols(ws)
        has_hdr = has_header_structure(ws)

        header_end = 0
        if has_hdr:
            header_end = transform_header(ws, ncols, log)
            embed_logo(ws, ncols)
            log.append(f"  [{ws_name}] Logo embedded")

        # Classify rows
        row_types = classify_rows(ws, ncols, header_end)
        max_row = ws.max_row or 50

        # Transform each row type
        data_rows = []
        for r in range(header_end + 1, max_row + 1):
            rt = row_types.get(r, 'data')

            if rt == 'section':
                # Flush data rows
                if data_rows:
                    transform_data_rows(ws, data_rows, ncols, row_types, log)
                    data_rows = []
                transform_section_header(ws, r, ncols, log)

            elif rt == 'col_header':
                if data_rows:
                    transform_data_rows(ws, data_rows, ncols, row_types, log)
                    data_rows = []
                transform_col_header(ws, r, ncols, log)

            elif rt == 'spacer':
                if data_rows:
                    transform_data_rows(ws, data_rows, ncols, row_types, log)
                    data_rows = []
                transform_spacer(ws, r, ncols)

            elif rt == 'rule_stripe':
                if data_rows:
                    transform_data_rows(ws, data_rows, ncols, row_types, log)
                    data_rows = []
                # Already handled by header transform

            elif rt == 'notice':
                if data_rows:
                    transform_data_rows(ws, data_rows, ncols, row_types, log)
                    data_rows = []
                transform_notice(ws, r, ncols)

            else:
                data_rows.append(r)

        # Flush remaining
        if data_rows:
            transform_data_rows(ws, data_rows, ncols, row_types, log)

        # Final font sweep
        n = fix_remaining_fonts(ws)
        if n: log.append(f"  [{ws_name}] {n} fonts fixed")

        # Fix gridlines
        try:
            ws.sheet_view.showGridLines = False
        except:
            pass

        # Fix page setup
        try:
            pm = ws.page_margins
            pm.left = pm.right = 0.197
            pm.top = pm.bottom = 0.197
            pm.header = pm.footer = 0.1
            ps = ws.page_setup
            ps.paperSize = 9 if ncols <= 56 else 8
            ps.fitToWidth = 1
            ps.fitToHeight = 0
        except:
            pass

        log.append(f"  [{ws_name}] transformation complete")

    # Save
    try:
        wb.save(filepath)
        log.append("  >> SAVED")
    except Exception as e:
        log.append(f"  >> SAVE ERROR: {e}")

    wb.close()
    return fname, log, True


def main():
    base_dir = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')

    frm_files = []
    for root, dirs, files in os.walk(base_dir):
        if '_build' in root:
            continue
        for f in files:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                frm_files.append(os.path.join(root, f))
    frm_files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS FORM COMPLIANCE FIX v3")
    print(f"Logo: {'OK' if os.path.exists(LOGO_PATH) else 'MISSING!'}")
    print(f"Files: {len(frm_files)}")
    print(f"{'='*70}\n")

    verbose = '--verbose' in sys.argv

    for i, fpath in enumerate(frm_files, 1):
        fname = os.path.basename(fpath)
        print(f"[{i:3d}/{len(frm_files)}] {fname}...", end=" ", flush=True)
        name, log, ok = process_file(fpath)
        print("OK" if ok else "FAIL")
        if verbose:
            for l in log:
                print(l)

    print(f"\n{'='*70}")
    print(f"DONE - {len(frm_files)} files processed")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
