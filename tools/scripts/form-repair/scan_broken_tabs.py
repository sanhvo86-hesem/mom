#!/usr/bin/env python3
"""
Scan ALL forms to find broken/corrupted tabs.
A tab is "broken" if:
1. Has content but wrong column widths (not matching master template grid)
2. Has very few merges relative to content (layout destroyed)
3. Has images but no proper header structure
4. Title/header text is truncated or misaligned
"""
import os, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
STD_WIDTH = 2.33


def check_grid_compliance(ws):
    """Check if column widths follow the master template grid."""
    ncols = 0
    non_std = 0
    for col_idx in range(1, 100):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width is not None:
            if abs(dim.width - STD_WIDTH) < 0.5:
                ncols += 1
            else:
                non_std += 1
                if ncols > 5:
                    break
    return ncols, non_std


def check_header_structure(ws):
    """Check if the sheet has a proper master template header."""
    # Check row 1-5 for header content
    has_title = False
    has_logo_space = False
    has_meta = False

    for row in range(1, 7):
        for col in range(1, 60):
            cell = ws.cell(row=row, column=col)
            val = str(cell.value or '')
            if len(val) > 10 and any(kw in val.upper() for kw in
                ['CHECKLIST', 'REPORT', 'LOG', 'RECORD', 'FORM', 'REGISTER',
                 'TRACKER', 'SHEET', 'INDEX', 'PLAN', 'MATRIX', 'REVIEW',
                 'REQUEST', 'MINUTES', 'LABEL', 'TAG', 'SURVEY', 'CERTIFICATE',
                 'SCORECARD', 'INVOICE', 'DISPATCH', 'PACKAGING', 'SHIPPING',
                 'INSPECTION', 'AUDIT', 'ASSESSMENT', 'EVALUATION',
                 'VERIFICATION', 'MAINTENANCE', 'TRAINING', 'ATTENDANCE',
                 'ANALYSIS', 'SCAR', 'NCR', 'CAPA', 'PDCA', 'FAI',
                 'OUTSOURCE', 'CALIBRATION', 'GAGE', 'SPC', 'AQL',
                 'TOOL', 'MACHINE', 'ENVIRONMENT', 'CLEANLINESS',
                 'INCIDENT', 'LIGHTING', 'HELIUM', 'ULTRASONIC', 'VACUUM',
                 'FOD', 'SMED', 'COSTING', 'SETUP', 'DFM', 'ENGINEERING',
                 'CLASSIFICATION', 'SEMICONDUCTOR', 'DOWNTIME',
                 'WIP', 'SHIFT', 'DAILY', 'COMPLETION', 'KICKOFF',
                 'COMMUNICATION', 'DISRUPTION', 'RISK', 'PFMEA',
                 'CONTROL', 'ACCESS', 'DOCUMENT', 'MASTER',
                 'CONTRACT', 'RFQ', 'JOB', 'ORDER', 'COMPLAINT',
                 'CUSTOMER', 'PROPERTY', 'RMA', 'PURCHASE',
                 'SUPPLIER', 'HOLD', 'PLANNING', 'SAFETY',
                 'SKILLS', 'COMPETENCE', 'PERFORMANCE',
                 'MANAGEMENT', 'INTERNAL', 'LAYERED', 'ACTION',
                 'FINDING', 'DECISION', 'CONDITION',
                 'CLEAN', 'LEAK', 'ENTRY', 'GOWNING', 'BATCH',
                 'COMPATIBLE', 'BUILD', 'BAGGING', 'CLEARANCE',
                 'CERTIFICATION', 'CRASH', 'TOOLING',
                 'HISTORY', 'MEASURING', 'EQUIPMENT',
                 'BIAS', 'LINEARITY', 'STABILITY', 'ATTRIBUTE',
                 'SPECIAL', 'CHARACTERISTIC', 'SATISFACTION',
                 'RECEIVING', 'IQC', 'FINAL', 'PART', 'LOCATION',
                 'CONTEXT', 'INTERESTED', 'CLIMATE', 'CSR',
                 'LESSON', 'ECR', 'ECO', 'CHANGE', 'CONFIGURATION',
                 'BUSINESS', 'QUARTERLY', 'M365', 'PEER', 'PILOT',
                 'DEPLOYMENT', 'CAPABILITY', 'FLOW', 'REQUIREMENT',
                 'WORK TRANSFER', 'PACKET', 'QUICK', 'PRE-RUN',
                 'PREVENTIVE', 'IMPACT', 'RELEASE', 'BASELINE']):
                has_title = True

    # Check for merged header cells (logo zone, title zone, meta zone)
    header_merges = 0
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 6:
            header_merges += 1

    return has_title, header_merges


def scan_all():
    """Scan all files and report broken tabs."""
    files = []
    for root, dirs, fnames in os.walk(BASE):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    broken_tabs = []
    ok_tabs = 0
    total_tabs = 0

    for fp in files:
        fname = os.path.basename(fp)
        code = fname.split('_')[0]

        try:
            wb = load_workbook(fp, data_only=True)
        except Exception as e:
            print(f"ERROR loading {fname}: {e}")
            continue

        for ws in wb.worksheets:
            if ws.sheet_state == 'veryHidden':
                continue

            total_tabs += 1
            sn = ws.title

            # Skip LISTS tabs
            if 'list' in sn.lower():
                ok_tabs += 1
                continue

            ncols, non_std = check_grid_compliance(ws)
            has_title, header_merges = check_header_structure(ws)
            merge_count = len(list(ws.merged_cells.ranges))
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            has_images = len(ws._images) if hasattr(ws, '_images') else 0

            # Determine if broken
            is_broken = False
            reasons = []

            # Rule 1: Has content but grid is wrong
            if max_row > 3 and ncols < 20 and non_std > 5:
                is_broken = True
                reasons.append(f"grid:{ncols}std/{non_std}non-std")

            # Rule 2: Has content but very few merges (layout destroyed)
            if max_row > 10 and merge_count < 5 and has_images:
                is_broken = True
                reasons.append(f"low-merges:{merge_count}")

            # Rule 3: Is a support tab (not first tab) with header but broken grid
            if ws != wb.worksheets[0]:
                if max_row > 5 and ncols < 15 and max_col > 5:
                    # Check if it has any content worth keeping
                    content_cells = 0
                    for row in range(1, min(20, max_row + 1)):
                        for col in range(1, min(10, max_col + 1)):
                            if ws.cell(row=row, column=col).value is not None:
                                content_cells += 1
                    if content_cells > 5 and ncols < 15:
                        is_broken = True
                        reasons.append(f"support-tab-broken-grid")

            # Rule 4: Second+ tab with header image but truncated title
            if ws != wb.worksheets[0] and has_images and not has_title and max_row > 5:
                is_broken = True
                reasons.append("no-title-detected")

            if is_broken:
                broken_tabs.append({
                    'file': fname,
                    'code': code,
                    'tab': sn,
                    'reasons': reasons,
                    'ncols': ncols,
                    'merges': merge_count,
                    'rows': max_row,
                    'cols': max_col,
                    'images': has_images,
                })
            else:
                ok_tabs += 1

        wb.close()

    # Report
    print(f"{'='*80}")
    print(f"BROKEN TAB SCAN RESULTS")
    print(f"  Total visible tabs: {total_tabs}")
    print(f"  OK tabs: {ok_tabs}")
    print(f"  Broken tabs: {len(broken_tabs)}")
    print(f"{'='*80}")

    if broken_tabs:
        print(f"\nBROKEN TABS:")
        print(f"{'Code':<10} {'Tab':<30} {'Grid':<12} {'Merges':<8} {'Rows':<6} {'Reasons'}")
        print("-" * 100)
        for bt in broken_tabs:
            grid_str = f"{bt['ncols']}std"
            print(f"{bt['code']:<10} {bt['tab']:<30} {grid_str:<12} {bt['merges']:<8} {bt['rows']:<6} {', '.join(bt['reasons'])}")


if __name__ == '__main__':
    scan_all()
