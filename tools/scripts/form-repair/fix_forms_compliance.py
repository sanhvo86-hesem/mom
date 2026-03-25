#!/usr/bin/env python3
"""
HESEM QMS Form Compliance Fixer
Audits and fixes all FRM-*.xlsx files against the master template spec.
"""
import os
import sys
import glob
import copy
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, Protection, numbers
)
from openpyxl.utils import get_column_letter

# ─── MASTER TEMPLATE CONSTANTS ───────────────────────────────────────────────

# Colors
HESEM_BLUE   = "2C9CD7"
DARK_NAVY    = "0C2D48"
MED_BLUE     = "1565C0"
WHITE        = "FFFFFF"
LABEL_BG     = "F0F7FC"
META_LABEL   = "E4F0F9"
STRIPE_BG    = "F7FBFF"
HINT_BG      = "F8FAFB"
NOTICE_BG    = "FFF8E1"
THIN_BORDER  = "C5D9E8"
SIG_TEXT     = "BBCCCC"
LGRAY        = "7A8FA6"
GOLD_TEXT    = "7B4F00"
NOTICE_BORDER= "F9A825"
BODY_TEXT    = "1A2733"
SUBTITLE_TEXT= "1565C0"

# Fonts
FONT_FAMILY = "Segoe UI"

# Standard column width
STD_COL_WIDTH = 2.33

# Margins (inches)
MARGIN_LR = 0.197
MARGIN_TB = 0.197
MARGIN_HEADER = 0.1
MARGIN_FOOTER = 0.1

# Row heights
HEIGHT_RULE_STRIPE = 1.5
HEIGHT_HEADER_ROW  = 15.0
HEIGHT_SECTION_HDR = 18.0
HEIGHT_COL_HDR     = 18.0
HEIGHT_DATA_ROW    = 20.0
HEIGHT_SPACER      = 3.0
HEIGHT_SIGNATURE   = 26.0
HEIGHT_NOTICE      = 26.0

# Border styles
MEDIUM_BORDER_SIDE = Side(style='medium', color=HESEM_BLUE)
THIN_BORDER_SIDE   = Side(style='thin', color=THIN_BORDER)
NO_BORDER_SIDE     = Side(style=None)
NOTICE_BORDER_SIDE = Side(style='thin', color=NOTICE_BORDER)

# Fills
FILL_HESEM_BLUE  = PatternFill('solid', fgColor=HESEM_BLUE)
FILL_DARK_NAVY   = PatternFill('solid', fgColor=DARK_NAVY)
FILL_MED_BLUE    = PatternFill('solid', fgColor=MED_BLUE)
FILL_WHITE       = PatternFill('solid', fgColor=WHITE)
FILL_LABEL       = PatternFill('solid', fgColor=LABEL_BG)
FILL_META_LABEL  = PatternFill('solid', fgColor=META_LABEL)
FILL_STRIPE      = PatternFill('solid', fgColor=STRIPE_BG)
FILL_NOTICE      = PatternFill('solid', fgColor=NOTICE_BG)
FILL_NONE        = PatternFill(fill_type=None)

# Fonts
FONT_TITLE       = Font(name=FONT_FAMILY, size=14, bold=True, color=DARK_NAVY)
FONT_SECTION_HDR = Font(name=FONT_FAMILY, size=9, bold=True, color=WHITE)
FONT_COL_HDR     = Font(name=FONT_FAMILY, size=9, bold=True, color=WHITE)
FONT_LABEL       = Font(name=FONT_FAMILY, size=8, bold=True, color=DARK_NAVY)
FONT_BODY        = Font(name=FONT_FAMILY, size=8, bold=False, color=BODY_TEXT)
FONT_META_LABEL  = Font(name=FONT_FAMILY, size=8, bold=True, color=DARK_NAVY)
FONT_META_VALUE  = Font(name=FONT_FAMILY, size=8, bold=False, color=BODY_TEXT)
FONT_SUBTITLE    = Font(name=FONT_FAMILY, size=8, bold=False, color=SUBTITLE_TEXT)
FONT_TAGLINE     = Font(name=FONT_FAMILY, size=7, bold=False, color=LGRAY)
FONT_CRITERIA    = Font(name=FONT_FAMILY, size=7, bold=False, color=LGRAY)
FONT_NOTICE      = Font(name=FONT_FAMILY, size=7, bold=False, color=GOLD_TEXT)
FONT_SIG         = Font(name=FONT_FAMILY, size=8, bold=False, color=SIG_TEXT)
FONT_AUTONUM     = Font(name=FONT_FAMILY, size=8, bold=True, color=DARK_NAVY)

# Alignments
ALIGN_CENTER     = Alignment(horizontal='center', vertical='center', wrapText=True)
ALIGN_LEFT       = Alignment(horizontal='left', vertical='center', wrapText=True)
ALIGN_RIGHT_CTR  = Alignment(horizontal='right', vertical='center', wrapText=True)
ALIGN_SIG        = Alignment(horizontal='center', vertical='bottom', wrapText=False)

# ─── DETECTION HELPERS ────────────────────────────────────────────────────────

def color_hex(color):
    """Extract hex color string from openpyxl color object."""
    if color is None:
        return None
    if hasattr(color, 'rgb') and color.rgb and color.rgb != '00000000':
        rgb = str(color.rgb)
        if len(rgb) == 8:  # ARGB
            return rgb[2:].upper()
        return rgb.upper()
    if hasattr(color, 'theme') and color.theme is not None:
        return f"theme:{color.theme}"
    if hasattr(color, 'indexed') and color.indexed is not None:
        return f"indexed:{color.indexed}"
    return None

def fill_hex(fill):
    """Get fill color hex."""
    if fill is None or fill.fill_type is None:
        return None
    return color_hex(fill.fgColor)

def font_color_hex(font):
    """Get font color hex."""
    if font is None or font.color is None:
        return None
    return color_hex(font.color)

def detect_ncols(ws):
    """Detect the grid width (number of columns) based on column widths."""
    count = 0
    for col_idx in range(1, 200):
        col_letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(col_letter)
        if dim and dim.width is not None and abs(dim.width - STD_COL_WIDTH) < 0.5:
            count += 1
        elif count > 10:
            break
    # Round to nearest standard
    if count <= 44:
        return 40  # A4 Portrait
    elif count <= 60:
        return 56  # A4 Landscape / A3 Portrait
    else:
        return 82  # A3 Landscape

def get_header_zones(ncols):
    """Return (logo_end, title_start, title_end, meta_lbl_start, meta_lbl_end, meta_val_start, meta_val_end)."""
    if ncols <= 40:
        return 8, 9, 30, 31, 34, 35, 40
    elif ncols <= 56:
        return 11, 12, 42, 43, 48, 49, 56
    else:
        return 16, 17, 62, 63, 70, 71, 82

def is_rule_stripe_row(ws, row):
    """Check if this row is a rule stripe (row 1 or 7 typically)."""
    cell = ws.cell(row=row, column=1)
    fh = fill_hex(cell.fill)
    if fh and fh.upper() == HESEM_BLUE.upper():
        return True
    return False

def is_section_header_row(ws, row, ncols):
    """Check if row is a section header (full-width merge with MED_BLUE or similar bg)."""
    cell = ws.cell(row=row, column=1)
    fh = fill_hex(cell.fill)
    if fh and fh.upper() in (MED_BLUE.upper(), HESEM_BLUE.upper(), DARK_NAVY.upper()):
        # Check if it's merged across most columns
        for mg in ws.merged_cells.ranges:
            if mg.min_row == row and mg.max_row == row and (mg.max_col - mg.min_col + 1) >= ncols * 0.8:
                return True
        # Even without merge, if it's a colored full row
        if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
            return True
    return False

def is_column_header_row(ws, row, ncols):
    """Check if row is a column header row (DARK_NAVY bg)."""
    cell = ws.cell(row=row, column=1)
    fh = fill_hex(cell.fill)
    if fh and fh.upper() == DARK_NAVY.upper():
        return True
    # Check a few cells across
    for c in [1, 3, 5, 8]:
        if c <= ncols:
            fh2 = fill_hex(ws.cell(row=row, column=c).fill)
            if fh2 and fh2.upper() == DARK_NAVY.upper():
                return True
    return False

def is_sub_section_header(ws, row, ncols):
    """L2 sub-section: HESEM_BLUE bg."""
    for c in [1, 2]:
        fh = fill_hex(ws.cell(row=row, column=c).fill)
        if fh and fh.upper() == HESEM_BLUE.upper():
            # Make sure it's not rule stripe (height check)
            rd = ws.row_dimensions.get(row)
            if rd and rd.height and rd.height > 5:
                return True
    return False

def is_spacer_row(ws, row):
    """Spacer rows have height ~3pt and no content."""
    rd = ws.row_dimensions.get(row)
    if rd and rd.height is not None and rd.height <= 5:
        return True
    return False

def is_notice_row(ws, row):
    """Notice bar: NOTICE_BG fill."""
    cell = ws.cell(row=row, column=1)
    fh = fill_hex(cell.fill)
    if fh and fh.upper() == NOTICE_BG.upper().replace('#',''):
        return True
    return False

def is_signature_area(ws, row, ncols):
    """Detect signature area rows."""
    cell = ws.cell(row=row, column=1)
    val = str(cell.value or '').lower()
    if any(kw in val for kw in ['prepared by', 'reviewed by', 'approved by',
                                  'chuẩn bị', 'xem xét', 'phê duyệt',
                                  'người lập', 'người duyệt', 'name', 'signature', 'date']):
        return True
    fh = font_color_hex(cell.font)
    if fh and fh.upper() == SIG_TEXT.upper():
        return True
    return False


# ─── FIX FUNCTIONS ────────────────────────────────────────────────────────────

def fix_font_family(cell, issues):
    """Ensure cell uses Segoe UI font."""
    if cell.font and cell.font.name and cell.font.name != FONT_FAMILY:
        old_name = cell.font.name
        new_font = copy.copy(cell.font)
        new_font = Font(
            name=FONT_FAMILY,
            size=cell.font.size,
            bold=cell.font.bold,
            italic=cell.font.italic,
            underline=cell.font.underline,
            strike=cell.font.strikethrough,
            color=cell.font.color
        )
        cell.font = new_font
        issues.append(f"  Font: {old_name} -> {FONT_FAMILY}")
        return True
    return False

def fix_gridlines(ws, issues):
    """Hide gridlines."""
    view = ws.sheet_view if hasattr(ws, 'sheet_view') else None
    if hasattr(ws, 'views') and ws.views:
        view = ws.views.sheetView[0] if hasattr(ws.views, 'sheetView') else None

    if hasattr(ws, 'sheet_view'):
        if ws.sheet_view.showGridLines != False:
            ws.sheet_view.showGridLines = False
            issues.append("  Grid lines: shown -> hidden")
            return True
    return False

def fix_page_setup(ws, ncols, issues):
    """Fix page margins, paper size, fit to page."""
    changed = False

    # Margins
    if ws.page_margins:
        pm = ws.page_margins
        if abs(pm.left - MARGIN_LR) > 0.01:
            pm.left = MARGIN_LR
            changed = True
        if abs(pm.right - MARGIN_LR) > 0.01:
            pm.right = MARGIN_LR
            changed = True
        if abs(pm.top - MARGIN_TB) > 0.01:
            pm.top = MARGIN_TB
            changed = True
        if abs(pm.bottom - MARGIN_TB) > 0.01:
            pm.bottom = MARGIN_TB
            changed = True
        if abs(pm.header - MARGIN_HEADER) > 0.01:
            pm.header = MARGIN_HEADER
            changed = True
        if abs(pm.footer - MARGIN_FOOTER) > 0.01:
            pm.footer = MARGIN_FOOTER
            changed = True

    # Paper size
    ps = ws.page_setup
    if ncols <= 40 or ncols <= 56:
        target_paper = 9  # A4
    else:
        target_paper = 8  # A3

    if ps.paperSize != target_paper:
        ps.paperSize = target_paper
        changed = True

    # Fit to page
    if ps.fitToWidth != 1:
        ps.fitToWidth = 1
        changed = True
    if ps.fitToHeight != 0:
        ps.fitToHeight = 0
        changed = True

    if changed:
        issues.append("  Page setup: fixed margins/paper/fit")
    return changed

def fix_column_widths(ws, ncols, issues):
    """Set all grid columns to standard width."""
    changed = False
    for col_idx in range(1, ncols + 1):
        col_letter = get_column_letter(col_idx)
        dim = ws.column_dimensions[col_letter]
        if dim.width is None or abs(dim.width - STD_COL_WIDTH) > 0.3:
            old_w = dim.width
            dim.width = STD_COL_WIDTH
            changed = True
    if changed:
        issues.append(f"  Column widths: normalized to {STD_COL_WIDTH}")
    return changed

def fix_rule_stripe(ws, row, ncols, issues):
    """Fix rule stripe row formatting."""
    changed = False
    rd = ws.row_dimensions[row]
    if rd.height is None or abs(rd.height - HEIGHT_RULE_STRIPE) > 0.5:
        rd.height = HEIGHT_RULE_STRIPE
        changed = True

    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        fh = fill_hex(cell.fill)
        if fh is None or fh.upper() != HESEM_BLUE.upper():
            cell.fill = FILL_HESEM_BLUE
            changed = True
        # No borders on rule stripes
        cell.border = Border()

    if changed:
        issues.append(f"  Row {row}: rule stripe fixed (h=1.5, fill=HESEM_BLUE, no borders)")
    return changed

def fix_header_rows(ws, ncols, issues):
    """Fix header area rows 2-6 formatting."""
    changed = False
    zones = get_header_zones(ncols)
    logo_end, title_start, title_end, meta_lbl_start, meta_lbl_end, meta_val_start, meta_val_end = zones

    for row in range(2, 7):
        rd = ws.row_dimensions[row]
        if rd.height is None or abs(rd.height - HEIGHT_HEADER_ROW) > 1:
            rd.height = HEIGHT_HEADER_ROW
            changed = True

        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            fix_font_family(cell, [])  # silent font fix

            # Logo zone
            if c <= logo_end:
                if fill_hex(cell.fill) != WHITE.upper():
                    cell.fill = FILL_WHITE
                    changed = True

            # Title zone
            elif title_start <= c <= title_end:
                if row <= 4:  # Title rows
                    if cell.font.name != FONT_FAMILY or cell.font.size != 14:
                        cell.font = FONT_TITLE
                        changed = True
                    cell.alignment = ALIGN_CENTER
                elif row == 5:  # QMS subtitle
                    cell.font = FONT_SUBTITLE
                    cell.alignment = ALIGN_CENTER
                    changed = True
                elif row == 6:  # Tagline
                    cell.font = FONT_TAGLINE
                    cell.alignment = ALIGN_CENTER
                    changed = True

            # Meta label zone
            elif meta_lbl_start <= c <= meta_lbl_end:
                cell.fill = FILL_META_LABEL
                cell.font = FONT_META_LABEL
                cell.alignment = ALIGN_RIGHT_CTR
                changed = True

            # Meta value zone
            elif meta_val_start <= c <= meta_val_end:
                cell.fill = FILL_WHITE
                cell.font = FONT_META_VALUE
                cell.alignment = ALIGN_LEFT
                changed = True

    if changed:
        issues.append(f"  Rows 2-6: header zones fixed (logo/title/meta)")
    return changed

def fix_section_header(ws, row, ncols, level, issues):
    """Fix section header row formatting."""
    changed = False
    rd = ws.row_dimensions[row]

    if level == 1:  # L1 Primary
        target_fill = FILL_MED_BLUE
        target_font = FONT_SECTION_HDR
        target_height = HEIGHT_SECTION_HDR
    elif level == 2:  # L2 Sub-section
        target_fill = FILL_HESEM_BLUE
        target_font = Font(name=FONT_FAMILY, size=8, bold=True, color=WHITE)
        target_height = HEIGHT_SECTION_HDR
    else:
        return False

    if rd.height is None or abs(rd.height - target_height) > 1:
        rd.height = target_height
        changed = True

    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = target_fill
        cell.font = target_font
        cell.alignment = ALIGN_LEFT
        changed = True

    # Medium outline border
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        top = MEDIUM_BORDER_SIDE
        bottom = MEDIUM_BORDER_SIDE
        left = MEDIUM_BORDER_SIDE if c == 1 else NO_BORDER_SIDE
        right = MEDIUM_BORDER_SIDE if c == ncols else NO_BORDER_SIDE
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    if changed:
        issues.append(f"  Row {row}: L{level} section header fixed")
    return changed

def fix_column_header(ws, row, ncols, issues):
    """Fix column header row formatting."""
    changed = False
    rd = ws.row_dimensions[row]
    if rd.height is None or abs(rd.height - HEIGHT_COL_HDR) > 1:
        rd.height = HEIGHT_COL_HDR
        changed = True

    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_DARK_NAVY
        cell.font = FONT_COL_HDR
        cell.alignment = ALIGN_CENTER
        changed = True

    if changed:
        issues.append(f"  Row {row}: column header fixed (DARK_NAVY bg, 9pt bold white)")
    return changed

def fix_data_row(ws, row, ncols, is_even, issues):
    """Fix data row formatting (label+input pattern or table data)."""
    changed = False
    rd = ws.row_dimensions[row]
    if rd.height is None or abs(rd.height - HEIGHT_DATA_ROW) > 2:
        rd.height = HEIGHT_DATA_ROW
        changed = True

    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)

        # Fix font family
        if cell.font and cell.font.name and cell.font.name != FONT_FAMILY:
            new_font = Font(
                name=FONT_FAMILY,
                size=cell.font.size or 8,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=cell.font.color
            )
            cell.font = new_font
            changed = True

        # Fix font size if too large or too small
        if cell.font and cell.font.size and cell.font.size > 10:
            new_font = Font(
                name=FONT_FAMILY,
                size=8,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=cell.font.color
            )
            cell.font = new_font
            changed = True

        # Ensure vertical center alignment
        if cell.alignment and cell.alignment.vertical != 'center':
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or 'left',
                vertical='center',
                wrapText=True
            )
            changed = True

    return changed

def fix_spacer_row(ws, row, ncols, issues):
    """Fix spacer row."""
    changed = False
    rd = ws.row_dimensions[row]
    if rd.height is None or abs(rd.height - HEIGHT_SPACER) > 1:
        rd.height = HEIGHT_SPACER
        changed = True

    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_NONE
        cell.border = Border()

    if changed:
        issues.append(f"  Row {row}: spacer fixed (h=3)")
    return changed

def fix_notice_row(ws, row, ncols, issues):
    """Fix notice bar row."""
    changed = False
    rd = ws.row_dimensions[row]
    if rd.height is None or abs(rd.height - HEIGHT_NOTICE) > 2:
        rd.height = HEIGHT_NOTICE
        changed = True

    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_NOTICE
        cell.font = FONT_NOTICE
        cell.alignment = ALIGN_LEFT
        # Notice border
        cell.border = Border(
            top=NOTICE_BORDER_SIDE,
            bottom=NOTICE_BORDER_SIDE,
            left=NOTICE_BORDER_SIDE if c == 1 else NO_BORDER_SIDE,
            right=NOTICE_BORDER_SIDE if c == ncols else NO_BORDER_SIDE
        )
        changed = True

    if changed:
        issues.append(f"  Row {row}: notice bar fixed")
    return changed

def fix_signature_rows(ws, rows, ncols, issues):
    """Fix signature area rows."""
    changed = False
    for row in rows:
        rd = ws.row_dimensions[row]
        cell = ws.cell(row=row, column=1)
        val = str(cell.value or '').lower()

        # Label row
        is_label = any(kw in val for kw in ['prepared', 'reviewed', 'approved',
                                              'chuẩn bị', 'xem xét', 'phê duyệt',
                                              'name', 'tên', 'chữ ký', 'ngày'])
        if is_label:
            if rd.height is None or abs(rd.height - HEIGHT_DATA_ROW) > 2:
                rd.height = HEIGHT_DATA_ROW
                changed = True
        else:
            # Signature input rows
            if rd.height is None or abs(rd.height - HEIGHT_SIGNATURE) > 2:
                rd.height = HEIGHT_SIGNATURE
                changed = True

        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            if cell.font and cell.font.name != FONT_FAMILY:
                cell.font = Font(
                    name=FONT_FAMILY,
                    size=cell.font.size or 8,
                    bold=cell.font.bold,
                    color=cell.font.color
                )
                changed = True

    if changed:
        issues.append(f"  Rows {rows[0]}-{rows[-1]}: signature area fixed")
    return changed

def fix_all_fonts_in_sheet(ws, max_row, ncols, issues):
    """Sweep ALL used cells and fix any non-Segoe UI fonts.
    Skips non-primary cells in merged ranges (their font is irrelevant)."""
    count = 0

    # Build set of non-primary merged cells for fast lookup
    merged_non_primary = set()
    for mg in ws.merged_cells.ranges:
        for row in range(mg.min_row, mg.max_row + 1):
            for col in range(mg.min_col, mg.max_col + 1):
                if row != mg.min_row or col != mg.min_col:
                    merged_non_primary.add((row, col))

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            # Skip non-primary merged cells - font doesn't persist
            if (cell.row, cell.column) in merged_non_primary:
                continue
            if cell.font and cell.font.name and cell.font.name != FONT_FAMILY:
                has_content = cell.value is not None
                has_fill = cell.fill and cell.fill.fill_type is not None
                has_border = cell.border and any(
                    getattr(cell.border, s) and getattr(cell.border, s).style
                    for s in ['top', 'bottom', 'left', 'right']
                )
                if has_content or has_fill or has_border:
                    cell.font = Font(
                        name=FONT_FAMILY,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strikethrough,
                        color=cell.font.color
                    )
                    count += 1
    if count > 0:
        issues.append(f"  Font sweep: {count} cells changed to Segoe UI")
    return count > 0

def fix_all_borders_in_sheet(ws, max_row, ncols, issues):
    """Fix border colors and styles throughout the sheet.
    Only medium (HESEM_BLUE) and thin (THIN_BORDER) allowed."""
    count = 0
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.border is None:
                continue

            border = cell.border
            new_sides = {}
            modified = False

            for side_name in ['top', 'bottom', 'left', 'right']:
                side = getattr(border, side_name, None)
                if side is None or side.style is None:
                    new_sides[side_name] = side
                    continue

                side_color = color_hex(side.color) if side.color else None

                if side.style == 'medium':
                    # Medium borders should be HESEM_BLUE
                    if side_color and side_color.upper() != HESEM_BLUE.upper():
                        new_sides[side_name] = MEDIUM_BORDER_SIDE
                        modified = True
                    else:
                        new_sides[side_name] = side
                elif side.style == 'thin':
                    # Thin borders should be THIN_BORDER (except notice rows)
                    if is_notice_row(ws, cell.row):
                        if side_color and side_color.upper() != NOTICE_BORDER.upper():
                            new_sides[side_name] = NOTICE_BORDER_SIDE
                            modified = True
                        else:
                            new_sides[side_name] = side
                    else:
                        if side_color and side_color.upper() != THIN_BORDER.upper() and side_color.upper() != HESEM_BLUE.upper():
                            new_sides[side_name] = THIN_BORDER_SIDE
                            modified = True
                        else:
                            new_sides[side_name] = side
                elif side.style in ('thick', 'double', 'hair', 'dotted', 'dashed',
                                     'dashDot', 'dashDotDot', 'mediumDashed',
                                     'mediumDashDot', 'mediumDashDotDot', 'slantDashDot'):
                    # Non-standard border styles -> convert to thin
                    new_sides[side_name] = THIN_BORDER_SIDE
                    modified = True
                else:
                    new_sides[side_name] = side

            if modified:
                cell.border = Border(
                    top=new_sides.get('top'),
                    bottom=new_sides.get('bottom'),
                    left=new_sides.get('left'),
                    right=new_sides.get('right')
                )
                count += 1

    if count > 0:
        issues.append(f"  Border sweep: {count} cells corrected")
    return count > 0


# ─── MAIN AUDIT & FIX ────────────────────────────────────────────────────────

def audit_and_fix_workbook(filepath, dry_run=False):
    """Audit and fix a single workbook."""
    filename = os.path.basename(filepath)
    issues = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return filename, [f"  ERROR: Cannot open: {e}"], False

    any_changed = False

    # Fix workbook default font via _named_styles
    try:
        for ns in wb._named_styles:
            if ns.font and ns.font.name and ns.font.name != FONT_FAMILY:
                ns.font = Font(
                    name=FONT_FAMILY,
                    size=ns.font.size,
                    bold=ns.font.bold,
                    italic=ns.font.italic,
                    color=ns.font.color
                )
                any_changed = True
                issues.append(f"  Named style '{ns.name}' font -> Segoe UI")
    except Exception:
        pass

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # Skip hidden sheets and LISTS sheet
        if ws_name.upper() in ('LISTS', 'DROPDOWN', 'DATA'):
            continue

        ws_issues = []
        ncols = detect_ncols(ws)
        max_row = ws.max_row or 50

        # 1. Fix gridlines
        try:
            fix_gridlines(ws, ws_issues)
        except:
            pass

        # 2. Fix page setup
        fix_page_setup(ws, ncols, ws_issues)

        # 3. Fix column widths
        fix_column_widths(ws, ncols, ws_issues)

        # 4. Identify and fix rows by type
        rule_stripe_rows = []
        section_header_rows = []
        col_header_rows = []
        spacer_rows = []
        notice_rows = []
        sig_rows = []
        data_rows = []

        for row in range(1, max_row + 1):
            if is_rule_stripe_row(ws, row):
                rule_stripe_rows.append(row)
            elif is_section_header_row(ws, row, ncols):
                section_header_rows.append(row)
            elif is_sub_section_header(ws, row, ncols):
                section_header_rows.append(row)
            elif is_column_header_row(ws, row, ncols):
                col_header_rows.append(row)
            elif is_spacer_row(ws, row):
                spacer_rows.append(row)
            elif is_notice_row(ws, row):
                notice_rows.append(row)
            elif is_signature_area(ws, row, ncols):
                sig_rows.append(row)
            elif row >= 8:
                data_rows.append(row)

        # Fix rule stripes
        for r in rule_stripe_rows:
            if fix_rule_stripe(ws, r, ncols, ws_issues):
                any_changed = True

        # Fix header (rows 2-6)
        if 1 in rule_stripe_rows or any(r <= 6 for r in rule_stripe_rows):
            if fix_header_rows(ws, ncols, ws_issues):
                any_changed = True

        # Fix section headers
        for r in section_header_rows:
            fh = fill_hex(ws.cell(row=r, column=1).fill)
            level = 1
            if fh and fh.upper() == HESEM_BLUE.upper():
                rd = ws.row_dimensions.get(r)
                if rd and rd.height and rd.height > 5:
                    level = 2
            if fix_section_header(ws, r, ncols, level, ws_issues):
                any_changed = True

        # Fix column headers
        for r in col_header_rows:
            if fix_column_header(ws, r, ncols, ws_issues):
                any_changed = True

        # Fix spacer rows
        for r in spacer_rows:
            if fix_spacer_row(ws, r, ncols, ws_issues):
                any_changed = True

        # Fix notice rows
        for r in notice_rows:
            if fix_notice_row(ws, r, ncols, ws_issues):
                any_changed = True

        # Fix signature rows
        if sig_rows:
            if fix_signature_rows(ws, sig_rows, ncols, ws_issues):
                any_changed = True

        # Fix data rows
        for idx, r in enumerate(data_rows):
            if fix_data_row(ws, r, ncols, idx % 2 == 1, ws_issues):
                any_changed = True

        # 5. Global border sweep (before font sweep, as it may create cells)
        if fix_all_borders_in_sheet(ws, max_row, ncols, ws_issues):
            any_changed = True

        # 6. Global font sweep (LAST - catches all cells including those created by border sweep)
        if fix_all_fonts_in_sheet(ws, max_row, ncols, ws_issues):
            any_changed = True

        if ws_issues:
            issues.append(f"  Sheet '{ws_name}':")
            issues.extend(ws_issues)

    if any_changed and not dry_run:
        try:
            wb.save(filepath)
            issues.append(f"  >> SAVED")
        except Exception as e:
            issues.append(f"  >> SAVE ERROR: {e}")
    elif not any_changed:
        issues.append(f"  (no changes needed)")

    wb.close()
    return filename, issues, any_changed


def main():
    base_dir = r"C:\Users\TEST4\qms.hesem.com.vn\04-Bieu-Mau"
    dry_run = '--dry-run' in sys.argv

    # Collect all FRM files (exclude _build)
    frm_files = []
    for root, dirs, files in os.walk(base_dir):
        # Skip _build directory
        if '_build' in root:
            continue
        for f in files:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                frm_files.append(os.path.join(root, f))

    frm_files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS FORM COMPLIANCE {'AUDIT' if dry_run else 'FIX'}")
    print(f"Master Template: hesem-master-template.xlsx v4.2")
    print(f"Files to process: {len(frm_files)}")
    print(f"{'='*70}\n")

    total_fixed = 0
    total_ok = 0
    results = []

    for i, fpath in enumerate(frm_files, 1):
        fname = os.path.basename(fpath)
        print(f"[{i:3d}/{len(frm_files)}] {fname}...", end=" ", flush=True)

        filename, issues, changed = audit_and_fix_workbook(fpath, dry_run=dry_run)

        if changed:
            total_fixed += 1
            print("FIXED")
        else:
            total_ok += 1
            print("OK")

        results.append((filename, issues, changed))

    # Summary
    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"{'='*70}")
    print(f"Total files:  {len(frm_files)}")
    print(f"Fixed:        {total_fixed}")
    print(f"Already OK:   {total_ok}")
    print(f"{'='*70}\n")

    # Detailed report
    print("DETAILED REPORT:")
    print("-"*70)
    for filename, issues, changed in results:
        status = "FIXED" if changed else "OK"
        print(f"\n[{status}] {filename}")
        for iss in issues:
            print(iss)

    return 0 if total_fixed == 0 else 1


if __name__ == '__main__':
    main()
