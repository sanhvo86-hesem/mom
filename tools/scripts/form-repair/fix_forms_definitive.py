#!/usr/bin/env python3
"""
HESEM QMS Form Definitive Fixer
Transforms all FRM files to match frm-000-master-template.xlsx EXACTLY.

TARGET spec (from actual master template):
  Borders:         thin/#0078D4 (blue outlines) + thin/#E2E8F0 (internal gray)
  Section headers: fill=#EFF6FF, font=9pt bold #1B3A6B, left-aligned
  Column headers:  fill=#005A9E, font=9pt bold #FFFFFF, center
  Labels:          fill=#F1F5F9, font=8pt bold #334155
  Inputs:          fill=#FFFFFF, font=8pt #1A2733 or similar
  Spacers:         height=4, fill=white, NO borders
  Signature label: fill=#F1F5F9, all 4 borders=blue, font=8pt bold #334155, center
  Signature box:   fill=#FFFFFF, all 4 borders=blue, font=8pt #CBD5E1, center, bottom
  Notice bar:      fill=#EFF6FF, all 4 borders=blue, font=7pt #334155
  No rule stripes.
  Header: rows 1-5 (logo/title/meta), row 6=spacer(4pt)
"""
import os
import sys
import glob as globmod

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

FONT = "Segoe UI"

# ─── TARGET COLORS ────────────────────────────────────────────────────────────
# Border
BLUE_BRD = "0078D4"
GRAY_BRD = "E2E8F0"

BB = Side(style='thin', color=BLUE_BRD)
BG = Side(style='thin', color=GRAY_BRD)
BN = Side(style=None)

# Fills
SECTION_FILL = PatternFill('solid', fgColor='EFF6FF')
COLHDR_FILL  = PatternFill('solid', fgColor='005A9E')
LABEL_FILL   = PatternFill('solid', fgColor='F1F5F9')
WHITE_FILL   = PatternFill('solid', fgColor='FFFFFF')
NONE_FILL    = PatternFill(fill_type=None)

# Fonts
SECTION_FONT = Font(name=FONT, size=9, bold=True, color='1B3A6B')
COLHDR_FONT  = Font(name=FONT, size=9, bold=True, color='FFFFFF')
LABEL_FONT   = Font(name=FONT, size=8, bold=True, color='334155')
INPUT_FONT   = Font(name=FONT, size=8, bold=False, color='1A2733')
SIG_FONT     = Font(name=FONT, size=8, bold=False, color='CBD5E1')
NOTICE_FONT  = Font(name=FONT, size=7, bold=False, color='334155')
TITLE_FONT   = Font(name=FONT, size=15, bold=True, color='1B3A6B')
SUB_FONT     = Font(name=FONT, size=8, bold=False, color='0078D4')
TAG_FONT     = Font(name=FONT, size=7, bold=False, color='64748B')
METALBL_FONT = Font(name=FONT, size=8, bold=True, color='334155')
METAVAL_FONT = Font(name=FONT, size=8, bold=False, color='1A2733')

# Alignments
A_LEFT   = Alignment(horizontal='left', vertical='center', wrapText=True)
A_CENTER = Alignment(horizontal='center', vertical='center', wrapText=True)
A_RIGHT  = Alignment(horizontal='right', vertical='center', wrapText=True)
A_SIG    = Alignment(horizontal='center', vertical='bottom', wrapText=False)

STD_COL = 2.33
LOGO_PATH = os.path.join(os.path.dirname(__file__), 'assets', 'hesem-logo.png')

# ─── OLD COLORS TO DETECT ────────────────────────────────────────────────────
OLD_SECTION_BGS  = {'1565C0', 'EFF6FF', 'E8F0FE', 'DCEEFB', 'EFF6FF'}
OLD_COLHDR_BGS   = {'0C2D48', '005A9E', '1B4472', '1F3864'}
OLD_LABEL_BGS    = {'F0F7FC', 'F1F5F9', 'E4F0F9'}
OLD_RULE_STRIPE  = '2C9CD7'
OLD_NOTICE_BGS   = {'FFF8E1', 'EFF6FF'}

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def color_hex(c):
    if c is None: return None
    if hasattr(c, 'rgb') and c.rgb and str(c.rgb) != '00000000':
        s = str(c.rgb)
        return s[2:].upper() if len(s) == 8 else s.upper()
    return None

def fill_c(cell):
    if cell.fill and cell.fill.fill_type:
        return color_hex(cell.fill.fgColor)
    return None

def detect_ncols(ws):
    count = 0
    for i in range(1, 200):
        d = ws.column_dimensions.get(get_column_letter(i))
        if d and d.width is not None and abs(d.width - STD_COL) < 0.5:
            count += 1
        elif count > 10:
            break
    return 40 if count <= 44 else (56 if count <= 60 else 82)

def header_zones(n):
    if n <= 40: return 8, 9, 30, 31, 34, 35, 40
    elif n <= 56: return 11, 12, 42, 43, 48, 49, 56
    else: return 16, 17, 62, 63, 70, 71, 82

def has_header(ws):
    """STRICT header detection: requires logo zone merge + title + Form Code.
    Supports different grid widths (20-col, 40-col, 56-col, 82-col)."""
    has_logo = False
    has_title = False
    has_formcode = False

    for mg in ws.merged_cells.ranges:
        # Logo zone: starts at col 1, spans at least 3 rows in first 5 rows
        if mg.min_col == 1 and mg.min_row <= 2 and (mg.max_row - mg.min_row) >= 2:
            has_logo = True
        # Title zone: not col 1, spans at least 2 cols, has text content
        if mg.min_col >= 2 and mg.max_col >= mg.min_col + 3 and mg.min_row <= 3 and (mg.max_row - mg.min_row) >= 1:
            cell = ws.cell(row=mg.min_row, column=mg.min_col)
            if cell.value and isinstance(cell.value, str) and len(cell.value) >= 5:
                has_title = True

    # Check for "Form Code" text anywhere in rows 1-2
    for r in range(1, 3):
        for c in range(1, min(ws.max_column or 60, 100) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value and 'Form Code' in str(cell.value):
                has_formcode = True
                break
        if has_formcode:
            break

    return has_logo and has_title and has_formcode

def row_merges(ws, row):
    result = []
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= row <= mg.max_row:
            result.append((mg.min_col, mg.max_col))
    result.sort()
    return result

def merge_ends(ws, row):
    return set(mx for _, mx in row_merges(ws, row))

def merged_non_primary(ws):
    s = set()
    for mg in ws.merged_cells.ranges:
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r != mg.min_row or c != mg.min_col:
                    s.add((r, c))
    return s

# ─── CLASSIFY ─────────────────────────────────────────────────────────────────

def classify(ws, row, ncols):
    rd = ws.row_dimensions.get(row)
    h = rd.height if rd else None
    c1 = ws.cell(row=row, column=1)
    fh = fill_c(c1)
    val = str(c1.value or '')

    # Rule stripe (to be converted to spacer)
    if fh == OLD_RULE_STRIPE and h and h <= 3:
        return 'rule_stripe'

    # Spacer
    if h and h <= 5 and not val.strip():
        return 'spacer'

    # Section header
    if fh and fh in OLD_SECTION_BGS:
        for mg in ws.merged_cells.ranges:
            if mg.min_row == row and mg.max_row == row and (mg.max_col - mg.min_col + 1) >= ncols * 0.5:
                # Check it's not a notice
                if 'NOTICE' in val.upper() or 'LƯU Ý' in val.upper():
                    return 'notice'
                return 'section'
        if val.strip() and len(val.strip()) > 3:
            if 'NOTICE' in val.upper():
                return 'notice'
            return 'section'

    # Column header
    if fh and fh in OLD_COLHDR_BGS:
        return 'col_header'
    for cc in [2, 3, 5, 8]:
        if cc <= ncols:
            if fill_c(ws.cell(row=row, column=cc)) in OLD_COLHDR_BGS:
                return 'col_header'

    return 'data'

# ─── TRANSFORMS ───────────────────────────────────────────────────────────────

def fix_header(ws, ncols):
    """Transform header: fix colors, borders, add logo.
    ADAPTIVE: detects actual merge zones from the sheet."""
    # Detect actual zones from existing merges
    logo_max_col = 1
    title_min_col = 2
    title_max_col = ncols
    meta_min_col = ncols + 1  # no meta zone by default
    meta_lbl_end = ncols
    meta_val_start = ncols

    for mg in ws.merged_cells.ranges:
        if mg.min_row > 5:
            continue
        # Logo zone: starts at col 1, multi-row
        if mg.min_col == 1 and (mg.max_row - mg.min_row) >= 2:
            logo_max_col = max(logo_max_col, mg.max_col)
        # Form Code / meta labels
        cell = ws.cell(row=mg.min_row, column=mg.min_col)
        val = str(cell.value or '')
        if 'Form Code' in val or 'Revision' in val or 'Eff. Date' in val or 'Owner' in val or 'Approved' in val:
            meta_min_col = min(meta_min_col, mg.min_col)
            meta_lbl_end = max(meta_lbl_end if meta_lbl_end != ncols else 0, mg.max_col)
        # Meta values: merge right after meta labels on same row
        if mg.min_col > meta_lbl_end and mg.min_row <= 5 and meta_min_col <= ncols:
            fh = fill_c(cell)
            if fh == 'FFFFFF' or fh is None:
                meta_val_start = min(meta_val_start, mg.min_col)

    # Detect meta value zone from merges adjacent to Form Code
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 5 and mg.min_col > logo_max_col:
            cell = ws.cell(row=mg.min_row, column=mg.min_col)
            val = str(cell.value or '')
            if val and ('FRM-' in val or 'V0' in val or 'Rev' in val or 'YYYY' in val or 'Manager' in val):
                meta_val_start = min(meta_val_start, mg.min_col)

    title_min_col = logo_max_col + 1
    if meta_min_col <= ncols:
        title_max_col = meta_min_col - 1
        if meta_val_start > meta_lbl_end:
            meta_lbl_end_actual = meta_val_start - 1
        else:
            meta_lbl_end_actual = meta_lbl_end
    else:
        meta_lbl_end_actual = ncols

    hdr_start = 1
    hdr_end = 5

    for r in range(hdr_start, hdr_end + 1):
        ws.row_dimensions[r].height = 15.0
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)

            # Logo zone
            if c <= logo_max_col:
                cell.fill = WHITE_FILL

            # Title zone
            elif title_min_col <= c <= title_max_col:
                cell.fill = WHITE_FILL
                if r <= 3:
                    cell.font = TITLE_FONT
                    cell.alignment = A_CENTER
                elif r == 4:
                    cell.font = SUB_FONT
                    cell.alignment = A_CENTER
                elif r == 5:
                    cell.font = TAG_FONT
                    cell.alignment = A_CENTER

            # Meta label zone
            elif meta_min_col <= ncols and meta_min_col <= c <= meta_lbl_end_actual:
                cell.fill = LABEL_FILL
                cell.font = METALBL_FONT
                cell.alignment = A_RIGHT

            # Meta value zone
            elif meta_val_start <= c <= ncols:
                cell.fill = WHITE_FILL
                cell.font = METAVAL_FONT
                cell.alignment = A_LEFT

            # BORDERS
            top = BB if r == hdr_start else BN
            bottom = BB if r == hdr_end else BN
            left = BB if c == 1 else BN
            right = BB if c == ncols else BN

            # Zone dividers
            if c == logo_max_col: right = BB
            if c == logo_max_col + 1: left = BB
            if meta_min_col <= ncols and c == meta_min_col: left = BB

            # Meta internal gray borders
            if meta_min_col <= ncols and c >= meta_min_col:
                if r > hdr_start: top = BG
                if c == meta_lbl_end_actual: right = BG

            try:
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)
            except:
                pass

    # Embed logo - centered in logo zone
    if os.path.exists(LOGO_PATH):
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker

        ws._images = []
        img = XlImage(LOGO_PATH)

        zone_w_mm = logo_max_col * 5.03
        zone_h_mm = 75 * 0.353  # 5 rows x 15pt

        asp = 52.0 / 15.0
        avail_w = zone_w_mm * 0.80
        avail_h = zone_h_mm * 0.55
        if avail_w / asp <= avail_h:
            w_mm, h_mm = avail_w, avail_w / asp
        else:
            h_mm, w_mm = avail_h, avail_h * asp

        img.width = w_mm / 25.4 * 96
        img.height = h_mm / 25.4 * 96

        # Center in logo zone
        x_off = int((zone_w_mm - w_mm) / 2 * 36000)
        y_off = int((zone_h_mm - h_mm) / 2 * 36000)

        marker = AnchorMarker(col=0, colOff=x_off,
                              row=hdr_start - 1, rowOff=y_off)
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        w_emu = int(img.width * 9525)
        h_emu = int(img.height * 9525)

        img.anchor = OneCellAnchor(_from=marker,
                                    ext=XDRPositiveSize2D(w_emu, h_emu))
        ws.add_image(img)


def fix_spacer(ws, row, ncols):
    """Convert to proper spacer: 4pt, white, no borders."""
    # Unmerge if needed
    to_remove = [mg for mg in ws.merged_cells.ranges if mg.min_row == row and mg.max_row == row]
    for mg in to_remove:
        ws.unmerge_cells(str(mg))

    ws.row_dimensions[row].height = 4.0
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = WHITE_FILL
        try:
            cell.value = None
        except:
            pass
        try:
            cell.border = Border()
        except:
            pass


def fix_section(ws, row, ncols):
    """Section header: fill=EFF6FF, font=9pt bold #1B3A6B, all borders=blue."""
    ws.row_dimensions[row].height = 17.0
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = A_LEFT
        try:
            cell.border = Border(top=BB, bottom=BB,
                                 left=BB if c == 1 else BN,
                                 right=BB if c == ncols else BN)
        except:
            pass


def fix_col_header(ws, row, ncols):
    """Column header: fill=005A9E, font=9pt bold white, T=blue B=gray R=gray."""
    ws.row_dimensions[row].height = 17.0
    me = merge_ends(ws, row)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = COLHDR_FILL
        cell.font = COLHDR_FONT
        cell.alignment = A_CENTER
        right = BB if c == ncols else (BG if c in me else BN)
        try:
            cell.border = Border(top=BB, bottom=BG,
                                 left=BB if c == 1 else BN,
                                 right=right)
        except:
            pass


def fix_data_block(ws, rows, ncols):
    """Data rows: fix colors + borders.
    First row: T=blue, B=gray. Middle: T=none, B=gray. Last: T=none, B=blue.
    L=blue, R=blue. Internal R=gray at merge boundaries."""
    if not rows:
        return
    first, last = rows[0], rows[-1]

    for r in rows:
        ws.row_dimensions[r].height = 20.0
        me = merge_ends(ws, r)

        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)

            # Fix fill colors
            fc = fill_c(cell)
            if fc and fc in OLD_LABEL_BGS:
                cell.fill = LABEL_FILL
                cell.font = LABEL_FONT
            elif fc == 'FFFFFF' or fc is None:
                # Input cells stay white, fix font
                if cell.font and cell.font.name:
                    cell.font = Font(name=FONT, size=cell.font.size or 8,
                                    bold=cell.font.bold, italic=cell.font.italic,
                                    color=cell.font.color)

            # Borders
            if r == first and r == last:
                top, bottom = BB, BB
            elif r == first:
                top, bottom = BB, BG
            elif r == last:
                top, bottom = BN, BB
            else:
                top, bottom = BN, BG

            left = BB if c == 1 else BN
            right = BB if c == ncols else (BG if c in me else BN)

            try:
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)
            except:
                pass


def fix_signature(ws, rows, ncols):
    """Signature rows: all borders=blue. Fix fill/font."""
    for r in rows:
        me = merge_ends(ws, r)
        ms = set(mc for mc, _ in row_merges(ws, r))

        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            fc = fill_c(cell)

            # Signature label (LABEL fill)
            if fc and fc in OLD_LABEL_BGS:
                cell.fill = LABEL_FILL
                cell.font = LABEL_FONT
                cell.alignment = A_CENTER
            # Signature placeholder (white)
            elif fc == 'FFFFFF' or (cell.font and color_hex(cell.font.color) in ('CBD5E1', 'BBCCCC')):
                cell.fill = WHITE_FILL
                cell.font = SIG_FONT
                cell.alignment = A_SIG
                ws.row_dimensions[r].height = 26.0

            left = BB if (c == 1 or c in ms) else BN
            right = BB if (c == ncols or c in me) else BN
            try:
                cell.border = Border(top=BB, bottom=BB, left=left, right=right)
            except:
                pass


def fix_notice(ws, row, ncols):
    """Notice bar: fill=EFF6FF, borders=blue, font=7pt #334155."""
    ws.row_dimensions[row].height = 24.0
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL  # Same EFF6FF as section headers
        cell.font = NOTICE_FONT
        cell.alignment = A_LEFT
        try:
            cell.border = Border(top=BB, bottom=BB,
                                 left=BB if c == 1 else BN,
                                 right=BB if c == ncols else BN)
        except:
            pass


def fix_fonts_sweep(ws):
    """Final sweep: fix remaining non-Segoe UI fonts."""
    mnp = merged_non_primary(ws)
    count = 0
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if (cell.row, cell.column) in mnp:
                continue
            if cell.font and cell.font.name and cell.font.name != FONT:
                has = (cell.value is not None or
                       (cell.fill and cell.fill.fill_type is not None) or
                       (cell.border and any(getattr(cell.border, s) and getattr(cell.border, s).style
                                           for s in ['top','bottom','left','right'])))
                if has:
                    cell.font = Font(name=FONT, size=cell.font.size, bold=cell.font.bold,
                                    italic=cell.font.italic, underline=cell.font.underline,
                                    strike=cell.font.strikethrough, color=cell.font.color)
                    count += 1
    return count


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def process(filepath):
    fname = os.path.basename(filepath)
    log = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, [f"ERROR: {e}"], False

    # Fix Normal style
    try:
        for ns in wb._named_styles:
            if ns.font and ns.font.name and ns.font.name != FONT:
                ns.font = Font(name=FONT, size=ns.font.size, bold=ns.font.bold,
                              italic=ns.font.italic, color=ns.font.color)
    except:
        pass

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # STRICT: Only process sheets with proper 5-row header
        # All other sheets: only fix fonts, do NOT touch colors/borders/structure
        hdr = has_header(ws)
        if not hdr:
            fix_fonts_sweep(ws)
            log.append(f"  [{ws_name}] fonts only (no header)")
            continue

        ncols = detect_ncols(ws)
        max_row = ws.max_row or 50

        # 1. Fix header (we know it has one from strict check above)
        fix_header(ws, ncols)
        log.append(f"  [{ws_name}] header + logo")

        # Find spacer after header
        hdr_end = 6  # default
        for r in range(5, 10):
            rt = classify(ws, r, ncols)
            if rt in ('spacer', 'rule_stripe'):
                hdr_end = r
                break

        # 2. Process body rows
        block = []
        sig_rows = []

        for r in range(max(hdr_end, 1), max_row + 1):
            rt = classify(ws, r, ncols)

            if rt in ('spacer', 'rule_stripe'):
                if block: fix_data_block(ws, block, ncols); block = []
                if sig_rows: fix_signature(ws, sig_rows, ncols); sig_rows = []
                fix_spacer(ws, r, ncols)

            elif rt == 'section':
                if block: fix_data_block(ws, block, ncols); block = []
                if sig_rows: fix_signature(ws, sig_rows, ncols); sig_rows = []
                fix_section(ws, r, ncols)

            elif rt == 'col_header':
                if block: fix_data_block(ws, block, ncols); block = []
                if sig_rows: fix_signature(ws, sig_rows, ncols); sig_rows = []
                fix_col_header(ws, r, ncols)

            elif rt == 'notice':
                if block: fix_data_block(ws, block, ncols); block = []
                if sig_rows: fix_signature(ws, sig_rows, ncols); sig_rows = []
                fix_notice(ws, r, ncols)

            else:
                # Check signature
                c1val = str(ws.cell(row=r, column=1).value or '').lower()
                fn_clr = color_hex(ws.cell(row=r, column=1).font.color) if ws.cell(row=r, column=1).font and ws.cell(row=r, column=1).font.color else None
                is_sig = any(kw in c1val for kw in ['prepared', 'reviewed', 'approved',
                                                      'name', 'signature', 'date',
                                                      'chuẩn bị', 'xem xét', 'phê duyệt'])
                is_sig_box = fn_clr in ('CBD5E1', 'BBCCCC')

                if is_sig or is_sig_box:
                    if block: fix_data_block(ws, block, ncols); block = []
                    sig_rows.append(r)
                else:
                    if sig_rows: fix_signature(ws, sig_rows, ncols); sig_rows = []
                    block.append(r)

        if block: fix_data_block(ws, block, ncols)
        if sig_rows: fix_signature(ws, sig_rows, ncols)

        # 3. Font sweep
        n = fix_fonts_sweep(ws)

        # 4. Page setup
        try:
            ws.sheet_view.showGridLines = False
        except: pass
        try:
            pm = ws.page_margins
            pm.left = pm.right = 0.197
            pm.top = pm.bottom = 0.197
            pm.header = pm.footer = 0.1
            ps = ws.page_setup
            ps.paperSize = 9 if ncols <= 56 else 8
            ps.fitToWidth = 1
            ps.fitToHeight = 0
        except: pass

        # 5. Fix footer font size to 7pt (matching master template)
        try:
            footer_text = str(ws.oddFooter or '')
            if footer_text and '&7' not in footer_text:
                # Add &7 font size after &C (centered)
                if footer_text.startswith('&C'):
                    footer_text = '&C&"-,Regular"&7' + footer_text[2:]
                else:
                    footer_text = '&C&"-,Regular"&7' + footer_text
                ws.oddFooter = footer_text
        except: pass

        log.append(f"  [{ws_name}] done")

    try:
        wb.save(filepath)
        log.append("  >> SAVED")
    except Exception as e:
        log.append(f"  >> SAVE ERROR: {e}")

    wb.close()
    return fname, log, True


def main():
    base = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS DEFINITIVE FIX")
    print(f"Template: frm-000-master-template.xlsx")
    print(f"Logo: {'OK' if os.path.exists(LOGO_PATH) else 'MISSING'}")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    v = '--verbose' in sys.argv
    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        print(f"[{i:3d}/{len(files)}] {fn}...", end=" ", flush=True)
        _, log, ok = process(fp)
        print("OK" if ok else "FAIL")
        if v:
            for l in log: print(l)

    print(f"\n{'='*70}")
    print(f"DONE - {len(files)} files")
    print(f"{'='*70}")

if __name__ == '__main__':
    main()
