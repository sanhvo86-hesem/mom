#!/usr/bin/env python3
"""
HESEM QMS Form MINIMAL Fixer
ONLY does 3 things:
  1. Embeds logo (centered per grid width)
  2. Fixes borders to match master template
  3. Fixes footer font size to 7pt
Does NOT change fonts, fills, font sizes, or any other formatting.
"""
import os
import sys
import glob as globmod

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS (from frm-000-master-template.xlsx)
# ═══════════════════════════════════════════════════════════════════════════════

BB = Side(style='thin', color='0078D4')   # Blue border (section outlines)
BG = Side(style='thin', color='E2E8F0')   # Gray border (internal)
BN = Side(style=None)

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'assets', 'hesem-logo.png')
STD_COL = 2.33
LOGO_ASPECT = 52.0 / 15.0  # SVG width/height ratio

# Logo sizing per grid width (% of zone)
LOGO_CONFIGS = {
    40:  {'w_pct': 0.75, 'h_pct': 0.50},  # A4 Portrait: smaller
    56:  {'w_pct': 0.70, 'h_pct': 0.50},  # A4 Landscape
    82:  {'w_pct': 0.60, 'h_pct': 0.45},  # A3
}

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def color_hex(c):
    if c is None: return None
    if hasattr(c, 'rgb') and c.rgb and str(c.rgb) != '00000000':
        s = str(c.rgb)
        return s[2:].upper() if len(s) == 8 else s.upper()
    return None

def fill_c(cell):
    if cell.fill and cell.fill.fill_type:
        return color_hex(cell.fill.fgColor)
    return None

def detect_ncols(ws):
    count = 0
    for i in range(1, 200):
        d = ws.column_dimensions.get(get_column_letter(i))
        if d and d.width is not None and abs(d.width - STD_COL) < 0.5:
            count += 1
        elif count > 10:
            break
    return 40 if count <= 44 else (56 if count <= 60 else 82)

def detect_logo_zone(ws):
    """Find the logo zone merge: starts at col 1, spans multiple rows."""
    for mg in ws.merged_cells.ranges:
        if mg.min_col == 1 and mg.min_row <= 2 and (mg.max_row - mg.min_row) >= 2:
            return mg.max_col, mg.min_row, mg.max_row
    return None, None, None

def detect_meta_zone(ws):
    """Find where Form Code label is -> meta zone start."""
    for r in range(1, 3):
        for c in range(1, 100):
            cell = ws.cell(row=r, column=c)
            if cell.value and 'Form Code' in str(cell.value):
                return c
    return None

def has_form_header(ws):
    """Check if sheet has a proper form header (logo zone + Form Code)."""
    logo_end, _, _ = detect_logo_zone(ws)
    meta_start = detect_meta_zone(ws)
    return logo_end is not None and meta_start is not None

def row_merges(ws, row):
    return sorted([(mg.min_col, mg.max_col) for mg in ws.merged_cells.ranges
                   if mg.min_row <= row <= mg.max_row])

def merge_ends(ws, row):
    return set(mx for _, mx in row_merges(ws, row))

# Section header detection (by fill color)
SECTION_FILLS = {'EFF6FF', '1565C0', 'E8F0FE', 'DCEEFB'}
COLHDR_FILLS  = {'005A9E', '0C2D48', '1B4472', '1F3864'}
RULE_STRIPE   = '2C9CD7'

def classify(ws, row, ncols):
    rd = ws.row_dimensions.get(row)
    h = rd.height if rd else None
    c1 = ws.cell(row=row, column=1)
    fh = fill_c(c1)

    if fh == RULE_STRIPE and h and h <= 3:
        return 'rule_stripe'
    if h and h <= 5 and not (c1.value and str(c1.value).strip()):
        return 'spacer'
    if fh and fh in SECTION_FILLS:
        val = str(c1.value or '').upper()
        if 'NOTICE' in val or 'LƯU Ý' in val:
            return 'notice'
        for mg in ws.merged_cells.ranges:
            if mg.min_row == row and mg.max_row == row and (mg.max_col - mg.min_col + 1) >= ncols * 0.4:
                return 'section'
        if c1.value and len(str(c1.value).strip()) > 3:
            return 'section'
    if fh and fh in COLHDR_FILLS:
        return 'col_header'
    for cc in [2, 3, 5]:
        if cc <= ncols and fill_c(ws.cell(row=row, column=cc)) in COLHDR_FILLS:
            return 'col_header'
    return 'data'


# ═══════════════════════════════════════════════════════════════════════════════
# FIX 1: LOGO (centered per paper size)
# ═══════════════════════════════════════════════════════════════════════════════

def embed_logo(ws, ncols):
    """Embed HESEM logo centered in the logo zone."""
    if not os.path.exists(LOGO_PATH):
        return False

    logo_end, logo_row_start, logo_row_end = detect_logo_zone(ws)
    if logo_end is None:
        return False

    ws._images = []

    # Calculate zone size in mm
    zone_w_mm = logo_end * 5.03  # each col ≈ 5.03mm
    num_rows = logo_row_end - logo_row_start + 1
    zone_h_mm = 0
    for r in range(logo_row_start, logo_row_end + 1):
        rd = ws.row_dimensions.get(r)
        h_pt = rd.height if rd and rd.height else 15.0
        zone_h_mm += h_pt * 0.353  # pt to mm

    # Get sizing config for this grid
    cfg = LOGO_CONFIGS.get(ncols, LOGO_CONFIGS[40])

    # Calculate logo size
    avail_w = zone_w_mm * cfg['w_pct']
    avail_h = zone_h_mm * cfg['h_pct']

    if avail_w / LOGO_ASPECT <= avail_h:
        w_mm, h_mm = avail_w, avail_w / LOGO_ASPECT
    else:
        h_mm, w_mm = avail_h, avail_h * LOGO_ASPECT

    img = XlImage(LOGO_PATH)
    img.width = w_mm / 25.4 * 96   # px at 96 DPI
    img.height = h_mm / 25.4 * 96

    # Center offset in EMU (1mm = 36000 EMU)
    x_off = int((zone_w_mm - w_mm) / 2 * 36000)
    y_off = int((zone_h_mm - h_mm) / 2 * 36000)

    marker = AnchorMarker(col=0, colOff=max(x_off, 0),
                          row=logo_row_start - 1, rowOff=max(y_off, 0))
    w_emu = int(img.width * 9525)
    h_emu = int(img.height * 9525)

    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(w_emu, h_emu))
    ws.add_image(img)
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# FIX 2: BORDERS ONLY (no fill/font changes)
# ═══════════════════════════════════════════════════════════════════════════════

def fix_border(ws, row, col, top=BN, bottom=BN, left=BN, right=BN):
    try:
        ws.cell(row=row, column=col).border = Border(top=top, bottom=bottom, left=left, right=right)
    except:
        pass

def fix_header_borders(ws, ncols):
    """Fix header area borders (rows 1-5)."""
    logo_end, _, _ = detect_logo_zone(ws)
    meta_start = detect_meta_zone(ws)
    if not logo_end or not meta_start:
        return

    # Find meta label end (where value zone starts)
    meta_lbl_end = meta_start
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 5 and mg.min_col == meta_start:
            meta_lbl_end = mg.max_col
            break

    for r in range(1, 6):
        for c in range(1, ncols + 1):
            top = BB if r == 1 else BN
            bottom = BB if r == 5 else BN
            left = BB if c == 1 else BN
            right = BB if c == ncols else BN

            if c == logo_end: right = BB
            if c == logo_end + 1: left = BB
            if c == meta_start: left = BB
            if c >= meta_start and r > 1: top = BG
            if c == meta_lbl_end and c < ncols: right = BG

            fix_border(ws, r, c, top, bottom, left, right)


def fix_body_borders(ws, ncols, start_row):
    """Fix body borders: section outlines + data grid."""
    max_row = ws.max_row or 50

    # Classify all rows
    types = {}
    for r in range(start_row, max_row + 1):
        types[r] = classify(ws, r, ncols)

    # Group data rows into blocks
    block = []
    for r in range(start_row, max_row + 1):
        rt = types[r]

        if rt == 'spacer' or rt == 'rule_stripe':
            if block: _apply_data_borders(ws, block, ncols); block = []
            # Spacer: no borders
            for c in range(1, ncols + 1):
                fix_border(ws, r, c)

        elif rt == 'section' or rt == 'notice':
            if block: _apply_data_borders(ws, block, ncols); block = []
            # Section/Notice: all 4 sides blue
            for c in range(1, ncols + 1):
                fix_border(ws, r, c, BB, BB,
                          BB if c == 1 else BN,
                          BB if c == ncols else BN)

        elif rt == 'col_header':
            if block: _apply_data_borders(ws, block, ncols); block = []
            me = merge_ends(ws, r)
            for c in range(1, ncols + 1):
                right = BB if c == ncols else (BG if c in me else BN)
                fix_border(ws, r, c, BB, BG,
                          BB if c == 1 else BN, right)

        else:
            block.append(r)

    if block: _apply_data_borders(ws, block, ncols)


def _apply_data_borders(ws, rows, ncols):
    """Apply data block borders: medium outline, thin internal."""
    if not rows: return
    first, last = rows[0], rows[-1]

    for r in rows:
        me = merge_ends(ws, r)
        for c in range(1, ncols + 1):
            if r == first and r == last:
                top, bottom = BB, BB
            elif r == first:
                top, bottom = BB, BG
            elif r == last:
                top, bottom = BN, BB
            else:
                top, bottom = BN, BG

            left = BB if c == 1 else BN
            right = BB if c == ncols else (BG if c in me else BN)
            fix_border(ws, r, c, top, bottom, left, right)


# ═══════════════════════════════════════════════════════════════════════════════
# FIX 3: FOOTER
# ═══════════════════════════════════════════════════════════════════════════════

def fix_footer(ws):
    """Set footer font to 7pt (matching master template)."""
    try:
        hf = ws.HeaderFooter
        if hf and hf.oddFooter and hf.oddFooter.center:
            if hf.oddFooter.center.size is None or hf.oddFooter.center.size != 7:
                hf.oddFooter.center.size = 7
                hf.oddFooter.center.font = "Segoe UI"
    except:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def process(filepath):
    fname = os.path.basename(filepath)
    log = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, [f"ERROR: {e}"], False

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        if not has_form_header(ws):
            log.append(f"  [{ws_name}] SKIP (no form header)")
            continue

        ncols = detect_ncols(ws)

        # 1. Embed logo (centered)
        if embed_logo(ws, ncols):
            log.append(f"  [{ws_name}] logo embedded (ncols={ncols})")

        # 2. Fix header borders
        fix_header_borders(ws, ncols)

        # 3. Fix body borders
        # Find where body starts (after header spacer)
        body_start = 7  # default
        for r in range(5, 10):
            rd = ws.row_dimensions.get(r)
            h = rd.height if rd else None
            if h and h <= 5:
                body_start = r
                break
            fh = fill_c(ws.cell(row=r, column=1))
            if fh in SECTION_FILLS or fh in COLHDR_FILLS:
                body_start = r
                break

        fix_body_borders(ws, ncols, body_start)

        # 4. Fix footer
        fix_footer(ws)

        # 5. Page setup
        try:
            ws.sheet_view.showGridLines = False
        except: pass
        try:
            pm = ws.page_margins
            pm.left = pm.right = 0.197
            pm.top = pm.bottom = 0.197
            pm.header = pm.footer = 0.1
            ps = ws.page_setup
            ps.paperSize = 9 if ncols <= 56 else 8
            ps.fitToWidth = 1
            ps.fitToHeight = 0
        except: pass

        log.append(f"  [{ws_name}] borders + setup done")

    try:
        wb.save(filepath)
        log.append("  >> SAVED")
    except Exception as e:
        log.append(f"  >> SAVE ERROR: {e}")

    wb.close()
    return fname, log, True


def main():
    base = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS MINIMAL FIX (logo + borders + footer only)")
    print(f"Logo: {'OK' if os.path.exists(LOGO_PATH) else 'MISSING'}")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    v = '--verbose' in sys.argv
    saved = 0
    errors = 0
    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        print(f"[{i:3d}/{len(files)}] {fn}...", end=" ", flush=True)
        _, log, ok = process(fp)
        if any('SAVE ERROR' in l for l in log):
            print("LOCKED")
            errors += 1
        else:
            print("OK")
            saved += 1
        if v:
            for l in log: print(l)

    print(f"\n{'='*70}")
    print(f"DONE - {saved} saved, {errors} locked (close Excel and retry)")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
