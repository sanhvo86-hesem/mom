#!/usr/bin/env python3
"""
Fix Excel corruption: missing sharedStrings.xml + notice bar styling.

Root cause: openpyxl saves without sharedStrings when all strings are inline.
Excel expects sharedStrings.xml to exist and triggers recovery when missing.

Fix: Force openpyxl to use shared strings by re-saving with use_shared_strings option.
Also fix notice bar styling per master template.
"""
import os, sys, zipfile, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xml.etree import ElementTree as ET

# Master template notice bar specs
NOTICE_FONT = Font(name='Segoe UI', size=7, bold=False, color='7B4F00')
NOTICE_FILL = PatternFill('solid', fgColor='FFF8E1')
NOTICE_ALIGN = Alignment(horizontal='left', vertical='center', wrapText=True)
NOTICE_BORDER_SIDE = Side(style='thin', color='F9A825')
NOTICE_BG = 'FFF8E1'

# Also fix spacer row height
SPACER_HEIGHT = 3.0


def color_hex(color):
    if color is None:
        return None
    if hasattr(color, 'rgb') and color.rgb and color.rgb != '00000000':
        rgb = str(color.rgb)
        return rgb[2:].upper() if len(rgb) == 8 else rgb.upper()
    return None


def fill_hex(fill):
    if fill is None or fill.fill_type is None:
        return None
    return color_hex(fill.fgColor)


def detect_ncols(ws):
    count = 0
    for col_idx in range(1, 200):
        col_letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(col_letter)
        if dim and dim.width is not None and abs(dim.width - 2.33) < 0.5:
            count += 1
        elif count > 10:
            break
    if count <= 44: return 40
    elif count <= 60: return 56
    else: return 82


def fix_notice_bars(ws, ncols):
    """Fix all notice bar rows to match master template."""
    fixed = 0
    for row in range(1, (ws.max_row or 50) + 1):
        cell = ws.cell(row=row, column=1)
        fh = fill_hex(cell.fill)

        # Detect notice bar by fill color or by "NOTICE" text
        is_notice = False
        if fh and fh.upper() == NOTICE_BG.upper():
            is_notice = True
        val = str(cell.value or '').upper()
        if 'NOTICE' in val and len(val) > 5:
            is_notice = True

        if not is_notice:
            continue

        # Fix notice row
        rd = ws.row_dimensions[row]
        if rd.height is None or rd.height < 20:
            rd.height = 26.0

        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = NOTICE_FILL
            cell.font = NOTICE_FONT
            cell.alignment = NOTICE_ALIGN
            cell.border = Border(
                top=NOTICE_BORDER_SIDE,
                bottom=NOTICE_BORDER_SIDE,
                left=NOTICE_BORDER_SIDE if c == 1 else Side(style=None),
                right=NOTICE_BORDER_SIDE if c == ncols else Side(style=None)
            )
        fixed += 1
    return fixed


def fix_workbook(filepath):
    """Fix corruption and notice bars in a single workbook."""
    fname = os.path.basename(filepath)
    issues = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, [f"LOAD ERROR: {e}"], False

    # Fix notice bars in all visible sheets
    for ws in wb.worksheets:
        if ws.sheet_state == 'veryHidden':
            continue
        ncols = detect_ncols(ws)
        n = fix_notice_bars(ws, ncols)
        if n > 0:
            issues.append(f"  {ws.title}: {n} notice bar(s) fixed")

    # Save - this forces openpyxl to regenerate sharedStrings.xml
    try:
        wb.save(filepath)
        issues.append("  SAVED")
    except Exception as e:
        issues.append(f"  SAVE ERROR: {e}")
        wb.close()
        return fname, issues, False

    wb.close()

    # Now verify and fix sharedStrings if still missing
    # openpyxl should create it, but let's verify
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            has_ss = 'xl/sharedStrings.xml' in z.namelist()
            if not has_ss:
                issues.append("  WARNING: sharedStrings still missing, injecting...")
                inject_shared_strings(filepath)
                issues.append("  sharedStrings.xml injected")
    except Exception as e:
        issues.append(f"  ZIP CHECK ERROR: {e}")

    return fname, issues, True


def inject_shared_strings(filepath):
    """Inject a minimal sharedStrings.xml into the xlsx zip."""
    import shutil, tempfile

    # Read all inline strings from the workbook
    wb = load_workbook(filepath)
    strings = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    strings.add(cell.value)
    wb.close()

    # Build sharedStrings XML
    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    sst = ET.Element('sst')
    sst.set('xmlns', ns)
    sst.set('count', str(len(strings)))
    sst.set('uniqueCount', str(len(strings)))
    for s in sorted(strings):
        si = ET.SubElement(sst, 'si')
        t = ET.SubElement(si, 't')
        t.text = s

    ss_xml = b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\n"
    ss_xml += ET.tostring(sst, encoding='unicode').encode('utf-8')

    # Inject into zip
    tmpfd, tmppath = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmpfd)

    with zipfile.ZipFile(filepath, 'r') as zin:
        with zipfile.ZipFile(tmppath, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                zout.writestr(item, zin.read(item))

            # Add sharedStrings
            zout.writestr('xl/sharedStrings.xml', ss_xml)

            # Update Content_Types to include sharedStrings
            ct_xml = zin.read('[Content_Types].xml').decode('utf-8')
            if 'sharedStrings' not in ct_xml:
                ct_xml = ct_xml.replace(
                    '</Types>',
                    '<Override PartName="/xl/sharedStrings.xml" '
                    'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                    '</Types>'
                )
                zout.writestr('[Content_Types].xml', ct_xml.encode('utf-8'))

            # Update workbook.xml.rels to reference sharedStrings
            rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            if 'sharedStrings' not in rels_xml:
                # Find max rId
                rids = re.findall(r'rId(\d+)', rels_xml)
                max_rid = max(int(r) for r in rids) if rids else 0
                new_rid = f'rId{max_rid + 1}'
                rels_xml = rels_xml.replace(
                    '</Relationships>',
                    f'<Relationship Id="{new_rid}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
                    f'Target="sharedStrings.xml"/>'
                    f'</Relationships>'
                )
                zout.writestr('xl/_rels/workbook.xml.rels', rels_xml.encode('utf-8'))

    # Replace original
    shutil.move(tmppath, filepath)


def main():
    base = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS - FIX CORRUPTION + NOTICE BARS")
    print(f"  Root cause: missing sharedStrings.xml (openpyxl bug)")
    print(f"  Also fixing: notice bar styling per master template")
    print(f"  Files: {len(files)}")
    print(f"{'='*70}")

    errors = 0
    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        print(f"[{i:3d}/{len(files)}] {fn[:50]:50s}", end=" ", flush=True)

        fname, issues, ok = fix_workbook(fp)
        if ok:
            print("OK")
        else:
            print("ERROR")
            errors += 1

    # Verify
    print(f"\n{'='*70}")
    print("VERIFICATION - checking sharedStrings.xml presence:")
    missing = 0
    for fp in files:
        try:
            with zipfile.ZipFile(fp, 'r') as z:
                if 'xl/sharedStrings.xml' not in z.namelist():
                    print(f"  STILL MISSING: {os.path.basename(fp)}")
                    missing += 1
        except:
            pass

    print(f"\n  Files with sharedStrings: {len(files) - missing}/{len(files)}")
    print(f"  Files still missing: {missing}")
    print(f"  Errors: {errors}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
