import openpyxl, json, sys

def analyze_form(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    result = {"file": filepath, "sheets": []}
    for ws in wb.worksheets:
        sheet_data = {"name": ws.title, "dimensions": ws.dimensions,
            "merged_cells": [str(m) for m in ws.merged_cells.ranges],
            "data_validations": [], "rows": []}
        for dv in ws.data_validations.dataValidation:
            sheet_data["data_validations"].append({"type": dv.type, "formula1": str(dv.formula1) if dv.formula1 else None, "cells": str(dv.sqref)})
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 100), values_only=False):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append({"cell": cell.coordinate, "value": str(cell.value), "is_formula": str(cell.value).startswith("=") if isinstance(cell.value, str) else False})
            if row_data:
                sheet_data["rows"].append(row_data)
        result["sheets"].append(sheet_data)
    print(json.dumps(result, indent=2, ensure_ascii=False))

if __name__ == "__main__":
    analyze_form(sys.argv[1])
