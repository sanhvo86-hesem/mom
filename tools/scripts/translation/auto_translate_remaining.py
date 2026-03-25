#!/usr/bin/env python3
"""
Auto-fill Vietnamese translations for words marked Yes but missing translation.
For ambiguous/difficult words, mark yellow for user review.
"""
import openpyxl
from openpyxl.styles import PatternFill

# Comprehensive auto-translation dictionary for QMS/CNC manufacturing context
AUTO = {
    # ── Common verbs ──
    'know': 'biết',
    'leave': 'rời / nghỉ phép',
    'accept': 'chấp nhận',
    'accepted': 'đã chấp nhận',
    'adjust': 'điều chỉnh',
    'allowed': 'cho phép',
    'apply': 'áp dụng',
    'book': 'sổ / đặt',
    'carry': 'mang / chuyển',
    'challenge': 'thách thức',
    'collect': 'thu thập',
    'combine': 'kết hợp',
    'communicate': 'truyền đạt',
    'complete': 'hoàn thành',
    'confirm': 'xác nhận',
    'connect': 'kết nối',
    'convert': 'chuyển đổi',
    'coordinate': 'phối hợp',
    'correct': 'sửa / đúng',
    'define': 'xác định',
    'deliver': 'giao hàng',
    'demonstrate': 'chứng minh / trình diễn',
    'deploy': 'triển khai',
    'designate': 'chỉ định',
    'designated': 'được chỉ định',
    'detect': 'phát hiện',
    'develop': 'phát triển',
    'direct': 'trực tiếp / chỉ đạo',
    'distribute': 'phân phối',
    'distribution': 'phân phối',
    'enable': 'kích hoạt',
    'ensure': 'đảm bảo',
    'establish': 'thiết lập',
    'evaluate': 'đánh giá',
    'execute': 'thực thi',
    'extend': 'mở rộng',
    'follow': 'theo dõi / tuân theo',
    'freeze': 'đóng băng',
    'frozen': 'đã đóng băng',
    'handle': 'xử lý',
    'identify': 'nhận diện',
    'implement': 'triển khai',
    'implemented': 'đã triển khai',
    'improve': 'cải tiến',
    'inform': 'thông báo',
    'initiate': 'khởi tạo',
    'inspect': 'kiểm tra',
    'install': 'lắp đặt',
    'investigate': 'điều tra',
    'join': 'tham gia / nối',
    'maintain': 'duy trì / bảo trì',
    'manage': 'quản lý',
    'modify': 'sửa đổi',
    'monitor': 'giám sát',
    'notify': 'thông báo',
    'observe': 'quan sát',
    'observation': 'quan sát',
    'operate': 'vận hành',
    'participate': 'tham gia',
    'perform': 'thực hiện',
    'plan': 'lập kế hoạch',
    'predict': 'dự đoán',
    'prepare': 'chuẩn bị',
    'present': 'trình bày',
    'prevent': 'phòng ngừa',
    'prioritize': 'ưu tiên hóa',
    'process': 'xử lý / quy trình',
    'produce': 'sản xuất',
    'protect': 'bảo vệ',
    'provide': 'cung cấp',
    'pull': 'kéo',
    'push': 'đẩy',
    'read': 'đọc',
    'receive': 'nhận',
    'recommend': 'khuyến nghị',
    'record': 'ghi nhận',
    'redirect': 'chuyển hướng',
    'reduce': 'giảm',
    'refer': 'tham chiếu',
    'reject': 'từ chối',
    'release': 'phát hành',
    'remove': 'loại bỏ',
    'replace': 'thay thế',
    'report': 'báo cáo',
    'request': 'yêu cầu',
    'resolve': 'giải quyết',
    'respond': 'phản hồi',
    'restore': 'khôi phục',
    'retain': 'lưu giữ',
    'retained': 'đã lưu giữ',
    'retrieve': 'truy xuất',
    'review': 'rà soát',
    'revise': 'sửa đổi',
    'schedule': 'lập lịch',
    'schedules': 'lịch trình',
    'scheduling': 'lập lịch',
    'secure': 'bảo mật',
    'select': 'chọn',
    'send': 'gửi',
    'sign': 'ký',
    'specify': 'quy định',
    'start': 'bắt đầu',
    'startup': 'khởi động',
    'store': 'lưu trữ',
    'submit': 'nộp / gửi',
    'suggest': 'đề xuất',
    'support': 'hỗ trợ',
    'suspend': 'tạm ngưng',
    'test': 'kiểm nghiệm',
    'trace': 'truy vết',
    'track': 'theo dõi',
    'train': 'đào tạo',
    'transfer': 'chuyển giao',
    'upgrade': 'nâng cấp',
    'validate': 'xác nhận hiệu lực',
    'verify': 'xác minh',
    'visit': 'tham quan / kiểm tra',
    'wait': 'chờ',
    'waiting': 'chờ',
    'walk': 'đi kiểm tra',
    'withdraw': 'rút / thu hồi',
    'write': 'viết / ghi',
    'writing': 'viết',

    # ── Past tense / -ed forms ──
    'suspected': 'nghi ngờ',
    'disputed': 'tranh chấp',
    'promised': 'đã cam kết',
    'annealed': 'đã ủ (annealed)',
    'blocked': 'bị chặn',
    'reviewed': 'đã rà soát',
    'shipped': 'đã giao hàng',
    'unapproved': 'chưa phê duyệt',

    # ── Plural / -s forms ──
    'relations': 'quan hệ',
    'cases': 'trường hợp',
    'instructions': 'hướng dẫn',
    'packages': 'gói hàng',
    'patterns': 'mẫu / quy luật',
    'blockers': 'vấn đề chặn',
    'threads': 'ren',
    'columns': 'cột',
    'gaps': 'khoảng trống / thiếu sót',
    'boards': 'bảng',
    'errors': 'lỗi',
    'attributes': 'thuộc tính',
    'mismatches': 'không khớp',
    'payments': 'thanh toán',
    'utilities': 'tiện ích',
    'assets': 'tài sản',
    'features': 'tính năng / đặc điểm',
    'services': 'dịch vụ',
    'heads': 'trưởng bộ phận',
    'contacts': 'liên hệ',
    'days': 'ngày',
    'keys': 'khóa / phím',

    # ── Nouns - Business ──
    'absence': 'vắng mặt / nghỉ phép',
    'capex': 'chi phí đầu tư (CAPEX)',
    'cashflow': 'dòng tiền',
    'conversion': 'chuyển đổi',
    'correspondence': 'thư từ / giao dịch',
    'coupon': 'phiếu / mẫu thử (coupon)',
    'demand': 'nhu cầu',
    'disposal': 'thanh lý / xử lý',
    'inheritance': 'kế thừa',
    'penalty': 'phạt',
    'proposal': 'đề xuất',
    'quoting': 'báo giá',
    'rating': 'xếp hạng / đánh giá',
    'reasoning': 'lý luận',
    'recon': 'đối soát',
    'stewardship': 'quản lý tài sản',
    'survey': 'khảo sát',
    'timing': 'thời điểm',
    'trust': 'tin cậy',
    'annual': 'hàng năm',
    'industry': 'ngành công nghiệp',
    'recent': 'gần đây',
    'historical': 'lịch sử',
    'destination': 'điểm đến',
    'applicability': 'phạm vi áp dụng',
    'worksheet': 'bảng tính / phiếu công việc',

    # ── Nouns - Manufacturing / CNC ──
    'conveyor': 'băng tải',
    'corrosion': 'ăn mòn',
    'fitting': 'phụ kiện nối ống',
    'furnace': 'lò nung',
    'marker': 'bút đánh dấu',
    'plane': 'mặt phẳng',
    'printer': 'máy in',
    'scanner': 'máy quét',
    'spike': 'đột biến / tăng đột ngột',
    'spill': 'tràn / đổ',
    'stamp': 'tem / dấu',
    'stress': 'ứng suất',
    'tube': 'ống',
    'wafer': 'tấm bán dẫn (wafer)',
    'weld': 'hàn',
    'wire': 'dây',
    'port': 'cổng',
    'mock': 'mô phỏng / giả lập',
    'markup': 'đánh dấu / ghi chú trên bản vẽ',
    'filename': 'tên tập tin',
    'regrind': 'mài lại',
    'reinspection': 'kiểm tra lại',
    'recheck': 'kiểm tra lại',
    'replan': 'lập kế hoạch lại',
    'positional': 'vị trí',
    'hand': 'tay / thủ công',
    'lift': 'nâng',
    'tactile': 'xúc giác / cảm ứng',
    'dirty': 'bẩn',
    'facing': 'tiện mặt đầu (facing)',
    'tiering': 'phân tầng',

    # ── Nouns - Quality ──
    'formal': 'chính thức',
    'paperwork': 'giấy tờ / hồ sơ',
    'guardrail': 'rào chắn / quy định bảo vệ',
    'pain': 'vấn đề khó khăn',
    'trap': 'bẫy / lỗi tiềm ẩn',
    'opening': 'mở / lỗ mở',
    'downgrade': 'hạ cấp',
    'malware': 'phần mềm độc hại',
    'falsification': 'giả mạo',

    # ── IT / System ──
    'antivirus': 'chống vi-rút',
    'cyber': 'an ninh mạng',
    'media': 'phương tiện',
    'object': 'đối tượng',
    'static': 'tĩnh',
    'deck': 'bộ trình chiếu',
    'watchlist': 'danh sách theo dõi',
    'blind': 'mù / ẩn',
    'requester': 'người yêu cầu',
    'semantic': 'ngữ nghĩa',
    'audience': 'đối tượng',
    'keying': 'nhập liệu',
    'semi': 'bán',

    # ── Roles ──
    'buyer': 'nhân viên mua hàng',
    'sales': 'kinh doanh',
    'coordinator': 'điều phối viên',
    'front': 'tuyến đầu',

    # ── Status / Adjectives ──
    'unknown': 'không xác định',
    'poor': 'kém',
    'outside': 'bên ngoài',
    'general': 'chung / tổng quát',
    'primary': 'chính / sơ cấp',
    'public': 'công khai',
    'reading': 'đọc / chỉ số đo',

    # ── Proper names / Abbreviations - keep ──
    'PPAP': 'PPAP',
    'REST': 'REST',
    'SWOT': 'SWOT',
    'HMLV': 'HMLV',
    'CAPAID': 'CAPAID',
    'NCRDefectType': 'NCRDefectType',
    'NCRDisposition': 'NCRDisposition',
    'Hastelloy': 'Hastelloy',
    'Ceramic': 'Gốm sứ (Ceramic)',
    'Entra': 'Entra',
    'Golden': 'Vàng / Chuẩn (Golden)',
    'Ethics': 'Đạo đức',
    'Surveillance': 'Giám sát',
    'Consulted': 'Được tham vấn',
    'Periodic': 'Định kỳ',
    'Hypercare': 'Hỗ trợ tăng cường (Hypercare)',
    'Execute': 'Thực thi',
    'Upgrade': 'Nâng cấp',
    'Waiting': 'Đang chờ',
    'SALES': 'KINH DOANH',
    'THPT': 'THPT',
    'Front': 'Tuyến đầu',
    'Survey': 'Khảo sát',
    'Public': 'Công khai',

    # ── More common words (1-4 occurrences) ──
    'additional': 'bổ sung',
    'adequate': 'đầy đủ / thỏa đáng',
    'alternative': 'thay thế',
    'appropriate': 'phù hợp',
    'available': 'khả dụng',
    'aware': 'nhận thức',
    'capable': 'có khả năng',
    'certain': 'chắc chắn',
    'clear': 'rõ ràng',
    'consistent': 'nhất quán',
    'continuous': 'liên tục',
    'critical': 'quan trọng',
    'current': 'hiện tại',
    'effective': 'hiệu quả',
    'efficient': 'hiệu quả',
    'essential': 'thiết yếu',
    'existing': 'hiện có',
    'expected': 'dự kiến',
    'final': 'cuối cùng',
    'global': 'toàn cầu / toàn bộ',
    'immediate': 'ngay lập tức',
    'important': 'quan trọng',
    'initial': 'ban đầu',
    'key': 'chính / then chốt',
    'known': 'đã biết',
    'local': 'nội bộ / tại chỗ',
    'main': 'chính',
    'major': 'lớn / nghiêm trọng',
    'mandatory': 'bắt buộc',
    'maximum': 'tối đa',
    'minimum': 'tối thiểu',
    'minor': 'nhỏ',
    'necessary': 'cần thiết',
    'new': 'mới',
    'normal': 'bình thường',
    'original': 'gốc / ban đầu',
    'overall': 'tổng thể',
    'permanent': 'vĩnh viễn',
    'planned': 'đã lên kế hoạch',
    'potential': 'tiềm ẩn',
    'proper': 'đúng cách',
    'relevant': 'liên quan',
    'remaining': 'còn lại',
    'required': 'bắt buộc',
    'responsible': 'chịu trách nhiệm',
    'safe': 'an toàn',
    'significant': 'đáng kể',
    'similar': 'tương tự',
    'specific': 'cụ thể',
    'suitable': 'phù hợp',
    'temporary': 'tạm thời',
    'total': 'tổng',
    'typical': 'điển hình',
    'unique': 'duy nhất',
    'valid': 'hợp lệ',
    'various': 'nhiều / khác nhau',
    'visible': 'nhìn thấy được',
    'weekly': 'hàng tuần',
    'monthly': 'hàng tháng',
    'quarterly': 'hàng quý',
    'daily': 'hàng ngày',

    # ── More nouns ──
    'ability': 'khả năng',
    'approach': 'cách tiếp cận',
    'aspect': 'khía cạnh',
    'background': 'nền / bối cảnh',
    'basis': 'cơ sở',
    'benefit': 'lợi ích',
    'category': 'danh mục',
    'circumstance': 'hoàn cảnh',
    'complexity': 'độ phức tạp',
    'component': 'linh kiện / bộ phận',
    'consequence': 'hậu quả',
    'consideration': 'cân nhắc',
    'context': 'bối cảnh',
    'detail': 'chi tiết',
    'element': 'yếu tố',
    'emphasis': 'nhấn mạnh',
    'example': 'ví dụ',
    'extent': 'mức độ',
    'focus': 'trọng tâm',
    'format': 'định dạng',
    'foundation': 'nền tảng',
    'frequency': 'tần suất',
    'importance': 'tầm quan trọng',
    'instance': 'trường hợp',
    'intent': 'mục đích',
    'issue': 'vấn đề',
    'issues': 'vấn đề',
    'knowledge': 'kiến thức',
    'lesson': 'bài học',
    'method': 'phương pháp',
    'nature': 'bản chất',
    'outcome': 'kết quả',
    'overview': 'tổng quan',
    'owner': 'chủ sở hữu',
    'period': 'giai đoạn',
    'perspective': 'góc nhìn',
    'principle': 'nguyên tắc',
    'purpose': 'mục đích',
    'reference': 'tham chiếu',
    'resource': 'nguồn lực',
    'response': 'phản hồi',
    'result': 'kết quả',
    'role': 'vai trò',
    'section': 'phần',
    'situation': 'tình huống',
    'source': 'nguồn',
    'stage': 'giai đoạn',
    'structure': 'cấu trúc',
    'summary': 'tóm tắt',
    'system': 'hệ thống',
    'task': 'nhiệm vụ',
    'technique': 'kỹ thuật',
    'topic': 'chủ đề',
    'type': 'loại',
    'value': 'giá trị',
    'volume': 'khối lượng',
}

# Words that are too ambiguous to auto-translate
DIFFICULT = {
    'chai', 'cali', 'upsert', 'thru', 'case', 'CUSTOM', 'calc', 'comp',
    'eofe', 'SSee', 'Trao', 'khoang', 'linh', 'xanh', 'xoay', 'trao',
    'chen', 'ninh', 'dang', 'ngang', 'Ranh', 'Kinh', 'Trung', 'Bieu',
    'Khoanh', 'xung',
}


def main():
    wb = openpyxl.load_workbook('tools/remaining-english-words.xlsx')
    ws = wb.active
    yellow_fill = PatternFill('solid', fgColor='FFFF00')
    yes_fill = PatternFill('solid', fgColor='C6EFCE')

    filled = 0
    yellowed = 0

    for row_idx in range(2, ws.max_row + 1):
        word = ws.cell(row=row_idx, column=2).value
        vi = ws.cell(row=row_idx, column=4).value
        yn = ws.cell(row=row_idx, column=5).value

        if not word:
            continue

        word_str = str(word).strip()
        yn_str = str(yn).strip().lower() if yn else ''

        # Only process Yes rows without Vietnamese
        if yn_str == 'yes' and not vi:
            # Check auto-translate
            translation = AUTO.get(word_str, AUTO.get(word_str.lower(), ''))

            if word_str in DIFFICULT:
                ws.cell(row=row_idx, column=4).fill = yellow_fill
                ws.cell(row=row_idx, column=5).fill = yellow_fill
                ws.cell(row=row_idx, column=6).value = 'Cần anh quyết định'
                yellowed += 1
            elif translation:
                ws.cell(row=row_idx, column=4).value = translation
                ws.cell(row=row_idx, column=4).fill = yes_fill
                filled += 1
            else:
                # Try lowercase lookup
                lower = word_str.lower()
                if lower in AUTO:
                    ws.cell(row=row_idx, column=4).value = AUTO[lower]
                    ws.cell(row=row_idx, column=4).fill = yes_fill
                    filled += 1
                else:
                    # Mark yellow - difficult
                    ws.cell(row=row_idx, column=4).fill = yellow_fill
                    ws.cell(row=row_idx, column=5).fill = yellow_fill
                    ws.cell(row=row_idx, column=6).value = 'Cần anh quyết định'
                    yellowed += 1

    wb.save('tools/remaining-english-words.xlsx')
    print(f'Done!')
    print(f'  Auto-filled: {filled}')
    print(f'  Marked yellow (khó): {yellowed}')
    print(f'  Remaining untouched: {2621 - filled - yellowed}')


if __name__ == '__main__':
    main()
