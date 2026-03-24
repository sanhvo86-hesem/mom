#!/usr/bin/env python3
"""
Translate English QMS terms to Vietnamese across all HTML files.
Reads dictionary from tools/qms-terminology-dictionary.xlsx.
Only translates text nodes (not HTML attributes, CSS, JS, file paths).
"""
import re
import os
import sys
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DICT_PATH = os.path.join(BASE_DIR, 'tools', 'qms-terminology-dictionary.xlsx')
SKIP_DIRS = {'.git', '_build', '.claude', '_Deleted', 'node_modules', '.vscode'}

# ── Load dictionary from Excel ──────────────────────────────────────────────
def load_dictionary():
    wb = openpyxl.load_workbook(DICT_PATH)
    entries = []  # (english, vietnamese, is_phrase)
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            _stt, en, vi, yn = row[0], row[1], row[2], row[3]
            if not en or not vi or not yn:
                continue
            if str(yn).strip().lower() != 'yes':
                continue
            en = str(en).strip()
            vi = str(vi).strip()
            if en == vi:
                continue
            is_phrase = ' ' in en or '-' in en
            entries.append((en, vi, is_phrase))
    wb.close()
    return entries

# ── Build ordered replacement list ──────────────────────────────────────────
def build_replacements(entries):
    """Sort: longest first, then phrases before single words."""
    # Deduplicate (keep first occurrence)
    seen = set()
    unique = []
    for en, vi, is_phrase in entries:
        key = en.lower()
        if key not in seen:
            seen.add(key)
            unique.append((en, vi, is_phrase))

    # Sort by length descending (longest match first)
    unique.sort(key=lambda x: -len(x[0]))

    replacements = []
    for en, vi, is_phrase in unique:
        # Build regex with word boundaries
        # For metadata labels ending with ":", match exactly
        if en.endswith(':'):
            pattern = re.compile(re.escape(en))
        else:
            pattern = re.compile(r'\b' + re.escape(en) + r'\b')
        replacements.append((pattern, vi, en))

        # Also add capitalized/lowercase variants if the entry is lowercase
        if en[0].islower():
            cap_en = en[0].upper() + en[1:]
            cap_vi = vi[0].upper() + vi[1:]
            if cap_en.endswith(':'):
                cap_pattern = re.compile(re.escape(cap_en))
            else:
                cap_pattern = re.compile(r'\b' + re.escape(cap_en) + r'\b')
            replacements.append((cap_pattern, cap_vi, cap_en))
        # If entry starts with uppercase, also add lowercase variant
        elif en[0].isupper() and not en.isupper():
            low_en = en[0].lower() + en[1:]
            low_vi = vi[0].lower() + vi[1:]
            if low_en.endswith(':'):
                low_pattern = re.compile(re.escape(low_en))
            else:
                low_pattern = re.compile(r'\b' + re.escape(low_en) + r'\b')
            replacements.append((low_pattern, low_vi, low_en))

    # Add plural forms for common single-word terms
    plural_map = {}
    for pattern, vi, en in replacements:
        if ' ' not in en and '-' not in en and not en.endswith(':') and not en.endswith('s'):
            plural_en = en + 's'
            if plural_en.lower() not in seen:
                plural_map[plural_en] = vi  # Vietnamese doesn't pluralize

    for en, vi in plural_map.items():
        p = re.compile(r'\b' + re.escape(en) + r'\b')
        replacements.append((p, vi, en))

    # Re-sort by pattern length descending after adding variants
    replacements.sort(key=lambda x: -len(x[2]))
    return replacements

# ── HTML text-node translation engine ───────────────────────────────────────
SKIP_BLOCKS_RE = re.compile(
    r'(<style[\s>].*?</style>|<script[\s>].*?</script>)',
    re.DOTALL | re.IGNORECASE
)
TAG_RE = re.compile(r'(<[^>]*>)')

def translate_html(html, replacements):
    """Translate only text nodes in HTML, preserving tags/attributes/CSS/JS."""
    # Protect style/script blocks
    protected = {}
    counter = [0]
    def protect(m):
        key = f'\x00P{counter[0]}\x00'
        protected[key] = m.group(0)
        counter[0] += 1
        return key
    html = SKIP_BLOCKS_RE.sub(protect, html)

    # Split into tags and text nodes
    parts = TAG_RE.split(html)

    changes = 0
    for i in range(len(parts)):
        if i % 2 == 0 and parts[i].strip():  # text node
            original = parts[i]
            text = original
            for pattern, vi, _en in replacements:
                text = pattern.sub(vi, text)
            if text != original:
                parts[i] = text
                changes += 1

    result = ''.join(parts)

    # Restore protected blocks
    for key, value in protected.items():
        result = result.replace(key, value)

    return result, changes

# ── Detect remaining English words ──────────────────────────────────────────
ENGLISH_WORD_RE = re.compile(r'\b[A-Za-z]{4,}\b')
# Known words to ignore
IGNORE_WORDS = set()
def _load_ignore():
    # Vietnamese words that look English (no diacritics)
    viet = {
        'theo', 'trong', 'thay', 'giao', 'quan', 'minh', 'ngay', 'gian',
        'nhanh', 'danh', 'dung', 'truy', 'khai', 'sang', 'sinh', 'then',
        'tham', 'khao', 'trinh', 'dich', 'bang', 'cung', 'phat', 'hanh',
        'hien', 'dung', 'tieu', 'chuan', 'pham', 'nham', 'phuc', 'doan',
        'tang', 'giam', 'thuc', 'hieu', 'chua', 'chinh', 'chap', 'nhan',
        'thong', 'bao', 'canh', 'dong', 'trai', 'phai', 'tren', 'duoi',
        'giua', 'ngoai', 'trong', 'truoc', 'sau', 'dien', 'bien', 'lang',
        'nghe', 'tinh', 'toan', 'hoan', 'thanh', 'cong', 'viec', 'hang',
        'khach', 'xuat', 'nhap', 'chot', 'kiem', 'soat', 'theo', 'cach',
        'phan', 'tich', 'danh', 'muc', 'bang', 'lieu', 'dieu', 'khien',
        'huong', 'nhieu', 'dong', 'giai', 'quyet', 'dinh', 'luong',
        'tuong', 'ung', 'dung', 'trinh', 'hien', 'khoi', 'dong', 'chay',
        'chuyen', 'huong', 'bien', 'dich', 'giao', 'tiep', 'nhan', 'xong',
        'xuat', 'thuoc', 'loai', 'rieng', 'chung', 'nhom', 'dung', 'thoi',
        'nhap', 'hach', 'toan', 'doanh', 'nghiep', 'phong', 'nghi', 'dinh',
        'quyen', 'luong', 'luot', 'cuoi', 'cung', 'mien', 'dich', 'kiem',
        'nghiem', 'chung', 'nhan', 'thuc', 'hanh', 'luyen', 'tap', 'diem',
        'chan', 'dung', 'chan', 'phep', 'nuoc', 'cong', 'nhan', 'vien',
        'chuc', 'danh', 'truong', 'phong', 'giam', 'doc', 'lanh', 'dao',
        'soat', 'xuat', 'hang', 'phan', 'phoi', 'tong', 'hieu', 'qua',
        'nang', 'suat', 'chat', 'luong', 'tham', 'chieu', 'khung',
        'khoan', 'mang', 'luoi', 'diem', 'nhan', 'vong', 'quanh', 'quay',
        'chon', 'bien', 'phap', 'cach', 'thuc', 'tang', 'cuong', 'thuc',
        'dien', 'tich', 'luong', 'nghia', 'dinh', 'huong', 'dieu', 'chinh',
        'tham', 'khao', 'chan', 'doan', 'phuc', 'khac', 'lien', 'quan',
        'quang', 'chung', 'lich', 'hoach', 'tich', 'luy', 'tong', 'ket',
        'tieu', 'bang', 'dung', 'luong', 'tham', 'khao', 'lieu', 'phap',
        'phap', 'luat', 'chon', 'loc', 'chia', 'tach', 'phoi', 'ghep',
    }
    # Abbreviations & proper names
    tech = {
        'qms', 'qa', 'qc', 'ncr', 'capa', 'dcr', 'sop', 'frm', 'annex',
        'iso', 'kpi', 'raci', 'ssot', 'fai', 'fmea', 'pfmea', 'spc',
        'msa', 'ipqc', 'copq', 'rfq', 'bom', 'cnc', 'erp', 'ehs', 'hse',
        'uat', 'url', 'pdf', 'html', 'css', 'json', 'xlsx', 'csv', 'api',
        'setup', 'traveler', 'balloon', 'epicor', 'sharepoint', 'microsoft',
        'power', 'automate', 'teams', 'outlook', 'excel', 'word', 'azure',
        'hesem', 'vietnam', 'email', 'online', 'offline', 'server', 'cloud',
        'backup', 'dashboard', 'portal', 'login', 'logout', 'admin', 'site',
        'menu', 'header', 'footer', 'sidebar', 'icon', 'link', 'click',
        'font', 'bold', 'italic', 'color', 'size', 'width', 'height',
        'version', 'date', 'time', 'page', 'total', 'count', 'index',
        'table', 'cell', 'form', 'field', 'label', 'value', 'option',
        'class', 'style', 'script', 'layout', 'design', 'image', 'video',
        'chart', 'graph', 'status', 'active', 'pending', 'closed', 'draft',
        'true', 'false', 'null', 'void', 'type', 'name', 'text', 'data',
        'code', 'file', 'path', 'list', 'item', 'step', 'note', 'info',
        'sop', 'cert', 'org', 'pur', 'whs', 'mnt', 'sal', 'fin', 'ops',
        'nadcap', 'iatf', 'sync', 'scan', 'print', 'copy', 'paste',
        'sort', 'filter', 'search', 'find', 'view', 'edit', 'delete',
        'save', 'load', 'send', 'upload', 'download', 'import', 'export',
        'start', 'stop', 'pause', 'reset', 'clear', 'done', 'fail', 'pass',
        'skip', 'next', 'prev', 'back', 'home', 'main', 'test', 'demo',
        'temp', 'final', 'ready', 'control', 'system', 'module', 'section',
        'category', 'level', 'phase', 'stage', 'round', 'cycle', 'flow',
        'rule', 'policy', 'plan', 'goal', 'target', 'metric', 'score',
        'template', 'select', 'input', 'output', 'submit', 'cancel',
        'alert', 'modal', 'toast', 'badge', 'card', 'panel', 'group',
        'shift', 'lead', 'owner', 'role', 'team', 'dept', 'tier',
        'gate', 'pack', 'ship', 'part', 'tool', 'spec', 'gage',
        'clean', 'logic', 'matrix', 'sheet', 'fixture', 'machine',
        'source', 'program', 'academy', 'planner', 'supervisor',
        'engineer', 'manager', 'director', 'approved', 'responsible',
        'person', 'warehouse', 'engineering', 'finance', 'outsource',
    }
    IGNORE_WORDS.update(viet)
    IGNORE_WORDS.update(tech)
_load_ignore()

def detect_remaining_english(html):
    """Find English words in text nodes that might need translation."""
    html_clean = SKIP_BLOCKS_RE.sub('', html)
    parts = TAG_RE.split(html_clean)
    found = {}
    for i in range(len(parts)):
        if i % 2 == 0 and parts[i].strip():
            for m in ENGLISH_WORD_RE.finditer(parts[i]):
                word = m.group()
                if word.lower() not in IGNORE_WORDS and len(word) >= 4:
                    found[word] = found.get(word, 0) + 1
    return found

# ── Main ────────────────────────────────────────────────────────────────────
def main():
    dry_run = '--dry-run' in sys.argv
    detect = '--detect' in sys.argv

    print(f"Loading dictionary from {DICT_PATH}...")
    entries = load_dictionary()
    print(f"  Loaded {len(entries)} translation entries (Yes only)")

    replacements = build_replacements(entries)
    print(f"  Built {len(replacements)} replacement patterns (incl. variants)")

    # Collect HTML files
    html_files = []
    for root, dirs, files in os.walk(BASE_DIR):
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
        for f in files:
            if f.endswith('.html'):
                html_files.append(os.path.join(root, f))
    print(f"  Found {len(html_files)} HTML files")

    total_changes = 0
    total_files_changed = 0
    all_new_words = {}

    for filepath in sorted(html_files):
        rel = os.path.relpath(filepath, BASE_DIR)
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()

        new_content, changes = translate_html(content, replacements)

        if changes > 0:
            total_changes += changes
            total_files_changed += 1
            print(f"  [{changes:3d} changes] {rel}")
            if not dry_run:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(new_content)

        # Detect remaining English words (always on translated content)
        if detect:
            remaining = detect_remaining_english(new_content)
            for word, count in remaining.items():
                all_new_words[word] = all_new_words.get(word, 0) + count

    print(f"\n{'DRY RUN ' if dry_run else ''}SUMMARY:")
    print(f"  Files changed: {total_files_changed}/{len(html_files)}")
    print(f"  Total text nodes modified: {total_changes}")

    if detect and all_new_words:
        # Sort by frequency
        sorted_words = sorted(all_new_words.items(), key=lambda x: -x[1])
        print(f"\n  Remaining English words ({len(sorted_words)} unique):")
        for word, count in sorted_words[:100]:
            print(f"    {word}: {count}x")

        # Save to file for review
        detect_path = os.path.join(BASE_DIR, 'tools', 'remaining_english_words.txt')
        with open(detect_path, 'w', encoding='utf-8') as f:
            for word, count in sorted_words:
                f.write(f"{word}\t{count}\n")
        print(f"\n  Full list saved to: {detect_path}")

    return all_new_words

if __name__ == '__main__':
    main()
