#!/usr/bin/env python3
"""
Read remaining_english_words.txt and create an Excel with suggested translations.
Filter out Vietnamese words, code/technical terms, proper names, etc.
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Vietnamese words without diacritics that look English
VIETNAMESE_WORDS = {
    'kinh', 'nguy', 'trung', 'sung', 'trang', 'treo', 'nhau', 'thuc',
    'bien', 'nhat', 'luong', 'bang', 'dung', 'theo', 'trong', 'thay',
    'giao', 'quan', 'minh', 'ngay', 'gian', 'nhanh', 'danh', 'truy',
    'khai', 'sang', 'sinh', 'then', 'tham', 'khao', 'trinh', 'dich',
    'cung', 'phat', 'hanh', 'hien', 'tieu', 'chuan', 'pham', 'nham',
    'phuc', 'doan', 'tang', 'giam', 'thuc', 'hieu', 'chua', 'chinh',
    'chap', 'nhan', 'thong', 'dong', 'trai', 'phai', 'tren', 'duoi',
    'giua', 'ngoai', 'truoc', 'dien', 'lang', 'nghe', 'tinh', 'toan',
    'hoan', 'thanh', 'cong', 'viec', 'hang', 'khach', 'xuat', 'nhap',
    'chot', 'kiem', 'soat', 'cach', 'phan', 'tich', 'muc', 'lieu',
    'dieu', 'khien', 'huong', 'nhieu', 'giai', 'quyet', 'dinh',
    'tuong', 'ung', 'khoi', 'chay', 'chuyen', 'tiep', 'xong',
    'thuoc', 'loai', 'rieng', 'chung', 'nhom', 'thoi', 'hach',
    'doanh', 'nghiep', 'phong', 'nghi', 'quyen', 'luot', 'cuoi',
    'mien', 'nghiem', 'luyen', 'chan', 'phep', 'nuoc', 'vien',
    'chuc', 'truong', 'doc', 'lanh', 'pho', 'phoi', 'tong',
    'nang', 'suat', 'chat', 'chieu', 'khung', 'khoan', 'mang',
    'luoi', 'vong', 'quanh', 'quay', 'chon', 'tang', 'cuong',
    'nghia', 'phap', 'lien', 'quang', 'lich', 'hoach', 'luy',
    'nhung', 'duoc', 'khong', 'chua', 'nhac', 'vang', 'tien',
    'cang', 'rang', 'nguoi', 'them', 'dung', 'xung', 'quanh',
}

# Code/system terms to skip
CODE_TERMS = {
    'JobNum', 'SSCC', 'JOBCODE', 'NCREL', 'MaNV', 'PartNum', 'JobOper',
    'DocID', 'RevNum', 'BaseID', 'FormID', 'RecID', 'UserID', 'GUID',
    'xmlns', 'href', 'onclick', 'thead', 'tbody', 'colspan', 'rowspan',
    'HTTPS', 'HTTP', 'JSON', 'HTML', 'XLSX', 'DOCX', 'PPTX',
}

# Epicor/ERP field names (CamelCase patterns)
CAMEL_CASE_RE = re.compile(r'^[a-z]+[A-Z]')
ALL_CAPS_CODE_RE = re.compile(r'^[A-Z]{2,}[a-z]*[A-Z]')

# Suggested translations for common QMS/manufacturing words
SUGGESTED = {
    # High frequency
    'action': 'hành động',
    'change': 'thay đổi',
    'risk': 'rủi ro',
    'report': 'báo cáo',
    'check': 'kiểm tra',
    'baseline': 'mốc chuẩn',
    'packet': 'gói hồ sơ',
    'shipment': 'lô giao hàng',
    'package': 'gói hàng',
    'route': 'lộ trình gia công',
    'decision': 'quyết định',
    'heat': 'mẻ nấu (heat number)',
    'line': 'dòng / chuyền',
    'folder': 'thư mục',
    'cost': 'chi phí',
    'access': 'truy cập',
    'escape': 'bỏ qua / thoát',
    'impact': 'tác động',
    'first': 'đầu tiên',
    'dispatch': 'điều phối / xuất hàng',
    'capability': 'khả năng',
    'closure': 'kết thúc / đóng',
    'gating': 'kiểm soát cổng',
    'reaction': 'phản ứng / ứng phó',
    'wrong': 'sai',
    'mismatch': 'không khớp',
    'trigger': 'kích hoạt',
    'drift': 'trôi / sai lệch',
    'rubric': 'tiêu chí đánh giá',
    'deputy': 'phó',
    'freeze': 'đóng băng',
    'performance': 'hiệu suất',
    'life': 'tuổi thọ / vòng đời',
    'snapshot': 'ảnh chụp trạng thái',
    'governance': 'quản trị',
    'workbook': 'sổ tay',
    'request': 'yêu cầu',
    'routing': 'lộ trình gia công',
    'update': 'cập nhật',
    'safety': 'an toàn',
    'library': 'thư viện',
    'condition': 'điều kiện',
    'critical': 'quan trọng / nghiêm trọng',
    'rate': 'tỷ lệ',
    'trace': 'truy vết',
    'mapping': 'sơ đồ ánh xạ',
    'board': 'bảng',
    'downtime': 'thời gian ngừng máy',
    'skill': 'kỹ năng',
    'cleanliness': 'vệ sinh',
    'blueprint': 'bản thiết kế',
    'vacuum': 'chân không',
    'discipline': 'kỷ luật / chuyên ngành',
    'event': 'sự kiện',
    'metadata': 'siêu dữ liệu',
    'cycle': 'chu kỳ',
    'effective': 'hiệu lực',
    'aging': 'cũ / lão hóa',
    'alert': 'cảnh báo',
    'assign': 'giao / phân công',
    'backup': 'sao lưu',
    'block': 'chặn / khóa',
    'budget': 'ngân sách',
    'cancel': 'hủy',
    'chain': 'chuỗi',
    'classify': 'phân loại',
    'collect': 'thu thập',
    'combine': 'kết hợp',
    'commit': 'cam kết',
    'communicate': 'truyền đạt',
    'complete': 'hoàn thành',
    'confirm': 'xác nhận',
    'connect': 'kết nối',
    'correct': 'sửa / đúng',
    'coverage': 'phạm vi bao phủ',
    'create': 'tạo',
    'current': 'hiện tại',
    'daily': 'hàng ngày',
    'damage': 'hư hỏng',
    'defect': 'khuyết tật',
    'define': 'xác định',
    'delay': 'chậm trễ',
    'design': 'thiết kế',
    'detect': 'phát hiện',
    'develop': 'phát triển',
    'digital': 'số hóa',
    'distribute': 'phân phối',
    'enable': 'kích hoạt',
    'ensure': 'đảm bảo',
    'entry': 'mục nhập',
    'estimate': 'ước tính',
    'evaluate': 'đánh giá',
    'exception': 'ngoại lệ',
    'execute': 'thực thi',
    'extend': 'mở rộng',
    'external': 'bên ngoài',
    'failure': 'hỏng / lỗi',
    'feedback': 'phản hồi',
    'final': 'cuối cùng',
    'follow': 'theo dõi',
    'frequency': 'tần suất',
    'function': 'chức năng',
    'global': 'toàn bộ / toàn cầu',
    'grade': 'cấp / hạng',
    'guide': 'hướng dẫn',
    'handoff': 'bàn giao',
    'identify': 'nhận diện',
    'implement': 'triển khai',
    'improve': 'cải tiến',
    'incomplete': 'chưa hoàn thành',
    'inform': 'thông báo',
    'initial': 'ban đầu',
    'install': 'lắp đặt',
    'instruct': 'chỉ dẫn',
    'interface': 'giao diện',
    'internal': 'nội bộ',
    'investigate': 'điều tra',
    'issue': 'vấn đề / phát hành',
    'label': 'nhãn',
    'manage': 'quản lý',
    'manual': 'sổ tay / thủ công',
    'method': 'phương pháp',
    'milestone': 'mốc quan trọng',
    'modify': 'sửa đổi',
    'monitor': 'giám sát',
    'move': 'di chuyển',
    'notify': 'thông báo',
    'objective': 'mục tiêu',
    'occur': 'xảy ra',
    'overdue': 'quá hạn',
    'oversight': 'giám sát',
    'owner': 'chủ sở hữu',
    'permit': 'giấy phép',
    'practice': 'thực hành',
    'prevent': 'phòng ngừa',
    'priority': 'ưu tiên',
    'progress': 'tiến độ',
    'project': 'dự án',
    'protect': 'bảo vệ',
    'provide': 'cung cấp',
    'publish': 'công bố',
    'purchase': 'mua hàng',
    'range': 'phạm vi / dải',
    'receive': 'nhận',
    'recommend': 'khuyến nghị',
    'refer': 'tham chiếu',
    'reject': 'từ chối',
    'remove': 'loại bỏ',
    'repair': 'sửa chữa',
    'replace': 'thay thế',
    'resolve': 'giải quyết',
    'response': 'phản hồi',
    'restore': 'khôi phục',
    'result': 'kết quả',
    'retain': 'lưu giữ',
    'return': 'trả lại',
    'sample': 'mẫu',
    'schedule': 'lịch trình',
    'segregate': 'cách ly',
    'sequence': 'trình tự',
    'severity': 'mức nghiêm trọng',
    'sign': 'ký',
    'signature': 'chữ ký',
    'special': 'đặc biệt',
    'stack': 'chồng / xếp',
    'store': 'lưu trữ',
    'submit': 'nộp / gửi',
    'supply': 'cung ứng',
    'support': 'hỗ trợ',
    'suspend': 'tạm ngưng',
    'task': 'nhiệm vụ',
    'tolerance': 'dung sai',
    'track': 'theo dõi',
    'transfer': 'chuyển giao',
    'trend': 'xu hướng',
    'turnover': 'doanh thu / tỷ lệ nghỉ việc',
    'validate': 'xác nhận hiệu lực',
    'verify': 'xác minh',
    'volume': 'khối lượng',
    'withdraw': 'rút / thu hồi',
    'yield': 'tỷ lệ đạt',
    # Roles
    'Estimator': 'Người định giá',
    'Shipping': 'Giao vận',
    'Receiving': 'Nhận hàng',
    'Technician': 'Kỹ thuật viên',
    'Programmer': 'Lập trình viên (CNC)',
    'Competency': 'Năng lực',
    'Breakdown': 'Sự cố hỏng máy',
    'Metrology': 'Đo lường',
    'Planner': 'Người lập kế hoạch',
    'Scenario': 'Tình huống',
    'Planning': 'Kế hoạch',
    'Purchasing': 'Mua hàng',
    'Supply': 'Cung ứng',
    'Chain': 'Chuỗi',
    'Management': 'Quản lý',
    'Department': 'Phòng ban',
    'Service': 'Dịch vụ',
    'Sales': 'Kinh doanh',
    'Cash': 'Tiền mặt',
    'HOLD': 'TẠM GIỮ',
    'BLOCK': 'CHẶN',
    'Kaizen': 'Kaizen',
    'processor': 'người xử lý',
    'datum': 'mốc chuẩn (datum)',
    'yoke': 'giá đỡ / yoke',
    'family': 'nhóm / họ (part family)',
}

MORE_IGNORE = {
    # CSS/HTML property words that leak into text detection
    'left', 'right', 'bold', 'none', 'auto', 'margin', 'padding',
    'border', 'solid', 'hidden', 'visible', 'fixed', 'relative',
    'absolute', 'flex', 'grid', 'wrap', 'overflow', 'display',
    'color', 'background', 'transparent', 'inherit', 'transform',
    'transition', 'opacity', 'position', 'float', 'clear',
    # Common HTML/tech context
    'link', 'href', 'class', 'style', 'width', 'height', 'align',
    'span', 'cell', 'table', 'thead', 'tbody', 'form', 'input',
    'output', 'submit', 'button', 'option', 'select', 'textarea',
    'image', 'icon', 'font', 'size', 'page', 'index', 'click',
    'hover', 'focus', 'active', 'disabled', 'readonly', 'required',
    # Common verbs/prepositions already well-known
    'with', 'from', 'this', 'that', 'what', 'when', 'where', 'which',
    'into', 'over', 'under', 'after', 'before', 'between', 'through',
    'above', 'below', 'each', 'every', 'both', 'also', 'only', 'just',
    'more', 'most', 'less', 'than', 'very', 'much', 'such', 'like',
    'will', 'shall', 'must', 'should', 'would', 'could', 'might',
    'been', 'being', 'have', 'having', 'does', 'doing', 'make',
    'made', 'take', 'taken', 'give', 'given', 'come', 'came',
    'going', 'gone', 'keep', 'kept', 'show', 'shown', 'used',
    'using', 'work', 'need', 'want', 'know', 'known',
    'same', 'other', 'some', 'many', 'these', 'those',
    'about', 'still', 'even', 'here', 'there', 'back', 'well',
    'them', 'they', 'their', 'your', 'yours', 'were', 'said',
    # Short function words
    'else', 'case', 'true', 'false', 'null', 'void', 'type',
    'name', 'text', 'data', 'code', 'file', 'path', 'list',
    'item', 'step', 'note', 'info', 'warn', 'error', 'test',
    'demo', 'temp', 'draft', 'date', 'time', 'year', 'month',
    'week', 'hour', 'mins', 'secs', 'unit', 'mode', 'rule',
    'goal', 'plan', 'role', 'team', 'user', 'lead', 'tier',
    'gate', 'pack', 'ship', 'part', 'tool', 'spec', 'gage',
    'view', 'edit', 'save', 'load', 'send', 'move', 'copy',
    'sort', 'find', 'open', 'down', 'next', 'prev', 'home',
    'main', 'done', 'fail', 'pass', 'skip', 'stop', 'line',
    'fast', 'slow', 'long', 'full', 'half', 'high', 'last',
    'best', 'good', 'poor', 'okay', 'fine', 'safe',
}

def is_code_term(word):
    """Check if word is a code/system term."""
    if word in CODE_TERMS:
        return True
    if word.lower() in MORE_IGNORE:
        return True
    if CAMEL_CASE_RE.match(word):
        return True
    if ALL_CAPS_CODE_RE.match(word):
        return True
    if word.isupper() and len(word) <= 6:
        return True
    return False

def main():
    with open('tools/remaining_english_words.txt', 'r', encoding='utf-8') as f:
        lines = f.readlines()

    words = []
    for line in lines:
        parts = line.strip().split('\t')
        if len(parts) == 2:
            word, count = parts[0], int(parts[1])
            if count >= 20 and len(word) >= 4 and word.isascii() and word.isalpha():
                # Skip Vietnamese words
                if word.lower() in VIETNAMESE_WORDS:
                    continue
                # Skip code terms
                if is_code_term(word):
                    continue
                words.append((word, count))

    print(f"Filtered to {len(words)} English words (3+ occurrences)")

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tu tieng Anh con sot'

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)
    yes_fill = PatternFill('solid', fgColor='C6EFCE')
    no_fill = PatternFill('solid', fgColor='FFC7CE')
    maybe_fill = PatternFill('solid', fgColor='FFEB9C')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['STT', 'Từ tiếng Anh', 'Số lần xuất hiện', 'Bản dịch gợi ý', 'Dịch (Yes/No)', 'Ghi chú']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, (word, count) in enumerate(words, 2):
        suggested = SUGGESTED.get(word, SUGGESTED.get(word.lower(), ''))
        translate = 'Yes' if suggested else ''

        ws.cell(row=i, column=1, value=i-1).font = cell_font
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=i, column=1).border = thin_border

        ws.cell(row=i, column=2, value=word).font = cell_font
        ws.cell(row=i, column=2).alignment = cell_align
        ws.cell(row=i, column=2).border = thin_border

        c = ws.cell(row=i, column=3, value=count)
        c.font = cell_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

        ws.cell(row=i, column=4, value=suggested).font = cell_font
        ws.cell(row=i, column=4).alignment = cell_align
        ws.cell(row=i, column=4).border = thin_border

        c = ws.cell(row=i, column=5, value=translate)
        c.font = cell_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        if translate == 'Yes':
            c.fill = yes_fill
        elif translate == 'No':
            c.fill = no_fill
        else:
            c.fill = maybe_fill
        c.border = thin_border

        ws.cell(row=i, column=6, value='').font = cell_font
        ws.cell(row=i, column=6).alignment = cell_align
        ws.cell(row=i, column=6).border = thin_border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 40
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{ws.max_row}'

    out = 'tools/remaining-english-words.xlsx'
    wb.save(out)
    print(f"Saved: {out}")
    print(f"  Total words: {len(words)}")
    print(f"  With suggested translation: {sum(1 for w,c in words if w in SUGGESTED or w.lower() in SUGGESTED)}")
    print(f"  Need your review: {sum(1 for w,c in words if w not in SUGGESTED and w.lower() not in SUGGESTED)}")

if __name__ == '__main__':
    main()
