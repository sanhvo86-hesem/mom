#!/usr/bin/env python3
"""
Rebuild broken tabs v2 - FULL master template compliance.
Each support tab gets the SAME header structure as the main tab:
  Row 1-5: Logo zone | Title zone | Meta zone
  Row 6: Rule stripe (1.5pt)
  Row 7+: Content (section header + table)

Copies header from main tab to support tab for consistency.
"""
import os, sys, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

FONT_FAMILY = 'Segoe UI'
HESEM_BLUE = '2C9CD7'
DARK_NAVY = '0C2D48'
MED_BLUE = '1565C0'
BODY_TEXT = '1A2733'
THIN_CLR = 'C5D9E8'
LABEL_BG = 'F0F7FC'
EFF6FF = 'EFF6FF'

MEDIUM_SIDE = Side(style='medium', color=HESEM_BLUE)
THIN_SIDE = Side(style='thin', color=THIN_CLR)
NO_SIDE = Side(style=None)

BASE = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
LOGO = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau', '00-FORM-DESIGN-SYSTEM', 'hesem-logo.png')

# Check if logo exists, try alternate paths
if not os.path.exists(LOGO):
    # Try to find it
    for root, dirs, files in os.walk(os.path.dirname(__file__)):
        for f in files:
            if 'hesem' in f.lower() and 'logo' in f.lower() and f.endswith('.png'):
                LOGO = os.path.join(root, f)
                break


def detect_ncols(ws):
    """Detect grid width from main tab."""
    count = 0
    for col_idx in range(1, 200):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width is not None and abs(dim.width - 2.33) < 0.5:
            count += 1
        elif count > 10:
            break
    if count <= 44: return 40
    elif count <= 60: return 56
    else: return 82


def get_header_zones(ncols):
    """Return column boundaries for header zones."""
    if ncols <= 40:
        return {
            'logo': (1, 8), 'title': (9, 30),
            'meta_lbl': (31, 34), 'meta_val': (35, 40),
            'ncols': 40
        }
    elif ncols <= 56:
        return {
            'logo': (1, 11), 'title': (12, 42),
            'meta_lbl': (43, 48), 'meta_val': (49, 56),
            'ncols': 56
        }
    else:
        return {
            'logo': (1, 16), 'title': (17, 62),
            'meta_lbl': (63, 70), 'meta_val': (71, 82),
            'ncols': 82
        }


def copy_header_from_main(main_ws, target_ws, ncols, tab_title):
    """Copy header structure (rows 1-6) from main tab to target tab."""
    zones = get_header_zones(ncols)

    # Ensure column widths match
    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        target_ws.column_dimensions[letter].width = 2.33

    # Hide grid lines
    target_ws.sheet_view.showGridLines = False

    # Set page setup
    target_ws.page_margins.left = 0.197
    target_ws.page_margins.right = 0.197
    target_ws.page_margins.top = 0.197
    target_ws.page_margins.bottom = 0.197

    # Clear target rows 1-6
    for row in range(1, 7):
        for col in range(1, ncols + 1):
            cell = target_ws.cell(row=row, column=col)
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill('solid', fgColor='FFFFFF')

    # Copy cell values and styles from main tab rows 1-5
    for row in range(1, 6):
        target_ws.row_dimensions[row].height = 15.0
        for col in range(1, ncols + 1):
            src = main_ws.cell(row=row, column=col)
            dst = target_ws.cell(row=row, column=col)

            # Copy value
            if src.value is not None:
                dst.value = src.value

            # Copy font
            if src.font:
                dst.font = Font(
                    name=src.font.name or FONT_FAMILY,
                    size=src.font.size,
                    bold=src.font.bold,
                    italic=src.font.italic,
                    color=src.font.color
                )

            # Copy fill
            if src.fill and src.fill.fill_type:
                dst.fill = copy.copy(src.fill)

            # Copy alignment
            if src.alignment:
                dst.alignment = copy.copy(src.alignment)

            # Copy border
            if src.border:
                dst.border = copy.copy(src.border)

    # Row 6: Rule stripe
    target_ws.row_dimensions[6].height = 1.5
    for col in range(1, ncols + 1):
        cell = target_ws.cell(row=6, column=col)
        cell.fill = PatternFill('solid', fgColor=HESEM_BLUE)
        cell.border = Border()

    # Copy merges from main tab (rows 1-6 only)
    # First remove existing merges in rows 1-6
    for mg in list(target_ws.merged_cells.ranges):
        if mg.min_row <= 6:
            target_ws.unmerge_cells(str(mg))

    # Add merges from main tab
    for mg in main_ws.merged_cells.ranges:
        if mg.min_row <= 6 and mg.max_row <= 6:
            try:
                target_ws.merge_cells(
                    start_row=mg.min_row, start_column=mg.min_col,
                    end_row=mg.max_row, end_column=mg.max_col
                )
            except:
                pass

    # Override title text for this specific tab
    title_start = zones['title'][0]
    title_cell = target_ws.cell(row=1, column=title_start)
    if title_cell.value:
        # Keep original title from main tab
        pass

    # Try to add logo image
    try:
        if os.path.exists(LOGO):
            img = XlImage(LOGO)
            logo_end_col = zones['logo'][1]
            # Size logo to fit zone: ~55x55px for small zone, ~70x70 for larger
            logo_size = 55 if ncols <= 40 else 65
            img.width = logo_size
            img.height = logo_size
            # Center in logo zone
            logo_width_mm = logo_end_col * 5.03  # each col ~5mm
            offset_x_mm = (logo_width_mm - logo_size * 0.264) / 2  # px to mm
            img.anchor = 'B2'  # Place at B2 for centering
            target_ws.add_image(img)
    except:
        pass


def build_table(ws, start_row, ncols, layout, data_rows, section_title, section_desc):
    """Build a table section starting at start_row."""

    row = start_row

    # Section header
    ws.row_dimensions[row].height = 18.0
    # Remove existing merges in this area
    for mg in list(ws.merged_cells.ranges):
        if mg.min_row >= start_row:
            try:
                ws.unmerge_cells(str(mg))
            except:
                pass

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.value = section_title
    cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=MED_BLUE)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=MED_BLUE)
        ws.cell(row=row, column=c).border = Border(
            top=MEDIUM_SIDE, bottom=MEDIUM_SIDE,
            left=MEDIUM_SIDE if c == 1 else NO_SIDE,
            right=MEDIUM_SIDE if c == ncols else NO_SIDE
        )
    row += 1

    # Column headers
    ws.row_dimensions[row].height = 18.0
    for header, start, end in layout:
        # Scale columns to ncols
        if ncols != 40:
            scale = ncols / 40
            start = max(1, int((start - 1) * scale) + 1)
            end = min(ncols, int(end * scale))

        ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        cell = ws.cell(row=row, column=start)
        cell.value = header
        cell.font = Font(name=FONT_FAMILY, size=8, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=DARK_NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        for c in range(start, end + 1):
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=DARK_NAVY)
            ws.cell(row=row, column=c).border = Border(
                top=MEDIUM_SIDE, bottom=MEDIUM_SIDE,
                left=MEDIUM_SIDE if c == 1 else (THIN_SIDE if c == start else NO_SIDE),
                right=MEDIUM_SIDE if c == ncols else (THIN_SIDE if c == end else NO_SIDE)
            )
    row += 1

    # Data rows
    anchor_row = row
    for i in range(data_rows):
        is_even = i % 2 == 1
        ws.row_dimensions[row].height = 20.0
        fill = PatternFill('solid', fgColor='F7FBFF') if is_even else PatternFill('solid', fgColor='FFFFFF')

        for col_idx, (header, start, end) in enumerate(layout):
            if ncols != 40:
                scale = ncols / 40
                start = max(1, int((start - 1) * scale) + 1)
                end = min(ncols, int(end * scale))

            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            cell = ws.cell(row=row, column=start)

            if header == '#':
                cell.value = f"=ROW()-ROW($A${anchor_row})+1"
                cell.font = Font(name=FONT_FAMILY, size=8, bold=True, color=DARK_NAVY)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = Font(name=FONT_FAMILY, size=8, bold=False, color=BODY_TEXT)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

            for c in range(start, end + 1):
                ws.cell(row=row, column=c).fill = fill
                is_last_data = (i == data_rows - 1)
                ws.cell(row=row, column=c).border = Border(
                    top=THIN_SIDE,
                    bottom=MEDIUM_SIDE if is_last_data else THIN_SIDE,
                    left=MEDIUM_SIDE if c == 1 else (THIN_SIDE if c == start else NO_SIDE),
                    right=MEDIUM_SIDE if c == ncols else (THIN_SIDE if c == end else NO_SIDE)
                )
        row += 1

    return row


# ─── TABLE LAYOUTS (40-col base, will be scaled) ──────────────

LAYOUT_ACTIONS = [
    ('#', 1, 2), ('Issue / Requirement', 3, 12),
    ('Required Action', 13, 24), ('Owner', 25, 29),
    ('Due Date', 30, 34), ('Status', 35, 37), ('Verified by', 38, 40),
]

LAYOUT_CONDITIONS = [
    ('#', 1, 2), ('Condition / Exception', 3, 12),
    ('Required Action', 13, 24), ('Owner', 25, 29),
    ('Due Date', 30, 34), ('Status', 35, 37), ('Verified by', 38, 40),
]

LAYOUT_8D_TRACKER = [
    ('#', 1, 2), ('8D Step / Action', 3, 14),
    ('Owner', 15, 19), ('Due Date', 20, 24),
    ('Status', 25, 28), ('Effectiveness', 29, 33),
    ('Verified by', 34, 37), ('Notes', 38, 40),
]

LAYOUT_RAW_DATA = [
    ('#', 1, 2), ('Part', 3, 6), ('Appraiser', 7, 10),
    ('Trial 1', 11, 14), ('Trial 2', 15, 18), ('Trial 3', 19, 22),
    ('Average', 23, 27), ('Range', 28, 32), ('Decision', 33, 40),
]

LAYOUT_CAPABILITY = [
    ('#', 1, 2), ('Characteristic', 3, 12),
    ('USL', 13, 16), ('LSL', 17, 20), ('Mean', 21, 24),
    ('Sigma', 25, 28), ('Cp', 29, 31), ('Cpk', 32, 34),
    ('Decision', 35, 40),
]


# ─── TAB CONFIGS ─────────────────────────────────────────────

TAB_CONFIGS = {
    ('FRM-403', 'ACTIONS'): ('ACTION TRACKER', LAYOUT_ACTIONS, 10),
    ('FRM-406', 'ACTIONS'): ('SCAR ACTION TRACKER', LAYOUT_ACTIONS, 10),
    ('FRM-411', 'ACTIONS'): ('ACTION TRACKER', LAYOUT_ACTIONS, 10),
    ('FRM-651', 'ACTIONS'): ('NCR ACTION TRACKER', LAYOUT_ACTIONS, 10),
    ('FRM-408', 'CONDITIONS_ACTIONS'): ('CONDITIONS / ACTIONS', LAYOUT_CONDITIONS, 8),
    ('FRM-409', 'CONDITIONS_ACTIONS'): ('AUDIT FINDINGS / ACTIONS', LAYOUT_CONDITIONS, 10),
    ('FRM-702', 'CONDITIONS_ACTIONS'): ('SHIPPING CONDITIONS / ACTIONS', LAYOUT_CONDITIONS, 8),
    ('FRM-707', 'CONDITIONS_ACTIONS'): ('PACKAGING CONDITIONS / ACTIONS', LAYOUT_CONDITIONS, 8),
    ('FRM-709', 'CONDITIONS_ACTIONS'): ('CLEAN-PACK CONDITIONS / ACTIONS', LAYOUT_CONDITIONS, 8),
    ('FRM-715', 'CONDITIONS_ACTIONS'): ('VAC-BUILD CONDITIONS / ACTIONS', LAYOUT_CONDITIONS, 8),
    ('FRM-721', 'CONDITIONS_ACTIONS'): ('FOD CLEARANCE ACTIONS', LAYOUT_CONDITIONS, 8),
    ('FRM-652', 'ACTION_TRACKER'): ('CAPA / 8D ACTION TRACKER', LAYOUT_8D_TRACKER, 12),
    ('FRM-653', 'ACTION_TRACKER'): ('A3 PDCA ACTION TRACKER', LAYOUT_8D_TRACKER, 10),
    ('FRM-611', 'RAW_DATA'): ('GR&R RAW DATA', LAYOUT_RAW_DATA, 30),
    ('FRM-612', 'RAW_DATA'): ('BIAS / LINEARITY RAW DATA', LAYOUT_RAW_DATA, 20),
    ('FRM-613', 'RAW_DATA'): ('ATTRIBUTE MSA RAW DATA', LAYOUT_RAW_DATA, 20),
    ('FRM-631', 'CAPABILITY_REVIEW'): ('PROCESS CAPABILITY REVIEW', LAYOUT_CAPABILITY, 15),
}


def find_file(code):
    for root, dirs, files in os.walk(BASE):
        if '_build' in root: continue
        for f in files:
            if f.startswith(code) and f.endswith('.xlsx') and not f.startswith('~'):
                return os.path.join(root, f)
    return None


def main():
    print(f"{'='*70}")
    print("REBUILD TABS v2 - Full Master Template Compliance")
    print(f"Tabs to rebuild: {len(TAB_CONFIGS)}")
    print(f"{'='*70}")

    rebuilt = errors = 0

    for (code, tab_name), (section_title, layout, data_rows) in sorted(TAB_CONFIGS.items()):
        filepath = find_file(code)
        if not filepath:
            print(f"  [{code}] FILE NOT FOUND")
            errors += 1
            continue

        try:
            wb = load_workbook(filepath)

            if tab_name not in wb.sheetnames:
                print(f"  [{code}] Tab '{tab_name}' not found")
                errors += 1
                wb.close()
                continue

            main_ws = wb[wb.sheetnames[0]]  # Main tab
            target_ws = wb[tab_name]         # Tab to rebuild
            ncols = detect_ncols(main_ws)

            # Step 1: Clear ALL content in target
            for mg in list(target_ws.merged_cells.ranges):
                try:
                    target_ws.unmerge_cells(str(mg))
                except:
                    pass

            max_r = max(target_ws.max_row or 1, 50)
            for row in range(1, max_r + 1):
                for col in range(1, ncols + 1):
                    cell = target_ws.cell(row=row, column=col)
                    cell.value = None
                    cell.font = Font(name=FONT_FAMILY, size=8)
                    cell.fill = PatternFill()
                    cell.border = Border()
                    cell.alignment = Alignment()

            # Step 2: Copy header from main tab (rows 1-6)
            copy_header_from_main(main_ws, target_ws, ncols, section_title)

            # Step 3: Build table starting at row 7
            build_table(target_ws, 7, ncols, layout, data_rows, section_title, '')

            wb.save(filepath)
            wb.close()
            print(f"  [{code}] {tab_name:25s} REBUILT (ncols={ncols}, {data_rows} rows)")
            rebuilt += 1

        except Exception as e:
            print(f"  [{code}] ERROR: {e}")
            import traceback; traceback.print_exc()
            errors += 1

    print(f"\n{'='*70}")
    print(f"Rebuilt: {rebuilt}, Errors: {errors}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
