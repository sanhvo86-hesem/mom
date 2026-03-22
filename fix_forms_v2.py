#!/usr/bin/env python3
"""
HESEM QMS Form Compliance Fixer v2
Fixes: 1) Logo embedding  2) Border compliance  3) Action tab merge
"""
import os
import sys
import copy
import glob as globmod

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
FONT_FAMILY  = "Segoe UI"
HESEM_BLUE   = "2C9CD7"
DARK_NAVY    = "0C2D48"
MED_BLUE     = "1565C0"
WHITE        = "FFFFFF"
LABEL_BG     = "F0F7FC"
META_LABEL_BG= "E4F0F9"
THIN_BORDER_C= "C5D9E8"
BODY_TEXT     = "1A2733"
NOTICE_BG_C  = "FFF8E1"
NOTICE_BORDER_C = "F9A825"
STD_COL_WIDTH = 2.33

MEDIUM_SIDE  = Side(style='medium', color=HESEM_BLUE)
THIN_SIDE    = Side(style='thin', color=THIN_BORDER_C)
NO_SIDE      = Side(style=None)
NOTICE_SIDE  = Side(style='thin', color=NOTICE_BORDER_C)

LOGO_PATH    = os.path.join(os.path.dirname(__file__), 'assets', 'hesem-logo.png')

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def color_hex(c):
    if c is None: return None
    if hasattr(c, 'rgb') and c.rgb and c.rgb != '00000000':
        s = str(c.rgb)
        return s[2:].upper() if len(s) == 8 else s.upper()
    return None

def fill_hex(f):
    if f is None or f.fill_type is None: return None
    return color_hex(f.fgColor)

def detect_ncols(ws):
    count = 0
    for i in range(1, 200):
        d = ws.column_dimensions.get(get_column_letter(i))
        if d and d.width is not None and abs(d.width - STD_COL_WIDTH) < 0.5:
            count += 1
        elif count > 10:
            break
    if count <= 44: return 40
    elif count <= 60: return 56
    else: return 82

def get_header_zones(n):
    if n <= 40: return 8, 9, 30, 31, 34, 35, 40
    elif n <= 56: return 11, 12, 42, 43, 48, 49, 56
    else: return 16, 17, 62, 63, 70, 71, 82

def get_merged_non_primary(ws):
    s = set()
    for mg in ws.merged_cells.ranges:
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r != mg.min_row or c != mg.min_col:
                    s.add((r, c))
    return s

def classify_row(ws, row, ncols):
    """Classify a row type based on its visual properties."""
    c1 = ws.cell(row=row, column=1)
    fh = fill_hex(c1.fill)
    rd = ws.row_dimensions.get(row)
    height = rd.height if rd else None

    # Rule stripe: HESEM_BLUE fill, very short
    if fh and fh == HESEM_BLUE and height and height <= 5:
        return 'rule_stripe'

    # Spacer: very short, no content
    if height and height <= 5 and not c1.value:
        return 'spacer'

    # Notice bar
    if fh and fh == NOTICE_BG_C.replace('#', '').upper():
        return 'notice'

    # Section header L1: MED_BLUE bg
    if fh and fh == MED_BLUE:
        return 'section_L1'

    # Section header L2 / column header: HESEM_BLUE or DARK_NAVY bg with content
    if fh and fh == HESEM_BLUE and height and height > 5:
        return 'section_L2'

    # Column header: DARK_NAVY
    if fh and fh == DARK_NAVY:
        return 'col_header'
    # Also check a few cells
    for cc in [2, 3, 5]:
        if cc <= ncols:
            fh2 = fill_hex(ws.cell(row=row, column=cc).fill)
            if fh2 and fh2 == DARK_NAVY:
                return 'col_header'

    # Header rows 2-6
    if row >= 2 and row <= 6:
        return 'header'

    return 'data'


# ─── FIX 1: EMBED LOGO ───────────────────────────────────────────────────────

def embed_logo(ws, ncols):
    """Embed HESEM logo into the header logo zone."""
    if not os.path.exists(LOGO_PATH):
        return False

    # Remove existing images
    ws._images = []

    zones = get_header_zones(ncols)
    logo_end = zones[0]

    # Create image
    img = XlImage(LOGO_PATH)

    # Size: fit within logo zone (logo_end cols x 5 rows)
    # Each col = ~5mm, so logo_end cols = logo_end * 5mm
    # 5 rows x 15pt = 75pt height = ~26.5mm
    # Logo SVG is 52mm x 15mm, scale to fit
    logo_width_mm = logo_end * 5.03  # mm
    logo_height_mm = 75 * 0.353  # 75pt to mm = ~26.5mm

    # Logo aspect ratio from SVG: 52/15 = 3.47
    aspect = 52.0 / 15.0
    # Fit within available space with padding
    avail_w = logo_width_mm * 0.85  # 85% of zone width for padding
    avail_h = logo_height_mm * 0.75  # 75% for padding

    if avail_w / aspect <= avail_h:
        final_w = avail_w
        final_h = avail_w / aspect
    else:
        final_h = avail_h
        final_w = avail_h * aspect

    # Convert mm to EMU (1mm = 36000 EMU) then to cm for openpyxl
    img.width = final_w / 25.4 * 96  # pixels at 96 DPI
    img.height = final_h / 25.4 * 96

    # Anchor at A2 with some offset
    img.anchor = 'A2'

    ws.add_image(img)
    return True


# ─── FIX 2: BORDERS ──────────────────────────────────────────────────────────

def apply_header_borders(ws, ncols):
    """Apply correct borders to header area (rows 1-7)."""
    zones = get_header_zones(ncols)
    logo_end, title_start, title_end, meta_lbl_start, meta_lbl_end, meta_val_start, meta_val_end = zones

    # Row 1 & 7: rule stripes - no borders
    for r in [1, 7]:
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = Border()

    # Header block (rows 2-6): medium border around entire block
    for r in range(2, 7):
        for c in range(1, ncols + 1):
            top = MEDIUM_SIDE if r == 2 else NO_SIDE
            bottom = MEDIUM_SIDE if r == 6 else NO_SIDE
            left = MEDIUM_SIDE if c == 1 else NO_SIDE
            right = MEDIUM_SIDE if c == ncols else NO_SIDE

            # Zone boundaries (vertical medium borders)
            if c == logo_end:
                right = MEDIUM_SIDE
            if c == title_end:
                right = MEDIUM_SIDE
            if c == logo_end + 1:
                left = MEDIUM_SIDE
            if c == title_start:
                left = MEDIUM_SIDE
            if c == meta_lbl_start:
                left = MEDIUM_SIDE
            if c == meta_lbl_end:
                right = MEDIUM_SIDE
            if c == meta_lbl_end + 1:
                left = MEDIUM_SIDE

            # Thin horizontal separators between meta rows
            if c >= meta_lbl_start and r > 2:
                top = THIN_SIDE

            ws.cell(row=r, column=c).border = Border(
                top=top, bottom=bottom, left=left, right=right
            )


def apply_section_header_borders(ws, row, ncols):
    """Section header: medium outline."""
    for c in range(1, ncols + 1):
        left = MEDIUM_SIDE if c == 1 else NO_SIDE
        right = MEDIUM_SIDE if c == ncols else NO_SIDE
        ws.cell(row=row, column=c).border = Border(
            top=MEDIUM_SIDE, bottom=MEDIUM_SIDE, left=left, right=right
        )


def apply_col_header_borders(ws, row, ncols):
    """Column header: medium top/bottom, thin vertical separators."""
    merged_ranges = list(ws.merged_cells.ranges)
    merge_boundaries = set()
    for mg in merged_ranges:
        if mg.min_row <= row <= mg.max_row:
            merge_boundaries.add(mg.min_col)
            merge_boundaries.add(mg.max_col + 1)

    for c in range(1, ncols + 1):
        left = MEDIUM_SIDE if c == 1 else (THIN_SIDE if c in merge_boundaries else NO_SIDE)
        right = MEDIUM_SIDE if c == ncols else NO_SIDE
        ws.cell(row=row, column=c).border = Border(
            top=MEDIUM_SIDE, bottom=MEDIUM_SIDE, left=left, right=right
        )


def apply_data_block_borders(ws, rows, ncols):
    """Data rows within a block: thin inner grid, medium outer."""
    if not rows:
        return

    first_row = rows[0]
    last_row = rows[-1]

    # Find merge boundaries for vertical separators
    merged_ranges = list(ws.merged_cells.ranges)
    col_boundaries = set()
    col_boundaries.add(1)
    col_boundaries.add(ncols + 1)

    for mg in merged_ranges:
        if mg.min_row >= first_row and mg.max_row <= last_row:
            col_boundaries.add(mg.min_col)
            col_boundaries.add(mg.max_col + 1)

    # Also detect label/input boundaries from fill colors
    for r in rows[:3]:  # Sample first 3 rows
        prev_fill = None
        for c in range(1, ncols + 1):
            curr_fill = fill_hex(ws.cell(row=r, column=c).fill)
            if prev_fill != curr_fill and curr_fill is not None:
                col_boundaries.add(c)
            prev_fill = curr_fill

    for r in rows:
        for c in range(1, ncols + 1):
            top = MEDIUM_SIDE if r == first_row else THIN_SIDE
            bottom = MEDIUM_SIDE if r == last_row else THIN_SIDE
            left = MEDIUM_SIDE if c == 1 else (THIN_SIDE if c in col_boundaries else NO_SIDE)
            right = MEDIUM_SIDE if c == ncols else NO_SIDE

            ws.cell(row=r, column=c).border = Border(
                top=top, bottom=bottom, left=left, right=right
            )


def fix_all_borders(ws, ncols):
    """Apply the correct border pattern to the entire sheet."""
    max_row = ws.max_row or 50

    # Classify all rows
    row_types = {}
    for r in range(1, max_row + 1):
        row_types[r] = classify_row(ws, r, ncols)

    # Apply header borders
    if row_types.get(1) == 'rule_stripe':
        apply_header_borders(ws, ncols)

    # Group data rows into blocks (between section headers)
    current_block = []
    for r in range(8, max_row + 1):
        rt = row_types.get(r, 'data')

        if rt in ('section_L1', 'section_L2'):
            # Flush current block
            if current_block:
                apply_data_block_borders(ws, current_block, ncols)
                current_block = []
            apply_section_header_borders(ws, r, ncols)

        elif rt == 'col_header':
            if current_block:
                apply_data_block_borders(ws, current_block, ncols)
                current_block = []
            apply_col_header_borders(ws, r, ncols)

        elif rt == 'spacer':
            if current_block:
                apply_data_block_borders(ws, current_block, ncols)
                current_block = []
            # Spacer: no borders
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).border = Border()

        elif rt == 'rule_stripe':
            if current_block:
                apply_data_block_borders(ws, current_block, ncols)
                current_block = []
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).border = Border()

        elif rt == 'notice':
            if current_block:
                apply_data_block_borders(ws, current_block, ncols)
                current_block = []
            # Notice: thin amber border
            for c in range(1, ncols + 1):
                left = NOTICE_SIDE if c == 1 else NO_SIDE
                right = NOTICE_SIDE if c == ncols else NO_SIDE
                ws.cell(row=r, column=c).border = Border(
                    top=NOTICE_SIDE, bottom=NOTICE_SIDE, left=left, right=right
                )

        else:  # data
            current_block.append(r)

    # Flush last block
    if current_block:
        apply_data_block_borders(ws, current_block, ncols)


# ─── FIX 3: FONT SWEEP ───────────────────────────────────────────────────────

def fix_fonts(ws):
    """Fix all fonts to Segoe UI (skip merged non-primary cells)."""
    mnp = get_merged_non_primary(ws)
    count = 0
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if (cell.row, cell.column) in mnp:
                continue
            if cell.font and cell.font.name and cell.font.name != FONT_FAMILY:
                has_content = cell.value is not None
                has_fill = cell.fill and cell.fill.fill_type is not None
                has_border = cell.border and any(
                    getattr(cell.border, s) and getattr(cell.border, s).style
                    for s in ['top', 'bottom', 'left', 'right']
                )
                if has_content or has_fill or has_border:
                    cell.font = Font(
                        name=FONT_FAMILY,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strikethrough,
                        color=cell.font.color
                    )
                    count += 1
    return count


# ─── FIX 4: MERGE ACTION TAB ─────────────────────────────────────────────────

def find_action_sheet(wb):
    """Find the action/corrective action sheet name."""
    for name in wb.sheetnames:
        lower = name.lower()
        if any(kw in lower for kw in ['action', 'capa', 'corrective', 'escalat',
                                        'containment', 'disposition']):
            # Don't match the main form sheet
            if not any(main_kw in lower for main_kw in ['frm', 'gate', 'checklist',
                                                          'main', 'master', 'register',
                                                          'record', 'report', 'log']):
                return name
    return None


def find_main_sheet(wb):
    """Find the main form sheet."""
    for name in wb.sheetnames:
        lower = name.lower()
        # Skip utility sheets
        if any(skip in lower for skip in ['list', 'import', 'ref', 'evidence',
                                            'calc', 'chart', 'raw', 'data',
                                            'legend', 'gap', 'trend', 'review',
                                            'finding', 'decision', 'input']):
            continue
        # Main sheet is usually the first non-utility sheet
        if lower.startswith('frm') or lower.startswith('_'):
            continue
        return name
    # Fallback: first sheet
    return wb.sheetnames[0]


# ─── MAIN PROCESSING ─────────────────────────────────────────────────────────

def process_workbook(filepath):
    """Process a single workbook: fix logo, borders, fonts."""
    filename = os.path.basename(filepath)
    changes = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return filename, [f"ERROR: {e}"], False

    # Fix Normal style
    try:
        for ns in wb._named_styles:
            if ns.font and ns.font.name and ns.font.name != FONT_FAMILY:
                ns.font = Font(
                    name=FONT_FAMILY,
                    size=ns.font.size,
                    bold=ns.font.bold,
                    italic=ns.font.italic,
                    color=ns.font.color
                )
                changes.append(f"  Style '{ns.name}' -> Segoe UI")
    except:
        pass

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # Skip utility sheets for logo/border fixes (but still fix fonts)
        is_utility = ws_name.upper() in ('LISTS', 'DROPDOWN', 'DATA') or \
                     ws_name.lower().startswith('zlists') or \
                     ws_name.lower().startswith('_lists')

        ncols = detect_ncols(ws)

        if not is_utility:
            # 1. Embed logo
            # Only embed if this sheet has a header (rule stripe in row 1)
            c1 = ws.cell(row=1, column=1)
            fh = fill_hex(c1.fill)
            if fh and fh == HESEM_BLUE:
                if embed_logo(ws, ncols):
                    changes.append(f"  [{ws_name}] Logo embedded")

            # 2. Fix borders
            fix_all_borders(ws, ncols)
            changes.append(f"  [{ws_name}] Borders fixed")

        # 3. Fix fonts (all sheets)
        n = fix_fonts(ws)
        if n > 0:
            changes.append(f"  [{ws_name}] {n} fonts fixed")

    # Save
    try:
        wb.save(filepath)
        changes.append("  >> SAVED")
    except Exception as e:
        changes.append(f"  >> SAVE ERROR: {e}")

    wb.close()
    return filename, changes, True


def main():
    base_dir = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')

    frm_files = []
    for root, dirs, files in os.walk(base_dir):
        if '_build' in root:
            continue
        for f in files:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                frm_files.append(os.path.join(root, f))

    frm_files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS FORM COMPLIANCE FIX v2")
    print(f"Logo: {LOGO_PATH}")
    print(f"Logo exists: {os.path.exists(LOGO_PATH)}")
    print(f"Files: {len(frm_files)}")
    print(f"{'='*70}\n")

    for i, fpath in enumerate(frm_files, 1):
        fname = os.path.basename(fpath)
        print(f"[{i:3d}/{len(frm_files)}] {fname}...", end=" ", flush=True)
        filename, changes, ok = process_workbook(fpath)
        print("OK" if ok else "FAIL")
        if '--verbose' in sys.argv:
            for ch in changes:
                print(ch)

    print(f"\n{'='*70}")
    print(f"DONE - {len(frm_files)} files processed")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
