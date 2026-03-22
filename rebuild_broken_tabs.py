#!/usr/bin/env python3
"""
Rebuild all 17 broken tabs to follow master template grid system.

Problem: Tabs have 1 cell = 1 column with 2.33 width = unreadable.
Solution: Merge multiple grid columns per field to create proper widths.

Grid: 40 cols (A4P) = 200mm usable. Each col = 5mm.

Table layouts per tab type:
  ACTIONS (8 fields):        #(2) | Issue(8) | Action(8) | Owner(5) | Due(5) | Status(4) | Evidence(4) | Notes(4)
  CONDITIONS_ACTIONS (8):    #(2) | Condition(8) | Action(8) | Owner(5) | Due(5) | Status(4) | Evidence(4) | Notes(4)
  ACTION_TRACKER (8):        #(2) | 8D Step(8) | Owner(5) | Due(5) | Status(4) | Effectiveness(4) | Evidence(4) | Notes(8)
  RAW_DATA (9):              #(2) | Part(4) | Appraiser(4) | T1(4) | T2(4) | T3(4) | Avg(4) | Range(4) | Decision(10)
  CAPABILITY_REVIEW (varied): Custom layout
"""
import os, sys, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Master template styles
FONT_FAMILY = 'Segoe UI'
HESEM_BLUE = '2C9CD7'
DARK_NAVY = '0C2D48'
MED_BLUE = '1565C0'
BODY_TEXT = '1A2733'
THIN_BORDER_COLOR = 'C5D9E8'
LABEL_BG = 'F0F7FC'

FONT_TITLE = Font(name=FONT_FAMILY, size=11, bold=True, color=DARK_NAVY)
FONT_DESC = Font(name=FONT_FAMILY, size=7, bold=False, color='7A8FA6')
FONT_COL_HDR = Font(name=FONT_FAMILY, size=8, bold=True, color='FFFFFF')
FONT_BODY = Font(name=FONT_FAMILY, size=8, bold=False, color=BODY_TEXT)
FONT_AUTONUM = Font(name=FONT_FAMILY, size=8, bold=True, color=DARK_NAVY)

FILL_DARK_NAVY = PatternFill('solid', fgColor=DARK_NAVY)
FILL_WHITE = PatternFill('solid', fgColor='FFFFFF')
FILL_STRIPE = PatternFill('solid', fgColor='F7FBFF')
FILL_LABEL = PatternFill('solid', fgColor=LABEL_BG)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrapText=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrapText=True)

MEDIUM_SIDE = Side(style='medium', color=HESEM_BLUE)
THIN_SIDE = Side(style='thin', color=THIN_BORDER_COLOR)
NO_SIDE = Side(style=None)

BASE = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')

# ─── LAYOUT DEFINITIONS ──────────────────────────────────────────

# Each layout: list of (header_text, col_start, col_end)
# Using 1-indexed columns, 40 cols total

LAYOUT_ACTIONS = [
    ('#', 1, 2),
    ('Issue / Requirement', 3, 10),
    ('Required Action', 11, 20),
    ('Owner', 21, 25),
    ('Due Date', 26, 30),
    ('Status', 31, 34),
    ('Verified by', 35, 37),
    ('Notes', 38, 40),
]

LAYOUT_CONDITIONS = [
    ('#', 1, 2),
    ('Condition / Exception', 3, 10),
    ('Required Action', 11, 20),
    ('Owner', 21, 25),
    ('Due Date', 26, 30),
    ('Status', 31, 34),
    ('Verified by', 35, 37),
    ('Notes', 38, 40),
]

LAYOUT_8D_TRACKER = [
    ('#', 1, 2),
    ('8D Step / Action', 3, 12),
    ('Owner', 13, 17),
    ('Due Date', 18, 22),
    ('Status', 23, 26),
    ('Effectiveness', 27, 30),
    ('Evidence Ref', 31, 35),
    ('Notes', 36, 40),
]

LAYOUT_RAW_DATA = [
    ('#', 1, 2),
    ('Part', 3, 6),
    ('Appraiser', 7, 10),
    ('Trial 1', 11, 14),
    ('Trial 2', 15, 18),
    ('Trial 3', 19, 22),
    ('Average', 23, 26),
    ('Range', 27, 30),
    ('Decision', 31, 40),
]

LAYOUT_CAPABILITY = [
    ('#', 1, 2),
    ('Characteristic', 3, 12),
    ('USL', 13, 16),
    ('LSL', 17, 20),
    ('Mean', 21, 24),
    ('Sigma', 25, 28),
    ('Cp', 29, 31),
    ('Cpk', 32, 34),
    ('Decision', 35, 40),
]


def get_layout(tab_name):
    """Get layout for a tab type."""
    lower = tab_name.lower()
    if 'action_tracker' in lower or '8d' in lower:
        return LAYOUT_8D_TRACKER
    elif 'condition' in lower:
        return LAYOUT_CONDITIONS
    elif 'action' in lower:
        return LAYOUT_ACTIONS
    elif 'raw_data' in lower:
        return LAYOUT_RAW_DATA
    elif 'capability' in lower:
        return LAYOUT_CAPABILITY
    return LAYOUT_ACTIONS  # default


def rebuild_tab(ws, tab_name, title_text, desc_text, data_rows=15):
    """Rebuild a broken tab with proper master template grid layout."""

    layout = get_layout(tab_name)
    ncols = 40

    # Clear existing merges
    for mg in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mg))

    # Clear all existing content and styles in the used area
    for row in range(1, (ws.max_row or 20) + 5):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.font = FONT_BODY
            cell.fill = FILL_WHITE
            cell.border = Border()
            cell.alignment = ALIGN_LEFT

    # ─── ROW 1: Section header (title) ───
    ws.row_dimensions[1].height = 18.0
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1)
    cell.value = title_text
    cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=MED_BLUE)
    cell.alignment = ALIGN_LEFT
    # Medium border on section header
    for c in range(1, ncols + 1):
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor=MED_BLUE)
        ws.cell(row=1, column=c).border = Border(
            top=MEDIUM_SIDE, bottom=MEDIUM_SIDE,
            left=MEDIUM_SIDE if c == 1 else NO_SIDE,
            right=MEDIUM_SIDE if c == ncols else NO_SIDE
        )

    # ─── ROW 2: Description ───
    ws.row_dimensions[2].height = 14.0
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    cell = ws.cell(row=2, column=1)
    cell.value = desc_text
    cell.font = FONT_DESC
    cell.fill = FILL_LABEL
    cell.alignment = ALIGN_LEFT

    # ─── ROW 3: Spacer ───
    ws.row_dimensions[3].height = 3.0

    # ─── ROW 4: Column headers ───
    ws.row_dimensions[4].height = 18.0
    for header, start, end in layout:
        ws.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
        cell = ws.cell(row=4, column=start)
        cell.value = header
        cell.font = FONT_COL_HDR
        cell.fill = FILL_DARK_NAVY
        cell.alignment = ALIGN_CENTER

        # Apply fill and borders to all cells in merge
        for c in range(start, end + 1):
            ws.cell(row=4, column=c).fill = FILL_DARK_NAVY
            top = MEDIUM_SIDE
            bottom = MEDIUM_SIDE
            left = MEDIUM_SIDE if c == 1 else (THIN_SIDE if c == start else NO_SIDE)
            right = MEDIUM_SIDE if c == ncols else (THIN_SIDE if c == end else NO_SIDE)
            ws.cell(row=4, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)

    # ─── ROWS 5+: Data rows ───
    for row_idx in range(5, 5 + data_rows):
        is_even = (row_idx - 5) % 2 == 1
        ws.row_dimensions[row_idx].height = 20.0

        for col_idx, (header, start, end) in enumerate(layout):
            ws.merge_cells(start_row=row_idx, start_column=start, end_row=row_idx, end_column=end)
            cell = ws.cell(row=row_idx, column=start)

            if header == '#':
                # Auto-number formula
                cell.value = f"=ROW()-ROW($A$5)+1"
                cell.font = FONT_AUTONUM
                cell.alignment = ALIGN_CENTER
                cell.fill = FILL_STRIPE if is_even else FILL_WHITE
            else:
                cell.font = FONT_BODY
                cell.alignment = ALIGN_LEFT
                cell.fill = FILL_STRIPE if is_even else FILL_WHITE

            # Borders for all cells in merge
            for c in range(start, end + 1):
                ws.cell(row=row_idx, column=c).fill = FILL_STRIPE if is_even else FILL_WHITE
                left_b = MEDIUM_SIDE if c == 1 else (THIN_SIDE if c == start else NO_SIDE)
                right_b = MEDIUM_SIDE if c == ncols else (THIN_SIDE if c == end else NO_SIDE)
                top_b = THIN_SIDE
                bottom_b = MEDIUM_SIDE if row_idx == 4 + data_rows else THIN_SIDE
                ws.cell(row=row_idx, column=c).border = Border(
                    top=top_b, bottom=bottom_b, left=left_b, right=right_b
                )


# ─── TAB CONFIGURATIONS ──────────────────────────────────────────

TAB_CONFIGS = {
    # ACTIONS tabs
    ('FRM-403', 'ACTIONS'): ('ACTION TRACKER', 'Record outsource issues, required actions, owners and due dates.', 10),
    ('FRM-406', 'ACTIONS'): ('SCAR ACTION TRACKER', 'Track corrective actions from supplier corrective action request.', 10),
    ('FRM-411', 'ACTIONS'): ('ACTION TRACKER', 'Record incoming verification issues and required follow-up actions.', 10),
    ('FRM-651', 'ACTIONS'): ('NCR ACTION TRACKER', 'Track containment, disposition, and corrective actions from NCR.', 10),

    # CONDITIONS_ACTIONS tabs
    ('FRM-408', 'CONDITIONS_ACTIONS'): ('CONDITIONS / ACTIONS', 'Record open conditions and required actions for requirements flow-down.', 8),
    ('FRM-409', 'CONDITIONS_ACTIONS'): ('AUDIT FINDINGS / ACTIONS', 'Record supplier audit findings with corrective actions and follow-up.', 10),
    ('FRM-702', 'CONDITIONS_ACTIONS'): ('SHIPPING CONDITIONS / ACTIONS', 'Record any shipping exceptions, holds, or corrective actions.', 8),
    ('FRM-707', 'CONDITIONS_ACTIONS'): ('PACKAGING CONDITIONS / ACTIONS', 'Record packaging exceptions and corrective actions.', 8),
    ('FRM-709', 'CONDITIONS_ACTIONS'): ('CLEAN-PACK CONDITIONS / ACTIONS', 'Record clean packaging exceptions and contamination concerns.', 8),
    ('FRM-715', 'CONDITIONS_ACTIONS'): ('VAC-BUILD CONDITIONS / ACTIONS', 'Record vacuum-compatible build exceptions and deviations.', 8),
    ('FRM-721', 'CONDITIONS_ACTIONS'): ('FOD CLEARANCE ACTIONS', 'Record FOD findings, tool accountability gaps, and corrective actions.', 8),

    # ACTION_TRACKER tabs
    ('FRM-652', 'ACTION_TRACKER'): ('CAPA / 8D ACTION TRACKER', 'Track D3-D7 actions: containment, root cause fix, prevention, verification.', 12),
    ('FRM-653', 'ACTION_TRACKER'): ('A3 PDCA ACTION TRACKER', 'Track Plan-Do-Check-Act actions with owners and effectiveness review.', 10),

    # RAW_DATA tabs
    ('FRM-611', 'RAW_DATA'): ('GR&R RAW DATA', 'Enter trial measurements: parts x appraisers x trials for repeatability & reproducibility.', 30),
    ('FRM-612', 'RAW_DATA'): ('BIAS / LINEARITY RAW DATA', 'Enter reference values and measured values for bias and linearity study.', 20),
    ('FRM-613', 'RAW_DATA'): ('ATTRIBUTE MSA RAW DATA', 'Enter attribute inspection decisions for agreement analysis.', 20),

    # CAPABILITY_REVIEW
    ('FRM-631', 'CAPABILITY_REVIEW'): ('PROCESS CAPABILITY REVIEW', 'Review Cp/Cpk results for each monitored characteristic.', 15),
}


def main():
    print(f"{'='*70}")
    print("REBUILD BROKEN TABS - Master Template Grid System")
    print(f"Tabs to rebuild: {len(TAB_CONFIGS)}")
    print(f"{'='*70}")

    rebuilt = 0
    errors = 0

    for (code, tab_name), (title, desc, rows) in sorted(TAB_CONFIGS.items()):
        # Find file
        filepath = None
        for root, dirs, files in os.walk(BASE):
            if '_build' in root:
                continue
            for f in files:
                if f.startswith(code) and f.endswith('.xlsx') and not f.startswith('~'):
                    filepath = os.path.join(root, f)
                    break

        if not filepath:
            print(f"  [{code}] FILE NOT FOUND")
            errors += 1
            continue

        try:
            wb = load_workbook(filepath)
            if tab_name in wb.sheetnames:
                ws = wb[tab_name]
                rebuild_tab(ws, tab_name, title, desc, rows)
                wb.save(filepath)
                print(f"  [{code}] {tab_name:25s} REBUILT ({rows} data rows)")
                rebuilt += 1
            else:
                print(f"  [{code}] Tab '{tab_name}' not found")
                errors += 1
            wb.close()
        except Exception as e:
            print(f"  [{code}] ERROR: {e}")
            errors += 1

    print(f"\n{'='*70}")
    print(f"Rebuilt: {rebuilt}, Errors: {errors}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
