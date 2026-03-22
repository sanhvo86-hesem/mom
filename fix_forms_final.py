#!/usr/bin/env python3
"""
HESEM QMS Form Compliance Fixer - FINAL
Based on actual frm-000-master-template.xlsx analysis.

Border spec:  thin/#0078D4 (section boundaries) + thin/#E2E8F0 (internal)
Colors:       Keep existing (forms already match template)
Logo:         Embed hesem-logo.png into header zone A2
Fonts:        Segoe UI everywhere
"""
import os
import sys
import glob as globmod
import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS (from frm-000-master-template.xlsx)
# ═══════════════════════════════════════════════════════════════════════════════

FONT_NAME = "Segoe UI"

# Border colors
BLUE_BORDER  = "0078D4"   # Section boundaries
GRAY_BORDER  = "E2E8F0"   # Internal dividers

# Border sides
B_BLUE  = Side(style='thin', color=BLUE_BORDER)
B_GRAY  = Side(style='thin', color=GRAY_BORDER)
B_NONE  = Side(style=None)

# Fill colors (for detection only - we keep existing)
SECTION_HDR_BG = "EFF6FF"
COL_HDR_BG     = "005A9E"
LABEL_BG       = "F1F5F9"
WHITE_BG       = "FFFFFF"

STD_COL_WIDTH  = 2.33
LOGO_PATH      = os.path.join(os.path.dirname(__file__), 'assets', 'hesem-logo.png')

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def color_hex(c):
    if c is None: return None
    if hasattr(c, 'rgb') and c.rgb and str(c.rgb) != '00000000':
        s = str(c.rgb)
        return s[2:].upper() if len(s) == 8 else s.upper()
    return None

def fill_color(cell):
    if cell.fill and cell.fill.fill_type:
        return color_hex(cell.fill.fgColor)
    return None

def detect_ncols(ws):
    count = 0
    for i in range(1, 200):
        d = ws.column_dimensions.get(get_column_letter(i))
        if d and d.width is not None and abs(d.width - STD_COL_WIDTH) < 0.5:
            count += 1
        elif count > 10:
            break
    if count <= 44: return 40
    elif count <= 60: return 56
    else: return 82

def get_header_zones(n):
    if n <= 40: return 8, 9, 30, 31, 34, 35, 40
    elif n <= 56: return 11, 12, 42, 43, 48, 49, 56
    else: return 16, 17, 62, 63, 70, 71, 82

def has_header(ws):
    """Check if sheet has standard header (merged title in rows 1-5)."""
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 3 and mg.max_row <= 5:
            cell = ws.cell(row=mg.min_row, column=mg.min_col)
            if cell.value and isinstance(cell.value, str) and len(cell.value) > 5:
                return True
    return False

def get_merge_map(ws):
    """Build dict: (row, col) -> merge_range for all merged cells."""
    m = {}
    for mg in ws.merged_cells.ranges:
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                m[(r, c)] = mg
    return m

def get_row_merges(ws, row):
    """Get list of (min_col, max_col) for merges in this row."""
    merges = []
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= row <= mg.max_row:
            merges.append((mg.min_col, mg.max_col))
    merges.sort()
    return merges

def merged_non_primary(ws):
    s = set()
    for mg in ws.merged_cells.ranges:
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r != mg.min_row or c != mg.min_col:
                    s.add((r, c))
    return s


# ═══════════════════════════════════════════════════════════════════════════════
# ROW CLASSIFICATION
# ═══════════════════════════════════════════════════════════════════════════════

def classify_row(ws, row, ncols):
    """Classify a single row."""
    rd = ws.row_dimensions.get(row)
    h = rd.height if rd else None
    c1 = ws.cell(row=row, column=1)
    fh = fill_color(c1)

    # Spacer: height <= 5, usually merged full width
    if h and h <= 5:
        return 'spacer'

    # Section header: EFF6FF bg, full-width merge
    if fh == SECTION_HDR_BG:
        for mg in ws.merged_cells.ranges:
            if mg.min_row == row and mg.max_row == row and (mg.max_col - mg.min_col + 1) >= ncols * 0.5:
                return 'section'
        # Even without merge, if it has section header text
        if c1.value and isinstance(c1.value, str) and len(c1.value.strip()) > 3:
            return 'section'

    # Column header: 005A9E bg
    if fh == COL_HDR_BG:
        return 'col_header'
    for cc in [2, 3, 5, 8]:
        if cc <= ncols:
            if fill_color(ws.cell(row=row, column=cc)) == COL_HDR_BG:
                return 'col_header'

    # Notice bar (same bg as section header but at end of sheet, with notice text)
    if fh == SECTION_HDR_BG:
        val = str(c1.value or '').upper()
        if 'NOTICE' in val or 'LƯU Ý' in val:
            return 'notice'

    return 'data'


# ═══════════════════════════════════════════════════════════════════════════════
# BORDER APPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

def apply_border_to_cell(ws, row, col, top=None, bottom=None, left=None, right=None):
    """Apply border sides to a cell, keeping any sides not specified."""
    cell = ws.cell(row=row, column=col)
    try:
        cell.border = Border(
            top=top if top is not None else B_NONE,
            bottom=bottom if bottom is not None else B_NONE,
            left=left if left is not None else B_NONE,
            right=right if right is not None else B_NONE
        )
    except AttributeError:
        pass  # MergedCell - skip


def apply_header_borders(ws, ncols):
    """Apply borders to header area (rows 1/2 to 5/6)."""
    zones = get_header_zones(ncols)
    logo_end, title_start, title_end, meta_lbl_start, meta_lbl_end, meta_val_start, meta_val_end = zones

    # Determine header rows (usually 2-6 or 1-5)
    # Check if row 1 has content or if header starts at row 2
    c1 = ws.cell(row=1, column=1)
    if fill_color(c1) == WHITE_BG or c1.value is not None:
        hdr_start = 1
    else:
        hdr_start = 2
    hdr_end = hdr_start + 4  # 5 rows total

    for r in range(hdr_start, hdr_end + 1):
        is_first = (r == hdr_start)
        is_last = (r == hdr_end)

        for c in range(1, ncols + 1):
            top = B_BLUE if is_first else B_NONE
            bottom = B_BLUE if is_last else B_NONE
            left = B_BLUE if c == 1 else B_NONE
            right = B_BLUE if c == ncols else B_NONE

            # Zone dividers (vertical blue borders)
            if c == logo_end:
                right = B_BLUE
            if c == logo_end + 1:
                left = B_BLUE
            if c == title_end:
                # Gray divider between title and meta
                pass
            if c == meta_lbl_start:
                left = B_BLUE

            # Meta area: gray internal dividers between rows and between label/value
            if c >= meta_lbl_start:
                if not is_first:
                    top = B_GRAY
                if c == meta_lbl_end:
                    right = B_GRAY
                if c == meta_val_start:
                    left = B_NONE  # Gray right on label serves as divider

            apply_border_to_cell(ws, r, c, top, bottom, left, right)


def apply_section_header_borders(ws, row, ncols):
    """Section header: all 4 sides = blue."""
    for c in range(1, ncols + 1):
        apply_border_to_cell(ws, row, c,
                             top=B_BLUE, bottom=B_BLUE,
                             left=B_BLUE if c == 1 else B_NONE,
                             right=B_BLUE if c == ncols else B_NONE)


def apply_col_header_borders(ws, row, ncols):
    """Column header: T=blue, B=gray, internal R=gray between columns."""
    merges = get_row_merges(ws, row)

    # Build set of merge boundaries
    merge_ends = set()
    for mc, mx in merges:
        merge_ends.add(mx)

    for c in range(1, ncols + 1):
        right = B_BLUE if c == ncols else (B_GRAY if c in merge_ends else B_NONE)
        apply_border_to_cell(ws, row, c,
                             top=B_BLUE, bottom=B_GRAY,
                             left=B_BLUE if c == 1 else B_NONE,
                             right=right)


def apply_data_block_borders(ws, block_rows, ncols):
    """Apply borders to a contiguous block of data rows.

    Rules:
    - First row: T=blue, B=gray
    - Middle rows: T=none, B=gray
    - Last row: T=none, B=blue
    - Left (col 1): L=blue
    - Right (col ncols): R=blue
    - Internal column dividers: R=gray at merge boundaries
    """
    if not block_rows:
        return

    first = block_rows[0]
    last = block_rows[-1]

    # Collect all merge end-columns across the block for vertical dividers
    all_merge_ends = set()
    for r in block_rows:
        merges = get_row_merges(ws, r)
        for mc, mx in merges:
            all_merge_ends.add(mx)

    for r in block_rows:
        is_first = (r == first)
        is_last = (r == last)

        # Get this row's merge ends
        row_merge_ends = set()
        for mc, mx in get_row_merges(ws, r):
            row_merge_ends.add(mx)

        for c in range(1, ncols + 1):
            # Horizontal borders
            if is_first and is_last:
                top = B_BLUE
                bottom = B_BLUE
            elif is_first:
                top = B_BLUE
                bottom = B_GRAY
            elif is_last:
                top = B_NONE
                bottom = B_BLUE
            else:
                top = B_NONE
                bottom = B_GRAY

            # Vertical borders
            left = B_BLUE if c == 1 else B_NONE

            if c == ncols:
                right = B_BLUE
            elif c in row_merge_ends and c < ncols:
                right = B_GRAY
            else:
                right = B_NONE

            apply_border_to_cell(ws, r, c, top, bottom, left, right)


def apply_spacer_borders(ws, row, ncols):
    """Spacer: no borders, ensure height=4."""
    ws.row_dimensions[row].height = 4.0
    for c in range(1, ncols + 1):
        apply_border_to_cell(ws, row, c, B_NONE, B_NONE, B_NONE, B_NONE)


def apply_signature_borders(ws, sig_rows, ncols):
    """Signature area: all 4 sides = blue on each merged block."""
    for r in sig_rows:
        merges = get_row_merges(ws, r)
        merge_ends = set(mx for _, mx in merges)
        merge_starts = set(mc for mc, _ in merges)

        for c in range(1, ncols + 1):
            # Each signature block gets blue borders all around
            top = B_BLUE
            bottom = B_BLUE
            left = B_BLUE if c == 1 or c in merge_starts else B_NONE
            right = B_BLUE if c == ncols or c in merge_ends else B_NONE
            apply_border_to_cell(ws, r, c, top, bottom, left, right)


def apply_notice_borders(ws, row, ncols):
    """Notice bar: all 4 sides = blue."""
    for c in range(1, ncols + 1):
        apply_border_to_cell(ws, row, c,
                             top=B_BLUE, bottom=B_BLUE,
                             left=B_BLUE if c == 1 else B_NONE,
                             right=B_BLUE if c == ncols else B_NONE)


# ═══════════════════════════════════════════════════════════════════════════════
# LOGO EMBEDDING
# ═══════════════════════════════════════════════════════════════════════════════

def embed_logo(ws, ncols):
    """Embed HESEM logo in header logo zone."""
    if not os.path.exists(LOGO_PATH):
        return False

    # Remove existing images
    ws._images = []

    zones = get_header_zones(ncols)
    logo_end = zones[0]

    img = XlImage(LOGO_PATH)

    # Size: fit within logo zone with padding
    zone_w_mm = logo_end * 5.03
    zone_h_mm = 75 * 0.353  # 5 rows x 15pt

    aspect = 52.0 / 15.0  # SVG original ratio
    avail_w = zone_w_mm * 0.90
    avail_h = zone_h_mm * 0.70

    if avail_w / aspect <= avail_h:
        w_mm, h_mm = avail_w, avail_w / aspect
    else:
        h_mm, w_mm = avail_h, avail_h * aspect

    img.width = w_mm / 25.4 * 96
    img.height = h_mm / 25.4 * 96

    # Detect header start row
    c1 = ws.cell(row=1, column=1)
    if fill_color(c1) == WHITE_BG or c1.value is not None:
        anchor = 'A1'
    else:
        anchor = 'A2'

    img.anchor = anchor
    ws.add_image(img)
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# FONT FIX
# ═══════════════════════════════════════════════════════════════════════════════

def fix_fonts(ws):
    """Fix all fonts to Segoe UI."""
    mnp = merged_non_primary(ws)
    count = 0
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if (cell.row, cell.column) in mnp:
                continue
            if cell.font and cell.font.name and cell.font.name != FONT_NAME:
                has_stuff = (cell.value is not None or
                            (cell.fill and cell.fill.fill_type is not None) or
                            (cell.border and any(
                                getattr(cell.border, s) and getattr(cell.border, s).style
                                for s in ['top', 'bottom', 'left', 'right'])))
                if has_stuff:
                    cell.font = Font(
                        name=FONT_NAME,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strikethrough,
                        color=cell.font.color
                    )
                    count += 1
    return count


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN PROCESSING
# ═══════════════════════════════════════════════════════════════════════════════

def process_file(filepath):
    fname = os.path.basename(filepath)
    log = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return fname, [f"ERROR: {e}"], False

    # Fix Normal style
    try:
        for ns in wb._named_styles:
            if ns.font and ns.font.name and ns.font.name != FONT_NAME:
                ns.font = Font(name=FONT_NAME, size=ns.font.size,
                              bold=ns.font.bold, italic=ns.font.italic,
                              color=ns.font.color)
    except:
        pass

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # Skip utility sheets
        skip = ['lists', 'dropdown', 'data', 'zlists', '_lists', '_import',
                'ref_import', 'import_ref']
        if any(ws_name.lower().startswith(s) or ws_name.lower() == s for s in skip):
            fix_fonts(ws)
            continue

        ncols = detect_ncols(ws)
        max_row = ws.max_row or 50
        hdr = has_header(ws)

        # 1. Fix header borders + embed logo
        if hdr:
            apply_header_borders(ws, ncols)
            embed_logo(ws, ncols)
            log.append(f"  [{ws_name}] Header borders + logo")

        # 2. Classify and fix body rows
        # Find header end
        hdr_end = 0
        if hdr:
            # Header is rows 1-5 or 2-6, followed by a spacer
            for r in range(1, 10):
                rt = classify_row(ws, r, ncols)
                if rt == 'spacer' and r >= 5:
                    hdr_end = r
                    break
            if hdr_end == 0:
                hdr_end = 6  # fallback

        # Classify body rows
        current_block = []
        sig_rows = []

        for r in range(hdr_end + 1, max_row + 1):
            rt = classify_row(ws, r, ncols)

            if rt == 'spacer':
                if current_block:
                    apply_data_block_borders(ws, current_block, ncols)
                    current_block = []
                apply_spacer_borders(ws, r, ncols)

            elif rt == 'section':
                if current_block:
                    apply_data_block_borders(ws, current_block, ncols)
                    current_block = []
                apply_section_header_borders(ws, r, ncols)

            elif rt == 'col_header':
                if current_block:
                    apply_data_block_borders(ws, current_block, ncols)
                    current_block = []
                apply_col_header_borders(ws, r, ncols)

            elif rt == 'notice':
                if current_block:
                    apply_data_block_borders(ws, current_block, ncols)
                    current_block = []
                apply_notice_borders(ws, r, ncols)

            else:
                # Check if this is a signature label or box
                c1val = str(ws.cell(row=r, column=1).value or '').lower()
                if any(kw in c1val for kw in ['prepared', 'reviewed', 'approved',
                                                'name', 'signature', 'date',
                                                'chuẩn bị', 'xem xét', 'phê duyệt',
                                                'chữ ký', 'ngày']):
                    if current_block:
                        apply_data_block_borders(ws, current_block, ncols)
                        current_block = []
                    sig_rows.append(r)
                else:
                    # Check if this row continues a signature block
                    fc = fill_color(ws.cell(row=r, column=1))
                    fn_color = color_hex(ws.cell(row=r, column=1).font.color) if ws.cell(row=r, column=1).font and ws.cell(row=r, column=1).font.color else None
                    if fn_color == 'CBD5E1':  # Signature placeholder color
                        sig_rows.append(r)
                    else:
                        if sig_rows:
                            apply_signature_borders(ws, sig_rows, ncols)
                            sig_rows = []
                        current_block.append(r)

        # Flush remaining
        if current_block:
            apply_data_block_borders(ws, current_block, ncols)
        if sig_rows:
            apply_signature_borders(ws, sig_rows, ncols)

        # 3. Fix fonts
        n = fix_fonts(ws)
        if n: log.append(f"  [{ws_name}] {n} fonts fixed")

        # 4. Page setup
        try:
            ws.sheet_view.showGridLines = False
        except:
            pass
        try:
            pm = ws.page_margins
            pm.left = pm.right = 0.197
            pm.top = pm.bottom = 0.197
            pm.header = pm.footer = 0.1
            ps = ws.page_setup
            ps.paperSize = 9 if ncols <= 56 else 8
            ps.fitToWidth = 1
            ps.fitToHeight = 0
        except:
            pass

        log.append(f"  [{ws_name}] borders complete")

    # Save
    try:
        wb.save(filepath)
        log.append("  >> SAVED")
    except Exception as e:
        log.append(f"  >> SAVE ERROR: {e}")

    wb.close()
    return fname, log, True


def main():
    base_dir = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')

    frm_files = []
    for root, dirs, files in os.walk(base_dir):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root:
            continue
        for f in files:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                frm_files.append(os.path.join(root, f))
    frm_files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS FORM FIX - FINAL (borders + logo + fonts)")
    print(f"Template: frm-000-master-template.xlsx")
    print(f"Logo: {'OK' if os.path.exists(LOGO_PATH) else 'MISSING!'}")
    print(f"Files: {len(frm_files)}")
    print(f"{'='*70}\n")

    verbose = '--verbose' in sys.argv

    for i, fpath in enumerate(frm_files, 1):
        fname = os.path.basename(fpath)
        print(f"[{i:3d}/{len(frm_files)}] {fname}...", end=" ", flush=True)
        name, log, ok = process_file(fpath)
        print("OK" if ok else "FAIL")
        if verbose:
            for l in log:
                print(l)

    print(f"\n{'='*70}")
    print(f"DONE - {len(frm_files)} files processed")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
