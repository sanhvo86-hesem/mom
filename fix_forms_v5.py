#!/usr/bin/env python3
"""
HESEM QMS Form Fixer v5 - SAFE COLOR + BORDER + LOGO
Strategy:
  - Map OLD fill colors -> NEW fill colors (cell-by-cell, only if fill exists)
  - Map OLD font colors -> NEW font colors (preserve size/bold/italic)
  - Fix borders: thin/#0078D4 + thin/#E2E8F0
  - Embed logo: sized to ACTUAL logo zone merge, centered
  - Footer: 7pt
  - NEVER override font size or bold - preserve original typography
"""
import os, sys, glob as globmod
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter

# ═══════ CONSTANTS ═══════
BB = Side(style='thin', color='0078D4')
BG = Side(style='thin', color='E2E8F0')
BN = Side(style=None)

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'assets', 'hesem-logo.png')
STD_COL = 2.33
LOGO_ASPECT = 52.0 / 15.0

# Fill color mapping: OLD -> NEW (master template target)
FILL_MAP = {
    '1565C0': 'EFF6FF',   # section header: dark blue -> light blue
    '0C2D48': '005A9E',   # col header: dark navy -> medium blue
    'F0F7FC': 'F1F5F9',   # labels: slight tone shift
    'E4F0F9': 'F1F5F9',   # meta labels -> standard label
    'FFF8E1': 'EFF6FF',   # notice bar -> same as section header
    '2C9CD7': 'FFFFFF',   # rule stripe -> white (convert to spacer)
}

# Font color mapping: OLD -> NEW
FONT_MAP = {
    '0C2D48': '1B3A6B',   # dark navy text -> standard navy
    'BBCCCC': 'CBD5E1',   # sig placeholder -> standard placeholder
    '7B4F00': '334155',   # gold notice text -> standard text
    '7A8FA6': '64748B',   # gray text -> standard gray
}

# Section/col header fills for detection
SECTION_FILLS = {'EFF6FF', '1565C0', 'E8F0FE', 'DCEEFB'}
COLHDR_FILLS = {'005A9E', '0C2D48', '1B4472', '1F3864'}

# ═══════ HELPERS ═══════
def chex(c):
    if c is None: return None
    if hasattr(c, 'rgb') and c.rgb and str(c.rgb) != '00000000':
        s = str(c.rgb); return s[2:].upper() if len(s) == 8 else s.upper()
    return None

def fhex(cell):
    return chex(cell.fill.fgColor) if cell.fill and cell.fill.fill_type else None

def detect_ncols(ws):
    n = 0
    for i in range(1, 200):
        d = ws.column_dimensions.get(get_column_letter(i))
        if d and d.width is not None and abs(d.width - STD_COL) < 0.5: n += 1
        elif n > 10: break
    return 40 if n <= 44 else (56 if n <= 60 else 82)

def detect_logo_zone(ws):
    for mg in ws.merged_cells.ranges:
        if mg.min_col == 1 and mg.min_row <= 2 and (mg.max_row - mg.min_row) >= 2:
            return mg.max_col, mg.min_row, mg.max_row
    return None, None, None

def detect_meta_start(ws):
    for r in range(1, 3):
        for c in range(1, 100):
            cell = ws.cell(row=r, column=c)
            if cell.value and 'Form Code' in str(cell.value): return c
    return None

def has_header(ws):
    le, _, _ = detect_logo_zone(ws)
    ms = detect_meta_start(ws)
    return le is not None and ms is not None

def row_merge_ends(ws, row):
    return set(mg.max_col for mg in ws.merged_cells.ranges if mg.min_row <= row <= mg.max_row)

def classify(ws, row, ncols):
    rd = ws.row_dimensions.get(row)
    h = rd.height if rd else None
    c1 = ws.cell(row=row, column=1)
    fh = fhex(c1)
    if fh == '2C9CD7' and h and h <= 3: return 'rule_stripe'
    if h and h <= 5 and not (c1.value and str(c1.value).strip()): return 'spacer'
    if fh and fh in SECTION_FILLS:
        val = str(c1.value or '').upper()
        if 'NOTICE' in val or 'LƯU Ý' in val: return 'notice'
        for mg in ws.merged_cells.ranges:
            if mg.min_row == row and mg.max_row == row and (mg.max_col - mg.min_col + 1) >= ncols * 0.4:
                return 'section'
        if c1.value and len(str(c1.value).strip()) > 3: return 'section'
    if fh and fh in COLHDR_FILLS: return 'col_header'
    for cc in [2, 3, 5]:
        if cc <= ncols and fhex(ws.cell(row=row, column=cc)) in COLHDR_FILLS: return 'col_header'
    return 'data'

def merged_non_primary(ws):
    s = set()
    for mg in ws.merged_cells.ranges:
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                if r != mg.min_row or c != mg.min_col: s.add((r, c))
    return s

# ═══════ SAFE COLOR TRANSFORM ═══════
def safe_transform_colors(ws, ncols):
    """Map fill and font colors to master template palette.
    PRESERVES font size, bold, italic. Only changes color values."""
    mnp = merged_non_primary(ws)
    max_row = ws.max_row or 50

    for row_cells in ws.iter_rows(min_row=1, max_row=max_row, max_col=ncols):
        for cell in row_cells:
            if (cell.row, cell.column) in mnp:
                continue

            # Map fill color
            fc = fhex(cell)
            old_fill_was_dark = fc in ('1565C0', '0C2D48')
            if fc and fc in FILL_MAP:
                cell.fill = PatternFill('solid', fgColor=FILL_MAP[fc])

            # Map font color (preserve size, bold, italic)
            if cell.font and cell.font.color:
                fn_c = chex(cell.font.color)
                new_color = None

                # Explicit font color mapping
                if fn_c and fn_c in FONT_MAP:
                    new_color = FONT_MAP[fn_c]

                # When fill changed dark->light, white text must become dark
                # 1565C0 (dark blue) -> EFF6FF (light): white text -> #1B3A6B
                # 0C2D48 (dark navy) -> 005A9E (medium): white text stays white
                if old_fill_was_dark and fn_c == 'FFFFFF':
                    if fc == '1565C0':  # section header: dark->light
                        new_color = '1B3A6B'
                    # 0C2D48 col header: still dark bg, white text OK

                if new_color:
                    cell.font = Font(
                        name=cell.font.name or 'Segoe UI',
                        size=cell.font.size,      # PRESERVE
                        bold=cell.font.bold,       # PRESERVE
                        italic=cell.font.italic,   # PRESERVE
                        underline=cell.font.underline,
                        strike=cell.font.strikethrough,
                        color=new_color
                    )

            # Fix font name to Segoe UI (preserve everything else)
            if cell.font and cell.font.name and cell.font.name != 'Segoe UI':
                has_stuff = cell.value is not None or (cell.fill and cell.fill.fill_type)
                if has_stuff:
                    cell.font = Font(
                        name='Segoe UI',
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        strike=cell.font.strikethrough,
                        color=cell.font.color
                    )

    # Rule stripe rows -> spacer (h=4, white fill)
    for r in range(1, max_row + 1):
        c1 = ws.cell(row=r, column=1)
        fc = fhex(c1)
        rd = ws.row_dimensions.get(r)
        h = rd.height if rd else None
        if fc == '2C9CD7' and h and h <= 3:
            ws.row_dimensions[r].height = 4.0
            # Unmerge first
            to_rm = [mg for mg in ws.merged_cells.ranges if mg.min_row == r and mg.max_row == r]
            for mg in to_rm:
                ws.unmerge_cells(str(mg))
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill('solid', fgColor='FFFFFF')
                try: cell.value = None
                except: pass

# ═══════ BORDERS ═══════
def fb(ws, r, c, t=BN, b=BN, l=BN, ri=BN):
    try: ws.cell(row=r, column=c).border = Border(top=t, bottom=b, left=l, right=ri)
    except: pass

def fix_header_borders(ws, ncols):
    le, _, _ = detect_logo_zone(ws)
    ms = detect_meta_start(ws)
    if not le or not ms: return
    mle = ms
    for mg in ws.merged_cells.ranges:
        if mg.min_row <= 5 and mg.min_col == ms: mle = mg.max_col; break

    for r in range(1, 6):
        for c in range(1, ncols + 1):
            t = BB if r == 1 else BN
            b = BB if r == 5 else BN
            l = BB if c == 1 else BN
            ri = BB if c == ncols else BN
            if c == le: ri = BB
            if c == le + 1: l = BB
            if c == ms: l = BB
            if c >= ms and r > 1: t = BG
            if c == mle and c < ncols: ri = BG
            fb(ws, r, c, t, b, l, ri)

def fix_body_borders(ws, ncols, start):
    max_row = ws.max_row or 50
    types = {r: classify(ws, r, ncols) for r in range(start, max_row + 1)}
    block = []
    for r in range(start, max_row + 1):
        rt = types[r]
        if rt in ('spacer', 'rule_stripe'):
            if block: _db(ws, block, ncols); block = []
            for c in range(1, ncols + 1): fb(ws, r, c)
        elif rt in ('section', 'notice'):
            if block: _db(ws, block, ncols); block = []
            for c in range(1, ncols + 1):
                fb(ws, r, c, BB, BB, BB if c == 1 else BN, BB if c == ncols else BN)
        elif rt == 'col_header':
            if block: _db(ws, block, ncols); block = []
            me = row_merge_ends(ws, r)
            for c in range(1, ncols + 1):
                fb(ws, r, c, BB, BG, BB if c == 1 else BN,
                   BB if c == ncols else (BG if c in me else BN))
        else:
            block.append(r)
    if block: _db(ws, block, ncols)

def _db(ws, rows, ncols):
    if not rows: return
    f, la = rows[0], rows[-1]
    for r in rows:
        me = row_merge_ends(ws, r)
        for c in range(1, ncols + 1):
            if r == f and r == la: t, b = BB, BB
            elif r == f: t, b = BB, BG
            elif r == la: t, b = BN, BB
            else: t, b = BN, BG
            fb(ws, r, c, t, b, BB if c == 1 else BN,
               BB if c == ncols else (BG if c in me else BN))

# ═══════ LOGO ═══════
def embed_logo(ws, ncols):
    if not os.path.exists(LOGO_PATH): return False
    le, lr_start, lr_end = detect_logo_zone(ws)
    if le is None: return False

    ws._images = []

    # Zone size in mm
    zw = le * 5.03
    zh = sum((ws.row_dimensions[r].height or 15.0) * 0.353
             for r in range(lr_start, lr_end + 1))

    # Size logo to fit zone (responsive per grid)
    pct = 0.70 if ncols >= 56 else 0.65  # smaller for wider grids
    aw, ah = zw * pct, zh * 0.50
    if aw / LOGO_ASPECT <= ah: w, h = aw, aw / LOGO_ASPECT
    else: h, w = ah, ah * LOGO_ASPECT

    img = XlImage(LOGO_PATH)
    img.width = w / 25.4 * 96
    img.height = h / 25.4 * 96

    # Center offset (EMU, 1mm=36000)
    xo = max(int((zw - w) / 2 * 36000), 0)
    yo = max(int((zh - h) / 2 * 36000), 0)

    marker = AnchorMarker(col=0, colOff=xo, row=lr_start - 1, rowOff=yo)
    img.anchor = OneCellAnchor(_from=marker,
                                ext=XDRPositiveSize2D(int(img.width * 9525), int(img.height * 9525)))
    ws.add_image(img)
    return True

# ═══════ FOOTER ═══════
def fix_footer(ws):
    try:
        hf = ws.HeaderFooter
        if hf and hf.oddFooter and hf.oddFooter.center:
            hf.oddFooter.center.size = 7
            hf.oddFooter.center.font = "Segoe UI"
    except: pass

# ═══════ MAIN ═══════
def process(filepath):
    fname = os.path.basename(filepath)
    log = []
    try: wb = load_workbook(filepath)
    except Exception as e: return fname, [f"ERROR: {e}"], False

    # Fix Normal style
    try:
        for ns in wb._named_styles:
            if ns.font and ns.font.name and ns.font.name != 'Segoe UI':
                ns.font = Font(name='Segoe UI', size=ns.font.size, bold=ns.font.bold,
                              italic=ns.font.italic, color=ns.font.color)
    except: pass

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        hdr = has_header(ws)

        if not hdr:
            log.append(f"  [{ws_name}] SKIP (no header)")
            continue

        ncols = detect_ncols(ws)

        # 1. Safe color transform (fills + font colors)
        safe_transform_colors(ws, ncols)

        # 2. Logo
        embed_logo(ws, ncols)

        # 3. Header borders
        fix_header_borders(ws, ncols)

        # 4. Body borders
        body_start = 7
        for r in range(5, 10):
            rd = ws.row_dimensions.get(r)
            h = rd.height if rd else None
            fh = fhex(ws.cell(row=r, column=1))
            if (h and h <= 5) or fh in SECTION_FILLS or fh in COLHDR_FILLS:
                body_start = r; break
        fix_body_borders(ws, ncols, body_start)

        # 5. Footer
        fix_footer(ws)

        # 6. Page setup
        try: ws.sheet_view.showGridLines = False
        except: pass
        try:
            pm = ws.page_margins
            pm.left = pm.right = 0.197; pm.top = pm.bottom = 0.197
            pm.header = pm.footer = 0.1
            ps = ws.page_setup; ps.paperSize = 9 if ncols <= 56 else 8
            ps.fitToWidth = 1; ps.fitToHeight = 0
        except: pass

        log.append(f"  [{ws_name}] colors+borders+logo done")

    try:
        wb.save(filepath); log.append("  >> SAVED")
    except Exception as e:
        log.append(f"  >> SAVE ERROR: {e}")
    wb.close()
    return fname, log, True

def main():
    base = os.path.join(os.path.dirname(__file__), '04-Bieu-Mau')
    files = []
    for root, dirs, fnames in os.walk(base):
        if '_build' in root or '00-FORM-DESIGN-SYSTEM' in root: continue
        for f in fnames:
            if f.startswith('FRM-') and f.endswith('.xlsx') and not f.startswith('~'):
                files.append(os.path.join(root, f))
    files.sort()

    print(f"{'='*70}")
    print(f"HESEM QMS v5 (safe colors + borders + logo + footer)")
    print(f"Files: {len(files)}")
    print(f"{'='*70}\n")

    v = '--verbose' in sys.argv
    ok_count = err_count = 0
    for i, fp in enumerate(files, 1):
        fn = os.path.basename(fp)
        print(f"[{i:3d}/{len(files)}] {fn}...", end=" ", flush=True)
        _, log, ok = process(fp)
        if any('SAVE ERROR' in l for l in log):
            print("LOCKED"); err_count += 1
        else:
            print("OK"); ok_count += 1
        if v:
            for l in log: print(l)

    print(f"\n{'='*70}")
    print(f"DONE - {ok_count} saved, {err_count} locked")
    print(f"{'='*70}")

if __name__ == '__main__':
    main()
